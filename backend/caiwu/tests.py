import json
from decimal import Decimal

from django.test import Client, TestCase, override_settings
from openpyxl import Workbook

from caiwu.models import (
    BUSINESS_UNITS,
    FinancialEntry,
    FinancialTarget,
    ImportBatch,
    L1Category,
    L2Category,
    L3Category,
)
from caiwu.views import (
    _aggregate_report,
    _allocate_unalloc,
    _compute_l1_name_map,
    _compute_pl_check,
    _detect_dept_ledger,
    _detect_project_ledger,
    _get_published_batches,
    _make_token,
    _parse_dept_ledger_rows,
    _parse_project_ledger,
)
from paikuan.models import PaikuanUser


REV = '\u4e3b\u8425\u4e1a\u52a1\u6536\u5165'
COST = '\u4e3b\u8425\u4e1a\u52a1\u6210\u672c'
TAX = '\u7a0e\u91d1\u6210\u672c'
OPERATING_GROSS = '\u8fd0\u8425\u6bdb\u5229'
SALES_EXP = '\u9500\u552e\u8d39\u7528'
MGMT_EXP = '\u7ba1\u7406\u8d39\u7528'
FIN_EXP = '\u8d22\u52a1\u8d39\u7528'
NONOP_REV = '\u8425\u4e1a\u5916\u6536\u5165'
NONOP_EXP = '\u8425\u4e1a\u5916\u652f\u51fa'
OPERATING_PROFIT = '\u7ecf\u8425\u6bdb\u5229'
GROUP_MGMT = '\u96c6\u56e2\u7ba1\u7406\u8d39\u7528'
NET_PROFIT = '\u7ecf\u8425\u51c0\u5229'

L1_SEEDS = [
    (REV, 10, False, 1, True),
    (COST, 20, False, -1, True),
    (TAX, 30, False, -1, False),
    (OPERATING_GROSS, 40, True, 1, False),
    (SALES_EXP, 50, False, -1, True),
    (MGMT_EXP, 60, False, -1, True),
    (FIN_EXP, 70, False, -1, False),
    (NONOP_REV, 80, False, 1, False),
    (NONOP_EXP, 90, False, -1, False),
    (OPERATING_PROFIT, 100, True, 1, False),
    (GROUP_MGMT, 110, False, -1, True),
    (NET_PROFIT, 120, True, 1, False),
]

BASE_AMOUNTS = {
    REV: '1000.00',
    COST: '600.00',
    TAX: '50.00',
    SALES_EXP: '30.00',
    MGMT_EXP: '40.00',
    FIN_EXP: '10.00',
    NONOP_REV: '20.00',
    NONOP_EXP: '5.00',
    GROUP_MGMT: '25.00',
}

CURRENT_AMOUNTS = {
    REV: '1200.00',
    COST: '650.00',
    TAX: '60.00',
    SALES_EXP: '35.00',
    MGMT_EXP: '45.00',
    FIN_EXP: '12.00',
    NONOP_REV: '25.00',
    NONOP_EXP: '8.00',
    GROUP_MGMT: '30.00',
}


@override_settings(ALLOWED_HOSTS=['testserver', 'localhost', '127.0.0.1'])
class CaiwuCalculationLogicTests(TestCase):
    # caiwu 已并入 default 库（平台整合阶段1），测试不再需要独立的 caiwu alias。
    databases = {'default'}

    @classmethod
    def setUpTestData(cls):
        for name, sort_order, is_calculated, sign, is_profit_driver in L1_SEEDS:
            L1Category.objects.update_or_create(
                name=name,
                defaults={
                    'sort_order': sort_order,
                    'is_calculated': is_calculated,
                    'sign': sign,
                    'is_profit_driver': is_profit_driver,
                },
            )
        # Unified platform account (Stage 2+3): auth + uploaded_by both use it.
        cls.admin = PaikuanUser(
            phone='13900000000',
            name='Finance Admin',
            role='super_admin',
            job_title='finance_director',
            departments=[],
            is_active=True,
            is_approved=True,
        )
        cls.admin.set_password('Test123456')
        cls.admin.save()

    def setUp(self):
        self.client = Client()
        self.bu = BUSINESS_UNITS[0]
        self.l1 = {c.name: c for c in L1Category.objects.order_by('sort_order', 'id')}

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {_make_token(self.admin)}'}

    def create_batch(
        self,
        year=2026,
        month=5,
        batch_type=ImportBatch.TYPE_DEPT,
        status=ImportBatch.STATUS_PUBLISHED,
        amounts=None,
        bu=None,
    ):
        batch = ImportBatch.objects.create(
            business_unit=bu or self.bu,
            year=year,
            month=month,
            batch_type=batch_type,
            status=status,
            uploaded_by=self.admin,
            row_count=len(amounts or {}),
            file_name='logic-test.xlsx',
        )
        for name, amount in (amounts or {}).items():
            FinancialEntry.objects.create(
                batch=batch,
                l1=self.l1[name],
                amount=Decimal(str(amount)),
            )
        return batch

    def response_json(self, resp):
        return json.loads(resp.content.decode('utf-8'))

    def test_batch_uploader_is_unified_account(self):
        # Regression: after the platform merge the uploader FK targets
        # PaikuanUser; to_dict() must surface that account's name, not None.
        batch = self.create_batch(amounts={REV: '1.00'})
        self.assertEqual(batch.uploaded_by_id, self.admin.id)
        self.assertEqual(batch.to_dict()['uploaded_by'], self.admin.name)

    def test_l1_formula_chain_computes_bottom_line(self):
        l1_cats = list(L1Category.objects.order_by('sort_order', 'id'))
        raw_by_id = {self.l1[name].id: Decimal(amount) for name, amount in BASE_AMOUNTS.items()}

        name_map, _ = _compute_l1_name_map(l1_cats, raw_by_id)

        self.assertEqual(name_map[OPERATING_GROSS], 350.0)
        self.assertEqual(name_map[OPERATING_PROFIT], 285.0)
        self.assertEqual(name_map[NET_PROFIT], 260.0)

    def test_report_aggregates_department_detail_only(self):
        self.create_batch(amounts=BASE_AMOUNTS, batch_type=ImportBatch.TYPE_DEPT)
        self.create_batch(
            amounts={REV: '9999.00', COST: '1.00'},
            batch_type=ImportBatch.TYPE_PL,
        )

        rows = _aggregate_report(_get_published_batches([self.bu], 2026, 5), 1)
        by_name = {row['l1_name']: row['amount'] for row in rows}

        self.assertEqual(by_name[REV], 1000.0)
        self.assertEqual(by_name[COST], 600.0)
        self.assertEqual(by_name[NET_PROFIT], 260.0)

    def test_publish_replaces_same_period_and_type_only(self):
        old_dept = self.create_batch(amounts=BASE_AMOUNTS, batch_type=ImportBatch.TYPE_DEPT)
        old_pl = self.create_batch(amounts={REV: '9999.00'}, batch_type=ImportBatch.TYPE_PL)
        draft_dept = self.create_batch(
            amounts=CURRENT_AMOUNTS,
            batch_type=ImportBatch.TYPE_DEPT,
            status=ImportBatch.STATUS_DRAFT,
        )

        resp = self.client.put(f'/api/cw/batches/{draft_dept.id}/publish', **self.auth())

        self.assertEqual(resp.status_code, 200, self.response_json(resp))
        self.assertFalse(ImportBatch.objects.filter(id=old_dept.id).exists())
        self.assertTrue(ImportBatch.objects.filter(id=old_pl.id, status=ImportBatch.STATUS_PUBLISHED).exists())
        draft_dept.refresh_from_db()
        self.assertEqual(draft_dept.status, ImportBatch.STATUS_PUBLISHED)

    def test_waterfall_factors_bridge_net_profit_delta(self):
        self.create_batch(year=2026, month=4, amounts=BASE_AMOUNTS)
        self.create_batch(year=2026, month=5, amounts=CURRENT_AMOUNTS)

        resp = self.client.get(
            '/api/cw/charts/waterfall',
            {
                'bu': self.bu,
                'year': 2026,
                'month': 5,
                'compare_year': 2026,
                'compare_month': 4,
            },
            **self.auth(),
        )
        payload = self.response_json(resp)

        self.assertEqual(resp.status_code, 200, payload)
        data = payload['data']
        base_total = Decimal(str(data['base_period']['total']))
        current_total = Decimal(str(data['current_period']['total']))
        factor_delta = sum(Decimal(str(f['delta'])) for f in data['factors'])
        waterfall_delta = sum(
            Decimal(str(w['value']))
            for w in data['waterfall']
            if w['type'] in ('increase', 'decrease')
        )

        self.assertEqual(base_total, Decimal('260.0'))
        self.assertEqual(current_total, Decimal('385.0'))
        self.assertEqual(factor_delta, current_total - base_total)
        self.assertEqual(waterfall_delta, current_total - base_total)

    def test_waterfall_bridge_closes_when_current_period_is_a_loss(self):
        # Regression: loss period (negative terminal). The terminal bar is drawn
        # 0→value (below the axis); the last decrease factor must land exactly on
        # it, i.e. base + Σ(waterfall factor bars) == current total — no floating gap.
        self.create_batch(year=2026, month=4, amounts=BASE_AMOUNTS)
        loss_amounts = dict(BASE_AMOUNTS, **{REV: '100.00', COST: '900.00'})
        self.create_batch(year=2026, month=5, amounts=loss_amounts)

        resp = self.client.get(
            '/api/cw/charts/waterfall',
            {'bu': self.bu, 'year': 2026, 'month': 5,
             'compare_year': 2026, 'compare_month': 4},
            **self.auth(),
        )
        payload = self.response_json(resp)
        self.assertEqual(resp.status_code, 200, payload)
        data = payload['data']

        base_total = Decimal(str(data['base_period']['total']))
        current_total = Decimal(str(data['current_period']['total']))
        self.assertLess(current_total, 0)  # genuine loss

        bars = data['waterfall']
        self.assertEqual(bars[0]['type'], 'base')
        self.assertEqual(bars[-1]['type'], 'total')
        factor_sum = sum(
            Decimal(str(b['value'])) for b in bars
            if b['type'] in ('increase', 'decrease')
        )
        # Bridge closes: terminal value is reached exactly by the cumulative.
        self.assertEqual(base_total + factor_sum, current_total)
        self.assertEqual(Decimal(str(bars[-1]['value'])), current_total)

    def test_closed_period_ledger_parser_excludes_carry_forward_and_summary_rows(self):
        wb = Workbook()
        ws = wb.active
        ws.append([
            '\u90e8\u95e8\u540d\u79f0',
            '\u79d1\u76ee\u7f16\u7801',
            '\u79d1\u76ee\u540d\u79f0',
            '\u6458\u8981',
            '\u501f\u65b9',
            '\u8d37\u65b9',
        ])
        ws.append([self.bu, '6001.01', '\u9500\u552e\u6536\u5165', '\u51ed\u8bc1001', 0, 1000])
        ws.append([self.bu, '6001.01', '\u9500\u552e\u6536\u5165', '\u7ed3\u8f6c\u635f\u76ca', 1000, 0])
        ws.append([self.bu, '6001.01', '\u9500\u552e\u6536\u5165', '\u672c\u671f\u5408\u8ba1', 1000, 1000])
        ws.append([self.bu, '6401.01', '\u8fd0\u8f93\u6210\u672c', '\u51ed\u8bc1002', 600, 0])
        ws.append([self.bu, '6401.01', '\u8fd0\u8f93\u6210\u672c', '\u672c\u5e74\u7d2f\u8ba1', 600, 0])

        data_start, col_map = _detect_dept_ledger(ws)
        parsed, errors = _parse_dept_ledger_rows(
            ws,
            data_start,
            col_map,
            self.bu,
            self.l1,
            {c.name: c for c in L2Category.objects.filter(business_unit=self.bu)},
            {(c.l1_category_id, c.name): c for c in L3Category.objects.filter(business_unit=self.bu)},
        )
        by_name = {}
        for row in parsed:
            by_name[row['l1_name']] = by_name.get(row['l1_name'], Decimal('0')) + row['amount']

        self.assertEqual(errors, [])
        self.assertEqual(by_name[REV], Decimal('1000'))
        self.assertEqual(by_name[COST], Decimal('600'))
        self.assertEqual(len(parsed), 2)

    def test_6602_99_03_maps_to_group_management_fee(self):
        """集团管理费用仅取 6602.99.03 本科目（集团管理费分摊/收回）；其同级的
        6602.99.01 培训费、6602.99.02 会议费等属普通管理费用，按 6602 前缀归入
        管理费用，不计入集团管理费。科目名称含「集团管理费用」者另由名称兜底覆盖。"""
        wb = Workbook()
        ws = wb.active
        ws.append([
            '部门名称',   # 部门名称
            '科目编码',   # 科目编码
            '科目名称',   # 科目名称
            '摘要',               # 摘要
            '借方',               # 借方
            '贷方',               # 贷方
        ])
        # 6602.01 普通管理费 → 管理费用
        ws.append([self.bu, '6602.01', '办公费', '凭证001', 100, 0])
        # 6602.99.03 集团管理费用本科目 → 集团管理费用（即使名称不含也按编码归集）
        ws.append([self.bu, '6602.99.03', '分摊费用', '凭证002', 200, 0])
        # 6602.99.01 培训费等同级子目 → 回落 6602 前缀 → 管理费用（不计集团管理费）
        ws.append([self.bu, '6602.99.01', '外部咨询培训费', '凭证003', 50, 0])

        data_start, col_map = _detect_dept_ledger(ws)
        parsed, errors = _parse_dept_ledger_rows(
            ws, data_start, col_map, self.bu, self.l1,
            {c.name: c for c in L2Category.objects.filter(business_unit=self.bu)},
            {(c.l1_category_id, c.name): c for c in L3Category.objects.filter(business_unit=self.bu)},
        )
        by_name = {}
        for row in parsed:
            by_name[row['l1_name']] = by_name.get(row['l1_name'], Decimal('0')) + row['amount']

        self.assertEqual(errors, [])
        # 管理费用 sign=-1 → 办公费100 + 培训费50 = 150；集团管理费用 = 仅 6602.99.03 的 200
        self.assertEqual(by_name[MGMT_EXP], Decimal('150'))
        self.assertEqual(by_name[GROUP_MGMT], Decimal('200'))

    def test_hq_import_excludes_finance_dept(self):
        """集团总部导入时整段剔除「财务金融」部门（供应链金融独立条线），
        其收入/成本/费用均不计入集团总部报表；其他部门正常计入。"""
        self.assertEqual(self.bu, '集团总部')   # BUSINESS_UNITS[0]
        wb = Workbook()
        ws = wb.active
        ws.append(['部门名称', '科目编码', '科目名称', '摘要', '借方', '贷方'])
        # 财务金融：收入80 / 成本30 / 财务费用5 —— 应被整段剔除
        ws.append([self.bu, '6001', '服务费收入', '凭证001', 0, 80])
        ws.append([self.bu, '6401', '主营业务成本', '凭证002', 30, 0])
        ws.append([self.bu, '6603', '利息支出', '凭证003', 5, 0])
        # 行政园区：园区收入100 —— 应正常计入
        ws.append(['行政园区', '6001', '仓库租赁收入', '凭证004', 0, 100])

        # 重新读 dept 列：第一行表头里「部门名称」需要对上财务金融
        for ri, dept in [(2, '财务金融'), (3, '财务金融'), (4, '财务金融'), (5, '行政园区')]:
            ws.cell(ri, 1, dept)

        data_start, col_map = _detect_dept_ledger(ws)
        parsed, errors = _parse_dept_ledger_rows(
            ws, data_start, col_map, self.bu, self.l1,
            {c.name: c for c in L2Category.objects.filter(business_unit=self.bu)},
            {(c.l1_category_id, c.name): c for c in L3Category.objects.filter(business_unit=self.bu)},
        )
        by_name = {}
        for row in parsed:
            by_name[row['l1_name']] = by_name.get(row['l1_name'], Decimal('0')) + row['amount']

        self.assertEqual(errors, [])
        # 仅行政园区的园区收入100计入；财务金融全部剔除
        self.assertEqual(by_name.get(REV), Decimal('100'))
        self.assertNotIn(COST, by_name)
        # 剔除的部门不应创建二级项目部
        self.assertFalse(L2Category.objects.filter(business_unit=self.bu, name='财务金融').exists())

    def test_report_excludes_finance_dept_retroactively(self):
        """报表聚合层剔除集团总部的财务金融——即使历史已发布批次里仍含该部门，
        报表也不计入（无需重新导入即对旧数据生效）。"""
        self.assertEqual(self.bu, '集团总部')
        batch = ImportBatch.objects.create(
            business_unit=self.bu, year=2026, month=5,
            batch_type=ImportBatch.TYPE_DEPT, status=ImportBatch.STATUS_PUBLISHED,
            uploaded_by=self.admin, row_count=2, file_name='hq.xlsx',
        )
        fin = L2Category.objects.create(business_unit=self.bu, name='财务金融')
        park = L2Category.objects.create(business_unit=self.bu, name='行政园区')
        # 财务金融收入80（应剔除）+ 行政园区收入100（应保留）
        FinancialEntry.objects.create(batch=batch, l1=self.l1[REV], l2=fin, amount=Decimal('80'))
        FinancialEntry.objects.create(batch=batch, l1=self.l1[REV], l2=park, amount=Decimal('100'))

        rows = _aggregate_report(_get_published_batches([self.bu], 2026, 5), 1)
        rev = next((r['amount'] for r in rows if r['l1_name'] == REV), None)
        self.assertEqual(rev, 100.0)   # 仅行政园区100，财务金融80被剔除

        # 二级明细里也不应出现财务金融
        rows2 = _aggregate_report(_get_published_batches([self.bu], 2026, 5), 2)
        rev_row = next(r for r in rows2 if r['l1_name'] == REV)
        l2names = {c['l2_name'] for c in rev_row['children']}
        self.assertIn('行政园区', l2names)
        self.assertNotIn('财务金融', l2names)

    def test_pl_check_matches_report_kpis(self):
        """导入预览的「数据核对」KPI（_compute_pl_check）应与发布后的财务报表
        （_aggregate_report）逐项一致——回归用户反馈的「导入核对显示异常但报表正常」。"""
        parsed_rows = []
        for name, amt in BASE_AMOUNTS.items():
            parsed_rows.append({
                'l1': self.l1[name], 'l2': None, 'l3': None,
                'amount': Decimal(amt),
                'l1_name': name, 'l2_name': '', 'l3_name': '',
            })
        pl = _compute_pl_check(parsed_rows)
        kpi = {r['name']: r['amount'] for r in pl['kpis']}
        self.assertEqual(kpi[OPERATING_GROSS], 350.0)
        self.assertEqual(kpi[OPERATING_PROFIT], 285.0)
        self.assertEqual(kpi[NET_PROFIT], 260.0)

        # 与发布后报表逐项一致
        self.create_batch(amounts=BASE_AMOUNTS, batch_type=ImportBatch.TYPE_DEPT)
        report = {r['l1_name']: r['amount']
                  for r in _aggregate_report(_get_published_batches([self.bu], 2026, 5), 1)}
        for r in pl['l1_summary']:
            self.assertAlmostEqual(r['amount'], report.get(r['name'], 0), places=2,
                                   msg=f"{r['name']} 预览={r['amount']} 报表={report.get(r['name'])}")

    def test_project_ledger_parse_and_allocate(self):
        """项目核算明细账（维度=项目名称）解析：6001→收入(贷-借)、6401→成本(借-贷)，
        跳过小计/结转损益行；以及未挂成本按收入比例分摊。"""
        wb = Workbook()
        ws = wb.active
        ws.append(['核算维度明细账'])
        ws.append(['账簿信息'])
        ws.append(['序号', '项目名称', '科目编码', '科目名称', '会计期间',
                   '记账日期', '业务日期', '凭证字号', '摘要', '币种', '借方', '贷方'])
        # 甲项目：收入1000(贷)、成本300(借)
        ws.append([1, '甲', '6001.01', '主营收入', '2026年5期', '2026-05-10', None, '记1', '直客收入', '人民币', 0, 1000])
        ws.append([2, '甲', '6401.01', '运输成本', '2026年5期', '2026-05-11', None, '记2', '成本', '人民币', 300, 0])
        # 乙项目：收入500(贷)
        ws.append([3, '乙', '6001.01', '主营收入', '2026年5期', '2026-05-12', None, '记3', '直客收入', '人民币', 0, 500])
        # 未挂项目「无」：成本600（待分摊）
        ws.append([4, '无', '6401.01', '分摊成本', '2026年5期', '2026-05-20', None, '记4', '公共成本', '人民币', 600, 0])
        # 小计 & 结转损益行：应跳过
        ws.append([5, '甲', '6001.01', '主营收入', '2026年5期', '2026-05-31', None, None, '本期合计', '人民币', 0, 1000])
        ws.append([6, '甲', '6001.01', '主营收入', '2026年5期', '2026-05-31', None, '记9', '结转损益', '人民币', 1000, 0])

        ds, cm = _detect_project_ledger(ws)
        self.assertIsNotNone(ds)
        agg = _parse_project_ledger(ws, ds, cm)
        self.assertEqual(agg['甲']['revenue'], Decimal('1000'))
        self.assertEqual(agg['甲']['cost'], Decimal('300'))
        self.assertEqual(agg['乙']['revenue'], Decimal('500'))
        self.assertEqual(agg['无']['cost'], Decimal('600'))

        # 分摊：未挂成本600 按收入(甲1000:乙500=2:1)分摊 → 甲+400、乙+200
        rows = [
            {'project_name': '甲', 'revenue': 1000.0, 'cost': 300.0, 'sales_exp': 0.0, 'mgmt_exp': 0.0, 'margin': 700.0, 'margin_rate': 70.0},
            {'project_name': '乙', 'revenue': 500.0, 'cost': 0.0, 'sales_exp': 0.0, 'mgmt_exp': 0.0, 'margin': 500.0, 'margin_rate': 100.0},
        ]
        unalloc = {'revenue': 0.0, 'cost': 600.0, 'sales_exp': 0.0, 'mgmt_exp': 0.0}
        out = _allocate_unalloc(rows, unalloc)
        by = {r['project_name']: r for r in out}
        self.assertEqual(by['甲']['cost'], 700.0)   # 300 + 400
        self.assertEqual(by['甲']['margin'], 300.0)  # 1000 - 700
        self.assertEqual(by['乙']['cost'], 200.0)    # 0 + 200
        self.assertEqual(by['乙']['margin'], 300.0)  # 500 - 200

    def test_template_has_no_profit_loss_sheet(self):
        """利润表已下线：下载模板不应再含「利润表模板」页（仅部门明细相关页）。"""
        import io
        from openpyxl import load_workbook
        resp = self.client.get('/api/cw/batches/template', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        wb = load_workbook(io.BytesIO(resp.content))
        self.assertNotIn('利润表模板', wb.sheetnames)


@override_settings(ALLOWED_HOSTS=['testserver', 'localhost', '127.0.0.1'])
class CaiwuUnifiedPermissionTests(TestCase):
    """Stage 2-4: caiwu auth/permissions are driven by the paikuan platform.
    These lock in the cross-module behaviour (page gating + shared cache)."""
    databases = {'default'}

    @classmethod
    def setUpTestData(cls):
        for name, sort_order, is_calculated, sign, is_profit_driver in L1_SEEDS:
            L1Category.objects.update_or_create(
                name=name,
                defaults={'sort_order': sort_order, 'is_calculated': is_calculated,
                          'sign': sign, 'is_profit_driver': is_profit_driver},
            )
        cls.bu = BUSINESS_UNITS[0]
        # super_admin (paikuan) — manages permissions
        cls.admin = PaikuanUser(phone='13700000000', name='Admin', role='super_admin',
                                job_title='', departments=[], is_active=True, is_approved=True)
        cls.admin.set_password('Test123456'); cls.admin.save()
        # finance_director — should get full 财务分析 access by default
        cls.fin = PaikuanUser(phone='13700000001', name='Finance', role='viewer',
                              job_title='finance_director', departments=[cls.bu],
                              is_active=True, is_approved=True)
        cls.fin.set_password('Test123456'); cls.fin.save()
        # cashier — should have NO 财务分析 access by default
        cls.cashier = PaikuanUser(phone='13700000002', name='Cashier', role='viewer',
                                  job_title='cashier', departments=[cls.bu],
                                  is_active=True, is_approved=True)
        cls.cashier.set_password('Test123456'); cls.cashier.save()

    def setUp(self):
        self.client = Client()
        # Clear caches so a prior test's stored JobPermission doesn't leak.
        from paikuan.views import _invalidate_perm_cache as pk_inv
        from caiwu.views import _invalidate_perm_cache as cw_inv
        pk_inv(); cw_inv()

    def hdr(self, user):
        return {'HTTP_AUTHORIZATION': f'Bearer {_make_token(user)}'}

    def jj(self, resp):
        return json.loads(resp.content.decode('utf-8'))

    def test_finance_director_can_access_report(self):
        resp = self.client.get('/api/cw/report',
                               {'year': 2026, 'month': 5, 'bu': self.bu, 'level': 1},
                               **self.hdr(self.fin))
        self.assertEqual(resp.status_code, 200, self.jj(resp))

    def test_cashier_denied_report_and_upload(self):
        r1 = self.client.get('/api/cw/report',
                             {'year': 2026, 'month': 5, 'bu': self.bu, 'level': 1},
                             **self.hdr(self.cashier))
        self.assertEqual(r1.status_code, 403, self.jj(r1))
        r2 = self.client.post('/api/cw/batches/upload', {'bu': self.bu, 'year': 2026, 'month': 5},
                              **self.hdr(self.cashier))
        self.assertEqual(r2.status_code, 403, self.jj(r2))

    def test_cashier_denied_new_ai_and_project_endpoints(self):
        """无财务分析权限者不能访问 项目毛利 / 驾驶舱知识库 / 技能（回归权限是否跟上新功能）。"""
        r = self.client.get('/api/cw/project-margin',
                            {'bu': self.bu, 'year': 2026, 'month': 5}, **self.hdr(self.cashier))
        self.assertEqual(r.status_code, 403, self.jj(r))
        for url in ('/api/cw/cockpit/knowledge', '/api/cw/cockpit/skills'):
            self.assertEqual(self.client.get(url, **self.hdr(self.cashier)).status_code, 403)
        chat = self.client.post('/api/cw/cockpit/ai-chat/stream',
                                data=json.dumps({'year': 2026, 'month': 5, 'bu': self.bu,
                                                 'messages': [{'role': 'user', 'content': 'hi'}]}),
                                content_type='application/json', **self.hdr(self.cashier))
        self.assertEqual(chat.status_code, 403, self.jj(chat))

    def test_paikuan_permission_edit_invalidates_caiwu_cache(self):
        # finance_director starts with caiwu_report access
        before = self.client.get('/api/cw/report',
                                 {'year': 2026, 'month': 5, 'bu': self.bu, 'level': 1},
                                 **self.hdr(self.fin))
        self.assertEqual(before.status_code, 200, self.jj(before))

        # super_admin disables 财务分析·报表 for finance_director via the paikuan UI path
        perms = self.jj(self.client.get('/api/pk/permissions', **self.hdr(self.admin)))
        cfg = next(j['config'] for j in perms['data']['jobs'] if j['job_title'] == 'finance_director')
        cfg['pages']['caiwu_report'] = False
        put = self.client.put('/api/pk/permissions/finance_director',
                              data=json.dumps({'config': cfg}),
                              content_type='application/json', **self.hdr(self.admin))
        self.assertEqual(put.status_code, 200, self.jj(put))

        # Change must take effect immediately (caiwu cache invalidated)
        after = self.client.get('/api/cw/report',
                                {'year': 2026, 'month': 5, 'bu': self.bu, 'level': 1},
                                **self.hdr(self.fin))
        self.assertEqual(after.status_code, 403, self.jj(after))


@override_settings(ALLOWED_HOSTS=['testserver', 'localhost', '127.0.0.1'])
class CaiwuMetricsAndTargetsTests(TestCase):
    """指标管理 / 财务驾驶舱：目标录入校验 + 完成情况取数（达成率/环比/同比/YTD）。"""
    databases = {'default'}

    @classmethod
    def setUpTestData(cls):
        for name, sort_order, is_calculated, sign, is_profit_driver in L1_SEEDS:
            L1Category.objects.update_or_create(
                name=name,
                defaults={'sort_order': sort_order, 'is_calculated': is_calculated,
                          'sign': sign, 'is_profit_driver': is_profit_driver},
            )
        cls.admin = PaikuanUser(phone='13900000009', name='Metrics Admin',
                                role='super_admin', job_title='finance_director',
                                departments=[], is_active=True, is_approved=True)
        cls.admin.set_password('Test123456')
        cls.admin.save()

    def setUp(self):
        self.client = Client()
        self.bu = BUSINESS_UNITS[1]   # 一个真实事业部（非集团总部）
        self.l1 = {c.name: c for c in L1Category.objects.all()}

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {_make_token(self.admin)}'}

    def jj(self, resp):
        return json.loads(resp.content.decode('utf-8'))

    def mk(self, year, month, rev, cost):
        b = ImportBatch.objects.create(
            business_unit=self.bu, year=year, month=month,
            batch_type=ImportBatch.TYPE_DEPT, status=ImportBatch.STATUS_PUBLISHED,
            uploaded_by=self.admin, row_count=2, file_name='m.xlsx')
        FinancialEntry.objects.create(batch=b, l1=self.l1[REV], amount=Decimal(str(rev)))
        FinancialEntry.objects.create(batch=b, l1=self.l1[COST], amount=Decimal(str(cost)))
        return b

    def post_targets(self, items, year=2026):
        return self.client.post('/api/cw/targets',
                                data=json.dumps({'year': year, 'items': items}),
                                content_type='application/json', **self.auth())

    # ── 目标录入校验 ──────────────────────────────────────────────────────────
    def _full_year_items(self, monthly_rev, monthly_prof, annual_rev, annual_prof,
                         monthly_gross=0, annual_gross=0):
        items = [{'business_unit': self.bu, 'month': m,
                  'target_revenue': monthly_rev, 'target_profit': monthly_prof,
                  'target_gross_profit': monthly_gross}
                 for m in range(1, 13)]
        items.append({'business_unit': self.bu, 'month': 0,
                      'target_revenue': annual_rev, 'target_profit': annual_prof,
                      'target_gross_profit': annual_gross})
        return items

    def test_month_sum_must_equal_annual(self):
        # 12×100 = 1200 收入；年度填 1300 → 拒绝
        bad = self.post_targets(self._full_year_items(100, 10, 1300, 120, 8, 96))
        self.assertEqual(bad.status_code, 400, self.jj(bad))
        self.assertIn('收入', self.jj(bad)['error'])
        # 经营净利不符 → 拒绝
        bad2 = self.post_targets(self._full_year_items(100, 10, 1200, 999, 8, 96))
        self.assertEqual(bad2.status_code, 400, self.jj(bad2))
        self.assertIn('经营净利', self.jj(bad2)['error'])
        # 经营毛利不符 → 拒绝
        bad3 = self.post_targets(self._full_year_items(100, 10, 1200, 120, 8, 999))
        self.assertEqual(bad3.status_code, 400, self.jj(bad3))
        self.assertIn('经营毛利', self.jj(bad3)['error'])
        # 全部一致 → 通过
        good = self.post_targets(self._full_year_items(100, 10, 1200, 120, 8, 96))
        self.assertEqual(good.status_code, 200, self.jj(good))
        self.assertEqual(self.jj(good)['data']['saved'], 13)

    def test_validation_merges_with_existing_db(self):
        # 数据库已存 12 个月（合计 1200 收入）
        for m in range(1, 13):
            FinancialTarget.objects.create(business_unit=self.bu, year=2026, month=m,
                                           target_revenue=Decimal('100'), target_profit=Decimal('10'))
        # 仅提交一个与库内 12 月不符的年度目标 → 应被拒绝（基于合并后的状态）
        bad = self.post_targets([{'business_unit': self.bu, 'month': 0,
                                  'target_revenue': 2000, 'target_profit': 120}])
        self.assertEqual(bad.status_code, 400, self.jj(bad))
        # 提交相符的年度目标 → 通过
        good = self.post_targets([{'business_unit': self.bu, 'month': 0,
                                   'target_revenue': 1200, 'target_profit': 120}])
        self.assertEqual(good.status_code, 200, self.jj(good))

    # ── 完成情况取数 ──────────────────────────────────────────────────────────
    def test_metrics_rate_mom_yoy_ytd(self):
        self.mk(2026, 4, 150, 100)   # 上月：利润 50
        self.mk(2026, 5, 200, 130)   # 本月：利润 70
        self.mk(2025, 5, 180, 120)   # 去年同月：利润 60
        FinancialTarget.objects.create(business_unit=self.bu, year=2026, month=5,
                                       target_revenue=Decimal('250'), target_profit=Decimal('60'))
        FinancialTarget.objects.create(business_unit=self.bu, year=2026, month=0,
                                       target_revenue=Decimal('3000'), target_profit=Decimal('700'))
        resp = self.client.get('/api/cw/metrics',
                               {'year': 2026, 'month': 5, 'bu': self.bu}, **self.auth())
        self.assertEqual(resp.status_code, 200, self.jj(resp))
        m = next(b for b in self.jj(resp)['data']['bus'] if b['business_unit'] == self.bu)
        self.assertAlmostEqual(m['month']['actual_revenue'], 200)
        self.assertAlmostEqual(m['month']['actual_profit'], 70)
        self.assertAlmostEqual(m['month']['revenue_rate'], 80.0)        # 200/250
        self.assertAlmostEqual(m['month']['profit_rate'], 116.7, places=1)  # 70/60
        self.assertAlmostEqual(m['month']['revenue_mom'], 33.3, places=1)   # (200-150)/150
        self.assertAlmostEqual(m['month']['profit_mom'], 40.0, places=1)    # (70-50)/50
        self.assertAlmostEqual(m['month']['revenue_yoy'], 11.1, places=1)   # (200-180)/180
        self.assertAlmostEqual(m['ytd']['actual_revenue'], 350)            # 150+200
        self.assertAlmostEqual(m['ytd']['actual_profit'], 120)            # 50+70

    def test_cockpit_overview_and_12_month_trend(self):
        self.mk(2026, 5, 200, 130)
        FinancialTarget.objects.create(business_unit=self.bu, year=2026, month=5,
                                       target_revenue=Decimal('250'), target_profit=Decimal('60'))
        resp = self.client.get('/api/cw/cockpit',
                               {'year': 2026, 'month': 5, 'bu': self.bu}, **self.auth())
        self.assertEqual(resp.status_code, 200, self.jj(resp))
        data = self.jj(resp)['data']
        self.assertEqual(len(data['trend']), 12)
        m5 = next(t for t in data['trend'] if t['month'] == 5)
        self.assertAlmostEqual(m5['actual_revenue'], 200)
        self.assertAlmostEqual(m5['target_revenue'], 250)
        self.assertAlmostEqual(data['overview']['month']['revenue_rate'], 80.0)
        # MoM/YoY 键应始终存在（无对比期时为 None）
        self.assertIn('revenue_mom', data['overview']['month'])

    # ── 驾驶舱全局 AI 分析（mock 掉外部模型调用）─────────────────────────────
    def test_cockpit_ai_uses_pro_model_and_group_scope(self):
        from unittest import mock
        from django.conf import settings
        self.mk(2026, 5, 200, 130)
        captured = {}

        def fake_chat(messages, timeout=90, model=None, max_tokens=1800):
            captured['model'] = model
            captured['max_tokens'] = max_tokens
            captured['prompt'] = messages[-1]['content']
            return '【模拟分析】全集团经营稳健。'

        with mock.patch('caiwu.views._deepseek_chat', fake_chat):
            resp = self.client.post(
                '/api/cw/cockpit/ai-analysis',
                data=json.dumps({'year': 2026, 'month': 5}),
                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, self.jj(resp))
        data = self.jj(resp)['data']
        self.assertEqual(data['model'], settings.DEEPSEEK_PRO_MODEL)
        self.assertEqual(data['scope'], '全集团')
        self.assertIn('analysis', data)
        # 用更强模型 + 更大 token 预算，提示词带全集团口径
        self.assertEqual(captured['model'], settings.DEEPSEEK_PRO_MODEL)
        self.assertGreaterEqual(captured['max_tokens'], 3000)
        self.assertIn('全集团', captured['prompt'])

    def test_cockpit_ai_no_data_returns_error(self):
        from unittest import mock
        with mock.patch('caiwu.views._deepseek_chat') as m:
            resp = self.client.post(
                '/api/cw/cockpit/ai-analysis',
                data=json.dumps({'year': 2026, 'month': 7}),
                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 400, self.jj(resp))
        m.assert_not_called()   # 无数据时不应调用外部模型

    def test_cockpit_ai_stream_emits_sse_frames(self):
        from unittest import mock
        from django.conf import settings
        self.mk(2026, 5, 200, 130)
        captured = {}

        def fake_stream(messages, model=None, max_tokens=1800, timeout=300):
            captured['model'] = model
            captured['max_tokens'] = max_tokens
            yield ('reasoning', '先看全集团达成')
            yield ('answer', '## 总览\n')
            yield ('answer', '集团收入500万。')

        with mock.patch('caiwu.views._deepseek_stream', fake_stream):
            resp = self.client.post(
                '/api/cw/cockpit/ai-analysis/stream',
                data=json.dumps({'year': 2026, 'month': 5, 'bu': self.bu}),
                content_type='application/json', **self.auth())
            self.assertEqual(resp.status_code, 200)
            self.assertEqual(resp['Content-Type'], 'text/event-stream')
            self.assertEqual(resp['X-Accel-Buffering'], 'no')
            # 必须在 patch 生效期间消费惰性生成器，否则会落到真实的流式调用。
            body = b''.join(resp.streaming_content).decode('utf-8')
        events = [json.loads(fr[5:].strip())
                  for fr in body.split('\n\n') if fr.strip().startswith('data:')]
        types = [e['type'] for e in events]
        self.assertEqual(types[0], 'meta')
        self.assertEqual(types[-1], 'done')
        self.assertIn('reasoning', types)
        self.assertIn('answer', types)
        answer = ''.join(e['delta'] for e in events if e['type'] == 'answer')
        self.assertEqual(answer, '## 总览\n集团收入500万。')
        self.assertEqual(captured['model'], settings.DEEPSEEK_PRO_MODEL)
        self.assertGreaterEqual(captured['max_tokens'], 3000)

    def test_cockpit_ai_chat_stream(self):
        """业财融合对话：带历史的多轮提问，SSE 流式返回，且用 PRO 模型 + 数据上下文。"""
        from unittest import mock
        from django.conf import settings
        self.mk(2026, 5, 200, 130)
        captured = {}

        def fake_raw(messages, tools=None, model=None, timeout=90, max_tokens=1800):
            captured['model'] = model
            captured['messages'] = messages
            return {'content': '根据数据，本月利润达标。'}

        with mock.patch('caiwu.views._deepseek_chat_raw', fake_raw):
            resp = self.client.post(
                '/api/cw/cockpit/ai-chat/stream',
                data=json.dumps({'year': 2026, 'month': 5, 'bu': self.bu, 'messages': [
                    {'role': 'user', 'content': '本月经营如何？'},
                    {'role': 'assistant', 'content': '收入200万。'},
                    {'role': 'user', 'content': '利润达标吗？'},
                ]}),
                content_type='application/json', **self.auth())
            self.assertEqual(resp.status_code, 200)
            self.assertEqual(resp['Content-Type'], 'text/event-stream')
            body = b''.join(resp.streaming_content).decode('utf-8')

        events = [json.loads(fr[5:].strip())
                  for fr in body.split('\n\n') if fr.strip().startswith('data:')]
        types = [e['type'] for e in events]
        self.assertEqual(types[0], 'meta')
        self.assertEqual(types[-1], 'done')
        answer = ''.join(e['delta'] for e in events if e['type'] == 'answer')
        self.assertEqual(answer, '根据数据，本月利润达标。')
        # 含 system 人设 + 数据上下文 + 完整对话历史（末条为用户提问）
        msgs = captured['messages']
        self.assertEqual(msgs[0]['role'], 'system')
        self.assertIn('经营数据上下文', msgs[1]['content'])
        self.assertEqual(msgs[-1]['role'], 'user')
        self.assertEqual(msgs[-1]['content'], '利润达标吗？')

    def test_cockpit_ai_chat_requires_user_question(self):
        self.mk(2026, 5, 200, 130)
        resp = self.client.post(
            '/api/cw/cockpit/ai-chat/stream',
            data=json.dumps({'year': 2026, 'month': 5, 'bu': self.bu, 'messages': [
                {'role': 'assistant', 'content': '历史回答'},
            ]}),
            content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 400)

    def test_cockpit_knowledge_crud_and_injection(self):
        """知识库：增/查/删，且被注入对话上下文（让助手越用越聪明）。"""
        from unittest import mock
        self.mk(2026, 5, 200, 130)
        r = self.client.post(
            '/api/cw/cockpit/knowledge',
            data=json.dumps({'content': '阔展大客户Q2流失，收入下滑属预期', 'scope': '全集团', 'kind': 'background'}),
            content_type='application/json', **self.auth())
        self.assertEqual(r.status_code, 200, r.content)
        kid = r.json()['data']['id']

        lst = self.client.get('/api/cw/cockpit/knowledge', **self.auth())
        self.assertTrue(any(k['id'] == kid for k in lst.json()['data']['items']))

        captured = {}

        def fake_raw(messages, tools=None, model=None, timeout=90, max_tokens=1800):
            captured['messages'] = messages
            return {'content': 'ok'}

        with mock.patch('caiwu.views._deepseek_chat_raw', fake_raw):
            resp = self.client.post(
                '/api/cw/cockpit/ai-chat/stream',
                data=json.dumps({'year': 2026, 'month': 5, 'bu': self.bu,
                                 'messages': [{'role': 'user', 'content': '背景如何'}]}),
                content_type='application/json', **self.auth())
            b''.join(resp.streaming_content)
        joined = '\n'.join(m['content'] for m in captured['messages'])
        self.assertIn('知识库', joined)
        self.assertIn('大客户Q2流失', joined)

        d = self.client.delete(f'/api/cw/cockpit/knowledge/{kid}', **self.auth())
        self.assertEqual(d.status_code, 200)

    def test_cockpit_knowledge_distill(self):
        """AI 自我提炼：把一段分析提炼成知识入库（来源标记 ai）。"""
        from unittest import mock

        def fake_chat(messages, timeout=90, model=None, max_tokens=1800):
            return '{"title":"应收风险","content":"逾期集中在大东，需加强催收。"}'

        with mock.patch('caiwu.views._deepseek_chat', fake_chat):
            r = self.client.post(
                '/api/cw/cockpit/knowledge/distill',
                data=json.dumps({'text': '大东逾期20万，账龄拉长……', 'scope': '全集团'}),
                content_type='application/json', **self.auth())
        self.assertEqual(r.status_code, 200, r.content)
        self.assertEqual(r.json()['data']['title'], '应收风险')
        self.assertEqual(r.json()['data']['source'], 'ai')

    def test_knowledge_import_text_raw(self):
        """文件导入（原文切块）：文本文件 → 知识条目。"""
        from django.core.files.uploadedfile import SimpleUploadedFile
        f = SimpleUploadedFile('notes.txt', '第一条经营背景说明\n第二条口径说明'.encode('utf-8'),
                               content_type='text/plain')
        r = self.client.post('/api/cw/cockpit/knowledge/import',
                             {'file': f, 'scope': '全集团', 'mode': 'raw'}, **self.auth())
        self.assertEqual(r.status_code, 200, r.content)
        self.assertGreaterEqual(r.json()['data']['created'], 1)

    def test_knowledge_import_distill(self):
        """文件导入（AI 提炼）：文档 → 多条知识。"""
        from unittest import mock
        from django.core.files.uploadedfile import SimpleUploadedFile

        def fake_chat(messages, timeout=90, model=None, max_tokens=1800):
            return '[{"title":"背景A","content":"要点A"},{"title":"背景B","content":"要点B"}]'

        f = SimpleUploadedFile('doc.md', '# 标题\n这里是一些经营文档内容'.encode('utf-8'))
        with mock.patch('caiwu.views._deepseek_chat', fake_chat):
            r = self.client.post('/api/cw/cockpit/knowledge/import',
                                 {'file': f, 'scope': '全集团', 'mode': 'distill'}, **self.auth())
        self.assertEqual(r.status_code, 200, r.content)
        self.assertEqual(r.json()['data']['created'], 2)
        items = self.client.get('/api/cw/cockpit/knowledge', **self.auth()).json()['data']['items']
        self.assertTrue(any(k['source'] == 'ai' and k['title'] == '背景A' for k in items))

    def test_chat_function_calling_tool_then_answer(self):
        """function-calling：模型先调用 save_knowledge 工具，再给出最终回答。"""
        from unittest import mock
        self.mk(2026, 5, 200, 130)
        calls = {'n': 0}

        def fake_raw(messages, tools=None, model=None, timeout=90, max_tokens=1800):
            calls['n'] += 1
            if calls['n'] == 1:
                return {'content': '', 'tool_calls': [{'id': 'c1', 'function': {
                    'name': 'save_knowledge',
                    'arguments': '{"content":"工具写入的知识Z","scope":"全集团"}'}}]}
            return {'content': '已记录并回答。'}

        with mock.patch('caiwu.views._deepseek_chat_raw', fake_raw):
            resp = self.client.post(
                '/api/cw/cockpit/ai-chat/stream',
                data=json.dumps({'year': 2026, 'month': 5, 'bu': self.bu,
                                 'messages': [{'role': 'user', 'content': '把这条记下来'}]}),
                content_type='application/json', **self.auth())
            body = b''.join(resp.streaming_content).decode('utf-8')
        events = [json.loads(fr[5:].strip()) for fr in body.split('\n\n') if fr.strip().startswith('data:')]
        self.assertIn('tool', [e['type'] for e in events])
        answer = ''.join(e['delta'] for e in events if e['type'] == 'answer')
        self.assertIn('已记录并回答', answer)
        items = self.client.get('/api/cw/cockpit/knowledge', **self.auth()).json()['data']['items']
        self.assertTrue(any(k['content'] == '工具写入的知识Z' for k in items))

    def test_chat_function_calling_generate_report(self):
        """function-calling：模型调用 generate_report（终止型技能），直接流式产出报告。"""
        from unittest import mock
        self.mk(2026, 5, 200, 130)

        def fake_raw(messages, tools=None, model=None, timeout=90, max_tokens=1800):
            return {'content': '', 'tool_calls': [{'id': 'r1', 'function': {
                'name': 'generate_report', 'arguments': '{"period":"month"}'}}]}

        def fake_stream(messages, model=None, max_tokens=1800, timeout=300):
            yield ('answer', '【正文】本月经营稳健。')

        with mock.patch('caiwu.views._deepseek_chat_raw', fake_raw), \
                mock.patch('caiwu.views._deepseek_stream', fake_stream):
            resp = self.client.post(
                '/api/cw/cockpit/ai-chat/stream',
                data=json.dumps({'year': 2026, 'month': 5, 'bu': self.bu,
                                 'messages': [{'role': 'user', 'content': '生成本月经营分析报告'}]}),
                content_type='application/json', **self.auth())
            body = b''.join(resp.streaming_content).decode('utf-8')
        events = [json.loads(fr[5:].strip()) for fr in body.split('\n\n') if fr.strip().startswith('data:')]
        answer = ''.join(e['delta'] for e in events if e['type'] == 'answer')
        self.assertIn('经营分析报告', answer)   # 技能注入的标题
        self.assertIn('本月经营稳健', answer)
        self.assertEqual(events[-1]['type'], 'done')

    def test_agent_skills_list_and_run(self):
        """Agent 技能：列表含基础技能，且可执行 写入/检索/清理 知识库。"""
        r = self.client.get('/api/cw/cockpit/skills', **self.auth())
        self.assertEqual(r.status_code, 200, r.content)
        names = [s['name'] for s in r.json()['data']['skills']]
        for n in ('save_knowledge', 'search_knowledge', 'forget_knowledge'):
            self.assertIn(n, names)

        run = self.client.post(
            '/api/cw/cockpit/skills/run',
            data=json.dumps({'name': 'save_knowledge', 'args': {'content': '技能写入的知识X', 'scope': '全集团'}}),
            content_type='application/json', **self.auth())
        self.assertEqual(run.status_code, 200, run.content)

        s = self.client.post(
            '/api/cw/cockpit/skills/run',
            data=json.dumps({'name': 'search_knowledge', 'args': {'query': '技能写入'}}),
            content_type='application/json', **self.auth())
        self.assertTrue(any('技能写入' in x['content'] for x in s.json()['data']))

        f = self.client.post(
            '/api/cw/cockpit/skills/run',
            data=json.dumps({'name': 'forget_knowledge', 'args': {'query': '技能写入'}}),
            content_type='application/json', **self.auth())
        self.assertGreaterEqual(f.json()['data']['deleted'], 1)

    def test_knowledge_import_dedup(self):
        """同范围内内容完全相同的导入不重复入库。"""
        from django.core.files.uploadedfile import SimpleUploadedFile
        payload = '同一条经营背景内容用于去重测试'
        for _ in range(2):
            f = SimpleUploadedFile('n.txt', payload.encode('utf-8'))
            self.client.post('/api/cw/cockpit/knowledge/import',
                             {'file': f, 'scope': '全集团', 'mode': 'raw'}, **self.auth())
        items = self.client.get('/api/cw/cockpit/knowledge', **self.auth()).json()['data']['items']
        self.assertEqual(sum(1 for k in items if k['content'] == payload), 1)

    def test_report_matrix_and_export(self):
        """月度矩阵报表：1月→最后已发布月，各月金额+合计；导出返回美化 xlsx。"""
        self.mk(2026, 1, 100, 60)
        self.mk(2026, 2, 200, 130)
        r = self.client.get('/api/cw/report/matrix',
                            {'year': 2026, 'bu': self.bu, 'level': 1}, **self.auth())
        self.assertEqual(r.status_code, 200, r.content)
        d = r.json()['data']
        self.assertEqual(d['months'], [1, 2])
        self.assertEqual(d['last_month'], 2)
        rev = next(x for x in d['rows'] if x['l1_name'] == REV)
        self.assertEqual(rev['values'], [100.0, 200.0])
        self.assertEqual(rev['total'], 300.0)
        # 计算行（经营净利）各月之和=合计
        net = next(x for x in d['rows'] if x['l1_name'] == NET_PROFIT)
        self.assertAlmostEqual(net['total'], net['values'][0] + net['values'][1], places=2)
        # 导出
        ex = self.client.get('/api/cw/report/export',
                             {'year': 2026, 'bu': self.bu, 'level': 1}, **self.auth())
        self.assertEqual(ex.status_code, 200, ex.content)
        self.assertIn('spreadsheet', ex['Content-Type'])

    def test_report_ai_stream_emits_answer_frames(self):
        from unittest import mock
        self.mk(2026, 5, 200, 130)

        def fake_stream(messages, model=None, max_tokens=1800, timeout=300):
            # 快模型只产出正文（无 reasoning_content）
            yield ('answer', '本月经营')
            yield ('answer', '稳健。')

        with mock.patch('caiwu.views._deepseek_stream', fake_stream):
            resp = self.client.post(
                '/api/cw/report/ai-analysis/stream',
                data=json.dumps({'year': 2026, 'month': 5, 'bus': [self.bu]}),
                content_type='application/json', **self.auth())
            self.assertEqual(resp.status_code, 200)
            self.assertEqual(resp['Content-Type'], 'text/event-stream')
            body = b''.join(resp.streaming_content).decode('utf-8')
        events = [json.loads(fr[5:].strip())
                  for fr in body.split('\n\n') if fr.strip().startswith('data:')]
        types = [e['type'] for e in events]
        self.assertEqual(types[0], 'meta')
        self.assertEqual(types[-1], 'done')
        answer = ''.join(e['delta'] for e in events if e['type'] == 'answer')
        self.assertEqual(answer, '本月经营稳健。')

    def test_cockpit_ai_stream_no_data_is_json_error(self):
        # 无数据时应在开流前以普通 JSON 错误返回，而非 event-stream
        from unittest import mock
        with mock.patch('caiwu.views._deepseek_stream') as m:
            resp = self.client.post(
                '/api/cw/cockpit/ai-analysis/stream',
                data=json.dumps({'year': 2024, 'month': 3}),
                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 400)
        self.assertNotEqual(resp['Content-Type'], 'text/event-stream')
        m.assert_not_called()
