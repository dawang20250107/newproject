import io
import json
import logging
import datetime
import functools
from decimal import Decimal, InvalidOperation
from urllib.parse import quote

from django.http import HttpResponse, JsonResponse, StreamingHttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from django.db.models import Sum, Q
from django.conf import settings
from django.utils import timezone
import jwt

from caiwu.models import (
    L1Category, L2Category, L3Category,
    ImportBatch, FinancialEntry, FinancialTarget, ProjectMargin, CockpitKnowledge,
    BUSINESS_UNITS, VALID_BUSINESS_UNITS, JOB_TITLES,
)
from paikuan.models import PaikuanUser, JobPermission as PaikuanJobPermission

logger = logging.getLogger('caiwu')

BUILD_VERSION = '2026-05-24.2'

EXCEL_HEADERS = ['一级科目', '二级项目部', '三级科目明细', '借方(元)', '贷方(元)']
IMPORT_SIZE_LIMIT = 5 * 1024 * 1024  # 5 MB

# Hardcoded KXT P&L calculation formulas.
# Keys are L1 category names; each lambda receives a name->float dict
# (built in sort_order so earlier results are available to later ones).
_CALC_FORMULAS = {
    '运营毛利': lambda m: (
        m.get('主营业务收入', 0) - m.get('主营业务成本', 0) - m.get('税金成本', 0)
    ),
    '经营毛利': lambda m: (
        m.get('运营毛利', 0)
        - m.get('销售费用', 0) - m.get('管理费用', 0) - m.get('财务费用', 0)
        + m.get('营业外收入', 0) - m.get('营业外支出', 0)
    ),
    '经营净利': lambda m: m.get('经营毛利', 0) - m.get('集团管理费用', 0),
}


def _compute_l1_name_map(l1_cats, raw_by_id):
    """
    Build name->float dict for all L1 categories (sorted by sort_order).
    Raw rows are filled from raw_by_id; calculated rows use _CALC_FORMULAS.
    Returns (name_map, id_map) where id_map is l1_id->float.
    """
    name_map = {}
    id_map = {}
    for l1 in l1_cats:
        if l1.is_calculated:
            formula = _CALC_FORMULAS.get(l1.name)
            if formula:
                val = formula(name_map)
                name_map[l1.name] = val
                id_map[l1.id] = val
        else:
            val = raw_by_id.get(l1.id)
            if val is not None:
                name_map[l1.name] = float(val)
                id_map[l1.id] = float(val)
    return name_map, id_map


# ── granular job-title permission system (mirrors paikuan) ────────────────────

# View-controllable report/chart elements. Each job title can be granted or
# denied visibility of each one (super_admin always sees everything).
PERM_FIELD_DEFS = [
    {'key': 'report_l1',       'label': '一级科目报表'},
    {'key': 'report_l2',       'label': '二级项目部明细'},
    {'key': 'report_l3',       'label': '三级科目明细'},
    {'key': 'amount',          'label': '金额数据'},
    {'key': 'chart_trend',     'label': '走势折线图'},
    {'key': 'chart_waterfall', 'label': '因素瀑布图'},
    {'key': 'export',          'label': '导出 Excel'},
]
FIELD_KEYS = [f['key'] for f in PERM_FIELD_DEFS]
PAGE_DEFS = [
    {'key': 'report',  'label': '财务报表'},
    {'key': 'data',    'label': '数据加工'},
    {'key': 'charts',  'label': '报表分析'},
    {'key': 'metrics', 'label': '指标管理'},
    {'key': 'cockpit', 'label': '财务驾驶舱'},
]
PAGE_KEYS = [p['key'] for p in PAGE_DEFS]


def _all_fields(value):
    return {k: value for k in FIELD_KEYS}


def _caiwu_perms_from_pk(pk):
    """Project paikuan's unified perms dict onto the caiwu-shaped subset."""
    from paikuan.views import CAIWU_FIELD_KEYS
    pk_pages = pk.get('pages', {})
    return {
        'pages': {
            'report':  bool(pk_pages.get('caiwu_report',  False)),
            'data':    bool(pk_pages.get('caiwu_data',    False)),
            'charts':  bool(pk_pages.get('caiwu_charts',  False)),
            'metrics': bool(pk_pages.get('caiwu_metrics', False)),
            'cockpit': bool(pk_pages.get('caiwu_cockpit', False)),
        },
        'view':        dict(pk.get('caiwu_view', {k: True for k in CAIWU_FIELD_KEYS})),
        'can_upload':  bool(pk.get('caiwu_upload',  False)),
        'can_publish': bool(pk.get('caiwu_publish', False)),
        'can_delete':  bool(pk.get('caiwu_delete',  False)),
    }


def default_job_config(job):
    """caiwu perm subset of paikuan's default job config (no stored overrides)."""
    from paikuan.views import default_job_config as pk_default
    return _caiwu_perms_from_pk(pk_default(job))


def _invalidate_perm_cache(job=None):
    """Caiwu keeps no perm cache of its own — it derives perms on demand from
    paikuan's cached get_job_perms. Kept as a no-op so cross-module callers and
    the standalone permission_detail endpoint don't break."""
    return None


def get_job_perms(job):
    """Effective caiwu config for a job title.

    Derived on demand from paikuan's get_job_perms (itself cached and
    invalidated whenever a JobPermission is edited). A single cache in paikuan
    avoids a second, separately-invalidated cache drifting out of sync.
    """
    from paikuan.views import get_job_perms as pk_get_job_perms
    return _caiwu_perms_from_pk(pk_get_job_perms(job))


def full_perms():
    return {'pages': {k: True for k in PAGE_KEYS}, 'view': _all_fields(True),
            'can_upload': True, 'can_publish': True, 'can_delete': True}


def effective_perms(user):
    """Perms object sent to the client. Works with PaikuanUser (super_admin gets full access)."""
    cfg = full_perms() if user.role == 'super_admin' else get_job_perms(user.job_title)
    return {**cfg, 'is_admin': user.role == 'super_admin',
            'fields': PERM_FIELD_DEFS, 'pages_meta': PAGE_DEFS}


def get_request_perms(request):
    """Effective perms for the authed request. None means full access (super_admin)."""
    if request.pk_role == 'super_admin':
        return None
    jt = getattr(request, 'pk_job', '') or ''
    return get_job_perms(jt)


def _can_view(request, field_key):
    p = get_request_perms(request)
    return True if p is None else bool(p['view'].get(field_key, True))


def _page_denied(request, page):
    """Return an err() response if the request's job lacks access to a page, else None."""
    p = get_request_perms(request)
    if p is not None and not p['pages'].get(page, True):
        return err('无访问权限', 403, 403)
    return None


# ── helpers ─────────────────────────────────────────────────────────────────

def ok(data=None, **extra):
    body = {'code': 0}
    if data is not None:
        body['data'] = data
    body.update(extra)
    return JsonResponse(body)


def err(msg, status=400, code=400):
    return JsonResponse({'code': code, 'error': msg}, status=status)


def _parse_json(request):
    try:
        return json.loads(request.body)
    except Exception:
        return {}


def _make_token(user):
    payload = {
        'uid': user.id,
        'role': user.role,
        'job': user.job_title,
        'departments': user.departments,
        'exp': datetime.datetime.utcnow() + datetime.timedelta(hours=24),
        'iat': datetime.datetime.utcnow(),
    }
    return jwt.encode(payload, settings.JWT_SECRET, algorithm='HS256')


def _build_excel_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    ascii_fallback = filename.encode('ascii', 'ignore').decode() or 'export.xlsx'
    encoded = quote(filename.encode('utf-8'), safe='')
    response['Content-Disposition'] = (
        f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{encoded}"
    )
    return response


# ── auth decorator ───────────────────────────────────────────────────────────

def cw_required(roles=None):
    """Auth decorator backed by PaikuanUser (Stage 2+3 unified accounts)."""
    def decorator(func):
        @functools.wraps(func)
        @csrf_exempt
        def wrapper(request, *args, **kwargs):
            token = request.headers.get('Authorization', '').replace('Bearer ', '').strip()
            if not token:
                return err('未认证', 401, 401)
            try:
                payload = jwt.decode(token, settings.JWT_SECRET, algorithms=['HS256'])
                uid = payload['uid']
            except jwt.ExpiredSignatureError:
                return err('Token已过期', 401, 401)
            except jwt.InvalidTokenError:
                return err('Token无效', 401, 401)
            user = PaikuanUser.objects.filter(id=uid).first()
            if not user:
                return err('用户不存在', 401, 401)
            if not user.is_active:
                return err('账号已停用', 401, 401)
            if not user.is_approved:
                return err('账号待审批，请联系管理员', 403, 403)
            request.pk_user = user
            request.pk_uid = user.id
            request.pk_role = user.role
            request.pk_job = user.job_title
            request.pk_depts = user.departments or []
            if roles and request.pk_role not in roles:
                return err('权限不足', 403, 403)
            return func(request, *args, **kwargs)
        return wrapper
    return decorator


def _bu_filter(qs, request):
    """Row visibility: super_admin sees all; everyone else sees assigned 事业部."""
    if request.pk_role == 'super_admin':
        return qs
    return qs.filter(business_unit__in=request.pk_depts)


def _can_access_bu(request, bu):
    if request.pk_role == 'super_admin':
        return True
    return bu in request.pk_depts


def _can_upload(request):
    p = get_request_perms(request)
    return True if p is None else bool(p.get('can_upload'))


def _can_publish(request):
    p = get_request_perms(request)
    return True if p is None else bool(p.get('can_publish'))


def _can_delete(request):
    p = get_request_perms(request)
    return True if p is None else bool(p.get('can_delete'))


# ── auth views ───────────────────────────────────────────────────────────────

@csrf_exempt
def register(request):
    """Disabled — account management moved to paikuan platform."""
    return err('财务分析账号已与排款平台统一，请通过排款平台登录', 410, 410)


@csrf_exempt
def login(request):
    """Disabled — use paikuan platform login (/api/pk/login)."""
    return err('财务分析账号已与排款平台统一，请通过排款平台登录', 410, 410)


@csrf_exempt
def registration_status(request):
    """Disabled — use paikuan platform."""
    return err('财务分析账号已与排款平台统一', 410, 410)


@cw_required()
def me(request):
    if request.method != 'GET':
        return err('方法不允许', 405)
    user = PaikuanUser.objects.filter(id=request.pk_uid).first()
    if not user:
        return err('用户不存在', 404)
    token = _make_token(user)
    return ok({'token': token, 'user': user.to_dict(), 'permissions': effective_perms(user)})


# ── user management ──────────────────────────────────────────────────────────

@cw_required(roles=['super_admin'])
def users(request):
    """User management moved to paikuan platform."""
    return err('账号管理已统一至排款平台，请通过排款平台管理用户', 410, 410)


@cw_required(roles=['super_admin'])
def user_detail(request, uid):
    return err('账号管理已统一至排款平台', 410, 410)


@cw_required(roles=['super_admin'])
def user_approve(request, uid):
    return err('账号管理已统一至排款平台', 410, 410)


@cw_required(roles=['super_admin'])
def user_reject(request, uid):
    return err('账号管理已统一至排款平台', 410, 410)


# ── job-title permission config (super_admin) ────────────────────────────────

@cw_required(roles=['super_admin'])
def permissions(request):
    """GET: full caiwu permission matrix (sourced from paikuan JobPermission)."""
    if request.method != 'GET':
        return err('方法不允许', 405)
    jobs = [
        {'job_title': key, 'label': label, 'config': get_job_perms(key)}
        for key, label in JOB_TITLES.items()
    ]
    return ok({'fields': PERM_FIELD_DEFS, 'pages': PAGE_DEFS, 'jobs': jobs})


@cw_required(roles=['super_admin'])
def permission_detail(request, job):
    """PUT: update caiwu-specific keys inside paikuan's JobPermission."""
    if request.method != 'PUT':
        return err('方法不允许', 405)
    if job not in JOB_TITLES:
        return err('职务无效', 404)
    body = _parse_json(request)
    cfg = body.get('config') or {}
    obj, _ = PaikuanJobPermission.objects.get_or_create(job_title=job)
    existing = dict(obj.config or {})
    # Write caiwu-specific fields into paikuan's config namespace
    existing['caiwu_view'] = {k: bool(cfg.get('view', {}).get(k, True)) for k in FIELD_KEYS}
    existing.setdefault('pages', {})
    existing['pages']['caiwu_report'] = bool(cfg.get('pages', {}).get('report', True))
    existing['pages']['caiwu_data']   = bool(cfg.get('pages', {}).get('data',   True))
    existing['pages']['caiwu_charts'] = bool(cfg.get('pages', {}).get('charts', True))
    existing['caiwu_upload']  = bool(cfg.get('can_upload', False))
    existing['caiwu_publish'] = bool(cfg.get('can_publish', False))
    existing['caiwu_delete']  = bool(cfg.get('can_delete', False))
    obj.config = existing
    obj.save()
    # Perms live in paikuan's JobPermission + cache (single source of truth).
    from paikuan.views import _invalidate_perm_cache as _pk_invalidate
    _pk_invalidate(job)
    return ok({'job_title': job, 'config': get_job_perms(job)})


# ── category management ──────────────────────────────────────────────────────

@cw_required(roles=['super_admin'])
def categories_l1(request):
    if request.method == 'GET':
        return ok([c.to_dict() for c in L1Category.objects.all()])
    if request.method == 'POST':
        body = _parse_json(request)
        name = (body.get('name') or '').strip()
        if not name:
            return err('科目名称不能为空')
        if L1Category.objects.filter(name=name).exists():
            return err('该科目已存在')
        sign = int(body.get('sign', 1))
        if sign not in (1, -1):
            sign = 1
        cat = L1Category.objects.create(
            name=name,
            sort_order=body.get('sort_order', 0),
            is_profit_driver=bool(body.get('is_profit_driver', False)),
            is_calculated=bool(body.get('is_calculated', False)),
            sign=sign,
        )
        return ok(cat.to_dict())
    return err('方法不允许', 405)


@cw_required(roles=['super_admin'])
def category_l1_detail(request, cid):
    try:
        cat = L1Category.objects.get(id=cid)
    except L1Category.DoesNotExist:
        return err('科目不存在', 404)
    if request.method == 'PUT':
        body = _parse_json(request)
        name = (body.get('name') or '').strip()
        if not name:
            return err('科目名称不能为空')
        if L1Category.objects.filter(name=name).exclude(id=cid).exists():
            return err('该名称已被其他科目使用')
        sign = int(body.get('sign', cat.sign))
        if sign not in (1, -1):
            sign = cat.sign
        cat.name = name
        cat.sort_order = body.get('sort_order', cat.sort_order)
        cat.is_profit_driver = bool(body.get('is_profit_driver', cat.is_profit_driver))
        cat.is_calculated = bool(body.get('is_calculated', cat.is_calculated))
        cat.sign = sign
        cat.save()
        return ok(cat.to_dict())
    if request.method == 'DELETE':
        if cat.entries.exists():
            return err('该科目已有数据，无法删除')
        cat.delete()
        return ok({'deleted': cid})
    return err('方法不允许', 405)


@cw_required(roles=['super_admin'])
def categories_l2(request):
    bu = request.GET.get('bu', '')
    if request.method == 'GET':
        qs = L2Category.objects.all()
        if bu:
            qs = qs.filter(business_unit=bu)
        return ok([c.to_dict() for c in qs])
    if request.method == 'POST':
        body = _parse_json(request)
        bu = (body.get('business_unit') or bu or '').strip()
        name = (body.get('name') or '').strip()
        if not bu or not name:
            return err('事业部和项目部名称不能为空')
        if bu not in VALID_BUSINESS_UNITS:
            return err('无效的事业部')
        if L2Category.objects.filter(business_unit=bu, name=name).exists():
            return err('该项目部已存在')
        cat = L2Category.objects.create(
            business_unit=bu, name=name,
            sort_order=body.get('sort_order', 0),
        )
        return ok(cat.to_dict())
    return err('方法不允许', 405)


@cw_required(roles=['super_admin'])
def category_l2_detail(request, cid):
    try:
        cat = L2Category.objects.get(id=cid)
    except L2Category.DoesNotExist:
        return err('项目部不存在', 404)
    if request.method == 'PUT':
        body = _parse_json(request)
        name = (body.get('name') or '').strip()
        if not name:
            return err('名称不能为空')
        if L2Category.objects.filter(business_unit=cat.business_unit, name=name).exclude(id=cid).exists():
            return err('该名称已存在')
        cat.name = name
        cat.sort_order = body.get('sort_order', cat.sort_order)
        cat.save()
        return ok(cat.to_dict())
    if request.method == 'DELETE':
        if cat.entries.exists():
            return err('该项目部已有数据，无法删除')
        cat.delete()
        return ok({'deleted': cid})
    return err('方法不允许', 405)


@cw_required(roles=['super_admin'])
def categories_l3(request):
    bu = request.GET.get('bu', '')
    if request.method == 'GET':
        qs = L3Category.objects.select_related('l1_category').all()
        if bu:
            qs = qs.filter(business_unit=bu)
        return ok([c.to_dict() for c in qs])
    if request.method == 'POST':
        body = _parse_json(request)
        bu = (body.get('business_unit') or bu or '').strip()
        name = (body.get('name') or '').strip()
        l1_id = body.get('l1_category_id')
        if not bu or not name:
            return err('事业部和科目明细不能为空')
        if bu not in VALID_BUSINESS_UNITS:
            return err('无效的事业部')
        l1 = None
        if l1_id:
            try:
                l1 = L1Category.objects.get(id=l1_id)
            except L1Category.DoesNotExist:
                return err('一级科目不存在')
        if L3Category.objects.filter(business_unit=bu, l1_category=l1, name=name).exists():
            return err('该科目明细已存在')
        cat = L3Category.objects.create(
            business_unit=bu, l1_category=l1, name=name,
            sort_order=body.get('sort_order', 0),
            kingdee_code=(body.get('kingdee_code') or '').strip(),
        )
        return ok(cat.to_dict())
    return err('方法不允许', 405)


@cw_required(roles=['super_admin'])
def category_l3_detail(request, cid):
    try:
        cat = L3Category.objects.get(id=cid)
    except L3Category.DoesNotExist:
        return err('科目明细不存在', 404)
    if request.method == 'PUT':
        body = _parse_json(request)
        name = (body.get('name') or '').strip()
        if not name:
            return err('名称不能为空')
        l1_id = body.get('l1_category_id', cat.l1_category_id)
        l1 = None
        if l1_id:
            try:
                l1 = L1Category.objects.get(id=l1_id)
            except L1Category.DoesNotExist:
                return err('一级科目不存在')
        if L3Category.objects.filter(
            business_unit=cat.business_unit, l1_category=l1, name=name,
        ).exclude(id=cid).exists():
            return err('该名称已存在')
        cat.name = name
        cat.l1_category = l1
        cat.sort_order = body.get('sort_order', cat.sort_order)
        cat.kingdee_code = (body.get('kingdee_code') or '').strip()
        cat.save()
        return ok(cat.to_dict())
    if request.method == 'DELETE':
        if cat.entries.exists():
            return err('该科目已有数据，无法删除')
        cat.delete()
        return ok({'deleted': cid})
    return err('方法不允许', 405)


# ── batch / import ───────────────────────────────────────────────────────────

@cw_required()
def batches(request):
    if request.method != 'GET':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'data')
    if denied:
        return denied
    qs = _bu_filter(ImportBatch.objects.all(), request)
    year = request.GET.get('year')
    month = request.GET.get('month')
    status = request.GET.get('status')
    if year:
        qs = qs.filter(year=year)
    if month:
        qs = qs.filter(month=month)
    if status:
        qs = qs.filter(status=status)
    return ok([b.to_dict() for b in qs[:200]])


@cw_required()
def batch_template(request):
    if request.method != 'GET':
        return err('方法不允许', 405)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return err('服务器缺少 openpyxl 依赖')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '财务数据导入模板'

    header_fill = PatternFill(start_color='C96342', end_color='C96342', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Example rows (revenue + cost)
    ws.cell(row=2, column=1, value='主营业务收入')
    ws.cell(row=2, column=2, value='甲项目部（可选）')
    ws.cell(row=2, column=3, value='（可选）')
    ws.cell(row=2, column=4, value=0)         # 借方
    ws.cell(row=2, column=5, value=1000000.00)  # 贷方（收入在贷方）
    ws.cell(row=3, column=1, value='主营业务成本')
    ws.cell(row=3, column=2, value='甲项目部（可选）')
    ws.cell(row=3, column=3, value='（可选）')
    ws.cell(row=3, column=4, value=600000.00)  # 借方（成本在借方）
    ws.cell(row=3, column=5, value=0)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    # Instruction sheet
    ws2 = wb.create_sheet('填写说明')
    instructions = [
        ['列名', '说明', '是否必填'],
        ['一级科目', '必须与系统中已配置的一级科目名称完全一致', '必填'],
        ['二级项目部', '该事业部下的项目部名称，系统会自动创建不存在的项目部', '选填'],
        ['三级科目明细', '成本费用科目明细，系统会自动创建不存在的明细', '选填'],
        ['借方(元)', '会计借方发生额（成本费用在此列）', '二选一'],
        ['贷方(元)', '会计贷方发生额（收入在此列）', '二选一'],
        ['', '', ''],
        ['注意：借贷方向说明', '', ''],
        ['收入类科目（主营业务收入等）', '贷方填金额，借方填0', ''],
        ['成本费用类科目', '借方填金额，贷方填0', ''],
        ['红字冲销（负数）', '直接在对应列填负数', ''],
        ['', '', ''],
        ['注意事项', '', ''],
        ['1. 上传时须指定事业部、年份、月份', '', ''],
        ['2. 同一事业部同月份只能有一个已发布的数据批次', '', ''],
        ['3. 发布后旧数据自动归档，新数据生效', '', ''],
    ]
    for row in instructions:
        ws2.append(row)
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 40
    ws2.column_dimensions['C'].width = 10

    return _build_excel_response(wb, '财务数据导入模板.xlsx')


# ── Kingdee 部门明细表 format detection ──────────────────────────────────────

_KD_ACCT_KW   = ['科目名称', '科目全称', '一级科目名称', '一级科目']
_KD_DEPT_KW   = ['部门']
_KD_DEBIT_KW  = ['本期借方', '借方发生额', '本期发生(借', '发生额(借', '本月发生额(借']
_KD_CREDIT_KW = ['本期贷方', '贷方发生额', '本期发生(贷', '发生额(贷', '本月发生额(贷']
_KD_AMT_KW    = ['本期金额', '本月发生额', '发生额合计', '本年金额']
_KD_L3_KW     = ['明细科目', '三级科目', '科目明细', '子目', '小科目']
_KD_SKIP_KW   = ['合计', '小计', '总计', '期末', '期初', '余额', '科目代码', '科目编码', '编码']
# 核算维度明细账中的小计/汇总行摘要（非业务发生额，解析时跳过）
_LEDGER_SUMMARY_ROWS = {'期初余额', '本期合计', '本年累计', '期末余额'}

# ── 金蝶科目编码 → KXT 一级科目映射（核算维度明细账 / 部门明细表）──────────────
# 按科目编码前缀（第一段）归类到 KXT 一级科目
_KD_CODE_L1 = {
    '6001': '主营业务收入', '6051': '主营业务收入',
    '6301': '营业外收入',
    '6401': '主营业务成本', '6402': '主营业务成本',
    '6403': '税金成本',
    '6601': '销售费用',
    '6602': '管理费用',
    '6603': '财务费用',
    '6711': '营业外支出',
}
# 科目名称精确包含覆盖（优先于编码前缀）——集团管理费用在金蝶里挂在 6602 管理费用下，
# 但 KXT 体系中需单列以正确推算经营净利。
_KD_NAME_L1 = {
    '集团管理费用': '集团管理费用',
}
# 多段科目编码覆盖（优先于单段编码前缀）——6602.99 及其子明细属集团管理费，
# 金蝶挂在 6602 管理费用下，KXT 体系需单列以正确推算经营净利。
_KD_CODE_L1_SPECIFIC = {
    '6602.99': '集团管理费用',
}


def _detect_kingdee_format(ws):
    """Scan first 12 rows for Kingdee-style headers.
    Returns (data_start_row, col_map) or (None, {}).
    col_map keys: account, dept, debit(opt), credit(opt), amount(opt), l3(opt)
    """
    def hit(val, kws):
        return any(k in val for k in kws)

    for ri in range(1, min(13, ws.max_row + 1)):
        cm = {}
        for ci in range(1, min(ws.max_column + 1, 40)):
            v = str(ws.cell(row=ri, column=ci).value or '').strip()
            if not v:
                continue
            if 'account' not in cm and hit(v, _KD_ACCT_KW):
                cm['account'] = ci
            elif 'dept' not in cm and hit(v, _KD_DEPT_KW):
                cm['dept'] = ci
            elif 'debit' not in cm and hit(v, _KD_DEBIT_KW):
                cm['debit'] = ci
            elif 'credit' not in cm and hit(v, _KD_CREDIT_KW):
                cm['credit'] = ci
            elif 'amount' not in cm and hit(v, _KD_AMT_KW):
                cm['amount'] = ci
            elif 'l3' not in cm and hit(v, _KD_L3_KW):
                cm['l3'] = ci
        has_amount = ('debit' in cm and 'credit' in cm) or 'amount' in cm
        if 'account' in cm and 'dept' in cm and has_amount:
            return ri + 1, cm
    return None, {}


def _parse_kingdee_rows(ws, data_start, cm, bu, l1_map, l2_map, l3_map):
    """Parse rows using detected Kingdee column map. Returns (parsed_rows, errors)."""
    errors = []
    parsed = []

    for ri in range(data_start, ws.max_row + 1):
        def cv(col):
            return str(ws.cell(row=ri, column=col).value or '').strip()

        def cn(col):
            try:
                return Decimal(str(ws.cell(row=ri, column=col).value or 0))
            except (InvalidOperation, TypeError):
                return Decimal(0)

        acct = cv(cm['account'])
        dept = cv(cm['dept'])

        if not acct and not dept:
            continue
        if not acct or any(k in acct for k in _KD_SKIP_KW):
            continue

        # Accounting direction: revenue (sign=+1) grows on credit; cost (sign=-1) grows on debit.
        # Formula: sign * (credit − debit) yields correct positive magnitude for normal entries
        # and naturally handles 红字 (negative debit/credit reversal entries).
        # Zero-amount rows are skipped below — do NOT abs() here so reversals reduce totals.
        if 'debit' in cm and 'credit' in cm:
            raw_net = cn(cm['credit']) - cn(cm['debit'])
        else:
            raw_net = cn(cm['amount'])

        if raw_net == 0:
            continue

        # Match L1: exact first, then partial (name-in-cell or cell-in-name)
        l1 = l1_map.get(acct)
        matched_name = acct
        if not l1:
            for name, cat in l1_map.items():
                if name in acct or acct in name:
                    l1 = cat
                    matched_name = name
                    break
        if not l1:
            errors.append(f'第{ri}行：科目"{acct}"未匹配到一级科目（已跳过）')
            continue
        if l1.is_calculated:
            continue

        l2 = None
        if dept:
            if dept not in l2_map:
                l2_obj = L2Category(business_unit=bu, name=dept, sort_order=len(l2_map))
                l2_obj.save()
                l2_map[dept] = l2_obj
            l2 = l2_map[dept]

        l3 = None
        l3_name = cv(cm['l3']) if 'l3' in cm else ''
        if l3_name:
            key = (l1.id, l3_name)
            if key not in l3_map:
                l3_obj = L3Category(
                    business_unit=bu, l1_category=l1,
                    name=l3_name, sort_order=len(l3_map),
                )
                l3_obj.save()
                l3_map[key] = l3_obj
            l3 = l3_map[key]

        # Apply accounting sign: revenue sign=+1 → amount=credit-debit; cost sign=-1 → amount=debit-credit
        amount = Decimal(str(l1.sign)) * raw_net

        parsed.append({
            'l1': l1, 'l2': l2, 'l3': l3, 'amount': amount,
            'l1_name': matched_name, 'l2_name': dept, 'l3_name': l3_name,
        })

    return parsed, errors


# ── 金蝶 核算维度明细账（部门明细表）解析 ─────────────────────────────────────

def _detect_dept_ledger(ws):
    """Detect Kingdee 核算维度明细账 layout. Returns (data_start, col_map) or (None, {}).
    col_map keys: dept, code, name, debit, credit, summary
    """
    for ri in range(1, min(8, ws.max_row + 1)):
        cm = {}
        for ci in range(1, min(ws.max_column + 1, 20)):
            v = str(ws.cell(row=ri, column=ci).value or '').strip()
            if v == '部门名称':
                cm['dept'] = ci
            elif v == '科目编码':
                cm['code'] = ci
            elif v == '科目名称':
                cm['name'] = ci
            elif v == '摘要':
                cm['summary'] = ci
            elif v == '借方':
                cm['debit'] = ci
            elif v == '贷方':
                cm['credit'] = ci
        if all(k in cm for k in ('dept', 'name', 'debit', 'credit', 'summary')):
            return ri + 1, cm
    return None, {}


# ── 项目核算明细账（按项目维度）→ 项目毛利 ─────────────────────────────────────
# 金蝶科目编码前缀 → 项目毛利分类。收入类按 (贷-借)，成本/费用类按 (借-贷)。
_PM_CAT_BY_PREFIX = {
    '6001': 'revenue', '6051': 'revenue',
    '6401': 'cost', '6402': 'cost',
    '6601': 'sales_exp',
    '6602': 'mgmt_exp',
}
# 未挂项目池：金蝶里项目维度为「无」或空
_PM_UNALLOCATED = {'无', '', '（无）', '(无)', '未指定'}


def _detect_project_ledger(ws):
    """识别金蝶「核算维度明细账（按项目）」：维度列为「项目名称」。
    返回 (data_start, col_map{project,code,name,summary,debit,credit}) 或 (None, {})。"""
    for ri in range(1, min(8, ws.max_row + 1)):
        cm = {}
        for ci in range(1, min(ws.max_column + 1, 20)):
            v = str(ws.cell(row=ri, column=ci).value or '').strip()
            if v == '项目名称':
                cm['project'] = ci
            elif v == '科目编码':
                cm['code'] = ci
            elif v == '科目名称':
                cm['name'] = ci
            elif v == '摘要':
                cm['summary'] = ci
            elif v == '借方':
                cm['debit'] = ci
            elif v == '贷方':
                cm['credit'] = ci
        if all(k in cm for k in ('project', 'code', 'debit', 'credit', 'summary')):
            return ri + 1, cm
    return None, {}


def _parse_project_ledger(ws, data_start, cm):
    """汇总项目核算明细账 → {project_name: {revenue, cost, sales_exp, mgmt_exp}}。
    跳过 期初/合计/累计 等小计行与「结转损益」；按科目前缀归类、按方向取净额。"""
    def _dec(ri, col):
        try:
            return Decimal(str(ws.cell(row=ri, column=col).value or 0))
        except (InvalidOperation, TypeError):
            return Decimal(0)

    agg = {}
    for ri in range(data_start, ws.max_row + 1):
        summ = str(ws.cell(row=ri, column=cm['summary']).value or '').strip()
        if summ in _LEDGER_SUMMARY_ROWS or '结转损益' in summ:
            continue
        code = str(ws.cell(row=ri, column=cm['code']).value or '').strip()
        if not code:
            continue
        cat = _PM_CAT_BY_PREFIX.get(code.split('.')[0])
        if not cat:
            continue
        project = str(ws.cell(row=ri, column=cm['project']).value or '').strip()
        debit = _dec(ri, cm['debit'])
        credit = _dec(ri, cm['credit'])
        net = (credit - debit) if cat == 'revenue' else (debit - credit)
        bucket = agg.setdefault(project, {'revenue': Decimal('0'), 'cost': Decimal('0'),
                                          'sales_exp': Decimal('0'), 'mgmt_exp': Decimal('0')})
        bucket[cat] += net
    return agg


def _map_dept_ledger_l1(code, name, l1_map):
    """Map a 部门明细账 row to an L1 category. 科目名称 override → 编码前缀。
    Returns (L1Category|None, matched_name|None).
    """
    for kw, l1name in _KD_NAME_L1.items():
        if kw in name:
            return l1_map.get(l1name), l1name
    code_s = str(code).strip()
    # 多段编码精确/前缀匹配优先（如 6602.99 集团管理费用及其子明细 6602.99.xx）
    for spec, l1name in _KD_CODE_L1_SPECIFIC.items():
        if code_s == spec or code_s.startswith(spec + '.'):
            return l1_map.get(l1name), l1name
    prefix = code_s.split('.')[0]
    l1name = _KD_CODE_L1.get(prefix)
    if l1name:
        return l1_map.get(l1name), l1name
    return None, None


def _parse_dept_ledger_rows(ws, data_start, cm, bu, l1_map, l2_map, l3_map):
    """Parse Kingdee 核算维度明细账 by aggregating business transaction rows per account.

    Closed-period exports embed a 「结转损益」 carry-forward entry that reverses each
    account into the P&L, so the 「本期合计」 subtotal nets to ZERO for every account.
    We therefore sum the individual transaction rows and EXCLUDE 结转损益 (and the
    期初/合计/累计 summary rows) to recover the real period movement.
    部门名称 → L2, 科目名称 → L3, 编码前缀 → L1. Returns (parsed_rows, errors).
    """
    errors = []

    def _dec(ri, col):
        try:
            return Decimal(str(ws.cell(row=ri, column=col).value or 0))
        except (InvalidOperation, TypeError):
            return Decimal(0)

    # Sum (credit - debit) of business rows per (dept, code, name), preserving order.
    agg = {}
    order = []
    for ri in range(data_start, ws.max_row + 1):
        summ = str(ws.cell(row=ri, column=cm['summary']).value or '').strip()
        if summ in _LEDGER_SUMMARY_ROWS or '结转损益' in summ:
            continue
        code = str(ws.cell(row=ri, column=cm['code']).value or '').strip() if 'code' in cm else ''
        name = str(ws.cell(row=ri, column=cm['name']).value or '').strip()
        if not code and not name:
            continue
        dept = str(ws.cell(row=ri, column=cm['dept']).value or '').strip()
        key = (dept, code, name)
        if key not in agg:
            agg[key] = Decimal(0)
            order.append(key)
        agg[key] += _dec(ri, cm['credit']) - _dec(ri, cm['debit'])

    parsed = []
    for dept, code, name in order:
        raw_net = agg[(dept, code, name)]
        if raw_net == 0:
            continue

        l1, l1name = _map_dept_ledger_l1(code, name, l1_map)
        if not l1:
            errors.append(f'科目"{code} {name}"未匹配到一级科目（已跳过）')
            continue
        if l1.is_calculated:
            continue
        amount = Decimal(str(l1.sign)) * raw_net

        l2 = None
        if dept:
            if dept not in l2_map:
                obj = L2Category(business_unit=bu, name=dept, sort_order=len(l2_map))
                obj.save()
                l2_map[dept] = obj
            l2 = l2_map[dept]

        l3 = None
        if name:
            key = (l1.id, name)
            if key not in l3_map:
                obj = L3Category(business_unit=bu, l1_category=l1, name=name, sort_order=len(l3_map))
                obj.save()
                l3_map[key] = obj
            l3 = l3_map[key]

        parsed.append({
            'l1': l1, 'l2': l2, 'l3': l3, 'amount': amount,
            'l1_name': l1name, 'l2_name': dept, 'l3_name': name,
        })

    return parsed, errors


def _parse_template_rows(ws, bu, l1_map, l2_map, l3_map):
    """Parse the KXT 5-column template format (L1 / L2 / L3 / 借方 / 贷方).
    Also accepts the legacy 4-column format (L1 / L2 / L3 / 金额) for backward compatibility.
    """
    errors = []
    parsed = []

    # Detect old 4-column vs new 5-column format from header row
    h4 = str(ws.cell(row=1, column=4).value or '').strip()
    h5 = str(ws.cell(row=1, column=5).value or '').strip()
    five_col = bool(h5)  # if column 5 has a header it's the new format

    def _to_dec(val):
        try:
            return Decimal(str(val))
        except (InvalidOperation, TypeError):
            return Decimal(0)

    for ri in range(2, ws.max_row + 1):
        l1_name = str(ws.cell(row=ri, column=1).value or '').strip()
        l2_name = str(ws.cell(row=ri, column=2).value or '').strip()
        l3_name = str(ws.cell(row=ri, column=3).value or '').strip()

        if five_col:
            debit_raw = ws.cell(row=ri, column=4).value
            credit_raw = ws.cell(row=ri, column=5).value
            if not l1_name and debit_raw is None and credit_raw is None:
                continue
        else:
            amt_raw = ws.cell(row=ri, column=4).value
            if not l1_name and amt_raw is None:
                continue

        if not l1_name:
            errors.append(f'第{ri}行：一级科目不能为空')
            continue
        if l1_name not in l1_map:
            errors.append(f'第{ri}行：一级科目"{l1_name}"不存在，请先在设置中添加')
            continue
        l1 = l1_map[l1_name]
        if l1.is_calculated:
            continue

        # Compute amount using accounting direction (same as Kingdee parser)
        if five_col:
            debit = _to_dec(debit_raw)
            credit = _to_dec(credit_raw)
            raw_net = credit - debit
        else:
            # Legacy format: positive amount supplied directly (no sign adjustment needed)
            try:
                raw_net = Decimal(str(ws.cell(row=ri, column=4).value))
                # For backward compat: assume amount already represents the correct magnitude
                # (revenue positive, cost positive). Convert to sign * (credit-debit) space:
                # amount = sign * raw_net / sign = raw_net, so just use raw_net directly.
                # We still apply sign so totals are consistent: sign * raw_net / sign = raw_net / 1
                # Actually for legacy we store raw_net directly without sign adjustment.
                amount = raw_net
            except (InvalidOperation, TypeError):
                errors.append(f'第{ri}行：金额格式错误')
                continue
            l2 = None
            l3 = None
            if l2_name:
                if l2_name not in l2_map:
                    l2_obj = L2Category(business_unit=bu, name=l2_name, sort_order=len(l2_map))
                    l2_obj.save()
                    l2_map[l2_name] = l2_obj
                l2 = l2_map[l2_name]
            if l3_name:
                key = (l1.id, l3_name)
                if key not in l3_map:
                    l3_obj = L3Category(business_unit=bu, l1_category=l1, name=l3_name, sort_order=len(l3_map))
                    l3_obj.save()
                    l3_map[key] = l3_obj
                l3 = l3_map[key]
            parsed.append({'l1': l1, 'l2': l2, 'l3': l3, 'amount': amount,
                           'l1_name': l1_name, 'l2_name': l2_name, 'l3_name': l3_name})
            continue

        # Five-column path: apply accounting direction
        amount = Decimal(str(l1.sign)) * raw_net

        l2 = None
        l3 = None
        if l2_name:
            if l2_name not in l2_map:
                l2_obj = L2Category(business_unit=bu, name=l2_name, sort_order=len(l2_map))
                l2_obj.save()
                l2_map[l2_name] = l2_obj
            l2 = l2_map[l2_name]
        if l3_name:
            key = (l1.id, l3_name)
            if key not in l3_map:
                l3_obj = L3Category(business_unit=bu, l1_category=l1, name=l3_name, sort_order=len(l3_map))
                l3_obj.save()
                l3_map[key] = l3_obj
            l3 = l3_map[key]

        parsed.append({
            'l1': l1, 'l2': l2, 'l3': l3, 'amount': amount,
            'l1_name': l1_name, 'l2_name': l2_name, 'l3_name': l3_name,
        })

    return parsed, errors


def _parse_json_rows(data_str, bu, l1_map, l2_map, l3_map):
    """Parse JSON array import. Each element: {l1, l2?, l3?, debit?, credit?, amount?}.
    Returns (parsed_rows, errors).
    """
    errors = []
    parsed = []
    try:
        rows = json.loads(data_str)
        if not isinstance(rows, list):
            return [], ['JSON 格式错误：根节点需为数组']
    except json.JSONDecodeError as e:
        return [], [f'JSON 解析失败：{e}']

    def _to_dec(v):
        try:
            return Decimal(str(v))
        except (InvalidOperation, TypeError):
            return Decimal(0)

    for i, row in enumerate(rows, 1):
        l1_name = str(row.get('l1', '')).strip()
        l2_name = str(row.get('l2', '')).strip()
        l3_name = str(row.get('l3', '')).strip()

        if not l1_name:
            errors.append(f'第{i}项：缺少 l1 字段')
            continue
        l1 = l1_map.get(l1_name)
        if not l1:
            for k, cat in l1_map.items():
                if k in l1_name or l1_name in k:
                    l1 = cat
                    l1_name = k
                    break
        if not l1:
            errors.append(f'第{i}项：l1="{l1_name}" 未匹配到一级科目')
            continue
        if l1.is_calculated:
            continue

        if 'amount' in row:
            amount = _to_dec(row['amount'])
        else:
            debit = _to_dec(row.get('debit', 0))
            credit = _to_dec(row.get('credit', 0))
            amount = Decimal(str(l1.sign)) * (credit - debit)

        l2 = None
        if l2_name:
            if l2_name not in l2_map:
                obj = L2Category(business_unit=bu, name=l2_name, sort_order=len(l2_map))
                obj.save()
                l2_map[l2_name] = obj
            l2 = l2_map[l2_name]

        l3 = None
        if l3_name:
            key = (l1.id, l3_name)
            if key not in l3_map:
                obj = L3Category(business_unit=bu, l1_category=l1, name=l3_name, sort_order=len(l3_map))
                obj.save()
                l3_map[key] = obj
            l3 = l3_map[key]

        parsed.append({
            'l1': l1, 'l2': l2, 'l3': l3, 'amount': amount,
            'l1_name': l1_name, 'l2_name': l2_name, 'l3_name': l3_name,
        })

    return parsed, errors


def _compute_pl_check(parsed_rows):
    """Compute P&L KPIs + L1/L2 summaries from uploaded rows for reconciliation display."""
    from collections import defaultdict

    raw = defaultdict(Decimal)
    l2_totals = defaultdict(lambda: defaultdict(Decimal))
    for r in parsed_rows:
        raw[r['l1_name']] += r['amount']
        l2_totals[r['l1_name']][r['l2_name'] or '（无项目部）'] += r['amount']

    l1_cats = list(L1Category.objects.order_by('sort_order', 'id').all())
    name_map = {}
    l1_summary = []

    for l1 in l1_cats:
        if l1.is_calculated:
            formula = _CALC_FORMULAS.get(l1.name)
            val = float(formula(name_map)) if formula else 0.0
        else:
            val = float(raw.get(l1.name, 0))
        name_map[l1.name] = val
        if val != 0 or l1.is_calculated:
            l1_summary.append({
                'name': l1.name, 'amount': round(val, 2),
                'is_calculated': l1.is_calculated,
            })

    l2_summary = []
    for l1_name, depts in l2_totals.items():
        for dept_name, amt in sorted(depts.items()):
            if float(amt) != 0:
                l2_summary.append({
                    'l1_name': l1_name,
                    'l2_name': dept_name,
                    'amount': round(float(amt), 2),
                })

    KPI_ORDER = ['主营业务收入', '主营业务成本', '运营毛利', '经营毛利', '经营净利']
    nm = {r['name']: r for r in l1_summary}
    kpis = [nm[k] for k in KPI_ORDER if k in nm]

    return {'kpis': kpis, 'l1_summary': l1_summary, 'l2_summary': l2_summary}


@cw_required()
def batch_upload(request):
    if request.method != 'POST':
        return err('方法不允许', 405)
    if not _can_upload(request):
        return err('权限不足', 403)

    bu = request.POST.get('bu', '').strip()
    year_str = request.POST.get('year', '').strip()
    month_str = request.POST.get('month', '').strip()

    if not bu:
        return err('请指定事业部')
    if bu not in VALID_BUSINESS_UNITS:
        return err(f'无效的事业部：{bu}')
    if not _can_access_bu(request, bu):
        return err('您无权为该事业部上传数据', 403)

    try:
        year = int(year_str)
        month = int(month_str)
        assert 2000 <= year <= 2100 and 1 <= month <= 12
    except Exception:
        return err('年份或月份无效')

    f = request.FILES.get('file')
    if not f:
        return err('请上传文件')
    if f.size > IMPORT_SIZE_LIMIT:
        return err(f'文件超过5MB限制（当前{f.size // 1024 // 1024}MB）')

    l1_map = {c.name: c for c in L1Category.objects.order_by('sort_order', 'id')}
    if not l1_map:
        return err('系统尚未配置一级科目，请先在「设置」中添加一级科目')

    fname = f.name.lower()
    l2_map = {c.name: c for c in L2Category.objects.filter(business_unit=bu)}
    l3_map = {(c.l1_category_id, c.name): c for c in L3Category.objects.filter(business_unit=bu)}

    # ── JSON：部门明细（数组格式）──────────────────────────────────────────────
    if fname.endswith('.json'):
        data_str = f.read().decode('utf-8', errors='replace')
        parsed_rows, errors = _parse_json_rows(data_str, bu, l1_map, l2_map, l3_map)
        batch_type = ImportBatch.TYPE_DEPT
        fmt = 'json'

    # ── Excel：金蝶核算维度明细账 / KXT模板 = 部门明细 ─────────────────────────────
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
        except Exception:
            return err('文件格式错误，请使用Excel(.xlsx)格式')

        batch_type = ImportBatch.TYPE_DEPT
        ledger_start, ledger_cm = _detect_dept_ledger(ws)
        row1_vals = [str(ws.cell(row=1, column=c).value or '').strip() for c in range(1, 6)]
        is_kxt = (row1_vals[:5] == EXCEL_HEADERS or row1_vals[:4] == ['一级科目', '二级项目部', '三级科目明细', '金额(元)'])

        if ledger_start is not None:
            parsed_rows, errors = _parse_dept_ledger_rows(ws, ledger_start, ledger_cm, bu, l1_map, l2_map, l3_map)
            fmt = 'kingdee_ledger'
        elif is_kxt:
            parsed_rows, errors = _parse_template_rows(ws, bu, l1_map, l2_map, l3_map)
            fmt = 'template'
        else:
            data_start, cm = _detect_kingdee_format(ws)
            if data_start is None:
                return err(
                    '无法识别文件格式。支持：\n'
                    '① 金蝶核算维度明细账（部门明细表，含"部门名称""科目编码""借方""贷方"列）\n'
                    '② KXT模板（借方/贷方两列）'
                )
            parsed_rows, errors = _parse_kingdee_rows(ws, data_start, cm, bu, l1_map, l2_map, l3_map)
            fmt = 'kingdee'

    # Non-fatal warnings: show first 10 but don't block import
    warnings = errors[:10] if errors else []

    if not parsed_rows:
        hint = '；'.join(errors[:3]) if errors else '文件中没有有效数据行'
        return err(hint)

    # Compute P&L reconciliation preview (推算利润指标，仅展示部门明细推算的利润指标)
    pl_check = _compute_pl_check(parsed_rows) if batch_type == ImportBatch.TYPE_DEPT else {}

    # Create draft batch + entries atomically
    with transaction.atomic(using='default'):  # caiwu 已并入 default 主库(整合阶段1)
        batch = ImportBatch.objects.create(
            business_unit=bu, year=year, month=month,
            batch_type=batch_type,
            status=ImportBatch.STATUS_DRAFT,
            uploaded_by=request.pk_user,  # unified platform account (PaikuanUser)
            row_count=len(parsed_rows),
            file_name=f.name,
        )
        FinancialEntry.objects.bulk_create([
            FinancialEntry(batch=batch, l1=r['l1'], l2=r['l2'], l3=r['l3'], amount=r['amount'])
            for r in parsed_rows
        ])

    return ok({
        'batch': batch.to_dict(),
        'row_count': len(parsed_rows),
        'fmt': fmt,
        'warnings': warnings,
        'pl_check': pl_check,
    })


@cw_required()
def batch_publish(request, bid):
    if request.method != 'PUT':
        return err('方法不允许', 405)
    if not _can_publish(request):
        return err('权限不足', 403)
    try:
        batch = ImportBatch.objects.get(id=bid)
    except ImportBatch.DoesNotExist:
        return err('批次不存在', 404)
    if not _can_access_bu(request, batch.business_unit):
        return err('您无权操作该事业部数据', 403)
    if batch.status == ImportBatch.STATUS_PUBLISHED:
        return err('该批次已发布')

    # Use a transaction + row-level lock so concurrent publish requests for
    # the same BU+month don't both delete each other's data.
    with transaction.atomic(using='default'):  # caiwu 已并入 default 主库(整合阶段1)
        # Re-fetch with lock to guard against concurrent publish
        try:
            batch = ImportBatch.objects.select_for_update().get(id=bid)
        except ImportBatch.DoesNotExist:
            return err('批次不存在', 404)
        if batch.status == ImportBatch.STATUS_PUBLISHED:
            return err('该批次已发布')

        # Retire any previously published batch for same BU+month+type.
        # (部门明细表 and 利润表 coexist — only replace the same table type.)
        old_published = ImportBatch.objects.select_for_update().filter(
            business_unit=batch.business_unit,
            year=batch.year,
            month=batch.month,
            batch_type=batch.batch_type,
            status=ImportBatch.STATUS_PUBLISHED,
        )
        for old in old_published:
            old.entries.all().delete()
            old.delete()

        batch.status = ImportBatch.STATUS_PUBLISHED
        batch.published_at = timezone.now()
        batch.save()
    return ok(batch.to_dict())


@cw_required()
def batch_detail(request, bid):
    try:
        batch = ImportBatch.objects.get(id=bid)
    except ImportBatch.DoesNotExist:
        return err('批次不存在', 404)
    if not _can_access_bu(request, batch.business_unit):
        return err('无权访问', 403)

    if request.method == 'GET':
        return ok(batch.to_dict())

    if request.method == 'DELETE':
        if not _can_delete(request):
            return err('权限不足', 403)
        batch.entries.all().delete()
        batch.delete()
        return ok({'deleted': bid})

    return err('方法不允许', 405)


@cw_required()
def batch_submission_status(request):
    """GET /batches/submission-status?year=&month=
    Returns per-BU completion status: 部门明细表 是否已提交/发布。
    """
    if request.method != 'GET':
        return err('方法不允许', 405)
    year_s = request.GET.get('year', '').strip()
    month_s = request.GET.get('month', '').strip()
    try:
        year = int(year_s)
        month = int(month_s)
        assert 2000 <= year <= 2100 and 1 <= month <= 12
    except Exception:
        return err('年份或月份无效')

    # Visible BUs depend on role/job
    if request.pk_role in ('super_admin', 'manager') or request.pk_job == 'general_manager':
        bus = BUSINESS_UNITS
    else:
        bus = [b for b in (request.pk_depts or []) if b in VALID_BUSINESS_UNITS]

    # Fetch all batches for this period across visible BUs
    batches_qs = ImportBatch.objects.filter(
        business_unit__in=bus, year=year, month=month,
    ).values('business_unit', 'batch_type', 'status', 'id', 'uploaded_at', 'uploaded_by__name')

    # Build lookup: bu → {batch_type → {status, id, ...}}
    lookup = {bu: {} for bu in bus}
    for b in batches_qs:
        bu = b['business_unit']
        btype = b['batch_type']
        existing = lookup[bu].get(btype)
        # Keep the most recent / published one
        if existing is None or b['status'] == 'published':
            lookup[bu][btype] = {
                'id': b['id'],
                'status': b['status'],
                'uploaded_at': b['uploaded_at'].isoformat() if b['uploaded_at'] else None,
                'uploaded_by': b['uploaded_by__name'],
            }

    result = []
    for bu in bus:
        types = lookup[bu]
        dept_info = types.get(ImportBatch.TYPE_DEPT)
        complete = dept_info is not None and dept_info['status'] == 'published'
        result.append({
            'bu': bu,
            'department_detail': dept_info,
            'complete': complete,
        })

    return ok({'year': year, 'month': month, 'bus': result})


# ── 项目毛利（业财融合）──────────────────────────────────────────────────────────

@cw_required()
def project_margin_upload(request):
    """上传金蝶「核算维度明细账（按项目）」→ 按项目汇总收入/成本/费用。
    同 (事业部+年月) 重复上传时整体替换。"""
    if request.method != 'POST':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'charts')
    if denied:
        return denied
    if not _can_upload(request):
        return err('权限不足', 403)
    bu = request.POST.get('bu', '').strip()
    if not bu or bu not in VALID_BUSINESS_UNITS:
        return err('请选择有效的事业部')
    if not _can_access_bu(request, bu):
        return err('您无权为该事业部上传数据', 403)
    try:
        year = int(request.POST.get('year', ''))
        month = int(request.POST.get('month', ''))
        assert 2000 <= year <= 2100 and 1 <= month <= 12
    except Exception:
        return err('年份或月份无效')
    f = request.FILES.get('file')
    if not f:
        return err('请上传文件')
    if f.size > IMPORT_SIZE_LIMIT:
        return err('文件过大（上限5MB）')
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
    except Exception:
        return err('文件格式错误，请使用Excel(.xlsx)格式')

    data_start, cm = _detect_project_ledger(ws)
    if data_start is None:
        return err('无法识别为「核算维度明细账（按项目）」：需含「项目名称」「科目编码」「借方」「贷方」「摘要」列')

    agg = _parse_project_ledger(ws, data_start, cm)
    rows = []
    for project, b in agg.items():
        if all(v == 0 for v in b.values()):
            continue
        rows.append(ProjectMargin(
            business_unit=bu, year=year, month=month, project_name=project or '无',
            revenue=b['revenue'], cost=b['cost'],
            sales_exp=b['sales_exp'], mgmt_exp=b['mgmt_exp'],
            uploaded_by=request.pk_user,
        ))
    if not rows:
        return err('未解析到任何项目的收入/成本数据（请确认导出含 6001/6401 等科目）')

    with transaction.atomic(using='default'):
        ProjectMargin.objects.filter(business_unit=bu, year=year, month=month).delete()
        ProjectMargin.objects.bulk_create(rows)

    return ok({'business_unit': bu, 'year': year, 'month': month, 'project_count': len(rows)})


def _allocate_unalloc(rows, unalloc):
    """把未挂项目池(unalloc) 的成本/费用按各项目收入比例分摊；收入为正者参与分摊。
    返回新的 rows（每项目 cost/sales_exp/mgmt_exp 已加分摊额）。"""
    total_rev = sum(r['revenue'] for r in rows if r['revenue'] > 0)
    if total_rev <= 0:
        return rows  # 无收入可依据，不分摊（未挂池仍单列在 summary）
    for r in rows:
        share = (r['revenue'] / total_rev) if r['revenue'] > 0 else 0
        r['cost'] = round(r['cost'] + unalloc['cost'] * share, 2)
        r['sales_exp'] = round(r['sales_exp'] + unalloc['sales_exp'] * share, 2)
        r['mgmt_exp'] = round(r['mgmt_exp'] + unalloc['mgmt_exp'] * share, 2)
        r['margin'] = round(r['revenue'] - r['cost'], 2)
        r['margin_rate'] = round(r['margin'] / r['revenue'] * 100, 1) if r['revenue'] else None
    return rows


@cw_required()
def project_margin(request):
    """项目毛利透视。GET ?bu=&year=&month=&mode=direct|allocated。
    direct：未挂项目「无」单列为未分摊池；allocated：未挂成本按收入比例分摊到各项目。"""
    if request.method != 'GET':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'charts')
    if denied:
        return denied
    bu = request.GET.get('bu', '').strip()
    if not bu or bu not in VALID_BUSINESS_UNITS:
        return err('请选择有效的事业部')
    if not _can_access_bu(request, bu):
        return err('无权访问', 403)
    try:
        year = int(request.GET.get('year', ''))
        month = int(request.GET.get('month', ''))
    except Exception:
        return err('年份或月份无效')
    mode = request.GET.get('mode', 'direct').strip() or 'direct'

    qs = ProjectMargin.objects.filter(business_unit=bu, year=year, month=month)
    rows, unalloc = [], {'revenue': 0.0, 'cost': 0.0, 'sales_exp': 0.0, 'mgmt_exp': 0.0}
    for m in qs:
        d = m.to_dict()
        if (m.project_name or '').strip() in _PM_UNALLOCATED:
            unalloc['revenue'] += d['revenue']; unalloc['cost'] += d['cost']
            unalloc['sales_exp'] += d['sales_exp']; unalloc['mgmt_exp'] += d['mgmt_exp']
        else:
            rows.append(d)

    if mode == 'allocated':
        rows = _allocate_unalloc(rows, unalloc)

    rows.sort(key=lambda r: r['margin'], reverse=True)
    total_rev = sum(r['revenue'] for r in rows)
    total_cost = sum(r['cost'] for r in rows)
    # 收入是否按项目核算：若各项目收入几乎为 0、收入全在未挂池 → 本事业部不适用项目毛利
    revenue_by_project = total_rev > 0
    # direct 口径下，未挂池成本不进各项目，但计入整体合计的"未分摊成本"
    grand_cost = total_cost + (unalloc['cost'] if mode == 'direct' else 0)
    grand_rev = total_rev + (unalloc['revenue'] if mode == 'direct' else 0)
    grand_margin = grand_rev - grand_cost
    summary = {
        'project_count': len(rows),
        'total_revenue': round(grand_rev, 2),
        'total_cost': round(grand_cost, 2),
        'total_margin': round(grand_margin, 2),
        'margin_rate': round(grand_margin / grand_rev * 100, 1) if grand_rev else None,
        'unalloc_cost': round(unalloc['cost'], 2),
        'unalloc_revenue': round(unalloc['revenue'], 2),
        'has_data': qs.exists(),
        'revenue_by_project': revenue_by_project,
    }
    return ok({'mode': mode, 'business_unit': bu, 'year': year, 'month': month,
               'rows': rows, 'summary': summary})


# ── report ───────────────────────────────────────────────────────────────────

def _get_published_batches(bu_list, year, month, batch_type=ImportBatch.TYPE_DEPT):
    """Return published ImportBatch objects for given BUs and period.

    Reports/charts aggregate the 部门明细表 (TYPE_DEPT) only — the 利润表 (TYPE_PL)
    is a reconciliation reference and must NOT be summed alongside it (double-count).
    Pass batch_type=None to include all types.
    """
    qs = ImportBatch.objects.filter(
        business_unit__in=bu_list,
        year=year, month=month,
        status=ImportBatch.STATUS_PUBLISHED,
    )
    if batch_type:
        qs = qs.filter(batch_type=batch_type)
    return qs


def _aggregate_report(batches_qs, level):
    """Return hierarchical report rows including calculated L1 rows."""
    from collections import defaultdict

    batch_ids = list(batches_qs.values_list('id', flat=True))
    if not batch_ids:
        return []

    l1_cats = list(L1Category.objects.order_by('sort_order', 'id').all())

    # Raw aggregation by l1_id
    raw_agg = (
        FinancialEntry.objects
        .filter(batch_id__in=batch_ids)
        .values('l1_id')
        .annotate(amount=Sum('amount'))
    )
    raw_by_id = {r['l1_id']: float(r['amount']) for r in raw_agg}

    # Compute all L1 amounts including calculated rows
    name_map, id_map = _compute_l1_name_map(l1_cats, raw_by_id)

    if level == 1:
        rows = []
        for l1 in l1_cats:
            amt = id_map.get(l1.id)
            if amt is None:
                continue
            rows.append({
                'l1_id': l1.id, 'l1_name': l1.name,
                'amount': amt, 'is_calculated': l1.is_calculated,
            })
        return rows

    # For levels 2/3: fetch raw l2/l3 breakdown; calculated rows have no children
    if level == 2:
        l2_agg = (
            FinancialEntry.objects
            .filter(batch_id__in=batch_ids)
            .values('l1_id', 'l2_id', 'l2__name')
            .annotate(amount=Sum('amount'))
        )
        l1_children = defaultdict(list)
        for r in l2_agg:
            l1_children[r['l1_id']].append({
                'l2_id': r['l2_id'],
                'l2_name': r['l2__name'] or '（无项目部）',
                'amount': float(r['amount']),
            })

        rows = []
        for l1 in l1_cats:
            amt = id_map.get(l1.id)
            if amt is None:
                continue
            rows.append({
                'l1_id': l1.id, 'l1_name': l1.name,
                'amount': amt, 'is_calculated': l1.is_calculated,
                'children': [] if l1.is_calculated else sorted(
                    l1_children.get(l1.id, []), key=lambda x: x['l2_name'],
                ),
            })
        return rows

    if level == 3:
        l3_agg = (
            FinancialEntry.objects
            .filter(batch_id__in=batch_ids)
            .values('l1_id', 'l2_id', 'l2__name', 'l3_id', 'l3__name')
            .annotate(amount=Sum('amount'))
        )
        l2_totals = defaultdict(Decimal)
        l3_rows = defaultdict(list)
        l2_meta = {}
        for r in l3_agg:
            key = (r['l1_id'], r['l2_id'])
            l2_totals[key] += r['amount']
            l2_meta[key] = r['l2__name']
            l3_rows[key].append({
                'l3_id': r['l3_id'],
                'l3_name': r['l3__name'] or '（无明细）',
                'amount': float(r['amount']),
            })

        rows = []
        for l1 in l1_cats:
            amt = id_map.get(l1.id)
            if amt is None:
                continue
            if l1.is_calculated:
                children = []
            else:
                children = []
                for key, tot in l2_totals.items():
                    if key[0] != l1.id:
                        continue
                    l2_id = key[1]
                    children.append({
                        'l2_id': l2_id,
                        'l2_name': l2_meta.get(key) or '（无项目部）',
                        'amount': float(tot),
                        'children': l3_rows[key],
                    })
            rows.append({
                'l1_id': l1.id, 'l1_name': l1.name,
                'amount': amt, 'is_calculated': l1.is_calculated,
                'children': children,
            })
        return rows

    return []


def _report_trend(bu_list, year, month, level, n=6):
    """报表各层级最近 n 个月金额序列 + 环比，用于行内迷你趋势线 / 环比。

    在所查看的 level 上逐月聚合并按节点身份归集（id 跨月稳定）：
      返回 {'l1': {l1_id: (trend, mom)},
            'l2': {(l1_id, l2_id): (trend, mom)},          # level>=2
            'l3': {(l1_id, l2_id, l3_id): (trend, mom)}}   # level>=3
    trend 为 [oldest…newest] 长度 n，缺失月补 0.0；上月为 0/缺失时 mom=None。
    """
    from collections import defaultdict
    seq = []
    yy, mm = year, month
    for _ in range(n):
        seq.append((yy, mm))
        mm -= 1
        if mm == 0:
            mm, yy = 12, yy - 1
    seq.reverse()  # oldest → newest

    s1 = defaultdict(lambda: [0.0] * n)
    s2 = defaultdict(lambda: [0.0] * n)
    s3 = defaultdict(lambda: [0.0] * n)
    for idx, (yy, mm) in enumerate(seq):
        rows = _aggregate_report(_get_published_batches(bu_list, yy, mm), level)
        for r in rows:
            s1[r['l1_id']][idx] = float(r['amount'])
            if level >= 2:
                for l2 in r.get('children', []):
                    s2[(r['l1_id'], l2['l2_id'])][idx] = float(l2['amount'])
                    if level >= 3:
                        for l3 in l2.get('children', []):
                            s3[(r['l1_id'], l2['l2_id'], l3['l3_id'])][idx] = float(l3['amount'])

    def _mom(series):
        prev = series[-2] if len(series) >= 2 else 0.0
        return round((series[-1] - prev) / abs(prev) * 100, 1) if prev else None

    out = {'l1': {k: (v, _mom(v)) for k, v in s1.items()}}
    if level >= 2:
        out['l2'] = {k: (v, _mom(v)) for k, v in s2.items()}
    if level >= 3:
        out['l3'] = {k: (v, _mom(v)) for k, v in s3.items()}
    return out


@cw_required()
def report(request):
    if request.method != 'GET':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'report')
    if denied:
        return denied

    year_s = request.GET.get('year', '')
    month_s = request.GET.get('month', '')
    bu_param = request.GET.get('bu', '')
    try:
        level = int(request.GET.get('level', '1'))
        assert level in (1, 2, 3)
    except Exception:
        level = 1

    # Clamp drill-down to what this job may view.
    if level >= 3 and not _can_view(request, 'report_l3'):
        level = 2
    if level >= 2 and not _can_view(request, 'report_l2'):
        level = 1

    try:
        year = int(year_s)
        month = int(month_s)
        assert 2000 <= year <= 2100 and 1 <= month <= 12
    except Exception:
        return err('年份或月份无效')

    # Determine accessible BUs
    if request.pk_role == 'super_admin':
        accessible = BUSINESS_UNITS
    else:
        accessible = [d for d in request.pk_depts if d in VALID_BUSINESS_UNITS]

    if bu_param:
        if bu_param not in accessible:
            return err('无权访问该事业部', 403)
        bu_list = [bu_param]
    else:
        bu_list = accessible

    batches_qs = _get_published_batches(bu_list, year, month)
    rows = _aggregate_report(batches_qs, level)
    # Use 经营净利 as headline total; fall back to sum of raw rows
    profit_row = next((r for r in rows if r['l1_name'] == '经营净利'), None)
    total = profit_row['amount'] if profit_row else sum(
        (r['amount'] for r in rows if not r.get('is_calculated')), Decimal(0)
    )
    total_label = '经营净利' if profit_row else '合计金额'

    # Previous month KPIs for MoM comparison (level=1 only, lightweight)
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    prev_batches = _get_published_batches(bu_list, prev_year, prev_month)
    prev_rows_l1 = _aggregate_report(prev_batches, 1)
    prev_kpis = {r['l1_name']: float(r['amount']) for r in prev_rows_l1}

    # 行内环比 + 迷你趋势线：按所查看层级，为 L1/L2/L3 各节点附 mom/trend
    tr = _report_trend(bu_list, year, month, level, 6)
    for r in rows:
        t1 = tr['l1'].get(r['l1_id'])
        r['trend'], r['mom'] = t1 if t1 else (None, None)
        for l2 in r.get('children', []):
            t2 = tr.get('l2', {}).get((r['l1_id'], l2['l2_id']))
            l2['trend'], l2['mom'] = t2 if t2 else (None, None)
            for l3 in l2.get('children', []):
                t3 = tr.get('l3', {}).get((r['l1_id'], l2['l2_id'], l3['l3_id']))
                l3['trend'], l3['mom'] = t3 if t3 else (None, None)

    return ok({
        'rows': rows, 'total': float(total), 'total_label': total_label,
        'level': level, 'year': year, 'month': month, 'bu': bu_param or None,
        'prev_kpis': prev_kpis, 'prev_year': prev_year, 'prev_month': prev_month,
    })


@cw_required()
def report_export(request):
    if request.method != 'GET':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'report')
    if denied:
        return denied
    if not _can_view(request, 'export'):
        return err('无导出权限', 403, 403)

    year_s = request.GET.get('year', '')
    month_s = request.GET.get('month', '')
    bu_param = request.GET.get('bu', '')
    try:
        level = int(request.GET.get('level', '1'))
        assert level in (1, 2, 3)
    except Exception:
        level = 1

    if level >= 3 and not _can_view(request, 'report_l3'):
        level = 2
    if level >= 2 and not _can_view(request, 'report_l2'):
        level = 1

    try:
        year = int(year_s)
        month = int(month_s)
        assert 2000 <= year <= 2100 and 1 <= month <= 12
    except Exception:
        return err('年份或月份无效')

    if request.pk_role == 'super_admin':
        accessible = BUSINESS_UNITS
    else:
        accessible = [d for d in request.pk_depts if d in VALID_BUSINESS_UNITS]

    if bu_param:
        if bu_param not in accessible:
            return err('无权访问该事业部', 403)
        bu_list = [bu_param]
    else:
        bu_list = accessible

    batches_qs = _get_published_batches(bu_list, year, month)
    rows = _aggregate_report(batches_qs, level)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return err('服务器缺少 openpyxl 依赖')

    wb = openpyxl.Workbook()
    ws = wb.active
    level_label = {1: '一级', 2: '二级', 3: '三级'}[level]
    ws.title = f'{year}年{month}月{level_label}报表'

    header_fill = PatternFill(start_color='C96342', end_color='C96342', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    if level == 1:
        headers = ['一级科目', '金额(元)']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            ws.append([row['l1_name'], float(row['amount'])])
    elif level == 2:
        headers = ['一级科目', '二级项目部', '金额(元)']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            ws.append([row['l1_name'], '', float(row['amount'])])
            for child in row.get('children', []):
                ws.append(['', child['l2_name'], float(child['amount'])])
    else:
        headers = ['一级科目', '二级项目部', '三级科目明细', '金额(元)']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            ws.append([row['l1_name'], '', '', float(row['amount'])])
            for l2 in row.get('children', []):
                ws.append(['', l2['l2_name'], '', float(l2['amount'])])
                for l3 in l2.get('children', []):
                    ws.append(['', '', l3['l3_name'], float(l3['amount'])])

    bu_label = bu_param or '全部事业部'
    filename = f'财务报表_{bu_label}_{year}年{month}月_{level_label}.xlsx'
    return _build_excel_response(wb, filename)


# ── 指标管理 & 财务驾驶舱 ──────────────────────────────────────────────────────
#
# Headline metrics are revenue (主营业务收入) and profit (经营净利). Targets are
# entered manually (FinancialTarget); actuals come from the aggregated, published
# 部门明细表. month=0 holds the annual target; 1-12 hold monthly targets.

REVENUE_L1 = '主营业务收入'
PROFIT_L1 = '经营净利'
GROSS_PROFIT_L1 = '经营毛利'


def _collect_actuals(bu_list, years):
    """Pull all published 部门明细表 actuals in ONE grouped query.

    Returns {(business_unit, year, month): name_map}, where name_map maps each L1
    category name → float (calculated rows like 经营净利 included). Periods/BUs with
    no published entries are simply absent. This replaces the per-BU/per-period
    _aggregate_report fan-out (dozens of round-trips) with a single query plus one
    L1Category fetch, so metrics/cockpit run in ~3 queries instead of ~50-90.
    """
    from collections import defaultdict
    agg = (
        FinancialEntry.objects
        .filter(
            batch__status=ImportBatch.STATUS_PUBLISHED,
            batch__batch_type=ImportBatch.TYPE_DEPT,
            batch__business_unit__in=bu_list,
            batch__year__in=list(years),
        )
        .values('batch__business_unit', 'batch__year', 'batch__month', 'l1_id')
        .annotate(amount=Sum('amount'))
    )
    raw = defaultdict(dict)
    for r in agg:
        key = (r['batch__business_unit'], r['batch__year'], r['batch__month'])
        raw[key][r['l1_id']] = float(r['amount'])

    l1_cats = list(L1Category.objects.order_by('sort_order', 'id'))
    out = {}
    for key, raw_by_id in raw.items():
        name_map, _ = _compute_l1_name_map(l1_cats, raw_by_id)
        out[key] = name_map
    return out


def _rp(actuals, bu, year, month):
    """(revenue, profit, gross_profit) for one BU/period from the actuals map."""
    nm = actuals.get((bu, year, month))
    if not nm:
        return None, None, None
    return nm.get(REVENUE_L1), nm.get(PROFIT_L1), nm.get(GROSS_PROFIT_L1)


def _period_group(actuals, bus, year, month):
    """(revenue, profit, gross_profit) summed across BUs for one period."""
    tuples = [_rp(actuals, bu, year, month) for bu in bus]
    rev   = _sum_opt([t[0] for t in tuples])
    prof  = _sum_opt([t[1] for t in tuples])
    gross = _sum_opt([t[2] for t in tuples])
    return rev, prof, gross


def _prev_period(year, month):
    """Calendar month before (year, month)."""
    return (year, month - 1) if month > 1 else (year - 1, 12)


def _rate(actual, target):
    """Achievement rate (%), rounded; None if not computable.

    使用 (1 + (actual - target)/|target|) * 100：
    - 目标为正时与 actual/target*100 完全等价（向后兼容）；
    - 目标为负（计划性亏损）时方向正确——实际亏损更小/转盈 → 达成率更高，
      实际亏损更大 → 达成率更低，不再出现 actual/target 因负负得正导致的反向；
    - 实际为负而目标为正时得到负达成率，如实反映"不仅没达标还亏损"。
    """
    if target in (None, 0) or actual is None:
        return None
    return round((1 + (actual - target) / abs(float(target))) * 100, 1)


def _chg(cur, prev):
    """Percent change cur vs prev; None when prev is missing or zero."""
    if cur is None or prev is None or prev == 0:
        return None
    return round((cur - prev) / abs(prev) * 100, 1)


def _accessible_bus(request):
    if request.pk_role == 'super_admin':
        return list(BUSINESS_UNITS)
    return [d for d in request.pk_depts if d in VALID_BUSINESS_UNITS]


def _resolve_bu_list(request, bu_param):
    """(bu_list, error_response). bu_param empty → all accessible BUs."""
    accessible = _accessible_bus(request)
    if bu_param:
        if bu_param not in accessible:
            return None, err('无权访问该事业部', 403)
        return [bu_param], None
    return accessible, None


def _parse_year_month(request):
    year = int(request.GET.get('year', ''))
    month = int(request.GET.get('month', ''))
    assert 2000 <= year <= 2100 and 1 <= month <= 12
    return year, month


def _metric_block(target_rev, actual_rev, target_prof, actual_prof,
                  target_gross=None, actual_gross=None,
                  rev_prev=None, prof_prev=None, gross_prev=None,
                  rev_yoy=None, prof_yoy=None, gross_yoy=None,
                  with_chg=False):
    """Assemble one revenue/profit metric block: target, actual, rate, and
    (when with_chg) MoM/YoY change keys — always present for a uniform shape."""
    block = {
        'target_revenue': float(target_rev) if target_rev is not None else 0.0,
        'actual_revenue': actual_rev,
        'revenue_rate': _rate(actual_rev, target_rev),
        'target_profit': float(target_prof) if target_prof is not None else 0.0,
        'actual_profit': actual_prof,
        'profit_rate': _rate(actual_prof, target_prof),
        'target_gross_profit': float(target_gross) if target_gross is not None else 0.0,
        'actual_gross_profit': actual_gross,
        'gross_profit_rate': _rate(actual_gross, target_gross),
    }
    if with_chg:
        block['revenue_mom'] = _chg(actual_rev, rev_prev)
        block['profit_mom'] = _chg(actual_prof, prof_prev)
        block['gross_profit_mom'] = _chg(actual_gross, gross_prev)
        block['revenue_yoy'] = _chg(actual_rev, rev_yoy)
        block['profit_yoy'] = _chg(actual_prof, prof_yoy)
        block['gross_profit_yoy'] = _chg(actual_gross, gross_yoy)
    return block


def _bu_metrics(bu, year, month, tgt_index, actuals):
    """Full metric bundle (month block + YTD block) for one BU, from the
    pre-collected actuals map (no per-BU DB round-trips)."""
    prev_year, prev_month = _prev_period(year, month)

    rev_m,  prof_m,  gross_m  = _rp(actuals, bu, year, month)
    rev_pm, prof_pm, gross_pm = _rp(actuals, bu, prev_year, prev_month)
    rev_ym, prof_ym, gross_ym = _rp(actuals, bu, year - 1, month)
    # YTD = sum of monthly actuals Jan..month (revenue/profit/gross all additive).
    rev_ytd   = _sum_opt([_rp(actuals, bu, year, m)[0] for m in range(1, month + 1)])
    prof_ytd  = _sum_opt([_rp(actuals, bu, year, m)[1] for m in range(1, month + 1)])
    gross_ytd = _sum_opt([_rp(actuals, bu, year, m)[2] for m in range(1, month + 1)])

    m_tgt = tgt_index.get((bu, month), {})
    a_tgt = tgt_index.get((bu, FinancialTarget.MONTH_ANNUAL), {})

    return {
        'business_unit': bu,
        'month': _metric_block(
            m_tgt.get('target_revenue'), rev_m,
            m_tgt.get('target_profit'), prof_m,
            target_gross=m_tgt.get('target_gross_profit'), actual_gross=gross_m,
            rev_prev=rev_pm, prof_prev=prof_pm, gross_prev=gross_pm,
            rev_yoy=rev_ym, prof_yoy=prof_ym, gross_yoy=gross_ym,
            with_chg=True),
        'ytd': _metric_block(
            a_tgt.get('target_revenue'), rev_ytd,
            a_tgt.get('target_profit'), prof_ytd,
            target_gross=a_tgt.get('target_gross_profit'), actual_gross=gross_ytd),
    }


def _sum_opt(values):
    """Sum treating None as 0; return None only if every value is None."""
    present = [v for v in values if v is not None]
    return sum(present) if present else None


def _aggregate_total(bu_blocks, key):
    """Build a group-total block by summing per-BU blocks for 'month' or 'ytd'."""
    rev_t   = _sum_opt([b[key]['target_revenue'] for b in bu_blocks])
    rev_a   = _sum_opt([b[key]['actual_revenue'] for b in bu_blocks])
    prof_t  = _sum_opt([b[key]['target_profit'] for b in bu_blocks])
    prof_a  = _sum_opt([b[key]['actual_profit'] for b in bu_blocks])
    gross_t = _sum_opt([b[key]['target_gross_profit'] for b in bu_blocks])
    gross_a = _sum_opt([b[key]['actual_gross_profit'] for b in bu_blocks])
    return {
        'target_revenue': rev_t or 0.0, 'actual_revenue': rev_a,
        'revenue_rate': _rate(rev_a, rev_t),
        'target_profit': prof_t or 0.0, 'actual_profit': prof_a,
        'profit_rate': _rate(prof_a, prof_t),
        'target_gross_profit': gross_t or 0.0, 'actual_gross_profit': gross_a,
        'gross_profit_rate': _rate(gross_a, gross_t),
    }


def _attach_group_chg(block, bus, year, month, actuals):
    """Add group-level MoM/YoY to a month total block (rates can't be summed, so
    recompute from consolidated actuals)."""
    prev_year, prev_month = _prev_period(year, month)
    rev_m,  prof_m,  gross_m  = _period_group(actuals, bus, year, month)
    rev_pm, prof_pm, gross_pm = _period_group(actuals, bus, prev_year, prev_month)
    rev_ym, prof_ym, gross_ym = _period_group(actuals, bus, year - 1, month)
    block['revenue_mom'] = _chg(rev_m, rev_pm)
    block['profit_mom'] = _chg(prof_m, prof_pm)
    block['gross_profit_mom'] = _chg(gross_m, gross_pm)
    block['revenue_yoy'] = _chg(rev_m, rev_ym)
    block['profit_yoy'] = _chg(prof_m, prof_ym)
    block['gross_profit_yoy'] = _chg(gross_m, gross_ym)
    return block


def _load_target_index(bus, year):
    """{(bu, month): {target_revenue, target_profit, target_gross_profit}} for the year."""
    idx = {}
    for t in FinancialTarget.objects.filter(business_unit__in=bus, year=year):
        idx[(t.business_unit, t.month)] = {
            'target_revenue': float(t.target_revenue),
            'target_profit': float(t.target_profit),
            'target_gross_profit': float(t.target_gross_profit),
        }
    return idx


@cw_required()
def targets(request):
    """GET ?year=&bu= → editable target rows; POST upsert a batch of targets."""
    denied = _page_denied(request, 'metrics')
    if denied:
        return denied

    if request.method == 'GET':
        try:
            year = int(request.GET.get('year', ''))
            assert 2000 <= year <= 2100
        except Exception:
            return err('年份无效')
        bus, e = _resolve_bu_list(request, request.GET.get('bu', ''))
        if e:
            return e
        rows = list(
            FinancialTarget.objects.filter(business_unit__in=bus, year=year)
            .order_by('business_unit', 'month')
        )
        return ok({'year': year, 'targets': [r.to_dict() for r in rows]})

    if request.method == 'POST':
        if not (request.pk_role == 'super_admin' or _can_upload(request)):
            return err('无录入权限', 403, 403)
        body = _parse_json(request)
        try:
            year = int(body.get('year'))
            assert 2000 <= year <= 2100
        except Exception:
            return err('年份无效')
        items = body.get('items') or []
        if not isinstance(items, list):
            return err('items 格式错误')

        accessible = _accessible_bus(request)
        uid = getattr(request, 'pk_uid', None)

        # ── parse all items first, then validate before touching the DB ────────
        parsed = []  # (bu, month, rev_dec, prof_dec)
        for it in items:
            bu = it.get('business_unit')
            if bu not in accessible:
                return err(f'无权编辑事业部：{bu}', 403)
            try:
                mo = int(it.get('month', 0))
                assert 0 <= mo <= 12
            except Exception:
                return err('月份无效')
            try:
                rev   = Decimal(str(it.get('target_revenue', 0) or 0))
                prof  = Decimal(str(it.get('target_profit', 0) or 0))
                gross = Decimal(str(it.get('target_gross_profit', 0) or 0))
            except (InvalidOperation, ValueError):
                return err('目标金额无效')
            parsed.append((bu, mo, rev, prof, gross))

        # ── month-sum == annual validation ────────────────────────────────────
        TOLERANCE = Decimal('1.00')
        from collections import defaultdict
        affected = {bu for bu, *_ in parsed}
        merged = defaultdict(dict)  # bu -> {month: (rev, prof, gross)}
        for t in FinancialTarget.objects.filter(year=year, business_unit__in=affected):
            merged[t.business_unit][t.month] = (
                t.target_revenue, t.target_profit, t.target_gross_profit)
        for bu, mo, rev, prof, gross in parsed:
            merged[bu][mo] = (rev, prof, gross)
        for bu, mp in merged.items():
            if 0 not in mp or any(m not in mp for m in range(1, 13)):
                continue
            for col, label in ((0, '收入'), (1, '经营净利'), (2, '经营毛利')):
                s = sum(mp[m][col] for m in range(1, 13))
                if abs(s - mp[0][col]) > TOLERANCE:
                    delta = float(s - mp[0][col]) / 10000
                    return err(f'{bu}：月度{label}目标合计与年度目标不符'
                               f'（差额 {delta:+.2f} 万元），请修正后保存')

        with transaction.atomic():
            for bu, mo, rev, prof, gross in parsed:
                FinancialTarget.objects.update_or_create(
                    business_unit=bu, year=year, month=mo,
                    defaults={
                        'target_revenue': rev, 'target_profit': prof,
                        'target_gross_profit': gross,
                        'updated_by_id': uid,
                    },
                )
        return ok({'saved': len(parsed)})

    return err('方法不允许', 405)


# ── Target template / upload / export ────────────────────────────────────────

_TGT_SECTIONS = [
    ('收入目标（万元）',    'target_revenue',      '收入'),
    ('经营毛利目标（万元）', 'target_gross_profit', '经营毛利'),
    ('经营净利目标（万元）', 'target_profit',       '经营净利'),
]
_MONTH_LABELS = ['1月','2月','3月','4月','5月','6月','7月','8月','9月','10月','11月','12月']


def _build_target_wb(bus, year, existing_map):
    """Build an openpyxl workbook with 3 stacked sections (收入/净利/毛利).
    existing_map = {(bu, month): {target_revenue, target_profit, target_gross_profit}}
    For a blank template pass {} as existing_map.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = '指标模板'

    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill  = PatternFill('solid', fgColor='C96342')
    sub_font  = Font(bold=True, size=10)
    sub_fill  = PatternFill('solid', fgColor='F3E8E0')
    col_font  = Font(bold=True, size=9, color='555555')
    col_fill  = PatternFill('solid', fgColor='FAF3EF')
    thin      = Side(style='thin', color='DDDDDD')
    thin_bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal='center', vertical='center')
    right_al  = Alignment(horizontal='right', vertical='center')

    row = 1
    # Year banner
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
    c = ws.cell(row=row, column=1, value=f'{year} 年度经营指标（万元）')
    c.font = Font(bold=True, size=13); c.alignment = center
    row += 1

    for sect_label, _field, _short in _TGT_SECTIONS:
        # Section header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
        c = ws.cell(row=row, column=1, value=sect_label)
        c.font = sub_font; c.fill = sub_fill; c.alignment = center
        row += 1

        # Column header
        ws.cell(row=row, column=1, value='事业部').font = col_font
        ws.cell(row=row, column=1).fill = col_fill
        ws.cell(row=row, column=1).alignment = center
        for mi, lbl in enumerate(_MONTH_LABELS, 2):
            c = ws.cell(row=row, column=mi, value=lbl)
            c.font = col_font; c.fill = col_fill; c.alignment = center
        c = ws.cell(row=row, column=14, value='年度目标')
        c.font = col_font; c.fill = col_fill; c.alignment = center
        c = ws.cell(row=row, column=15, value='月合计')
        c.font = col_font; c.fill = col_fill; c.alignment = center
        row += 1

        # BU data rows
        for bu in bus:
            ws.cell(row=row, column=1, value=bu).font = Font(bold=True, size=10)
            ws.cell(row=row, column=1).alignment = Alignment(vertical='center')
            for mi in range(1, 13):
                val = existing_map.get((bu, mi), {}).get(_field)
                c = ws.cell(row=row, column=mi + 1,
                            value=round(val / 10000, 2) if val is not None else None)
                c.number_format = '#,##0.00'
                c.alignment = right_al
            annual_val = existing_map.get((bu, 0), {}).get(_field)
            c = ws.cell(row=row, column=14,
                        value=round(annual_val / 10000, 2) if annual_val is not None else None)
            c.number_format = '#,##0.00'; c.alignment = right_al
            # Month-sum formula
            col_start = get_column_letter(2)
            col_end   = get_column_letter(13)
            ws.cell(row=row, column=15,
                    value=f'=SUM({col_start}{row}:{col_end}{row})').number_format = '#,##0.00'
            ws.cell(row=row, column=15).alignment = right_al
            # Apply border to all cells in row
            for col in range(1, 16):
                ws.cell(row=row, column=col).border = thin_bdr
            row += 1

        row += 1  # blank separator

    # Column widths
    ws.column_dimensions['A'].width = 16
    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = 9
    ws.column_dimensions['N'].width = 10
    ws.column_dimensions['O'].width = 10

    return wb


@cw_required()
def targets_template(request):
    """GET ?year= → download blank Excel template for target entry."""
    if request.method != 'GET':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'metrics')
    if denied:
        return denied
    try:
        year = int(request.GET.get('year', yearCST()))
        assert 2000 <= year <= 2100
    except Exception:
        return err('年份无效')
    bus, e = _resolve_bu_list(request, '')
    if e:
        return e
    wb = _build_target_wb(bus, year, {})
    return _build_excel_response(wb, f'{year}年指标模板.xlsx')


@cw_required()
def targets_export(request):
    """GET ?year= → export current targets as filled Excel."""
    if request.method != 'GET':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'metrics')
    if denied:
        return denied
    try:
        year = int(request.GET.get('year', yearCST()))
        assert 2000 <= year <= 2100
    except Exception:
        return err('年份无效')
    bus, e = _resolve_bu_list(request, '')
    if e:
        return e
    rows = FinancialTarget.objects.filter(business_unit__in=bus, year=year)
    existing_map = {}
    for t in rows:
        existing_map[(t.business_unit, t.month)] = {
            'target_revenue': float(t.target_revenue),
            'target_profit': float(t.target_profit),
            'target_gross_profit': float(t.target_gross_profit),
        }
    wb = _build_target_wb(bus, year, existing_map)
    return _build_excel_response(wb, f'{year}年指标导出.xlsx')


@cw_required()
def targets_upload(request):
    """POST multipart/form-data: file=<xlsx> year=<int> → bulk upsert targets."""
    if request.method != 'POST':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'metrics')
    if denied:
        return denied
    if not (request.pk_role == 'super_admin' or _can_upload(request)):
        return err('无录入权限', 403, 403)

    try:
        year = int(request.POST.get('year', ''))
        assert 2000 <= year <= 2100
    except Exception:
        return err('年份无效')

    f = request.FILES.get('file')
    if not f:
        return err('缺少文件')
    if f.size > IMPORT_SIZE_LIMIT:
        return err('文件过大（超过 5 MB）')

    accessible = _accessible_bus(request)
    uid = getattr(request, 'pk_uid', None)

    try:
        from openpyxl import load_workbook
        wb = load_workbook(f, data_only=True)
        ws = wb.active
    except Exception as ex:
        return err(f'文件解析失败：{ex}')

    # Detect section boundaries: column A contains the section header keyword.
    section_map = {s[0]: s[1] for s in _TGT_SECTIONS}
    valid_bus = set(accessible)
    current_field = None
    parsed = {}  # {(bu, month): {field: Decimal}}

    for row in ws.iter_rows():
        a_val = str(row[0].value or '').strip()
        if a_val in section_map:
            current_field = section_map[a_val]
            continue
        if current_field is None or not a_val:
            continue
        bu = a_val
        if bu not in valid_bus or a_val in ('事业部', '月合计', '年合计'):
            continue
        if bu not in valid_bus:
            continue
        # Columns B-M = months 1-12, N = annual (month=0)
        slots = [(i + 1, row[i + 1]) for i in range(12)]  # (month 1-12, cell)
        slots.append((0, row[13] if len(row) > 13 else None))  # annual
        for month_no, cell in slots:
            if cell is None:
                continue
            try:
                wan_val = cell.value
                if wan_val is None:
                    amount = Decimal('0')
                else:
                    amount = Decimal(str(float(wan_val))) * 10000
            except (InvalidOperation, ValueError, TypeError):
                continue
            key = (bu, month_no)
            if key not in parsed:
                parsed[key] = {}
            parsed[key][current_field] = amount

    if not parsed:
        return err('未从文件中解析到任何有效数据，请确认使用正确的指标模板')

    # Validate sum == annual for each BU × metric
    TOLERANCE = Decimal('1.00')
    from collections import defaultdict
    merged = defaultdict(lambda: defaultdict(dict))  # {bu: {month: {field: val}}}
    existing_qs = FinancialTarget.objects.filter(
        year=year, business_unit__in=accessible)
    for t in existing_qs:
        merged[t.business_unit][t.month] = {
            'target_revenue': t.target_revenue,
            'target_profit': t.target_profit,
            'target_gross_profit': t.target_gross_profit,
        }
    for (bu, mo), fields in parsed.items():
        for field, val in fields.items():
            if bu not in merged or mo not in merged[bu]:
                merged[bu][mo] = {}
            merged[bu][mo][field] = val

    for bu, mp in merged.items():
        if 0 not in mp or any(m not in mp for m in range(1, 13)):
            continue
        for col_field, label in (
            ('target_revenue', '收入'),
            ('target_profit', '经营净利'),
            ('target_gross_profit', '经营毛利'),
        ):
            annual = mp[0].get(col_field, Decimal('0'))
            s = sum(mp[m].get(col_field, Decimal('0')) for m in range(1, 13))
            if annual != 0 and abs(s - annual) > TOLERANCE:
                delta = float(s - annual) / 10000
                return err(f'{bu}：月度{label}合计与年度目标不符（差额 {delta:+.2f} 万元）')

    # Upsert
    saved = 0
    with transaction.atomic():
        for (bu, mo), fields in parsed.items():
            if not fields:
                continue
            FinancialTarget.objects.update_or_create(
                business_unit=bu, year=year, month=mo,
                defaults={**fields, 'updated_by_id': uid},
            )
            saved += 1

    return ok({'saved': saved})


def yearCST():
    from django.utils import timezone
    import datetime
    now = timezone.now().astimezone(datetime.timezone(datetime.timedelta(hours=8)))
    return now.year


@cw_required()
def metrics(request):
    """Per-BU revenue/profit targets vs actuals with achievement, MoM and YoY."""
    if request.method != 'GET':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'metrics')
    if denied:
        return denied
    try:
        year, month = _parse_year_month(request)
    except Exception:
        return err('年份或月份无效')

    bus, e = _resolve_bu_list(request, request.GET.get('bu', ''))
    if e:
        return e

    tgt_index = _load_target_index(bus, year)
    actuals = _collect_actuals(bus, {year, year - 1})
    bu_rows = [_bu_metrics(bu, year, month, tgt_index, actuals) for bu in bus]
    total = {
        'business_unit': '全集团',
        'month': _attach_group_chg(_aggregate_total(bu_rows, 'month'), bus, year, month, actuals),
        'ytd': _aggregate_total(bu_rows, 'ytd'),
    }
    return ok({'year': year, 'month': month, 'bus': bu_rows, 'total': total})


@cw_required()
def cockpit(request):
    """Group-level KPI overview + per-BU bars + 12-month trend (actual vs target)."""
    if request.method != 'GET':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'cockpit')
    if denied:
        return denied
    try:
        year, month = _parse_year_month(request)
    except Exception:
        return err('年份或月份无效')

    bus, e = _resolve_bu_list(request, request.GET.get('bu', ''))
    if e:
        return e

    tgt_index = _load_target_index(bus, year)
    actuals = _collect_actuals(bus, {year, year - 1})
    bu_rows = [_bu_metrics(bu, year, month, tgt_index, actuals) for bu in bus]

    overview = {
        # Headline cards: consolidated targets/actuals/rates + group MoM/YoY.
        'month': _attach_group_chg(_aggregate_total(bu_rows, 'month'), bus, year, month, actuals),
        'ytd': _aggregate_total(bu_rows, 'ytd'),
    }

    # 12-month trend: group actual revenue/profit + group target lines per month.
    trend = []
    for mo in range(1, 13):
        rev_a, prof_a, gross_a = _period_group(actuals, bus, year, mo)
        rev_t   = _sum_opt([tgt_index.get((bu, mo), {}).get('target_revenue') for bu in bus])
        prof_t  = _sum_opt([tgt_index.get((bu, mo), {}).get('target_profit') for bu in bus])
        gross_t = _sum_opt([tgt_index.get((bu, mo), {}).get('target_gross_profit') for bu in bus])
        trend.append({
            'month': mo,
            'actual_revenue': rev_a, 'actual_profit': prof_a, 'actual_gross_profit': gross_a,
            'target_revenue': rev_t, 'target_profit': prof_t, 'target_gross_profit': gross_t,
            'has_data': rev_a is not None or prof_a is not None,
        })

    return ok({
        'year': year, 'month': month,
        'overview': overview,
        'bus': bu_rows,
        'trend': trend,
    })


def _fmt_rate(v):
    return '—' if v is None else f'{v:.1f}%'


def _fmt_signed_pct(v):
    return '—' if v is None else f'{"+" if v >= 0 else ""}{v:.1f}%'


def _cockpit_data_lines(year, month, bus, bu_rows, ov_m, ov_y, actuals):
    """全集团 + 各事业部 + 12个月趋势的数据明细行（财务侧），供报告 prompt 与对话上下文复用。"""
    L = []
    L.append('【全集团合并概览】')
    L.append(f'  当月收入：{_fmt_wan(ov_m["actual_revenue"])}'
             f'（达成率{_fmt_rate(ov_m["revenue_rate"])}，环比{_fmt_signed_pct(ov_m.get("revenue_mom"))}，'
             f'同比{_fmt_signed_pct(ov_m.get("revenue_yoy"))}）')
    L.append(f'  当月利润：{_fmt_wan(ov_m["actual_profit"])}'
             f'（达成率{_fmt_rate(ov_m["profit_rate"])}，环比{_fmt_signed_pct(ov_m.get("profit_mom"))}，'
             f'同比{_fmt_signed_pct(ov_m.get("profit_yoy"))}）')
    L.append(f'  年度累计收入：{_fmt_wan(ov_y["actual_revenue"])}（年度目标达成率{_fmt_rate(ov_y["revenue_rate"])}）')
    L.append(f'  年度累计利润：{_fmt_wan(ov_y["actual_profit"])}（年度目标达成率{_fmt_rate(ov_y["profit_rate"])}）')

    L.append('')
    L.append('【各事业部表现（当月 / 年度累计）】')
    shown = 0
    for r in bu_rows:
        m, y = r['month'], r['ytd']
        if m['actual_revenue'] is None and m['actual_profit'] is None \
                and y['actual_revenue'] is None and y['actual_profit'] is None:
            continue
        shown += 1
        L.append(
            f'  {r["business_unit"]}：'
            f'当月收入{_fmt_wan(m["actual_revenue"])}(达成{_fmt_rate(m["revenue_rate"])}，'
            f'环比{_fmt_signed_pct(m.get("revenue_mom"))}，同比{_fmt_signed_pct(m.get("revenue_yoy"))})；'
            f'当月利润{_fmt_wan(m["actual_profit"])}(达成{_fmt_rate(m["profit_rate"])})；'
            f'YTD收入{_fmt_wan(y["actual_revenue"])}(达成{_fmt_rate(y["revenue_rate"])})，'
            f'YTD利润{_fmt_wan(y["actual_profit"])}(达成{_fmt_rate(y["profit_rate"])})'
        )
    if shown == 0:
        L.append('  （所选范围内各事业部均无已发布数据）')

    L.append('')
    L.append('【近12个月全集团趋势（实际收入/实际利润）】')
    tr = []
    for mo in range(1, 13):
        rev_a, prof_a, _gross_a = _period_group(actuals, bus, year, mo)
        if rev_a is None and prof_a is None:
            continue
        tr.append(f'{mo}月:收入{_fmt_wan(rev_a)}/利润{_fmt_wan(prof_a)}')
    L.append('  ' + ('；'.join(tr) if tr else '无'))
    return '\n'.join(L)


def _build_cockpit_prompt(year, month, bus, bu_rows, ov_m, ov_y, actuals):
    """Compose the group-level cockpit analysis prompt from consolidated +
    per-BU metrics and the 12-month group trend."""
    data = _cockpit_data_lines(year, month, bus, bu_rows, ov_m, ov_y, actuals)
    return f"""你正在为集团管理层解读 {year}年{month}月 的「财务驾驶舱」。以下是全集团及各事业部的经营数据（口径：已发布部门明细表，收入=主营业务收入，利润=经营净利；集团总部为成本中心，本身无收入）。

{data}

请站在全集团高度，输出一份综合、全面的经营分析报告，包含：

1. **集团经营总览**：本月与年度累计的整体经营态势——收入/利润规模、目标达成进度、环比同比趋势的综合研判。
2. **事业部横向对比**：识别领跑与掉队的事业部，分析达成率分化与利润贡献结构，指出谁是增长引擎、谁在拖累集团。
3. **目标达成与缺口**：结合当前节奏研判全年能否达标，量化关键缺口与达成压力最大的事业部。
4. **风险预警**：亏损或利润大幅下滑的事业部、异常波动、成本/费用异常、过度依赖单一事业部等系统性风险。
5. **战略与资源配置建议**：3-5条面向集团决策的具体建议（资源倾斜、整改重点、目标校准等）。

要求：专业、有数据支撑、有洞察，避免空话套话；分点清晰，适合集团领导决策参考，篇幅 800-1200 字。"""


def _build_report_messages(year, month, bus, period):
    """生成集团 月度/年度 经营分析报告的 messages（供 generate_report 技能用）。"""
    tgt_index = _load_target_index(bus, year)
    actuals = _collect_actuals(bus, {year, year - 1})
    bu_rows = [_bu_metrics(bu, year, month, tgt_index, actuals) for bu in bus]
    ov_m = _attach_group_chg(_aggregate_total(bu_rows, 'month'), bus, year, month, actuals)
    ov_y = _aggregate_total(bu_rows, 'ytd')
    data = _cockpit_data_lines(year, month, bus, bu_rows, ov_m, ov_y, actuals)
    if period == 'year':
        head = f'你正在为集团管理层撰写 {year}年度 经营分析报告（截至{month}月的累计即年度口径）。'
        focus = ('请输出年度经营分析报告：1)年度经营总览（累计收入/利润/目标达成）；'
                 '2)各事业部年度贡献与分化；3)全年目标完成研判与关键缺口；4)结构性/系统性风险；'
                 '5)面向来年的战略与资源配置建议。篇幅 1000-1500 字。')
    else:
        head = f'你正在为集团管理层撰写 {year}年{month}月 月度经营分析报告。'
        focus = ('请输出月度经营分析报告：1)当月经营总览；2)事业部横向对比；3)目标达成与缺口；'
                 '4)风险预警；5)下月行动建议。篇幅 800-1200 字。')
    prompt = (f'{head}\n口径：已发布部门明细表，收入=主营业务收入，利润=经营净利；集团总部为成本中心。\n\n'
              f'{data}\n\n{focus}\n要求：专业、有数据支撑、有洞察、分点清晰，避免空话套话。')
    return [{'role': 'system', 'content': _COCKPIT_AI_SYSTEM}, {'role': 'user', 'content': prompt}]


def _chunk_text(s, n=48):
    s = s or ''
    for i in range(0, len(s), n):
        yield s[i:i + n]


_COCKPIT_AI_SYSTEM = (
    '你是集团CFO级别的资深财务与经营分析专家，擅长从全集团高度做综合诊断、'
    '事业部横向对比与战略建议，用中文专业作答。'
)


def _cockpit_ai_prepare(request):
    """Shared validation + prompt build for both cockpit AI endpoints.

    Returns ((messages, scope), None) on success, or (None, err_response)."""
    if not settings.DEEPSEEK_API_KEY:
        return None, err('AI 分析未配置（缺少 DEEPSEEK_API_KEY）', 503)

    body = _parse_json(request)
    try:
        year = int(body.get('year'))
        month = int(body.get('month'))
        assert 2000 <= year <= 2100 and 1 <= month <= 12
    except Exception:
        return None, err('年份或月份无效')

    bus, e = _resolve_bu_list(request, (body.get('bu') or '').strip())
    if e:
        return None, e

    tgt_index = _load_target_index(bus, year)
    actuals = _collect_actuals(bus, {year, year - 1})
    bu_rows = [_bu_metrics(bu, year, month, tgt_index, actuals) for bu in bus]
    if not any(b['month']['actual_revenue'] is not None or b['ytd']['actual_revenue'] is not None
               for b in bu_rows):
        return None, err(f'{year}年{month}月该范围内暂无已发布数据，无法进行AI分析')

    ov_m = _attach_group_chg(_aggregate_total(bu_rows, 'month'), bus, year, month, actuals)
    ov_y = _aggregate_total(bu_rows, 'ytd')
    prompt = _build_cockpit_prompt(year, month, bus, bu_rows, ov_m, ov_y, actuals)
    messages = [
        {'role': 'system', 'content': _COCKPIT_AI_SYSTEM},
        {'role': 'user', 'content': prompt},
    ]
    scope = '全集团' if len(bus) > 1 else (bus[0] if bus else '全集团')
    return (messages, scope), None


@cw_required()
def cockpit_ai_analysis(request):
    """POST /cockpit/ai-analysis — 全集团综合分析（一次性返回，用 PRO 模型）。"""
    if request.method != 'POST':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'cockpit')
    if denied:
        return denied
    prep, e = _cockpit_ai_prepare(request)
    if e:
        return e
    messages, scope = prep
    try:
        text = _deepseek_chat(messages, timeout=180,
                              model=settings.DEEPSEEK_PRO_MODEL, max_tokens=3200)
        return ok({'analysis': text, 'model': settings.DEEPSEEK_PRO_MODEL, 'scope': scope})
    except Exception as ex:
        logger.error(f'Cockpit AI error: {ex}')
        return err(f'AI分析服务暂时不可用，请稍后重试（{str(ex)[:80]}）', 503)


@cw_required()
def cockpit_ai_analysis_stream(request):
    """POST /cockpit/ai-analysis/stream — 全集团综合分析，SSE 流式逐字推送。

    先流推理过程(reasoning)、再流正文(answer)，让领导秒级看到进度，不再干等。
    数据校验/无权限/无数据/无APIKey 仍在开流前以普通 JSON 错误返回。"""
    if request.method != 'POST':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'cockpit')
    if denied:
        return denied
    prep, e = _cockpit_ai_prepare(request)
    if e:
        return e
    messages, scope = prep
    model = settings.DEEPSEEK_PRO_MODEL

    def gen():
        yield _sse_event({'type': 'meta', 'scope': scope, 'model': model})
        try:
            for kind, delta in _deepseek_stream(messages, model=model,
                                                max_tokens=3200, timeout=300):
                yield _sse_event({'type': kind, 'delta': delta})
            yield _sse_event({'type': 'done'})
        except Exception as ex:
            logger.error(f'Cockpit AI stream error: {ex}')
            yield _sse_event({'type': 'error', 'error': f'AI分析服务暂时不可用（{str(ex)[:80]}）'})

    resp = StreamingHttpResponse(gen(), content_type='text/event-stream')
    resp['Cache-Control'] = 'no-cache'
    resp['X-Accel-Buffering'] = 'no'   # 关掉 nginx 缓冲，确保逐块下推
    return resp


# ── 业财融合 经营问答 Agent（财务驾驶舱内置对话）─────────────────────────────────

_COCKPIT_CHAT_SYSTEM = (
    '你是集团CFO级的「业财融合」经营分析智能助手，为集团管理层的经营决策赋能。'
    '你掌握全集团各事业部的财务数据（收入/成本/利润、目标达成、同比环比、12个月趋势）'
    '与业务数据（应收未收/回款/逾期、项目毛利等），擅长把"业务动因"与"财务结果"打通，'
    '做归因分析、风险预警与可执行建议。回答要求：'
    '①只依据【经营数据上下文】中的数据作答，数据缺失就如实说明、绝不编造数字；'
    '②中文、专业、简洁，有数据支撑、有洞察、给可执行建议；'
    '③用 Markdown（标题/列表/加粗）组织答案，避免空话套话；'
    '④口径：财务=已发布部门明细表（收入=主营业务收入，利润=经营净利，集团总部为成本中心）。'
)


def _build_ar_business_summary(bus, year, month):
    """业财融合「业」侧：应收未收 / 逾期账龄 / 本期回款（按交付部门）+ 逾期 Top 项目。
    最佳努力，异常或无数据返回空串。"""
    try:
        import datetime as _dt
        from ar.models import ARRecord, ARPayment
        today = _dt.date.today()
        out_map = {r['delivery_dept']: r['s'] for r in ARRecord.objects
                   .filter(delivery_dept__in=bus, outstanding_amount__gt=0)
                   .values('delivery_dept').annotate(s=Sum('outstanding_amount'))}
        od_map = {r['delivery_dept']: r['s'] for r in ARRecord.objects
                  .filter(delivery_dept__in=bus, outstanding_amount__gt=0, due_date__lt=today)
                  .values('delivery_dept').annotate(s=Sum('outstanding_amount'))}
        pay_map = {r['ar_record__delivery_dept']: r['s'] for r in ARPayment.objects
                   .filter(ar_record__delivery_dept__in=bus,
                           payment_date__year=year, payment_date__month=month)
                   .values('ar_record__delivery_dept').annotate(s=Sum('amount'))}
        if not (out_map or od_map or pay_map):
            return ''
        lines = ['【应收 / 回款（业务侧，交付部门口径，未收为当前快照）】']
        for bu in bus:
            o, od, p = out_map.get(bu), od_map.get(bu), pay_map.get(bu)
            if not (o or od or p):
                continue
            lines.append(f'  {bu}：未收余额{_fmt_wan(o)}（其中逾期{_fmt_wan(od)}）；{month}月回款{_fmt_wan(p)}')
        # 逾期账龄分布（全范围）
        od_qs = ARRecord.objects.filter(delivery_dept__in=bus, outstanding_amount__gt=0,
                                        due_date__lt=today).only('outstanding_amount', 'due_date')
        buckets = {'30天内': 0, '30-90天': 0, '90天以上': 0}
        for r in od_qs.iterator():
            days = (today - r.due_date).days if r.due_date else 0
            amt = float(r.outstanding_amount or 0)
            if days <= 30:
                buckets['30天内'] += amt
            elif days <= 90:
                buckets['30-90天'] += amt
            else:
                buckets['90天以上'] += amt
        if any(buckets.values()):
            lines.append('  逾期账龄：' + '；'.join(
                f'{k} {_fmt_wan(v)}' for k, v in buckets.items() if v))
        # 逾期 Top5 项目
        top = (ARRecord.objects.filter(delivery_dept__in=bus, outstanding_amount__gt=0,
                                       due_date__lt=today)
               .select_related('project').order_by('-outstanding_amount')[:5])
        names = []
        for r in top:
            nm = (r.project.short_name if r.project_id else None) or '—'
            names.append(f'{nm}({r.delivery_dept}) 逾期{_fmt_wan(r.outstanding_amount)}')
        if names:
            lines.append('  逾期最大项目：' + '；'.join(names))
        return '\n'.join(lines) if len(lines) > 1 else ''
    except Exception:
        return ''


def _build_project_margin_summary(bus, year, month):
    """业财融合补充：项目毛利（若已导入金蝶按项目核算账）。列各事业部毛利 Top3。"""
    try:
        from collections import defaultdict
        rows = list(ProjectMargin.objects.filter(business_unit__in=bus, year=year, month=month))
        if not rows:
            return ''
        by_bu = defaultdict(list)
        for m in rows:
            if (m.project_name or '').strip() in _PM_UNALLOCATED:
                continue
            by_bu[m.business_unit].append(m)
        if not by_bu:
            return ''
        lines = [f'【项目毛利（{year}年{month}月，收入−主营成本，直接口径）】']
        for bu in bus:
            ms = by_bu.get(bu)
            if not ms:
                continue
            ms.sort(key=lambda x: float(x.revenue) - float(x.cost), reverse=True)
            def _fmt(x):
                mg = float(x.revenue) - float(x.cost)
                rev = float(x.revenue)
                rate = f'，毛利率{mg / rev * 100:.0f}%' if rev else ''
                return f'{x.project_name} 毛利{_fmt_wan(mg)}(收入{_fmt_wan(rev)}{rate})'
            top = '；'.join(_fmt(x) for x in ms[:3])
            line = f'  {bu}：共{len(ms)}个项目，毛利Top：{top}'
            # 亏损项目（毛利<0）
            losers = [x for x in ms if float(x.revenue) - float(x.cost) < 0]
            if losers:
                losers.sort(key=lambda x: float(x.revenue) - float(x.cost))
                line += '；亏损项目：' + '、'.join(
                    f'{x.project_name}({_fmt_wan(float(x.revenue) - float(x.cost))})'
                    for x in losers[:3])
            lines.append(line)
        return '\n'.join(lines) if len(lines) > 1 else ''
    except Exception:
        return ''


_KNOWLEDGE_KIND_LABEL = {'insight': '洞察', 'background': '背景', 'rule': '口径'}


def _build_knowledge_context(bus):
    """注入已积累的经营知识库（长期记忆），让助手延续历史判断、越用越懂业务。"""
    try:
        scopes = set(bus) | {'全集团'}
        rows = list(CockpitKnowledge.objects.filter(scope__in=scopes)
                    .order_by('-pinned', '-created_at')[:40])
        if not rows:
            return ''
        lines = ['【已积累的经营知识库（历史沉淀，供延续判断与背景参考；若与最新数据冲突，以数据为准）】']
        for k in rows:
            tag = _KNOWLEDGE_KIND_LABEL.get(k.kind, k.kind)
            sc = '' if k.scope == '全集团' else f'[{k.scope}]'
            ttl = (k.title + '：') if k.title else ''
            lines.append(f'  ·（{tag}）{sc}{ttl}{k.content}')
        return '\n'.join(lines)
    except Exception:
        return ''


def _build_cockpit_data_pack(year, month, bus, bu_rows, ov_m, ov_y, actuals):
    """汇集"财+业"全口径经营数据，作为对话 Agent 的事实上下文。"""
    scope = '全集团' if len(bus) > 1 else (bus[0] if bus else '全集团')
    parts = [
        f'数据期间：{year}年{month}月　|　分析范围：{scope}　|　事业部清单：{"、".join(bus)}',
        '',
        _cockpit_data_lines(year, month, bus, bu_rows, ov_m, ov_y, actuals),
    ]
    ar = _build_ar_business_summary(bus, year, month)
    if ar:
        parts += ['', ar]
    pm = _build_project_margin_summary(bus, year, month)
    if pm:
        parts += ['', pm]
    return '\n'.join(parts)


def _cockpit_chat_prepare(request):
    """校验 + 组装对话 messages（system + 数据上下文 + 历史）。返回 ((messages, scope), None) 或 (None, err)。"""
    if not settings.DEEPSEEK_API_KEY:
        return None, err('AI 助手未配置（缺少 DEEPSEEK_API_KEY）', 503)
    body = _parse_json(request)
    try:
        year = int(body.get('year'))
        month = int(body.get('month'))
        assert 2000 <= year <= 2100 and 1 <= month <= 12
    except Exception:
        return None, err('年份或月份无效')
    bus, e = _resolve_bu_list(request, (body.get('bu') or '').strip())
    if e:
        return None, e
    # 供技能（如生成报告）默认取用当前对话的期间/范围
    request.chat_year, request.chat_month, request.chat_bus = year, month, bus

    # 清洗对话历史：仅保留 user/assistant，限制条数与单条长度，末条必须是用户提问
    history = []
    for m in (body.get('messages') or [])[-16:]:
        role = m.get('role')
        content = (m.get('content') or '').strip()
        if role in ('user', 'assistant') and content:
            history.append({'role': role, 'content': content[:4000]})
    if not history or history[-1]['role'] != 'user':
        return None, err('缺少用户提问')

    tgt_index = _load_target_index(bus, year)
    actuals = _collect_actuals(bus, {year, year - 1})
    bu_rows = [_bu_metrics(bu, year, month, tgt_index, actuals) for bu in bus]
    ov_m = _attach_group_chg(_aggregate_total(bu_rows, 'month'), bus, year, month, actuals)
    ov_y = _aggregate_total(bu_rows, 'ytd')
    data_pack = _build_cockpit_data_pack(year, month, bus, bu_rows, ov_m, ov_y, actuals)

    messages = [
        {'role': 'system', 'content': _COCKPIT_CHAT_SYSTEM},
        {'role': 'system', 'content': f'【经营数据上下文】\n{data_pack}'},
    ]
    knowledge = _build_knowledge_context(bus)
    if knowledge:
        messages.append({'role': 'system', 'content': knowledge})
    from caiwu import agent_skills
    brief = agent_skills.skills_brief()
    if brief:
        messages.append({'role': 'system', 'content':
                         f'你具备以下可调用技能（function-calling）：{brief}。'
                         '当用户的意图明确对应某个技能时（如"生成/出一份月度或年度经营分析报告""把这条记进知识库"），'
                         '请调用相应技能完成，而不是仅用文字描述；其余经营问答正常作答即可。'})
    messages += history
    scope = '全集团' if len(bus) > 1 else (bus[0] if bus else '全集团')
    return (messages, scope), None


@cw_required()
def cockpit_ai_chat_stream(request):
    """POST /cockpit/ai-chat/stream — 业财融合经营问答（多轮对话，SSE 流式，PRO 模型）。
    入参：{year, month, bu, messages:[{role,content}...]}（末条为用户提问）。"""
    if request.method != 'POST':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'cockpit')
    if denied:
        return denied
    prep, e = _cockpit_chat_prepare(request)
    if e:
        return e
    messages, scope = prep
    from caiwu import agent_skills
    tool_model = settings.DEEPSEEK_MODEL   # function-calling 走支持 tools 的对话模型

    def gen():
        yield _sse_event({'type': 'meta', 'scope': scope, 'model': tool_model})
        try:
            tools = agent_skills.agent_tools()
            convo = list(messages)
            for _step in range(4):  # 工具调用循环上限，防止无限调用
                msg = _deepseek_chat_raw(convo, tools=tools, model=tool_model,
                                         timeout=120, max_tokens=2000)
                tool_calls = msg.get('tool_calls')
                if not tool_calls:
                    for chunk in _chunk_text(msg.get('content') or '（未返回内容）'):
                        yield _sse_event({'type': 'answer', 'delta': chunk})
                    yield _sse_event({'type': 'done'})
                    return
                convo.append({'role': 'assistant', 'content': msg.get('content') or '',
                              'tool_calls': tool_calls})
                terminal_done = False
                for tc in tool_calls:
                    fn = tc.get('function') or {}
                    name = fn.get('name')
                    try:
                        a = json.loads(fn.get('arguments') or '{}')
                    except Exception:
                        a = {}
                    sk = agent_skills.get_skill(name)
                    yield _sse_event({'type': 'tool', 'name': name,
                                      'label': sk['label'] if sk else (name or '技能')})
                    if sk and sk.get('terminal') and sk.get('stream_handler'):
                        for kind, delta in sk['stream_handler'](request, a):
                            yield _sse_event({'type': kind, 'delta': delta})
                        terminal_done = True
                        break
                    try:
                        res = sk['handler'](request, a) if sk else {'ok': False, 'error': '未知技能'}
                    except Exception as ex:
                        res = {'ok': False, 'error': str(ex)[:120]}
                    convo.append({'role': 'tool', 'tool_call_id': tc.get('id'),
                                  'content': json.dumps(res, ensure_ascii=False)})
                if terminal_done:
                    yield _sse_event({'type': 'done'})
                    return
            yield _sse_event({'type': 'answer', 'delta': '（处理步骤过多，请把问题说得更具体些）'})
            yield _sse_event({'type': 'done'})
        except Exception as ex:
            logger.error(f'Cockpit chat stream error: {ex}')
            yield _sse_event({'type': 'error', 'error': f'AI助手暂时不可用（{str(ex)[:80]}）'})

    resp = StreamingHttpResponse(gen(), content_type='text/event-stream')
    resp['Cache-Control'] = 'no-cache'
    resp['X-Accel-Buffering'] = 'no'
    return resp


# ── 经营知识库（让 Agent 越用越聪明：长期记忆 + 自我提炼）─────────────────────────

def _knowledge_visible_scopes(request, bu):
    if bu and bu in VALID_BUSINESS_UNITS:
        return ['全集团', bu]
    if request.pk_role in ('super_admin', 'manager', 'general_manager'):
        visible = list(BUSINESS_UNITS)
    else:
        visible = [b for b in (request.pk_depts or []) if b in VALID_BUSINESS_UNITS]
    return ['全集团'] + visible


@cw_required()
def cockpit_knowledge(request):
    """GET 列出（按范围）/ POST 新增 一条经营知识。"""
    denied = _page_denied(request, 'cockpit')
    if denied:
        return denied
    if request.method == 'GET':
        scopes = _knowledge_visible_scopes(request, (request.GET.get('bu') or '').strip())
        rows = CockpitKnowledge.objects.filter(scope__in=scopes).order_by('-pinned', '-created_at')[:200]
        return ok({'items': [k.to_dict() for k in rows]})
    if request.method == 'POST':
        body = _parse_json(request)
        content = (body.get('content') or '').strip()
        if not content:
            return err('内容不能为空')
        scope = (body.get('scope') or '全集团').strip() or '全集团'
        if scope != '全集团' and scope not in VALID_BUSINESS_UNITS:
            scope = '全集团'
        if scope != '全集团' and not _can_access_bu(request, scope):
            return err('无权写入该事业部知识', 403)
        kind = body.get('kind') if body.get('kind') in ('insight', 'background', 'rule') else 'insight'
        k = CockpitKnowledge.objects.create(
            scope=scope, kind=kind, title=(body.get('title') or '')[:120].strip(),
            content=content[:2000], source=(body.get('source') or 'user'),
            pinned=bool(body.get('pinned')), created_by=request.pk_user)
        return ok(k.to_dict())
    return err('方法不允许', 405)


@cw_required()
def cockpit_knowledge_detail(request, kid):
    """PUT 置顶/编辑 / DELETE 删除 一条知识。"""
    denied = _page_denied(request, 'cockpit')
    if denied:
        return denied
    try:
        k = CockpitKnowledge.objects.get(id=kid)
    except CockpitKnowledge.DoesNotExist:
        return err('知识不存在', 404)
    if request.method == 'DELETE':
        if request.pk_role != 'super_admin' and k.created_by_id != request.pk_uid:
            return err('仅创建者或管理员可删除', 403)
        k.delete()
        return ok({'deleted': kid})
    if request.method == 'PUT':
        body = _parse_json(request)
        if 'pinned' in body:
            k.pinned = bool(body['pinned'])
        if (body.get('content') or '').strip():
            k.content = body['content'].strip()[:2000]
        if 'title' in body:
            k.title = (body['title'] or '')[:120].strip()
        k.save()
        return ok(k.to_dict())
    return err('方法不允许', 405)


def _extract_file_text(f):
    """从上传文件提取纯文本，尽量支持多格式。返回 (text, error)。
    原生支持：txt/md/csv/tsv/json/log/html 等文本类、xlsx/xls；
    pdf/docx 需服务器装 pypdf / python-docx（已列入 requirements，未装时给出提示）。"""
    name = (f.name or '').lower()
    raw = f.read()

    def _decode(b):
        for enc in ('utf-8', 'gb18030', 'latin-1'):
            try:
                return b.decode(enc)
            except Exception:
                continue
        return b.decode('utf-8', 'ignore')

    if name.endswith(('.xlsx', '.xlsm', '.xls')):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
            out = []
            for ws in wb.worksheets:
                out.append(f'# {ws.title}')
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c not in (None, '')]
                    if cells:
                        out.append('\t'.join(cells))
                    if len(out) > 6000:
                        break
            return '\n'.join(out), None
        except Exception as e:
            return '', f'Excel 解析失败：{e}'
    if name.endswith('.pdf'):
        try:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(raw))
            return '\n'.join((p.extract_text() or '') for p in reader.pages), None
        except ImportError:
            return '', 'PDF 解析需服务器安装 pypdf（已列入 requirements，部署后可用）'
        except Exception as e:
            return '', f'PDF 解析失败：{e}'
    if name.endswith(('.docx',)):
        try:
            import docx
            d = docx.Document(io.BytesIO(raw))
            return '\n'.join(p.text for p in d.paragraphs), None
        except ImportError:
            return '', 'Word 解析需服务器安装 python-docx（已列入 requirements，部署后可用）'
        except Exception as e:
            return '', f'Word 解析失败：{e}'
    # 其余按文本解码（txt/md/markdown/csv/tsv/json/log/html/yaml/…）
    return _decode(raw), None


@cw_required()
def cockpit_knowledge_import(request):
    """上传文件 → 提取文本 → （默认）AI 提炼为多条经营知识入库；支持大量格式。
    入参（multipart）：file、scope、mode(distill|raw)。"""
    if request.method != 'POST':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'cockpit')
    if denied:
        return denied
    if not _can_upload(request):
        return err('无导入权限（需上传权限）', 403)
    f = request.FILES.get('file')
    if not f:
        return err('请上传文件')
    if f.size > 12 * 1024 * 1024:
        return err('文件过大（上限12MB）')
    scope = (request.POST.get('scope') or '全集团').strip() or '全集团'
    if scope != '全集团' and scope not in VALID_BUSINESS_UNITS:
        scope = '全集团'
    if scope != '全集团' and not _can_access_bu(request, scope):
        return err('无权写入该事业部知识', 403)
    mode = request.POST.get('mode', 'distill')

    text, e = _extract_file_text(f)
    if e:
        return err(e)
    text = (text or '').strip()
    if not text:
        return err('未能从文件中提取到文本内容')
    text = text[:24000]

    entries = []
    if mode == 'distill' and settings.DEEPSEEK_API_KEY:
        sys = ('你是经营分析助手。从用户提供的文档内容中提炼出若干条（最多10条）'
               '可长期复用的经营知识/背景/口径。只输出 JSON 数组：'
               '[{"title":"≤16字小标题","content":"≤140字要点，保留关键数字与结论"}…]，只输出JSON。')
        try:
            raw_out = _deepseek_chat([{'role': 'system', 'content': sys},
                                      {'role': 'user', 'content': text}],
                                     timeout=120, max_tokens=2000)
            s = raw_out.strip()
            arr = json.loads(s[s.find('['): s.rfind(']') + 1])
            for it in arr[:10]:
                c = (it.get('content') or '').strip()
                if c:
                    entries.append((( it.get('title') or '')[:120], c[:2000]))
        except Exception as ex:
            logger.error(f'knowledge import distill error: {ex}')
            entries = []  # 回退到原文切块

    if not entries:
        # 原文切块（每块≤1500字），标题用文件名
        base = (f.name or '导入').rsplit('.', 1)[0][:80]
        chunk, n = [], 0
        cur = ''
        for para in text.split('\n'):
            if len(cur) + len(para) > 1500:
                if cur.strip():
                    n += 1
                    entries.append((f'{base} #{n}', cur.strip()[:2000]))
                cur = ''
            cur += para + '\n'
        if cur.strip():
            n += 1
            entries.append((f'{base} #{n}', cur.strip()[:2000]))
        entries = entries[:30]

    if not entries:
        return err('未能从文件生成知识条目')

    # 去重：同范围内内容完全一致的不重复入库
    existing = set(CockpitKnowledge.objects.filter(scope=scope).values_list('content', flat=True))
    src = 'ai' if (mode == 'distill' and settings.DEEPSEEK_API_KEY) else 'user'
    objs, seen = [], set(existing)
    for t, c in entries:
        if c in seen:
            continue
        seen.add(c)
        objs.append(CockpitKnowledge(scope=scope, kind='background', title=t, content=c,
                                     source=src, created_by=request.pk_user))
    if not objs:
        return err('文件内容均已在知识库中（已去重）')
    CockpitKnowledge.objects.bulk_create(objs)
    return ok({'created': len(objs), 'skipped': len(entries) - len(objs),
               'mode': mode, 'scope': scope, 'file': f.name})


@cw_required()
def cockpit_knowledge_distill(request):
    """把一段 AI 分析自我提炼为一条可长期复用的经营知识并入库（Agent 自我总结、积累）。"""
    if request.method != 'POST':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'cockpit')
    if denied:
        return denied
    if not settings.DEEPSEEK_API_KEY:
        return err('AI 未配置', 503)
    body = _parse_json(request)
    text = (body.get('text') or '').strip()
    if not text:
        return err('缺少待提炼内容')
    scope = (body.get('scope') or '全集团').strip() or '全集团'
    if scope != '全集团' and scope not in VALID_BUSINESS_UNITS:
        scope = '全集团'
    if scope != '全集团' and not _can_access_bu(request, scope):
        return err('无权写入该事业部知识', 403)
    sys = ('你是经营分析助手。把用户给的一段分析提炼为一条简洁、可长期复用的经营知识。'
           '只输出 JSON：{"title":"≤16字小标题","content":"≤140字要点，保留关键数字与结论，不要套话"}。')
    try:
        raw = _deepseek_chat([{'role': 'system', 'content': sys},
                              {'role': 'user', 'content': text[:4000]}],
                             timeout=60, max_tokens=400)
    except Exception as ex:
        logger.error(f'knowledge distill error: {ex}')
        return err('提炼失败，请稍后重试', 503)
    title, content = '', ''
    try:
        s = raw.strip()
        obj = json.loads(s[s.find('{'): s.rfind('}') + 1])
        title = (obj.get('title') or '')[:120]
        content = (obj.get('content') or '').strip()
    except Exception:
        content = text[:280]
    if not content:
        content = text[:280]
    k = CockpitKnowledge.objects.create(scope=scope, kind='insight', title=title,
                                        content=content[:2000], source='ai',
                                        created_by=request.pk_user)
    return ok(k.to_dict())


# ── Agent 技能（为后续可调用技能/工具预留的注册表 + 列表/执行接口）──────────────
from caiwu import agent_skills  # noqa: E402


@agent_skills.register_skill(
    'save_knowledge', '记入知识库', '把一条经营知识/背景/口径存入知识库（长期记忆）',
    {'content': '知识内容(必填)', 'scope': '范围(事业部或全集团,可选)', 'kind': 'insight|background|rule(可选)'},
    tool=True)
def _skill_save_knowledge(request, args):
    content = (args.get('content') or '').strip()
    if not content:
        return {'ok': False, 'error': '内容为空'}
    scope = (args.get('scope') or '全集团').strip() or '全集团'
    if scope != '全集团' and (scope not in VALID_BUSINESS_UNITS or not _can_access_bu(request, scope)):
        scope = '全集团'
    kind = args.get('kind') if args.get('kind') in ('insight', 'background', 'rule') else 'insight'
    k = CockpitKnowledge.objects.create(scope=scope, kind=kind, title=(args.get('title') or '')[:120],
                                        content=content[:2000], source='ai', created_by=request.pk_user)
    return {'ok': True, 'data': k.to_dict()}


@agent_skills.register_skill(
    'search_knowledge', '检索知识库', '按关键词检索已积累的经营知识', {'query': '关键词'},
    tool=True)
def _skill_search_knowledge(request, args):
    q = (args.get('query') or '').strip()
    qs = CockpitKnowledge.objects.filter(scope__in=_knowledge_visible_scopes(request, ''))
    if q:
        qs = qs.filter(Q(content__icontains=q) | Q(title__icontains=q))
    return {'ok': True, 'data': [k.to_dict() for k in qs[:20]]}


@agent_skills.register_skill(
    'forget_knowledge', '清理知识库', '按关键词删除知识条目（用于对话式清理）', {'query': '关键词'})
def _skill_forget_knowledge(request, args):
    q = (args.get('query') or '').strip()
    if not q:
        return {'ok': False, 'error': '需提供关键词'}
    qs = CockpitKnowledge.objects.filter(scope__in=_knowledge_visible_scopes(request, '')).filter(
        Q(content__icontains=q) | Q(title__icontains=q))
    if request.pk_role != 'super_admin':
        qs = qs.filter(created_by_id=request.pk_uid)
    ids = list(qs.values_list('id', flat=True))
    qs.delete()
    return {'ok': True, 'data': {'deleted': len(ids), 'ids': ids}}


def _resolve_report_args(request, args):
    import datetime as _dt
    today = _dt.date.today()
    try:
        year = int(args.get('year') or getattr(request, 'chat_year', 0) or 0)
    except Exception:
        year = 0
    try:
        month = int(args.get('month') or getattr(request, 'chat_month', 0) or 0)
    except Exception:
        month = 0
    bus = list(getattr(request, 'chat_bus', None) or [])
    if not bus:
        bus, _e = _resolve_bu_list(request, '')
    bu_arg = (args.get('bu') or '').strip()
    if bu_arg and bu_arg in VALID_BUSINESS_UNITS and _can_access_bu(request, bu_arg):
        bus = [bu_arg]
    period = 'year' if (args.get('period') or '').strip() in ('year', '年', '全年', '年度', 'annual') else 'month'
    if not (2000 <= year <= 2100):
        year = today.year
    if not (1 <= month <= 12):
        month = today.month
    return year, month, bus, period


def _skill_generate_report_stream(request, args):
    year, month, bus, period = _resolve_report_args(request, args)
    label = '年度' if period == 'year' else f'{month}月'
    yield ('answer', f'### 集团 {year}年{label} 经营分析报告\n\n')
    for kind, delta in _deepseek_stream(_build_report_messages(year, month, bus, period),
                                        model=settings.DEEPSEEK_PRO_MODEL, max_tokens=3500, timeout=300):
        yield (kind, delta)


@agent_skills.register_skill(
    'generate_report', '生成经营分析报告', '生成集团全年或某月的经营分析报告（数据驱动，含目标达成、事业部对比、风险与建议）',
    {'period': 'year 或 month（必填，年度填year）', 'year': '年份(可选)', 'month': '月份(可选)', 'bu': '事业部(可选,默认全集团)'},
    tool=True, terminal=True, stream_handler=_skill_generate_report_stream)
def _skill_generate_report(request, args):
    year, month, bus, period = _resolve_report_args(request, args)
    text = _deepseek_chat(_build_report_messages(year, month, bus, period),
                          timeout=180, model=settings.DEEPSEEK_PRO_MODEL, max_tokens=3500)
    return {'ok': True, 'data': {'report': text, 'year': year, 'month': month, 'period': period}}


@cw_required()
def cockpit_skills(request):
    """GET 列出 Agent 可用技能（供 UI 展示 / 后续 function-calling）。"""
    denied = _page_denied(request, 'cockpit')
    if denied:
        return denied
    return ok({'skills': agent_skills.list_skills()})


@cw_required()
def cockpit_skill_run(request):
    """POST {name, args} 执行一个技能（为对话即操作预留的统一入口）。"""
    if request.method != 'POST':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'cockpit')
    if denied:
        return denied
    body = _parse_json(request)
    skill = agent_skills.get_skill(body.get('name'))
    if not skill:
        return err('技能不存在', 404)
    try:
        res = skill['handler'](request, body.get('args') or {})
    except Exception as ex:
        logger.error(f'skill run error: {ex}')
        return err(f'技能执行失败（{str(ex)[:80]}）')
    if not res.get('ok'):
        return err(res.get('error') or '执行失败')
    return ok(res.get('data'))


# ── charts ───────────────────────────────────────────────────────────────────

@cw_required()
def chart_trend(request):
    """
    ?bu=&year=
    Returns 12-month trend for the selected BU/year.
    Aggregates by month: total revenue (first L1), total entries sum per month.
    Frontend decides which L1 to display.
    """
    if request.method != 'GET':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'charts')
    if denied:
        return denied
    if not _can_view(request, 'chart_trend'):
        return err('无访问权限', 403, 403)

    bu = request.GET.get('bu', '').strip()
    year_s = request.GET.get('year', '')

    try:
        year = int(year_s)
        assert 2000 <= year <= 2100
    except Exception:
        return err('年份无效')

    if not bu:
        return err('请指定事业部')
    if not _can_access_bu(request, bu):
        return err('无权访问该事业部', 403)

    l1_cats = list(L1Category.objects.order_by('sort_order', 'id'))

    # For each month, get published batch and aggregate by L1 (including calculated rows)
    result = []
    for month in range(1, 13):
        batches_qs = _get_published_batches([bu], year, month)
        batch_ids = list(batches_qs.values_list('id', flat=True))
        if batch_ids:
            agg = (
                FinancialEntry.objects
                .filter(batch_id__in=batch_ids)
                .values('l1_id')
                .annotate(amount=Sum('amount'))
            )
            raw_by_id = {r['l1_id']: float(r['amount']) for r in agg}
            _, id_map = _compute_l1_name_map(l1_cats, raw_by_id)
            month_data = id_map
        else:
            month_data = {}
        result.append({
            'month': month,
            'has_data': bool(batch_ids),
            'by_l1': month_data,
        })

    return ok({
        'year': year,
        'bu': bu,
        'months': result,
        'l1_categories': [c.to_dict() for c in l1_cats],
    })


@cw_required()
def chart_waterfall(request):
    """
    ?bu=&year=&month=&compare_year=&compare_month=
    Factor analysis between two periods using is_profit_driver L1 categories.
    Returns waterfall chart data.
    """
    if request.method != 'GET':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'charts')
    if denied:
        return denied
    if not _can_view(request, 'chart_waterfall'):
        return err('无访问权限', 403, 403)

    bu = request.GET.get('bu', '').strip()
    year_s = request.GET.get('year', '')
    month_s = request.GET.get('month', '')
    cmp_year_s = request.GET.get('compare_year', year_s)
    cmp_month_s = request.GET.get('compare_month', '')

    if not bu:
        return err('请指定事业部')
    if not _can_access_bu(request, bu):
        return err('无权访问该事业部', 403)

    try:
        year = int(year_s)
        month = int(month_s)
        cmp_year = int(cmp_year_s)
        cmp_month = int(cmp_month_s)
        assert all(2000 <= y <= 2100 for y in (year, cmp_year))
        assert all(1 <= m <= 12 for m in (month, cmp_month))
    except Exception:
        return err('年份或月份无效')

    l1_cats = list(L1Category.objects.order_by('sort_order', 'id'))

    def _get_l1_totals(yr, mo):
        batches_qs = _get_published_batches([bu], yr, mo)
        batch_ids = list(batches_qs.values_list('id', flat=True))
        if not batch_ids:
            return {}, {}
        agg = (
            FinancialEntry.objects
            .filter(batch_id__in=batch_ids)
            .values('l1_id')
            .annotate(amount=Sum('amount'))
        )
        raw_by_id = {r['l1_id']: float(r['amount']) for r in agg}
        name_map, id_map = _compute_l1_name_map(l1_cats, raw_by_id)
        return name_map, id_map

    base_name_map, base_id_map = _get_l1_totals(cmp_year, cmp_month)
    curr_name_map, curr_id_map = _get_l1_totals(year, month)

    def _safe_sum(id_map, cats):
        # Signed sum of raw rows == 经营净利 by construction (revenue +1, cost -1).
        # Must apply sign, else costs would add to (instead of subtract from) profit.
        return sum((id_map.get(l1.id, 0) or 0) * l1.sign for l1 in cats if not l1.is_calculated)

    # Headline profit = 经营净利 if available; otherwise sum raw entries
    base_total = base_name_map.get('经营净利') if '经营净利' in base_name_map else _safe_sum(base_id_map, l1_cats)
    curr_total = curr_name_map.get('经营净利') if '经营净利' in curr_name_map else _safe_sum(curr_id_map, l1_cats)
    base_total = base_total or 0
    curr_total = curr_total or 0

    # Factor analysis: ALL raw L1 categories (not calculated rows) form a complete
    # bridge so base + Σdeltas == current exactly. is_profit_driver only flags
    # which to emphasise; every changed raw item must appear or the bridge breaks.
    drivers = [l1 for l1 in l1_cats if not l1.is_calculated]

    factors = []
    for l1 in drivers:
        base_v = base_id_map.get(l1.id) or 0
        curr_v = curr_id_map.get(l1.id) or 0
        raw_delta = curr_v - base_v
        profit_delta = raw_delta * l1.sign  # sign=-1 (cost) → cost increase → negative profit impact
        factors.append({
            'l1_id': l1.id,
            'name': l1.name,
            'base': round(base_v, 2),
            'current': round(curr_v, 2),
            'delta': round(profit_delta, 2),
            'is_driver': l1.is_profit_driver,
        })

    # Waterfall: base → factor deltas → current total
    waterfall = [
        {'name': f'{cmp_year}年{cmp_month}月', 'value': round(base_total, 2), 'type': 'base'},
    ]
    for f in factors:
        if f['delta'] != 0:  # skip zero-delta factors to reduce noise
            waterfall.append({
                'name': f['name'],
                'value': f['delta'],
                'type': 'increase' if f['delta'] >= 0 else 'decrease',
            })
    # Close the bridge: the headline total (经营净利) is a calculated row, while the
    # factor bars are signed deltas of the raw rows. With the standard taxonomy these
    # reconcile exactly, but stray raw categories outside 经营净利's formula leave a
    # residual — surface it as an explicit factor so base + Σfactors == current and the
    # last bar always lands on the terminal column instead of floating short of it.
    residual = round((curr_total - base_total) - sum(f['delta'] for f in factors), 2)
    if abs(residual) >= 0.01:
        waterfall.append({
            'name': '其他',
            'value': residual,
            'type': 'increase' if residual >= 0 else 'decrease',
        })
    waterfall.append({'name': f'{year}年{month}月', 'value': round(curr_total, 2), 'type': 'total'})

    return ok({
        'bu': bu,
        'base_period': {'year': cmp_year, 'month': cmp_month, 'total': base_total},
        'current_period': {'year': year, 'month': month, 'total': curr_total},
        'factors': factors,
        'waterfall': waterfall,
    })


# ── AI analysis (DeepSeek) ───────────────────────────────────────────────────

def _deepseek_chat(messages, timeout=90, model=None, max_tokens=1800):
    """Call DeepSeek chat completion API. Returns response text or raises.

    `model` overrides settings.DEEPSEEK_MODEL — used by the cockpit's group-level
    analysis to invoke the stronger DEEPSEEK_PRO_MODEL with a larger token budget.
    """
    import requests as req_lib
    resp = req_lib.post(
        f'{settings.DEEPSEEK_BASE_URL}/chat/completions',
        headers={
            'Authorization': f'Bearer {settings.DEEPSEEK_API_KEY}',
            'Content-Type': 'application/json',
        },
        json={
            'model': model or settings.DEEPSEEK_MODEL,
            'messages': messages,
            'temperature': 0.3,
            'max_tokens': max_tokens,
        },
        timeout=timeout,
    )
    resp.raise_for_status()
    return resp.json()['choices'][0]['message']['content']


def _deepseek_chat_raw(messages, tools=None, model=None, timeout=90, max_tokens=1800):
    """调用 DeepSeek（支持 function-calling），返回完整 message dict（含可能的 tool_calls）。"""
    import requests as req_lib
    payload = {
        'model': model or settings.DEEPSEEK_MODEL,
        'messages': messages,
        'temperature': 0.3,
        'max_tokens': max_tokens,
    }
    if tools:
        payload['tools'] = tools
        payload['tool_choice'] = 'auto'
    resp = req_lib.post(
        f'{settings.DEEPSEEK_BASE_URL}/chat/completions',
        headers={'Authorization': f'Bearer {settings.DEEPSEEK_API_KEY}',
                 'Content-Type': 'application/json'},
        json=payload, timeout=timeout)
    resp.raise_for_status()
    return resp.json()['choices'][0]['message']


def _deepseek_stream(messages, model=None, max_tokens=1800, timeout=300):
    """Yield (kind, delta) from a streaming DeepSeek completion.

    kind ∈ {'reasoning', 'answer'}: reasoner models emit reasoning_content first
    (the chain-of-thought) then content (the final answer). Streaming both lets the
    UI show activity within seconds instead of waiting for the whole response.
    """
    import requests as req_lib
    resp = req_lib.post(
        f'{settings.DEEPSEEK_BASE_URL}/chat/completions',
        headers={
            'Authorization': f'Bearer {settings.DEEPSEEK_API_KEY}',
            'Content-Type': 'application/json',
        },
        json={
            'model': model or settings.DEEPSEEK_MODEL,
            'messages': messages,
            'temperature': 0.3,
            'max_tokens': max_tokens,
            'stream': True,
        },
        timeout=timeout,
        stream=True,
    )
    resp.raise_for_status()
    # Parse SSE lines as raw bytes → utf-8 (line boundaries never split multibyte chars).
    for raw in resp.iter_lines(decode_unicode=False):
        if not raw:
            continue
        line = raw.decode('utf-8', 'ignore')
        if not line.startswith('data:'):
            continue
        payload = line[5:].strip()
        if payload == '[DONE]':
            break
        try:
            delta = json.loads(payload)['choices'][0]['delta']
        except (ValueError, KeyError, IndexError):
            continue
        rc = delta.get('reasoning_content')
        if rc:
            yield ('reasoning', rc)
        c = delta.get('content')
        if c:
            yield ('answer', c)


def _sse_event(obj):
    """Encode one Server-Sent-Events frame."""
    return f'data: {json.dumps(obj, ensure_ascii=False)}\n\n'


def _fmt_wan(v):
    if v is None:
        return '无数据'
    try:
        v = float(v)
    except (TypeError, ValueError):
        return '无数据'
    sign = '-' if v < 0 else ''
    abs_v = abs(v)
    if abs_v >= 1e8:
        return f'{sign}{abs_v/1e8:.2f}亿元'
    if abs_v >= 1e4:
        return f'{sign}{abs_v/1e4:.2f}万元'
    return f'{sign}{abs_v:.2f}元'


_REPORT_AI_SYSTEM = '你是一位专业的企业财务分析师，擅长财务报表解读和经营决策分析，用中文回答。'


def _report_ai_prepare(request):
    """Shared validation + prompt build for the report AI endpoints.
    Returns (messages, None) on success or (None, err_response).

    Body: {year, month, bu? | bus?}"""
    if not settings.DEEPSEEK_API_KEY:
        return None, err('AI 分析未配置（缺少 DEEPSEEK_API_KEY）', 503)

    body = _parse_json(request)
    try:
        year = int(body.get('year'))
        month = int(body.get('month'))
        assert 2000 <= year <= 2100 and 1 <= month <= 12
    except Exception:
        return None, err('年份或月份无效')

    # Resolve BU scope from request, clamped to what this user may access.
    if request.pk_role == 'super_admin':
        accessible = list(BUSINESS_UNITS)
    else:
        accessible = [d for d in request.pk_depts if d in VALID_BUSINESS_UNITS]

    single_bu = (body.get('bu') or '').strip()
    req_bus = body.get('bus') or []
    if single_bu:
        if single_bu not in accessible:
            return None, err('无权访问该事业部', 403)
        bu_list = [single_bu]
        bu = single_bu
        bu_scope = ''
    elif req_bus:
        bu_list = [b for b in req_bus if b in accessible]
        if not bu_list:
            return None, err('无权访问所选事业部', 403)
        bu = '全集团' if len(bu_list) > 1 else bu_list[0]
        bu_scope = 'all' if len(bu_list) > 1 else ''
    else:
        bu_list = accessible
        bu = '全集团' if len(bu_list) != 1 else bu_list[0]
        bu_scope = 'all' if len(bu_list) > 1 else ''

    # Aggregate report data server-side (single source of truth).
    rows = _aggregate_report(_get_published_batches(bu_list, year, month), 1)
    if not rows:
        return None, err(f'{year}年{month}月该范围内暂无已发布数据，无法进行AI分析')

    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    prev_rows = _aggregate_report(_get_published_batches(bu_list, prev_year, prev_month), 1)
    prev_kpis = {r['l1_name']: float(r['amount']) for r in prev_rows}

    # Build KPI summary with MoM change
    KPI_NAMES = ['主营业务收入', '主营业务成本', '运营毛利', '经营毛利', '经营净利']
    kpi_map = {r['l1_name']: r.get('amount') for r in rows}

    kpi_lines = []
    for name in KPI_NAMES:
        curr = kpi_map.get(name)
        prev = prev_kpis.get(name)
        line = f'  {name}：{_fmt_wan(curr)}'
        if curr is not None and prev is not None and prev != 0:
            delta_pct = (float(curr) - float(prev)) / abs(float(prev)) * 100
            trend = '↑' if delta_pct > 0 else '↓'
            line += f'（环比{trend}{abs(delta_pct):.1f}%，上月{_fmt_wan(prev)}）'
        elif curr is not None and prev is not None:
            line += f'（上月{_fmt_wan(prev)}，环比基数为零）'
        kpi_lines.append(line)

    # Other L1 items
    other_lines = []
    for r in rows:
        if r['l1_name'] not in KPI_NAMES and not r.get('is_calculated'):
            other_lines.append(f'  {r["l1_name"]}：{_fmt_wan(r.get("amount"))}')

    scope_note = ''
    if bu_scope == 'all':
        scope_note = '（以下为全集团合并口径，已排除集团总部内部往来，代表5个核心事业部合并情况）'

    prompt = f"""你是一位资深企业财务分析师，请对以下 {year}年{month}月 {bu} 的财务数据进行专业分析。{scope_note}

【关键财务指标】
{chr(10).join(kpi_lines)}

【其他科目明细】
{chr(10).join(other_lines) if other_lines else '  （无其他科目数据）'}

请从纯财务角度（业财融合分析将在后续版本中加入）进行以下分析：

1. **财务状况综合评价**（2-3句话，突出核心亮点或问题）
2. **关键指标分析**（重点分析盈利能力、成本管控，如有环比数据请分析变动趋势）
3. **风险提示**（负利润/异常波动/成本偏高等需警示的情况，如无风险可简要说明）
4. **决策建议**（2-3条具体可操作的财务管理建议）

要求：语言简明专业，适合集团领导决策参考，不超过600字。"""

    messages = [
        {'role': 'system', 'content': _REPORT_AI_SYSTEM},
        {'role': 'user', 'content': prompt},
    ]
    return messages, None


@cw_required()
def report_ai_analysis(request):
    """POST /report/ai-analysis — 单期财务分析（一次性返回）。"""
    if request.method != 'POST':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'report')
    if denied:
        return denied
    messages, e = _report_ai_prepare(request)
    if e:
        return e
    try:
        text = _deepseek_chat(messages)
        return ok({'analysis': text})
    except Exception as ex:
        logger.error(f'DeepSeek AI error: {ex}')
        return err(f'AI分析服务暂时不可用，请稍后重试（{str(ex)[:80]}）', 503)


@cw_required()
def report_ai_analysis_stream(request):
    """POST /report/ai-analysis/stream — 单期财务分析，SSE 流式逐字推送。"""
    if request.method != 'POST':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'report')
    if denied:
        return denied
    messages, e = _report_ai_prepare(request)
    if e:
        return e

    def gen():
        yield _sse_event({'type': 'meta'})
        try:
            for kind, delta in _deepseek_stream(messages):
                yield _sse_event({'type': kind, 'delta': delta})
            yield _sse_event({'type': 'done'})
        except Exception as ex:
            logger.error(f'DeepSeek AI stream error: {ex}')
            yield _sse_event({'type': 'error', 'error': f'AI分析服务暂时不可用（{str(ex)[:80]}）'})

    resp = StreamingHttpResponse(gen(), content_type='text/event-stream')
    resp['Cache-Control'] = 'no-cache'
    resp['X-Accel-Buffering'] = 'no'
    return resp


@cw_required()
def chart_ai_analysis(request):
    """POST /charts/ai-analysis
    Body: {chart_type: 'trend'|'waterfall', bu, year, month?, data: {...}}
    """
    if request.method != 'POST':
        return err('方法不允许', 405)
    denied = _page_denied(request, 'charts')
    if denied:
        return denied
    if not settings.DEEPSEEK_API_KEY:
        return err('AI 分析未配置（缺少 DEEPSEEK_API_KEY）', 503)

    body = _parse_json(request)
    chart_type = body.get('chart_type', 'trend')
    bu = body.get('bu', '未知事业部')
    year = body.get('year')
    data = body.get('data', {})

    if chart_type == 'trend':
        # Build trend narrative from 12-month data
        months_data = data.get('months', [])
        l1_cats = data.get('l1_categories', [])
        selected_ids = body.get('selected_l1_ids', [])
        sel_cats = [c for c in l1_cats if c['id'] in selected_ids] or l1_cats[:3]

        trend_lines = []
        for cat in sel_cats:
            vals = []
            for m in months_data:
                v = m.get('by_l1', {}).get(str(cat['id']))
                vals.append(f'{m["month"]}月:{_fmt_wan(v)}' if v is not None else f'{m["month"]}月:无')
            trend_lines.append(f'  {cat["name"]}：{", ".join(vals)}')

        prompt = f"""{year}年 {bu} 收入/利润走势数据如下：

{chr(10).join(trend_lines)}

请从财务角度分析：
1. 全年走势特征（旺季/淡季规律、增长趋势）
2. 月度异常波动（如某月大幅上升/下降的原因判断）
3. 结合该事业部全局情况的简要评价
4. 一条具体的改善建议

要求：简明专业，不超过300字。"""

    elif chart_type == 'waterfall':
        base = data.get('base_period', {})
        curr = data.get('current_period', {})
        factors = data.get('factors', [])
        factor_lines = [f'  {f["name"]}：变动{_fmt_wan(f["delta"])}（对利润影响）' for f in factors if f.get('delta')]

        prompt = f"""{bu} 利润因素分析（{base.get("year")}年{base.get("month")}月 vs {curr.get("year")}年{curr.get("month")}月）：

  基期经营净利：{_fmt_wan(base.get("total"))}
  当期经营净利：{_fmt_wan(curr.get("total"))}
  净变动：{_fmt_wan((curr.get("total") or 0) - (base.get("total") or 0))}

【驱动因素分解】
{chr(10).join(factor_lines) if factor_lines else "  暂无驱动因素数据"}

请从财务角度分析：
1. 利润变动的主要驱动因素评价
2. 正/负因素各自的关键点
3. 结合该事业部全局情况的评价
4. 一条针对性建议

要求：简明专业，不超过250字。"""
    else:
        return err('未知图表类型')

    try:
        text = _deepseek_chat([
            {'role': 'system', 'content': '你是一位专业的企业财务分析师，擅长图表数据解读，用中文简洁回答。'},
            {'role': 'user', 'content': prompt},
        ])
        return ok({'analysis': text})
    except Exception as e:
        logger.error(f'DeepSeek chart AI error: {e}')
        return err(f'AI分析暂时不可用（{str(e)[:80]}）', 503)
