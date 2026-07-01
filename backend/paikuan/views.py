import io
import json
import logging
import datetime
import functools
import re
import threading
import time
from decimal import Decimal, InvalidOperation
from urllib.parse import quote

from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction, IntegrityError
from django.db.models import F, Q, Sum, Count, Case, When, ExpressionWrapper, Value, OuterRef, Subquery, IntegerField
from django.db.models import DecimalField as DjDecimalField
from django.db.models.functions import Coalesce, Greatest
from django.conf import settings
from django.utils import timezone
import jwt

from paikuan.models import (PaikuanUser, Payment, PaymentInstallment, PaymentPlanItem,
                            JobPermission, ApprovalRecord, PaymentChangeLog)
from paikuan.list_filters import build_filter_q, resolve_sort

logger = logging.getLogger('paikuan')

DEPARTMENTS = [
    '集团总部', '劳务事业部', '运输事业部', '自营事业部',
    '阔展事业部', '多式联运事业部', '供应链事业部',
]
VALID_DEPARTMENTS = set(DEPARTMENTS)

# Job titles double as the permission "roles" configured by super_admin.
JOB_TITLES = {
    'finance_director': '财务总监',
    'finance_bp': '财务BP',
    'chief_cashier': '总出纳',
    'cashier': '出纳',
    'general_manager': '总经理',
    'gm_assistant': '总经理助理',
    'settlement_accountant': '结算会计',
    'sales_bp': '销售BP',
}

# ── Excel 风格列头筛选/排序的列注册表（白名单）──────────────────────────────────
# 只有此处登记的字段才能被前端 filters/sort 命中。type 决定可用运算符；
# col 为 ORM lookup；multi=True 表示反向关联(需 .distinct())；sortable=False 禁排序。
APPROVAL_FILTER_REGISTRY = {
    'applicant':          {'type': 'text',   'col': 'applicant'},
    'department':         {'type': 'enum',   'col': 'department'},
    'secondary_dept':     {'type': 'text',   'col': 'secondary_dept'},
    'project_short_name': {'type': 'text',   'col': 'project_short_name'},
    'approval_number':    {'type': 'text',   'col': 'approval_number'},
    'g7_number':          {'type': 'text',   'col': 'g7_number'},
    'summary':            {'type': 'text',   'col': 'summary'},
    'notes':              {'type': 'text',   'col': 'notes'},
    'amount':             {'type': 'number', 'col': 'amount'},
    'scheduled_amount':   {'type': 'number', 'col': 'scheduled_amount'},
    'payee':              {'type': 'text',   'col': 'payee'},
    'status':             {'type': 'enum',   'col': 'status'},
}

# 审批管理「计算列」未排金额 = max(0, 申请额 − 已排额)。需先注解再走通用数值解析器。
APPROVAL_COMPUTED_REGISTRY = {
    'remaining_amount': {'type': 'number', 'col': 'remaining_calc'},
}

# 付款管理：本期只登记真实 DB 列（计算列 已付/剩余/逾期 下期再做注解筛选）。
# status 为计算口径，单独走既有 annotation 逻辑，不入通用解析器。
PAYMENT_FILTER_REGISTRY = {
    'department':         {'type': 'enum',   'col': 'department'},
    'secondary_dept':     {'type': 'text',   'col': 'secondary_dept'},
    'project_short_name': {'type': 'text',   'col': 'project_short_name'},
    'applicant':          {'type': 'text',   'col': 'applicant'},
    'approval_number':    {'type': 'text',   'col': 'approval_number'},
    'g7_number':          {'type': 'text',   'col': 'g7_number'},
    'project_desc':       {'type': 'text',   'col': 'project_desc'},
    'payee':              {'type': 'text',   'col': 'payee'},
    'notes':              {'type': 'text',   'col': 'notes'},
    'planned_date':       {'type': 'date',   'col': 'planned_date'},
    'total_amount':       {'type': 'number', 'col': 'total_amount'},
    'plan_adjustment':    {'type': 'number', 'col': 'plan_adjustment'},
    # 回款日期：反向关联到分期表，需去重；排序无意义故禁用
    'pay_date':           {'type': 'date',   'col': 'installments__pay_date',
                           'multi': True, 'sortable': False},
}

# 付款管理「计算列」筛选：已付/剩余为 SQL 注解后的数值列；逾期为派生是/否（单独处理）。
# 这些列不在 PAYMENT_FILTER_REGISTRY（真实列）中，需先注解再用本表走通用数值解析器。
# 注解名避免与模型 @property（total_paid / remaining）冲突：remaining→remaining_calc。
PAYMENT_COMPUTED_REGISTRY = {
    'paid':      {'type': 'number', 'col': 'paid'},
    'remaining': {'type': 'number', 'col': 'remaining_calc'},
}

# 审计日志：仅登记 AuditLog 上真实存在的 DB 列（result 为计算口径，
# 由既有 status_code 阈值逻辑处理，不入通用解析器）。
AUDITLOG_FILTER_REGISTRY = {
    'user_name':   {'type': 'text',   'col': 'user_name'},
    'module':      {'type': 'enum',   'col': 'module'},
    'method':      {'type': 'enum',   'col': 'method'},
    'path':        {'type': 'text',   'col': 'path'},
    'status_code': {'type': 'number', 'col': 'status_code'},
    'created_at':  {'type': 'date',   'col': 'created_at__date'},
}

# ── permission model ────────────────────────────────────────────────────────────
# Payment fields that can be individually permission-controlled. Each logical
# field maps to one or more columns in the serialized payment dict.
PAYMENT_FIELD_DEFS = [
    {'key': 'department',      'label': '部门',         'cols': ['department']},
    {'key': 'secondary_dept',  'label': '二级部门',      'cols': ['secondary_dept']},
    {'key': 'project_short_name', 'label': '项目简称',   'cols': ['project_short_name']},
    {'key': 'applicant',       'label': '申请人',        'cols': ['applicant']},
    {'key': 'approval_number', 'label': '审批单号',      'cols': ['approval_number']},
    {'key': 'g7_number',       'label': 'G7编号',        'cols': ['g7_number']},
    {'key': 'project_desc',    'label': '付款事项',      'cols': ['project_desc']},
    {'key': 'payee',           'label': '收款方',        'cols': ['payee']},
    {'key': 'total_amount',    'label': '计划总金额',    'cols': ['total_amount']},
    {'key': 'planned_date',    'label': '计划付款日期',  'cols': ['planned_date']},
    {'key': 'installments',    'label': '付款明细',     'cols': ['installments', 'total_paid', 'remaining', 'status']},
    {'key': 'notes',            'label': '备注',          'cols': ['notes']},
    {'key': 'plan_adjustment', 'label': '计划调整金额',   'cols': ['plan_adjustment']},
]
FIELD_KEYS = [f['key'] for f in PAYMENT_FIELD_DEFS]
PAGE_KEYS = [
    'dashboard', 'payments', 'approval_records', 'stats',
    'ar_projects', 'ar_records', 'ar_advance', 'ar_analytics', 'ar_cashflow', 'ar_budget',
    'caiwu_report', 'caiwu_data', 'caiwu_charts', 'caiwu_metrics', 'caiwu_cockpit',
]

# Caiwu report/chart element permission keys (mirrored from caiwu.views.PERM_FIELD_DEFS)
CAIWU_FIELD_DEFS = [
    {'key': 'report_l1',       'label': '一级科目报表'},
    {'key': 'report_l2',       'label': '二级项目部明细'},
    {'key': 'report_l3',       'label': '三级科目明细'},
    {'key': 'amount',          'label': '金额数据'},
    {'key': 'chart_trend',     'label': '走势折线图'},
    {'key': 'chart_waterfall', 'label': '因素瀑布图'},
    {'key': 'export',          'label': '导出 Excel'},
]
CAIWU_FIELD_KEYS = [f['key'] for f in CAIWU_FIELD_DEFS]

# 仅允许这些职务对审批单直接置为 approved/rejected；其余职务（操作员/出纳/结算会计等）
# 只能创建 pending、或取消自己的申请。super_admin 永远豁免此限制。
APPROVER_JOBS = {'finance_director', 'finance_bp', 'general_manager', 'gm_assistant'}


def is_approver(request):
    if getattr(request, 'pk_role', None) == 'super_admin':
        return True
    return getattr(request, 'pk_job', None) in APPROVER_JOBS

# ── AR field-level permission defs ───────────────────────────────────────────
# Per-field show/hide control for the AR module. Keys are namespaced (p_ for
# project台账 fields, r_ for 应收明细 fields) so project/record fields never
# collide. `cols` lists the serialized dict keys masked when the field is hidden.
# Defined here (not in ar/views.py) so the permission system has no dependency
# on the ar app, avoiding a circular import (ar imports paikuan, not vice versa).
AR_PROJECT_FIELD_DEFS = [
    {'key': 'p_contract_name',  'label': '客户名称',   'group': 'project', 'cols': ['customer_name']},
    {'key': 'p_short_name',     'label': '项目简称',   'group': 'project', 'cols': ['short_name']},
    {'key': 'p_delivery_dept',  'label': '交付部门',   'group': 'project', 'cols': ['delivery_dept']},
    {'key': 'p_sub_dept',       'label': '二级部门',   'group': 'project', 'cols': ['sub_dept']},
    {'key': 'p_business_mode',  'label': '业务模式',   'group': 'project', 'cols': ['business_mode']},
    {'key': 'p_customer_level', 'label': '客户等级',   'group': 'project', 'cols': ['customer_level']},
    {'key': 'p_sales_contact',  'label': '销售对接人', 'group': 'project', 'cols': ['sales_contact']},
    {'key': 'p_project_manager','label': '项目负责人', 'group': 'project', 'cols': ['project_manager']},
    {'key': 'p_has_contract',   'label': '有无合同',   'group': 'project', 'cols': ['has_contract']},
    {'key': 'p_contract_date',  'label': '签订日期',   'group': 'project', 'cols': ['contract_date']},
    {'key': 'p_account_period', 'label': '账期配置',   'group': 'project',
        'cols': ['reconciliation_days', 'invoice_wait_days', 'post_invoice_days', 'total_days']},
    {'key': 'p_invoice_config', 'label': '开票配置(模式/类型/税率)', 'group': 'project',
        'cols': ['invoice_mode', 'invoice_type', 'tax_rate']},
    {'key': 'p_notes',          'label': '项目备注',   'group': 'project', 'cols': ['notes']},
]
AR_RECORD_FIELD_DEFS = [
    {'key': 'r_estimated_amount',      'label': '预估上账金额', 'group': 'record', 'cols': ['estimated_amount']},
    {'key': 'r_actual_invoice_amount', 'label': '实际开票金额', 'group': 'record', 'cols': ['actual_invoice_amount']},
    {'key': 'r_tax_amount',            'label': '税额',         'group': 'record', 'cols': ['tax_amount']},
    {'key': 'r_invoice_date',          'label': '开票日期',     'group': 'record', 'cols': ['invoice_date']},
    {'key': 'r_account_diff',          'label': '账实差额',     'group': 'record', 'cols': ['account_diff_adjustment']},
    {'key': 'r_outstanding',           'label': '未回款金额',   'group': 'record', 'cols': ['outstanding_amount']},
    {'key': 'r_due_date',              'label': '应收日期/逾期', 'group': 'record',
        'cols': ['due_date', 'is_overdue', 'overdue_days', 'target_collection_date']},
    {'key': 'r_reconciliation',        'label': '对账信息',     'group': 'record',
        'cols': ['reconciliation_date', 'reconciliation_status']},
    {'key': 'r_invoice_status',        'label': '开票状态',     'group': 'record', 'cols': ['invoice_status']},
    {'key': 'r_payments',              'label': '回款记录',     'group': 'record', 'cols': ['payments']},
    {'key': 'r_notes',                 'label': '明细备注',     'group': 'record', 'cols': ['notes']},
]
# 预收预付字段级权限（adv_ 前缀）。cols 为隐藏时遮蔽的序列化 key。
AR_ADVANCE_FIELD_DEFS = [
    {'key': 'adv_counterparty',  'label': '往来单位',     'group': 'advance', 'cols': ['counterparty']},
    {'key': 'adv_amount',        'label': '预收/预付金额', 'group': 'advance', 'cols': ['advance_amount']},
    {'key': 'adv_writeoff',      'label': '核销信息',     'group': 'advance',
        'cols': ['written_off_amount', 'balance_amount', 'writeoff_status', 'writeoffs']},
    {'key': 'adv_expected_date', 'label': '预计核销/账龄', 'group': 'advance',
        'cols': ['expected_writeoff_date', 'pending_days', 'is_overdue', 'overdue_days']},
    {'key': 'adv_notes',         'label': '预收预付备注',  'group': 'advance', 'cols': ['notes']},
]
AR_FIELD_DEFS = AR_PROJECT_FIELD_DEFS + AR_RECORD_FIELD_DEFS + AR_ADVANCE_FIELD_DEFS
AR_FIELD_KEYS = [f['key'] for f in AR_FIELD_DEFS]

# ── 操作级权限（独立于"能不能新增记录"的细粒度动作开关）────────────────────────
# 解决"出纳要做预付核销，却只能靠放开整个 can_create"的粒度问题：
# 每个动作单独可配，与页面/字段/记录权限正交组合。
ACTION_DEFS = [
    {'key': 'wo_prepaid',      'label': '预付核销（排款冲抵/撤销）'},
    {'key': 'wo_receive',      'label': '预收核销（冲抵应收/批量/撤销）'},
    {'key': 'ar_collect',      'label': '应收回款录入（单笔/批次分摊）'},
    {'key': 'adv_installment', 'label': '预收预付收付登记（多次到账/付出）'},
]
ACTION_KEYS = [a['key'] for a in ACTION_DEFS]


def _all_actions(value):
    return {k: value for k in ACTION_KEYS}

# Ordered columns for Excel template / import / export.
# (perm_field_key, excel_header, db_column)
EXCEL_COLUMN_MAP = [
    ('department',      '部门',               'department'),
    ('secondary_dept',  '二级部门',            'secondary_dept'),
    ('project_short_name', '项目简称',         'project_short_name'),
    ('applicant',       '申请人',             'applicant'),
    ('approval_number', '审批单号',            'approval_number'),
    ('g7_number',       'G7编号',              'g7_number'),
    # 项目维度统一走「项目简称」列（与台账精确校验）。项目编号不再出现在
    # 模板/导入/导出；创建时由简称从台账自动带出（旧文件多出的编号列被忽略）
    ('project_desc',    '付款事项',            'project_desc'),
    ('payee',           '收款方',              'payee'),
    ('total_amount',    '计划总金额(元)',       'total_amount'),
    ('planned_date',    '计划付款日期',         'planned_date'),
    # installment columns are generated dynamically (see _visible_excel_cols)
    ('notes',           '备注',                'notes'),
    ('plan_adjustment', '计划调整金额(元)',     'plan_adjustment'),
]
_EXCEL_HEADER_TO_COL = {h: c for _, h, c in EXCEL_COLUMN_MAP}
_EXCEL_DATE_COLS = {'planned_date'}
# Default number of installment slots in the import template.
TEMPLATE_INSTALLMENT_SLOTS = 6

# Example row in the template carries this marker in 部门 so import skips it
# even when the user forgets to delete it.
EXAMPLE_ROW_MARKER = '示例-导入前请删除此行'


def _normalize_date(s):
    """Best-effort normalize a user-typed date string to ISO YYYY-MM-DD."""
    s = (str(s) or '').strip()
    if not s:
        return None
    s = s.replace('/', '-').replace('.', '-').replace('年', '-').replace('月', '-').replace('日', '')
    parts = [p for p in s.split('-') if p != '']
    try:
        if len(parts) >= 3:
            return f'{int(parts[0]):04d}-{int(parts[1]):02d}-{int(parts[2]):02d}'
    except ValueError:
        pass
    return s  # leave as-is; downstream save will reject if truly invalid


def _all_fields(value):
    return {k: value for k in FIELD_KEYS}


def _all_ar_fields(value):
    return {k: value for k in AR_FIELD_KEYS}


def _all_caiwu_fields(value):
    return {k: value for k in CAIWU_FIELD_KEYS}


def default_job_config(job):
    """Sensible starting permissions for each job title; super_admin can override."""
    # Non-caiwu pages True for most roles; caiwu pages default to off
    _non_cw_pages = {k: True for k in PAGE_KEYS if not k.startswith('caiwu_')}
    pages_all = {**_non_cw_pages, 'caiwu_report': False, 'caiwu_data': False, 'caiwu_charts': False,
                 'caiwu_metrics': False, 'caiwu_cockpit': False}
    ar_pages_all = {k: True for k in ('ar_projects', 'ar_records', 'ar_advance', 'ar_analytics', 'ar_cashflow', 'ar_budget')}
    ar_pages_cashier = {k: (k in ('ar_records', 'ar_advance', 'ar_cashflow', 'ar_budget')) for k in ar_pages_all}
    # Reusable caiwu capability blocks
    _cw_full = {
        'caiwu_view': _all_caiwu_fields(True),
        'caiwu_upload': True, 'caiwu_publish': True, 'caiwu_delete': True,
    }
    _cw_upload_no_del = {
        'caiwu_view': _all_caiwu_fields(True),
        'caiwu_upload': True, 'caiwu_publish': True, 'caiwu_delete': False,
    }
    _cw_readonly = {
        'caiwu_view': _all_caiwu_fields(True),
        'caiwu_upload': False, 'caiwu_publish': False, 'caiwu_delete': False,
    }
    if job == 'finance_director':
        pages = {**pages_all, 'caiwu_report': True, 'caiwu_data': True, 'caiwu_charts': True,
                 'caiwu_metrics': True, 'caiwu_cockpit': True}
        return {'pages': pages, 'view': _all_fields(True),
                'edit': _all_fields(True), 'ar_view': _all_ar_fields(True),
                'can_create': True, 'can_delete': True, 'ar_shared_only': False,
                'actions': _all_actions(True),
                **_cw_full}
    if job == 'finance_bp':
        pages = {**pages_all, 'caiwu_report': True, 'caiwu_data': True, 'caiwu_charts': True,
                 'caiwu_metrics': True, 'caiwu_cockpit': True}
        return {'pages': pages, 'view': _all_fields(True),
                'edit': _all_fields(True), 'ar_view': _all_ar_fields(True),
                'can_create': True, 'can_delete': False, 'ar_shared_only': False,
                'actions': _all_actions(True),
                **_cw_upload_no_del}
    if job == 'chief_cashier':
        edit = {k: (k in ('installments',)) for k in FIELD_KEYS}
        # 总出纳默认看不到税额
        ar_view = {**_all_ar_fields(True), 'r_tax_amount': False}
        return {'pages': {**pages_all, **ar_pages_all}, 'view': _all_fields(True),
                'edit': edit, 'ar_view': ar_view,
                'can_create': True, 'can_delete': False, 'ar_shared_only': False,
                'actions': _all_actions(True),
                **_cw_readonly}
    if job == 'cashier':
        edit = {k: (k in ('installments',)) for k in FIELD_KEYS}
        base_pages = {'dashboard': True, 'payments': True, 'approval_records': True, 'stats': False}
        # 出纳默认看不到税额与账实差额
        ar_view = {**_all_ar_fields(True), 'r_tax_amount': False, 'r_account_diff': False}
        return {'pages': {**base_pages, **ar_pages_cashier},
                'view': _all_fields(True), 'edit': edit, 'ar_view': ar_view,
                'can_create': False, 'can_delete': False, 'ar_shared_only': False,
                'actions': {'wo_prepaid': True, 'wo_receive': False,
                            'ar_collect': True, 'adv_installment': True},
                **_cw_readonly}
    if job == 'general_manager':
        # 总经理：全量查看，无编辑/创建；财务分析只读
        pages = {**pages_all, 'caiwu_report': True, 'caiwu_data': False, 'caiwu_charts': True,
                 'caiwu_metrics': True, 'caiwu_cockpit': True}
        return {'pages': pages, 'view': _all_fields(True),
                'edit': _all_fields(False), 'ar_view': _all_ar_fields(True),
                'can_create': False, 'can_delete': False, 'ar_shared_only': False,
                'actions': _all_actions(False),
                **_cw_readonly}
    if job == 'gm_assistant':
        # 总经理助理：全量查看 + 可新增（协助登记），不可删除
        return {'pages': pages_all, 'view': _all_fields(True),
                'edit': _all_fields(True), 'ar_view': _all_ar_fields(True),
                'can_create': True, 'can_delete': False, 'ar_shared_only': False,
                'actions': _all_actions(True),
                **_cw_readonly}
    if job == 'settlement_accountant':
        # 结算会计：聚焦应收/对账/开票，可编辑应收主数据；付款条数据只读
        edit = {k: False for k in FIELD_KEYS}
        return {'pages': pages_all, 'view': _all_fields(True),
                'edit': edit, 'ar_view': _all_ar_fields(True),
                'can_create': False, 'can_delete': False, 'ar_can_create': True,
                'ar_shared_only': False,
                'actions': {'wo_prepaid': False, 'wo_receive': True,
                            'ar_collect': True, 'adv_installment': True},
                **_cw_readonly}
    if job == 'sales_bp':
        # 销售BP：仅可见共享业务，AR 只读，无付款操作
        ar_pages = {k: (k in ('ar_projects', 'ar_records')) for k in _non_cw_pages}
        pages = {**{k: False for k in _non_cw_pages}, **ar_pages,
                 'dashboard': True, 'caiwu_report': False, 'caiwu_data': False, 'caiwu_charts': False}
        return {'pages': pages, 'view': _all_fields(False),
                'edit': _all_fields(False), 'ar_view': _all_ar_fields(True),
                'can_create': False, 'can_delete': False, 'ar_shared_only': True,
                'actions': _all_actions(False),
                **_cw_readonly}
    # Unknown / no job title → read-only minimum, no caiwu access.
    return {'pages': pages_all, 'view': _all_fields(True),
            'edit': _all_fields(False), 'ar_view': _all_ar_fields(True),
            'can_create': False, 'can_delete': False, 'ar_shared_only': False,
            'actions': _all_actions(False),
            **_cw_readonly}


_perm_cache: dict = {}
_perm_cache_lock = threading.Lock()
# 进程内缓存带短 TTL：权限修改在多 worker 部署下，未命中保存请求的 worker 也能在
# 数秒内自愈（_invalidate_perm_cache 只清当前进程，跨进程靠 TTL 兜底），避免"刚授权
# 仍提示无权限"。读多写少，TTL 期内仍走缓存不增加 DB 压力。
_PERM_CACHE_TTL = 5.0  # seconds


def _invalidate_perm_cache(job=None):
    with _perm_cache_lock:
        if job:
            _perm_cache.pop(job, None)
        else:
            _perm_cache.clear()


def get_job_perms(job):
    """Effective config for a job title = defaults merged with stored overrides.
    Results are cached per-process to avoid repeated DB lookups."""
    now = time.monotonic()
    with _perm_cache_lock:
        ent = _perm_cache.get(job)
        if ent and ent[0] > now:
            return ent[1]
    base = default_job_config(job)
    try:
        rp = JobPermission.objects.filter(job_title=job).first()
    except Exception:
        # paikuan_job_permissions table may not exist yet (migration pending).
        logger.warning('get_job_perms: DB unavailable for %s, using defaults', job)
        return base
    if not (rp and rp.config):
        result = dict(base)
        # 能力位规范化：保证返回值始终带 ar_can_create（多数职务基线无此键 → False），
        # 让权限配置页统一可读可改，避免 undefined 造成误判。
        result.setdefault('ar_can_create', base.get('ar_can_create', False))
    else:
        cfg = rp.config
        view = dict(base['view']);  view.update(cfg.get('view', {}))
        edit = dict(base['edit']);  edit.update(cfg.get('edit', {}))
        pages = dict(base['pages']); pages.update(cfg.get('pages', {}))
        ar_view = dict(base.get('ar_view', _all_ar_fields(True)))
        ar_view.update(cfg.get('ar_view', {}))
        actions = dict(base.get('actions', _all_actions(False)))
        actions.update(cfg.get('actions', {}))
        caiwu_view = dict(base.get('caiwu_view', _all_caiwu_fields(True)))
        caiwu_view.update(cfg.get('caiwu_view', {}))
        result = {
            'pages': pages, 'view': view, 'edit': edit, 'ar_view': ar_view,
            'actions': actions,
            'can_create': bool(cfg.get('can_create', base['can_create'])),
            'can_delete': bool(cfg.get('can_delete', base['can_delete'])),
            # 应收专属写权限（结算会计等）：必须在合并时保留，否则一旦超管在权限
            # 配置页对该职务做任何自定义覆盖，ar_can_create 会丢失 → 写入再次被拒。
            'ar_can_create': bool(cfg.get('ar_can_create', base.get('ar_can_create', False))),
            'ar_shared_only': bool(cfg.get('ar_shared_only', base.get('ar_shared_only', False))),
            'caiwu_view': caiwu_view,
            'caiwu_upload':  bool(cfg.get('caiwu_upload',  base.get('caiwu_upload', False))),
            'caiwu_publish': bool(cfg.get('caiwu_publish', base.get('caiwu_publish', False))),
            'caiwu_delete':  bool(cfg.get('caiwu_delete',  base.get('caiwu_delete', False))),
        }
    with _perm_cache_lock:
        _perm_cache[job] = (time.monotonic() + _PERM_CACHE_TTL, result)
    return result


def full_perms():
    return {'pages': {k: True for k in PAGE_KEYS}, 'view': _all_fields(True),
            'edit': _all_fields(True), 'ar_view': _all_ar_fields(True),
            'can_create': True, 'can_delete': True, 'ar_can_create': True,
            'ar_shared_only': False,
            'actions': _all_actions(True),
            'caiwu_view': _all_caiwu_fields(True),
            'caiwu_upload': True, 'caiwu_publish': True, 'caiwu_delete': True}


def effective_perms(user):
    """Perms object sent to the client (super_admin gets full access)."""
    cfg = full_perms() if user.role == 'super_admin' else get_job_perms(user.job_title)
    return {**cfg, 'is_admin': user.role == 'super_admin',
            'fields': PAYMENT_FIELD_DEFS, 'ar_fields': AR_FIELD_DEFS,
            'action_defs': ACTION_DEFS}


_AR_DEFS_BY_GROUP = {
    'project': AR_PROJECT_FIELD_DEFS,
    'record': AR_RECORD_FIELD_DEFS,
    'advance': AR_ADVANCE_FIELD_DEFS,
}


def apply_ar_view_mask(d, perms, group):
    """Null out AR dict fields the role cannot view. `group` is 'project',
    'record' or 'advance'. perms None (super_admin) → unchanged."""
    if perms is None:
        return d
    ar_view = perms.get('ar_view') or {}
    defs = _AR_DEFS_BY_GROUP.get(group, AR_RECORD_FIELD_DEFS)
    for f in defs:
        if not ar_view.get(f['key'], True):
            for c in f['cols']:
                if c in d:
                    d[c] = None
    return d


def get_request_perms(request):
    """Effective perms for the authed request. None means full access (super_admin)."""
    if request.pk_role == 'super_admin':
        return None
    jt = getattr(request, 'pk_job', '') or ''
    if not jt:
        u = PaikuanUser.objects.filter(id=request.pk_uid).only('job_title').first()
        jt = u.job_title if u else ''
    return get_job_perms(jt)


def apply_view_mask(d, perms):
    """Null out payment-dict fields the role cannot view."""
    if perms is None:
        return d
    view = perms['view']
    hide_cols = set()
    for f in PAYMENT_FIELD_DEFS:
        if not view.get(f['key'], True):
            hide_cols.update(f['cols'])
    for c in hide_cols:
        if c in d:
            d[c] = None
    inst_hidden = not view.get('installments', True)
    if inst_hidden:
        d['total_paid'] = None
    if inst_hidden or not view.get('total_amount', True):
        d['remaining'] = None
    return d


# ── helpers ───────────────────────────────────────────────────────────────────

def _no_store(resp):
    # API data is per-user and mutates frequently (e.g. user deletion); never let
    # a browser or proxy serve a stale copy that would resurrect deleted records.
    resp['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    resp['Pragma'] = 'no-cache'
    return resp


def ok(data=None):
    return _no_store(JsonResponse({'code': 0, 'data': data}))


def err(msg, status=400, code=-1):
    return _no_store(JsonResponse({'code': code, 'error': msg}, status=status))


def parse_body(request):
    try:
        return json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return {}


def make_token(user):
    payload = {
        'uid': user.id,
        'role': user.role,
        'job': user.job_title,
        'departments': user.departments,
        'exp': datetime.datetime.utcnow() + datetime.timedelta(hours=24),
        'iat': datetime.datetime.utcnow(),
    }
    return jwt.encode(payload, settings.JWT_SECRET, algorithm='HS256')


def pk_required(roles=None):
    def decorator(func):
        @functools.wraps(func)
        def wrapper(request, *args, **kwargs):
            token = request.headers.get('Authorization', '').replace('Bearer ', '').strip()
            if not token:
                return err('未认证', 401, 401)
            try:
                payload = jwt.decode(token, settings.JWT_SECRET, algorithms=['HS256'])
                uid = payload['uid']
            except jwt.ExpiredSignatureError:
                return err('Token已过期', 401, 401)
            except (jwt.InvalidTokenError, KeyError):
                return err('Token无效', 401, 401)
            user = PaikuanUser.objects.filter(id=uid).first()
            if not user:
                return err('用户不存在或登录已失效', 401, 401)
            if not user.is_active:
                return err('账号已停用，请重新登录', 401, 401)
            if not user.is_approved:
                return err('账号待审批，请联系管理员', 403, -2)
            # 改密码即踢旧会话：token 签发时间早于最近一次改密 → 失效
            if user.pwd_changed_at:
                iat = payload.get('iat')
                if iat and int(iat) < int(user.pwd_changed_at.timestamp()):
                    return err('密码已修改，请重新登录', 401, 401)
            request.pk_user = user
            request.pk_uid = user.id
            request.pk_role = user.role
            request.pk_job = user.job_title
            request.pk_depts = user.departments or []
            if roles and request.pk_role not in roles:
                return err('权限不足', 403, 403)
            return func(request, *args, **kwargs)
        return wrapper
    return decorator


def _payments_page_denied(request, perms=None):
    perms = get_request_perms(request) if perms is None else perms
    if perms is not None and not perms['pages'].get('payments', True):
        return err('无访问权限', 403, 403)
    return None


def dept_filter(qs, request):
    """Row visibility: super_admin sees all; everyone else sees assigned departments.
    Optional ?depts=A,B,C further narrows the set (intersected with pk_depts)."""
    raw = request.GET.get('depts', '').strip()
    requested = [d for d in raw.split(',') if d.strip()] if raw else []

    if request.pk_role == 'super_admin':
        if requested:
            return qs.filter(department__in=requested)
        return qs

    allowed = set(request.pk_depts or [])
    if requested:
        active = [d for d in requested if d in allowed]
        if active:
            return qs.filter(department__in=active)
    return qs.filter(department__in=request.pk_depts)


def can_write_dept(request, dept):
    if request.pk_role == 'super_admin':
        return True
    return dept in request.pk_depts


_DEC = DjDecimalField(max_digits=15, decimal_places=2)


def _paid_subq():
    """Correlated subquery: sum of installment amounts for each Payment row."""
    return Coalesce(
        Subquery(
            PaymentInstallment.objects
                .filter(payment_id=OuterRef('pk'))
                .values('payment_id')
                .annotate(s=Sum('pay_amount'))
                .values('s'),
            output_field=_DEC,
        ),
        Value(Decimal('0')),
        output_field=_DEC,
    )


def _paid_expr():
    """Annotation expression for total paid amount (from installments subtable)."""
    return _paid_subq()


def _plan_count_subq():
    """Correlated subquery: number of plan-item batches per Payment row（供轻量列表取批次数，
    免预取 plan_items 子表）。"""
    return Coalesce(
        Subquery(
            PaymentPlanItem.objects
                .filter(payment_id=OuterRef('pk'))
                .values('payment_id')
                .annotate(c=Count('id'))
                .values('c'),
            output_field=IntegerField(),
        ),
        Value(0),
        output_field=IntegerField(),
    )


def _paid_subq_windowed(pay_start, pay_end):
    """Correlated subquery: sum of installment amounts within an optional
    pay_date window. 台账「已付」按付款日期窗口筛选时与「付款流水」同口径——
    仅统计落在窗口内的实付，避免把窗口外的历史分期也计入（否则两表对不上）。"""
    sub = PaymentInstallment.objects.filter(payment_id=OuterRef('pk'))
    if pay_start:
        sub = sub.filter(pay_date__gte=pay_start)
    if pay_end:
        sub = sub.filter(pay_date__lte=pay_end)
    return Coalesce(
        Subquery(
            sub.values('payment_id').annotate(s=Sum('pay_amount')).values('s'),
            output_field=_DEC,
        ),
        Value(Decimal('0')),
        output_field=_DEC,
    )


def _safe_iso_date(s):
    """Parse 'YYYY-MM-DD' → date, or None on any failure (defensive)."""
    try:
        return datetime.date.fromisoformat(s)
    except Exception:
        return None


def _remaining_expr():
    """SQL expression for remaining = total_amount - paid (requires 'paid' annotation)."""
    return ExpressionWrapper(F('total_amount') - F('paid'), output_field=_DEC)


# 求和字段位宽放宽，避免跨多月结转累加时溢出 _DEC 的 15 位上限。
_DEC18 = DjDecimalField(max_digits=18, decimal_places=2)


def _not_settled_q(paid_field='_paid'):
    """未结清：已付 + 预付冲抵 仍小于有效计划（调整额优先）。
    预付冲抵的现金在预付时已流出，剩余应付须扣除——与资金池「刚性待付」口径一致。"""
    return (
        Q(plan_adjustment__isnull=True,
          **{f'{paid_field}__lt': F('total_amount') - F('prepaid_offset_amount')}) |
        Q(plan_adjustment__isnull=False,
          **{f'{paid_field}__lt': F('plan_adjustment') - F('prepaid_offset_amount')})
    )


def _effective_remaining_expr(paid_field='_paid'):
    """剩余应付：有效计划（调整额优先）− 已付 − 预付冲抵（与资金池口径一致）。"""
    return Case(
        When(plan_adjustment__isnull=False,
             then=ExpressionWrapper(
                 F('plan_adjustment') - F(paid_field) - F('prepaid_offset_amount'),
                 output_field=_DEC18)),
        default=ExpressionWrapper(
            F('total_amount') - F(paid_field) - F('prepaid_offset_amount'),
            output_field=_DEC18),
        output_field=_DEC18,
    )


# ── version / deploy check ──────────────────────────────────────────────────────
# Bump BUILD_VERSION whenever backend behaviour changes so a deploy can be verified
# by opening /api/pk/version in a browser (no auth required).
BUILD_VERSION = '2026-06-11.1'


@csrf_exempt
def version(request):
    return ok({
        'version': BUILD_VERSION,
        'features': {
            'hard_delete_user': True,      # DELETE /users/<id> permanently removes the row
            'stats_dept_filter': True,     # GET /stats?depts=A,B filters by department
            'no_store_headers': True,      # API responses are never cached
        },
    })


# ── auth ──────────────────────────────────────────────────────────────────────

# 常见弱密码黑名单（小写比对）：纯弱口令直接拒绝，不指望用户自觉
_WEAK_PASSWORDS = {
    '12345678', '123456789', '1234567890', '88888888', '66668888', '11111111',
    '00000000', 'password', 'password1', 'passw0rd', 'abc12345', 'abcd1234',
    'qwerty123', 'qwertyui', 'admin123', 'iloveyou', 'a1234567', '1qaz2wsx',
    'zaq12wsx', '123qwe123', 'aa123456', 'woaini123', '123456abc', 'abc123456',
}


def _validate_password_strength(pwd, phone='', name=''):
    """密码强度校验：返回错误文案（不通过）或 None（通过）。
    用于注册 / 自助改密 / 超管重置三处，口径统一。"""
    pwd = pwd or ''
    if len(pwd) < 8:
        return '密码至少8位'
    if len(pwd) > 128:
        return '密码过长（最多128位）'
    has_letter = any(c.isalpha() for c in pwd)
    has_digit = any(c.isdigit() for c in pwd)
    if not (has_letter and has_digit):
        return '密码须同时包含字母和数字'
    if len(set(pwd)) <= 2:
        return '密码过于简单（字符种类太少）'
    low = pwd.lower()
    if low in _WEAK_PASSWORDS:
        return '密码过于常见，请更换'
    if phone and phone in pwd:
        return '密码不能包含手机号'
    if name and name.lower() and name.lower() in low:
        return '密码不能包含姓名'
    return None


@csrf_exempt
def register(request):
    if request.method != 'POST':
        return err('Method not allowed', 405)
    data = parse_body(request)
    phone = (data.get('phone') or '').strip()
    password = (data.get('password') or '').strip()
    name = (data.get('name') or '').strip()
    job_title = (data.get('job_title') or '').strip()
    depts = data.get('departments') or []

    if not phone or not password or not name:
        return err('手机号、密码和姓名均必填')
    if not phone.isdigit() or len(phone) < 8:
        return err('手机号格式有误')
    pwd_err = _validate_password_strength(password, phone=phone, name=name)
    if pwd_err:
        return err(pwd_err)
    if not job_title or job_title not in JOB_TITLES:
        return err('请选择有效职务')
    if not isinstance(depts, list) or len(depts) == 0:
        return err('请至少选择一个部门')

    # Phone uniqueness
    existing = PaikuanUser.objects.filter(phone=phone).first()
    if existing:
        if existing.name != name:
            return err(f'该手机号已被"{existing.name}"注册，姓名不符')
        return err('该手机号已注册')
    # Name uniqueness — each person should have exactly one account
    if PaikuanUser.objects.filter(name=name).exists():
        return err('该姓名已被注册，如有疑问请联系管理员')

    is_first = not PaikuanUser.objects.exists()
    role = 'super_admin' if is_first else 'viewer'

    user = PaikuanUser(
        phone=phone, name=name, role=role,
        job_title=job_title,
        departments=depts if role != 'super_admin' else [],
        is_approved=is_first,
    )
    user.set_password(password)
    user.save()

    if is_first:
        return ok({'token': make_token(user), 'user': user.to_dict(),
                   'permissions': effective_perms(user), 'pending': False})
    return ok({'pending': True, 'message': '注册成功！请等待管理员审批后登录'})


@csrf_exempt
def login(request):
    if request.method != 'POST':
        return err('Method not allowed', 405)
    data = parse_body(request)
    phone = (data.get('phone') or '').strip()
    password = (data.get('password') or '').strip()
    if not phone or not password:
        return err('手机号和密码均必填')
    # 防暴力破解：同账号15分钟内失败≥5次 → 临时锁定（基于审计日志，多进程安全）
    try:
        from paikuan.models import AuditLog
        _cutoff = timezone.now() - datetime.timedelta(minutes=15)
        _failed = AuditLog.objects.filter(
            path__endswith='/login', status_code__gte=400,
            created_at__gte=_cutoff, user__phone=phone).count()
        if _failed >= 5:
            return err('失败次数过多，账号已临时锁定，请15分钟后再试', 429, 429)
    except Exception:
        pass
    try:
        user = PaikuanUser.objects.get(phone=phone)
    except PaikuanUser.DoesNotExist:
        return err('手机号或密码错误', 401, 401)
    if not user.is_active:
        return err('账号已停用，请联系管理员', 401, 401)
    if not user.check_password(password):
        return err('手机号或密码错误', 401, 401)
    if not user.is_approved:
        return err('账号待审批，请联系管理员', 403, -2)
    return ok({'token': make_token(user), 'user': user.to_dict(),
               'permissions': effective_perms(user),
               'must_change_password': user.must_change_password})


@csrf_exempt
def registration_status(request):
    """Public polling endpoint so a pending registrant can detect approval."""
    if request.method != 'GET':
        return err('Method not allowed', 405)
    phone = (request.GET.get('phone') or '').strip()
    if not phone:
        return err('缺少手机号')
    user = PaikuanUser.objects.filter(phone=phone).first()
    if not user:
        return ok({'status': 'none'})
    if not user.is_active:
        return ok({'status': 'rejected'})
    if user.is_approved:
        return ok({'status': 'approved'})
    return ok({'status': 'pending'})


@pk_required()
def me(request):
    if request.method != 'GET':
        return err('Method not allowed', 405)
    try:
        user = PaikuanUser.objects.get(id=request.pk_uid)
    except PaikuanUser.DoesNotExist:
        return err('用户不存在', 404)
    return ok({'user': user.to_dict(), 'permissions': effective_perms(user)})


@csrf_exempt
@pk_required()
def change_password(request):
    """自助修改密码：验旧密码 + 新密码强度校验。改密即刷新 pwd_changed_at（踢掉其它
    会话）并清除强制改密标志；返回新 token 让当前会话无缝续上。"""
    if request.method != 'POST':
        return err('Method not allowed', 405)
    try:
        user = PaikuanUser.objects.get(id=request.pk_uid)
    except PaikuanUser.DoesNotExist:
        return err('用户不存在', 404)
    data = parse_body(request)
    old_pwd = (data.get('old_password') or '').strip()
    new_pwd = (data.get('new_password') or '').strip()
    if not user.check_password(old_pwd):
        return err('原密码不正确', 400)
    if new_pwd == old_pwd:
        return err('新密码不能与原密码相同')
    pwd_err = _validate_password_strength(new_pwd, phone=user.phone, name=user.name)
    if pwd_err:
        return err(pwd_err)
    user.set_password(new_pwd)
    user.pwd_changed_at = timezone.now()
    user.must_change_password = False
    user.save(update_fields=['password_hash', 'pwd_changed_at', 'must_change_password', 'updated_at'])
    return ok({'token': make_token(user), 'user': user.to_dict()})


# ── payments ──────────────────────────────────────────────────────────────────

@csrf_exempt
@pk_required()
def payments(request):
    perms = get_request_perms(request)
    denied = _payments_page_denied(request, perms)
    if denied:
        return denied
    if request.method == 'GET':
        return _list_payments(request)
    if request.method == 'POST':
        return _create_payment(request)
    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def prepaid_balance(request):
    """排款页联动：查可冲抵本排款的「预付」未核销余额。

    匹配口径（并集）：①挂在该项目（按项目编号/项目简称定位台账）下的预付；
    ②散单预付（未挂项目）且往来单位 = 排款收款方（payee，忽略大小写精确匹配）。
    只读；按用户可见部门作用域过滤。
    """
    perms = get_request_perms(request)
    denied = _payments_page_denied(request, perms)
    if denied:
        return denied
    if request.method != 'GET':
        return err('Method not allowed', 405)
    project_no = (request.GET.get('project_no') or '').strip()
    short_name = (request.GET.get('short_name') or '').strip()
    payee = (request.GET.get('payee') or '').strip()
    empty = {'project_no': project_no, 'matched': False, 'count': 0,
             'total_balance': '0.00', 'items': []}
    if not (project_no or short_name or payee):
        return ok(empty)
    from ar.models import AdvanceRecord, ARProject
    proj = None
    if project_no:
        proj = ARProject.objects.filter(project_no=project_no).first()
    if proj is None and short_name:
        proj = ARProject.objects.filter(short_name=short_name).order_by('-id').first()

    match = Q()
    if proj is not None:
        match |= Q(project=proj)
    if payee:
        match |= Q(project__isnull=True, counterparty__iexact=payee)
    if not match:
        return ok(empty)
    qs = AdvanceRecord.objects.filter(match, direction='预付', balance_amount__gt=0)
    if request.pk_role != 'super_admin':
        qs = qs.filter(delivery_dept__in=request.pk_depts)
    agg = qs.aggregate(total=Sum('balance_amount'), cnt=Count('id'))
    total = (agg['total'] or Decimal('0')).quantize(Decimal('0.01'))
    items = [{
        'id': r.id,
        'counterparty': r.counterparty,
        'occur_date': str(r.occur_date) if r.occur_date else None,
        'advance_amount': str(r.advance_amount),
        'balance_amount': str(r.balance_amount),
    } for r in qs.order_by('-balance_amount')[:20]]
    return ok({
        'project_no': project_no,
        'matched': True,
        'short_name': proj.short_name if proj else '',
        'count': agg['cnt'] or 0,
        'total_balance': str(total),
        'items': items,
    })


def _parse_plan_item_body(data):
    """解析计划批次入参（编辑/追加共用）→ (planned_date, amount, notes) 或 (None, err)。"""
    raw_date = (data.get('planned_date') or '').strip()
    if not raw_date:
        return None, '计划日期必填'
    try:
        planned_date = datetime.date.fromisoformat(raw_date)
    except (ValueError, TypeError):
        return None, '计划日期格式有误（应为 YYYY-MM-DD）'
    try:
        amount = Decimal(str(data.get('amount')))
    except (InvalidOperation, ValueError, TypeError):
        return None, '计划金额格式有误'
    # Decimal('NaN')/('Infinity') 能构造但后续比较会抛 InvalidOperation → 显式拦掉，避免 500
    if not amount.is_finite():
        return None, '计划金额格式有误'
    if amount <= 0:
        return None, '计划金额必须大于0'
    notes = (data.get('notes') or '').strip()[:200]
    return (planned_date, amount, notes), None


@csrf_exempt
@pk_required()
def payment_plan_items(request, pk):
    """POST /payments/<pk>/plan-items — 给某条付款管理追加一批计划排款（排款管理）。

    与审批侧「分批排款」同口径：汇总金额/计划日（取最后一次排款批次日）随之派生刷新；
    该付款若来源于审批，审批的已排款累计与归档状态同步对账。若已挂审批，本批不得使累计
    排款额超过审批申请金额（剩余可排）。"""
    perms = get_request_perms(request)
    denied = _payments_page_denied(request, perms)
    if denied:
        return denied
    if request.method != 'POST':
        return err('Method not allowed', 405)
    if perms is not None and not perms.get('can_create'):
        return err('无排款操作权限', 403, 403)
    try:
        p = Payment.objects.get(pk=pk)
    except Payment.DoesNotExist:
        return err('排款记录不存在', 404)
    if not can_write_dept(request, p.department):
        return err('无权操作该部门', 403, 403)
    parsed, error = _parse_plan_item_body(parse_body(request))
    if error:
        return err(error)
    planned_date, amount, notes = parsed
    approval_id = p.approval_id   # FK 一经建立即不可变，非锁读取用于确定锁序归属
    with transaction.atomic():
        # 统一锁序：先审批后付款（与 _schedule_one 一致，避免 AB-BA 死锁）；
        # 剩余可排在锁内重算，消除 TOCTOU（并发追加各自读快照导致超额）。
        rec = (ApprovalRecord.objects.select_for_update().filter(pk=approval_id).first()
               if approval_id else None)
        p = Payment.objects.select_for_update().get(pk=pk)
        if rec:
            current = (PaymentPlanItem.objects.filter(payment__approval_id=approval_id)
                       .aggregate(s=Sum('amount'))['s'] or Decimal('0'))
            remaining = (rec.amount or Decimal('0')) - current
            if amount > remaining:
                return err(f'本批 {amount} 超过剩余可排 {remaining}'
                           f'（申请 {rec.amount} − 已排 {current}）', 400, 400)
        last = p.plan_items.order_by('-seq').first()
        seq = (last.seq + 1) if last else 1
        PaymentPlanItem.objects.create(payment=p, seq=seq, planned_date=planned_date,
                                       amount=amount, notes=notes or f'追加排款 第{seq}批')
        _sync_payment_plan(p)
        _reconcile_approval_schedule(approval_id)
        _record_plan_item_log(p, request,
                              new_desc=f'追加第{seq}批 {planned_date} ¥{amount}')
    p.refresh_from_db()
    return ok({'payment': p.to_dict(),
               'message': f'已追加计划批次 {amount}，汇总与审批已排款同步'})


@csrf_exempt
@pk_required()
def payment_plan_item_detail(request, pk, iid):
    """计划批次的编辑 / 撤销（排款管理）。

    PUT    /payments/<pk>/plan-items/<iid> — 编辑某批计划的日期/金额/备注。
    DELETE /payments/<pk>/plan-items/<iid> — 撤销某批计划。

    两者都会派生刷新汇总（total_amount=批次之和、planned_date=最后一次排款批次日），并把来源
    审批的已排款累计与归档状态重算对账。护栏：调整/撤销后计划合计不得低于 已付+预付
    冲抵；已挂审批时累计排款不得超过申请金额；最后一批不可单独撤销（请删除整条排款）。"""
    perms = get_request_perms(request)
    denied = _payments_page_denied(request, perms)
    if denied:
        return denied
    if request.method not in ('PUT', 'DELETE'):
        return err('Method not allowed', 405)
    if perms is not None and not perms.get('can_create'):
        return err('无排款操作权限', 403, 403)
    # 归属审批（FK 一经建立即不可变）：非锁读取仅用于确定锁序与存在性
    try:
        approval_id = Payment.objects.values_list('approval_id', flat=True).get(pk=pk)
    except Payment.DoesNotExist:
        return err('排款记录不存在', 404)

    # PUT 入参在锁外解析（无需持锁）
    if request.method == 'PUT':
        parsed, error = _parse_plan_item_body(parse_body(request))
        if error:
            return err(error)
        planned_date, amount, notes = parsed

    with transaction.atomic():
        # 统一锁序：先审批后付款（与 _schedule_one / 追加批次一致，避免 AB-BA 死锁）；
        # 所有护栏在锁内以最新数据重算，消除 TOCTOU（并发编辑/撤销读旧快照越界）。
        rec = (ApprovalRecord.objects.select_for_update().filter(pk=approval_id).first()
               if approval_id else None)
        try:
            p = Payment.objects.select_for_update().get(pk=pk)
        except Payment.DoesNotExist:
            return err('排款记录不存在', 404)
        if not can_write_dept(request, p.department):
            return err('无权操作该部门', 403, 403)
        try:
            item = PaymentPlanItem.objects.get(pk=iid, payment_id=pk)
        except PaymentPlanItem.DoesNotExist:
            return err('计划批次不存在（可能已撤销）', 404)
        paid = p.total_paid + (p.prepaid_offset_amount or Decimal('0'))

        if request.method == 'DELETE':
            if p.plan_items.count() <= 1:
                return err('最后一批计划不可单独撤销——如需作废整条排款请使用删除记录')
            after = (p.total_amount or Decimal('0')) - item.amount
            if after < paid:
                return err(f'撤销后计划合计 {after} 将低于 已付+冲抵 {paid}，不能撤销；'
                           f'请先删除对应付款明细')
            amt = item.amount
            seq_n = item.seq
            date_n = item.planned_date
            item.delete()
            _sync_payment_plan(p)
            _reconcile_approval_schedule(approval_id)
            _record_plan_item_log(p, request,
                                  old_desc=f'撤回第{seq_n}批 {date_n} ¥{amt}')
            p.refresh_from_db()
            return ok({'payment': p.to_dict(),
                       'message': f'已撤销该批计划 {amt}，汇总与审批已排款同步回退'})

        # PUT — 编辑批次
        old_date = item.planned_date
        old_amount = item.amount
        old_seq = item.seq
        after = (p.total_amount or Decimal('0')) - item.amount + amount
        if after < paid:
            return err(f'调整后计划合计 {after} 将低于 已付+冲抵 {paid}，不能调整；'
                       f'请先删除对应付款明细或调高金额')
        if rec and after > (rec.amount or Decimal('0')):
            return err(f'调整后累计排款 {after} 超过审批申请金额 {rec.amount}；'
                       f'如需超额请先修改审批记录的申请金额', 400, 400)
        item.planned_date = planned_date
        item.amount = amount
        item.notes = notes
        item.save(update_fields=['planned_date', 'amount', 'notes'])
        _sync_payment_plan(p)
        _reconcile_approval_schedule(approval_id)
        _record_plan_item_log(p, request,
                              old_desc=f'第{old_seq}批 {old_date} ¥{old_amount}',
                              new_desc=f'第{old_seq}批 {planned_date} ¥{amount}')
        p.refresh_from_db()
        return ok({'payment': p.to_dict(),
                   'message': f'已调整该批计划为 {planned_date} · {amount}，汇总与审批已排款同步'})


@csrf_exempt
@pk_required()
def payment_offsets(request, pk):
    """GET /payments/<pk>/offsets — 该排款已关联的预付核销列表（供行内反向核销）。"""
    perms = get_request_perms(request)
    denied = _payments_page_denied(request, perms)
    if denied:
        return denied
    if request.method != 'GET':
        return err('Method not allowed', 405)
    try:
        p = Payment.objects.get(pk=pk)
    except Payment.DoesNotExist:
        return err('排款记录不存在', 404)
    if not can_write_dept(request, p.department) and request.pk_role != 'super_admin':
        # 只读列表按部门可见性放行：dept_filter 同口径
        if p.department not in (request.pk_depts or []):
            return err('无权访问', 403)
    from ar.models import AdvanceWriteoff
    items = [{
        'id': w.id,
        'advance_id': w.advance_record_id,
        'counterparty': w.advance_record.counterparty,
        'amount': str(w.amount),
        'writeoff_date': str(w.writeoff_date) if w.writeoff_date else None,
        'notes': w.notes,
    } for w in (AdvanceWriteoff.objects.select_related('advance_record')
                .filter(payment_id=pk).order_by('writeoff_date', 'id'))]
    return ok({'items': items, 'total_offset': str(p.prepaid_offset_amount or 0)})


def _payment_status_bucket_q(status):
    """单个付款计划状态桶 → Q（口径与前端 StatusBadge / to_dict 一致，含预付冲抵）。"""
    if status == 'pending':
        return Q(paid=Decimal('0'), plan_adjustment__isnull=True,
                 prepaid_offset_amount=Decimal('0'))
    if status == 'settled':
        return (Q(paid__gte=F('total_amount') - F('prepaid_offset_amount')) |
                Q(plan_adjustment__isnull=False,
                  paid__gte=F('plan_adjustment') - F('prepaid_offset_amount')))
    if status == 'partial':
        return Q(paid__gt=Decimal('0'),
                 paid__lt=F('total_amount') - F('prepaid_offset_amount'),
                 plan_adjustment__isnull=True)
    if status == 'overdue':
        return Q(planned_date__lt=datetime.date.today()) & _not_settled_q('paid')
    if status == 'adjusted':
        return Q(plan_adjustment__isnull=False,
                 paid__lt=F('plan_adjustment') - F('prepaid_offset_amount'))
    return None


def _apply_payment_status_filter(qs, status_q, paid_annotated, hide_settled=False):
    """付款管理状态筛选：口径与前端 StatusBadge / to_dict 完全一致（含预付冲抵）。
    供 _list_payments 与 payment_export 共用，避免两处逻辑漂移。
    status_q 支持逗号分隔的多状态（多选并集，OR）。
    hide_settled=True 时排除已付清记录（进入页面默认非已付清视图）。
    """
    statuses = [s.strip() for s in (status_q or '').split(',') if s.strip()]
    if not statuses and not hide_settled:
        return qs
    if not paid_annotated:
        qs = qs.annotate(paid=_paid_expr())
    if not statuses and hide_settled:
        settled_q = (
            Q(paid__gte=F('total_amount') - F('prepaid_offset_amount')) |
            Q(plan_adjustment__isnull=False,
              paid__gte=F('plan_adjustment') - F('prepaid_offset_amount'))
        )
        return qs.exclude(settled_q)
    # 多选状态：各桶 Q 求并集
    combined = None
    for s in statuses:
        sub = _payment_status_bucket_q(s)
        if sub is not None:
            combined = sub if combined is None else (combined | sub)
    if combined is not None:
        qs = qs.filter(combined)
    return qs


def _apply_payment_computed_filters(qs, request):
    """付款管理「计算列」筛选 + 排序：已付(paid)/剩余(remaining)/逾期(overdue)。

    口径与 Payment.to_dict 显示值严格一致：
      · 已付   = 分期回款合计（_paid_expr）
      · 剩余   = max(0, 有效计划 − 已付 − 预付冲抵)（Greatest 钳到 0，与 property 一致）
      · 逾期=是 = 计划日期 < 今天 且 未结清（复用 _not_settled_q）
    仅在请求确有计算列筛选/排序时才注解，避免拖慢默认列表。
    返回 (qs, paid_annotated) 供下游 status 注解去重，防止 paid 重复注解报错。
    """
    raw = request.GET.get('filters', '')
    spec = {}
    try:
        if raw:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                spec = parsed
    except (ValueError, TypeError):
        spec = {}
    sort_f = (request.GET.get('sort') or '').strip()

    if not (any(k in spec for k in ('paid', 'remaining', 'overdue')) or sort_f in ('paid', 'remaining')):
        return qs, False

    qs = qs.annotate(paid=_paid_expr()).annotate(
        remaining_calc=Greatest(Value(Decimal('0'), output_field=_DEC18),
                                _effective_remaining_expr('paid'), output_field=_DEC18))

    # 已付/剩余：复用通用数值解析器（注解名即 col）
    fq, _ = build_filter_q(raw, PAYMENT_COMPUTED_REGISTRY)
    if fq:
        qs = qs.filter(fq)

    # 逾期(是/否)：派生布尔，单独构造
    ov = spec.get('overdue')
    if isinstance(ov, dict) and ov.get('op') == 'in':
        vals = {str(x) for x in (ov.get('value') or [])}
        overdue_q = Q(planned_date__lt=datetime.date.today()) & _not_settled_q('paid')
        if '是' in vals and '否' not in vals:
            qs = qs.filter(overdue_q)
        elif '否' in vals and '是' not in vals:
            qs = qs.exclude(overdue_q)
        # 同时勾选是+否 或 都不勾 → 不约束

    if sort_f in ('paid', 'remaining'):
        prefix = '-' if (request.GET.get('order') or '').lower() == 'desc' else ''
        col = 'remaining_calc' if sort_f == 'remaining' else 'paid'
        qs = qs.order_by(f'{prefix}{col}')

    return qs, True


def _parse_numbers(raw):
    """把用户粘贴的多单号串按任意分隔符（空格/换行/Tab/+/逗号/分号/顿号/竖线，中英标点）
    拆成去重列表，供「批量单号筛选」用。空串返回 []。"""
    if not raw:
        return []
    parts = re.split(r'[\s,+;|，、；／/]+', str(raw).strip())
    seen, out = set(), []
    for p in parts:
        p = p.strip()
        if p and p not in seen:
            seen.add(p)
            out.append(p)
    return out


def _payments_filtered_qs(request):
    """付款列表「筛选后、未分页」queryset（与列表完全同口径）：部门作用域 + 顶部筛选 +
    列头筛选 + 计算列筛选 + 状态筛选 + 排序。返回 (qs, paid_annotated)。
    供列表分页、跨页全选取 ID、运输单号复制等共用——保证「看到什么就操作/复制什么」。"""
    qs = Payment.objects.select_related('created_by').prefetch_related('installments', 'plan_items').filter(deleted_at__isnull=True)
    qs = dept_filter(qs, request)

    dept = request.GET.get('dept', '').strip()
    status_q = request.GET.get('status', '').strip()
    hide_settled = request.GET.get('hide_settled') == '1'
    start = request.GET.get('start_date', '').strip()
    end = request.GET.get('end_date', '').strip()
    pay_start = request.GET.get('pay_date_start', '').strip()
    pay_end = request.GET.get('pay_date_end', '').strip()
    g7_no = request.GET.get('g7_number', '').strip()
    q = request.GET.get('q', '').strip()

    if dept:
        qs = qs.filter(department=dept)
    if start:
        qs = qs.filter(planned_date__gte=start)
    if end:
        qs = qs.filter(planned_date__lte=end)
    if pay_start:
        qs = qs.filter(installments__pay_date__gte=pay_start).distinct()
    if pay_end:
        qs = qs.filter(installments__pay_date__lte=pay_end).distinct()
    if g7_no:
        qs = qs.filter(g7_number__icontains=g7_no)
    # 重点付款筛选：priority=1 仅看已标记
    if request.GET.get('priority') == '1':
        qs = qs.filter(is_priority=True)
    # 批量单号筛选：numbers=以逗号分隔的单号集合（前端把任意分隔符归一化后传入），
    # 命中 G7编号(对账单号) 或 审批编号 任一即保留。
    nums = _parse_numbers(request.GET.get('numbers', ''))
    if nums:
        qs = qs.filter(Q(g7_number__in=nums) | Q(approval_number__in=nums))
    if q:
        qs = qs.filter(
            Q(project_desc__icontains=q) | Q(payee__icontains=q) |
            Q(approval_number__icontains=q) | Q(g7_number__icontains=q) |
            Q(department__icontains=q) | Q(applicant__icontains=q)
        )

    # 列头精确筛选（filters JSON，白名单解析）+ 列头排序
    fq, fq_distinct = build_filter_q(request.GET.get('filters', ''), PAYMENT_FILTER_REGISTRY)
    if fq:
        qs = qs.filter(fq)
        if fq_distinct:
            qs = qs.distinct()
    sort_by = resolve_sort(request.GET.get('sort'), request.GET.get('order'), PAYMENT_FILTER_REGISTRY)
    if sort_by:
        qs = qs.order_by(sort_by)

    # 计算列（已付/剩余/逾期）筛选与排序：按需注解，复用既有金额口径表达式
    qs, _paid_annotated = _apply_payment_computed_filters(qs, request)
    # Status filter — shared helper 与 export 口径完全一致
    qs = _apply_payment_status_filter(qs, status_q, _paid_annotated, hide_settled=hide_settled)
    return qs, _paid_annotated


def _list_payments(request):
    qs, _paid_annotated = _payments_filtered_qs(request)
    pay_start = request.GET.get('pay_date_start', '').strip()
    pay_end = request.GET.get('pay_date_end', '').strip()

    try:
        page = max(1, int(request.GET.get('page', 1)))
        size = min(200, max(1, int(request.GET.get('size', 50))))
    except ValueError:
        page, size = 1, 50

    total = qs.count()

    # 已付口径：未加付款日期窗口时＝累计实付；加了窗口时＝窗口内实付（与「付款流水」对账一致）。
    # 「未结清/剩余」始终用累计实付（lifetime），以保证逾期/结清判断不受日期筛选影响。
    pay_window_active = bool(pay_start or pay_end)
    not_settled = _not_settled_q('_paid')
    effective_remaining = _effective_remaining_expr('_paid')
    annotated = qs.annotate(_paid=_paid_expr())
    if pay_window_active:
        annotated = annotated.annotate(_paid_win=_paid_subq_windowed(pay_start, pay_end))
        paid_total_expr = Sum('_paid_win')
    else:
        paid_total_expr = Sum('_paid')
    summary = annotated.aggregate(
        outstanding=Sum(effective_remaining, filter=not_settled),
        outstanding_count=Count('id', filter=not_settled),
        planned_total=Sum('total_amount'),
        paid_total=paid_total_expr,
    )
    outstanding_total = summary['outstanding'] or Decimal('0')
    outstanding_count = summary['outstanding_count'] or 0

    perms = get_request_perms(request)
    if pay_window_active:
        # 付款日期窗口：需逐行遍历分期算「窗口内实付」→ 保留预取、返回完整明细。
        page_slice = qs[(page - 1) * size: page * size]
        ws = _safe_iso_date(pay_start) if pay_start else None
        we = _safe_iso_date(pay_end) if pay_end else None
        items = []
        for p in page_slice:
            d = apply_view_mask(p.to_dict(), perms)
            # 实付列改为「窗口内实付」，与付款流水逐行对账；剩余/状态仍由 to_dict 按累计实付给出。
            if d.get('total_paid') is not None:
                win = sum((i.pay_amount for i in p.installments.all()
                           if (ws is None or i.pay_date >= ws)
                           and (we is None or i.pay_date <= we)), Decimal('0'))
                d['total_paid'] = str(win)
            items.append(d)
    else:
        # 轻量列表：分批/分期明细不内嵌（前端展开行时懒加载 GET /payments/<id>）。
        # 去掉两张子表预取，改用子查询注解已付/批次数 → 查询更省、响应体更小、渲染更快。
        page_slice = (qs.prefetch_related(None)
                        .annotate(_paid=_paid_subq(), _plan_cnt=_plan_count_subq())
                        [(page - 1) * size: page * size])
        items = [apply_view_mask(p.to_dict(light=True), perms) for p in page_slice]
    return ok({
        'items': items, 'total': total, 'page': page, 'size': size,
        'outstanding_total': str(outstanding_total),
        'outstanding_count': outstanding_count,
        'planned_total': str(summary['planned_total'] or Decimal('0')),
        'paid_total': str(summary['paid_total'] or Decimal('0')),
    })


def _clean_approval_number(raw):
    """清洗并校验审批编号：剔除空格、不可打印/控制/零宽等字符（导入 Excel 常见脏数据），
    清洗后空值占位为 21 个 0，其余须为 1–100 位「字母/数字/短横」（放宽纯数字限制，
    以兼容运输对账单号 ZD… 等含字母的编号）。返回 (cleaned, error)。"""
    s = ''.join(ch for ch in str(raw if raw is not None else '')
                if ch.isprintable() and not ch.isspace())
    if not s:
        return '0' * 21, None
    if not re.fullmatch(r'[0-9A-Za-z\-]{1,100}', s):
        return None, '审批编号清洗后须为1–100位字母/数字/短横（不填将自动占位为21个0）'
    return s, None


def _parse_payment_fields(data, payment=None):
    fields = {}

    def get(key, default=None):
        return data.get(key, getattr(payment, key, default) if payment else default)

    fields['department'] = (get('department') or '').strip()
    fields['secondary_dept'] = (get('secondary_dept') or '').strip()[:100]
    fields['project_short_name'] = (get('project_short_name') or '').strip()[:100]
    fields['applicant'] = (get('applicant') or '').strip()[:100]
    fields['approval_number'] = get('approval_number') or ''
    fields['g7_number'] = (get('g7_number') or '').strip()[:21]
    fields['project_no'] = (get('project_no') or '').strip()[:20]
    fields['project_desc'] = (get('project_desc') or '').strip()
    fields['payee'] = (get('payee') or '').strip()
    fields['notes'] = (get('notes') or '').strip()

    # plan_adjustment: nullable decimal — None means "no adjustment"
    pa_raw = get('plan_adjustment')
    if pa_raw not in (None, '', 0):
        pa_str = str(pa_raw).replace(',', '').replace('，', '').replace('¥', '').replace('￥', '').strip()
        try:
            pa_val = Decimal(pa_str or '0')
            if pa_val < Decimal('0'):
                return None, '计划调整金额不能为负数'
            fields['plan_adjustment'] = pa_val
        except (InvalidOperation, TypeError):
            return None, f'计划调整金额"{pa_raw}"格式有误，请填写纯数字金额（如 20000）'
    else:
        fields['plan_adjustment'] = None

    def _num(raw):
        s = str(raw if raw not in (None, '') else 0)
        s = s.replace(',', '').replace('，', '').replace('¥', '').replace('￥', '').strip()
        return Decimal(s or '0')

    try:
        fields['total_amount'] = _num(get('total_amount', 0))
    except (InvalidOperation, TypeError) as e:
        return None, f'计划总金额格式有误: {e}'
    if fields['total_amount'] < Decimal('0'):
        return None, '计划总金额不能为负数'
    if fields['total_amount'] <= Decimal('0'):
        return None, '计划总金额必须大于0'

    # Parse planned_date
    pd_val = data.get('planned_date') if 'planned_date' in data else getattr(payment, 'planned_date', None)
    fields['planned_date'] = pd_val or None
    if fields['planned_date']:
        try:
            datetime.date.fromisoformat(str(fields['planned_date']))
        except ValueError:
            return None, f'计划付款日期无效（{fields["planned_date"]}），请使用 YYYY-MM-DD 格式'

    # Parse installments list
    raw_insts = data.get('installments')
    if raw_insts is None and payment:
        # Keep existing installments when not provided on edit (handled separately)
        raw_insts = None
        fields['installments'] = None  # sentinel: do not update
    else:
        raw_insts = raw_insts or []
        parsed_insts = []
        total_paid = Decimal('0')
        for idx, item in enumerate(raw_insts, 1):
            try:
                amt = _num(item.get('pay_amount', 0))
            except (InvalidOperation, TypeError):
                return None, f'第{idx}次付款金额格式有误'
            if amt < Decimal('0'):
                return None, f'第{idx}次付款金额不能为负数'
            if amt <= Decimal('0'):
                return None, f'第{idx}次付款金额必须大于0'
            pay_date = (item.get('pay_date') or '').strip() or None
            if not pay_date:
                return None, f'第{idx}次付款日期必填'
            try:
                pay_date_d = datetime.date.fromisoformat(str(pay_date))
            except ValueError:
                return None, f'第{idx}次付款日期无效（{pay_date}），请使用 YYYY-MM-DD 格式'
            # 付款明细是已发生的现金事件：未来日期会让这笔钱既不算已付现金流、
            # 又被从刚性待付中扣掉，资金池两头都看不见
            if pay_date_d > datetime.date.today():
                return None, f'第{idx}次付款日期（{pay_date}）不能晚于今天——实际付款以发生日入账；计划性付款请用「计划付款日期」'
            total_paid += amt
            parsed_insts.append({
                'seq': idx,
                'pay_date': pay_date,
                'pay_amount': amt,
                'notes': (item.get('notes') or '').strip()[:200],
                'id': item.get('id'),  # for update — may be None for new rows
            })
        fields['installments'] = parsed_insts

    # ── 有效计划（调整额优先）与实付/冲抵的关系校验 ──────────────────────────
    # 覆盖：实付超出有效计划；把计划/调整额改到低于已付（会让剩余被截到0、状态假结清）
    effective_plan = (fields['plan_adjustment'] if fields['plan_adjustment'] is not None
                      else fields['total_amount'])
    if fields['installments'] is not None:
        paid_check = sum((i['pay_amount'] for i in fields['installments']), Decimal('0'))
    elif payment is not None:
        paid_check = payment.total_paid
    else:
        paid_check = Decimal('0')
    if paid_check > effective_plan:
        label = '计划调整金额' if fields['plan_adjustment'] is not None else '计划总金额'
        return None, (
            f'实付总额（{paid_check}元）超出{label}（{effective_plan}元），'
            '请核实金额后再提交'
        )

    if not fields['department']:
        return None, '部门必填'
    if fields['department'] not in VALID_DEPARTMENTS:
        return None, f'部门"{fields["department"]}"无效，请使用系统预设部门'
    if not fields['project_desc']:
        return None, '付款事项描述必填'
    if not fields['payee']:
        return None, '收款方必填'
    # Payee must be text (letters/Chinese/digits/common punctuation only — no HTML/control chars)
    if fields['payee'] and re.search(r'[<>{}\[\]|\\^~`@#$%&*=+]', fields['payee']):
        return None, '收款方仅允许输入文字内容，不能含特殊符号'
    if not fields['planned_date']:
        return None, '计划付款日期必填'
    # 审批单号：清洗（去空格/不可打印字符）后校验 1–100 位数字；空则占位为 21 个 0
    cleaned_no, err_no = _clean_approval_number(fields['approval_number'])
    if err_no:
        return None, err_no
    fields['approval_number'] = cleaned_no
    # 项目简称：填了就必须能在项目台账中找到（打通应收/现金流/分析/资金池的项目维度）
    err_psn = _validate_project_short_name(fields['project_short_name'], fields['department'])
    if err_psn:
        return None, err_psn
    # 项目编号已不再手填/导入：简称合法且编号为空时从台账自动带出，
    # 保持预付余额联动与历史按编号匹配的分析链路不断
    if fields['project_short_name'] and not fields['project_no']:
        from ar.models import ARProject
        no = (ARProject.objects.filter(short_name=fields['project_short_name'])
              .order_by('-id').values_list('project_no', flat=True).first())
        if no:
            fields['project_no'] = no[:20]

    return fields, None


def _validate_project_short_name(short_name, department=''):
    """校验项目简称必须与项目台账（ARProject.short_name）精确匹配。

    返回错误文案（str）或 None（通过）。留空视为「暂不关联项目」，直接通过。
    跨部门项目仅 is_shared（共享业务）的可被其他部门引用。
    """
    if not short_name:
        return None
    from ar.models import ARProject
    qs = ARProject.objects.filter(short_name=short_name)
    if not qs.exists():
        cand = list(ARProject.objects.filter(short_name__icontains=short_name)
                    .values_list('short_name', flat=True).distinct()[:5])
        hint = f'台账中相近的项目简称：{"、".join(cand)}。' if cand else ''
        return (f'项目简称「{short_name}」在项目台账中不存在。{hint}'
                f'请先在「项目台账」新建该项目，或通过输入框的模糊搜索从台账中选择'
                f'已有项目简称；留空表示暂不关联项目')
    if department and not qs.filter(
            Q(delivery_dept=department) | Q(is_shared=True)).exists():
        depts = sorted(set(qs.values_list('delivery_dept', flat=True)))
        return (f'项目简称「{short_name}」属于「{"/".join(d or "未填部门" for d in depts)}」，'
                f'与当前部门「{department}」不一致（非共享业务项目不可跨部门引用）。'
                f'请核对部门，或改填本部门项目台账中的项目简称')
    return None


# Field-name → Chinese-label map for change-log records.
_PAYMENT_FIELD_LABELS = {
    'department': '部门', 'secondary_dept': '二级部门', 'project_short_name': '项目简称',
    'applicant': '申请人', 'approval_number': '审批单号', 'g7_number': 'G7编号',
    'project_desc': '付款事项', 'payee': '收款方',
    'total_amount': '计划总金额', 'planned_date': '计划付款日期',
    'notes': '备注',
    'plan_adjustment': '计划调整金额',
}


def _find_duplicate_payment(fields, exclude_id=None):
    """Detect duplicate planned payments on same business key.

    Business key: department + approval_number + payee + planned_date + total_amount.
    Skip the rule when approval_number is blank or the 21-zero placeholder (low confidence).
    """
    no = fields.get('approval_number') or ''
    if not no or set(no) == {'0'}:
        return None
    # 仅与「在册」记录比对：软删记录已移入回收站、用户不可见，否则会以一条看不见的
    # 记录把合法重建判为重复（409），让财务陷入死胡同。与全局软删口径一致。
    qs = Payment.objects.filter(
        department=fields['department'],
        approval_number=fields['approval_number'],
        payee=fields['payee'],
        planned_date=fields['planned_date'],
        total_amount=fields['total_amount'],
        deleted_at__isnull=True,
    )
    if exclude_id:
        qs = qs.exclude(id=exclude_id)
    return qs.first()


def _record_payment_changes(payment, before_dict, after_dict, request, action='update'):
    """Persist field-level diffs into PaymentChangeLog."""
    user = PaikuanUser.objects.filter(id=request.pk_uid).only('id', 'name').first() \
        if getattr(request, 'pk_uid', None) else None
    operator_name = user.name if user else ''
    if action == 'create':
        PaymentChangeLog.objects.create(
            payment=payment, payment_id_snapshot=payment.id,
            action='create', operator=user, operator_name=operator_name,
            field_name='', field_label='整条记录',
            old_value='', new_value=f'新建排款 #{payment.id}',
        )
        return
    if action == 'delete':
        PaymentChangeLog.objects.create(
            payment=None, payment_id_snapshot=payment.id,
            action='delete', operator=user, operator_name=operator_name,
            field_name='', field_label='整条记录',
            old_value=f'排款 #{payment.id}', new_value='',
        )
        return
    # update: diff each tracked field
    logs = []
    for fname, label in _PAYMENT_FIELD_LABELS.items():
        old_v = before_dict.get(fname)
        new_v = after_dict.get(fname)
        if str(old_v if old_v is not None else '') != str(new_v if new_v is not None else ''):
            logs.append(PaymentChangeLog(
                payment=payment, payment_id_snapshot=payment.id,
                action='update', operator=user, operator_name=operator_name,
                field_name=fname, field_label=label,
                old_value=str(old_v if old_v is not None else ''),
                new_value=str(new_v if new_v is not None else ''),
            ))
    # diff installments as a whole
    old_insts = before_dict.get('installments_summary', '')
    new_insts = after_dict.get('installments_summary', '')
    if old_insts != new_insts:
        logs.append(PaymentChangeLog(
            payment=payment, payment_id_snapshot=payment.id,
            action='update', operator=user, operator_name=operator_name,
            field_name='installments', field_label='付款明细',
            old_value=old_insts, new_value=new_insts,
        ))
    if logs:
        PaymentChangeLog.objects.bulk_create(logs)


def _record_plan_item_log(payment, request, old_desc='', new_desc=''):
    """PaymentChangeLog entry for a plan-batch add/edit/withdraw (audit trail)."""
    user = (PaikuanUser.objects.filter(id=request.pk_uid).only('id', 'name').first()
            if getattr(request, 'pk_uid', None) else None)
    PaymentChangeLog.objects.create(
        payment=payment, payment_id_snapshot=payment.id,
        action='update', operator=user, operator_name=user.name if user else '',
        field_name='plan_items', field_label='计划批次',
        old_value=old_desc, new_value=new_desc,
    )


def _save_installments(payment, parsed_insts):
    """Replace all installments for a payment with the provided list."""
    payment.installments.all().delete()
    if parsed_insts:
        PaymentInstallment.objects.bulk_create([
            PaymentInstallment(
                payment=payment,
                seq=inst['seq'],
                pay_date=inst['pay_date'],
                pay_amount=inst['pay_amount'],
                notes=inst['notes'],
            )
            for inst in parsed_insts
        ])


def _installments_summary(insts_qs):
    """Human-readable summary of installments for change-log."""
    rows = list(insts_qs.order_by('seq', 'pay_date'))
    if not rows:
        return '无'
    return '；'.join(f'{i.pay_date} ¥{i.pay_amount}' for i in rows)


def _sync_payment_plan(p):
    """计划明细 → 汇总派生：total_amount=明细之和、planned_date=最后一次排款批次的计划日。
    多次分批排款时，汇总计划日以「最后一次排款」（seq 最大的批次）的计划日期为准；
    明细为空时不动（直建/导入路径会先建首条明细）。"""
    items = list(p.plan_items.all())
    if not items:
        return
    last = max(items, key=lambda i: i.seq)
    p.total_amount = sum(i.amount for i in items)
    p.planned_date = last.planned_date
    p.save(update_fields=['total_amount', 'planned_date', 'updated_at'])


def _reconcile_approval_schedule(approval_id):
    """审批「已排款」与「归档」状态 ←→ 付款管理「计划批次」的单一正源对账。

    口径：审批.已排款 == 该审批关联付款管理的全部计划批次(PaymentPlanItem)之和。
    任何会改变计划批次的付款管理写操作（编辑计划额/删除整条排款/编辑或撤销单批/
    追加批次）之后都应调用本函数，把 scheduled_amount 重算为批次之和、并据此重置
    archived（approved 记录按是否排满；rejected/canceled 维持终态归档）。

    幂等：以「批次之和」为正源整体重算，多次调用结果一致，不会因增量漂移失真。
    必须在改动 plan_items / 删除 Payment 之后调用（删除会级联清空批次→对账归零→
    审批回到可排款态）。应在 transaction.atomic 内调用以与本次写操作同生共死。"""
    if not approval_id:
        return
    rec = ApprovalRecord.objects.select_for_update().filter(pk=approval_id).first()
    if not rec:
        return
    scheduled = (PaymentPlanItem.objects
                 .filter(payment__approval_id=approval_id, payment__deleted_at__isnull=True)
                 .aggregate(s=Sum('amount'))['s'] or Decimal('0'))
    rec.scheduled_amount = scheduled
    # rejected/canceled 为终态，保持归档；approved 记录按是否排满申请金额决定归档
    if rec.status in {'rejected', 'canceled'}:
        rec.archived = True
    else:
        rec.archived = scheduled >= (rec.amount or Decimal('0'))
    rec.save(update_fields=['scheduled_amount', 'archived', 'updated_at'])


def _ensure_plan_item(p):
    """直建/导入的付款：生成首条计划明细（=计划日期+计划金额），维持派生不变量。"""
    if p.planned_date and not p.plan_items.exists():
        PaymentPlanItem.objects.create(payment=p, seq=1, planned_date=p.planned_date,
                                       amount=p.total_amount or 0)


def _create_payment(request):
    perms = get_request_perms(request)
    if perms is not None and not perms['can_create']:
        return err('无新增权限', 403, 403)
    data = parse_body(request)
    fields, error = _parse_payment_fields(data)
    if error:
        return err(error)
    if not can_write_dept(request, fields['department']):
        return err('无权操作该部门', 403, 403)
    parsed_insts = fields.pop('installments') or []
    # Idempotency: block duplicate submissions on the same business key.
    dup = _find_duplicate_payment(fields)
    if dup:
        return err(
            f'重复排款：已有相同审批单号({fields["approval_number"]})/收款方/计划日期/计划金额的排款记录 #{dup.id}',
            409, 409,
        )
    try:
        with transaction.atomic():
            p = Payment(created_by_id=request.pk_uid, updated_by_id=request.pk_uid, **fields)
            p.save()
            _ensure_plan_item(p)
            _save_installments(p, parsed_insts)
            _record_payment_changes(p, {}, {}, request, action='create')
    except IntegrityError:
        dup = _find_duplicate_payment(fields)
        ref = f' #{dup.id}' if dup else ''
        return err(
            f'重复排款：已有相同审批单号/收款方/计划日期/金额的排款记录{ref}',
            409, 409,
        )
    p_fresh = Payment.objects.prefetch_related('installments', 'plan_items').select_related('created_by').get(id=p.id)
    return ok(p_fresh.to_dict())


def _editable_payment_payload(data, perms):
    """Drop non-editable payment fields before merge+validation on update."""
    if perms is None:
        return data
    editable_cols = set()
    for f in PAYMENT_FIELD_DEFS:
        if perms['edit'].get(f['key'], False):
            editable_cols.update(f['cols'])
    # 'installments' is a virtual key — always include it when its permission is granted
    if 'installments' in editable_cols and 'installments' in data:
        editable_cols.add('installments')
    return {k: v for k, v in data.items() if k in editable_cols or k == 'installments'}


@csrf_exempt
@pk_required()
def payment_detail(request, pk):
    try:
        p = Payment.objects.select_related('created_by').prefetch_related('installments', 'plan_items').get(id=pk)
    except Payment.DoesNotExist:
        return err('记录不存在', 404)

    if not dept_filter(Payment.objects.filter(id=pk), request).exists():
        return err('无权访问', 403, 403)

    perms = get_request_perms(request)
    denied = _payments_page_denied(request, perms)
    if denied:
        return denied

    if request.method == 'GET':
        return ok(apply_view_mask(p.to_dict(), perms))

    if request.method == 'PUT':
        # Record-level access: super_admin or write access to this department.
        if not can_write_dept(request, p.department):
            return err('无权编辑此记录', 403, 403)
        data = _editable_payment_payload(parse_body(request), perms)
        fields, error = _parse_payment_fields(data, payment=p)
        if error:
            return err(error)
        new_dept = fields['department']
        if new_dept != p.department and not can_write_dept(request, new_dept):
            return err('无权操作目标部门', 403, 403)
        parsed_insts = fields.pop('installments')  # None means "keep existing"
        # 多批计划的汇总为派生值：直接改总额/计划日会失真，须经计划批次操作
        plan_n = p.plan_items.count()
        if plan_n > 1:
            from decimal import Decimal as _D
            if _D(str(fields.get('total_amount', p.total_amount))) != p.total_amount:
                return err('该排款含多批计划（来自审批分批排款），计划总金额=各批之和，'
                           '不能直接修改；请在排款明细中撤销/追加批次')
            if str(fields.get('planned_date', p.planned_date)) != str(p.planned_date):
                return err('该排款含多批计划，计划日期=最后一次排款批次日期，不能直接修改')
        # 单批且来源审批：改后的计划额=该审批已排款，不得超过审批申请金额（防经台账超额排款）
        if plan_n <= 1 and p.approval_id:
            new_total = Decimal(str(fields.get('total_amount', p.total_amount)))
            _rec = ApprovalRecord.objects.filter(pk=p.approval_id).first()
            if _rec and new_total > (_rec.amount or Decimal('0')):
                return err(f'计划额 {new_total} 超过来源审批申请金额 {_rec.amount}；'
                           f'如需超额请先修改审批记录的申请金额', 400, 400)
        before_snapshot = {f: getattr(p, f) for f in _PAYMENT_FIELD_LABELS}
        before_snapshot['installments_summary'] = _installments_summary(p.installments)
        with transaction.atomic():
            for k, v in fields.items():
                setattr(p, k, v)
            p.updated_by_id = request.pk_uid
            p.save()
            # 单批计划：总额/日期变更同步到首条计划明细（维持 汇总=明细之和）
            if plan_n <= 1:
                item = p.plan_items.first()
                if item is None:
                    _ensure_plan_item(p)
                elif (item.amount != p.total_amount
                      or str(item.planned_date) != str(p.planned_date)):
                    item.amount = p.total_amount
                    item.planned_date = p.planned_date
                    item.save(update_fields=['amount', 'planned_date'])
            if parsed_insts is not None:
                _save_installments(p, parsed_insts)
            # 计划额变更（单批与首条明细同步后）→ 审批「已排款/归档」对账，保持口径一致
            _reconcile_approval_schedule(p.approval_id)
            after_snapshot = {f: getattr(p, f) for f in _PAYMENT_FIELD_LABELS}
            after_snapshot['installments_summary'] = _installments_summary(p.installments)
            _record_payment_changes(p, before_snapshot, after_snapshot, request, action='update')
        p_fresh = Payment.objects.prefetch_related('installments', 'plan_items').select_related('created_by').get(id=p.id)
        return ok(apply_view_mask(p_fresh.to_dict(), perms))

    if request.method == 'DELETE':
        if perms is not None:
            if not perms['can_delete']:
                return err('无删除权限', 403, 403)
            if not can_write_dept(request, p.department):
                return err('无权删除此记录', 403, 403)
        # 已关联预付核销的排款不可直接删：FK 置空后核销悬空、
        # 预付台账与资金池的冲抵口径会失去对应关系
        if p.prepaid_offsets.exists():
            return err('该排款已关联预付核销，不能直接删除；'
                       '请先到「预收预付」删除对应核销记录后再删本排款', 409, 409)
        approval_id = p.approval_id   # 删除前留存：级联清空计划批次后据此把审批回退为可排款
        with transaction.atomic():
            _record_payment_changes(p, {}, {}, request, action='delete')
            p.delete()
            # 整条排款删除 → 该审批已排款归零、归档回退，可重新排款
            _reconcile_approval_schedule(approval_id)
        return ok({'deleted': pk})

    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def payment_change_logs(request, pk):
    """Return the change-log timeline for a single payment."""
    if request.method != 'GET':
        return err('Method not allowed', 405)
    # Permission: viewer must have access to this row (same dept or super_admin).
    try:
        p = Payment.objects.get(id=pk)
    except Payment.DoesNotExist:
        # If the payment was deleted, still allow viewing logs (via snapshot id)
        # but only for super_admin to avoid leaking history.
        if request.pk_role != 'super_admin':
            return err('记录不存在', 404)
        logs = PaymentChangeLog.objects.filter(payment_id_snapshot=pk).order_by('-at')[:200]
        return ok({'items': [l.to_dict() for l in logs], 'payment_id': pk, 'deleted': True})
    if not dept_filter(Payment.objects.filter(id=pk), request).exists():
        return err('无权访问', 403, 403)
    logs = PaymentChangeLog.objects.filter(payment_id_snapshot=pk).order_by('-at')[:200]
    return ok({'items': [l.to_dict() for l in logs], 'payment_id': pk, 'deleted': False})


@csrf_exempt
@pk_required()
def payment_installments(request):
    """GET — 付款流水：列出 PaymentInstallment 明细，支持付款日期范围/部门/关键字筛选。"""
    if request.method != 'GET':
        return err('Method not allowed', 405)
    perms = get_request_perms(request)
    denied = _payments_page_denied(request, perms)
    if denied:
        return denied

    # 排除已软删除付款记录的分期：付款台账删除后，其付款流水须一并隐藏（回收站可还原）
    qs = PaymentInstallment.objects.select_related('payment').filter(payment__deleted_at__isnull=True)
    # Dept visibility: mirror dept_filter but via related field payment__department
    if request.pk_role != 'super_admin':
        qs = qs.filter(payment__department__in=request.pk_depts or [])

    dept = request.GET.get('dept', '').strip()
    start = request.GET.get('pay_date_start', '').strip()
    end = request.GET.get('pay_date_end', '').strip()
    g7_no = request.GET.get('g7_number', '').strip()
    q = request.GET.get('q', '').strip()

    if dept:
        qs = qs.filter(payment__department=dept)
    if start:
        qs = qs.filter(pay_date__gte=start)
    if end:
        qs = qs.filter(pay_date__lte=end)
    if g7_no:
        qs = qs.filter(payment__g7_number__icontains=g7_no)
    if q:
        qs = qs.filter(
            Q(payment__project_desc__icontains=q) | Q(payment__payee__icontains=q) |
            Q(payment__approval_number__icontains=q) | Q(payment__g7_number__icontains=q) |
            Q(payment__applicant__icontains=q)
        )

    qs = qs.order_by('-pay_date', '-id')

    try:
        page = max(1, int(request.GET.get('page', 1)))
        size = min(200, max(1, int(request.GET.get('size', 50))))
    except ValueError:
        page, size = 1, 50

    total = qs.count()
    total_amount_val = qs.aggregate(s=Sum('pay_amount'))['s'] or Decimal('0')

    can_view_amounts = perms is None or perms['view'].get('total_amount', True)

    items = []
    for inst in qs[(page - 1) * size: page * size]:
        p = inst.payment
        items.append({
            'id': inst.id,
            'payment_id': p.id,
            'seq': inst.seq,
            'pay_date': str(inst.pay_date),
            'pay_amount': str(inst.pay_amount) if can_view_amounts else None,
            'notes': inst.notes,
            'department': p.department,
            'secondary_dept': p.secondary_dept,
            'project_short_name': p.project_short_name,
            'project_desc': p.project_desc,
            'payee': p.payee,
            'approval_number': p.approval_number,
            'g7_number': p.g7_number,
            'planned_date': str(p.planned_date) if p.planned_date else None,
        })

    return ok({
        'items': items, 'total': total, 'page': page, 'size': size,
        'total_amount': str(total_amount_val) if can_view_amounts else None,
    })


def _approvals_filtered_qs(request):
    """审批列表「筛选后、未分页」queryset（与列表完全同口径）：部门作用域 + 全局关键字 +
    列头筛选 + 计算列注解 + 排序。供列表分页与跨页全选取 ID 共用。

    可见口径：仅隐藏「审批通过且已排满」的归档记录（已完全流转至付款管理）；
    已拒绝/已撤销虽为终态（archived=True）但仍需可在审批管理查阅，故不再一刀切按
    archived 隐藏。默认只看待审批/审批通过由前端状态列筛选控制（可清除以查看全部）。"""
    qs = dept_filter(ApprovalRecord.objects.all(), request).exclude(status='approved', archived=True).filter(deleted_at__isnull=True)
    # 批量单号筛选：命中 G7编号(对账单号) 或 审批编号 任一即保留
    _nums = _parse_numbers(request.GET.get('numbers', ''))
    if _nums:
        qs = qs.filter(Q(g7_number__in=_nums) | Q(approval_number__in=_nums))
    # 全局关键字（跨字段模糊）+ 列头精确筛选（filters JSON）+ 列头排序
    kw = request.GET.get('q', '').strip()
    if kw:
        qs = qs.filter(
            Q(applicant__icontains=kw) | Q(department__icontains=kw) |
            Q(secondary_dept__icontains=kw) | Q(project_short_name__icontains=kw) |
            Q(approval_number__icontains=kw) | Q(g7_number__icontains=kw) |
            Q(summary__icontains=kw) | Q(payee__icontains=kw)
        )
    # 未排金额为计算列：仅当筛选/排序用到时再注解（amount − scheduled，钳 0）
    spec_raw = request.GET.get('filters', '')
    sort_f = (request.GET.get('sort') or '').strip()
    if 'remaining_amount' in spec_raw or sort_f == 'remaining_amount':
        qs = qs.annotate(remaining_calc=Greatest(
            Value(Decimal('0'), output_field=_DEC18),
            ExpressionWrapper(F('amount') - F('scheduled_amount'), output_field=_DEC18),
            output_field=_DEC18))
    merged_reg = {**APPROVAL_FILTER_REGISTRY, **APPROVAL_COMPUTED_REGISTRY}
    fq, fq_distinct = build_filter_q(spec_raw, merged_reg)
    if fq:
        qs = qs.filter(fq)
        if fq_distinct:
            qs = qs.distinct()
    sort_by = resolve_sort(request.GET.get('sort'), request.GET.get('order'), merged_reg)
    if sort_by:
        qs = qs.order_by(sort_by)
    return qs


@csrf_exempt
@pk_required()
def approval_records(request):
    perms = get_request_perms(request)
    if perms is not None and not perms['pages'].get('approval_records', True):
        return err('无访问权限', 403, 403)
    if request.method == 'GET':
        qs = _approvals_filtered_qs(request)
        # 全量（跨页）合计：总申请 / 已排合计 / 未排合计。合计仅统计在途标的
        # （待审批 / 审批通过），已拒绝/已撤销为终态不计入；在途记录 remaining =
        # amount - scheduled_amount ≥ 0，故未排合计 = 总申请 - 已排合计。
        live_qs = qs.exclude(status__in=['rejected', 'canceled'])
        sums = live_qs.aggregate(s=Sum('amount'), sched=Sum('scheduled_amount'))
        total_amount = sums['s'] or Decimal('0')
        total_scheduled = sums['sched'] or Decimal('0')
        total_remaining = total_amount - total_scheduled
        page = max(1, int(request.GET.get('page', 1) or 1))
        size = min(200, max(1, int(request.GET.get('size', 50) or 50)))
        total = qs.count()
        items = [o.to_dict() for o in qs[(page - 1) * size: page * size]]
        return ok({'items': items, 'total': total, 'page': page, 'size': size,
                   'total_amount': str(total_amount),
                   'total_scheduled': str(total_scheduled),
                   'total_remaining': str(total_remaining)})
    if request.method == 'POST':
        if perms is not None and not perms.get('can_create'):
            return err('无新增权限', 403, 403)
        data = parse_body(request)
        applicant = (data.get('applicant') or '').strip()
        department = (data.get('department') or '').strip()
        secondary_dept = (data.get('secondary_dept') or '').strip()[:100]
        project_short_name = (data.get('project_short_name') or '').strip()[:100]
        approval_number = (data.get('approval_number') or '').strip()
        g7_number_val = (data.get('g7_number') or '').strip()[:21]
        summary = (data.get('summary') or '').strip()
        notes = (data.get('notes') or '').strip()[:500]
        payee = (data.get('payee') or '').strip()
        status = (data.get('status') or 'pending').strip()
        amount = Decimal(str(data.get('amount') or '0'))
        if not applicant:
            return err('申请人不能为空')
        if amount <= 0:
            return err('申请金额必须大于0')
        # 审批编号：清洗后校验 1–100 位数字；空则自动占位 21 个 0
        approval_number, err_no = _clean_approval_number(approval_number)
        if err_no:
            return err(err_no)
        if department not in VALID_DEPARTMENTS:
            return err('所属事业部无效')
        if status not in {'pending', 'approved', 'rejected', 'canceled'}:
            return err('审批状态无效')
        if not can_write_dept(request, department):
            return err('无权操作该部门', 403, 403)
        # 仅有审批权限职务可直接登记 approved/rejected；其它人新建只能为 pending
        if status != 'pending' and not is_approver(request):
            return err('当前职务无权直接登记非待审批状态，请先创建为"待审批"', 403, 403)
        err_psn = _validate_project_short_name(project_short_name, department)
        if err_psn:
            return err(err_psn)
        rec = ApprovalRecord.objects.create(
            applicant=applicant, department=department, approval_number=approval_number,
            g7_number=g7_number_val,
            secondary_dept=secondary_dept, project_short_name=project_short_name,
            summary=summary, notes=notes, amount=amount, payee=payee, status=status,
            created_by_id=request.pk_uid
        )
        return ok(rec.to_dict())
    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def approval_record_detail(request, pk):
    perms = get_request_perms(request)
    if perms is not None and not perms['pages'].get('approval_records', True):
        return err('无访问权限', 403, 403)
    try:
        rec = ApprovalRecord.objects.get(pk=pk)
    except ApprovalRecord.DoesNotExist:
        return err('记录不存在', 404)
    if not can_write_dept(request, rec.department):
        return err('无权操作该部门', 403, 403)
    if request.method == 'PUT':
        data = parse_body(request)
        # 归档即终态（已排款 / 已拒绝 / 已撤销）：金额/状态等不可再改。
        # 例外：二级部门/项目简称是补录性元数据（历史数据默认空白，允许操作栏补录），
        # 仅改这两个字段不影响审批标的与资金池在途口径
        meta_only = bool(data) and set(data.keys()) <= {'secondary_dept', 'project_short_name'}
        if rec.archived and not meta_only:
            return err('该审批记录已归档（已排款或已拒绝/已撤销），不可修改；如需变更请新建记录'
                       '（二级部门/项目简称两个补录字段除外）', 409, 409)
        # 写入能力闸口：与新建一致（防止只读岗位改写审批台账）
        if perms is not None and not (perms.get('can_create') or is_approver(request)):
            return err('无编辑权限', 403, 403)
        if 'secondary_dept' in data:
            rec.secondary_dept = (data.get('secondary_dept') or '').strip()[:100]
        if 'project_short_name' in data:
            psn = (data.get('project_short_name') or '').strip()[:100]
            err_psn = _validate_project_short_name(psn, rec.department)
            if err_psn:
                return err(err_psn)
            rec.project_short_name = psn
        for k in ('applicant', 'summary', 'payee'):
            if k in data:
                setattr(rec, k, (data.get(k) or '').strip())
        if 'notes' in data:
            rec.notes = (data.get('notes') or '').strip()[:500]
        if 'department' in data and data['department'] in VALID_DEPARTMENTS:
            if not can_write_dept(request, data['department']):
                return err('无权操作目标事业部', 403, 403)
            rec.department = data['department']
        if 'approval_number' in data:
            cleaned_no, err_no = _clean_approval_number(data['approval_number'])
            if err_no:
                return err(err_no)
            rec.approval_number = cleaned_no
        if 'g7_number' in data:
            rec.g7_number = (data.get('g7_number') or '').strip()[:21]
        if 'amount' in data:
            try:
                new_amount = Decimal(str(data['amount'] or '0'))
            except (InvalidOperation, ValueError, TypeError):
                return err('申请金额格式错误')
            if new_amount <= 0:
                return err('申请金额必须大于0')
            # 金额是审批的标的：已审批通过的记录改金额会让资金池「在途支出」
            # 被悄悄改写，须先退回待审批再改
            if new_amount != rec.amount and rec.status != 'pending':
                return err('仅「待审批」状态可修改金额；已审批的记录请先退回待审批再修改')
            rec.amount = new_amount
        if 'status' in data and data['status'] in {'pending', 'approved', 'rejected', 'canceled'}:
            new_status = data['status']
            # 仅审批权限职务可设置 approved/rejected；任何登记人均可取消自己的申请
            if new_status in {'approved', 'rejected'} and not is_approver(request):
                return err('当前职务无权审批/拒绝该记录', 403, 403)
            if new_status == 'canceled' and not is_approver(request) and rec.created_by_id != request.pk_uid:
                return err('仅原申请人或审批人可撤销', 403, 403)
            rec.status = new_status
            if rec.status in {'rejected', 'canceled'}:
                rec.archived = True
        rec.save()
        return ok(rec.to_dict())
    return err('Method not allowed', 405)


def _schedule_one(request, rec, planned_date, total_amount):
    """核心排款：把 total_amount@planned_date 排入审批 rec 对应的付款管理汇总记录。
    返回 (payment, None, 200) 成功 或 (None, error_message, status)。status 为建议 HTTP 码：
    校验类错误用 400，冲突/防重类用 409。调用方负责页面/部门权限校验；此处复核
    状态/剩余可排 + 防重 + 原子写入。单条与批量排款共用，保证口径一致。"""
    if rec.status != 'approved':
        return None, '仅审批通过记录可排款', 400
    if rec.archived:
        return None, '记录已归档', 409
    if total_amount is None or total_amount <= 0:
        return None, '计划金额必须大于0', 400
    # 分批排款：本次金额不得超过剩余可排（申请金额 − 已排款累计）
    remaining = (rec.amount or Decimal('0')) - (rec.scheduled_amount or Decimal('0'))
    if total_amount > remaining:
        return None, (f'本次排款 {total_amount} 超过剩余可排 {remaining}'
                      f'（申请 {rec.amount} − 已排 {rec.scheduled_amount}）。'
                      f'如需超额请先修改审批记录的申请金额'), 400
    # 一条审批 ↔ 一条付款管理汇总记录（经 approval 静默ID打通）：
    # 首次排款建汇总记录+首条计划明细；再次排款只追加计划明细，汇总派生刷新
    existing = Payment.objects.filter(approval=rec).first()
    if existing is None and rec.approval_number and set(rec.approval_number) != {'0'}:
        # 自动收养：本审批在静默ID链路上线前用旧模式排过款（独立记录、未挂链）。
        # 业务键唯一匹配时挂上链转为追加批次，历史数据自愈；多条匹配（旧模式
        # 分批生成了多条）时不猜，保持独立。占位审批号(全0)不收养，避免误挂。
        cand = Payment.objects.filter(
            approval__isnull=True, approval_number=rec.approval_number,
            payee=rec.payee, department=rec.department)
        if cand.count() == 1:
            existing = cand.first()
            existing.approval = rec
            existing.save(update_fields=['approval', 'updated_at'])
            _ensure_plan_item(existing)   # 旧记录可能缺首条计划明细
    if existing is None:
        # 防历史重复（approval 链路之前的数据按业务键兜底查重）
        fields = {
            'department': rec.department,
            'approval_number': rec.approval_number,
            'payee': rec.payee,
            'planned_date': planned_date,
            'total_amount': total_amount,
        }
        dup = _find_duplicate_payment(fields)
        if dup:
            return None, (
                f'重复排款：已有相同审批单号/收款方/计划日期/金额的排款记录 #{dup.id}。'
                f'若这是旧版本分批生成的多条历史记录，请在付款管理核对后调整其中一条的'
                f'计划日期或金额再排，或换一个计划日期排本批'), 409
    else:
        # 防双击：仅拦 10 秒内连续提交的完全相同批次（日期+金额）。
        # 同日同金额的多批是真实业务需求（如同一天给不同子事项各排一笔），放行。
        from django.utils import timezone as _tz
        recent = existing.plan_items.filter(
            planned_date=planned_date, amount=total_amount,
            created_at__gte=_tz.now() - datetime.timedelta(seconds=10))
        if recent.exists():
            return None, ('检测到 10 秒内重复提交相同批次（日期+金额），已拦截疑似误触；'
                          '如确需同日再排一笔相同金额，请稍候片刻再提交'), 409
    try:
        with transaction.atomic():
            # 统一锁序：先锁审批，再锁付款（与台账侧排款管理一致，避免 AB-BA 死锁）
            rec_locked = ApprovalRecord.objects.select_for_update().get(pk=rec.pk)
            if rec_locked.archived:
                return None, '记录已归档', 409
            # 已排款以「关联付款管理的计划批次之和」为正源（兼容历史 scheduled_amount
            # 漂移与旧记录收养：收养时已 _ensure_plan_item 物化首条批次，此处即可如实计入）
            already = (PaymentPlanItem.objects.filter(payment__approval_id=rec_locked.pk)
                       .aggregate(s=Sum('amount'))['s'] or Decimal('0'))
            remaining = (rec_locked.amount or Decimal('0')) - already
            if total_amount > remaining:
                return None, f'本次排款 {total_amount} 超过剩余可排 {remaining}', 400
            if existing is None:
                p = Payment.objects.create(
                    created_by_id=request.pk_uid,
                    updated_by_id=request.pk_uid,
                    approval=rec_locked,
                    department=rec.department,
                    secondary_dept=rec.secondary_dept,
                    project_short_name=rec.project_short_name,
                    applicant=rec.applicant,
                    approval_number=rec.approval_number,
                    g7_number=rec.g7_number,   # G7编号随审批带入付款（运输事业部即对账单号）
                    project_desc=rec.summary,
                    payee=rec.payee,
                    notes=rec.notes,           # 备注随审批带入付款台账（运输事业部即运单号）
                    total_amount=total_amount,
                    planned_date=planned_date,
                )
                PaymentPlanItem.objects.create(payment=p, seq=1,
                                               planned_date=planned_date, amount=total_amount,
                                               notes='审批排款 第1批')
                _record_payment_changes(p, {}, {}, request, action='create')
            else:
                p = Payment.objects.select_for_update().get(pk=existing.pk)
                last = p.plan_items.order_by('-seq').first()
                seq = (last.seq + 1) if last else 1
                PaymentPlanItem.objects.create(payment=p, seq=seq,
                                               planned_date=planned_date, amount=total_amount,
                                               notes=f'审批排款 第{seq}批')
                _sync_payment_plan(p)
            # 已排款/归档以批次之和对账（单一正源），替代易漂移的增量累加；
            # 排满申请金额才归档，未排满留在审批管理可继续分批
            _reconcile_approval_schedule(rec_locked.pk)
    except IntegrityError:
        return None, '重复排款：相同业务键的排款记录已存在', 409
    rec.refresh_from_db()
    p.refresh_from_db()
    return p, None, 200


@csrf_exempt
@pk_required()
def approval_record_schedule(request, pk):
    if request.method != 'POST':
        return err('Method not allowed', 405)
    # 须同时具备审批页与排款页权限 + 新增权限
    perms = get_request_perms(request)
    if perms is not None:
        if not perms['pages'].get('approval_records', True):
            return err('无审批记录访问权限', 403, 403)
        if not perms['pages'].get('payments', True):
            return err('无排款台账访问权限', 403, 403)
        if not perms.get('can_create'):
            return err('无新增排款权限', 403, 403)
    try:
        rec = ApprovalRecord.objects.get(pk=pk, archived=False)
    except ApprovalRecord.DoesNotExist:
        return err('记录不存在', 404)
    if not can_write_dept(request, rec.department):
        return err('无权操作该部门', 403, 403)
    data = parse_body(request)
    planned_date = data.get('planned_date')
    total_amount = Decimal(str(data.get('total_amount') or '0'))
    if not planned_date or total_amount <= 0:
        return err('计划日期和计划金额必填')
    p, error, status = _schedule_one(request, rec, planned_date, total_amount)
    if error:
        return err(error, status, status)
    return ok({'payment': p.to_dict(), 'archived': rec.archived,
               'scheduled_amount': str(rec.scheduled_amount),
               'remaining_amount': str(max(Decimal('0'), rec.amount - rec.scheduled_amount)),
               'message': ('已排满申请金额，记录归档' if rec.archived else
                           f'本次排款 {total_amount} 已流转付款管理，剩余可排 {rec.amount - rec.scheduled_amount}')})


@pk_required()
def approval_budget_check(request):
    """排款超预算预警：查询某部门某月的预算总额与已排款总额，返回是否超预算。
    GET ?dept=<部门>&month=<yyyy-mm>&amount=<本次拟排金额>
    budget 取自 AR PaymentBudget（按 delivery_dept + expected_date 月份合计）。
    scheduled 取自 PaymentPlanItem（按 department + planned_date 月份合计）。
    """
    dept = request.GET.get('dept', '').strip()
    month = request.GET.get('month', '').strip()  # yyyy-mm
    amount_str = request.GET.get('amount', '0').strip()
    if not dept or not month:
        return err('dept 和 month 必填')
    try:
        year, mon = int(month[:4]), int(month[5:7])
        proposed = Decimal(amount_str) if amount_str else Decimal('0')
    except (ValueError, InvalidOperation):
        return err('参数格式错误')

    # 已排款：本月该部门在付款管理已安排的批次金额合计（department 在父表 Payment 上）
    scheduled_agg = (PaymentPlanItem.objects
                     .filter(payment__department=dept,
                             payment__deleted_at__isnull=True,
                             planned_date__year=year,
                             planned_date__month=mon)
                     .aggregate(s=Sum('amount')))
    scheduled = scheduled_agg['s'] or Decimal('0')

    # 预算：取 AR PaymentBudget（按 delivery_dept + expected_date 月份合计）
    budget = None
    try:
        from ar.models import PaymentBudget
        budget_agg = (PaymentBudget.objects
                      .filter(delivery_dept=dept,
                              expected_date__year=year,
                              expected_date__month=mon)
                      .aggregate(s=Sum('amount')))
        if budget_agg['s'] is not None:
            budget = budget_agg['s']
    except Exception:
        pass

    if budget is None:
        return ok({'has_budget': False, 'scheduled': str(scheduled), 'proposed': str(proposed)})

    remaining = budget - scheduled
    remaining_after = remaining - proposed
    return ok({
        'has_budget': True,
        'budget': str(budget),
        'scheduled': str(scheduled),
        'proposed': str(proposed),
        'remaining': str(remaining),
        'remaining_after': str(remaining_after),
        'over': remaining_after < Decimal('0'),
        'over_by': str(abs(remaining_after)) if remaining_after < Decimal('0') else '0',
    })


@csrf_exempt
@pk_required()
def approval_records_bulk_schedule(request):
    """批量排款：对所选审批记录各排一笔（默认金额=剩余可排=申请金额，默认日期=今天）。
    逐条独立处理，单条失败不影响其它；返回汇总（成功条数/金额合计）与失败明细。"""
    if request.method != 'POST':
        return err('POST only', 405)
    perms = get_request_perms(request)
    if perms is not None:
        if not perms['pages'].get('approval_records', True):
            return err('无审批记录访问权限', 403, 403)
        if not perms['pages'].get('payments', True):
            return err('无排款台账访问权限', 403, 403)
        if not perms.get('can_create'):
            return err('无新增排款权限', 403, 403)
    body = parse_body(request)
    # 兼容两种入参：items=[{id, amount}] 逐条指定金额（卡片内可调整）；
    # 或仅 ids=[] 时各取「剩余可排」为默认金额。
    raw_items = body.get('items')
    amount_map = {}
    if isinstance(raw_items, list) and raw_items:
        ids = []
        for it in raw_items:
            try:
                rid = int(it.get('id'))
            except (ValueError, TypeError, AttributeError):
                continue
            ids.append(rid)
            amt_raw = it.get('amount')
            if amt_raw not in (None, ''):
                try:
                    amount_map[rid] = Decimal(str(amt_raw))
                except (InvalidOperation, ValueError):
                    amount_map[rid] = None   # 标记金额非法，循环内跳过
    else:
        ids = body.get('ids') or []
    if not isinstance(ids, list) or not ids:
        return err('请提供要排款的记录 ids')
    try:
        ids = [int(i) for i in ids]
    except (ValueError, TypeError):
        return err('ids 必须为整数列表')
    if len(ids) > 5000:
        return err('单次批量排款上限 5000 条，请缩小选择范围')
    planned_date = body.get('planned_date') or datetime.date.today().isoformat()
    qs = dept_filter(ApprovalRecord.objects.filter(pk__in=ids, archived=False), request)
    recs = {r.id: r for r in qs}
    scheduled, total, skipped = 0, Decimal('0'), []
    for rid in ids:
        rec = recs.get(rid)
        if rec is None:
            skipped.append({'id': rid, 'reason': '不存在/已归档/无权限'})
            continue
        if not can_write_dept(request, rec.department):
            skipped.append({'id': rid, 'reason': '无权操作该部门'})
            continue
        # 默认本次金额=剩余可排（首次排款即等于申请金额）；卡片内逐条调整时用所填金额
        remaining = (rec.amount or Decimal('0')) - (rec.scheduled_amount or Decimal('0'))
        amount = amount_map.get(rid, remaining) if rid in amount_map else remaining
        if amount is None:
            skipped.append({'id': rid, 'reason': '金额格式有误'})
            continue
        p, error, status = _schedule_one(request, rec, planned_date, amount)
        if error:
            skipped.append({'id': rid, 'reason': error})
            continue
        scheduled += 1
        total += amount
    return ok({
        'scheduled': scheduled,
        'total_amount': str(total),
        'skipped': skipped,
        'message': f'已排款 {scheduled} 条，合计 {total}'
                   + (f'；跳过 {len(skipped)} 条' if skipped else ''),
    })


@csrf_exempt
@pk_required()
def approval_records_bulk_delete(request):
    """批量删除审批记录（含单选）。已关联付款管理（已排款）的记录不删，避免悬空排款。
    始终受部门作用域与删除权限约束；单次上限 5000 条。"""
    if request.method != 'POST':
        return err('POST only', 405)
    perms = get_request_perms(request)
    if perms is not None:
        if not perms['pages'].get('approval_records', True):
            return err('无访问权限', 403, 403)
        if not perms.get('can_delete'):
            return err('无删除权限', 403, 403)
    body = parse_body(request)
    ids = body.get('ids') or []
    if not isinstance(ids, list) or not ids:
        return err('请提供要删除的记录 ids')
    try:
        ids = [int(i) for i in ids]
    except (ValueError, TypeError):
        return err('ids 必须为整数列表')
    if len(ids) > 5000:
        return err('单次删除上限 5000 条，请缩小选择范围')
    qs = dept_filter(ApprovalRecord.objects.filter(pk__in=ids), request)
    now = timezone.now()
    actor = getattr(request, 'pk_user', None)
    deleted, skipped = 0, []
    for rec in list(qs):
        if not can_write_dept(request, rec.department):
            skipped.append({'id': rec.id, 'reason': '无权操作该部门'})
            continue
        if rec.payments.exists():
            skipped.append({'id': rec.id, 'reason': '已关联付款管理（已排款），不能删除；请先在付款管理删除对应排款'})
            continue
        rec.deleted_at = now
        rec.deleted_by = actor
        rec.save(update_fields=['deleted_at', 'deleted_by'])
        deleted += 1
    return ok({'deleted': deleted, 'skipped': skipped,
               'message': f'已移入回收站 {deleted} 条' + (f'；跳过 {len(skipped)} 条' if skipped else '')})


@csrf_exempt
@pk_required()
def approval_schedule_detail(request, pk):
    """GET /approvals/<pk>/schedule-detail —
    返回该审批关联的付款管理计划批次，供审批管理「排款管理」面板展示和操作。"""
    perms = get_request_perms(request)
    if perms is not None and not perms['pages'].get('approval_records', True):
        return err('无访问权限', 403, 403)
    if request.method != 'GET':
        return err('Method not allowed', 405)
    try:
        rec = ApprovalRecord.objects.get(pk=pk)
    except ApprovalRecord.DoesNotExist:
        return err('记录不存在', 404)
    if request.pk_role != 'super_admin' and rec.department not in (request.pk_depts or []):
        return err('无权访问', 403)
    payment = (Payment.objects.prefetch_related('plan_items', 'installments')
               .filter(approval=rec).first())
    if not payment:
        return ok({'payment_id': None, 'plan_items': [], 'total_paid': '0',
                   'total_amount': '0', 'can_edit': False})
    can_edit = perms is None or bool(perms.get('can_create'))
    return ok({
        'payment_id': payment.id,
        'plan_items': [
            {'id': pi.id, 'seq': pi.seq, 'planned_date': str(pi.planned_date),
             'amount': str(pi.amount), 'notes': pi.notes,
             'created_at': pi.created_at.isoformat() if pi.created_at else None}
            for pi in payment.plan_items.all()
        ],
        'total_paid': str(payment.total_paid),
        'total_amount': str(payment.total_amount),
        'can_edit': can_edit,
    })


@csrf_exempt
@pk_required()
def approval_records_bulk_return_schedule(request):
    """POST /approvals/bulk-return-schedule —
    批量退回排款：删除所选审批关联的付款管理记录，已排款归零、归档回退，审批可重新排款。
    单条失败不影响其它；返回汇总（成功条数）与失败明细（含原因）。"""
    if request.method != 'POST':
        return err('POST only', 405)
    perms = get_request_perms(request)
    if perms is not None:
        if not perms['pages'].get('approval_records', True):
            return err('无访问权限', 403, 403)
        if not perms.get('can_create'):
            return err('无排款管理权限', 403, 403)
    body = parse_body(request)
    ids = body.get('ids') or []
    if not isinstance(ids, list) or not ids:
        return err('请提供要退回排款的审批记录 ids')
    try:
        ids = [int(i) for i in ids]
    except (ValueError, TypeError):
        return err('ids 必须为整数列表')
    if len(ids) > 200:
        return err('单次批量退回上限 200 条，请缩小选择范围')
    qs = dept_filter(ApprovalRecord.objects.filter(pk__in=ids), request)
    returned, skipped = 0, []
    for rec in list(qs):
        if not can_write_dept(request, rec.department):
            skipped.append({'id': rec.id, 'reason': '无权操作该部门'})
            continue
        payment = Payment.objects.filter(approval=rec).first()
        if not payment:
            skipped.append({'id': rec.id, 'reason': '未找到关联排款记录'})
            continue
        if payment.prepaid_offsets.exists():
            skipped.append({'id': rec.id,
                            'reason': '排款已关联预付核销，不能退回；请先删除核销记录'})
            continue
        with transaction.atomic():
            _record_payment_changes(payment, {}, {}, request, action='delete')
            payment.delete()
            _reconcile_approval_schedule(rec.id)
        returned += 1
    return ok({
        'returned': returned,
        'skipped': skipped,
        'message': (f'已退回 {returned} 条排款'
                    + (f'；跳过 {len(skipped)} 条' if skipped else '')),
    })


@csrf_exempt
@pk_required()
def approval_records_bulk_approve(request):
    """批量审批通过（含单选）：把所选「待审批」记录置为审批通过。
    与单条审批同口径：仅审批职务可操作（is_approver），受部门作用域约束；
    非「待审批」状态（已通过/已拒绝/已撤销/已归档）逐条跳过，不报错。单次上限 5000 条。"""
    if request.method != 'POST':
        return err('POST only', 405)
    perms = get_request_perms(request)
    if perms is not None and not perms['pages'].get('approval_records', True):
        return err('无访问权限', 403, 403)
    # 审批通过是终态决策，仅审批职务可操作（与单条 PUT status=approved 一致）
    if not is_approver(request):
        return err('当前职务无权审批', 403, 403)
    body = parse_body(request)
    ids = body.get('ids') or []
    if not isinstance(ids, list) or not ids:
        return err('请提供要审批的记录 ids')
    try:
        ids = [int(i) for i in ids]
    except (ValueError, TypeError):
        return err('ids 必须为整数列表')
    if len(ids) > 5000:
        return err('单次审批上限 5000 条，请缩小选择范围')
    qs = dept_filter(ApprovalRecord.objects.filter(pk__in=ids), request)
    recs = {r.id: r for r in qs}
    approved, skipped = 0, []
    for rid in ids:
        rec = recs.get(rid)
        if rec is None:
            skipped.append({'id': rid, 'reason': '不存在/无权限'})
            continue
        if not can_write_dept(request, rec.department):
            skipped.append({'id': rid, 'reason': '无权操作该部门'})
            continue
        if rec.status != 'pending':
            skipped.append({'id': rid, 'reason': '非「待审批」状态，已跳过'})
            continue
        rec.status = 'approved'
        rec.save(update_fields=['status', 'updated_at'])
        approved += 1
    return ok({'approved': approved, 'skipped': skipped,
               'message': f'已审批通过 {approved} 条' + (f'；跳过 {len(skipped)} 条' if skipped else '')})


@csrf_exempt
@pk_required()
def payments_mark_priority(request):
    """重点付款打标/清除。
    POST { ids:[...], value:true/false }：对所选记录设置/取消重点标记；
    POST { all:true }：一键清除当前部门作用域内全部重点标记。"""
    if request.method != 'POST':
        return err('POST only', 405)
    perms = get_request_perms(request)
    denied = _payments_page_denied(request, perms)
    if denied:
        return denied
    body = parse_body(request)
    if body.get('all'):
        # 一键清除全部标记（作用域内）：.update 高效且不触发 save()，is_priority 不影响 dedup_key
        n = dept_filter(Payment.objects.filter(deleted_at__isnull=True, is_priority=True), request) \
            .update(is_priority=False)
        return ok({'count': n, 'value': False})
    value = bool(body.get('value', True))
    ids = [int(i) for i in (body.get('ids') or [])]
    if not ids:
        return err('ids 必填或传 all:true')
    if len(ids) > 5000:
        return err('单次上限 5000 条')
    n = dept_filter(Payment.objects.filter(pk__in=ids, deleted_at__isnull=True), request) \
        .update(is_priority=value)
    return ok({'count': n, 'value': value})


@csrf_exempt
@pk_required()
def payments_bulk_delete(request):
    """批量删除付款管理记录（含单选）。已关联预付核销的记录不删（同单条删除口径）。
    始终受部门作用域与删除权限约束；单次上限 5000 条。"""
    if request.method != 'POST':
        return err('POST only', 405)
    perms = get_request_perms(request)
    denied = _payments_page_denied(request, perms)
    if denied:
        return denied
    if perms is not None and not perms.get('can_delete'):
        return err('无删除权限', 403, 403)
    body = parse_body(request)
    ids = body.get('ids') or []
    if not isinstance(ids, list) or not ids:
        return err('请提供要删除的记录 ids')
    try:
        ids = [int(i) for i in ids]
    except (ValueError, TypeError):
        return err('ids 必须为整数列表')
    if len(ids) > 5000:
        return err('单次删除上限 5000 条，请缩小选择范围')
    qs = dept_filter(Payment.objects.filter(pk__in=ids, deleted_at__isnull=True), request)
    now = timezone.now()
    actor = getattr(request, 'pk_user', None)
    deleted, skipped = 0, []
    for p in list(qs.prefetch_related('prepaid_offsets')):
        if not can_write_dept(request, p.department):
            skipped.append({'id': p.id, 'reason': '无权操作该部门'})
            continue
        if p.prepaid_offsets.exists():
            skipped.append({'id': p.id, 'reason': '已关联预付核销，不能删除；请先删除对应核销记录'})
            continue
        approval_id = p.approval_id
        with transaction.atomic():
            _record_payment_changes(p, {}, {}, request, action='delete')
            p.deleted_at = now
            p.deleted_by = actor
            p.save(update_fields=['deleted_at', 'deleted_by'])
            _reconcile_approval_schedule(approval_id)
        deleted += 1
    return ok({'deleted': deleted, 'skipped': skipped,
               'message': f'已移入回收站 {deleted} 条' + (f'；跳过 {len(skipped)} 条' if skipped else '')})


@csrf_exempt
@pk_required()
def payments_bulk_pay(request):
    """批量付款（批量编辑）：对所选付款记录各追加一笔付款明细（默认金额=剩余应付=计划金额，
    默认日期=今天）。逐条独立处理；返回汇总（成功条数/金额合计）与失败明细。"""
    if request.method != 'POST':
        return err('POST only', 405)
    perms = get_request_perms(request)
    denied = _payments_page_denied(request, perms)
    if denied:
        return denied
    # 记录付款=编辑「付款明细」字段：以该字段的编辑权限闸口
    if perms is not None and not perms['edit'].get('installments', False):
        return err('无登记付款明细的权限', 403, 403)
    body = parse_body(request)
    # 兼容两种入参：items=[{id, amount}] 逐条指定本次付款金额（卡片内可调整）；
    # 或仅 ids=[] 时各取「剩余应付」为默认金额。
    raw_items = body.get('items')
    amount_map = {}
    if isinstance(raw_items, list) and raw_items:
        ids = []
        for it in raw_items:
            try:
                rid = int(it.get('id'))
            except (ValueError, TypeError, AttributeError):
                continue
            ids.append(rid)
            amt_raw = it.get('amount')
            if amt_raw not in (None, ''):
                try:
                    amount_map[rid] = Decimal(str(amt_raw))
                except (InvalidOperation, ValueError):
                    amount_map[rid] = None   # 标记金额非法，循环内跳过
    else:
        ids = body.get('ids') or []
    if not isinstance(ids, list) or not ids:
        return err('请提供要付款的记录 ids')
    try:
        ids = [int(i) for i in ids]
    except (ValueError, TypeError):
        return err('ids 必须为整数列表')
    if len(ids) > 5000:
        return err('单次批量付款上限 5000 条，请缩小选择范围')
    pay_date = body.get('pay_date') or datetime.date.today().isoformat()
    notes = (body.get('notes') or '批量付款').strip()[:200]
    qs = dept_filter(Payment.objects.filter(pk__in=ids), request)
    paid_cnt, total, skipped = 0, Decimal('0'), []
    for p in list(qs.prefetch_related('installments')):
        if not can_write_dept(request, p.department):
            skipped.append({'id': p.id, 'reason': '无权操作该部门'})
            continue
        remaining = p.remaining  # 剩余应付（=计划金额 − 已付 − 预付冲抵）
        amount = amount_map.get(p.id, remaining) if p.id in amount_map else remaining
        if amount is None:
            skipped.append({'id': p.id, 'reason': '金额格式有误'})
            continue
        if amount <= 0:
            skipped.append({'id': p.id, 'reason': '本次付款金额必须大于0'})
            continue
        if amount > remaining:
            skipped.append({'id': p.id, 'reason': f'本次付款 {amount} 超过剩余应付 {remaining}'})
            continue
        before = {f: getattr(p, f) for f in _PAYMENT_FIELD_LABELS}
        before['installments_summary'] = _installments_summary(p.installments)
        with transaction.atomic():
            last = p.installments.order_by('-seq').first()
            seq = (last.seq + 1) if last else 1
            PaymentInstallment.objects.create(payment=p, seq=seq, pay_date=pay_date,
                                               pay_amount=amount, notes=notes)
            p.updated_by_id = request.pk_uid
            p.save(update_fields=['updated_by', 'updated_at'])
            after = {f: getattr(p, f) for f in _PAYMENT_FIELD_LABELS}
            after['installments_summary'] = _installments_summary(p.installments)
            _record_payment_changes(p, before, after, request, action='update')
        paid_cnt += 1
        total += amount
    return ok({'paid': paid_cnt, 'total_amount': str(total), 'skipped': skipped,
               'message': f'已付款 {paid_cnt} 条，合计 {total}'
                          + (f'；跳过 {len(skipped)} 条' if skipped else '')})


@pk_required()
def approval_template(request):
    if request.method != 'GET':
        return err('Method not allowed', 405)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return err('服务器缺少 openpyxl 依赖', 500)
    wb = Workbook()
    ws = wb.active
    ws.title = '审批记录导入模板'
    headers = ['申请人*', '所属事业部*', '二级部门', '项目简称', '审批编号*', 'G7编号', '摘要', '申请金额*', '收款主体', '审批状态*']
    header_fill = PatternFill(fill_type='solid', fgColor='C96342')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h); c.fill = header_fill; c.font = header_font; c.alignment = center
    ws.append(['示例张三', '运输事业部', '', '', '123456789012345678901', '', '示例摘要', 10000, '某供应商', '待审批'])
    return _build_excel_response(wb, '审批记录导入模板.xlsx')


def _read_import_rows(upload):
    """把上传的表格文件读成行列表（每行 tuple，rows[0] 为表头），兼容多种格式：
    .xlsx/.xlsm（openpyxl）、.csv（UTF-8/GBK 自适应编码）；旧版 .xls 给出可操作的改存提示。
    返回 (rows, error)。"""
    name = (getattr(upload, 'name', '') or '').lower()
    try:
        data = upload.read()
    except Exception:
        return None, '文件读取失败，请重试'
    if not data:
        return None, '文件为空'
    head = data[:4]
    # .xlsx/.xlsm 本质是 zip（magic 'PK\x03\x04'）
    if head[:2] == b'PK' or name.endswith(('.xlsx', '.xlsm')):
        try:
            import io
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            return rows, None
        except Exception:
            return None, '文件已损坏或不是有效的 .xlsx，请用 Excel 重新「另存为」.xlsx 后再试'
    # 旧版 .xls 二进制（OLE2，magic D0 CF 11 E0）—— openpyxl 读不了，给可操作提示
    if head == b'\xd0\xcf\x11\xe0' or name.endswith('.xls'):
        return None, '检测到旧版 .xls 文件：请在 Excel 中「另存为」→「Excel 工作簿(.xlsx)」或「CSV」后再导入'
    # 其余按 CSV 处理（编码自适应：UTF-8 带/不带 BOM、GBK）
    text = None
    for enc in ('utf-8-sig', 'gbk', 'utf-8'):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        return None, '无法识别的文件格式或编码，请上传 .xlsx，或 UTF-8/GBK 编码的 .csv'
    import csv
    import io
    rows = [tuple(row) for row in csv.reader(io.StringIO(text))]
    if not rows:
        return None, '文件为空'
    return rows, None


# ── 审批导入：解析/校验抽成复用函数，供「导入」与「导入预检」同口径调用 ──────────
_APPROVAL_COLUMNS = [
    {'key': 'applicant', 'label': '申请人'},
    {'key': 'department', 'label': '所属事业部'},
    {'key': 'amount', 'label': '申请金额'},
    {'key': 'approval_number', 'label': '审批编号'},
    {'key': 'g7_number', 'label': 'G7编号'},
    {'key': 'payee', 'label': '收款主体'},
    {'key': 'secondary_dept', 'label': '二级部门'},
    {'key': 'project_short_name', 'label': '项目简称'},
    {'key': 'summary', 'label': '摘要', 'wide': True},
    {'key': 'status', 'label': '审批状态'},
]


def _approval_reader(rows):
    """从行列表构造 (cv, max_row, missing)。cv(r,*names) 取规范化表头对应单元格文本。
    表头去掉 *＊空格后匹配并支持别名——模板(带*)、导出文件(不带*)、轻微改名都能识别。"""
    max_row = len(rows)
    max_col = max((len(r) for r in rows), default=0)

    def cell(r, c):  # 1-indexed
        row = rows[r - 1] if 0 < r <= max_row else ()
        return row[c - 1] if 0 < c <= len(row) else None

    def _norm(h):
        return str(h or '').strip().replace('*', '').replace('＊', '').replace(' ', '').replace('　', '')
    norm_col = {}
    for c in range(1, max_col + 1):
        key = _norm(cell(1, c))
        if key and key not in norm_col:
            norm_col[key] = c

    def cv(r, *names):
        for n in names:
            ci = norm_col.get(_norm(n))
            if ci:
                return str(cell(r, ci) or '').strip()
        return ''
    missing = not (norm_col.get('申请人') and (norm_col.get('申请金额') or norm_col.get('金额')))
    return cv, max_row, missing


def _approval_row_view(cv, r):
    """读一行为可编辑的清洗字段 dict（与导入同口径取值）。"""
    return {
        'applicant': cv(r, '申请人'),
        'department': cv(r, '所属事业部', '事业部', '部门'),
        'secondary_dept': cv(r, '二级部门'),
        'project_short_name': cv(r, '项目简称'),
        'approval_number': cv(r, '审批编号', '审批单号'),
        'g7_number': cv(r, 'G7编号'),
        'summary': cv(r, '摘要'),
        'amount': cv(r, '申请金额', '金额'),
        'payee': cv(r, '收款主体'),
        'status': cv(r, '审批状态', '状态') or '待审批',
    }


def _approval_row_is_blank(v):
    return not any([v['applicant'], v['department'], v['approval_number'], v['amount'],
                    v['summary'], v['payee']])


def _parse_approval_fields(data, request):
    """校验+规范化一行审批数据；返回 (fields, error)。与 approval_import 同口径。"""
    applicant = (data.get('applicant') or '').strip()
    dept = (data.get('department') or '').strip()
    no = (data.get('approval_number') or '').strip()
    status_cn = (data.get('status') or '待审批').strip()
    amount_raw = str(data.get('amount') or '').strip()
    if not applicant or not amount_raw:
        return None, '申请人和金额不能为空'
    no_clean, err_no = _clean_approval_number(no)
    if err_no:
        return None, f'{err_no}（当前“{no}”）'
    no = no_clean
    if status_cn not in {'待审批', 'pending'}:
        return None, f'仅允许导入“待审批”状态（当前“{status_cn}”）'
    if dept not in VALID_DEPARTMENTS:
        return None, f'所属事业部无效（当前“{dept or "空"}”）'
    if not can_write_dept(request, dept):
        return None, f'无权操作事业部 {dept}'
    try:
        amount = Decimal(str(amount_raw).replace(',', '').replace('，', '').replace('¥', '').strip())
        if amount <= 0:
            raise ValueError()
    except Exception:
        return None, f'申请金额无效（当前“{amount_raw}”）'
    # 项目简称：填了就必须与项目台账匹配，搜索不到直接拒绝该行并给出指导
    psn = (data.get('project_short_name') or '')[:100]
    err_psn = _validate_project_short_name(psn, dept)
    if err_psn:
        return None, err_psn
    return {
        'applicant': applicant, 'department': dept, 'approval_number': no,
        'g7_number': (data.get('g7_number') or '')[:21],
        'secondary_dept': (data.get('secondary_dept') or '')[:100], 'project_short_name': psn,
        'summary': data.get('summary') or '', 'amount': amount, 'payee': data.get('payee') or '',
        'status': 'pending',
    }, None


@csrf_exempt
@pk_required()
def approval_import(request):
    if request.method != 'POST':
        return err('Method not allowed', 405)
    perms = get_request_perms(request)
    if perms is not None:
        if not perms['pages'].get('approval_records', True):
            return err('无访问权限', 403, 403)
        if not perms.get('can_create'):
            return err('无新增权限', 403, 403)
    f = request.FILES.get('file')
    if not f:
        return err('请上传文件')
    rows, ferr = _read_import_rows(f)
    if ferr:
        return err(ferr)
    if not rows:
        return err('文件为空')
    cv, max_row, missing = _approval_reader(rows)
    # 必要列缺失（多为表头被删/改名/选错文件）→ 明确报错，而不是静默全跳过
    if missing:
        return err('未识别到必要表头（申请人 / 申请金额）。请使用最新模板，或确认首行表头与列内容未被删改、且未选错文件。')

    created = skipped = 0
    errors = []
    for r in range(2, max_row + 1):
        view = _approval_row_view(cv, r)
        # 整行空白：略过（容忍尾部空行）；模板示例行：静默跳过
        if _approval_row_is_blank(view):
            continue
        if view['applicant'].startswith('示例') or view['summary'].startswith('示例'):
            continue
        fields, error = _parse_approval_fields(view, request)
        if error:
            skipped += 1; errors.append(f'第{r}行: {error}'); continue
        ApprovalRecord.objects.create(created_by_id=request.pk_uid, **fields)
        created += 1
    return ok({'created': created, 'skipped': skipped, 'errors': errors})


_AI_REVIEW_APPROVAL_SYS = (
    '你是企业审批台账的数据质检助手。下面是一批待导入的审批记录（已通过基础格式校验）。'
    '请只挑出"疑似有问题"的行，找规则难以发现的软问题：申请人/收款主体像乱码/测试数据/占位符；'
    '申请金额明显异常（疑似多或少一个0、过大或过小）；同一审批编号下申请人/金额/收款主体相互矛盾；'
    '摘要与金额/部门明显不符；疑似重复行。严格只返回 JSON 数组，每个元素形如 '
    '{"row":行号,"field":"字段名","issue":"问题简述","suggestion":"修正建议(可空)","severity":"high|medium|low"}。'
    '没有发现问题就返回 []。不要输出 JSON 以外的任何文字。'
)


def _ai_review_records(records, system_prompt):
    """通用：用 DeepSeek 复核一批待导入记录，挑出规则覆盖不到的软问题。
    best-effort：无 key / 异常 / 无数据时返回 []。返回 [{'row','field','issue','suggestion','severity'}]。"""
    if not getattr(settings, 'DEEPSEEK_API_KEY', '') or not records:
        return []
    try:
        from caiwu.views import _deepseek_chat
    except Exception:
        return []
    import json as _json
    try:
        text = _deepseek_chat(
            [{'role': 'system', 'content': system_prompt},
             {'role': 'user', 'content': _json.dumps(records[:50], ensure_ascii=False)}],
            timeout=60, max_tokens=1500)
        arr = _json.loads(text[text.index('['):text.rindex(']') + 1])
    except Exception:
        return []
    out = []
    for it in arr if isinstance(arr, list) else []:
        if isinstance(it, dict) and it.get('row') is not None:
            sev = it.get('severity')
            out.append({
                'row': it.get('row'), 'field': str(it.get('field', ''))[:40],
                'issue': str(it.get('issue', ''))[:200],
                'suggestion': str(it.get('suggestion', ''))[:200],
                'severity': sev if sev in ('high', 'medium', 'low') else 'medium',
            })
    return out


def _ai_review_approvals(records):
    return _ai_review_records(records, _AI_REVIEW_APPROVAL_SYS)


def _approval_corrected_xlsx(rows):
    """把修正后的审批行写成标准模板格式 .xlsx 返回，可直接再次导入。"""
    from openpyxl import Workbook
    cols = [('申请人', 'applicant'), ('所属事业部', 'department'), ('二级部门', 'secondary_dept'),
            ('项目简称', 'project_short_name'), ('审批编号', 'approval_number'), ('摘要', 'summary'),
            ('申请金额', 'amount'), ('收款主体', 'payee'), ('审批状态', 'status')]
    wb = Workbook(); ws = wb.active; ws.title = '审批记录(修正版)'
    ws.append([h for h, _ in cols])
    for r in rows:
        d = r.get('data', r)
        ws.append([d.get(c, '') if d.get(c) is not None else '' for _, c in cols])
    return _build_excel_response(wb, '审批记录_修正版.xlsx')


@csrf_exempt
@pk_required()
def approval_import_precheck(request):
    """审批导入预检：①规则预检（复用 _parse_approval_fields，与导入同口径）②AI 复核疑点。
    只读不落库；只返回"需关注"的行供前端展示编辑，其余通过的行放 okRows 确认时一并导入。"""
    if request.method != 'POST':
        return err('Method not allowed', 405)
    perms = get_request_perms(request)
    if perms is not None:
        if not perms['pages'].get('approval_records', True):
            return err('无访问权限', 403, 403)
        if not perms.get('can_create'):
            return err('无新增权限', 403, 403)
    f = request.FILES.get('file')
    if not f:
        return err('请上传文件')
    if getattr(f, 'size', 0) > 5 * 1024 * 1024:
        return err('文件过大，请确认文件不超过5MB')
    rows, ferr = _read_import_rows(f)
    if ferr:
        return err(ferr)
    if not rows:
        return err('文件为空')
    PRECHECK_MAX = 10000
    if len(rows) - 1 > PRECHECK_MAX:
        return ok({'skipPrecheck': True, 'total': len(rows) - 1,
                   'reason': f'数据量较大（约 {len(rows) - 1} 行，超过 {PRECHECK_MAX} 行），已跳过 AI 预检，直接导入。'})
    cv, max_row, missing = _approval_reader(rows)
    if missing:
        return err('未识别到必要表头（申请人 / 申请金额）。请使用最新模板，或确认首行表头与列内容未被删改、且未选错文件。')

    report_rows, ai_input = [], []
    for r in range(2, max_row + 1):
        view = _approval_row_view(cv, r)
        if _approval_row_is_blank(view):
            continue
        if view['applicant'].startswith('示例') or view['summary'].startswith('示例'):
            continue
        fields, error = _parse_approval_fields(view, request)
        report_rows.append({'row': r, 'data': view, 'ruleIssue': error, 'warn': None, 'ai': []})
        if not error:
            ai_input.append({'row': r, 'applicant': view['applicant'], 'department': view['department'],
                             'amount': view['amount'], 'payee': view['payee'],
                             'approval_number': view['approval_number'], 'summary': view['summary']})

    by_row = {}
    for fnd in _ai_review_approvals(ai_input):
        by_row.setdefault(fnd['row'], []).append(fnd)
    for rr in report_rows:
        rr['ai'] = by_row.get(rr['row'], [])

    attention, ok_data, rule_errors, ai_findings = [], [], 0, 0
    for rr in report_rows:
        ai_findings += len(rr['ai'])
        if rr['ruleIssue']:
            rule_errors += 1; rr['status'] = 'error'; attention.append(rr)
        elif rr['ai']:
            rr['status'] = 'review'; attention.append(rr)
        else:
            ok_data.append(rr['data'])
    return ok({
        'total': len(report_rows), 'attention': len(attention), 'okCount': len(ok_data),
        'ruleErrors': rule_errors, 'warns': 0, 'aiFindings': ai_findings,
        'aiEnabled': bool(getattr(settings, 'DEEPSEEK_API_KEY', '')),
        'columns': _APPROVAL_COLUMNS, 'rows': attention, 'okRows': ok_data,
    })


@csrf_exempt
@pk_required()
def approval_import_apply(request):
    """对预检并经用户确认/修正后的审批行执行：mode=import 直接导入；mode=download 下载修正版。"""
    if request.method != 'POST':
        return err('Method not allowed', 405)
    perms = get_request_perms(request)
    if perms is not None:
        if not perms['pages'].get('approval_records', True):
            return err('无访问权限', 403, 403)
        if not perms.get('can_create'):
            return err('无新增权限', 403, 403)
    body = parse_body(request)
    in_rows = body.get('rows') or []
    ok_rows = body.get('okRows') or []
    rows = list(in_rows) + [{'data': d} for d in ok_rows]
    if not isinstance(rows, list) or not rows:
        return err('没有可处理的数据')
    if (body.get('mode') or 'import') == 'download':
        return _approval_corrected_xlsx(rows)

    created = skipped = 0
    errors = []
    for r in rows:
        rn = r.get('row', '?')
        data = dict(r.get('data') or r)
        fields, error = _parse_approval_fields(data, request)
        if error:
            skipped += 1; errors.append(f'第{rn}行: {error}'); continue
        ApprovalRecord.objects.create(created_by_id=request.pk_uid, **fields)
        created += 1
    msg = (f'全部 {created} 条成功导入' if created and not skipped
           else f'成功导入 {created} 条，跳过 {skipped} 条' if created
           else f'没有可导入的数据（跳过 {skipped} 条）')
    return ok({'created': created, 'skipped': skipped, 'errors': errors, 'message': msg})


@pk_required()
def approval_export(request):
    return _approval_export_core(request)


def _approval_export_core(request, export_cap=5000):
    if request.method != 'GET':
        return err('Method not allowed', 405)
    perms = get_request_perms(request)
    if perms is not None and not perms['pages'].get('approval_records', True):
        return err('无访问权限', 403, 403)
    # 与列表同口径：仅隐藏「审批通过且已排满」的归档记录，已拒绝/已撤销仍可导出
    qs = dept_filter(ApprovalRecord.objects.all(), request).exclude(status='approved', archived=True).filter(deleted_at__isnull=True)
    # 全局关键字 + 列头筛选 + 排序：导出与列表口径一致
    kw = request.GET.get('q', '').strip()
    if kw:
        qs = qs.filter(
            Q(applicant__icontains=kw) | Q(department__icontains=kw) |
            Q(secondary_dept__icontains=kw) | Q(project_short_name__icontains=kw) |
            Q(approval_number__icontains=kw) | Q(g7_number__icontains=kw) |
            Q(summary__icontains=kw) | Q(payee__icontains=kw)
        )
    fq, fq_distinct = build_filter_q(request.GET.get('filters', ''), APPROVAL_FILTER_REGISTRY)
    if fq:
        qs = qs.filter(fq)
        if fq_distinct:
            qs = qs.distinct()
    _sort_by = resolve_sort(request.GET.get('sort'), request.GET.get('order'), APPROVAL_FILTER_REGISTRY)
    if _sort_by:
        qs = qs.order_by(_sort_by)
    if qs.count() > export_cap:
        return err(f'导出超过{export_cap}行，请缩小筛选范围或使用后台导出')

    # Excel 公式注入防护：以 =+-@ 等开头的文本前置单引号转义
    _formula_chars = ('=', '+', '-', '@', '\t', '\r')

    def _safe(v):
        if isinstance(v, str) and v and v[0] in _formula_chars:
            return "'" + v
        return v
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = '审批记录'
    headers = ['申请人', '所属事业部', '二级部门', '项目简称', '审批编号', 'G7编号', '摘要', '备注', '申请金额',
               '已排款金额', '剩余可排', '收款主体', '审批状态']
    header_fill = PatternFill(fill_type='solid', fgColor='C96342')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h); c.fill = header_fill; c.font = header_font; c.alignment = center
    cn = {'pending': '待审批', 'approved': '审批通过', 'rejected': '已拒绝', 'canceled': '已撤销'}
    for o in qs:
        sched = float(o.scheduled_amount or 0)
        ws.append([_safe(o.applicant), o.department, _safe(o.secondary_dept),
                   _safe(o.project_short_name), _safe(o.approval_number),
                   _safe(o.g7_number), _safe(o.summary), _safe(o.notes), float(o.amount), sched,
                   max(0.0, float(o.amount) - sched), _safe(o.payee),
                   cn.get(o.status, o.status)])
    return _build_excel_response(wb, '审批记录.xlsx')


# ── dashboard ─────────────────────────────────────────────────────────────────

@pk_required()
def dashboard(request):
    if request.method != 'GET':
        return err('Method not allowed', 405)
    perms = get_request_perms(request)
    if perms is not None and not perms['pages'].get('dashboard', True):
        return err('无访问权限', 403, 403)
    today = timezone.localdate()
    show_amount = perms is None or perms['view'].get('total_amount', True)

    base = dept_filter(Payment.objects.all(), request).annotate(paid=_paid_expr())

    def money(v):
        return str(v if v is not None else Decimal('0')) if show_amount else None

    today_qs = base.filter(planned_date=today)
    pending_qs = base.filter(paid=Decimal('0'))
    overdue_qs = base.filter(paid__lt=F('total_amount'), planned_date__lt=today)

    # Each rollup is a single count+sum aggregate (indexed, no full table load).
    today_agg = today_qs.aggregate(c=Count('id'), s=Sum('total_amount'))
    # For pending records paid=0 → remaining == total_amount.
    pending_agg = pending_qs.aggregate(c=Count('id'), s=Sum('total_amount'))
    overdue_agg = overdue_qs.aggregate(c=Count('id'), s=Sum(_remaining_expr()))
    partial_agg = base.filter(paid__gt=Decimal('0'), paid__lt=F('total_amount')).aggregate(
        c=Count('id'), s=Sum(_remaining_expr()))

    today_payments = [
        apply_view_mask(p.to_dict(), perms)
        for p in today_qs.select_related('created_by').prefetch_related('installments', 'plan_items')[:50]
    ]

    return ok({
        'today_count': today_agg['c'],
        'today_amount': money(today_agg['s']),
        'pending_count': pending_agg['c'],
        'pending_amount': money(pending_agg['s']),
        'partial_count': partial_agg['c'],
        'partial_amount': money(partial_agg['s']),
        'overdue_count': overdue_agg['c'],
        'overdue_amount': money(overdue_agg['s']),
        'today_payments': today_payments,
    })


# ── stats ─────────────────────────────────────────────────────────────────────

@pk_required()
def stats(request):
    if request.method != 'GET':
        return err('Method not allowed', 405)
    perms = get_request_perms(request)
    if perms is not None and not perms['pages'].get('stats', True):
        return err('无访问权限', 403, 403)
    if perms is not None and not perms['view'].get('total_amount', True):
        return err('无金额查看权限，无法查看统计', 403, 403)
    today = timezone.localdate()
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
        if not (1 <= month <= 12):
            month = today.month
        if not (2000 <= year <= 2100):
            year = today.year
    except ValueError:
        year, month = today.year, today.month

    # 部门作用域 + 可选部门筛选先建基集，本期与前期结转共用同一过滤口径。
    base = dept_filter(Payment.objects.all(), request)
    depts_param = request.GET.get('depts', '').strip()
    if depts_param:
        selected = [d.strip() for d in depts_param.split(',') if d.strip()]
        if selected:
            base = base.filter(department__in=selected)

    qs = base.filter(
        planned_date__year=year, planned_date__month=month
    ).annotate(paid=_paid_expr())

    # 前期结转：计划日期早于本月、且仍未付清的排款，其剩余应付额结转到本期展示。
    month_start = datetime.date(year, month, 1)
    not_settled = _not_settled_q()
    eff_remaining = _effective_remaining_expr()
    carry_qs = base.filter(planned_date__lt=month_start).annotate(_paid=_paid_expr())

    D = Decimal

    # Overall totals + status breakdown in a single aggregate query.
    # 待付口径 = 计划 − 已付 − 预付冲抵（与付款管理列表/资金池一致）
    agg = qs.aggregate(
        total=Sum('total_amount'),
        paid_sum=Sum('paid'),
        offset_sum=Sum('prepaid_offset_amount'),
        cnt=Count('id'),
        settled=Count(Case(When(
            paid__gte=F('total_amount') - F('prepaid_offset_amount'), then=1))),
        partial=Count(Case(When(
            paid__gt=Decimal('0'),
            paid__lt=F('total_amount') - F('prepaid_offset_amount'), then=1))),
        pending=Count(Case(When(paid=Decimal('0'),
                                prepaid_offset_amount=Decimal('0'), then=1))),
    )
    total_amount = agg['total'] or D(0)
    total_paid = agg['paid_sum'] or D(0)
    total_remaining = total_amount - total_paid - (agg['offset_sum'] or D(0))
    completion_rate = round(float(total_paid / total_amount * 100), 1) if total_amount else 0.0

    # 前期结转合计（仅未付清部分的剩余应付额）。
    carry_agg = carry_qs.aggregate(
        carry=Sum(eff_remaining, filter=not_settled),
        carry_count=Count('id', filter=not_settled),
    )
    carryover_remaining = carry_agg['carry'] or D(0)
    carryover_count = carry_agg['carry_count'] or 0
    # 累计应付未付 = 本期未付 + 前期结转未付。
    total_outstanding = total_remaining + carryover_remaining

    # Per-department rollup grouped in SQL（本期 + 前期结转分别分组后合并）。
    dept_map = {}
    for r in qs.values('department').annotate(
            total=Sum('total_amount'), paid_sum=Sum('paid'),
            offset_sum=Sum('prepaid_offset_amount'), count=Count('id')):
        dept_map[r['department']] = {
            'total': r['total'] or D(0), 'paid': r['paid_sum'] or D(0),
            'offset': r['offset_sum'] or D(0),
            'count': r['count'], 'carry': D(0), 'carry_count': 0,
        }
    for r in carry_qs.values('department').annotate(
            carry=Sum(eff_remaining, filter=not_settled),
            carry_count=Count('id', filter=not_settled)):
        row = dept_map.setdefault(r['department'], {
            'total': D(0), 'paid': D(0), 'offset': D(0),
            'count': 0, 'carry': D(0), 'carry_count': 0})
        row['carry'] = r['carry'] or D(0)
        row['carry_count'] = r['carry_count'] or 0

    by_dept = []
    for dept, v in dept_map.items():
        t, pd, carry = v['total'], v['paid'], v['carry']
        by_dept.append({
            'dept': dept,
            'total': str(t),
            'paid': str(pd),
            'remaining': str(t - pd - v.get('offset', D(0))),
            'count': v['count'],
            'carry': str(carry),               # 前期结转未付
            'carry_count': v['carry_count'],
            'outstanding': str((t - pd) + carry),  # 本期未付 + 前期结转
            'completion_rate': round(float(pd / t * 100), 1) if t else 0.0,
        })
    # 按「应付合计（本期计划 + 前期结转）」降序，结转大的部门优先呈现。
    by_dept.sort(key=lambda x: Decimal(x['total']) + Decimal(x['carry']), reverse=True)

    return ok({
        'year': year, 'month': month,
        'total_amount': str(total_amount),
        'total_paid': str(total_paid),
        'total_remaining': str(total_remaining),
        'carryover_remaining': str(carryover_remaining),
        'carryover_count': carryover_count,
        'total_outstanding': str(total_outstanding),
        'completion_rate': completion_rate,
        'total_records': agg['cnt'],
        'by_dept': by_dept,
        'by_status': {
            'settled': agg['settled'],
            'partial': agg['partial'],
            'pending': agg['pending'],
        },
    })


# ── users ─────────────────────────────────────────────────────────────────────

@csrf_exempt
@pk_required(roles=['super_admin'])
def users(request):
    if request.method == 'GET':
        try:
            all_users = PaikuanUser.objects.order_by('is_approved', 'id')
            return ok([u.to_dict() for u in all_users])
        except Exception as e:
            logger.error('users list failed: %s', e)
            return err('用户列表加载失败，请确认已执行 python manage.py migrate', 500)
    # Super admin cannot create users directly — users register themselves
    return err('请通过注册流程创建用户', 405)


@csrf_exempt
@pk_required(roles=['super_admin'])
def user_detail(request, pk):
    try:
        user = PaikuanUser.objects.get(id=pk)
    except PaikuanUser.DoesNotExist:
        return err('用户不存在', 404)

    if request.method == 'PUT':
        data = parse_body(request)
        if 'name' in data:
            user.name = (data['name'] or '').strip()
        if 'role' in data:
            if data['role'] not in ('super_admin', 'manager', 'operator', 'viewer'):
                return err('角色无效')
            # Prevent creating a second super_admin
            if data['role'] == 'super_admin' and user.role != 'super_admin':
                return err('超级管理员只能有一位')
            # Prevent downgrading the only super_admin
            if user.role == 'super_admin' and data['role'] != 'super_admin':
                return err('不能修改超级管理员的角色')
            user.role = data['role']
        if 'departments' in data:
            depts = data['departments'] or []
            # Approved non-admin users must retain at least one department;
            # otherwise they'd become invisible to the system (no rows to see or edit).
            if user.is_approved and user.role != 'super_admin' and len(depts) == 0:
                return err('已审批用户至少需要分配一个部门')
            user.departments = depts
        if 'job_title' in data:
            user.job_title = data['job_title'] or ''
        if 'is_active' in data:
            user.is_active = bool(data['is_active'])
        if data.get('password'):
            pwd_err = _validate_password_strength(data['password'], phone=user.phone, name=user.name)
            if pwd_err:
                return err('新' + pwd_err if pwd_err.startswith('密码') else pwd_err)
            user.set_password(data['password'])
            user.pwd_changed_at = timezone.now()
            # 超管重置 = 临时密码：用户下次登录须先自助改密
            user.must_change_password = True
        user.save()
        return ok(user.to_dict())

    if request.method == 'DELETE':
        if user.id == request.pk_uid:
            return err('不能删除自己的账号')
        if user.role == 'super_admin':
            return err('不能删除超级管理员账号')
        user.delete()
        return ok({'deleted': pk})

    return err('Method not allowed', 405)


@csrf_exempt
@pk_required(roles=['super_admin'])
def approve_user(request, pk):
    if request.method != 'POST':
        return err('Method not allowed', 405)
    try:
        user = PaikuanUser.objects.get(id=pk)
    except PaikuanUser.DoesNotExist:
        return err('用户不存在', 404)
    except Exception as e:
        logger.error('approve_user get failed: %s', e)
        return err('操作失败，请先执行 python manage.py migrate', 500)
    if user.role == 'super_admin':
        return err('超级管理员无需审批')
    body = parse_body(request)
    # Super admin may adjust job title / departments at approval time.
    if body.get('job_title') in JOB_TITLES:
        user.job_title = body['job_title']
    if isinstance(body.get('departments'), list):
        user.departments = body['departments']
    user.is_approved = True
    user.is_active = True
    user.role = 'operator'   # regular member; field perms come from job_title
    try:
        user.approved_by_id = request.pk_uid
        user.approved_at = timezone.now()
        user.save()
    except Exception as e:
        logger.error('approve_user save failed: %s', e)
        return err('审批保存失败，请检查数据库迁移是否完整', 500)
    return ok(user.to_dict())


@csrf_exempt
@pk_required(roles=['super_admin'])
def reject_user(request, pk):
    """Reject a pending registration by deleting the account."""
    if request.method not in ('POST', 'DELETE'):
        return err('Method not allowed', 405)
    try:
        user = PaikuanUser.objects.get(id=pk)
    except PaikuanUser.DoesNotExist:
        return err('用户不存在', 404)
    except Exception as e:
        logger.error('reject_user get failed: %s', e)
        return err('操作失败，请先执行 python manage.py migrate', 500)
    if user.role == 'super_admin':
        return err('不能拒绝超级管理员')
    if user.is_approved:
        return err('该用户已审批，不能拒绝；如需禁用请使用停用')
    try:
        user.delete()
    except Exception as e:
        logger.error('reject_user delete failed: %s', e)
        return err('删除失败', 500)
    return ok({'rejected': pk})


# ── permissions ───────────────────────────────────────────────────────────────

@csrf_exempt
@pk_required(roles=['super_admin'])
def permissions(request):
    """GET: full permission matrix for all job titles + field metadata."""
    if request.method != 'GET':
        return err('Method not allowed', 405)
    jobs = [
        {'job_title': key, 'label': label, 'config': get_job_perms(key)}
        for key, label in JOB_TITLES.items()
    ]
    return ok({
        'fields': PAYMENT_FIELD_DEFS,
        'ar_fields': AR_FIELD_DEFS,
        'caiwu_fields': CAIWU_FIELD_DEFS,
        'action_defs': ACTION_DEFS,
        'pages': [
            {'key': 'dashboard',         'label': '今日工作台'},
            {'key': 'payments',          'label': '付款管理'},
            {'key': 'approval_records',  'label': '审批记录'},
            {'key': 'stats',             'label': '月度统计'},
            {'key': 'ar_projects',       'label': '项目台账'},
            {'key': 'ar_records',        'label': '应收明细'},
            {'key': 'ar_advance',        'label': '预收预付'},
            {'key': 'ar_analytics',      'label': '应收分析'},
            {'key': 'ar_cashflow',       'label': '现金流分析'},
            {'key': 'ar_budget',         'label': '预算管理'},
            {'key': 'caiwu_report',      'label': '财务分析·报表'},
            {'key': 'caiwu_data',        'label': '财务分析·数据加工'},
            {'key': 'caiwu_charts',      'label': '财务分析·图表'},
            {'key': 'caiwu_metrics',     'label': '财务分析·指标管理'},
            {'key': 'caiwu_cockpit',     'label': '财务分析·驾驶舱'},
        ],
        'jobs': jobs,
    })


@csrf_exempt
@pk_required(roles=['super_admin'])
def permission_detail(request, job):
    """PUT: update one job title's permission config."""
    if request.method != 'PUT':
        return err('Method not allowed', 405)
    if job not in JOB_TITLES:
        return err('职务无效', 404)
    body = parse_body(request)
    cfg = body.get('config') or {}
    _base = default_job_config(job)   # 用于补齐前端未显式传入的能力位默认值

    # Sanitize against known keys to avoid storing junk.
    clean = {
        'pages': {k: bool(cfg.get('pages', {}).get(k, True)) for k in PAGE_KEYS},
        'view': {k: bool(cfg.get('view', {}).get(k, True)) for k in FIELD_KEYS},
        'edit': {k: bool(cfg.get('edit', {}).get(k, False)) for k in FIELD_KEYS},
        'ar_view': {k: bool(cfg.get('ar_view', {}).get(k, True)) for k in AR_FIELD_KEYS},
        'can_create': bool(cfg.get('can_create', False)),
        'can_delete': bool(cfg.get('can_delete', False)),
        # 应收专属写权限：前端未传时回退到该职务的基线默认（结算会计=True），
        # 避免老前端保存时把该能力清掉；前端传了则以传入为准（可自由搭配）。
        'ar_can_create': bool(cfg.get('ar_can_create', _base.get('ar_can_create', False))),
        'ar_shared_only': bool(cfg.get('ar_shared_only', False)),
        # 操作级权限：前端未传某键时回退该职务基线默认，避免老前端保存清掉能力
        'actions': {k: bool(cfg.get('actions', {}).get(
            k, _base.get('actions', {}).get(k, False))) for k in ACTION_KEYS},
        'caiwu_view': {k: bool(cfg.get('caiwu_view', {}).get(k, True)) for k in CAIWU_FIELD_KEYS},
        'caiwu_upload':  bool(cfg.get('caiwu_upload', False)),
        'caiwu_publish': bool(cfg.get('caiwu_publish', False)),
        'caiwu_delete':  bool(cfg.get('caiwu_delete', False)),
    }
    obj, _ = JobPermission.objects.get_or_create(job_title=job)
    obj.config = clean
    obj.save()
    _invalidate_perm_cache(job)
    # caiwu reads its 财务分析 perms straight from this JobPermission via paikuan's
    # cached get_job_perms (it keeps no cache of its own), so clearing this cache
    # is enough for the change to take effect everywhere immediately.
    return ok({'job_title': job, 'config': get_job_perms(job)})


# ── Excel template / import / export ─────────────────────────────────────────

def _visible_excel_cols(perms, n_installments=TEMPLATE_INSTALLMENT_SLOTS):
    """Return [(excel_header, db_col)] for columns the user can view.
    Installment columns are generated dynamically for n_installments slots.
    db_col for installment date/amount uses the slot index as 'inst_date_{n}' / 'inst_amount_{n}'.
    """
    can_view_insts = perms is None or perms['view'].get('installments', True)
    cols = []
    for (fk, h, c) in EXCEL_COLUMN_MAP:
        if perms is None or perms['view'].get(fk, True):
            cols.append((h, c))
        # Insert installment columns right after planned_date
        if fk == 'planned_date' and can_view_insts:
            for i in range(1, n_installments + 1):
                cols.append((f'第{i}次付款日期', f'inst_date_{i}'))
                cols.append((f'第{i}次付款金额(元)', f'inst_amount_{i}'))
    return cols


def _build_excel_response(wb, filename, formula_guard=True):
    # 出口 chokepoint 统一公式注入防护：覆盖所有导出（含导入修正版下载等用户可控文本），
    # 新增导出点无需各自记得调用。运输零误差导出传 formula_guard=False（其自带更窄的
    # 逐格防护，且需保留原值还原契约）。
    if formula_guard:
        from wxcloudrun.excel_safe import sanitize_workbook
        sanitize_workbook(wb)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    # RFC 5987 encoding so Chinese filenames render correctly in all browsers.
    ascii_fallback = filename.encode('ascii', 'ignore').decode() or 'export.xlsx'
    encoded = quote(filename.encode('utf-8'), safe='')
    response['Content-Disposition'] = (
        f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{encoded}"
    )
    return response


@csrf_exempt
@pk_required()
def payment_template(request):
    """GET — download blank Excel import template (columns filtered by view perms)."""
    if request.method != 'GET':
        return err('Method not allowed', 405)
    perms = get_request_perms(request)
    denied = _payments_page_denied(request, perms)
    if denied:
        return denied
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return err('服务器缺少 openpyxl 依赖', 500)

    cols = _visible_excel_cols(perms)

    wb = Workbook()
    ws = wb.active
    ws.title = '排款导入模板'

    header_fill = PatternFill(fill_type='solid', fgColor='C96342')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    example_font = Font(italic=True, color='888888', size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Row 1: headers
    for col_idx, (header, _) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Row 2: example data (grayed out). The 部门 cell carries EXAMPLE_ROW_MARKER
    # so import skips this row even if the user forgets to delete it.
    example = {
        '部门': EXAMPLE_ROW_MARKER,
        '二级部门': '选填，如：华东项目部',
        '项目简称': '选填；须与项目台账的"项目简称"完全一致，搜索不到将拒绝导入',
        '申请人': '如：张三',
        '审批单号': '123456789012345678901',
        'G7编号': '',
        '付款事项': '如：工程款结算',
        '收款方': '如：某某公司',
        '计划总金额(元)': 100000,
        '计划付款日期': '2026-06-01',
        '第1次付款日期': '',
        '第1次付款金额(元)': 0,
        '备注': '本行为格式示例，导入前可删除',
        '计划调整金额(元)': '',
    }
    for col_idx, (header, _) in enumerate(cols, 1):
        cell = ws.cell(row=2, column=col_idx, value=example.get(header, ''))
        cell.font = example_font
        cell.alignment = Alignment(horizontal='left', vertical='center')

    # Column widths
    for col_idx, (header, _) in enumerate(cols, 1):
        if '日期' in header:
            w = 14
        elif '金额' in header:
            w = 16
        elif header in ('部门',):
            w = 14
        elif header in ('审批单号',):
            w = 16
        elif header in ('付款事项',):
            w = 28
        elif header in ('收款方',):
            w = 22
        elif header in ('备注',):
            w = 24
        elif header in ('申请人',):
            w = 12
        else:
            w = 14
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w

    ws.row_dimensions[1].height = 22

    return _build_excel_response(wb, 'paikuan_import_template.xlsx')


def _payment_header_map(rows):
    """解析表头行 → (col_pos, inst_date_cols, inst_amt_cols, missing_required_set)。
    付款导入 / 预检 共用，确保两者口径完全一致。"""
    header_row = [str(v).strip() if v is not None else '' for v in rows[0]]
    col_pos = {}
    for i, h in enumerate(header_row):
        if h in _EXCEL_HEADER_TO_COL:
            col_pos[_EXCEL_HEADER_TO_COL[h]] = i
    import re as _re
    inst_date_cols, inst_amt_cols = {}, {}
    for i, h in enumerate(header_row):
        m = _re.fullmatch(r'第(\d+)次付款日期', h)
        if m:
            inst_date_cols[int(m.group(1))] = i
        m2 = _re.fullmatch(r'第(\d+)次付款金额\(元\)', h)
        if m2:
            inst_amt_cols[int(m2.group(1))] = i
    missing = {'department', 'project_desc', 'payee', 'planned_date'} - set(col_pos.keys())
    return col_pos, inst_date_cols, inst_amt_cols, missing


def _payment_row_to_data(row, col_pos, inst_date_cols, inst_amt_cols):
    """单行单元格 → data dict（含 installments），与导入逻辑一致；不落库。"""
    def cell_val(col_name):
        idx = col_pos.get(col_name)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        if hasattr(v, 'strftime'):
            return v.strftime('%Y-%m-%d')
        if col_name in _EXCEL_DATE_COLS and isinstance(v, str) and v.strip():
            return _normalize_date(v)
        return v

    def cell_at(idx):
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return v.strftime('%Y-%m-%d') if hasattr(v, 'strftime') else v

    data = {col: cell_val(col) for col in col_pos}
    installments_raw = []
    for n in sorted(set(inst_date_cols) | set(inst_amt_cols)):
        raw_date = cell_at(inst_date_cols.get(n))
        raw_amt = cell_at(inst_amt_cols.get(n))
        if raw_date and isinstance(raw_date, str):
            raw_date = _normalize_date(str(raw_date))
        amt_str = str(raw_amt or '0').replace(',', '').replace('，', '').replace('¥', '').strip()
        try:
            amt = Decimal(amt_str or '0')
        except InvalidOperation:
            amt = Decimal('0')
        if raw_date and amt > Decimal('0'):
            installments_raw.append({'seq': n, 'pay_date': str(raw_date),
                                     'pay_amount': str(amt), 'notes': ''})
    data['installments'] = installments_raw
    return data


@csrf_exempt
@pk_required()
def payment_import(request):
    """POST multipart/form-data with field 'file' — bulk import payments from Excel."""
    if request.method != 'POST':
        return err('Method not allowed', 405)
    perms = get_request_perms(request)
    denied = _payments_page_denied(request, perms)
    if denied:
        return denied
    if perms is not None and not perms['can_create']:
        return err('无新增权限', 403, 403)

    upload = request.FILES.get('file')
    if not upload:
        return err('请上传文件（.xlsx / .csv）')

    MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB
    if upload.size > MAX_UPLOAD_BYTES:
        return err('文件过大，请确认文件不超过5MB')

    rows, ferr = _read_import_rows(upload)
    if ferr:
        return err(ferr)
    if not rows:
        return err('文件为空')

    # Map header position → db column name (base fields) + 动态分期列
    col_pos, inst_date_cols, inst_amt_cols, missing = _payment_header_map(rows)
    if missing:
        missing_labels = [h for _, h, c in EXCEL_COLUMN_MAP if c in missing]
        return err(f'模板缺少必要列：{", ".join(missing_labels)}，请重新下载模板')

    results = {'created': 0, 'skipped': 0, 'errors': []}

    for row_num, row in enumerate(rows[1:], start=2):
        if all(v is None or v == '' for v in row):
            continue

        dept_raw = row[col_pos['department']] if col_pos.get('department') is not None \
            and col_pos['department'] < len(row) else None
        if dept_raw and str(dept_raw).strip().startswith('示例'):
            continue

        data = _payment_row_to_data(row, col_pos, inst_date_cols, inst_amt_cols)

        fields, error = _parse_payment_fields(data)
        if error:
            results['errors'].append(f'第{row_num}行: {error}')
            results['skipped'] += 1
            continue

        if not can_write_dept(request, fields['department']):
            results['errors'].append(f'第{row_num}行: 无权操作部门"{fields["department"]}"')
            results['skipped'] += 1
            continue

        parsed_insts = fields.pop('installments') or []
        dup = _find_duplicate_payment(fields)
        if dup:
            results['errors'].append(
                f'第{row_num}行: 跳过重复（已存在排款 #{dup.id}，相同审批单号+收款方+计划日期+金额）')
            results['skipped'] += 1
            continue
        try:
            with transaction.atomic():
                p = Payment.objects.create(
                    created_by_id=request.pk_uid, updated_by_id=request.pk_uid, **fields)
                _ensure_plan_item(p)
                _save_installments(p, parsed_insts)
                _record_payment_changes(p, {}, {}, request, action='create')
            results['created'] += 1
        except Exception as e:
            results['errors'].append(f'第{row_num}行: 保存失败 ({e})')
            results['skipped'] += 1

    if results['created'] > 0 and results['skipped'] == 0:
        results['message'] = f'全部 {results["created"]} 条数据成功导入'
    elif results['created'] > 0:
        results['message'] = f'成功导入 {results["created"]} 条，跳过 {results["skipped"]} 条（含错误）'
    else:
        results['message'] = f'没有可导入的数据（跳过 {results["skipped"]} 条）'
    return ok(results)


_AI_REVIEW_PAYMENT_SYS = (
    '你是企业付款管理的数据质检助手。下面是一批待导入的付款记录（已通过基础格式校验）。'
    '请只挑出"疑似有问题"的行，找规则难以发现的软问题：收款方像乱码/测试数据/占位符；'
    '金额明显异常（疑似多或少一个0、过大或过小）；部门或收款方写法不一致（同一对象多种写法）；'
    '计划付款日期不合常理（过去太久或未来太远）；同一审批单号下收款方/金额相互矛盾；疑似重复行。'
    '严格只返回 JSON 数组，每个元素形如 '
    '{"row":行号,"field":"字段名","issue":"问题简述","suggestion":"修正建议(可空)","severity":"high|medium|low"}。'
    '没有发现问题就返回 []。不要输出 JSON 以外的任何文字。'
)


def _ai_review_payments(records):
    """用 DeepSeek 复核付款导入数据，挑出规则覆盖不到的"软问题"（疑似乱码/测试数据、
    金额异常、部门/收款方写法不一致、日期不合常理、同审批号矛盾、疑似重复等）。"""
    return _ai_review_records(records, _AI_REVIEW_PAYMENT_SYS)


def _payment_clean_view(data):
    """把解析后的 data 收敛为前端可编辑/回写的清洗字段。"""
    return {
        'department': data.get('department') or '',
        'project_short_name': data.get('project_short_name') or '',
        'project_desc': data.get('project_desc') or '',
        'payee': data.get('payee') or '',
        'total_amount': str(data.get('total_amount') or ''),
        'planned_date': str(data.get('planned_date') or ''),
        'approval_number': data.get('approval_number') or '',
        'g7_number': data.get('g7_number') or '',
        'notes': data.get('notes') or '',
        'installments': data.get('installments') or [],
    }


@csrf_exempt
@pk_required()
def payment_import_precheck(request):
    """导入预检：①规则预检（复用 _parse_payment_fields，与真实导入同口径）②AI 复核疑点。
    只读不落库；返回逐行报告供前端展示、编辑、确认。"""
    if request.method != 'POST':
        return err('Method not allowed', 405)
    perms = get_request_perms(request)
    denied = _payments_page_denied(request, perms)
    if denied:
        return denied
    if perms is not None and not perms['can_create']:
        return err('无新增权限', 403, 403)
    upload = request.FILES.get('file')
    if not upload:
        return err('请上传文件（.xlsx / .csv）')
    if upload.size > 5 * 1024 * 1024:
        return err('文件过大，请确认文件不超过5MB')
    rows, ferr = _read_import_rows(upload)
    if ferr:
        return err(ferr)
    if not rows:
        return err('文件为空')
    # 行数上限：超大文件不做逐行预检（逐行 DB 校验 + AI 复核 + 前端渲染成本高）。
    # 返回 skipPrecheck 让前端跳过 AI 介入、直接走常规导入（而非报错挡住用户）。
    PRECHECK_MAX = 10000
    if len(rows) - 1 > PRECHECK_MAX:
        return ok({'skipPrecheck': True, 'total': len(rows) - 1,
                   'reason': f'数据量较大（约 {len(rows) - 1} 行，超过 {PRECHECK_MAX} 行），'
                             f'已跳过 AI 预检，直接导入。'})
    col_pos, inst_date_cols, inst_amt_cols, missing = _payment_header_map(rows)
    if missing:
        missing_labels = [h for _, h, c in EXCEL_COLUMN_MAP if c in missing]
        return err(f'模板缺少必要列：{", ".join(missing_labels)}，请重新下载模板')

    report_rows, ai_input = [], []
    for row_num, row in enumerate(rows[1:], start=2):
        if all(v is None or v == '' for v in row):
            continue
        di = col_pos.get('department')
        dept_raw = row[di] if di is not None and di < len(row) else None
        if dept_raw and str(dept_raw).strip().startswith('示例'):
            continue
        data = _payment_row_to_data(row, col_pos, inst_date_cols, inst_amt_cols)
        fields, error = _parse_payment_fields(data)
        rule_issue, warn = error, None
        if not error:
            if not can_write_dept(request, fields['department']):
                rule_issue = f'无权操作部门"{fields["department"]}"'
            else:
                f2 = dict(fields); f2.pop('installments', None)
                dup = _find_duplicate_payment(f2)
                if dup:
                    warn = f'疑似重复（已存在排款 #{dup.id}）'
        clean = _payment_clean_view(data)
        report_rows.append({'row': row_num, 'data': clean,
                            'ruleIssue': rule_issue, 'warn': warn, 'ai': []})
        if not rule_issue:
            ai_input.append({'row': row_num, 'department': clean['department'],
                             'project_desc': clean['project_desc'], 'payee': clean['payee'],
                             'total_amount': clean['total_amount'],
                             'planned_date': clean['planned_date'],
                             'approval_number': clean['approval_number']})

    by_row = {}
    for fnd in _ai_review_payments(ai_input):
        by_row.setdefault(fnd['row'], []).append(fnd)
    for rr in report_rows:
        rr['ai'] = by_row.get(rr['row'], [])

    # 只把"需关注"的行（规则错误 / 疑似重复 / AI 疑点）返回前端展示编辑；其余已通过的行
    # 仅回传数据、不渲染，确认时一并导入 —— 这样 1 万行也不会卡（前端只渲染少量问题行）。
    columns = [
        {'key': 'department', 'label': '部门'},
        {'key': 'payee', 'label': '收款方'},
        {'key': 'total_amount', 'label': '计划金额'},
        {'key': 'planned_date', 'label': '计划日期'},
        {'key': 'approval_number', 'label': '审批单号'},
        {'key': 'project_short_name', 'label': '项目简称'},
        {'key': 'project_desc', 'label': '付款事项'},
    ]
    attention, ok_data, rule_errors, warns, ai_findings = [], [], 0, 0, 0
    for r in report_rows:
        ai_findings += len(r['ai'])
        if r['ruleIssue']:
            rule_errors += 1
            r['status'] = 'error'
            attention.append(r)
        elif r['warn']:
            warns += 1
            r['status'] = 'warn'
            attention.append(r)
        elif r['ai']:
            r['status'] = 'review'
            attention.append(r)
        else:
            ok_data.append(r['data'])
    return ok({
        'total': len(report_rows),
        'attention': len(attention),
        'okCount': len(ok_data),
        'ruleErrors': rule_errors,
        'warns': warns,
        'aiFindings': ai_findings,
        'aiEnabled': bool(getattr(settings, 'DEEPSEEK_API_KEY', '')),
        'columns': columns,
        'rows': attention,   # 仅"需关注"的行
        'okRows': ok_data,   # 已通过、不展示、确认时一并导入
    })


def _payment_corrected_xlsx(rows):
    """把修正后的行写成标准模板格式的 .xlsx 返回（含动态分期列），可直接再次导入。"""
    from openpyxl import Workbook
    base = [(h, c) for _, h, c in EXCEL_COLUMN_MAP]
    max_inst = max((len(r.get('data', r).get('installments') or []) for r in rows), default=0)
    wb = Workbook(); ws = wb.active; ws.title = '付款管理(修正版)'
    headers = [h for h, _ in base]
    for n in range(1, max_inst + 1):
        headers += [f'第{n}次付款日期', f'第{n}次付款金额(元)']
    ws.append(headers)
    for r in rows:
        d = r.get('data', r)
        line = [d.get(c, '') if d.get(c) is not None else '' for _, c in base]
        insts = d.get('installments') or []
        for n in range(max_inst):
            it = insts[n] if n < len(insts) else {}
            line += [it.get('pay_date', ''), it.get('pay_amount', '')]
        ws.append(line)
    return _build_excel_response(wb, '付款管理_修正版.xlsx')


@csrf_exempt
@pk_required()
def payment_import_apply(request):
    """对预检并经用户确认/修正后的行执行：mode=import 直接导入；mode=download 下载修正版文件。"""
    if request.method != 'POST':
        return err('Method not allowed', 405)
    perms = get_request_perms(request)
    denied = _payments_page_denied(request, perms)
    if denied:
        return denied
    if perms is not None and not perms['can_create']:
        return err('无新增权限', 403, 403)
    body = parse_body(request)
    in_rows = body.get('rows') or []
    ok_rows = body.get('okRows') or []
    # 需关注行（可能已被用户就地修正）+ 已通过行（原样回传）一并处理
    rows = list(in_rows) + [{'data': d} for d in ok_rows]
    if not isinstance(rows, list) or not rows:
        return err('没有可处理的数据')
    if (body.get('mode') or 'import') == 'download':
        return _payment_corrected_xlsx(rows)

    results = {'created': 0, 'skipped': 0, 'errors': []}
    for r in rows:
        rn = r.get('row', '?')
        data = dict(r.get('data') or r)
        fields, error = _parse_payment_fields(data)
        if error:
            results['errors'].append(f'第{rn}行: {error}'); results['skipped'] += 1; continue
        if not can_write_dept(request, fields['department']):
            results['errors'].append(f'第{rn}行: 无权操作部门"{fields["department"]}"')
            results['skipped'] += 1; continue
        parsed_insts = fields.pop('installments') or []
        dup = _find_duplicate_payment(fields)
        if dup:
            results['errors'].append(f'第{rn}行: 跳过重复（已存在排款 #{dup.id}）')
            results['skipped'] += 1; continue
        try:
            with transaction.atomic():
                p = Payment.objects.create(
                    created_by_id=request.pk_uid, updated_by_id=request.pk_uid, **fields)
                _ensure_plan_item(p); _save_installments(p, parsed_insts)
                _record_payment_changes(p, {}, {}, request, action='create')
            results['created'] += 1
        except Exception as e:
            results['errors'].append(f'第{rn}行: 保存失败 ({e})'); results['skipped'] += 1
    if results['created'] and not results['skipped']:
        results['message'] = f'全部 {results["created"]} 条成功导入'
    elif results['created']:
        results['message'] = f'成功导入 {results["created"]} 条，跳过 {results["skipped"]} 条'
    else:
        results['message'] = f'没有可导入的数据（跳过 {results["skipped"]} 条）'
    return ok(results)


@csrf_exempt
@pk_required()
def payment_export(request):
    """GET — export filtered payment list to Excel (same filters as list view)."""
    return _payment_export_core(request)


def _payment_export_core(request, export_cap=5000):
    if request.method != 'GET':
        return err('Method not allowed', 405)
    perms = get_request_perms(request)
    denied = _payments_page_denied(request, perms)
    if denied:
        return denied
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return err('服务器缺少 openpyxl 依赖', 500)

    qs = Payment.objects.select_related('created_by').all()
    qs = dept_filter(qs, request)
    qs = qs.prefetch_related('installments', 'plan_items')

    dept = request.GET.get('dept', '').strip()
    status_q = request.GET.get('status', '').strip()
    hide_settled = request.GET.get('hide_settled') == '1'
    start = request.GET.get('start_date', '').strip()
    end = request.GET.get('end_date', '').strip()
    q_str = request.GET.get('q', '').strip()

    if dept:
        qs = qs.filter(department=dept)
    if start:
        qs = qs.filter(planned_date__gte=start)
    if end:
        qs = qs.filter(planned_date__lte=end)
    if q_str:
        qs = qs.filter(
            Q(project_desc__icontains=q_str) | Q(payee__icontains=q_str) |
            Q(approval_number__icontains=q_str) | Q(g7_number__icontains=q_str) |
            Q(department__icontains=q_str) | Q(applicant__icontains=q_str)
        )
    # 付款日期窗口：导出须与列表视图一致（此前漏掉该筛选，导出会忽略付款日期范围）。
    pay_start = request.GET.get('pay_date_start', '').strip()
    pay_end = request.GET.get('pay_date_end', '').strip()
    if pay_start:
        qs = qs.filter(installments__pay_date__gte=pay_start).distinct()
    if pay_end:
        qs = qs.filter(installments__pay_date__lte=pay_end).distinct()
    # 列头精确筛选 + 排序：导出与列表口径一致（筛了再导出）
    fq, fq_distinct = build_filter_q(request.GET.get('filters', ''), PAYMENT_FILTER_REGISTRY)
    if fq:
        qs = qs.filter(fq)
        if fq_distinct:
            qs = qs.distinct()
    _sort_by = resolve_sort(request.GET.get('sort'), request.GET.get('order'), PAYMENT_FILTER_REGISTRY)
    if _sort_by:
        qs = qs.order_by(_sort_by)
    # 计算列（已付/剩余/逾期）筛选与排序：导出与列表口径一致
    qs, _paid_annotated = _apply_payment_computed_filters(qs, request)
    qs = _apply_payment_status_filter(qs, status_q, _paid_annotated, hide_settled=hide_settled)

    # Reject rather than silently truncate: a truncated export is worse than no export.
    total_count = qs.count()
    if total_count > export_cap:
        return err(
            f'当前筛选结果共 {total_count} 条，超出导出上限（{export_cap} 条）。'
            '请缩小日期范围或添加其他筛选条件后重试，或使用后台导出。'
        )

    # Determine max installment count across the filtered set for dynamic columns.
    can_view_insts = perms is None or perms['view'].get('installments', True)
    can_view_amounts = perms is None or perms['view'].get('total_amount', True)
    all_payments = list(qs.prefetch_related('installments', 'plan_items').select_related('created_by'))
    if can_view_insts and all_payments:
        max_slots = max((p.installments.count() for p in all_payments), default=0)
    else:
        max_slots = 0
    cols = _visible_excel_cols(perms, n_installments=max_slots)

    wb = Workbook()
    ws = wb.active
    ws.title = '排款记录'

    header_fill = PatternFill(fill_type='solid', fgColor='C96342')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center = Alignment(horizontal='center', vertical='center')

    for col_idx, (header, _) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    status_label = {'pending': '⏳待付款', 'partial': '⚡部分付款', 'settled': '✅已付清', 'adjusted': '📋计划调整'}

    _FORMULA_CHARS = ('=', '+', '-', '@', '\t', '\r')

    def _safe_text(v):
        if isinstance(v, str) and v and v[0] in _FORMULA_CHARS:
            return "'" + v
        return v

    for row_idx, p in enumerate(all_payments, start=2):
        d = apply_view_mask(p.to_dict(), perms)
        # Build installment slot map: seq → {date, amount}
        insts = d.get('installments') or []
        inst_by_seq = {inst['seq']: inst for inst in insts}
        for col_idx, (_, db_col) in enumerate(cols, 1):
            if db_col.startswith('inst_date_'):
                n = int(db_col.split('_')[-1])
                val = inst_by_seq[n]['pay_date'] if n in inst_by_seq else ''
            elif db_col.startswith('inst_amount_'):
                n = int(db_col.split('_')[-1])
                val = float(inst_by_seq[n]['pay_amount']) if n in inst_by_seq else 0.0
            else:
                val = d.get(db_col)
                if val is None:
                    val = ''
                if db_col in ('total_amount', 'total_paid', 'remaining', 'plan_adjustment'):
                    try:
                        val = float(val) if val != '' else 0.0
                    except (ValueError, TypeError):
                        pass
                else:
                    val = _safe_text(val)
            ws.cell(row=row_idx, column=col_idx, value=val)

        # Append derived columns after visible cols（导出专用，不参与导入）：
        # 预付冲抵 / 剩余待付 / 状态。冲抵为核销派生值，导入它会破坏两侧台账平衡。
        extra_idx = len(cols) + 1
        if can_view_amounts:
            try:
                ws.cell(row=row_idx, column=extra_idx,
                        value=float(d.get('prepaid_offset_amount') or 0))
                ws.cell(row=row_idx, column=extra_idx + 1,
                        value=float(d.get('remaining') or 0))
            except (ValueError, TypeError):
                pass
            plan_items = d.get('plan_items') or []
            ws.cell(row=row_idx, column=extra_idx + 2,
                    value='；'.join(f"{pi['planned_date']}:{pi['amount']}" for pi in plan_items)
                    if len(plan_items) > 1 else '')
            extra_idx += 3
        ws.cell(row=row_idx, column=extra_idx,
                value=status_label.get(d.get('status', ''), ''))

    # Derived headers
    extra_headers = (['预付冲抵(元)', '剩余待付(元)', '计划明细(分批)'] if can_view_amounts else []) + ['状态']
    for off, h in enumerate(extra_headers):
        c = ws.cell(row=1, column=len(cols) + 1 + off, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    # Auto-width approximation
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 32)

    today = timezone.localdate().strftime('%Y%m%d')
    return _build_excel_response(wb, f'排款记录_{today}.xlsx')


# ── 运输事业部对账单 专用导入 / 导出 ────────────────────────────────────────────
# 运输事业部从自有系统导出的对账单结构与排款表完全不同（14 固定列、金额为负）。
# 导入（审批管理）：逐行转为「已通过」审批记录（金额取绝对值、映射部门/收款方），原始整行
#       存 ext_raw，以「对账单号」为唯一键去重。之后由用户在审批管理手动排款进入付款管理。
# 导出（付款管理）：按勾选 / 筛选取回已排款的运输付款记录，从来源审批的 ext_raw 原样还原原表，
#       逐字零误差，仅把「状态」列改写为「已结算」（限系统内已付清行），供回传运输自有系统。
TRANSPORT_SOURCE = 'transport'
TRANSPORT_DEPT = '运输事业部'
# 运输对账单原表头（顺序即导出列顺序，必须与来源系统完全一致）
TRANSPORT_HEADERS = [
    '序号', '所属组织', '对账单号', '运单号', '项目名称', '收支方式', '对账对象', '联系电话',
    '实际对账金额', '对账时间', '状态', '创建人', '创建时间', '备注', '单据类别',
    '账单调整', '对账金额', '结算方式', '实对网货服务费合计', '实对网货费用合计', '开户人',
]
TRANSPORT_KEY_COL = '对账单号'      # 去重唯一键
TRANSPORT_AMOUNT_COL = '实际对账金额'  # 取绝对值作为申请金额（已含账单调整的最终额）
TRANSPORT_STATUS_COL = '状态'       # 导出时改写为「已结算」的列
TRANSPORT_SEQ_COL = '序号'          # 导出时按输出行重排为 1,2,3…（不沿用原表零散序号）
TRANSPORT_SETTLED_LABEL = '已结算'
# 汇总/合计行标记：运输原表末尾常有一行「合计」（序号=合计、对账单号=「681 条」、
# 金额=全表总额）。它不是真实对账单，若不识别会被当成一条巨额付款导入。
TRANSPORT_SUMMARY_MARKERS = ('合计', '总计', '小计', '合 计')
# 运输对账单正常应为「已通过」（对账通过、待我方付款）。下列源状态不可导入：
#   已结算/已完成类 → 源系统已结算，再导入排款会重复付款；
#   作废/取消/驳回类 → 单据已失效，不应付款。
# 精确匹配（规范化后），避免误伤「未结算/待结算」（含「结算」二字但应可导入）。
TRANSPORT_BLOCK_SETTLED = {'已结算', '已完成', '已支付', '已付款', '结算完成', '付款完成'}
TRANSPORT_BLOCK_VOID = {'已作废', '作废', '已取消', '取消', '已关闭', '已驳回', '驳回', '已拒绝', '拒绝', '无效'}


def _transport_json_safe(v):
    """把 openpyxl 读出的单元格值转成 JSON 可序列化值（存入 ext_raw JSONField）。
    运输原表「对账时间/创建时间」是真正的 datetime → 直接入 JSON 会报
    "Object of type datetime is not JSON serializable"。统一转贴近原表的字符串。"""
    if isinstance(v, datetime.datetime):
        # 纯日期格（零点）只留日期，避免凭空多出 00:00:00；否则保留到秒
        if (v.hour, v.minute, v.second, v.microsecond) == (0, 0, 0, 0):
            return v.strftime('%Y-%m-%d')
        return v.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(v, datetime.date):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, datetime.time):
        return v.strftime('%H:%M:%S')
    if isinstance(v, Decimal):
        return float(v)
    return v


def _transport_analyze(rows, existing_bills):
    """运输对账单逐行分类（只读，不落库）。导入与导入预检共用同一口径，杜绝漂移。
    返回各类别清单 + 列漂移信息。existing_bills：系统内已存在的 ext_bill_no 集合。"""
    header = [str(h).strip() if h is not None else '' for h in rows[0]]
    pos = {h: i for i, h in enumerate(header)}
    report = {
        'header': header,
        'missing_required': [c for c in (TRANSPORT_KEY_COL, TRANSPORT_AMOUNT_COL) if c not in pos],
        # 列漂移：原表新增的非标准列 / 缺失的标准列
        'extra_columns': [h for h in header if h and h not in TRANSPORT_HEADERS],
        'missing_standard': [h for h in TRANSPORT_HEADERS if h not in pos],
        'total': 0,
        'ok': [], 'settled': [], 'voided': [],
        'dup_in_file': [], 'dup_in_system': [], 'bad': [], 'summary_rows': [],
    }
    if report['missing_required']:
        return report

    def cell(row, col):
        i = pos.get(col)
        return row[i] if (i is not None and i < len(row)) else None

    seen = set()
    for rn, row in enumerate(rows[1:], start=2):
        if row is None or all(v is None or str(v).strip() == '' for v in row):
            continue
        # 合计/汇总行识别：序号或对账单号含「合计/总计/小计」，或对账单号形如「681 条」
        # （条数统计而非真实单号）。这类行跳过——不计入总数、不导入，避免全表总额被当成
        # 一条巨额付款。真实对账单号形如 ZD202605310006，绝不含「条」或「合计」。
        seq_probe = str(cell(row, TRANSPORT_SEQ_COL) or '').strip()
        bill_probe = str(cell(row, TRANSPORT_KEY_COL) or '').strip()
        if (any(m in seq_probe for m in TRANSPORT_SUMMARY_MARKERS)
                or any(m in bill_probe for m in TRANSPORT_SUMMARY_MARKERS)
                or re.match(r'^\d[\d,]*\s*条$', bill_probe)):
            report.setdefault('summary_rows', []).append({'row': rn, 'bill_no': bill_probe})
            continue
        report['total'] += 1
        bill_no = str(cell(row, TRANSPORT_KEY_COL) or '').strip()
        if not bill_no:
            report['bad'].append({'row': rn, 'bill_no': '', 'reason': '缺少对账单号'})
            continue
        if bill_no in seen:
            report['dup_in_file'].append({'row': rn, 'bill_no': bill_no})
            continue
        if bill_no in existing_bills:
            report['dup_in_system'].append({'row': rn, 'bill_no': bill_no})
            continue
        src_status = str(cell(row, TRANSPORT_STATUS_COL) or '').strip()
        if src_status in TRANSPORT_BLOCK_SETTLED:
            report['settled'].append({'row': rn, 'bill_no': bill_no, 'src_status': src_status})
            continue
        if src_status in TRANSPORT_BLOCK_VOID:
            report['voided'].append({'row': rn, 'bill_no': bill_no, 'src_status': src_status})
            continue
        amt_raw = cell(row, TRANSPORT_AMOUNT_COL)
        try:
            amount = abs(Decimal(str(amt_raw).strip()))
        except (InvalidOperation, AttributeError, TypeError):
            report['bad'].append({'row': rn, 'bill_no': bill_no, 'reason': f'金额「{amt_raw}」无法解析'})
            continue
        if amount <= 0:
            report['bad'].append({'row': rn, 'bill_no': bill_no, 'reason': '金额为 0'})
            continue
        raw = {h: (_transport_json_safe(row[i]) if i < len(row) else None) for h, i in pos.items()}
        report['ok'].append({
            'row': rn, 'bill_no': bill_no, 'src_status': src_status,
            # 字段映射（运输原表列 → 我方记录字段），各列按目标字段 max_length 截断：
            #   对账单号 → 审批编号(approval_number，唯一键/复制/批量筛选主键)
            #   运单号   → G7编号(g7_number，可多张以「/」拼接)
            #   项目名称 → 项目简称(project_short_name)
            #   备注     → 备注(notes)
            #   收支方式 → 收款主体(payee)、对账对象 → 摘要(summary)
            'approval_number': bill_no[:21],
            'g7': str(cell(row, '运单号') or '').strip()[:255],
            'project_short_name': str(cell(row, '项目名称') or '').strip()[:100],
            'notes': str(cell(row, '备注') or '').strip()[:500],
            'payee': (str(cell(row, '收支方式') or '').strip() or bill_no)[:200],
            'summary': (str(cell(row, '对账对象') or '').strip() or f'运输对账 {bill_no}')[:500],
            'org': str(cell(row, '所属组织') or '').strip(),
            'amount': amount, 'raw': raw,
        })
        seen.add(bill_no)
    return report


def _transport_read_for_analyze(request):
    """运输导入/预检共用的鉴权 + 取文件 + 读行 + 取已存在单号。
    成功返回 (rows, existing_bills, None)；失败返回 (None, None, err_response)。"""
    perms = get_request_perms(request)
    if perms is not None:
        if not perms['pages'].get('approval_records', True):
            return None, None, err('无审批记录访问权限', 403, 403)
        if not perms.get('can_create'):
            return None, None, err('无新增权限', 403, 403)
    if not can_write_dept(request, TRANSPORT_DEPT):
        return None, None, err(f'无权操作部门"{TRANSPORT_DEPT}"', 403, 403)
    upload = request.FILES.get('file')
    if not upload:
        return None, None, err('请上传文件（.xlsx / .csv）')
    if upload.size > 5 * 1024 * 1024:
        return None, None, err('文件过大，请确认文件不超过5MB')
    rows, ferr = _read_import_rows(upload)
    if ferr:
        return None, None, err(ferr)
    if not rows or len(rows) < 2:
        return None, None, err('文件为空或缺少数据行')
    existing = set(ApprovalRecord.objects.filter(ext_source=TRANSPORT_SOURCE)
                   .values_list('ext_bill_no', flat=True))
    return rows, existing, None


@csrf_exempt
@pk_required()
def transport_import_precheck(request):
    """POST（审批管理）— 运输对账单导入预检：只读分析，返回逐类详单供前端确认。
    覆盖：将导入 / 源已结算（防重复付款）/ 源已作废 / 文件内重复 / 系统内重复 /
    数据异常 / 列漂移（新增列、缺标准列）。不落库。"""
    if request.method != 'POST':
        return err('Method not allowed', 405)
    rows, existing, resp = _transport_read_for_analyze(request)
    if resp:
        return resp
    report = _transport_analyze(rows, existing)
    if report['missing_required']:
        return err(f'运输对账单缺少必要列：{"、".join(report["missing_required"])}。请直接上传运输系统导出的原始表')

    def slim(i):
        return {'row': i['row'], 'bill_no': i['bill_no'],
                'payee': i['payee'], 'amount': str(i['amount']), 'src_status': i['src_status']}
    return ok({
        'total': report['total'],
        'will_import': len(report['ok']),
        'ok_sample': [slim(i) for i in report['ok'][:100]],
        'settled': report['settled'],
        'voided': report['voided'],
        'dup_in_file': report['dup_in_file'],
        'dup_in_system': report['dup_in_system'],
        'bad': report['bad'],
        'summary_rows': report['summary_rows'],
        'extra_columns': report['extra_columns'],
        'missing_standard': report['missing_standard'],
    })


@csrf_exempt
@pk_required()
def transport_import(request):
    """POST（审批管理）— 运输对账单导入：原表 → 「已通过」审批记录（金额取绝对值），
    按对账单号去重。源已结算/已作废单据跳过（防重复付款）。导入后手动排款流转付款管理。"""
    if request.method != 'POST':
        return err('Method not allowed', 405)
    rows, existing, resp = _transport_read_for_analyze(request)
    if resp:
        return resp
    report = _transport_analyze(rows, existing)
    if report['missing_required']:
        return err(f'运输对账单缺少必要列：{"、".join(report["missing_required"])}。请直接上传运输系统导出的原始表')

    results = {'created': 0, 'skipped': 0, 'duplicates': 0,
               'already_settled': 0, 'voided': 0, 'errors': []}
    importer = getattr(getattr(request, 'pk_user', None), 'name', '') or '运输导入'
    for item in report['ok']:
        bill_no = item['bill_no']
        try:
            with transaction.atomic():
                ApprovalRecord.objects.create(
                    created_by_id=request.pk_uid, applicant=importer,
                    department=TRANSPORT_DEPT, secondary_dept=item['org'],
                    approval_number=item['approval_number'],  # 审批编号 ← 对账单号
                    g7_number=item['g7'],                     # G7编号 ← 运单号
                    project_short_name=item['project_short_name'],  # 项目简称 ← 项目名称
                    summary=item['summary'][:500], notes=item['notes'],
                    amount=item['amount'], payee=item['payee'], status='approved',
                    ext_source=TRANSPORT_SOURCE, ext_bill_no=bill_no, ext_raw=item['raw'],
                )
            results['created'] += 1
        except IntegrityError:
            # DB 兜底唯一约束命中（并发）→ 计为重复
            results['duplicates'] += 1
            results['skipped'] += 1
        except Exception as e:
            results['errors'].append(f'第{item["row"]}行: 对账单 {bill_no} 保存失败 ({e})')
            results['skipped'] += 1

    # 汇总各跳过类别（与预检同口径）
    results['duplicates'] += len(report['dup_in_file']) + len(report['dup_in_system'])
    results['already_settled'] = len(report['settled'])
    results['voided'] = len(report['voided'])
    results['skipped'] += (len(report['dup_in_file']) + len(report['dup_in_system'])
                           + len(report['settled']) + len(report['voided']) + len(report['bad']))
    for s in report['settled']:
        results['errors'].append(
            f'第{s["row"]}行: 对账单 {s["bill_no"]} 源状态「{s["src_status"]}」已结算，跳过以防重复付款')
    for v in report['voided']:
        results['errors'].append(
            f'第{v["row"]}行: 对账单 {v["bill_no"]} 源状态「{v["src_status"]}」已作废/取消，跳过')
    for b in report['bad']:
        results['errors'].append(f'第{b["row"]}行: 对账单 {b["bill_no"] or "(空)"} {b["reason"]}，已跳过')
    for s in report['summary_rows']:
        results['errors'].append(f'第{s["row"]}行: 合计/汇总行（对账单号「{s["bill_no"]}」），已跳过')
    results['summary_rows'] = len(report['summary_rows'])

    parts = [f'成功导入 {results["created"]} 条（已建为「审批通过」记录，请在审批管理排款）']
    if results['duplicates']:
        parts.append(f'跳过重复 {results["duplicates"]} 条')
    if results['already_settled']:
        parts.append(f'跳过源已结算 {results["already_settled"]} 条（防重复付款）')
    if results['voided']:
        parts.append(f'跳过源已作废/取消 {results["voided"]} 条')
    if report['summary_rows']:
        parts.append(f'跳过合计行 {len(report["summary_rows"])} 条')
    if report['bad']:
        parts.append(f'跳过异常 {len(report["bad"])} 条')
    results['message'] = '；'.join(parts)
    return ok(results)


@csrf_exempt
@pk_required()
def transport_export(request):
    """GET（付款管理）— 运输对账单导出：按 ids（勾选付款记录）/ 筛选取回已排款的运输付款，
    从来源审批的 ext_raw 原样还原原表，仅把状态列改为「已结算」（限已付清行）。零误差回传。"""
    return _transport_export_core(request)


def _transport_export_core(request, export_cap=5000):
    if request.method != 'GET':
        return err('Method not allowed', 405)
    perms = get_request_perms(request)
    denied = _payments_page_denied(request, perms)
    if denied:
        return denied
    try:
        from openpyxl import Workbook
    except ImportError:
        return err('服务器缺少 openpyxl 依赖', 500)

    # 运输付款 = 来源审批为运输导入的付款记录（经手动排款流转而来）
    qs = Payment.objects.filter(approval__ext_source=TRANSPORT_SOURCE)
    qs = dept_filter(qs, request)
    qs = qs.select_related('approval').prefetch_related('installments')

    # 勾选导出优先：ids=逗号分隔的付款记录主键
    ids_raw = (request.GET.get('ids') or '').strip()
    if ids_raw:
        try:
            ids = [int(x) for x in ids_raw.split(',') if x.strip()]
        except ValueError:
            return err('ids 参数格式有误')
        qs = qs.filter(id__in=ids)
    else:
        # 否则走列表同款筛选（日期 / 关键词 / 列头筛选）
        start = request.GET.get('start_date', '').strip()
        end = request.GET.get('end_date', '').strip()
        q_str = request.GET.get('q', '').strip()
        if start:
            qs = qs.filter(planned_date__gte=start)
        if end:
            qs = qs.filter(planned_date__lte=end)
        if q_str:
            qs = qs.filter(
                Q(payee__icontains=q_str) | Q(approval__ext_bill_no__icontains=q_str) |
                Q(secondary_dept__icontains=q_str) | Q(notes__icontains=q_str)
            )
        fq, fq_distinct = build_filter_q(request.GET.get('filters', ''), PAYMENT_FILTER_REGISTRY)
        if fq:
            qs = qs.filter(fq)
            if fq_distinct:
                qs = qs.distinct()

    payments = list(qs)
    if not payments:
        return err('没有可导出的运输对账记录（请先在审批管理排款，再勾选付款记录或调整筛选）')
    if len(payments) > export_cap:
        return err(f'当前结果共 {len(payments)} 条，超出导出上限（{export_cap} 条），请缩小范围')

    # 一条审批（一张对账单）↔ 一条付款汇总记录；按对账单号去重，保证一单一行
    by_bill = {}
    for p in payments:
        rec = p.approval
        if rec is None or not isinstance(rec.ext_raw, dict) or not rec.ext_raw:
            continue
        key = rec.ext_bill_no or rec.id
        # 同单多条付款（理论上不会）取已结算优先
        if key not in by_bill or (p.status == 'settled' and by_bill[key].status != 'settled'):
            by_bill[key] = p
    rows_out = list(by_bill.values())
    if not rows_out:
        return err('选中的付款记录没有可还原的运输对账原始数据')

    # ── 导出统一为「标准表」：无论导入表列多列少，导出一律补齐为标准结构 ──
    #   · 标准列：始终按 TRANSPORT_HEADERS 全列、固定列序输出。
    #       - 列少的导入（缺标准列）→ 缺失列补空白单元，仍是完整标准表；
    #       - 列序乱（生产库 jsonb 不保证键序）→ 强制按标准列序，不漂移。
    #   · 非标准额外列：标准列之后，按「所有行」首次出现顺序并集追加，避免丢数据
    #       （列多的导入不丢列，但标准块永远在前、结构稳定、可被运输系统原样回导）。
    all_raws = [(r.approval.ext_raw or {}) for r in rows_out]
    canonical = list(TRANSPORT_HEADERS)        # 始终输出全部标准列（缺失留空）
    seen_cols = set(canonical)
    extras = []
    for raw in all_raws:
        for k in raw.keys():
            if k not in seen_cols:
                seen_cols.add(k)
                extras.append(k)
    headers = canonical + extras
    if TRANSPORT_STATUS_COL not in headers:
        headers.append(TRANSPORT_STATUS_COL)

    wb = Workbook()
    ws = wb.active
    ws.title = '运输对账单'
    # 零误差还原：表头为原表纯文本，不施加任何品牌底色/字体样式（与来源系统原表一致，
    # 仅状态列的「值」可改写）。导出 = 原表逐字 + 已结算行状态列改为「已结算」。
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)

    _FORMULA_CHARS = ('=', '+', '@')

    def _safe(v):
        # Excel 注入防护：仅对会被当成公式的纯文本前缀加引号，不动数字 → 不改变数值精度
        if isinstance(v, str) and v and v[0] in _FORMULA_CHARS:
            return "'" + v
        return v

    for ri, p in enumerate(rows_out, start=2):
        raw = p.approval.ext_raw
        settled = p.status == 'settled'
        for ci, h in enumerate(headers, 1):
            if h == TRANSPORT_SEQ_COL:
                # 序号按导出行顺序重排 1,2,3…：导出常是勾选/筛选/去重后的子集，
                # 沿用原表零散序号会断号，统一从 1 连续编号更规范。
                val = ri - 1
            elif h == TRANSPORT_STATUS_COL:
                # 状态列以「我方结算」为准（这是导出存在的意义）：
                #   已结算 → 写「已结算」；
                #   未结算 → 保留源原值（零误差）；但若源原值本身就是「已结算类」
                #   （历史遗留：在导入状态校验之前导入的脏数据），不沿用以免谎报已结算，
                #   改写为明确的「未结算」，杜绝「我方未付却显示已结算」的错配。
                if settled:
                    val = TRANSPORT_SETTLED_LABEL
                else:
                    orig = raw.get(h)
                    val = '未结算' if str(orig or '').strip() in TRANSPORT_BLOCK_SETTLED else orig
            else:
                val = raw.get(h)
            # 空单元格统一还原为空字符串，与来源系统原表表示一致（避免 None/'' 表象差异）
            if val is None:
                val = ''
            ws.cell(row=ri, column=ci, value=_safe(val))

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 36)

    today = timezone.localdate().strftime('%Y%m%d')
    # 运输导出零误差还原：保留其自带的更窄逐格防护（仅 = + @），不施加 chokepoint 全表扫描，
    # 以免给以「-」开头的原始文本前置单引号、破坏与运输系统的原样回导契约。
    return _build_excel_response(wb, f'运输对账单_已结算_{today}.xlsx', formula_guard=False)


# ── 异步导出（大数据量后台任务）──────────────────────────────────────────────
# 同步导出上限 5000 行，超出改后台异步生成：前端建任务→轮询状态→完成后下载。
# 文件字节存 ExportJob.file_data（DB），跨 gunicorn worker 共享，无需 Redis/Celery。
ASYNC_EXPORT_CAP = 100000   # 后台导出硬上限（保护内存与 DB BLOB 体积）

# 哪些列表筛选参数需要快照进 ExportJob.params 以便后台重建同口径 queryset。
_EXPORT_PARAM_KEYS = (
    'q', 'filters', 'sort', 'order', 'depts', 'dept', 'status', 'hide_settled',
    'start_date', 'end_date', 'pay_date_start', 'pay_date_end', 'g7_number', 'ids',
)


class _ReplayRequest:
    """后台导出 worker 用：模拟一个已认证的 GET 请求，承载导出核心读取的全部上下文
    （pk_* 身份 + GET 筛选参数）。导出核心只读不写，故无需真实 HttpRequest。"""
    method = 'GET'

    def __init__(self, user, params):
        self.pk_user = user
        self.pk_uid = user.id
        self.pk_role = user.role
        self.pk_job = user.job_title
        self.pk_depts = user.departments or []
        self.GET = params or {}


def _run_export_job(job_id):
    """后台线程：重建筛选 → 生成 xlsx → 字节落库。独立 DB 连接，结束务必关闭。"""
    from django.db import connection
    from paikuan.models import ExportJob
    _CORES = {'approvals': _approval_export_core, 'payments': _payment_export_core,
              'transport': _transport_export_core}
    try:
        job = ExportJob.objects.get(pk=job_id)
    except ExportJob.DoesNotExist:
        return
    try:
        job.status = 'running'
        job.save(update_fields=['status'])
        core = _CORES.get(job.kind)
        if core is None:
            raise ValueError(f'未知导出类型 {job.kind}')
        req = _ReplayRequest(job.created_by, job.params)
        resp = core(req, export_cap=ASYNC_EXPORT_CAP)
        if getattr(resp, 'status_code', 200) != 200:
            # 错误响应体为 JSON {error/msg}
            try:
                payload = json.loads(resp.content.decode('utf-8'))
                msg = payload.get('error') or payload.get('msg') or '导出失败'
            except Exception:
                msg = '导出失败'
            job.status = 'failed'
            job.error = msg
            job.finished_at = timezone.now()
            job.save(update_fields=['status', 'error', 'finished_at'])
            return
        content = resp.content
        # 从 Content-Disposition 还原文件名（_build_excel_response 已 URL 编码）
        job.file_data = content
        if not job.filename:
            job.filename = {'approvals': '审批记录.xlsx', 'payments': '排款记录.xlsx',
                            'transport': '运输对账单.xlsx'}.get(job.kind, '导出.xlsx')
        job.status = 'done'
        job.finished_at = timezone.now()
        job.save(update_fields=['file_data', 'filename', 'status', 'finished_at'])
    except Exception as e:
        logger.exception('export job %s failed', job_id)
        try:
            job.status = 'failed'
            job.error = str(e)[:500]
            job.finished_at = timezone.now()
            job.save(update_fields=['status', 'error', 'finished_at'])
        except Exception:
            pass
    finally:
        connection.close()


@csrf_exempt
@pk_required()
def export_create(request):
    """POST {kind, params} — 建后台导出任务并启动线程。返回 {id, status}。
    kind ∈ approvals|payments|transport；params 为列表筛选参数快照。"""
    from paikuan.models import ExportJob
    if request.method != 'POST':
        return err('POST only', 405)
    body = parse_body(request)
    kind = (body.get('kind') or '').strip()
    if kind not in ('approvals', 'payments', 'transport'):
        return err('未知导出类型')
    # 权限校验：与对应列表页一致
    perms = get_request_perms(request)
    if kind == 'approvals':
        if perms is not None and not perms['pages'].get('approval_records', True):
            return err('无访问权限', 403, 403)
    else:
        denied = _payments_page_denied(request, perms)
        if denied:
            return denied
    raw_params = body.get('params') or {}
    params = {k: str(raw_params[k]) for k in _EXPORT_PARAM_KEYS if k in raw_params and raw_params[k] not in (None, '')}
    job = ExportJob.objects.create(kind=kind, created_by=request.pk_user, params=params)
    t = threading.Thread(target=_run_export_job, args=(job.id,), daemon=True)
    t.start()
    return ok(job.to_dict())


@pk_required()
def export_status(request, pk):
    """GET — 单个导出任务状态（轮询用）。"""
    from paikuan.models import ExportJob
    job = ExportJob.objects.filter(pk=pk, created_by_id=request.pk_uid).first()
    if not job:
        return err('任务不存在', 404)
    return ok(job.to_dict())


@pk_required()
def export_list(request):
    """GET — 当前用户最近的导出任务（默认 20 条）。"""
    from paikuan.models import ExportJob
    jobs = ExportJob.objects.filter(created_by_id=request.pk_uid)[:20]
    return ok({'items': [j.to_dict() for j in jobs]})


@pk_required()
def export_download(request, pk):
    """GET — 下载已完成的导出文件字节。"""
    from paikuan.models import ExportJob
    job = ExportJob.objects.filter(pk=pk, created_by_id=request.pk_uid).first()
    if not job:
        return err('任务不存在', 404)
    if job.status != 'done' or not job.file_data:
        return err('文件尚未生成完成')
    resp = HttpResponse(
        bytes(job.file_data),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(job.filename)}"
    return resp


# 跨页全选上限：与各批量操作的单次上限对齐（批量删除/审批/排款/付款均为 1000）。
SELECT_ALL_CAP = 5000   # 跨页「全选筛选结果」批量操作上限（批量删除/标记/退回等）
# 运输单号复制上限：与导出上限对齐，足够覆盖整批对账。
G7_COPY_CAP = 5000


@pk_required()
def payments_select_ids(request):
    """跨页全选：返回当前筛选口径下的全部付款记录 ID（与列表同口径，去分页）。
    供「选择全部筛选结果」后做跨页批量操作。上限 SELECT_ALL_CAP，超出时 capped=True。"""
    if request.method != 'GET':
        return err('Method not allowed', 405)
    perms = get_request_perms(request)
    denied = _payments_page_denied(request, perms)
    if denied:
        return denied
    qs, _ = _payments_filtered_qs(request)
    # 清空排序避免 DISTINCT + 非选择列 ORDER BY 在部分数据库报错；多取 1 条用于判溢出
    ids = list(qs.order_by().values_list('id', flat=True).distinct()[:SELECT_ALL_CAP + 1])
    capped = len(ids) > SELECT_ALL_CAP
    ids = ids[:SELECT_ALL_CAP]
    return ok({'ids': ids, 'count': len(ids), 'capped': capped, 'cap': SELECT_ALL_CAP})


@pk_required()
def transport_g7_numbers(request):
    """运输专用：返回所选(ids)/当前筛选口径下付款记录的 审批编号（=对账单号）去重列表，
    供前端以「+」或空格连接复制到剪贴板。跨页全量，不受分页限制。上限 G7_COPY_CAP。
    注：运输映射调整后「单号」对应审批编号（对账单号），不再取 G7编号（现为运单号）。"""
    if request.method != 'GET':
        return err('Method not allowed', 405)
    perms = get_request_perms(request)
    denied = _payments_page_denied(request, perms)
    if denied:
        return denied
    ids_raw = (request.GET.get('ids') or '').strip()
    if ids_raw:
        try:
            ids = [int(x) for x in ids_raw.split(',') if x.strip()]
        except ValueError:
            return err('ids 参数格式有误')
        qs = dept_filter(Payment.objects.filter(pk__in=ids), request)
    else:
        qs, _ = _payments_filtered_qs(request)
    # 保序去重（清空排序以规避 DISTINCT/ORDER BY 跨库差异；顺序对拼接用途不敏感）
    seen, nums = set(), []
    for no in qs.order_by().values_list('approval_number', flat=True).iterator():
        no = (no or '').strip()
        # 跳过空/占位（全 0）审批编号
        if not no or set(no) == {'0'} or no in seen:
            continue
        seen.add(no)
        nums.append(no)
        if len(nums) >= G7_COPY_CAP:
            break
    return ok({'numbers': nums, 'count': len(nums), 'capped': len(nums) >= G7_COPY_CAP})


@pk_required()
def approvals_select_ids(request):
    """跨页全选：返回当前筛选口径下的全部审批记录 ID（与列表同口径，去分页）。
    供「选择全部筛选结果」后做跨页批量操作。上限 SELECT_ALL_CAP。"""
    if request.method != 'GET':
        return err('Method not allowed', 405)
    perms = get_request_perms(request)
    if perms is not None and not perms['pages'].get('approval_records', True):
        return err('无访问权限', 403, 403)
    qs = _approvals_filtered_qs(request)
    ids = list(qs.order_by().values_list('id', flat=True).distinct()[:SELECT_ALL_CAP + 1])
    capped = len(ids) > SELECT_ALL_CAP
    ids = ids[:SELECT_ALL_CAP]
    return ok({'ids': ids, 'count': len(ids), 'capped': capped, 'cap': SELECT_ALL_CAP})


# ── departments ───────────────────────────────────────────────────────────────

@csrf_exempt
@pk_required()
def trash_approvals(request):
    """回收站 — 审批管理（软删记录查询/恢复/彻底删除）。
    GET：列出被软删的审批记录。
    POST { action: 'restore'|'purge', ids: [...] }：还原或永久删除。
    """
    perms = get_request_perms(request)
    if perms is not None and not perms['pages'].get('approval_records', True):
        return err('无访问权限', 403, 403)
    qs = dept_filter(ApprovalRecord.objects.filter(deleted_at__isnull=False), request)
    # 页内事业部筛选（在全局作用域之内进一步收窄）；同时作用于列表与跨页 all 操作
    _dept = request.GET.get('dept', '').strip()
    if _dept:
        qs = qs.filter(department=_dept)

    if request.method == 'GET':
        try:
            page = max(1, int(request.GET.get('page', 1)))
            size = min(100, max(1, int(request.GET.get('size', 50))))
        except ValueError:
            page, size = 1, 50
        total = qs.count()
        rows = qs.order_by('-deleted_at')[(page-1)*size:page*size]
        return ok({'total': total, 'items': [
            {**r.to_dict(), 'deleted_at': r.deleted_at.isoformat() if r.deleted_at else None,
             'deleted_by_name': r.deleted_by.name if r.deleted_by else None}
            for r in rows
        ]})

    if request.method == 'POST':
        if perms is not None and not perms.get('can_delete'):
            return err('无删除权限', 403, 403)
        body = parse_body(request)
        action = body.get('action', 'restore')
        # 跨页全选：all=true 作用于当前作用域下「全部软删记录」（上限保护，超出分批处理）
        if body.get('all'):
            targets = list(qs.order_by('-deleted_at')[:SELECT_ALL_CAP])
        else:
            ids = [int(i) for i in (body.get('ids') or [])]
            if not ids:
                return err('ids 必填或传 all:true')
            targets = list(qs.filter(pk__in=ids))
        count = 0
        for rec in targets:
            if action == 'restore':
                rec.deleted_at = None
                rec.deleted_by = None
                rec.save(update_fields=['deleted_at', 'deleted_by'])
            elif action == 'purge':
                if rec.payments.exists():
                    continue
                rec.delete()
            count += 1
        return ok({'count': count, 'action': action})
    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def trash_payments(request):
    """回收站 — 付款管理（软删记录查询/恢复/彻底删除）。
    GET：列出被软删的付款记录。
    POST { action: 'restore'|'purge', ids: [...] }：还原或永久删除。
    """
    perms = get_request_perms(request)
    denied = _payments_page_denied(request, perms)
    if denied:
        return denied
    qs = dept_filter(Payment.objects.filter(deleted_at__isnull=False), request)
    # 页内事业部筛选（在全局作用域之内进一步收窄）；同时作用于列表与跨页 all 操作
    _dept = request.GET.get('dept', '').strip()
    if _dept:
        qs = qs.filter(department=_dept)

    if request.method == 'GET':
        try:
            page = max(1, int(request.GET.get('page', 1)))
            size = min(100, max(1, int(request.GET.get('size', 50))))
        except ValueError:
            page, size = 1, 50
        total = qs.count()
        rows = list(qs.order_by('-deleted_at')[(page-1)*size:page*size])
        return ok({'total': total, 'items': [
            {**p.to_dict(), 'deleted_at': p.deleted_at.isoformat() if p.deleted_at else None,
             'deleted_by_name': p.deleted_by.name if p.deleted_by else None}
            for p in rows
        ]})

    if request.method == 'POST':
        if perms is not None and not perms.get('can_delete'):
            return err('无删除权限', 403, 403)
        body = parse_body(request)
        action = body.get('action', 'restore')
        # 跨页全选：all=true 作用于当前作用域下「全部软删记录」（上限保护，超出分批处理）
        if body.get('all'):
            targets = list(qs.order_by('-deleted_at')[:SELECT_ALL_CAP])
        else:
            ids = [int(i) for i in (body.get('ids') or [])]
            if not ids:
                return err('ids 必填或传 all:true')
            targets = list(qs.filter(pk__in=ids))
        count = 0
        skipped = []
        for p in targets:
            if action == 'restore':
                approval_id = p.approval_id
                p.deleted_at = None
                p.deleted_by = None
                try:
                    # 还原会重算 dedup_key；若同业务键已有在册记录（删除后又新建过），
                    # 唯一约束会拦下，给出明确提示而非 500。
                    with transaction.atomic():
                        p.save(update_fields=['deleted_at', 'deleted_by'])
                except IntegrityError:
                    skipped.append({'id': p.id, 'reason': '相同业务键已有在册排款，无法还原'})
                    continue
                _reconcile_approval_schedule(approval_id)
            elif action == 'purge':
                if p.prepaid_offsets.exists():
                    continue
                approval_id = p.approval_id
                with transaction.atomic():
                    _record_payment_changes(p, {}, {}, request, action='delete')
                    p.delete()
                    _reconcile_approval_schedule(approval_id)
            count += 1
        return ok({'count': count, 'action': action, 'skipped': skipped})
    return err('Method not allowed', 405)


@pk_required()
def departments(request):
    if request.method != 'GET':
        return err('Method not allowed', 405)
    denied = _payments_page_denied(request)
    if denied:
        return denied
    # Only return canonical departments. Historical records with non-canonical names
    # (from before strict validation) should not pollute the department picker.
    merged = list(DEPARTMENTS)
    # Non-admins may only see/filter departments they are assigned to,
    # matching the row-level visibility enforced by dept_filter().
    if request.pk_role != 'super_admin':
        allowed = set(request.pk_depts or [])
        merged = [d for d in merged if d in allowed]
    return ok(merged)


# ── 操作审计日志（仅超管）───────────────────────────────────────────────────────

@csrf_exempt
@pk_required(roles=['super_admin'])
def audit_logs(request):
    """全系统操作审计查询：按操作人/模块/方法/结果/路径关键词/日期过滤，分页。"""
    from paikuan.models import AuditLog
    if request.method != 'GET':
        return err('Method not allowed', 405)
    qs = AuditLog.objects.all()
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(Q(path__icontains=q) | Q(user_name__icontains=q))
    module = request.GET.get('module', '').strip()
    if module:
        qs = qs.filter(module=module)
    method = request.GET.get('method', '').strip().upper()
    if method:
        qs = qs.filter(method=method)
    result = request.GET.get('result', '').strip()   # ok | fail
    if result == 'ok':
        qs = qs.filter(status_code__lt=400)
    elif result == 'fail':
        qs = qs.filter(status_code__gte=400)
    date_start = request.GET.get('date_start', '').strip()
    if date_start:
        qs = qs.filter(created_at__date__gte=date_start)
    date_end = request.GET.get('date_end', '').strip()
    if date_end:
        qs = qs.filter(created_at__date__lte=date_end)

    # Excel 风格列头筛选 + 排序（白名单解析；与上方旧参数向后兼容并存）
    fq, fq_distinct = build_filter_q(request.GET.get('filters', ''), AUDITLOG_FILTER_REGISTRY)
    if fq:
        qs = qs.filter(fq)
        if fq_distinct:
            qs = qs.distinct()
    sort_by = resolve_sort(request.GET.get('sort'), request.GET.get('order'), AUDITLOG_FILTER_REGISTRY)
    if sort_by:
        qs = qs.order_by(sort_by)

    total = qs.count()
    page = max(1, int(request.GET.get('page', 1) or 1))
    size = min(200, max(1, int(request.GET.get('size', 50) or 50)))
    items = [l.to_dict() for l in qs[(page - 1) * size: page * size]]
    return ok({'items': items, 'total': total, 'page': page, 'size': size})


@csrf_exempt
@pk_required(roles=['super_admin'])
def audit_logs_prune(request):
    """清理 N 天前的审计日志（默认保留180天），防止无限增长。"""
    from paikuan.models import AuditLog
    if request.method != 'POST':
        return err('Method not allowed', 405)
    data = parse_body(request)
    try:
        keep_days = max(30, min(3650, int(data.get('keep_days', 180))))
    except (ValueError, TypeError):
        keep_days = 180
    cutoff = timezone.now() - datetime.timedelta(days=keep_days)
    deleted, _ = AuditLog.objects.filter(created_at__lt=cutoff).delete()
    return ok({'deleted': deleted, 'keep_days': keep_days})
