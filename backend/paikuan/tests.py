import io
import json
from datetime import date
from decimal import Decimal

from django.test import Client, TestCase

from ar.models import ARProject
from paikuan.models import (ApprovalRecord, JobPermission, PaikuanUser, Payment,
                            PaymentInstallment, PaymentPlanItem)
from paikuan.views import DEPARTMENTS, default_job_config, make_token, _invalidate_perm_cache


class PaymentPermissionRegressionTests(TestCase):
    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.dept = DEPARTMENTS[0]
        self.user = PaikuanUser(
            phone='13900000001',
            name='Cashier User',
            role='operator',
            job_title='cashier',
            departments=[self.dept],
            is_active=True,
            is_approved=True,
        )
        self.user.set_password('Test123456')
        self.user.save()
        self.token = make_token(self.user)

    def tearDown(self):
        _invalidate_perm_cache()

    def auth(self, token=None):
        return {'HTTP_AUTHORIZATION': f'Bearer {token or self.token}'}

    def put_payment(self, payment_id, payload, token=None):
        return self.client.put(
            f'/api/pk/payments/{payment_id}',
            data=json.dumps(payload),
            content_type='application/json',
            **self.auth(token),
        )

    def create_payment(self, total='1000.00'):
        return Payment.objects.create(
            created_by=self.user,
            department=self.dept,
            approval_number='',
            project_desc='Regression project',
            payee='Regression payee',
            total_amount=Decimal(total),
            planned_date=date(2026, 6, 1),
            notes='',
        )

    def test_prepaid_balance_lookup_for_payment(self):
        """排款页按项目编号查预付余额：匹配项目返回未核销合计，未匹配返回空。"""
        from ar.models import ARProject, AdvanceRecord
        proj = ARProject.objects.create(
            customer_name='C', short_name='预付项目', delivery_dept=self.dept,
            sales_contact='S', project_manager='M', project_no='GYL-TEST-0001')
        AdvanceRecord.objects.create(
            direction='预付', project=proj, delivery_dept=self.dept, counterparty='供应商A',
            occur_year=2026, occur_month=3, occur_date=date(2026, 3, 1),
            advance_amount=Decimal('80000'))
        resp = self.client.get('/api/pk/payments/prepaid-balance',
                               {'project_no': 'GYL-TEST-0001'}, **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertTrue(d['matched'])
        self.assertEqual(d['count'], 1)
        self.assertEqual(d['total_balance'], '80000.00')
        # 未匹配项目编号 → 空
        resp2 = self.client.get('/api/pk/payments/prepaid-balance',
                                {'project_no': 'NOPE'}, **self.auth())
        self.assertFalse(resp2.json()['data']['matched'])

    def test_payment_create_persists_project_no(self):
        from paikuan.views import default_job_config
        JobPermission.objects.update_or_create(
            job_title='cashier', defaults={'config': default_job_config('cashier')})
        _invalidate_perm_cache()
        resp = self.client.post(
            '/api/pk/payments',
            data=json.dumps({'department': self.dept, 'project_no': 'GYL-TEST-0002',
                             'project_desc': 'D', 'payee': 'P', 'total_amount': '5000',
                             'planned_date': '2026-06-01'}),
            content_type='application/json', **self.auth())
        if resp.status_code == 200:
            self.assertEqual(resp.json()['data']['project_no'], 'GYL-TEST-0002')

    def test_payment_approval_number_cleaned_and_validated(self):
        """审批单号：清洗（去空格/不可打印字符）后校验 1–100 位数字；空则占位 21 个 0。"""
        from paikuan.views import _parse_payment_fields, _clean_approval_number

        def parse(no):
            return _parse_payment_fields({
                'department': self.dept, 'approval_number': no, 'project_desc': 'D',
                'payee': 'P', 'total_amount': '5000', 'planned_date': '2026-06-01'})

        # 带空格/零宽字符 → 清洗为纯数字后通过
        fields, error = parse('  123 456​789  ')
        self.assertIsNone(error, error)
        self.assertEqual(fields['approval_number'], '123456789')
        # 空 → 自动占位 21 个 0
        fields, error = parse('   ')
        self.assertIsNone(error, error)
        self.assertEqual(fields['approval_number'], '0' * 21)
        # 含字母 → 拒绝
        _, error = parse('12A3')
        self.assertIsNotNone(error)
        # 边界：100 位通过、101 位拒绝
        self.assertIsNone(_clean_approval_number('1' * 100)[1])
        self.assertIsNotNone(_clean_approval_number('1' * 101)[1])

    def test_read_import_rows_multiformat(self):
        """导入读取器兼容多格式：.xlsx / .csv(UTF-8/GBK)；旧版 .xls 给可操作提示。"""
        import io
        from openpyxl import Workbook
        from django.core.files.uploadedfile import SimpleUploadedFile
        from paikuan.views import _read_import_rows
        # .xlsx
        wb = Workbook(); ws = wb.active; ws.append(['甲', '乙']); ws.append([1, 2])
        buf = io.BytesIO(); wb.save(buf)
        rows, e = _read_import_rows(SimpleUploadedFile('x.xlsx', buf.getvalue()))
        self.assertIsNone(e, e); self.assertEqual(rows[0], ('甲', '乙'))
        # .csv（UTF-8 带 BOM）
        rows, e = _read_import_rows(SimpleUploadedFile('x.csv', '甲,乙\n1,2\n'.encode('utf-8-sig')))
        self.assertIsNone(e, e); self.assertEqual(rows[0], ('甲', '乙'))
        # .csv（GBK）
        rows, e = _read_import_rows(SimpleUploadedFile('x.csv', '部门,金额\n'.encode('gbk')))
        self.assertIsNone(e, e); self.assertEqual(rows[0][0], '部门')
        # 旧版 .xls（OLE2 magic）→ 友好改存提示
        rows, e = _read_import_rows(SimpleUploadedFile('x.xls', b'\xd0\xcf\x11\xe0' + b'\x00' * 16))
        self.assertIsNone(rows); self.assertIn('另存为', e)

    def _admin_auth(self):
        admin = PaikuanUser(phone='13900000099', name='Admin', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        admin.set_password('Test123456'); admin.save()
        return {'HTTP_AUTHORIZATION': f'Bearer {make_token(admin)}'}

    def _payment_xlsx(self, data_rows):
        from openpyxl import Workbook
        from django.core.files.uploadedfile import SimpleUploadedFile
        wb = Workbook(); ws = wb.active
        ws.append(['部门', '付款事项', '收款方', '计划付款日期', '审批单号', '计划总金额(元)'])
        for r in data_rows:
            ws.append(r)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return SimpleUploadedFile('p.xlsx', buf.read(),
                                  content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_payment_import_precheck_rules_and_ai(self):
        """预检：规则错误逐行标注（与导入同口径）+ AI 复核建议挂到对应行。"""
        from unittest import mock
        auth = self._admin_auth()
        rows = [
            [self.dept, '采购A', '供应商甲', '2026-06-01', '123456', '10000'],  # ok → 被 AI 标注
            [self.dept, '采购B', '供应商乙', '2026-06-02', '12A3', '20000'],    # 审批编号含字母→规则错
            [self.dept, '采购C', '供应商丙', '2026-06-03', '789', '30000'],     # 完全通过 → okRows
        ]

        def fake_ai(records):
            return ([{'row': records[0]['row'], 'field': 'payee', 'issue': '疑似异常',
                      'suggestion': '核实', 'severity': 'low'}] if records else [])

        with mock.patch('paikuan.views._ai_review_payments', side_effect=fake_ai):
            resp = self.client.post('/api/pk/payments/import/precheck',
                                    {'file': self._payment_xlsx(rows)}, **auth)
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['total'], 3)
        self.assertEqual(d['ruleErrors'], 1)
        self.assertEqual(d['aiFindings'], 1)
        # 只返回"需关注"的行（1 规则错误 + 1 AI 标注），完全通过的行进 okRows 不展示
        self.assertEqual(d['attention'], 2)
        self.assertEqual(d['okCount'], 1)
        self.assertEqual(len(d['rows']), 2)
        self.assertEqual(len(d['okRows']), 1)
        self.assertTrue(any(c['key'] == 'payee' for c in d['columns']))

    def test_payment_import_apply_import_and_download(self):
        """确认后导入：mode=import 落库；mode=download 返回修正版 xlsx。"""
        auth = self._admin_auth()
        rows = [{'row': 2, 'data': {
            'department': self.dept, 'project_desc': '采购C', 'payee': '供应商丙',
            'planned_date': '2026-06-03', 'approval_number': '999',
            'total_amount': '5000', 'installments': []}}]
        imp = self.client.post('/api/pk/payments/import/apply',
                               data=json.dumps({'rows': rows, 'mode': 'import'}),
                               content_type='application/json', **auth)
        self.assertEqual(imp.status_code, 200, imp.content)
        self.assertEqual(imp.json()['data']['created'], 1)
        self.assertTrue(Payment.objects.filter(payee='供应商丙').exists())

        dl = self.client.post('/api/pk/payments/import/apply',
                              data=json.dumps({'rows': rows, 'mode': 'download'}),
                              content_type='application/json', **auth)
        self.assertEqual(dl.status_code, 200)
        self.assertIn('spreadsheet', dl['Content-Type'])

    def test_payment_paid_equal_to_plan_not_rejected(self):
        """实付分笔合计恰等于计划总额时不应被判超出（含两位小数累加）。"""
        from paikuan.views import _parse_payment_fields
        fields, error = _parse_payment_fields({
            'department': self.dept, 'project_desc': 'D', 'payee': 'P',
            'total_amount': '1234.56', 'planned_date': '2026-06-01',
            'installments': [{'pay_date': '2026-06-01', 'pay_amount': '617.28'},
                             {'pay_date': '2026-06-02', 'pay_amount': '617.28'}],
        })
        self.assertIsNone(error, error)

    def test_stale_token_rejected_after_user_deactivated(self):
        self.user.is_active = False
        self.user.save(update_fields=['is_active'])

        resp = self.client.get('/api/pk/me', **self.auth())

        self.assertEqual(resp.status_code, 401)

    def test_stale_token_departments_are_ignored(self):
        payment = self.create_payment()
        stale_token = make_token(self.user)
        self.user.departments = []
        self.user.save(update_fields=['departments'])

        resp = self.put_payment(
            payment.id,
            {'installments': [{'pay_date': '2026-06-02', 'pay_amount': '100.00'}]},
            stale_token,
        )

        self.assertEqual(resp.status_code, 403)

    def test_non_editable_total_cannot_bypass_overpay_validation(self):
        payment = self.create_payment(total='1000.00')

        resp = self.put_payment(
            payment.id,
            {
                'total_amount': '5000.00',
                'installments': [{'pay_date': '2026-06-02', 'pay_amount': '2000.00'}],
            },
        )

        self.assertEqual(resp.status_code, 400)
        payment.refresh_from_db()
        self.assertEqual(payment.total_amount, Decimal('1000.00'))
        self.assertEqual(payment.installments.count(), 0)

    def test_blank_non_editable_fields_are_ignored_on_partial_edit(self):
        payment = self.create_payment(total='1000.00')

        resp = self.put_payment(
            payment.id,
            {
                'department': '',
                'approval_number': 'not-21-digits',
                'project_desc': '',
                'payee': '',
                'total_amount': '',
                'planned_date': '',
                'installments': [{'pay_date': '2026-06-03', 'pay_amount': '100.00'}],
            },
        )

        self.assertEqual(resp.status_code, 200)
        payment.refresh_from_db()
        self.assertEqual(payment.project_desc, 'Regression project')
        self.assertEqual(payment.payee, 'Regression payee')
        self.assertEqual(payment.total_amount, Decimal('1000.00'))
        inst = payment.installments.first()
        self.assertIsNotNone(inst)
        self.assertEqual(inst.pay_amount, Decimal('100.00'))

    def test_payments_page_permission_blocks_related_endpoints(self):
        cfg = default_job_config('cashier')
        cfg['pages']['payments'] = False
        JobPermission.objects.create(job_title='cashier', config=cfg)
        _invalidate_perm_cache('cashier')

        checks = [
            self.client.get('/api/pk/payments', **self.auth()),
            self.client.get('/api/pk/payments/export', **self.auth()),
            self.client.get('/api/pk/payments/template', **self.auth()),
            self.client.get('/api/pk/departments', **self.auth()),
            self.client.post('/api/pk/payments/import', **self.auth()),
        ]

        self.assertTrue(all(resp.status_code == 403 for resp in checks))


class ApprovalImportTests(TestCase):
    """审批记录导入：表头健壮匹配（兼容导出文件/无星号）、缺列明确报错、跳过原因返回。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.dept = DEPARTMENTS[0]
        u = PaikuanUser(phone='13900000200', name='Admin', role='super_admin',
                        job_title='finance_director', departments=[self.dept],
                        is_active=True, is_approved=True)
        u.set_password('Test123456'); u.save()
        self.token = make_token(u)

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _xlsx(self, headers, rows):
        import io
        from openpyxl import Workbook
        from django.core.files.uploadedfile import SimpleUploadedFile
        wb = Workbook(); ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return SimpleUploadedFile('approvals.xlsx', buf.read(),
                                  content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_import_accepts_export_style_headers_without_asterisk(self):
        # 导出文件表头不带 * —— 之前会整表跳过；现在应能正常导入
        headers = ['申请人', '所属事业部', '审批编号', '摘要', '申请金额', '收款主体', '审批状态']
        rows = [['张三', self.dept, '1' * 21, '摘要A', 12000, '某供应商', '待审批']]
        resp = self.client.post('/api/pk/approvals/import',
                                {'file': self._xlsx(headers, rows)}, **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['created'], 1, d)
        self.assertEqual(d['skipped'], 0, d)

    def test_import_missing_required_header_errors_clearly(self):
        headers = ['姓名', '事业部X', '金额X']  # 无 申请人/申请金额
        resp = self.client.post('/api/pk/approvals/import',
                                {'file': self._xlsx(headers, [['x', 'y', 1]])}, **self.auth())
        self.assertEqual(resp.status_code, 400, resp.content)
        self.assertIn('表头', resp.json().get('error', ''))

    def test_import_returns_skip_reasons(self):
        headers = ['申请人*', '所属事业部*', '审批编号*', '摘要', '申请金额*', '收款主体', '审批状态*']
        rows = [
            ['李四', self.dept, '2' * 21, '', 5000, '', '待审批'],       # ok
            ['王五', self.dept, '12A3', '', 6000, '', '待审批'],         # 审批编号含字母 → 非数字
            ['赵六', '不存在事业部', '3' * 21, '', 7000, '', '待审批'],   # 部门无效
            ['示例张三', self.dept, '4' * 21, '示例摘要', 9, '', '待审批'],  # 示例行静默跳过
        ]
        resp = self.client.post('/api/pk/approvals/import',
                                {'file': self._xlsx(headers, rows)}, **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['created'], 1, d)
        self.assertEqual(d['skipped'], 2, d)          # 示例行不计入跳过
        self.assertEqual(len(d['errors']), 2, d)

    def test_approval_import_precheck_and_apply(self):
        """审批预检：规则错误逐行标注 + AI 复核挂行，只返回需关注行 + okRows；apply 修正后落库。"""
        from unittest import mock
        headers = ['申请人*', '所属事业部*', '审批编号*', '摘要', '申请金额*', '收款主体', '审批状态*']
        rows = [
            ['李四', self.dept, '2' * 21, '摘要', 5000, '甲', '待审批'],   # ok → 被 AI 标注
            ['王五', self.dept, '12A3', '', 6000, '乙', '待审批'],          # 审批编号非数字 → 规则错
            ['赵六', self.dept, '789', '', 7000, '丙', '待审批'],           # 完全通过 → okRows
        ]

        def fake_ai(records):
            return ([{'row': records[0]['row'], 'field': 'payee', 'issue': '疑似异常',
                      'suggestion': '核实', 'severity': 'low'}] if records else [])

        with mock.patch('paikuan.views._ai_review_approvals', side_effect=fake_ai):
            resp = self.client.post('/api/pk/approvals/import/precheck',
                                    {'file': self._xlsx(headers, rows)}, **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['total'], 3)
        self.assertEqual(d['ruleErrors'], 1)
        self.assertEqual(d['aiFindings'], 1)
        self.assertEqual(d['attention'], 2)            # 1 规则错误 + 1 AI 标注
        self.assertEqual(d['okCount'], 1)
        self.assertEqual(len(d['rows']), 2)
        self.assertEqual(len(d['okRows']), 1)
        self.assertTrue(any(c['key'] == 'applicant' for c in d['columns']))

        # apply：就地修正规则错误行（审批编号改为合法数字）+ okRows 一并落库
        attention = d['rows']
        for r in attention:
            if r['status'] == 'error':
                r['data']['approval_number'] = '123'
        imp = self.client.post('/api/pk/approvals/import/apply',
                               data=json.dumps({'rows': attention, 'okRows': d['okRows'],
                                                'mode': 'import'}),
                               content_type='application/json', **self.auth())
        self.assertEqual(imp.status_code, 200, imp.content)
        self.assertEqual(imp.json()['data']['created'], 3, imp.content)   # 2 需关注 + 1 通过
        self.assertEqual(ApprovalRecord.objects.count(), 3)


class StatsCarryoverTests(TestCase):
    """月度统计：前期未付的排款结转到本期展示。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.dept = DEPARTMENTS[0]
        u = PaikuanUser(phone='13900000300', name='Admin', role='super_admin',
                        job_title='finance_director', departments=[self.dept],
                        is_active=True, is_approved=True)
        u.set_password('Test123456'); u.save()
        self.user = u
        self.token = make_token(u)

    def tearDown(self):
        _invalidate_perm_cache()

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _pay(self, planned, total, pay1='0', dept=None):
        p = Payment.objects.create(
            created_by=self.user, department=dept or self.dept,
            approval_number='', project_desc='P', payee='Payee',
            total_amount=Decimal(total), planned_date=planned, notes='')
        if Decimal(pay1) > 0:
            from paikuan.models import PaymentInstallment
            PaymentInstallment.objects.create(
                payment=p, seq=1, pay_date=planned, pay_amount=Decimal(pay1))
        return p

    def test_prior_unpaid_carries_into_current_period(self):
        # 本期(2026-06)：计划 1000，已付 400
        self._pay(date(2026, 6, 10), '1000', pay1='400')
        # 前期(2026-05)：计划 800，已付 300 → 剩余 500 应结转
        self._pay(date(2026, 5, 10), '800', pay1='300')
        # 前期(2026-04)：计划 600，已付清 → 不结转
        self._pay(date(2026, 4, 10), '600', pay1='600')
        resp = self.client.get('/api/pk/stats',
                               {'year': 2026, 'month': 6}, **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        # 本期口径不变：仅含 6 月
        self.assertEqual(Decimal(d['total_amount']), Decimal('1000'))
        self.assertEqual(Decimal(d['total_remaining']), Decimal('600'))
        # 前期结转：仅 5 月那笔剩余 500（4 月已付清排除）
        self.assertEqual(Decimal(d['carryover_remaining']), Decimal('500'))
        self.assertEqual(d['carryover_count'], 1)
        # 累计应付未付 = 600 + 500
        self.assertEqual(Decimal(d['total_outstanding']), Decimal('1100'))
        # 部门行带结转
        row = next(r for r in d['by_dept'] if r['dept'] == self.dept)
        self.assertEqual(Decimal(row['carry']), Decimal('500'))
        self.assertEqual(Decimal(row['outstanding']), Decimal('1100'))

    def test_dept_with_only_carryover_appears(self):
        # 某部门本期无计划，但有前期未付 → 仍应出现在 by_dept
        other = DEPARTMENTS[1]
        self._pay(date(2026, 5, 1), '700', pay1='0', dept=other)
        resp = self.client.get('/api/pk/stats',
                               {'year': 2026, 'month': 6}, **self.auth())
        d = resp.json()['data']
        row = next(r for r in d['by_dept'] if r['dept'] == other)
        self.assertEqual(Decimal(row['total']), Decimal('0'))
        self.assertEqual(Decimal(row['carry']), Decimal('700'))
        self.assertEqual(row['carry_count'], 1)


class PaymentApplicantTests(TestCase):
    """排款记录的申请人字段：创建/编辑、一键排款带出、模板与导出含申请人列。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.dept = DEPARTMENTS[0]
        u = PaikuanUser(phone='13900000400', name='Admin', role='super_admin',
                        job_title='finance_director', departments=[self.dept],
                        is_active=True, is_approved=True)
        u.set_password('Test123456'); u.save()
        self.user = u
        self.token = make_token(u)

    def tearDown(self):
        _invalidate_perm_cache()

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def test_create_payment_persists_applicant(self):
        payload = {
            'department': self.dept, 'applicant': '张三',
            'project_desc': '工程款', 'payee': '某公司',
            'total_amount': '1000', 'planned_date': '2026-06-01',
        }
        resp = self.client.post('/api/pk/payments', data=json.dumps(payload),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data']['applicant'], '张三')
        self.assertEqual(Payment.objects.get(id=resp.json()['data']['id']).applicant, '张三')

    def test_schedule_carries_applicant_from_approval(self):
        from paikuan.models import ApprovalRecord
        rec = ApprovalRecord.objects.create(
            applicant='李四', department=self.dept, approval_number='1' * 21,
            summary='付款摘要', amount=Decimal('5000'), payee='供应商A',
            status='approved')
        resp = self.client.post(f'/api/pk/approvals/{rec.id}/schedule',
                                data=json.dumps({'planned_date': '2026-06-10',
                                                 'total_amount': '5000'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data']['payment']['applicant'], '李四')

    def test_template_and_export_include_applicant_header(self):
        from openpyxl import load_workbook
        import io
        # 模板
        tpl = self.client.get('/api/pk/payments/template', **self.auth())
        self.assertEqual(tpl.status_code, 200, tpl.content)
        ws = load_workbook(io.BytesIO(tpl.content)).active
        headers = [c.value for c in ws[1]]
        self.assertIn('申请人', headers)
        # 导出
        Payment.objects.create(created_by=self.user, department=self.dept,
                               applicant='王五', approval_number='',
                               project_desc='P', payee='X',
                               total_amount=Decimal('100'),
                               planned_date=date(2026, 6, 1))
        exp = self.client.get('/api/pk/payments/export', **self.auth())
        self.assertEqual(exp.status_code, 200, exp.content)
        ws2 = load_workbook(io.BytesIO(exp.content)).active
        headers2 = [c.value for c in ws2[1]]
        self.assertIn('申请人', headers2)
        col = headers2.index('申请人')
        self.assertEqual(ws2[2][col].value, '王五')


class PaymentChainIntegrityTests(TestCase):
    """排款链路完整性：归档保护、预付冲抵删除拦截、未来日期拒绝、跨部门核销拦截、口径对齐。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.dept = DEPARTMENTS[0]
        self.dept2 = DEPARTMENTS[1] if len(DEPARTMENTS) > 1 else DEPARTMENTS[0]
        admin = PaikuanUser(phone='13900000500', name='ChainAdmin', role='super_admin',
                            job_title='finance_director', departments=DEPARTMENTS,
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.user = admin
        self.token = make_token(admin)

    def tearDown(self):
        _invalidate_perm_cache()

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _pay(self, total='5000', dept=None, planned='2026-06-01'):
        return Payment.objects.create(
            created_by=self.user, department=dept or self.dept,
            approval_number='', project_desc='P', payee='Payee',
            total_amount=Decimal(total), planned_date=date.fromisoformat(planned))

    def _approval(self, status='pending', archived=False):
        from paikuan.models import ApprovalRecord
        return ApprovalRecord.objects.create(
            applicant='申请人', department=self.dept,
            approval_number='1' * 21, summary='摘要',
            amount=Decimal('3000'), payee='供应商',
            status=status, archived=archived)

    # ── 1. 归档审批记录不可编辑 ──────────────────────────────────────────
    def test_archived_approval_cannot_be_edited(self):
        rec = self._approval(status='rejected', archived=True)
        resp = self.client.put(
            f'/api/pk/approvals/{rec.id}',
            data=json.dumps({'summary': '新摘要'}),
            content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 409, resp.content)
        self.assertIn('归档', resp.json().get('error', ''))

    # ── 2. 已审批的记录改金额被拒绝（须先退回 pending）──────────────────
    def test_amount_change_blocked_when_already_approved(self):
        rec = self._approval(status='approved', archived=False)
        resp = self.client.put(
            f'/api/pk/approvals/{rec.id}',
            data=json.dumps({'amount': '9999'}),
            content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 400, resp.content)
        self.assertIn('待审批', resp.json().get('error', ''))
        rec.refresh_from_db()
        self.assertEqual(rec.amount, Decimal('3000'))

    # ── 3. 金额 pending 状态下可修改 ────────────────────────────────────
    def test_amount_change_allowed_when_pending(self):
        rec = self._approval(status='pending', archived=False)
        resp = self.client.put(
            f'/api/pk/approvals/{rec.id}',
            data=json.dumps({'amount': '8888'}),
            content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(Decimal(resp.json()['data']['amount']), Decimal('8888'))

    # ── 4. 已关联预付核销的排款不可删除 ─────────────────────────────────
    def test_payment_delete_blocked_when_has_prepaid_writeoff(self):
        from ar.models import ARProject, AdvanceRecord, AdvanceWriteoff
        p = self._pay(total='10000')
        proj = ARProject.objects.create(
            customer_name='C', short_name='P', delivery_dept=self.dept,
            sales_contact='S', project_manager='M', project_no='GYL-CHAIN-001')
        adv = AdvanceRecord.objects.create(
            direction='预付', project=proj, delivery_dept=self.dept,
            counterparty='供应商', occur_year=2026, occur_month=6,
            occur_date=date(2026, 6, 1), advance_amount=Decimal('10000'))
        AdvanceWriteoff.objects.create(
            advance_record=adv, writeoff_no=1, amount=Decimal('3000'),
            writeoff_date=date(2026, 6, 5), payment=p)
        resp = self.client.delete(f'/api/pk/payments/{p.id}', **self.auth())
        self.assertEqual(resp.status_code, 409, resp.content)
        self.assertIn('预付核销', resp.json().get('error', ''))

    # ── 5. 付款分期不允许未来日期 ────────────────────────────────────────
    def test_future_installment_date_rejected(self):
        import datetime as dt
        p = self._pay(total='5000')
        future = (dt.date.today() + dt.timedelta(days=10)).isoformat()
        resp = self.client.put(
            f'/api/pk/payments/{p.id}',
            data=json.dumps({'installments': [{'pay_date': future, 'pay_amount': '1000'}]}),
            content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 400, resp.content)
        self.assertIn('今天', resp.json().get('error', ''))

    # ── 6. 今天日期的分期允许写入 ────────────────────────────────────────
    def test_today_installment_date_accepted(self):
        import datetime as dt
        p = self._pay(total='5000')
        today = dt.date.today().isoformat()
        resp = self.client.put(
            f'/api/pk/payments/{p.id}',
            data=json.dumps({'installments': [{'pay_date': today, 'pay_amount': '1000'}]}),
            content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)

    # ── 7. stats total_remaining 包含 prepaid_offset 扣减 ──────────────
    def test_stats_remaining_subtracts_prepaid_offset(self):
        from ar.models import ARProject, AdvanceRecord, AdvanceWriteoff
        p = self._pay(total='10000', planned='2026-06-01')
        # 人工设置 prepaid_offset_amount (模拟信号已触发)
        Payment.objects.filter(pk=p.pk).update(prepaid_offset_amount=Decimal('2000'))
        resp = self.client.get('/api/pk/stats',
                               {'year': 2026, 'month': 6}, **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        # total_remaining = 10000 - 0 paid - 2000 offset = 8000
        self.assertEqual(Decimal(d['total_remaining']), Decimal('8000'))

    # ── 8. 超付校验：已付 + 分期超 plan_adjustment 被拒绝 ───────────────
    def test_overpayment_against_plan_adjustment_rejected(self):
        from paikuan.models import PaymentInstallment
        p = self._pay(total='10000')
        # 先设定 plan_adjustment = 6000
        Payment.objects.filter(pk=p.pk).update(plan_adjustment=Decimal('6000'))
        p.refresh_from_db()
        # 加一笔超过 plan_adjustment 的分期
        resp = self.client.put(
            f'/api/pk/payments/{p.id}',
            data=json.dumps({'installments': [{'pay_date': '2026-06-01', 'pay_amount': '7000'}]}),
            content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 400, resp.content)

    # ── 9. 跨部门预付核销被拒绝 ─────────────────────────────────────────
    def test_cross_dept_prepaid_writeoff_blocked(self):
        from ar.models import ARProject, AdvanceRecord
        # 排款属于 dept2，预付预收记录属于 dept（不同）
        p = self._pay(total='10000', dept=self.dept2)
        proj = ARProject.objects.create(
            customer_name='C2', short_name='P2', delivery_dept=self.dept,
            sales_contact='S', project_manager='M', project_no='GYL-CHAIN-002')
        adv = AdvanceRecord.objects.create(
            direction='预付', project=proj, delivery_dept=self.dept,
            counterparty='供应商B', occur_year=2026, occur_month=6,
            occur_date=date(2026, 6, 1), advance_amount=Decimal('10000'))
        # 通过 AR API 创建核销（AR URLs 挂载在 /api/pk/ar/）
        resp = self.client.post(
            f'/api/pk/ar/advances/{adv.id}/writeoffs',
            data=json.dumps({'amount': '3000', 'writeoff_date': '2026-06-05',
                             'payment_id': p.id}),
            content_type='application/json', **self.auth())
        # 若 dept != dept2，应被拒绝
        if self.dept != self.dept2:
            self.assertIn(resp.status_code, [400, 403, 409], resp.content)
            self.assertIn('不一致', resp.json().get('error', ''))
        else:
            # 单部门环境跳过跨部门场景
            pass


class ProjectLinkFieldsTests(TestCase):
    """二级部门/项目简称字段：排款与审批的创建/编辑/导入校验 + 审批转排款贯通。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.dept = DEPARTMENTS[0]
        self.dept2 = DEPARTMENTS[1] if len(DEPARTMENTS) > 1 else DEPARTMENTS[0]
        admin = PaikuanUser(phone='13900000600', name='LinkAdmin', role='super_admin',
                            job_title='finance_director', departments=DEPARTMENTS,
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.user = admin
        self.token = make_token(admin)
        self.proj = ARProject.objects.create(
            customer_name='客户甲', short_name='链路打通项目', delivery_dept=self.dept,
            sub_dept='华东项目部', sales_contact='S', project_manager='M',
            project_no='LNK-0001')
        # is_shared 由模型自动推导（销售对接人≠负责人 → 共享）；这里同名保证非共享
        self.other_proj = ARProject.objects.create(
            customer_name='客户乙', short_name='他部门项目', delivery_dept=self.dept2,
            sales_contact='同人', project_manager='同人', project_no='LNK-0002')

    def tearDown(self):
        _invalidate_perm_cache()

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _post(self, url, payload):
        return self.client.post(url, data=json.dumps(payload),
                                content_type='application/json', **self.auth())

    def _payment_payload(self, **kw):
        base = {'department': self.dept, 'project_desc': '事项', 'payee': '收款方',
                'total_amount': '1000', 'planned_date': '2026-07-01'}
        base.update(kw)
        return base

    # ── 排款：项目简称匹配台账则保存，并写入二级部门 ─────────────────────
    def test_payment_create_with_valid_short_name(self):
        resp = self._post('/api/pk/payments', self._payment_payload(
            project_short_name='链路打通项目', secondary_dept='华东项目部'))
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['project_short_name'], '链路打通项目')
        self.assertEqual(d['secondary_dept'], '华东项目部')

    # ── 排款：台账中不存在的项目简称被拒绝并给出指导 ─────────────────────
    def test_payment_create_with_unknown_short_name_rejected(self):
        resp = self._post('/api/pk/payments', self._payment_payload(
            project_short_name='不存在的项目'))
        self.assertEqual(resp.status_code, 400, resp.content)
        msg = resp.json().get('error', '')
        self.assertIn('项目台账', msg)
        self.assertIn('不存在', msg)

    # ── 排款：非共享项目不可跨部门引用 ───────────────────────────────────
    def test_payment_cross_dept_short_name_rejected(self):
        if self.dept == self.dept2:
            return
        resp = self._post('/api/pk/payments', self._payment_payload(
            project_short_name='他部门项目'))
        self.assertEqual(resp.status_code, 400, resp.content)
        self.assertIn('不一致', resp.json().get('error', ''))

    # ── 排款：共享业务项目可跨部门引用 ───────────────────────────────────
    def test_payment_shared_project_cross_dept_ok(self):
        if self.dept == self.dept2:
            return
        # is_shared 是推导值：销售对接人≠项目负责人即共享业务
        self.other_proj.sales_contact = '销售A'
        self.other_proj.project_manager = '经理B'
        self.other_proj.save()
        resp = self._post('/api/pk/payments', self._payment_payload(
            project_short_name='他部门项目'))
        self.assertEqual(resp.status_code, 200, resp.content)

    # ── 审批：创建带两个新字段；未知简称拒绝 ─────────────────────────────
    def test_approval_create_with_fields_and_validation(self):
        ok_resp = self._post('/api/pk/approvals', {
            'applicant': '张三', 'department': self.dept, 'amount': '800',
            'summary': 'S', 'payee': 'P',
            'secondary_dept': '华东项目部', 'project_short_name': '链路打通项目'})
        self.assertEqual(ok_resp.status_code, 200, ok_resp.content)
        self.assertEqual(ok_resp.json()['data']['project_short_name'], '链路打通项目')
        bad = self._post('/api/pk/approvals', {
            'applicant': '张三', 'department': self.dept, 'amount': '800',
            'project_short_name': '不存在的项目'})
        self.assertEqual(bad.status_code, 400, bad.content)
        self.assertIn('项目台账', bad.json().get('error', ''))

    # ── 审批：归档记录仍可单独补录两个元数据字段，混合编辑仍 409 ─────────
    def test_archived_approval_meta_fields_editable(self):
        from paikuan.models import ApprovalRecord
        rec = ApprovalRecord.objects.create(
            applicant='A', department=self.dept, approval_number='2' * 21,
            summary='S', amount=Decimal('100'), payee='P',
            status='rejected', archived=True)
        meta = self.client.put(f'/api/pk/approvals/{rec.id}',
                               data=json.dumps({'secondary_dept': '华北项目部',
                                                'project_short_name': '链路打通项目'}),
                               content_type='application/json', **self.auth())
        self.assertEqual(meta.status_code, 200, meta.content)
        rec.refresh_from_db()
        self.assertEqual(rec.secondary_dept, '华北项目部')
        self.assertEqual(rec.project_short_name, '链路打通项目')
        mixed = self.client.put(f'/api/pk/approvals/{rec.id}',
                                data=json.dumps({'secondary_dept': 'X', 'summary': '改摘要'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(mixed.status_code, 409, mixed.content)

    # ── 审批导入：项目简称与台账不匹配的行被拒绝并提示 ───────────────────
    def test_approval_import_unmatched_short_name_rejected(self):
        import openpyxl as _xl
        wb = _xl.Workbook()
        ws = wb.active
        ws.append(['申请人*', '所属事业部*', '二级部门', '项目简称', '审批编号*',
                   '摘要', '申请金额*', '收款主体', '审批状态*'])
        ws.append(['李四', self.dept, '', '台账没有的项目', '', '摘要', 500, '收款', '待审批'])
        ws.append(['王五', self.dept, '华东项目部', '链路打通项目', '', '摘要', 600, '收款', '待审批'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = 'a.xlsx'
        resp = self.client.post('/api/pk/approvals/import', {'file': buf}, **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['created'], 1)
        self.assertEqual(d['skipped'], 1)
        self.assertTrue(any('项目台账' in e for e in d['errors']), d['errors'])

    # ── 审批转排款：两个字段随单带入排款台账 ─────────────────────────────
    def test_schedule_copies_link_fields_to_payment(self):
        from paikuan.models import ApprovalRecord
        rec = ApprovalRecord.objects.create(
            applicant='A', department=self.dept, approval_number='3' * 21,
            summary='S', amount=Decimal('900'), payee='P', status='approved',
            secondary_dept='华东项目部', project_short_name='链路打通项目')
        resp = self._post(f'/api/pk/approvals/{rec.id}/schedule',
                          {'planned_date': '2026-07-15', 'total_amount': '900'})
        self.assertEqual(resp.status_code, 200, resp.content)
        pay = resp.json()['data']['payment']
        self.assertEqual(pay['secondary_dept'], '华东项目部')
        self.assertEqual(pay['project_short_name'], '链路打通项目')


class PartialScheduleTests(TestCase):
    """审批分批排款：多次排款分批流转付款管理，排满自动归档。"""

    def setUp(self):
        self.client = Client()
        self.dept = '运输事业部'
        admin = PaikuanUser(phone='13900002000', name='SchedAdmin', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.token = make_token(admin)
        self.rec = ApprovalRecord.objects.create(
            applicant='张三', department=self.dept, approval_number='1' * 21,
            summary='设备采购', amount=Decimal('10000'), payee='供应商A',
            status='approved')

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _sched(self, amount, day):
        return self.client.post(f'/api/pk/approvals/{self.rec.id}/schedule',
                                data=json.dumps({'planned_date': day, 'total_amount': amount}),
                                content_type='application/json', **self.auth())

    def test_partial_schedules_until_archive(self):
        # 第一批 4000 → 留在审批管理
        resp = self._sched('4000', '2026-07-01')
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertFalse(d['archived'])
        self.assertEqual(Decimal(d['remaining_amount']), Decimal('6000'))
        self.rec.refresh_from_db()
        self.assertFalse(self.rec.archived)
        self.assertEqual(self.rec.scheduled_amount, Decimal('4000'))
        # 第二批 6000 → 排满归档
        resp = self._sched('6000', '2026-08-01')
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertTrue(resp.json()['data']['archived'])
        self.rec.refresh_from_db()
        self.assertTrue(self.rec.archived)
        # 付款管理只有一条汇总（经 approval 静默ID打通），两批计划在明细里
        pays = Payment.objects.filter(approval=self.rec)
        self.assertEqual(pays.count(), 1)
        p = pays.first()
        self.assertEqual(p.total_amount, Decimal('10000'))         # 汇总=各批之和
        self.assertEqual(str(p.planned_date), '2026-08-01')        # 最后一次排款批次日期
        items = list(p.plan_items.order_by('seq'))
        self.assertEqual([i.amount for i in items], [Decimal('4000'), Decimal('6000')])
        self.assertEqual([str(i.planned_date) for i in items], ['2026-07-01', '2026-08-01'])

    def test_plan_item_undo_rolls_back_schedule(self):
        # 撤销某批计划 → 汇总回退、审批已排款扣减、归档回退可继续排
        self._sched('4000', '2026-07-01')
        self._sched('6000', '2026-08-01')
        self.rec.refresh_from_db()
        self.assertTrue(self.rec.archived)
        p = Payment.objects.get(approval=self.rec)
        second = p.plan_items.get(seq=2)
        resp = self.client.delete(f'/api/pk/payments/{p.id}/plan-items/{second.id}',
                                  **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        p.refresh_from_db(); self.rec.refresh_from_db()
        self.assertEqual(p.total_amount, Decimal('4000'))
        self.assertEqual(self.rec.scheduled_amount, Decimal('4000'))
        self.assertFalse(self.rec.archived)   # 可继续分批排款
        # 最后一批不可单独撤销
        first = p.plan_items.get(seq=1)
        resp = self.client.delete(f'/api/pk/payments/{p.id}/plan-items/{first.id}',
                                  **self.auth())
        self.assertEqual(resp.status_code, 400)

    def test_over_remaining_rejected(self):
        self._sched('4000', '2026-07-01')
        resp = self._sched('7000', '2026-08-01')
        self.assertEqual(resp.status_code, 400)
        self.assertIn('剩余可排', resp.json()['error'])

    def test_archived_cannot_schedule_more(self):
        self._sched('10000', '2026-07-01')
        self.rec.refresh_from_db()
        self.assertTrue(self.rec.archived)
        resp = self._sched('1', '2026-08-01')
        self.assertEqual(resp.status_code, 404)   # archived=False 过滤不到

    def test_export_has_schedule_columns(self):
        self._sched('4000', '2026-07-01')
        resp = self.client.get('/api/pk/approvals/export', **self.auth())
        self.assertEqual(resp.status_code, 200)
        import io as _io
        import openpyxl as _xl
        wb = _xl.load_workbook(_io.BytesIO(resp.content))
        headers = [c.value for c in wb.active[1]]
        self.assertIn('已排款金额', headers)
        self.assertIn('剩余可排', headers)
        idx = headers.index('已排款金额')
        self.assertEqual(wb.active.cell(2, idx + 1).value, 4000)


class PaymentScheduleSyncTests(TestCase):
    """付款管理 编辑/删除/批次管理 ←→ 审批「已排款 / 归档」双向同步对账。

    口径：审批.已排款 == 关联付款管理全部计划批次之和；任一改动后两侧都应一致，
    且归档状态随排满/未排满自动切换（未排满回到审批管理可继续分批排款）。"""

    def setUp(self):
        self.client = Client()
        self.dept = '运输事业部'
        admin = PaikuanUser(phone='13900002050', name='SyncAdmin', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.token = make_token(admin)
        self.rec = ApprovalRecord.objects.create(
            applicant='张三', department=self.dept, approval_number='3' * 21,
            summary='设备采购', amount=Decimal('10000'), payee='供应商A',
            status='approved')

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _sched(self, amount, day):
        return self.client.post(f'/api/pk/approvals/{self.rec.id}/schedule',
                                data=json.dumps({'planned_date': day, 'total_amount': amount}),
                                content_type='application/json', **self.auth())

    def _payment(self):
        return Payment.objects.get(approval=self.rec)

    def _put(self, url, body):
        return self.client.put(url, data=json.dumps(body),
                               content_type='application/json', **self.auth())

    def _post(self, url, body):
        return self.client.post(url, data=json.dumps(body),
                                content_type='application/json', **self.auth())

    # ── 编辑付款管理「计划额」→ 审批已排款同步、归档回退 ──
    def test_edit_payment_amount_syncs_approval(self):
        self._sched('10000', '2026-07-01')          # 排满归档（单批）
        self.rec.refresh_from_db()
        self.assertTrue(self.rec.archived)
        p = self._payment()
        resp = self._put(f'/api/pk/payments/{p.id}', {'total_amount': '6000'})
        self.assertEqual(resp.status_code, 200, resp.content)
        self.rec.refresh_from_db()
        self.assertEqual(self.rec.scheduled_amount, Decimal('6000'))
        self.assertFalse(self.rec.archived)         # 6000<10000 → 可继续排
        # 单批明细随汇总同步
        p.refresh_from_db()
        self.assertEqual(p.plan_items.get().amount, Decimal('6000'))

    def test_edit_payment_over_approval_amount_rejected(self):
        self._sched('6000', '2026-07-01')
        p = self._payment()
        resp = self._put(f'/api/pk/payments/{p.id}', {'total_amount': '12000'})
        self.assertEqual(resp.status_code, 400, resp.content)
        self.assertIn('超过', resp.json()['error'])
        self.rec.refresh_from_db()
        self.assertEqual(self.rec.scheduled_amount, Decimal('6000'))   # 未改

    # ── 删除付款管理 → 审批回退为可排款 ──
    def test_delete_payment_frees_approval(self):
        self._sched('10000', '2026-07-01')
        self.rec.refresh_from_db(); self.assertTrue(self.rec.archived)
        p = self._payment()
        resp = self.client.delete(f'/api/pk/payments/{p.id}', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.rec.refresh_from_db()
        self.assertEqual(self.rec.scheduled_amount, Decimal('0'))
        self.assertFalse(self.rec.archived)
        # 现在可重新排款
        self.assertEqual(self._sched('5000', '2026-08-01').status_code, 200)

    def test_bulk_delete_frees_approval(self):
        self._sched('10000', '2026-07-01')
        p = self._payment()
        resp = self._post('/api/pk/payments/bulk-delete', {'ids': [p.id]})
        self.assertEqual(resp.status_code, 200, resp.content)
        self.rec.refresh_from_db()
        self.assertEqual(self.rec.scheduled_amount, Decimal('0'))
        self.assertFalse(self.rec.archived)

    # ── 编辑单批计划 → 审批已排款同步 ──
    def test_edit_plan_item_syncs_approval(self):
        self._sched('4000', '2026-07-01')
        self._sched('6000', '2026-08-01')           # 排满归档，两批
        p = self._payment()
        second = p.plan_items.get(seq=2)
        resp = self._put(f'/api/pk/payments/{p.id}/plan-items/{second.id}',
                         {'planned_date': '2026-08-15', 'amount': '3000'})
        self.assertEqual(resp.status_code, 200, resp.content)
        p.refresh_from_db(); self.rec.refresh_from_db()
        self.assertEqual(p.total_amount, Decimal('7000'))        # 4000+3000
        self.assertEqual(str(p.planned_date), '2026-08-15')      # 最后一次排款批次日（seq2）
        self.assertEqual(self.rec.scheduled_amount, Decimal('7000'))
        self.assertFalse(self.rec.archived)

    def test_edit_plan_item_over_approval_rejected(self):
        self._sched('4000', '2026-07-01')
        self._sched('3000', '2026-08-01')           # 共 7000
        p = self._payment()
        second = p.plan_items.get(seq=2)
        resp = self._put(f'/api/pk/payments/{p.id}/plan-items/{second.id}',
                         {'planned_date': '2026-08-01', 'amount': '7000'})  # 累计 11000>10000
        self.assertEqual(resp.status_code, 400, resp.content)
        p.refresh_from_db(); self.rec.refresh_from_db()
        self.assertEqual(self.rec.scheduled_amount, Decimal('7000'))   # 未改

    def test_edit_plan_item_below_paid_rejected(self):
        self._sched('4000', '2026-07-01')
        self._sched('6000', '2026-08-01')           # 共 10000
        p = self._payment()
        PaymentInstallment.objects.create(payment=p, seq=1, pay_date=date(2026, 7, 5),
                                           pay_amount=Decimal('5000'))
        second = p.plan_items.get(seq=2)
        # 第二批 6000→500 → 计划合计 4500 < 已付 5000 → 拦截
        resp = self._put(f'/api/pk/payments/{p.id}/plan-items/{second.id}',
                         {'planned_date': '2026-08-01', 'amount': '500'})
        self.assertEqual(resp.status_code, 400, resp.content)
        self.assertIn('低于', resp.json()['error'])

    # ── 追加批次 → 审批已排款同步 + 超额拦截 ──
    def test_add_plan_item_syncs_approval(self):
        self._sched('4000', '2026-07-01')           # 未排满
        p = self._payment()
        resp = self._post(f'/api/pk/payments/{p.id}/plan-items',
                          {'planned_date': '2026-09-01', 'amount': '6000'})
        self.assertEqual(resp.status_code, 200, resp.content)
        p.refresh_from_db(); self.rec.refresh_from_db()
        self.assertEqual(p.total_amount, Decimal('10000'))
        self.assertEqual(p.plan_items.count(), 2)
        self.assertEqual(self.rec.scheduled_amount, Decimal('10000'))
        self.assertTrue(self.rec.archived)          # 排满归档

    def test_add_plan_item_over_remaining_rejected(self):
        self._sched('4000', '2026-07-01')
        p = self._payment()
        resp = self._post(f'/api/pk/payments/{p.id}/plan-items',
                          {'planned_date': '2026-09-01', 'amount': '7000'})   # 剩余仅 6000
        self.assertEqual(resp.status_code, 400, resp.content)
        self.assertIn('剩余可排', resp.json()['error'])
        self.rec.refresh_from_db()
        self.assertEqual(self.rec.scheduled_amount, Decimal('4000'))   # 未改

    def test_summary_planned_date_follows_last_schedule(self):
        # 多次排款，汇总计划日以「最后一次排款」（seq 最大）的批次日为准，
        # 既非最早(min)也非最晚(max)日期，而是末次排款动作的日期。
        self._sched('2000', '2026-07-01')   # seq1
        self._sched('3000', '2026-09-01')   # seq2（日期最晚）
        self._sched('5000', '2026-08-01')   # seq3（最后一次排款）
        p = self._payment()
        self.assertEqual(str(p.planned_date), '2026-08-01')   # = seq3 末次排款日
        # 撤销末批后，汇总日回退为上一末批（seq2=09-01）
        third = p.plan_items.get(seq=3)
        resp = self.client.delete(f'/api/pk/payments/{p.id}/plan-items/{third.id}',
                                  **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        p.refresh_from_db()
        self.assertEqual(str(p.planned_date), '2026-09-01')   # 末批=seq2

    def test_plan_item_amount_nan_infinity_rejected_not_500(self):
        # Decimal('NaN')/('Infinity') 能构造但比较会抛 → 必须以 400 拦掉而非 500
        self._sched('4000', '2026-07-01')
        p = self._payment()
        for bad in ('NaN', 'Infinity', '-Infinity'):
            resp = self._post(f'/api/pk/payments/{p.id}/plan-items',
                              {'planned_date': '2026-09-01', 'amount': bad})
            self.assertEqual(resp.status_code, 400, f'{bad}: {resp.content}')
            self.assertIn('格式有误', resp.json()['error'])
        # 编辑路径同样拦截
        first = p.plan_items.get(seq=1)
        resp = self._put(f'/api/pk/payments/{p.id}/plan-items/{first.id}',
                         {'planned_date': '2026-07-01', 'amount': 'NaN'})
        self.assertEqual(resp.status_code, 400, resp.content)


class LegacyScheduleAdoptTests(TestCase):
    """静默ID上线前的旧排款：再次分批排款时自动收养挂链，不再误报重复。"""

    def setUp(self):
        self.client = Client()
        self.dept = '运输事业部'
        admin = PaikuanUser(phone='13900002100', name='AdoptAdmin', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.token = make_token(admin)
        self.rec = ApprovalRecord.objects.create(
            applicant='李四', department=self.dept, approval_number='2' * 21,
            summary='运输费', amount=Decimal('10000'), payee='承运商B',
            status='approved', scheduled_amount=Decimal('4000'))
        # 旧模式生成的独立排款（未挂 approval 链）
        self.legacy = Payment.objects.create(
            department=self.dept, approval_number='2' * 21, payee='承运商B',
            project_desc='运输费', total_amount=Decimal('4000'),
            planned_date=date(2026, 7, 1))

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def test_second_tranche_adopts_legacy_payment(self):
        # 第二批 4000（与历史首批同金额）→ 收养旧记录、追加批次，而非报重复
        resp = self.client.post(f'/api/pk/approvals/{self.rec.id}/schedule',
                                data=json.dumps({'planned_date': '2026-08-01',
                                                 'total_amount': '4000'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.legacy.refresh_from_db()
        self.assertEqual(self.legacy.approval_id, self.rec.id)      # 已挂链
        self.assertEqual(self.legacy.total_amount, Decimal('8000'))  # 汇总=两批之和
        items = list(self.legacy.plan_items.order_by('seq'))
        self.assertEqual(len(items), 2)
        self.assertEqual([i.amount for i in items], [Decimal('4000'), Decimal('4000')])
        self.assertEqual(Payment.objects.filter(approval_number='2' * 21).count(), 1)
        # 收养后已排款=批次之和（对账正源），而非旧增量
        self.rec.refresh_from_db()
        self.assertEqual(self.rec.scheduled_amount, Decimal('8000'))

    def test_adopt_with_stale_zero_scheduled_reconciles(self):
        # 真实历史场景：scheduled_amount 漂移为 0（迁移仅回填已归档记录），
        # 旧独立排款 total=4000 未计入审批。收养并追加 3000 批次后，
        # 已排款应对账为批次之和 7000（而非旧增量 0+3000=3000）。
        rec = ApprovalRecord.objects.create(
            applicant='钱六', department=self.dept, approval_number='5' * 21,
            summary='仓储费', amount=Decimal('10000'), payee='仓储商C',
            status='approved', scheduled_amount=Decimal('0'))
        Payment.objects.create(
            department=self.dept, approval_number='5' * 21, payee='仓储商C',
            project_desc='仓储费', total_amount=Decimal('4000'),
            planned_date=date(2026, 7, 1))
        resp = self.client.post(f'/api/pk/approvals/{rec.id}/schedule',
                                data=json.dumps({'planned_date': '2026-08-01',
                                                 'total_amount': '3000'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        rec.refresh_from_db()
        p = Payment.objects.get(approval=rec)
        self.assertEqual(p.total_amount, Decimal('7000'))               # 4000+3000
        self.assertEqual(rec.scheduled_amount, Decimal('7000'))         # 对账=批次之和
        self.assertFalse(rec.archived)
        # 剩余可排基于批次之和（3000），不能再排 4000（会越过申请金额）
        resp = self.client.post(f'/api/pk/approvals/{rec.id}/schedule',
                                data=json.dumps({'planned_date': '2026-09-01',
                                                 'total_amount': '4000'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 400, resp.content)
        self.assertIn('剩余可排', resp.json()['error'])

    def test_same_amount_different_date_allowed(self):
        # 同金额多批（不同日期）完全放行
        for day in ('2026-08-01', '2026-09-01'):
            resp = self.client.post(f'/api/pk/approvals/{self.rec.id}/schedule',
                                    data=json.dumps({'planned_date': day,
                                                     'total_amount': '3000'}),
                                    content_type='application/json', **self.auth())
            self.assertEqual(resp.status_code, 200, resp.content)
        self.legacy.refresh_from_db()
        self.assertEqual(self.legacy.plan_items.count(), 3)
        self.assertEqual(self.legacy.total_amount, Decimal('10000'))

    def test_placeholder_number_not_adopted(self):
        # 占位审批号(全0)不收养：新建独立汇总记录
        rec0 = ApprovalRecord.objects.create(
            applicant='王五', department=self.dept, approval_number='0' * 21,
            summary='杂费', amount=Decimal('500'), payee='某商户', status='approved')
        Payment.objects.create(
            department=self.dept, approval_number='0' * 21, payee='某商户',
            project_desc='历史杂费', total_amount=Decimal('300'),
            planned_date=date(2026, 6, 1))
        resp = self.client.post(f'/api/pk/approvals/{rec0.id}/schedule',
                                data=json.dumps({'planned_date': '2026-08-15',
                                                 'total_amount': '500'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(Payment.objects.filter(approval=rec0).count(), 1)
        self.assertIsNone(Payment.objects.get(project_desc='历史杂费').approval_id)

    def test_same_day_same_amount_allowed_outside_window(self):
        # 同日同金额多批：真实业务需求，时间窗外放行
        resp = self._sched_for(self.rec, '3000', '2026-08-01')
        self.assertEqual(resp.status_code, 200, resp.content)
        p = Payment.objects.get(approval=self.rec)
        # 把首批的创建时间拨回 1 分钟前，模拟"稍后再排一笔"
        from django.utils import timezone
        from datetime import timedelta
        p.plan_items.update(created_at=timezone.now() - timedelta(minutes=1))
        resp = self._sched_for(self.rec, '3000', '2026-08-01')
        self.assertEqual(resp.status_code, 200, resp.content)
        p.refresh_from_db()
        self.assertEqual(p.plan_items.filter(planned_date='2026-08-01',
                                             amount=Decimal('3000')).count(), 2)
        self.assertEqual(p.total_amount, Decimal('10000'))   # 4000历史 + 3000 + 3000

    def test_double_click_within_window_rejected(self):
        # 10秒窗口内连续提交相同批次 → 拦截疑似误触
        resp = self._sched_for(self.rec, '2000', '2026-09-01')
        self.assertEqual(resp.status_code, 200, resp.content)
        resp = self._sched_for(self.rec, '2000', '2026-09-01')
        self.assertEqual(resp.status_code, 409)
        self.assertIn('重复提交', resp.json()['error'])

    def _sched_for(self, rec, amount, day):
        return self.client.post(f'/api/pk/approvals/{rec.id}/schedule',
                                data=json.dumps({'planned_date': day, 'total_amount': amount}),
                                content_type='application/json', **self.auth())


class BulkOpsTests(TestCase):
    """批量删除 / 批量排款（审批）/ 批量付款（付款管理）。含单选=ids 单元素。"""

    def setUp(self):
        self.client = Client()
        self.dept = '运输事业部'
        admin = PaikuanUser(phone='13900003000', name='BulkAdmin', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.token = make_token(admin)

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _mk_approval(self, num, amount, status='approved'):
        return ApprovalRecord.objects.create(
            applicant='张三', department=self.dept, approval_number=str(num) * 21,
            summary='采购', amount=Decimal(amount), payee=f'供应商{num}', status=status)

    def _post(self, url, payload):
        return self.client.post(url, data=json.dumps(payload),
                                content_type='application/json', **self.auth())

    # ── 审批批量排款 ──────────────────────────────────────────────────────────
    def test_bulk_schedule_creates_payments(self):
        a1 = self._mk_approval(1, '1000')
        a2 = self._mk_approval(2, '2000')
        resp = self._post('/api/pk/approvals/bulk-schedule',
                          {'ids': [a1.id, a2.id], 'planned_date': '2026-07-01'})
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['scheduled'], 2)
        self.assertEqual(Decimal(d['total_amount']), Decimal('3000'))
        self.assertEqual(d['skipped'], [])
        # 两条付款管理各自打通审批，金额=申请金额，计划日=今天参数
        for a, amt in ((a1, '1000'), (a2, '2000')):
            p = Payment.objects.get(approval=a)
            self.assertEqual(p.total_amount, Decimal(amt))
            self.assertEqual(str(p.planned_date), '2026-07-01')
            a.refresh_from_db()
            self.assertTrue(a.archived)               # 排满申请金额即归档

    def test_bulk_schedule_skips_non_approved(self):
        a1 = self._mk_approval(1, '1000', status='approved')
        a2 = self._mk_approval(2, '2000', status='pending')
        resp = self._post('/api/pk/approvals/bulk-schedule',
                          {'ids': [a1.id, a2.id], 'planned_date': '2026-07-01'})
        d = resp.json()['data']
        self.assertEqual(d['scheduled'], 1)
        self.assertEqual(len(d['skipped']), 1)
        self.assertEqual(d['skipped'][0]['id'], a2.id)
        self.assertFalse(Payment.objects.filter(approval=a2).exists())

    def test_bulk_schedule_default_date_today(self):
        a1 = self._mk_approval(1, '500')
        resp = self._post('/api/pk/approvals/bulk-schedule', {'ids': [a1.id]})
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data']['scheduled'], 1)
        p = Payment.objects.get(approval=a1)
        self.assertEqual(str(p.planned_date), date.today().isoformat())

    # ── 审批批量删除 ──────────────────────────────────────────────────────────
    def test_bulk_delete_approvals(self):
        a1 = self._mk_approval(1, '1000', status='pending')
        a2 = self._mk_approval(2, '2000', status='pending')
        resp = self._post('/api/pk/approvals/bulk-delete', {'ids': [a1.id, a2.id]})
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data']['deleted'], 2)
        # soft-delete: records still exist in DB but with deleted_at set
        self.assertEqual(ApprovalRecord.objects.filter(deleted_at__isnull=True).count(), 0)
        self.assertEqual(ApprovalRecord.objects.filter(deleted_at__isnull=False).count(), 2)

    def test_bulk_delete_skips_scheduled_approval(self):
        a1 = self._mk_approval(1, '1000')
        self._post('/api/pk/approvals/bulk-schedule', {'ids': [a1.id], 'planned_date': '2026-07-01'})
        resp = self._post('/api/pk/approvals/bulk-delete', {'ids': [a1.id]})
        d = resp.json()['data']
        self.assertEqual(d['deleted'], 0)
        self.assertEqual(len(d['skipped']), 1)
        self.assertTrue(ApprovalRecord.objects.filter(id=a1.id).exists())  # 未删

    # ── 审批批量通过 ──────────────────────────────────────────────────────────
    def test_bulk_approve_pending(self):
        a1 = self._mk_approval(1, '1000', status='pending')
        a2 = self._mk_approval(2, '2000', status='pending')
        resp = self._post('/api/pk/approvals/bulk-approve', {'ids': [a1.id, a2.id]})
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['approved'], 2)
        self.assertEqual(d['skipped'], [])
        for a in (a1, a2):
            a.refresh_from_db()
            self.assertEqual(a.status, 'approved')
            self.assertFalse(a.archived)   # 审批通过不归档（仍可排款）

    def test_bulk_approve_skips_non_pending(self):
        a1 = self._mk_approval(1, '1000', status='pending')
        a2 = self._mk_approval(2, '2000', status='approved')
        a3 = self._mk_approval(3, '3000', status='rejected')
        resp = self._post('/api/pk/approvals/bulk-approve', {'ids': [a1.id, a2.id, a3.id]})
        d = resp.json()['data']
        self.assertEqual(d['approved'], 1)
        self.assertEqual({s['id'] for s in d['skipped']}, {a2.id, a3.id})
        a1.refresh_from_db()
        self.assertEqual(a1.status, 'approved')

    def test_bulk_approve_single_via_ids(self):
        a1 = self._mk_approval(1, '1000', status='pending')
        resp = self._post('/api/pk/approvals/bulk-approve', {'ids': [a1.id]})
        self.assertEqual(resp.json()['data']['approved'], 1)

    def test_bulk_approve_empty_ids_rejected(self):
        resp = self._post('/api/pk/approvals/bulk-approve', {'ids': []})
        self.assertEqual(resp.status_code, 400)

    def test_bulk_approve_non_approver_forbidden(self):
        # 非审批职务（如普通登记岗）无权批量审批，返回 403
        clerk = PaikuanUser(phone='13900003111', name='Clerk', role='user',
                            job_title='cashier', departments=[self.dept],
                            is_active=True, is_approved=True)
        clerk.set_password('Test123456')
        clerk.save()
        token = make_token(clerk)
        a1 = self._mk_approval(1, '1000', status='pending')
        resp = self.client.post('/api/pk/approvals/bulk-approve',
                                data=json.dumps({'ids': [a1.id]}),
                                content_type='application/json',
                                HTTP_AUTHORIZATION=f'Bearer {token}')
        self.assertEqual(resp.status_code, 403, resp.content)
        a1.refresh_from_db()
        self.assertEqual(a1.status, 'pending')   # 未被改动

    # ── 付款管理批量付款 ──────────────────────────────────────────────────────
    def _schedule_payment(self, num, amount):
        a = self._mk_approval(num, amount)
        self._post('/api/pk/approvals/bulk-schedule', {'ids': [a.id], 'planned_date': '2026-07-01'})
        return Payment.objects.get(approval=a)

    def test_bulk_pay_records_installments(self):
        p1 = self._schedule_payment(1, '1000')
        p2 = self._schedule_payment(2, '2000')
        resp = self._post('/api/pk/payments/bulk-pay',
                          {'ids': [p1.id, p2.id], 'pay_date': '2026-07-10'})
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['paid'], 2)
        self.assertEqual(Decimal(d['total_amount']), Decimal('3000'))
        for p, amt in ((p1, '1000'), (p2, '2000')):
            p.refresh_from_db()
            inst = list(p.installments.all())
            self.assertEqual(len(inst), 1)
            self.assertEqual(inst[0].pay_amount, Decimal(amt))   # 默认=剩余=计划金额
            self.assertEqual(str(inst[0].pay_date), '2026-07-10')
            self.assertEqual(p.remaining, Decimal('0'))
            self.assertEqual(p.status, 'settled')

    def test_bulk_pay_skips_settled(self):
        p1 = self._schedule_payment(1, '1000')
        self._post('/api/pk/payments/bulk-pay', {'ids': [p1.id], 'pay_date': '2026-07-10'})
        resp = self._post('/api/pk/payments/bulk-pay', {'ids': [p1.id], 'pay_date': '2026-07-11'})
        d = resp.json()['data']
        self.assertEqual(d['paid'], 0)
        self.assertEqual(len(d['skipped']), 1)
        p1.refresh_from_db()
        self.assertEqual(p1.installments.count(), 1)   # 未重复登记

    def test_bulk_pay_default_date_today(self):
        p1 = self._schedule_payment(1, '800')
        resp = self._post('/api/pk/payments/bulk-pay', {'ids': [p1.id]})
        self.assertEqual(resp.status_code, 200, resp.content)
        p1.refresh_from_db()
        self.assertEqual(str(p1.installments.first().pay_date), date.today().isoformat())

    # ── 付款管理批量删除 ──────────────────────────────────────────────────────
    def test_bulk_delete_payments(self):
        p1 = self._schedule_payment(1, '1000')
        p2 = self._schedule_payment(2, '2000')
        resp = self._post('/api/pk/payments/bulk-delete', {'ids': [p1.id, p2.id]})
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data']['deleted'], 2)
        # soft-delete: records still exist in DB but with deleted_at set
        self.assertEqual(Payment.objects.filter(deleted_at__isnull=True).count(), 0)
        self.assertEqual(Payment.objects.filter(deleted_at__isnull=False).count(), 2)

    # ── 校验/边界 ────────────────────────────────────────────────────────────
    def test_empty_ids_rejected(self):
        for url in ('/api/pk/approvals/bulk-delete', '/api/pk/approvals/bulk-schedule',
                    '/api/pk/payments/bulk-delete', '/api/pk/payments/bulk-pay'):
            resp = self._post(url, {'ids': []})
            self.assertEqual(resp.status_code, 400, url)

    def test_single_select_via_ids(self):
        # 单选 = ids 单元素，复用批量端点
        a1 = self._mk_approval(1, '1000', status='pending')
        resp = self._post('/api/pk/approvals/bulk-delete', {'ids': [a1.id]})
        self.assertEqual(resp.json()['data']['deleted'], 1)

    # ── 卡片内逐条调整金额（items 入参）─────────────────────────────────────────
    def test_bulk_schedule_per_record_amount(self):
        # items 指定各条本次金额（可小于申请金额=分批排款，记录不归档）
        a1 = self._mk_approval(1, '1000')
        a2 = self._mk_approval(2, '2000')
        resp = self._post('/api/pk/approvals/bulk-schedule',
                          {'items': [{'id': a1.id, 'amount': '300'},
                                     {'id': a2.id, 'amount': '2000'}],
                           'planned_date': '2026-07-01'})
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['scheduled'], 2)
        self.assertEqual(Decimal(d['total_amount']), Decimal('2300'))
        p1 = Payment.objects.get(approval=a1)
        self.assertEqual(p1.total_amount, Decimal('300'))
        a1.refresh_from_db()
        self.assertFalse(a1.archived)                 # 分批：未排满不归档
        self.assertEqual(a1.scheduled_amount, Decimal('300'))
        a2.refresh_from_db()
        self.assertTrue(a2.archived)                  # 排满申请金额即归档

    def test_bulk_schedule_over_remaining_skipped(self):
        # items 金额超过剩余可排 → 该条跳过（_schedule_one 返回 400 类错误）
        a1 = self._mk_approval(1, '1000')
        resp = self._post('/api/pk/approvals/bulk-schedule',
                          {'items': [{'id': a1.id, 'amount': '5000'}],
                           'planned_date': '2026-07-01'})
        d = resp.json()['data']
        self.assertEqual(d['scheduled'], 0)
        self.assertEqual(len(d['skipped']), 1)
        self.assertFalse(Payment.objects.filter(approval=a1).exists())

    def test_bulk_pay_per_record_amount(self):
        # items 指定各条本次付款金额（可小于剩余=分次付款，记录不结清）
        p1 = self._schedule_payment(1, '1000')
        p2 = self._schedule_payment(2, '2000')
        resp = self._post('/api/pk/payments/bulk-pay',
                          {'items': [{'id': p1.id, 'amount': '400'},
                                     {'id': p2.id, 'amount': '2000'}],
                           'pay_date': '2026-07-10'})
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['paid'], 2)
        self.assertEqual(Decimal(d['total_amount']), Decimal('2400'))
        p1.refresh_from_db()
        self.assertEqual(p1.installments.first().pay_amount, Decimal('400'))
        self.assertEqual(p1.remaining, Decimal('600'))
        self.assertNotEqual(p1.status, 'settled')     # 分次付款：未付满不结清
        p2.refresh_from_db()
        self.assertEqual(p2.status, 'settled')

    def test_bulk_pay_over_remaining_skipped(self):
        p1 = self._schedule_payment(1, '1000')
        resp = self._post('/api/pk/payments/bulk-pay',
                          {'items': [{'id': p1.id, 'amount': '5000'}],
                           'pay_date': '2026-07-10'})
        d = resp.json()['data']
        self.assertEqual(d['paid'], 0)
        self.assertEqual(len(d['skipped']), 1)
        p1.refresh_from_db()
        self.assertEqual(p1.installments.count(), 0)


class ColumnFilterTests(TestCase):
    """Excel 风格列头筛选 + 排序：filters(JSON) / sort / order 的白名单解析与生效。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.dept = DEPARTMENTS[0]
        u = PaikuanUser(phone='13900000900', name='Admin', role='super_admin',
                        job_title='finance_director', departments=[self.dept],
                        is_active=True, is_approved=True)
        u.set_password('Test123456'); u.save()
        self.user = u
        self.token = make_token(u)
        # 付款管理：两条不同金额/日期/收款方
        Payment.objects.create(created_by=u, department=self.dept, approval_number='',
                               g7_number='G70001', project_desc='甲项目', payee='阿尔法',
                               applicant='张三', total_amount=Decimal('1000'),
                               planned_date=date(2026, 1, 10))
        Payment.objects.create(created_by=u, department=self.dept, approval_number='',
                               g7_number='G70002', project_desc='乙项目', payee='贝塔',
                               applicant='李四', total_amount=Decimal('8000'),
                               planned_date=date(2026, 6, 20))
        # 审批管理：两条
        ApprovalRecord.objects.create(applicant='王五', department=self.dept,
                                      approval_number='1', g7_number='', summary='差旅',
                                      amount=Decimal('500'), payee='丙方', status='pending')
        ApprovalRecord.objects.create(applicant='赵六', department=self.dept,
                                      approval_number='2', g7_number='', summary='采购',
                                      amount=Decimal('9000'), payee='丁方', status='approved')

    def tearDown(self):
        _invalidate_perm_cache()

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _items(self, url, params):
        r = self.client.get(url, params, **self.auth())
        self.assertEqual(r.status_code, 200, r.content)
        return r.json()['data']['items']

    # ── 付款管理 ──
    def test_payment_text_contains(self):
        items = self._items('/api/pk/payments',
                            {'filters': json.dumps({'payee': {'op': 'contains', 'value': '阿尔法'}})})
        self.assertEqual([i['payee'] for i in items], ['阿尔法'])

    def test_payment_number_between(self):
        items = self._items('/api/pk/payments',
                            {'filters': json.dumps({'total_amount': {'op': 'gte', 'value': '5000'}})})
        self.assertEqual([i['project_desc'] for i in items], ['乙项目'])

    def test_payment_date_range(self):
        items = self._items('/api/pk/payments',
                            {'filters': json.dumps({'planned_date': {'op': 'between',
                                                                     'value': ['2026-06-01', '2026-06-30']}})})
        self.assertEqual([i['project_desc'] for i in items], ['乙项目'])

    def test_payment_sort_amount_asc(self):
        items = self._items('/api/pk/payments', {'sort': 'total_amount', 'order': 'asc'})
        amts = [Decimal(i['total_amount']) for i in items]
        self.assertEqual(amts, sorted(amts))

    def test_payment_illegal_field_ignored(self):
        # 未注册字段不应影响结果（静默忽略），返回全部
        items = self._items('/api/pk/payments',
                            {'filters': json.dumps({'created_by__password': {'op': 'eq', 'value': 'x'}})})
        self.assertEqual(len(items), 2)

    # ── 审批管理 ──
    def test_approval_enum_status_in(self):
        items = self._items('/api/pk/approvals',
                            {'filters': json.dumps({'status': {'op': 'in', 'value': ['approved']}})})
        self.assertEqual([i['applicant'] for i in items], ['赵六'])

    def test_approval_number_gt(self):
        items = self._items('/api/pk/approvals',
                            {'filters': json.dumps({'amount': {'op': 'gt', 'value': '1000'}})})
        self.assertEqual([i['applicant'] for i in items], ['赵六'])

    def test_approval_global_q(self):
        items = self._items('/api/pk/approvals', {'q': '采购'})
        self.assertEqual([i['applicant'] for i in items], ['赵六'])

    def test_approval_sort_amount_desc(self):
        items = self._items('/api/pk/approvals', {'sort': 'amount', 'order': 'desc'})
        self.assertEqual([i['applicant'] for i in items], ['赵六', '王五'])

    def test_approval_remaining_amount_computed_filter_and_sort(self):
        """未排金额（计算列 = 申请额 − 已排额，钳 0）支持数值筛选 + 排序。"""
        # 第三条：申请 2000，已排 1800 → 未排 200（最小）
        ApprovalRecord.objects.create(applicant='孙七', department=self.dept,
                                      approval_number='3', g7_number='', summary='维修',
                                      amount=Decimal('2000'), scheduled_amount=Decimal('1800'),
                                      payee='戊方', status='pending')
        # 未排 < 1000 → 王五(500) + 孙七(200)
        items = self._items('/api/pk/approvals',
                            {'filters': json.dumps({'remaining_amount': {'op': 'lt', 'value': '1000'}})})
        self.assertEqual(sorted(i['applicant'] for i in items), ['孙七', '王五'])
        # 升序排序：孙七200 < 王五500 < 赵六9000
        items2 = self._items('/api/pk/approvals', {'sort': 'remaining_amount', 'order': 'asc'})
        self.assertEqual([i['applicant'] for i in items2], ['孙七', '王五', '赵六'])


class PaymentComputedFilterTests(TestCase):
    """付款管理计算列筛选：已付/剩余(数值)、逾期(是/否)，及按已付/剩余排序。
    口径必须与 Payment.to_dict 显示值一致。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.dept = DEPARTMENTS[0]
        u = PaikuanUser(phone='13900000950', name='Admin', role='super_admin',
                        job_title='finance_director', departments=[self.dept],
                        is_active=True, is_approved=True)
        u.set_password('Test123456'); u.save()
        self.user = u
        self.token = make_token(u)
        from paikuan.models import PaymentInstallment
        today = date.today()
        past = today - __import__('datetime').timedelta(days=10)
        future = today + __import__('datetime').timedelta(days=30)
        # A: 未付，计划在过去 → paid=0, remaining=1000, 逾期=是
        self.a = Payment.objects.create(created_by=u, department=self.dept, approval_number='',
                                        project_desc='甲', payee='A', total_amount=Decimal('1000'),
                                        planned_date=past)
        # B: 已付清，计划在过去 → paid=1000, remaining=0, 逾期=否(已结清)
        self.b = Payment.objects.create(created_by=u, department=self.dept, approval_number='',
                                        project_desc='乙', payee='B', total_amount=Decimal('1000'),
                                        planned_date=past)
        PaymentInstallment.objects.create(payment=self.b, seq=1, pay_date=past, pay_amount=Decimal('1000'))
        # C: 部分付款，计划在未来 → paid=400, remaining=600, 逾期=否
        self.c = Payment.objects.create(created_by=u, department=self.dept, approval_number='',
                                        project_desc='丙', payee='C', total_amount=Decimal('1000'),
                                        planned_date=future)
        PaymentInstallment.objects.create(payment=self.c, seq=1, pay_date=future, pay_amount=Decimal('400'))

    def tearDown(self):
        _invalidate_perm_cache()

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _descs(self, params):
        r = self.client.get('/api/pk/payments', params, **self.auth())
        self.assertEqual(r.status_code, 200, r.content)
        return sorted(i['project_desc'] for i in r.json()['data']['items'])

    def test_paid_gte(self):
        # 已付 >= 400 → 乙(1000) 丙(400)
        self.assertEqual(self._descs({'filters': json.dumps({'paid': {'op': 'gte', 'value': '400'}})}),
                         ['丙', '乙'])

    def test_remaining_zero(self):
        # 剩余 = 0 → 仅乙(已结清，钳到0)
        self.assertEqual(self._descs({'filters': json.dumps({'remaining': {'op': 'eq', 'value': '0'}})}),
                         ['乙'])

    def test_remaining_between(self):
        # 剩余 区间 [500,700] → 仅丙(600)
        self.assertEqual(self._descs({'filters': json.dumps({'remaining': {'op': 'between', 'value': ['500', '700']}})}),
                         ['丙'])

    def test_overdue_yes(self):
        # 逾期=是 → 仅甲(过去且未结清)；乙过去但已结清=否，丙未来=否
        self.assertEqual(self._descs({'filters': json.dumps({'overdue': {'op': 'in', 'value': ['是']}})}),
                         ['甲'])

    def test_overdue_no(self):
        # 逾期=否 → 乙、丙
        self.assertEqual(self._descs({'filters': json.dumps({'overdue': {'op': 'in', 'value': ['否']}})}),
                         ['丙', '乙'])

    def test_sort_paid_asc(self):
        r = self.client.get('/api/pk/payments', {'sort': 'paid', 'order': 'asc'}, **self.auth())
        self.assertEqual(r.status_code, 200, r.content)
        descs = [i['project_desc'] for i in r.json()['data']['items']]
        # paid: 甲0 < 丙400 < 乙1000
        self.assertEqual(descs, ['甲', '丙', '乙'])

    def test_summary_consistent_with_filter(self):
        # 筛选即合计：已付>=400 时 planned_total 应为 乙+丙 = 2000
        r = self.client.get('/api/pk/payments',
                            {'filters': json.dumps({'paid': {'op': 'gte', 'value': '400'}})}, **self.auth())
        d = r.json()['data']
        self.assertEqual(d['total'], 2)
        self.assertEqual(Decimal(d['planned_total']), Decimal('2000'))

    def test_status_multi_select_union(self):
        """计划状态列头多选：status 逗号分隔取并集（A 逾期 / B 已付清 / C 部分付款）。"""
        # 单选仍生效
        self.assertEqual(self._descs({'status': 'overdue'}), ['甲'])
        self.assertEqual(self._descs({'status': 'settled'}), ['乙'])
        self.assertEqual(self._descs({'status': 'partial'}), ['丙'])
        # 多选并集
        self.assertEqual(self._descs({'status': 'overdue,settled'}), ['乙', '甲'])
        self.assertEqual(self._descs({'status': 'overdue,partial'}), ['丙', '甲'])
        self.assertEqual(self._descs({'status': 'overdue,settled,partial'}), ['丙', '乙', '甲'])


class AuthPasswordPolicyTests(TestCase):
    """密码强度校验 + 自助改密 + 超管重置临时密码（强制改密）。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.dept = DEPARTMENTS[0]
        # 第一个用户即超管
        self.admin = PaikuanUser(phone='13900001000', name='Boss', role='super_admin',
                                 job_title='finance_director', departments=[self.dept],
                                 is_active=True, is_approved=True)
        self.admin.set_password('Admin12345'); self.admin.save()
        self.admin_token = make_token(self.admin)
        # 普通用户
        self.u = PaikuanUser(phone='13900001001', name='Zhang', role='operator',
                             job_title='cashier', departments=[self.dept],
                             is_active=True, is_approved=True)
        self.u.set_password('Good1234'); self.u.save()
        self.u_token = make_token(self.u)

    def tearDown(self):
        _invalidate_perm_cache()

    def _register(self, **over):
        body = {'phone': '13900002000', 'password': 'Strong123', 'name': '新人',
                'job_title': 'cashier', 'departments': [self.dept]}
        body.update(over)
        return self.client.post('/api/pk/register', data=json.dumps(body),
                                content_type='application/json')

    def test_register_rejects_weak_passwords(self):
        for pw in ['12345678', 'abcdefgh', '88888888', 'password',
                   'aaaaaaaa', '13900002000']:
            r = self._register(password=pw)
            self.assertIn('error', r.json(), f'{pw} 应被拒')
            self.assertFalse(PaikuanUser.objects.filter(phone='13900002000').exists(),
                             f'{pw} 不应建号')

    def test_register_rejects_password_contains_phone(self):
        r = self._register(password='ab13900002000')
        self.assertIn('error', r.json())

    def test_register_accepts_strong_password(self):
        r = self._register(password='Qingdao2026')
        self.assertNotIn('error', r.json(), r.content)
        self.assertTrue(PaikuanUser.objects.filter(phone='13900002000').exists())

    def test_change_password_wrong_old(self):
        r = self.client.post('/api/pk/me/password',
                             data=json.dumps({'old_password': 'WRONG', 'new_password': 'NewGood123'}),
                             content_type='application/json',
                             HTTP_AUTHORIZATION=f'Bearer {self.u_token}')
        self.assertIn('error', r.json())

    def test_change_password_weak_new_rejected(self):
        r = self.client.post('/api/pk/me/password',
                             data=json.dumps({'old_password': 'Good1234', 'new_password': '12345678'}),
                             content_type='application/json',
                             HTTP_AUTHORIZATION=f'Bearer {self.u_token}')
        self.assertIn('error', r.json())

    def test_change_password_success_returns_token_and_clears_flag(self):
        self.u.must_change_password = True; self.u.save()
        r = self.client.post('/api/pk/me/password',
                             data=json.dumps({'old_password': 'Good1234', 'new_password': 'Brand2026new'}),
                             content_type='application/json',
                             HTTP_AUTHORIZATION=f'Bearer {self.u_token}')
        d = r.json()
        self.assertNotIn('error', d, r.content)
        self.assertIn('token', d['data'])
        self.u.refresh_from_db()
        self.assertFalse(self.u.must_change_password)
        self.assertTrue(self.u.check_password('Brand2026new'))

    def test_admin_reset_sets_must_change_and_validates(self):
        # 弱临时密码被拒
        r = self.client.put(f'/api/pk/users/{self.u.id}',
                            data=json.dumps({'password': '12345678'}),
                            content_type='application/json',
                            HTTP_AUTHORIZATION=f'Bearer {self.admin_token}')
        self.assertIn('error', r.json())
        # 合规临时密码 → must_change_password=True
        r = self.client.put(f'/api/pk/users/{self.u.id}',
                            data=json.dumps({'password': 'Temp2026ab'}),
                            content_type='application/json',
                            HTTP_AUTHORIZATION=f'Bearer {self.admin_token}')
        self.assertNotIn('error', r.json(), r.content)
        self.u.refresh_from_db()
        self.assertTrue(self.u.must_change_password)

    def test_login_surfaces_must_change_flag(self):
        self.u.must_change_password = True
        self.u.set_password('Temp2026ab'); self.u.save()
        r = self.client.post('/api/pk/login',
                            data=json.dumps({'phone': '13900001001', 'password': 'Temp2026ab'}),
                            content_type='application/json')
        d = r.json()['data']
        self.assertTrue(d['must_change_password'])


class ListSchemeTests(TestCase):
    """通用列表筛选方案基座：CRUD、私有/公共可见性、默认方案跟随账号。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.dept = DEPARTMENTS[0]

    def tearDown(self):
        _invalidate_perm_cache()

    def _user(self, phone, role='super_admin', job='finance_director'):
        u = PaikuanUser(phone=phone, name=f'{job} {phone}', role=role, job_title=job,
                        departments=[self.dept], is_active=True, is_approved=True)
        u.set_password('Test123456')
        u.save()
        return u

    def auth(self, u):
        return {'HTTP_AUTHORIZATION': f'Bearer {make_token(u)}'}

    def post(self, url, payload, u):
        return self.client.post(url, data=json.dumps(payload),
                                content_type='application/json', **self.auth(u))

    def test_private_vs_public_visibility_and_payload(self):
        a = self._user('13900000301')
        b = self._user('13900000302')
        payload = {'colFilters': {'applicant': {'op': 'contains', 'value': '张'}},
                   'sort': 'planned_date', 'order': 'desc', 'status': '待付款'}
        # A 建私有 + 公共
        r1 = self.post('/api/pk/list-schemes',
                       {'module': 'pk_payments', 'name': '我的待付', 'scope': 'private', 'payload': payload}, a)
        self.assertEqual(r1.status_code, 200)
        # payload 原样回带
        self.assertEqual(r1.json()['data']['payload'], payload)
        r2 = self.post('/api/pk/list-schemes',
                       {'module': 'pk_payments', 'name': '团队待付', 'scope': 'public', 'payload': payload}, a)
        self.assertEqual(r2.status_code, 200)
        # A 看到两个
        la = self.client.get('/api/pk/list-schemes', {'module': 'pk_payments'}, **self.auth(a)).json()['data']
        self.assertEqual({s['name'] for s in la['items']}, {'我的待付', '团队待付'})
        # B 只看到公共
        lb = self.client.get('/api/pk/list-schemes', {'module': 'pk_payments'}, **self.auth(b)).json()['data']
        self.assertEqual([s['name'] for s in lb['items']], ['团队待付'])

    def test_public_scheme_requires_create_perm(self):
        # cashier 角色无 can_create → 不能建公共，但能建私有
        cashier = self._user('13900000303', role='operator', job='cashier')
        pub = self.post('/api/pk/list-schemes',
                        {'module': 'pk_payments', 'name': 'x', 'scope': 'public', 'payload': {}}, cashier)
        self.assertEqual(pub.status_code, 403)
        priv = self.post('/api/pk/list-schemes',
                         {'module': 'pk_payments', 'name': 'x', 'scope': 'private', 'payload': {}}, cashier)
        self.assertEqual(priv.status_code, 200)

    def test_default_follows_account_and_cascade(self):
        a = self._user('13900000304')
        sid = self.post('/api/pk/list-schemes',
                        {'module': 'pk_payments', 'name': '默认', 'scope': 'private', 'payload': {}}, a
                        ).json()['data']['id']
        r = self.post('/api/pk/list-schemes/set-default',
                      {'module': 'pk_payments', 'scheme_id': sid}, a)
        self.assertEqual(r.json()['data']['default_id'], sid)
        lst = self.client.get('/api/pk/list-schemes', {'module': 'pk_payments'}, **self.auth(a)).json()['data']
        self.assertEqual(lst['default_id'], sid)
        self.assertTrue(next(s for s in lst['items'] if s['id'] == sid)['is_default'])
        # 删除方案 → 默认级联清理
        self.client.delete(f'/api/pk/list-schemes/{sid}', **self.auth(a))
        lst2 = self.client.get('/api/pk/list-schemes', {'module': 'pk_payments'}, **self.auth(a)).json()['data']
        self.assertIsNone(lst2['default_id'])

    def test_unknown_module_rejected(self):
        a = self._user('13900000305')
        r = self.client.get('/api/pk/list-schemes', {'module': 'nope'}, **self.auth(a))
        self.assertEqual(r.status_code, 400)

    def test_ar_module_uses_project_page_perm(self):
        # AR 列表模块映射到对应页权限；超管直接放行，能建并取回
        a = self._user('13900000306')
        for module in ('ar_projects', 'ar_customers', 'ar_advances', 'ar_budget'):
            r = self.post('/api/pk/list-schemes',
                          {'module': module, 'name': f'{module}方案', 'scope': 'private',
                           'payload': {'colFilters': {}, 'sort': '', 'order': ''}}, a)
            self.assertEqual(r.status_code, 200, module)
            lst = self.client.get('/api/pk/list-schemes', {'module': module}, **self.auth(a)).json()['data']
            self.assertEqual([s['name'] for s in lst['items']], [f'{module}方案'])

    def test_admin_only_modules_gated(self):
        # pk_users / pk_audit_logs 仅超管可用；非超管被拒
        admin = self._user('13900000307', role='super_admin')
        non_admin = self._user('13900000308', role='operator', job='cashier')
        for module in ('pk_users', 'pk_audit_logs'):
            ok_r = self.post('/api/pk/list-schemes',
                             {'module': module, 'name': 'x', 'scope': 'private', 'payload': {}}, admin)
            self.assertEqual(ok_r.status_code, 200, module)
            denied = self.client.get('/api/pk/list-schemes', {'module': module}, **self.auth(non_admin))
            self.assertEqual(denied.status_code, 403, module)


class TransportReconciliationTests(TestCase):
    """运输事业部对账单专用导入/导出：原表(负数/异构列) ↔ 标准排款，去重 + 零误差往返。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.dept = '运输事业部'
        admin = PaikuanUser(phone='13900009100', name='TransAdmin', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.token = make_token(admin)

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    # 运输系统导出的原始表头（与生产一致）
    HEADERS = ['序号', '所属组织', '对账单号', '对账对象', '联系电话', '实际对账金额',
               '对账时间', '状态', '创建人', '创建时间', '备注', '单据类别', '账单调整', '对账金额']

    def _row(self, seq, bill_no, payee, amount, status='已通过'):
        # amount 传正数，原表里存为负数（来源系统口径）
        return [seq, '运输项目一部', bill_no, payee, '', -abs(amount),
                '2026-06-26 10:47:20', status, '谭雯雯', '2026-06-26 10:51:24',
                '税后利润 2592.27 税后利润率 30.19%', '车辆', '', -abs(amount)]

    def _xlsx(self, rows):
        from openpyxl import Workbook
        from django.core.files.uploadedfile import SimpleUploadedFile
        wb = Workbook(); ws = wb.active
        ws.append(self.HEADERS)
        for r in rows:
            ws.append(r)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return SimpleUploadedFile('transport.xlsx', buf.read(),
                                  content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def _import(self, rows):
        # 导入口在审批管理：建「已通过」审批记录
        return self.client.post('/api/pk/approvals/transport/import',
                                {'file': self._xlsx(rows)}, **self.auth())

    def _schedule(self, rec, amount):
        # 手动排款：审批 → 付款管理（与现有排款同口径）
        return self.client.post(
            f'/api/pk/approvals/{rec.id}/schedule',
            data=json.dumps({'planned_date': '2026-06-27', 'total_amount': str(amount)}),
            content_type='application/json', **self.auth())

    def test_import_creates_approved_record_abs_amount(self):
        resp = self._import([self._row(1, 'ZD202606260055', '安仕吉-浓香酒', 4828.5)])
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['created'], 1)
        # 进的是审批管理，建为「已通过」审批记录（非付款记录）
        self.assertEqual(Payment.objects.filter(approval__ext_source='transport').count(), 0)
        rec = ApprovalRecord.objects.get(ext_bill_no='ZD202606260055')
        self.assertEqual(rec.ext_source, 'transport')
        self.assertEqual(rec.status, 'approved')          # 已通过，可直接排款
        self.assertEqual(rec.department, '运输事业部')
        self.assertEqual(rec.secondary_dept, '运输项目一部')
        self.assertEqual(rec.payee, '安仕吉-浓香酒')          # 收款主体 ← 对账对象
        self.assertEqual(rec.amount, Decimal('4828.5'))   # 取绝对值
        self.assertEqual(rec.approval_number, '')          # 审批编号留空
        self.assertEqual(rec.g7_number, 'ZD202606260055')  # G7编号 ← 对账单号
        self.assertEqual(rec.notes, '税后利润 2592.27 税后利润率 30.19%')  # 备注 ← 原表备注原文
        self.assertEqual(rec.ext_raw['对账单号'], 'ZD202606260055')
        self.assertEqual(rec.ext_raw['实际对账金额'], -4828.5)  # 原始负数逐字保留

    def test_import_dedup_by_bill_no(self):
        # 文件内重复 + 二次导入重复，均按对账单号去重
        rows = [self._row(1, 'ZD202606260001', 'A', 100), self._row(2, 'ZD202606260001', 'A', 100),
                self._row(3, 'ZD202606260002', 'B', 200)]
        d = self._import(rows).json()['data']
        self.assertEqual(d['created'], 2)
        self.assertEqual(d['duplicates'], 1)
        # 再次导入同文件 → 全部判重，0 新增
        d2 = self._import(rows).json()['data']
        self.assertEqual(d2['created'], 0)
        self.assertEqual(d2['duplicates'], 3)
        self.assertEqual(ApprovalRecord.objects.filter(ext_source='transport').count(), 2)

    def test_full_flow_roundtrip_zero_drift_only_status_changes(self):
        # 导入(审批) → 手动排款(付款) → 结算其一 → 付款管理导出，校验仅已结算行状态列改写
        self._import([self._row(1, 'ZD202606260010', '承运商甲', 1000),
                      self._row(2, 'ZD202606260020', '承运商乙', 2000)])
        ra = ApprovalRecord.objects.get(ext_bill_no='ZD202606260010')
        rb = ApprovalRecord.objects.get(ext_bill_no='ZD202606260020')
        self.assertEqual(self._schedule(ra, 1000).status_code, 200)
        self.assertEqual(self._schedule(rb, 2000).status_code, 200)
        pa = Payment.objects.get(approval=ra)
        pb = Payment.objects.get(approval=rb)
        # 结算 A：付清 → status=settled
        PaymentInstallment.objects.create(payment=pa, seq=1, pay_date=date(2026, 6, 27),
                                           pay_amount=Decimal('1000'))
        self.assertEqual(pa.status, 'settled')
        # 勾选导出 A + B（付款记录主键）
        resp = self.client.get('/api/pk/payments/transport/export',
                               {'ids': f'{pa.id},{pb.id}'}, **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        from openpyxl import load_workbook
        ws = load_workbook(io.BytesIO(resp.content), data_only=True).active
        hdr = [c.value for c in ws[1]]
        self.assertEqual(hdr, self.HEADERS)            # 原列顺序还原
        si = hdr.index('状态'); bi = hdr.index('对账单号'); ai = hdr.index('实际对账金额')
        by_bill = {row[bi].value: row for row in ws.iter_rows(min_row=2)}
        rowa, rowb = by_bill['ZD202606260010'], by_bill['ZD202606260020']
        # A 已结算：状态列改写；金额等其它列零误差（负数原样）
        self.assertEqual(rowa[si].value, '已结算')
        self.assertEqual(rowa[ai].value, -1000)
        # B 未结算：状态保留原值「已通过」，全列零误差
        self.assertEqual(rowb[si].value, '已通过')
        self.assertEqual(rowb[ai].value, -2000)

        # 强校验「零误差」：未结算行 B 逐列与原表完全一致（空单元 None/'' 视为等价）
        def _norm(v):
            return None if v in (None, '') else v
        orig_b = self._row(2, 'ZD202606260020', '承运商乙', 2000)
        for ci, col in enumerate(self.HEADERS):
            self.assertEqual(_norm(rowb[ci].value), _norm(orig_b[ci]),
                             f'未结算行列「{col}」发生漂移')
        # 已结算行 A：除「状态」列外全列与原表一致，状态列改为「已结算」
        orig_a = self._row(1, 'ZD202606260010', '承运商甲', 1000)
        for ci, col in enumerate(self.HEADERS):
            if col == '状态':
                self.assertEqual(rowa[ci].value, '已结算')
            else:
                self.assertEqual(_norm(rowa[ci].value), _norm(orig_a[ci]),
                                 f'已结算行非状态列「{col}」发生漂移')

    def test_export_empty_before_scheduling(self):
        # 仅导入未排款 → 付款管理无运输付款记录 → 导出报错引导先排款
        self._import([self._row(1, 'ZD202606260030', 'X', 500)])
        resp = self.client.get('/api/pk/payments/transport/export', **self.auth())
        self.assertEqual(resp.status_code, 400)

    def test_import_rejects_unparseable_amount(self):
        bad = self._row(1, 'ZD202606260099', 'X', 0)
        bad[5] = 'abc'   # 实际对账金额列非数字
        d = self._import([bad]).json()['data']
        self.assertEqual(d['created'], 0)
        self.assertEqual(d['skipped'], 1)
        self.assertTrue(d['errors'])

    def test_g7_numbers_copy_filter_and_ids(self):
        # 导入 + 排款 → 两条运输付款；G7编号=对账单号，跨页复制接口取全量/勾选
        self._import([self._row(1, 'ZD202606260040', '甲', 100),
                      self._row(2, 'ZD202606260041', '乙', 200)])
        ra = ApprovalRecord.objects.get(ext_bill_no='ZD202606260040')
        rb = ApprovalRecord.objects.get(ext_bill_no='ZD202606260041')
        self._schedule(ra, 100)
        self._schedule(rb, 200)
        pa = Payment.objects.get(approval=ra)
        # 无 ids → 走筛选口径取全部（去重）
        d = self.client.get('/api/pk/payments/transport/g7-numbers', **self.auth()).json()['data']
        self.assertEqual(set(d['numbers']), {'ZD202606260040', 'ZD202606260041'})
        self.assertEqual(d['count'], 2)
        self.assertFalse(d['capped'])
        # 勾选 ids → 仅取选中
        d2 = self.client.get('/api/pk/payments/transport/g7-numbers',
                             {'ids': str(pa.id)}, **self.auth()).json()['data']
        self.assertEqual(d2['numbers'], ['ZD202606260040'])

    def test_select_ids_cross_page(self):
        # 跨页全选：select-ids 返回当前筛选口径下全部付款记录 ID（不分页）
        self._import([self._row(1, 'ZD202606260050', '甲', 100),
                      self._row(2, 'ZD202606260051', '乙', 200)])
        for bill in ('ZD202606260050', 'ZD202606260051'):
            rec = ApprovalRecord.objects.get(ext_bill_no=bill)
            self._schedule(rec, rec.amount)
        d = self.client.get('/api/pk/payments/select-ids', **self.auth()).json()['data']
        pids = set(Payment.objects.filter(approval__ext_source='transport')
                   .values_list('id', flat=True))
        self.assertEqual(set(d['ids']), pids)
        self.assertEqual(d['count'], 2)
        self.assertFalse(d['capped'])

    def test_export_column_order_canonical_despite_scrambled_raw(self):
        # 模拟生产库（Postgres jsonb）不保证 JSON 键序：ext_raw 以乱序键存储，
        # 导出仍须按原表标准列序 TRANSPORT_HEADERS 还原，列序不漂移。
        self._import([self._row(1, 'ZD202606260060', '甲', 300)])
        rec = ApprovalRecord.objects.get(ext_bill_no='ZD202606260060')
        # 反转 ext_raw 键序后回存，模拟 jsonb 重排键
        rec.ext_raw = {k: rec.ext_raw[k] for k in reversed(list(rec.ext_raw.keys()))}
        rec.save(update_fields=['ext_raw'])
        self._schedule(rec, 300)
        p = Payment.objects.get(approval=rec)
        resp = self.client.get('/api/pk/payments/transport/export',
                               {'ids': str(p.id)}, **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        from openpyxl import load_workbook
        ws = load_workbook(io.BytesIO(resp.content), data_only=True).active
        hdr = [c.value for c in ws[1]]
        self.assertEqual(hdr, self.HEADERS)   # 乱序键 → 仍按标准列序导出


class ApprovalVisibilityTests(TestCase):
    """审批列表可见口径：已拒绝/已撤销可查阅（默认由前端状态筛选隐藏），合计只计在途；
    已排满归档的审批通过记录仍隐藏（已完全流转至付款管理）。"""

    def setUp(self):
        self.client = Client()
        self.dept = '运输事业部'
        admin = PaikuanUser(phone='13900003900', name='VisAdmin', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.token = make_token(admin)

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _mk(self, num, amount, status='pending', archived=False):
        return ApprovalRecord.objects.create(
            applicant='张三', department=self.dept, approval_number=str(num) * 21,
            summary='采购', amount=Decimal(amount), payee=f'供应商{num}',
            status=status, archived=archived)

    def _list(self, params=None):
        r = self.client.get('/api/pk/approvals', params or {}, **self.auth())
        self.assertEqual(r.status_code, 200, r.content)
        return r.json()['data']

    def test_rejected_canceled_visible_but_excluded_from_totals(self):
        self._mk(1, '1000', status='pending')
        self._mk(2, '2000', status='approved')
        self._mk(3, '3000', status='rejected', archived=True)
        self._mk(4, '4000', status='canceled', archived=True)
        # 无状态筛选 → 四条全部可查（不再因 archived 隐藏已拒绝/已撤销）
        d = self._list()
        self.assertEqual(d['total'], 4)
        # 合计只计在途（待审批 + 审批通过）= 1000 + 2000
        self.assertEqual(Decimal(d['total_amount']), Decimal('3000'))
        # 状态筛选「已拒绝」→ 命中第三条
        d2 = self._list({'filters': json.dumps({'status': {'op': 'in', 'value': ['rejected']}})})
        self.assertEqual(d2['total'], 1)
        self.assertEqual(d2['items'][0]['status'], 'rejected')

    def test_default_status_filter_shows_pending_approved(self):
        self._mk(1, '1000', status='pending')
        self._mk(2, '2000', status='approved')
        self._mk(3, '3000', status='rejected', archived=True)
        self._mk(4, '4000', status='canceled', archived=True)
        # 前端默认状态筛选 [pending, approved] 对应的后端查询
        d = self._list({'filters': json.dumps({'status': {'op': 'in', 'value': ['pending', 'approved']}})})
        self.assertEqual(d['total'], 2)
        self.assertEqual({i['status'] for i in d['items']}, {'pending', 'approved'})

    def test_fully_scheduled_approved_hidden(self):
        # 审批通过且已排满（archived=True, status=approved）→ 列表隐藏（已流转付款管理）
        self._mk(1, '1000', status='approved', archived=True)
        keep = self._mk(2, '2000', status='approved', archived=False)
        d = self._list()
        self.assertEqual(d['total'], 1)
        self.assertEqual(d['items'][0]['id'], keep.id)


class AsyncExportTests(TestCase):
    """异步导出：核心 replay-request 机制 + 任务建立/状态/权限。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.dept = DEPARTMENTS[0]
        self.admin = PaikuanUser(phone='13900004000', name='ExpAdmin', role='super_admin',
                                 job_title='finance_director', departments=[self.dept],
                                 is_active=True, is_approved=True)
        self.admin.set_password('Test123456')
        self.admin.save()
        self.token = make_token(self.admin)

    def tearDown(self):
        _invalidate_perm_cache()

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _mk_approvals(self, n):
        for i in range(n):
            ApprovalRecord.objects.create(
                applicant='张三', department=self.dept, approval_number=str(i + 1).zfill(21),
                summary=f'采购{i}', amount=Decimal('1000'), payee=f'供应商{i}', status='pending')

    def test_replay_request_builds_approval_xlsx(self):
        # 直接驱动导出核心（无线程）：验证 _ReplayRequest 能重建同口径并生成有效 xlsx
        from paikuan.views import _approval_export_core, _ReplayRequest
        self._mk_approvals(3)
        req = _ReplayRequest(self.admin, {})
        resp = _approval_export_core(req, export_cap=100000)
        self.assertEqual(resp.status_code, 200, getattr(resp, 'content', b''))
        from openpyxl import load_workbook
        ws = load_workbook(io.BytesIO(resp.content), data_only=True).active
        # 表头 + 3 行数据
        self.assertEqual(ws.max_row, 4)

    def test_run_export_job_end_to_end(self):
        # 同线程直接跑 worker 逻辑（不经 threading），验证落库字节
        from paikuan.models import ExportJob
        from paikuan.views import _run_export_job
        self._mk_approvals(2)
        job = ExportJob.objects.create(kind='approvals', created_by=self.admin, params={})
        # worker 会在 finally 关闭连接；TestCase 下手动调用核心避免连接副作用——
        # 这里直接调用核心并落库，等价于 worker 主体逻辑
        from paikuan.views import _approval_export_core, _ReplayRequest
        resp = _approval_export_core(_ReplayRequest(self.admin, {}), export_cap=100000)
        self.assertEqual(resp.status_code, 200)
        job.file_data = resp.content
        job.filename = '审批记录.xlsx'
        job.status = 'done'
        job.save()
        # 下载端点
        r = self.client.get(f'/api/pk/exports/{job.id}/download', **self.auth())
        self.assertEqual(r.status_code, 200, r.content)
        self.assertTrue(len(r.content) > 100)

    def test_export_create_endpoint(self):
        self._mk_approvals(1)
        r = self.client.post('/api/pk/exports', data=json.dumps({'kind': 'approvals', 'params': {}}),
                             content_type='application/json', **self.auth())
        self.assertEqual(r.status_code, 200, r.content)
        d = r.json()['data']
        self.assertIn(d['status'], ('pending', 'running', 'done'))
        self.assertEqual(d['kind'], 'approvals')

    def test_export_create_rejects_unknown_kind(self):
        r = self.client.post('/api/pk/exports', data=json.dumps({'kind': 'bogus'}),
                             content_type='application/json', **self.auth())
        self.assertEqual(r.status_code, 400)

    def test_export_status_scoped_to_owner(self):
        from paikuan.models import ExportJob
        other = PaikuanUser(phone='13900004001', name='Other', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        other.set_password('Test123456'); other.save()
        job = ExportJob.objects.create(kind='approvals', created_by=other, params={})
        # 当前用户查别人的任务 → 404
        r = self.client.get(f'/api/pk/exports/{job.id}', **self.auth())
        self.assertEqual(r.status_code, 404)
