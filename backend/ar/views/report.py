"""周期报表（周报 / 月报）业务域 — 集团财务中心 / 事业部的财务周期汇报。

站在业财融合视角，把分散在应收、预算、现金流各域的口径，按一个统一的「汇报期间」
切片，组装成一份可截图 / 可导出 Excel 的标准化财务报告。所有计算都是服务端单一
正源（前端只渲染、导出复用同一份计算），避免前后端口径漂移。

五大指标：
  1. 项目规模     ← ARProject（项目台账）
  2. 应收账款     ← ARRecord + ARPayment + ARAdjustment（资金运动表口径）
  3. 应收预算完成 ← CollectionBudget vs 实际现金回款
  4. 应付预算完成 ← PaymentBudget  vs 实际付款（月度口径）
  5. 现金流情况   ← 与「现金流分析」同口径（经营流入/流出/净额），按期间切片
"""
from ._common import *  # noqa: F401,F403

from django.db.models import OuterRef, Subquery

DEFAULT_REVIEWER = '李亚琳'
_ISO = datetime.date.fromisoformat


# ══════════════════════════════════════════════════════════════════════════════
# 期间与范围解析
# ══════════════════════════════════════════════════════════════════════════════

def _resolve_period(request, today):
    """解析汇报期间 → (period_type, start, end, year, month, week_of_month)。

    周报需要前端传 start_date/end_date（周一~周日）；月报可传 year/month 或 start/end。
    week_of_month 以期间起始日所在自然月的第几个 7 天块计（(day-1)//7+1）。
    """
    ptype = (request.GET.get('period') or 'monthly').strip()
    sd = (request.GET.get('start_date') or '').strip()
    ed = (request.GET.get('end_date') or '').strip()
    if sd and ed:
        start, end = _ISO(sd), _ISO(ed)
    elif ptype == 'weekly':
        weekday = today.weekday()  # 周一=0
        start = today - datetime.timedelta(days=weekday)
        end = start + datetime.timedelta(days=6)
    else:
        y = _int_param(request, 'year', today.year)
        m = _int_param(request, 'month', today.month)
        start = datetime.date(y, m, 1)
        end = datetime.date(y, m, calendar.monthrange(y, m)[1])
    year, month = start.year, start.month
    week_of_month = (start.day - 1) // 7 + 1
    # 周报标题归属：周一可能落在上月，前端按「所选月 + 周序」显式标注，后端优先采用，
    # 保证图表/Excel 标题与前端选择器一致（避免「跨月周」第X周错位）。
    if ptype == 'weekly':
        yp = request.GET.get('year', '').strip()
        mp = request.GET.get('month', '').strip()
        wp = request.GET.get('week', '').strip()
        if yp.isdigit():
            year = int(yp)
        if mp.isdigit():
            month = int(mp)
        if wp.isdigit():
            week_of_month = int(wp)
    return ptype, start, end, year, month, week_of_month


def _prev_period(ptype, start, end):
    """上一可比期间（周报=上一周；月报=上一月），用于「增减」对比。"""
    if ptype == 'weekly':
        p_end = start - datetime.timedelta(days=1)
        p_start = p_end - datetime.timedelta(days=6)
    else:
        p_end = start - datetime.timedelta(days=1)
        p_start = datetime.date(p_end.year, p_end.month, 1)
    return p_start, p_end


def _money(v):
    return float(v or 0)


def _rate(actual, base):
    a, b = Decimal(str(actual or 0)), Decimal(str(base or 0))
    if b == 0:
        return None
    return round(float(a / b * 100), 1)


# ══════════════════════════════════════════════════════════════════════════════
# 指标 1 — 项目规模
# ══════════════════════════════════════════════════════════════════════════════

def _projects_metric(depts, year, p_start, p_end, pv_start, pv_end):
    year_start = datetime.date(year, 1, 1)
    base = ARProject.objects.filter(delivery_dept__in=depts)
    rows = {}
    for d in depts:
        q = base.filter(delivery_dept=d)
        active = q.filter(status='运作中').count()
        total = q.count()
        ytd_new = q.filter(contract_date__gte=year_start, contract_date__lte=p_end).count()
        period_new = q.filter(contract_date__gte=p_start, contract_date__lte=p_end).count()
        prev_new = q.filter(contract_date__gte=pv_start, contract_date__lte=pv_end).count()
        rows[d] = {
            'active': active, 'total': total, 'ytd_new': ytd_new,
            'period_new': period_new, 'prev_new': prev_new,
            'delta': period_new - prev_new,
        }
    return rows


# ══════════════════════════════════════════════════════════════════════════════
# 指标 2 — 应收账款（资金运动表口径）
# ══════════════════════════════════════════════════════════════════════════════

def _ar_threshold(dept, t):
    """截至 t 日（含）的应收三要素：累计上账 / 累计账实差额 / 累计回款（全部来源）。
    未收余额 = 上账 + 账实差额 − 累计回款。账实差额按记录运作日归期（随上账确认）。"""
    rec = ARRecord.objects.filter(delivery_dept=dept, operation_date__lte=t).aggregate(
        est=Sum('estimated_amount'), adj=Sum('account_diff_adjustment'))
    pay = (ARPayment.objects
           .filter(ar_record__delivery_dept=dept, payment_date__lte=t)
           .aggregate(s=Sum('amount'))['s'] or Decimal('0'))
    est = rec['est'] or Decimal('0')
    adj = rec['adj'] or Decimal('0')
    return est, adj, pay


def _ar_metric(depts, p_start, p_end):
    day_before = p_start - datetime.timedelta(days=1)
    rows = {}
    for d in depts:
        # 期初 / 期末未收
        e0, a0, p0 = _ar_threshold(d, day_before)
        e1, a1, p1 = _ar_threshold(d, p_end)
        opening = e0 + a0 - p0
        closing = e1 + a1 - p1
        new_ar = e1 - e0            # 本期新增上账
        adj_in = a1 - a0            # 本期账实差额
        pay_all_in = p1 - p0        # 本期全部回款（含非现金核销）

        # 本期现金回款（仅 source=回款）与非现金核销（预收抵扣 + 内部往来）
        cash_in = (ARPayment.objects
                   .filter(ar_record__delivery_dept=d, payment_date__gte=p_start,
                           payment_date__lte=p_end)
                   .exclude(source__in=NON_CASH_PAYMENT_SOURCES)
                   .aggregate(s=Sum('amount'))['s'] or Decimal('0'))
        noncash_in = pay_all_in - cash_in

        # 期末账龄拆分：对 operation_date<=end 的记录逐条算截至 end 的未收，按是否逾期归并
        pay_to_end_sq = Subquery(
            ARPayment.objects.filter(ar_record=OuterRef('pk'), payment_date__lte=p_end)
            .values('ar_record').annotate(s=Sum('amount')).values('s'))
        recs = (ARRecord.objects
                .filter(delivery_dept=d, operation_date__lte=p_end)
                .annotate(paid_end=Coalesce(pay_to_end_sq, Value(Decimal('0'),
                          output_field=DecimalField(max_digits=18, decimal_places=2))))
                .annotate(out_end=F('estimated_amount') + F('account_diff_adjustment') - F('paid_end')))
        overdue = (recs.filter(due_date__lt=p_end, out_end__gt=0)
                   .aggregate(s=Sum('out_end'))['s'] or Decimal('0'))
        not_due = (recs.filter(Q(due_date__gte=p_end) | Q(due_date__isnull=True), out_end__gt=0)
                   .aggregate(s=Sum('out_end'))['s'] or Decimal('0'))

        # 本期到期 & 到期回款率
        due_recs = ARRecord.objects.filter(delivery_dept=d, due_date__gte=p_start, due_date__lte=p_end)
        due_amt = due_recs.aggregate(s=Sum(F('estimated_amount') + F('account_diff_adjustment')))['s'] or Decimal('0')
        due_collected = (ARPayment.objects
                         .filter(ar_record__delivery_dept=d, ar_record__due_date__gte=p_start,
                                 ar_record__due_date__lte=p_end, payment_date__lte=p_end)
                         .exclude(source__in=NON_CASH_PAYMENT_SOURCES)
                         .aggregate(s=Sum('amount'))['s'] or Decimal('0'))

        rows[d] = {
            'opening': _money(opening),
            'new_ar': _money(new_ar),
            'cash_in': _money(cash_in),
            'noncash_in': _money(noncash_in),
            'adj_in': _money(adj_in),
            'closing': _money(closing),
            'overdue': _money(overdue),
            'not_due': _money(not_due),
            'due_amt': _money(due_amt),
            'due_collected': _money(due_collected),
            'due_rate': _rate(due_collected, due_amt),
            # 回款率：本期现金回款 ÷ 本期可回收应收（期初未收 + 本期新增）
            'collect_rate': _rate(cash_in, opening + new_ar),
        }
    return rows


# ══════════════════════════════════════════════════════════════════════════════
# 指标 3 / 4 — 应收 / 应付预算完成
# ══════════════════════════════════════════════════════════════════════════════

def _collection_actual(depts_q, s, e):
    return (ARPayment.objects
            .filter(ar_record__delivery_dept__in=depts_q, payment_date__gte=s, payment_date__lte=e)
            .exclude(source__in=NON_CASH_PAYMENT_SOURCES)
            .aggregate(x=Sum('amount'))['x'] or Decimal('0'))


def _payment_actual(depts_q, s, e):
    return (PaymentInstallment.objects
            .filter(payment__department__in=depts_q, pay_date__gte=s, pay_date__lte=e)
            .aggregate(x=Sum('pay_amount'))['x'] or Decimal('0'))


def _budget_metric(depts, year, p_start, p_end, month_start, month_end):
    """应收预算按汇报期间口径；应付预算按所在自然月口径（付款预算仅维护到月）。"""
    year_start = datetime.date(year, 1, 1)
    rows = {}
    for d in depts:
        dl = [d]
        # 应收预算（期间 + 年累计）
        cb_p = CollectionBudget.objects.filter(delivery_dept=d, expected_date__gte=p_start,
                                               expected_date__lte=p_end).aggregate(s=Sum('amount'))['s'] or Decimal('0')
        cb_y = CollectionBudget.objects.filter(delivery_dept=d, expected_date__gte=year_start,
                                               expected_date__lte=p_end).aggregate(s=Sum('amount'))['s'] or Decimal('0')
        ca_p = _collection_actual(dl, p_start, p_end)
        ca_y = _collection_actual(dl, year_start, p_end)
        # 应付预算（所在月 + 年累计）
        pb_m = PaymentBudget.objects.filter(delivery_dept=d, expected_date__gte=month_start,
                                            expected_date__lte=month_end).aggregate(s=Sum('amount'))['s'] or Decimal('0')
        pb_y = PaymentBudget.objects.filter(delivery_dept=d, expected_date__gte=year_start,
                                            expected_date__lte=month_end).aggregate(s=Sum('amount'))['s'] or Decimal('0')
        pa_m = _payment_actual(dl, month_start, month_end)
        pa_y = _payment_actual(dl, year_start, month_end)
        rows[d] = {
            'coll_budget': _money(cb_p), 'coll_actual': _money(ca_p), 'coll_rate': _rate(ca_p, cb_p),
            'coll_budget_ytd': _money(cb_y), 'coll_actual_ytd': _money(ca_y), 'coll_rate_ytd': _rate(ca_y, cb_y),
            'pay_budget': _money(pb_m), 'pay_actual': _money(pa_m), 'pay_rate': _rate(pa_m, pb_m),
            'pay_budget_ytd': _money(pb_y), 'pay_actual_ytd': _money(pa_y), 'pay_rate_ytd': _rate(pa_y, pb_y),
        }
    return rows


# ══════════════════════════════════════════════════════════════════════════════
# 指标 5 — 现金流（与「现金流分析」同口径，按期间切片）
# ══════════════════════════════════════════════════════════════════════════════

def _cash_window(depts, s, e):
    """[s,e] 区间的经营现金流：流入=现金回款+预收；流出=实付(扣预付冲抵)+预付。"""
    coll = (ARPayment.objects
            .filter(ar_record__delivery_dept__in=depts, payment_date__gte=s, payment_date__lte=e)
            .exclude(source__in=NON_CASH_PAYMENT_SOURCES)
            .aggregate(x=Sum('amount'))['x'] or Decimal('0'))
    paid = (PaymentInstallment.objects
            .filter(payment__department__in=depts, pay_date__gte=s, pay_date__lte=e)
            .aggregate(x=Sum('pay_amount'))['x'] or Decimal('0'))
    # 预付核销冲抵：按首个实付日期（否则计划付款日）归期
    earliest = Subquery(PaymentInstallment.objects.filter(payment_id=OuterRef('pk'))
                        .order_by('pay_date').values('pay_date')[:1])
    offset = (Payment.objects
              .filter(department__in=depts, prepaid_offset_amount__gt=0)
              .annotate(fp=earliest)
              .annotate(ad=Coalesce('fp', 'planned_date'))
              .filter(ad__gte=s, ad__lte=e)
              .aggregate(x=Sum('prepaid_offset_amount'))['x'] or Decimal('0'))
    paid = max(Decimal('0'), paid - offset)
    adv = (AdvanceRecord.objects
           .filter(delivery_dept__in=depts, occur_date__gte=s, occur_date__lte=e)
           .values('direction').annotate(x=Sum('advance_amount')))
    adv_recv = adv_paid = Decimal('0')
    for r in adv:
        if r['direction'] == '预收':
            adv_recv += r['x'] or Decimal('0')
        else:
            adv_paid += r['x'] or Decimal('0')
    inflow = coll + adv_recv
    outflow = paid + adv_paid
    return {
        'collected': _money(coll), 'advance_received': _money(adv_recv),
        'paid': _money(paid), 'advance_paid': _money(adv_paid),
        'inflow': _money(inflow), 'outflow': _money(outflow),
        'net': _money(inflow - outflow),
    }


def _cash_metric(depts, year, p_start, p_end):
    year_start = datetime.date(year, 1, 1)
    period = _cash_window(depts, p_start, p_end)
    ytd = _cash_window(depts, year_start, p_end)
    return {'period': period, 'ytd': ytd}


# ══════════════════════════════════════════════════════════════════════════════
# 组装
# ══════════════════════════════════════════════════════════════════════════════

def _sum_rows(rows, depts, keys):
    """把多个事业部行按 key 求和，得到「合计」行（仅对金额/计数类 key）。"""
    out = {}
    for k in keys:
        out[k] = sum((rows[d].get(k) or 0) for d in depts)
    return out


def _compute_report(request, today):
    """计算一份完整周期报表，返回 dict（供 JSON 与 Excel 共用）。"""
    ptype, p_start, p_end, year, month, wom = _resolve_period(request, today)
    pv_start, pv_end = _prev_period(ptype, p_start, p_end)
    month_start = datetime.date(year, month, 1)
    month_end = datetime.date(year, month, calendar.monthrange(year, month)[1])

    # 权限范围
    if request.pk_role == 'super_admin':
        allowed = list(VALID_DEPARTMENTS)
    else:
        allowed = [d for d in (request.pk_depts or []) if d in VALID_DEPARTMENTS]

    # scope：group=可见范围合计；或指定单个事业部
    scope = (request.GET.get('scope') or '').strip()
    dept_param = (request.GET.get('dept') or '').strip()
    if dept_param and dept_param in allowed:
        depts = [dept_param]
        scope_kind = 'dept'
        scope_name = dept_param
    elif scope == 'group' or len(allowed) > 1:
        depts = list(allowed)
        scope_kind = 'group'
        scope_name = '集团财务中心' if request.pk_role == 'super_admin' else '本部门合计'
    else:
        depts = list(allowed)
        scope_kind = 'dept'
        scope_name = allowed[0] if allowed else '—'

    if not depts:
        return None, {'error': '无可见事业部范围'}

    projects = _projects_metric(depts, year, p_start, p_end, pv_start, pv_end)
    ar = _ar_metric(depts, p_start, p_end)
    budget = _budget_metric(depts, year, p_start, p_end, month_start, month_end)
    cash = _cash_metric(depts, year, p_start, p_end)

    # 合计行
    proj_total = _sum_rows(projects, depts, ['active', 'total', 'ytd_new', 'period_new', 'prev_new', 'delta'])
    ar_keys = ['opening', 'new_ar', 'cash_in', 'noncash_in', 'adj_in', 'closing',
               'overdue', 'not_due', 'due_amt', 'due_collected']
    ar_total = _sum_rows(ar, depts, ar_keys)
    ar_total['due_rate'] = _rate(ar_total['due_collected'], ar_total['due_amt'])
    ar_total['collect_rate'] = _rate(ar_total['cash_in'], ar_total['opening'] + ar_total['new_ar'])
    bud_keys = ['coll_budget', 'coll_actual', 'coll_budget_ytd', 'coll_actual_ytd',
                'pay_budget', 'pay_actual', 'pay_budget_ytd', 'pay_actual_ytd']
    bud_total = _sum_rows(budget, depts, bud_keys)
    bud_total['coll_rate'] = _rate(bud_total['coll_actual'], bud_total['coll_budget'])
    bud_total['coll_rate_ytd'] = _rate(bud_total['coll_actual_ytd'], bud_total['coll_budget_ytd'])
    bud_total['pay_rate'] = _rate(bud_total['pay_actual'], bud_total['pay_budget'])
    bud_total['pay_rate_ytd'] = _rate(bud_total['pay_actual_ytd'], bud_total['pay_budget_ytd'])

    reporter = getattr(request.pk_user, 'name', '') or '—'

    # 标题
    if ptype == 'weekly':
        period_label = f'周报（{year}年{month}月第{wom}周）'
    else:
        period_label = f'月报（{year}年{month}月）'
    title = f'{scope_name} {period_label}'

    # 可选范围（超管：集团 + 各事业部）
    scopes_available = []
    if request.pk_role == 'super_admin' or len(allowed) > 1:
        scopes_available.append({'kind': 'group', 'value': 'group',
                                 'label': '集团财务中心' if request.pk_role == 'super_admin' else '合计'})
    for d in allowed:
        scopes_available.append({'kind': 'dept', 'value': d, 'label': d})

    data = {
        'meta': {
            'title': title,
            'scope_kind': scope_kind,
            'scope_name': scope_name,
            'period_type': ptype,
            'period_label': period_label,
            'year': year, 'month': month, 'week_of_month': wom,
            'start_date': str(p_start), 'end_date': str(p_end),
            'month_start': str(month_start), 'month_end': str(month_end),
            'reporter': reporter,
            'reviewer': DEFAULT_REVIEWER,
            'report_date': str(today),
            'depts': depts,
            'is_multi': len(depts) > 1,
        },
        'scopes_available': scopes_available,
        'projects': {'rows': projects, 'total': proj_total},
        'ar': {'rows': ar, 'total': ar_total},
        'budget': {'rows': budget, 'total': bud_total},
        'cash': cash,
    }
    return data, None


# ══════════════════════════════════════════════════════════════════════════════
# 端点：JSON
# ══════════════════════════════════════════════════════════════════════════════

@csrf_exempt
@pk_required()
def periodic_report(request):
    denied = _page_denied(request, 'ar_analytics')
    if denied:
        return denied
    today = datetime.date.today()
    data, e = _compute_report(request, today)
    if e:
        return err(e['error'])
    return ok(data)


# ══════════════════════════════════════════════════════════════════════════════
# 端点：Excel 导出
# ══════════════════════════════════════════════════════════════════════════════

def _xlsx_report(data):
    from openpyxl.styles import Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '周期报表'
    m = data['meta']
    thin = Side(style='thin', color='D9C9B6')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    BR = 'C96342'   # 品牌色
    DK = '3C2011'

    def section(ws, r, text):
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = PatternFill('solid', fgColor=BR)
        c.alignment = Alignment(horizontal='left', vertical='center')
        return r + 1

    def hdr(ws, r, cols, color=DK):
        for i, h in enumerate(cols, 1):
            c = ws.cell(row=r, column=i, value=h)
            c.font = Font(bold=True, color='FFFFFF', size=10)
            c.fill = PatternFill('solid', fgColor=color)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border
        return r + 1

    def row(ws, r, vals, bold=False):
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = border
            if bold:
                c.font = Font(bold=True)
            if i > 1 and isinstance(v, (int, float)):
                c.alignment = Alignment(horizontal='right')
                c.number_format = '#,##0'
        return r + 1

    ncol = 9
    # 标题区
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    t = ws.cell(row=1, column=1, value=m['title'])
    t.font = Font(bold=True, size=16, color=DK)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    sub = ws.cell(row=2, column=1,
                  value=f"汇报人：{m['reporter']}    审核人：{m['reviewer']}    "
                        f"汇报日期：{m['report_date']}    取值期间：{m['start_date']} ~ {m['end_date']}")
    sub.alignment = Alignment(horizontal='center')
    sub.font = Font(size=10, color='7A5A40')

    r = 4
    P = data['projects']['total']
    r = section(ws, r, '一、项目规模')
    r = hdr(ws, r, ['事业部', '在运项目', '年初至今新签', '本期新签', '上期新签', '较上期增减', '项目总数'])
    rows = data['projects']['rows']
    for d in m['depts']:
        pr = rows[d]
        r = row(ws, r, [d, pr['active'], pr['ytd_new'], pr['period_new'], pr['prev_new'], pr['delta'], pr['total']])
    if m['is_multi']:
        r = row(ws, r, ['合计', P['active'], P['ytd_new'], P['period_new'], P['prev_new'], P['delta'], P['total']], bold=True)

    r += 1
    r = section(ws, r, '二、应收账款情况')
    r = hdr(ws, r, ['事业部', '期初未收', '本期新增', '现金回款', '非现金核销', '账实差额',
                    '期末未收', '其中逾期', '其中未到期'])
    arows = data['ar']['rows']
    for d in m['depts']:
        a = arows[d]
        r = row(ws, r, [d, a['opening'], a['new_ar'], a['cash_in'], a['noncash_in'], a['adj_in'],
                        a['closing'], a['overdue'], a['not_due']])
    if m['is_multi']:
        a = data['ar']['total']
        r = row(ws, r, ['合计', a['opening'], a['new_ar'], a['cash_in'], a['noncash_in'], a['adj_in'],
                        a['closing'], a['overdue'], a['not_due']], bold=True)
    # 到期回款
    r = hdr(ws, r, ['事业部', '本期到期应收', '本期到期已回', '到期回款率(%)'])
    for d in m['depts']:
        a = arows[d]
        r = row(ws, r, [d, a['due_amt'], a['due_collected'], a['due_rate']])
    if m['is_multi']:
        a = data['ar']['total']
        r = row(ws, r, ['合计', a['due_amt'], a['due_collected'], a['due_rate']], bold=True)

    r += 1
    r = section(ws, r, '三、应收预算完成')
    r = hdr(ws, r, ['事业部', '本期预算', '本期实际', '本期完成率(%)', '年度预算', '年度实际', '年度完成率(%)'])
    brows = data['budget']['rows']
    for d in m['depts']:
        b = brows[d]
        r = row(ws, r, [d, b['coll_budget'], b['coll_actual'], b['coll_rate'],
                        b['coll_budget_ytd'], b['coll_actual_ytd'], b['coll_rate_ytd']])
    if m['is_multi']:
        b = data['budget']['total']
        r = row(ws, r, ['合计', b['coll_budget'], b['coll_actual'], b['coll_rate'],
                        b['coll_budget_ytd'], b['coll_actual_ytd'], b['coll_rate_ytd']], bold=True)

    r += 1
    r = section(ws, r, '四、应付预算完成（月度口径）')
    r = hdr(ws, r, ['事业部', '本月预算', '本月实际', '本月完成率(%)', '年度预算', '年度实际', '年度完成率(%)'])
    for d in m['depts']:
        b = brows[d]
        r = row(ws, r, [d, b['pay_budget'], b['pay_actual'], b['pay_rate'],
                        b['pay_budget_ytd'], b['pay_actual_ytd'], b['pay_rate_ytd']])
    if m['is_multi']:
        b = data['budget']['total']
        r = row(ws, r, ['合计', b['pay_budget'], b['pay_actual'], b['pay_rate'],
                        b['pay_budget_ytd'], b['pay_actual_ytd'], b['pay_rate_ytd']], bold=True)

    r += 1
    r = section(ws, r, '五、现金流情况')
    r = hdr(ws, r, ['项目', '本期金额', '本年累计'])
    cp, cy = data['cash']['period'], data['cash']['ytd']
    r = row(ws, r, ['一、经营活动现金流入', cp['inflow'], cy['inflow']], bold=True)
    r = row(ws, r, ['　现金回款', cp['collected'], cy['collected']])
    r = row(ws, r, ['　预收款', cp['advance_received'], cy['advance_received']])
    r = row(ws, r, ['二、经营活动现金流出', cp['outflow'], cy['outflow']], bold=True)
    r = row(ws, r, ['　实付款项（扣预付冲抵）', cp['paid'], cy['paid']])
    r = row(ws, r, ['　预付款', cp['advance_paid'], cy['advance_paid']])
    r = row(ws, r, ['三、经营活动净现金流', cp['net'], cy['net']], bold=True)

    # 列宽
    ws.column_dimensions['A'].width = 22
    for col in 'BCDEFGHI':
        ws.column_dimensions[col].width = 14
    return wb


@csrf_exempt
@pk_required()
def periodic_report_export(request):
    denied = _page_denied(request, 'ar_analytics')
    if denied:
        return denied
    today = datetime.date.today()
    data, e = _compute_report(request, today)
    if e:
        return err(e['error'])
    reviewer_override = (request.GET.get('reviewer') or '').strip()
    if reviewer_override:
        data['meta']['reviewer'] = reviewer_override
    wb = _xlsx_report(data)
    fname = f"{data['meta']['title']}.xlsx"
    return _export_response(wb, fname)


__all__ = [n for n in dir() if not n.startswith('__')]
