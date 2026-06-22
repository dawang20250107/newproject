import io
import json
import logging
import re
import datetime
import calendar
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from urllib.parse import quote

logger = logging.getLogger(__name__)

from django.core.exceptions import ValidationError
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from django.db.models import Sum, Count, Q, F, Value, Case, When, IntegerField, CharField, DecimalField, Max, Min
from django.db.models.functions import TruncMonth, Coalesce
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from paikuan.views import (pk_required, ok, err, DEPARTMENTS, VALID_DEPARTMENTS,
                           get_request_perms, apply_ar_view_mask, AR_PROJECT_FIELD_DEFS,
                           AR_RECORD_FIELD_DEFS, AR_ADVANCE_FIELD_DEFS, _paid_subq,
                           _ai_review_records)
from ar.models import (ARProject, ARRecord, ARPayment, ARAdjustment,
                       NON_CASH_PAYMENT_SOURCES,
                       BatchInvoiceEvent,
                       CollectionBudget, PaymentBudget,
                       AdvanceRecord, AdvanceWriteoff, AdvanceInstallment,
                       Supplier, Customer,
                       Contract, ContractParty, ContractProject, ActionItem,
                       CashPoolConfig, CashPoolTransfer)
from paikuan.models import Payment, PaymentInstallment, ApprovalRecord

# ── AR page keys (must match PAGE_KEYS in paikuan/views.py) ───────────────────
AR_PAGE_KEYS = ['ar_projects', 'ar_records', 'ar_advance', 'ar_analytics', 'ar_cashflow', 'ar_budget']

ADVANCE_DIRECTIONS = ['预收', '预付']

EXAMPLE_ROW_MARKER = '示例-导入前请删除此行'
VALID_CUSTOMER_LEVELS = ['S级', 'A级', 'B级', 'C级', 'D级']

PRECHECK_MAX = 10000   # 超过此行数跳过预检、直接导入


def _ar_ai_review(records, system_prompt):
    """Best-effort AI review for AR module import rows. Wraps the generic paikuan helper."""
    return _ai_review_records(records, system_prompt)


def _ar_precheck_report(report_rows, columns):
    """将 [{row, data, ruleIssue, warn, ai}] 列表打包成前端预检 Modal 需要的 response shape。"""
    ai_findings = sum(len(r['ai']) for r in report_rows)
    attention, ok_data, rule_errors = [], [], 0
    for r in report_rows:
        if r['ruleIssue']:
            rule_errors += 1
            r['status'] = 'error'
            attention.append(r)
        elif r['ai']:
            r['status'] = 'review'
            attention.append(r)
        else:
            ok_data.append(r['data'])
    import django.conf as _dc
    ai_enabled = bool(getattr(_dc.settings, 'DEEPSEEK_API_KEY', ''))
    return {
        'total': len(report_rows), 'attention': len(attention), 'okCount': len(ok_data),
        'ruleErrors': rule_errors, 'warns': 0, 'aiFindings': ai_findings,
        'aiEnabled': ai_enabled, 'columns': columns,
        'rows': attention, 'okRows': [],   # AR: re-submit file on confirm, no JSON apply needed
    }
VALID_INVOICE_TYPES = ['专票', '普票', '不开票']


def _norm_header(h):
    """规范化 Excel 表头：去空格、去必填星号、去括号注释，使模板(带*)与
    导出(不带*)的表头可双向匹配。例：'合同名称*'、'开票模式(全额/差额)' →
    '合同名称'、'开票模式'。"""
    s = str(h or '').strip().replace('＊', '').replace('*', '')
    s = s.replace('（', '(').replace('）', ')')
    s = re.sub(r'\([^)]*\)', '', s)   # 去掉括号及其中的注释
    return s.strip()


def _parse_cycle_start_day(v):
    """对账周期起始日：1~28（1=自然月）。返回 (值, 错误文案)。空值按 1。"""
    if v in (None, ''):
        return 1, None
    try:
        d = int(float(v))
    except (TypeError, ValueError):
        return None, f'对账周期起始日"{v}"无效，须为 1~28 的整数（1=自然月）'
    if not (1 <= d <= 28):
        return None, f'对账周期起始日须为 1~28（当前 {d}）。29~31 在短月不存在，请用 28 或改约对账日'
    return d, None


def _match_project_by_short_name(short_name, dept='', allowed_depts=None):
    """Simple lookup used by non-import paths (project search, etc.)."""
    name = (short_name or '').strip()
    if not name:
        return None
    qs = ARProject.objects.filter(short_name=name)
    if dept:
        qs = qs.filter(delivery_dept=dept)
    elif allowed_depts is not None:
        qs = qs.filter(delivery_dept__in=allowed_depts)
    proj = qs.order_by('-id').first()
    if proj:
        return proj
    qs = ARProject.objects.filter(short_name__icontains=name)
    if dept:
        qs = qs.filter(delivery_dept=dept)
    elif allowed_depts is not None:
        qs = qs.filter(delivery_dept__in=allowed_depts)
    return qs.order_by('-id').first()


def _proj_candidate_dict(p):
    """歧义候选项的精简信息，供导入返回让用户挑选/区分。"""
    return {
        'project_no': p.project_no,
        'short_name': p.short_name,
        'customer_name': p.customer_name,
        'delivery_dept': p.delivery_dept,
    }


def _classify_project_for_import(short_name, customer_hint, project_no, allowed_depts, dept_hint=''):
    """只读判定项目归属（不建库、不创建草稿），用于导入前的歧义检查。

    返回 (status, payload)：
      'resolved'  -> payload = ARProject（唯一命中）
      'ambiguous' -> payload = 候选 list[dict]（多命中且无法区分，需人工指定）
      'bad_no'    -> payload = 填写的项目编号（填了但查不到/越权）
      'not_found' -> payload = None（找不到，写入阶段将自动建草稿）

    优先级：项目编号(确定性) > 精确简称(+客户名缩小) > 模糊简称(+客户名缩小)。
    """
    name = (short_name or '').strip()
    pno = (project_no or '').strip()
    dh = (dept_hint or '').strip()

    # ── 0) 项目编号：最高优先级、确定性指定 ────────────────────────────────
    if pno:
        qs = ARProject.objects.filter(project_no=pno)
        if allowed_depts is not None:
            qs = qs.filter(delivery_dept__in=allowed_depts)
        proj = qs.first()
        return ('resolved', proj) if proj else ('bad_no', pno)

    def _narrow(qs):
        # 交付部门是业务主键的一部分（短名+部门），优先用它区分同名跨部门项目
        if dh:
            nd = qs.filter(delivery_dept=dh)
            if nd.exists():
                qs = nd
        if customer_hint:
            n = qs.filter(customer_name__icontains=customer_hint)
            if n.exists():
                return n
        return qs

    # ── 1) 精确简称 ────────────────────────────────────────────────────────
    qs = ARProject.objects.filter(short_name=name)
    if allowed_depts is not None:
        qs = qs.filter(delivery_dept__in=allowed_depts)
    qs = _narrow(qs)
    cnt = qs.count()
    if cnt == 1:
        return 'resolved', qs.first()
    if cnt > 1:
        return 'ambiguous', [_proj_candidate_dict(p) for p in qs.order_by('-id')[:8]]

    # ── 2) 模糊简称 ────────────────────────────────────────────────────────
    qs2 = ARProject.objects.filter(short_name__icontains=name)
    if allowed_depts is not None:
        qs2 = qs2.filter(delivery_dept__in=allowed_depts)
    qs2 = _narrow(qs2)
    cnt2 = qs2.count()
    if cnt2 == 1:
        return 'resolved', qs2.first()
    if cnt2 > 1:
        return 'ambiguous', [_proj_candidate_dict(p) for p in qs2.order_by('-id')[:8]]

    # ── 3) 找不到 → 写入阶段自动建草稿 ─────────────────────────────────────
    return 'not_found', None


def _resolve_project_for_import(short_name, customer_hint, dept, allowed_depts, user,
                                proj_cache, project_no=''):
    """3-stage project resolution for bulk import.

    Returns (proj, match_type, warning_str | None).
    match_type: 'exact' | 'exact_multi' | 'fuzzy' | 'fuzzy_multi' | 'created'

    Uses proj_cache {(short_name, dept) -> result} to avoid creating duplicate
    draft projects when multiple rows share the same project name.
    若提供 project_no（项目编号），优先按编号精确命中，作为确定性指定。
    """
    name = (short_name or '').strip()
    if not name:
        return None, None, None
    cache_key = (name, dept or '')

    # 项目编号确定性命中（不进 proj_cache，按编号逐行解析）
    pno = (project_no or '').strip()
    if pno:
        qs_no = ARProject.objects.filter(project_no=pno)
        if allowed_depts is not None:
            qs_no = qs_no.filter(delivery_dept__in=allowed_depts)
        proj_no = qs_no.first()
        if proj_no:
            return proj_no, 'exact', None

    if cache_key in proj_cache:
        return proj_cache[cache_key]

    # ── Stage 1: 精确匹配 ──────────────────────────────────────────────────
    qs = ARProject.objects.filter(short_name=name)
    if allowed_depts is not None:
        qs = qs.filter(delivery_dept__in=allowed_depts)
    # 交付部门是业务主键的一部分：用它区分同名跨部门项目（在权限范围内进一步缩小）
    if dept:
        qs_dept = qs.filter(delivery_dept=dept)
        if qs_dept.exists():
            qs = qs_dept

    # Customer hint可进一步缩小范围（同名项目不同客户时）
    if customer_hint:
        qs_cust = qs.filter(customer_name__icontains=customer_hint)
        if qs_cust.exists():
            cnt = qs_cust.count()
            proj = qs_cust.order_by('-id').first()
            warn = (f'找到 {cnt} 个同名同客户项目，已取最新的「{proj.customer_name}」，请核查'
                    if cnt > 1 else None)
            result = (proj, 'exact' if cnt == 1 else 'exact_multi', warn)
            proj_cache[cache_key] = result
            return result

    if qs.exists():
        cnt = qs.count()
        proj = qs.order_by('-id').first()
        warn = (f'找到 {cnt} 个同名项目，已取最新的「{proj.customer_name}」，请核查'
                if cnt > 1 else None)
        result = (proj, 'exact' if cnt == 1 else 'exact_multi', warn)
        proj_cache[cache_key] = result
        return result

    # ── Stage 2: 模糊匹配（contains）──────────────────────────────────────
    qs2 = ARProject.objects.filter(short_name__icontains=name)
    if allowed_depts is not None:
        qs2 = qs2.filter(delivery_dept__in=allowed_depts)
    if dept:
        qs2_dept = qs2.filter(delivery_dept=dept)
        if qs2_dept.exists():
            qs2 = qs2_dept
    matches = list(qs2.order_by('-id')[:5])
    if matches:
        proj = matches[0]
        if len(matches) > 1:
            candidates = '、'.join(f'「{m.short_name}」' for m in matches[:3])
            warn = f'"{name}"模糊匹配到多个项目（{candidates}），已取最新的「{proj.short_name}」，建议核查'
            match_type = 'fuzzy_multi'
        else:
            warn = f'"{name}"未精确匹配，模糊命中「{proj.short_name}」，建议核查'
            match_type = 'fuzzy'
        result = (proj, match_type, warn)
        proj_cache[cache_key] = result
        return result

    # ── Stage 3: 自动创建草稿项目 ─────────────────────────────────────────
    proj_dept = dept or (list(allowed_depts)[0] if allowed_depts else '')
    creator_name = user.name if user else '待填'
    proj = ARProject(
        short_name=name,
        customer_name=customer_hint or name,
        delivery_dept=proj_dept,
        is_draft=True,
        sales_contact=creator_name,
        project_manager=creator_name,
        has_contract='无',
    )
    proj.save()
    warn = f'项目台账中未找到「{name}」，已自动创建草稿项目，请到项目台账补充完善'
    result = (proj, 'created', warn)
    proj_cache[cache_key] = result
    return result


def _normalize_date(s):
    """Best-effort normalize a user-typed date string to ISO YYYY-MM-DD.

    Accepts: 2026-01-15, 2026/1/15, 2026.1.15, 2026年1月15日, 26/1/15,
    20260115, datetime/date objects, Excel serial numbers.
    """
    import datetime as _dt
    if s is None:
        return None
    # datetime / date objects (openpyxl can return these)
    if isinstance(s, (_dt.datetime, _dt.date)):
        return s.strftime('%Y-%m-%d')
    # Excel serial date (number of days since 1899-12-30)
    if isinstance(s, (int, float)) and not isinstance(s, bool):
        try:
            base = _dt.date(1899, 12, 30)
            return (base + _dt.timedelta(days=int(s))).strftime('%Y-%m-%d')
        except (OverflowError, ValueError):
            return None
    s = str(s).strip()
    if not s or s.lower() == 'none':
        return None
    # 20260115 → 2026-01-15
    if re.fullmatch(r'\d{8}', s):
        try:
            return f'{s[:4]}-{s[4:6]}-{s[6:]}'
        except ValueError:
            pass
    # Drop chinese unit suffixes / unify separators
    cleaned = (s.replace('/', '-').replace('.', '-')
                .replace('年', '-').replace('月', '-').replace('日', '')
                .replace(' ', '').strip('-'))
    parts = [p for p in cleaned.split('-') if p]
    try:
        if len(parts) >= 3:
            y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
            # 2-digit year → 20xx
            if y < 100:
                y += 2000
            _dt.date(y, m, d)  # validate
            return f'{y:04d}-{m:02d}-{d:02d}'
    except (ValueError, TypeError):
        pass
    return None


def _dec(v, default=Decimal('0')):
    try:
        return Decimal(str(v))
    except (InvalidOperation, TypeError):
        return default


def _int_param(request, key, default):
    """GET 参数安全转 int：非法值回退默认值（避免 ?year=abc 之类把接口打成500）。"""
    try:
        return int(request.GET.get(key, default) or default)
    except (ValueError, TypeError):
        return int(default)


def _parse_body(request):
    try:
        return json.loads(request.body)
    except Exception:
        return {}


def _ar_dept_filter(qs, request, dept_field='delivery_dept', shared_field=None):
    """Filter queryset by the requesting user's allowed departments.
    Optional ?depts=A,B,C further narrows the set (intersected with pk_depts).
    When shared_field is provided and the user has ar_shared_only, restricts to
    shared projects only (shared_field=True)."""
    raw = request.GET.get('depts', '').strip()
    requested = [d for d in raw.split(',') if d.strip()] if raw else []

    if request.pk_role == 'super_admin':
        if requested:
            qs = qs.filter(**{f'{dept_field}__in': requested})
        return qs

    allowed = set(request.pk_depts or [])
    if not allowed:
        return qs.none()
    if requested:
        active = [d for d in requested if d in allowed]
        if active:
            qs = qs.filter(**{f'{dept_field}__in': active})
        else:
            qs = qs.filter(**{f'{dept_field}__in': request.pk_depts})
    else:
        qs = qs.filter(**{f'{dept_field}__in': request.pk_depts})

    if shared_field:
        from paikuan.views import get_request_perms
        perms = get_request_perms(request)
        if perms and perms.get('ar_shared_only'):
            qs = qs.filter(**{shared_field: True})
    return qs


def _page_denied(request, page_key):
    """Return error response if user lacks access to a page; None if allowed."""
    if request.pk_role == 'super_admin':
        return None
    from paikuan.views import get_job_perms
    jt = getattr(request, 'pk_job', '') or ''
    if not jt:
        from paikuan.models import PaikuanUser
        u = PaikuanUser.objects.filter(id=request.pk_uid).only('job_title').first()
        jt = u.job_title if u else ''
    perms = get_job_perms(jt)
    if not perms['pages'].get(page_key, True):
        return err('无权访问此模块', 403)
    return None


def _write_denied(request):
    perms = get_request_perms(request)
    # AR 写入：优先看 ar_can_create（结算会计等聚焦应收的岗位），回退到通用 can_create。
    if perms is not None and not (perms.get('ar_can_create') or perms.get('can_create', False)):
        return err('无写入权限', 403, 403)
    return None


def _action_denied(request, action_key):
    """操作级权限：核销/回款录入等细粒度动作的独立开关。

    actions[key] 显式配置时以其为准（True 放行 / False 拒绝）；
    旧配置无 actions 键时回退通用写权限（向后兼容，不会突然把人锁外面）。
    出纳等岗位可借此单独获得「预付核销」而不必放开整个 can_create。"""
    perms = get_request_perms(request)
    if perms is None:
        return None
    acts = perms.get('actions')
    if isinstance(acts, dict) and action_key in acts:
        if acts.get(action_key):
            return None
        return err('当前职务未开通此操作权限，请联系管理员在「权限配置·操作权限」中开通', 403, 403)
    return _write_denied(request)


def _delete_denied(request):
    perms = get_request_perms(request)
    if perms is not None and not perms.get('can_delete', False):
        return err('无删除权限', 403, 403)
    return None


def _dept_denied(request, dept, msg='无权操作此部门'):
    if request.pk_role == 'super_admin':
        return None
    if dept not in request.pk_depts:
        return err(msg, 403, 403)
    return None


def _sync_project_contracts(proj, contract_ids, request):
    """替换某项目挂靠的合同（项目侧维护多对多）。只挂用户有权部门的合同。"""
    proj.contract_links.all().delete()
    seen = set()
    for cid in (contract_ids or []):
        try:
            cid = int(cid)
        except (ValueError, TypeError):
            continue
        if cid in seen:
            continue
        ct = Contract.objects.filter(pk=cid).first()
        if not ct:
            continue
        if (request.pk_role != 'super_admin' and ct.delivery_dept
                and ct.delivery_dept not in request.pk_depts):
            continue
        seen.add(cid)
        ContractProject.objects.create(contract=ct, project=proj, is_primary=True)


def _project_contracts_list(proj):
    return [
        {'contract_id': cp.contract_id, 'name': cp.contract.name,
         'contract_no': cp.contract.contract_no, 'is_primary': cp.is_primary}
        for cp in proj.contract_links.select_related('contract').all()
    ]


def _object_dept_denied(request, obj, dept_field='delivery_dept'):
    return _dept_denied(request, getattr(obj, dept_field, ''), '无权访问')


def _can_ar_view(request, field_key):
    perms = get_request_perms(request)
    if perms is None:
        return True
    return (perms.get('ar_view') or {}).get(field_key, True)


def _ar_field_denied(request, field_key):
    if not _can_ar_view(request, field_key):
        return err('无权访问此字段', 403, 403)
    return None


_AR_PAYLOAD_DEFS_BY_GROUP = {
    'project': AR_PROJECT_FIELD_DEFS,
    'record': AR_RECORD_FIELD_DEFS,
    'advance': AR_ADVANCE_FIELD_DEFS,
}


def _ar_visible_payload(request, data, group, extra=()):
    perms = get_request_perms(request)
    if perms is None:
        return data
    defs = _AR_PAYLOAD_DEFS_BY_GROUP.get(group, AR_RECORD_FIELD_DEFS)
    ar_view = perms.get('ar_view') or {}
    cols = set(extra)
    for f in defs:
        if ar_view.get(f['key'], True):
            cols.update(f['cols'])
    return {k: v for k, v in data.items() if k in cols}


def _visible_ar_export_cols(request, columns):
    perms = get_request_perms(request)
    if perms is None:
        return columns
    ar_view = perms.get('ar_view') or {}
    return [col for col in columns if col[0] is None or ar_view.get(col[0], True)]


_XL_FORMULA_CHARS = ('=', '+', '-', '@', '\t', '\r')


def _export_response(wb, filename):
    # Excel 公式注入防护（与排款导出口径一致）：全部工作表的文本单元格，
    # 以 =+-@ 等开头的前置单引号转义。集中在导出总出口做，覆盖所有 AR 导出。
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v and v[0] in _XL_FORMULA_CHARS:
                    cell.value = "'" + v
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    safe = quote(filename, safe='')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{safe}"
    resp['Cache-Control'] = 'no-store'
    return resp


def _header_row(ws, headers, color='1565C0'):
    fill = PatternFill('solid', fgColor=color)
    font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')


# ── 共享：AR 记录过滤 / 状态 / 排序 / 条件助手（records、budget 等域共用）──
def _apply_record_filters(qs, request):
    """Shared dimension filters for AR records (used by list + kpi)."""
    project_id = request.GET.get('project_id', '').strip()
    if project_id:
        qs = qs.filter(project_id=int(project_id))
    dept = request.GET.get('dept', '').strip()
    if dept:
        qs = qs.filter(delivery_dept=dept)
    year = request.GET.get('year', '').strip()
    if year:
        qs = qs.filter(operation_year=int(year))
    month = request.GET.get('month', '').strip()
    if month:
        qs = qs.filter(operation_month=int(month))
    manager = request.GET.get('manager', '').strip()
    if manager:
        qs = qs.filter(project__project_manager__icontains=manager)
    is_shared = request.GET.get('is_shared', '').strip()
    if is_shared in ('1', 'true'):
        qs = qs.filter(project__is_shared=True)
    elif is_shared in ('0', 'false'):
        qs = qs.filter(project__is_shared=False)
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(project__short_name__icontains=q) |
            Q(project__customer_name__icontains=q) |
            Q(project__project_no__icontains=q) |
            Q(project__project_manager__icontains=q))
    return qs


def _apply_record_state_filters(qs, request, today=None):
    """Status / invoice / reconciliation / due-date filters for AR records.

    Used by the list view, the export view and the group-summary view so the
    same filter semantics apply everywhere. NOT used by the KPI endpoint, which
    computes its own status breakdown from the unfiltered (by-state) set.
    """
    if today is None:
        today = datetime.date.today()
    eomonth_today = datetime.date(today.year, today.month,
                                  calendar.monthrange(today.year, today.month)[1])

    status = request.GET.get('status', '').strip()
    if status == 'overdue':
        qs = qs.filter(outstanding_amount__gt=0, due_date__lt=today)
    elif status == 'current':
        qs = qs.filter(outstanding_amount__gt=0, due_date__gte=today,
                       due_date__lte=eomonth_today)
    elif status == 'not_due':
        qs = qs.filter(outstanding_amount__gt=0, due_date__gt=eomonth_today)
    elif status == 'settled':
        qs = qs.filter(outstanding_amount__lte=0)
    elif status == 'outstanding':
        qs = qs.filter(outstanding_amount__gt=0)

    inv_status = request.GET.get('invoice_status', '').strip()
    # 与 ARRecord.invoice_status 属性保持一致：已结清（outstanding<=0）优先级最高，
    # 因此「未开票」必须排除已结清记录（未开票 = 未开票 且 仍有未收余额）。
    if inv_status == '未开票':
        qs = qs.filter(actual_invoice_amount__isnull=True, outstanding_amount__gt=0)
    elif inv_status == '已结清':
        qs = qs.filter(outstanding_amount__lte=0)
    elif inv_status == '已开票':
        qs = qs.filter(actual_invoice_amount__isnull=False, outstanding_amount__gt=0)

    recon_status = request.GET.get('reconciliation_status', '').strip()
    if recon_status == '已对账':
        qs = qs.filter(Q(reconciliation_date__isnull=False) | Q(actual_invoice_amount__isnull=False))
    elif recon_status == '未对账':
        qs = qs.filter(reconciliation_date__isnull=True, actual_invoice_amount__isnull=True)

    # 责任状态（responsibility phase）— 按责任归属阶段过滤，与明细列的
    # post_invoice_status 责任链一致。逾期/等待中的细分依赖日期运算，这里只过滤到
    # 责任方所属阶段（settled / 对账阶段 / 开票阶段 / 票后回款阶段），可纯 DB 表达。
    responsibility = request.GET.get('responsibility', '').strip()
    if responsibility:
        no_inv = Q(project__invoice_type='不开票')
        if responsibility == 'settled':
            qs = qs.filter(outstanding_amount__lte=0)
        elif responsibility == 'post':
            # 票后/回款阶段（销售对接人责任）
            qs = qs.filter(
                Q(outstanding_amount__gt=0) & (
                    (no_inv & Q(reconciliation_date__isnull=False)) |
                    (~no_inv & Q(invoice_date__isnull=False))
                )
            )
        elif responsibility == 'invoice':
            # 待开票阶段（开票人责任）
            qs = qs.filter(
                Q(outstanding_amount__gt=0) & ~no_inv & Q(invoice_date__isnull=True) &
                (Q(reconciliation_date__isnull=False) | Q(actual_invoice_amount__isnull=False))
            )
        elif responsibility == 'recon':
            # 对账阶段（非销售责任）
            qs = qs.filter(
                Q(outstanding_amount__gt=0) & (
                    (no_inv & Q(reconciliation_date__isnull=True)) |
                    (~no_inv & Q(invoice_date__isnull=True) &
                     Q(reconciliation_date__isnull=True) & Q(actual_invoice_amount__isnull=True))
                )
            )

    # 回款筛选：pay_status='unpaid' 纯无回款；pay_include_unpaid=1 与日期区间做 OR
    # （"3月回款 + 含未结清"→传 pay_start/pay_end + pay_include_unpaid=1）
    pay_status = request.GET.get('pay_status', '').strip()
    pay_include_unpaid = request.GET.get('pay_include_unpaid', '') in ('1', 'true')
    pay_start = request.GET.get('pay_start', '').strip()
    pay_end = request.GET.get('pay_end', '').strip()
    has_date = bool(pay_start or pay_end)

    if has_date:
        date_q = Q()
        if pay_start:
            date_q &= Q(payments__payment_date__gte=pay_start)
        if pay_end:
            date_q &= Q(payments__payment_date__lte=pay_end)
        if pay_include_unpaid:
            qs = qs.filter(date_q | Q(outstanding_amount__gt=0))
        else:
            qs = qs.filter(date_q)
        qs = qs.distinct()
    elif pay_status == 'unpaid':
        qs = qs.filter(payments__isnull=True)
    elif pay_include_unpaid:
        qs = qs.filter(outstanding_amount__gt=0)
    elif pay_status == 'paid':
        qs = qs.filter(payments__isnull=False).distinct()
    return qs


# 应收明细可排序字段白名单：仅暴露已建索引/安全的列，映射到 ORM 字段（值为元组以支持
# 复合排序，如运作年+月）。前端传 sort=key（升序）或 sort=-key（降序）。
_AR_SORT_FIELDS = {
    'operation': ('operation_date',),
    'due_date': ('due_date',),
    'target_date': ('target_collection_date',),
    'outstanding': ('outstanding_amount',),
    'estimated': ('estimated_amount',),
    'invoiced': ('actual_invoice_amount',),
    'invoice_date': ('invoice_date',),
    'reconciliation_date': ('reconciliation_date',),
    'created': ('created_at',),
    'dept': ('delivery_dept',),
    'project_no': ('project__project_no',),
    'short_name': ('project__short_name',),
    'manager': ('project__project_manager',),
}


def _apply_record_sort(qs, request):
    """按白名单字段做服务端排序；空值统一排末尾，并追加 id 作为稳定次序保证分页确定。

    仅作用于列表查询集（不影响合计/聚合）。非法 sort 键安全忽略，回退模型默认排序。
    """
    raw = (request.GET.get('sort') or '').strip()
    if not raw:
        return qs
    desc = raw.startswith('-')
    key = raw[1:] if desc else raw
    fields = _AR_SORT_FIELDS.get(key)
    if not fields:
        return qs
    terms = [F(f).desc(nulls_last=True) if desc else F(f).asc(nulls_last=True) for f in fields]
    terms.append(F('id').desc() if desc else F('id').asc())  # 唯一键兜底，分页稳定
    return qs.order_by(*terms)


# ── 条件构建器（BI 式多条件筛选）─────────────────────────────────────────────
# 维度类条件仍走既有 flat 参数（dept/status/...，语义单一源）；这里只处理两类"新能力"：
#   date：任选日期字段 + 相对区间(本周/本月/上月/下月/本年/去年/自定义) + 含/不含
#   amt ：数值字段 + 运算符(=0/≠0/>0/<0/>/</区间)
# 前端把这两类条件序列化进 conditions(JSON 数组)下发；多条件 AND，可任意叠加。
_COND_DATE_FIELDS = {'due_date', 'payment_date', 'invoice_date', 'reconciliation_date',
                     'operation_date', 'target_collection_date'}
_COND_AMT_FIELDS = {'estimated_amount', 'outstanding_amount', 'tax_amount',
                    'actual_invoice_amount', 'account_diff_adjustment'}


def _relative_date_range(token, today):
    """相对区间 token → (start, end)（含端点）；未知 token 返回 None。服务端按 today 计算。"""
    y, m = today.year, today.month
    if token == 'this_week':
        start = today - datetime.timedelta(days=today.weekday())
        return start, start + datetime.timedelta(days=6)
    if token == 'this_month':
        return datetime.date(y, m, 1), datetime.date(y, m, calendar.monthrange(y, m)[1])
    if token == 'last_month':
        py, pm = (y - 1, 12) if m == 1 else (y, m - 1)
        return datetime.date(py, pm, 1), datetime.date(py, pm, calendar.monthrange(py, pm)[1])
    if token == 'next_month':
        ny, nm = (y + 1, 1) if m == 12 else (y, m + 1)
        return datetime.date(ny, nm, 1), datetime.date(ny, nm, calendar.monthrange(ny, nm)[1])
    if token == 'this_year':
        return datetime.date(y, 1, 1), datetime.date(y, 12, 31)
    if token == 'last_year':
        return datetime.date(y - 1, 1, 1), datetime.date(y - 1, 12, 31)
    if token == 'last_week':
        start = today - datetime.timedelta(days=today.weekday() + 7)
        return start, start + datetime.timedelta(days=6)
    if token == 'this_quarter':
        q = (m - 1) // 3
        qs, qe = q * 3 + 1, q * 3 + 3
        return datetime.date(y, qs, 1), datetime.date(y, qe, calendar.monthrange(y, qe)[1])
    if token == 'last_quarter':
        q = (m - 1) // 3
        if q == 0:
            py, lqs, lqe = y - 1, 10, 12
        else:
            py, lqs, lqe = y, (q - 1) * 3 + 1, (q - 1) * 3 + 3
        return datetime.date(py, lqs, 1), datetime.date(py, lqe, calendar.monthrange(py, lqe)[1])
    return None


def _dec_or_none(v):
    try:
        return Decimal(str(v))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _parse_ym(s):
    """解析 'YYYY-MM' → (year, month)；非法返回 None。"""
    try:
        ys, ms = str(s).split('-')[:2]
        return int(ys), int(ms)
    except (TypeError, ValueError, AttributeError):
        return None


def _ym_range_q(start_ym, end_ym):
    """运作年月区间 → 标量安全 Q（operation_year/month 双列比较）。
    仅给 start 时退化为单月；start/end 任意顺序都按从小到大归一。"""
    s = _parse_ym(start_ym) if start_ym else None
    e = _parse_ym(end_ym) if end_ym else None
    if s and not e:
        e = s
    if e and not s:
        s = e
    if not (s and e):
        return None
    if s > e:
        s, e = e, s
    (y1, m1), (y2, m2) = s, e
    if y1 == y2:
        return Q(operation_year=y1, operation_month__gte=m1, operation_month__lte=m2)
    q = Q(operation_year=y1, operation_month__gte=m1) | Q(operation_year=y2, operation_month__lte=m2)
    if y2 - y1 >= 2:
        q |= Q(operation_year__gt=y1, operation_year__lt=y2)
    return q


def _condition_q(c, today, eomonth_today):
    """把单个条件构建成「标量安全」的 Q（关联字段一律转 id IN 子查询），
    便于在 且/或 任意组合下安全参与布尔运算。非法条目返回 None。"""
    if not isinstance(c, dict):
        return None
    ctype = c.get('t')

    # ── 条件组（括号）：内部按 group.match(且/或) 组合，整体作为一个标量安全 Q ──
    # 支持「(A 且 B) 或 C」这类带括号的复合筛选，可与顶层条件任意嵌套一层。
    if ctype == 'group':
        inner = c.get('conditions')
        if not isinstance(inner, list):
            return None
        inner_any = (c.get('match') or 'all') == 'any'
        combined = None
        for sub in inner:
            q = _condition_q(sub, today, eomonth_today)
            if q is None:
                continue
            combined = q if combined is None else ((combined | q) if inner_any else (combined & q))
        return combined

    # ── 维度类 ─────────────────────────────────────────────────────────────
    if ctype == 'dim':
        field = c.get('field')
        # 运作年月：支持区间(value~end) + 含/不含(exclude)；在空值守卫之前处理
        if field == 'operation_ym':
            q = _ym_range_q(c.get('value'), c.get('end'))
            if q is None:
                return None
            return ~q if c.get('exclude') else q
        v = c.get('value')
        if v in (None, ''):
            return None
        if field == 'dept':
            return Q(delivery_dept=v)
        if field == 'project_id':
            try:
                return Q(project_id=int(v))
            except (TypeError, ValueError):
                return None
        if field == 'manager':
            return Q(project__project_manager__icontains=v)
        if field == 'is_shared':
            return Q(project__is_shared=(str(v) in ('1', 'true', 'True')))
        if field == 'operation_year':
            try:
                return Q(operation_year=int(v))
            except (TypeError, ValueError):
                return None
        if field == 'operation_month':
            try:
                return Q(operation_month=int(v))
            except (TypeError, ValueError):
                return None
        if field == 'q':
            return (Q(project__short_name__icontains=v) | Q(project__customer_name__icontains=v) |
                    Q(project__project_no__icontains=v) | Q(project__project_manager__icontains=v))
        if field == 'status':
            if v == 'overdue':
                return Q(outstanding_amount__gt=0, due_date__lt=today)
            if v == 'current':
                return Q(outstanding_amount__gt=0, due_date__gte=today, due_date__lte=eomonth_today)
            if v == 'not_due':
                return Q(outstanding_amount__gt=0, due_date__gt=eomonth_today)
            if v == 'settled':
                return Q(outstanding_amount__lte=0)
            if v == 'outstanding':
                return Q(outstanding_amount__gt=0)
            return None
        if field == 'reconciliation_status':
            if v == '已对账':
                return Q(reconciliation_date__isnull=False) | Q(actual_invoice_amount__isnull=False)
            if v == '未对账':
                return Q(reconciliation_date__isnull=True, actual_invoice_amount__isnull=True)
            return None
        if field == 'invoice_status':
            if v == '未开票':
                return Q(actual_invoice_amount__isnull=True, outstanding_amount__gt=0)
            if v == '已结清':
                return Q(outstanding_amount__lte=0)
            if v == '已开票':
                return Q(actual_invoice_amount__isnull=False, outstanding_amount__gt=0)
            return None
        if field == 'responsibility':
            no_inv = Q(project__invoice_type='不开票')
            if v == 'settled':
                return Q(outstanding_amount__lte=0)
            if v == 'post':
                return Q(outstanding_amount__gt=0) & (
                    (no_inv & Q(reconciliation_date__isnull=False)) |
                    (~no_inv & Q(invoice_date__isnull=False)))
            if v == 'invoice':
                return (Q(outstanding_amount__gt=0) & ~no_inv & Q(invoice_date__isnull=True) &
                        (Q(reconciliation_date__isnull=False) | Q(actual_invoice_amount__isnull=False)))
            if v == 'recon':
                return Q(outstanding_amount__gt=0) & (
                    (no_inv & Q(reconciliation_date__isnull=True)) |
                    (~no_inv & Q(invoice_date__isnull=True) &
                     Q(reconciliation_date__isnull=True) & Q(actual_invoice_amount__isnull=True)))
            return None
        return None

    # ── 日期类 ─────────────────────────────────────────────────────────────
    if ctype == 'date':
        field = c.get('field')
        if field not in _COND_DATE_FIELDS:
            return None
        rng = c.get('range')
        if rng == 'custom':
            start = _normalize_date(c.get('start')) or None
            end = _normalize_date(c.get('end')) or None
            if not (start or end):
                return None
        else:
            r = _relative_date_range(rng, today)
            if not r:
                return None
            start, end = r
        exclude = bool(c.get('exclude'))
        if field == 'payment_date':
            # 关联回款集 → 转 id IN 子查询，标量安全；不含=无该区间回款
            sub = ARPayment.objects.all()
            if start:
                sub = sub.filter(payment_date__gte=start)
            if end:
                sub = sub.filter(payment_date__lte=end)
            q = Q(id__in=sub.values('ar_record_id'))
        else:
            q = Q()
            if start:
                q &= Q(**{f'{field}__gte': start})
            if end:
                q &= Q(**{f'{field}__lte': end})
        return ~q if exclude else q

    # ── 金额类 ─────────────────────────────────────────────────────────────
    if ctype == 'amt':
        field = c.get('field')
        if field not in _COND_AMT_FIELDS:
            return None
        op = c.get('op')
        if op == 'gt0':
            return Q(**{f'{field}__gt': 0})
        if op == 'lt0':
            return Q(**{f'{field}__lt': 0})
        if op == 'eq0':
            return Q(**{field: 0})
        if op == 'ne0':
            return Q(**{f'{field}__gt': 0}) | Q(**{f'{field}__lt': 0})
        if op in ('gt', 'lt', 'eq'):
            v = _dec_or_none(c.get('value'))
            if v is None:
                return None
            return Q(**{{'gt': f'{field}__gt', 'lt': f'{field}__lt', 'eq': field}[op]: v})
        if op == 'between':
            lo = _dec_or_none(c.get('min'))
            hi = _dec_or_none(c.get('max'))
            q = Q()
            if lo is not None:
                q &= Q(**{f'{field}__gte': lo})
            if hi is not None:
                q &= Q(**{f'{field}__lte': hi})
            return q if q else None
        return None

    return None


def _apply_conditions(qs, request, today=None):
    """统一条件评估：conditions(JSON 数组) 经 match(all=且/any=或) 组合为单个 Q 后应用。

    所有条件均为标量安全 Q（关联字段已转 id IN 子查询），可在且/或下自由组合，
    无需 distinct。非法条目静默跳过，保证健壮与安全。
    """
    raw = (request.GET.get('conditions') or '').strip()
    if not raw:
        return qs
    try:
        conds = json.loads(raw)
        assert isinstance(conds, list)
    except (ValueError, AssertionError):
        return qs
    if today is None:
        today = datetime.date.today()
    eomonth_today = datetime.date(today.year, today.month,
                                  calendar.monthrange(today.year, today.month)[1])
    match_any = (request.GET.get('match') or 'all').strip() == 'any'

    combined = None
    for c in conds:
        q = _condition_q(c, today, eomonth_today)
        if q is None:
            continue
        if combined is None:
            combined = q
        else:
            combined = (combined | q) if match_any else (combined & q)
    if combined is not None:
        qs = qs.filter(combined)
    return qs


# 自动汇总本模块所有公开名（含单下划线助手）供各业务域 `from ._common import *`：
# dir() 在模块末尾返回当前命名空间全部名称；排除 dunder 即得到需要再导出的集合。
__all__ = [n for n in dir() if not n.startswith('__')]
