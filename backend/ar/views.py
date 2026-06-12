import io
import json
import logging
import re
import datetime
import calendar
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from urllib.parse import quote

logger = logging.getLogger(__name__)

from django.core.exceptions import ValidationError
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from django.db.models import Sum, Count, Q, F, Value, Case, When, IntegerField, CharField, DecimalField, Max, Min
from django.db.models.functions import TruncMonth, Coalesce
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from paikuan.views import (pk_required, ok, err, DEPARTMENTS, VALID_DEPARTMENTS,
                           get_request_perms, apply_ar_view_mask, AR_PROJECT_FIELD_DEFS,
                           AR_RECORD_FIELD_DEFS, AR_ADVANCE_FIELD_DEFS, _paid_subq)
from ar.models import (ARProject, ARRecord, ARPayment, ARAdjustment,
                       BatchInvoiceEvent,
                       CollectionBudget, PaymentBudget,
                       AdvanceRecord, AdvanceWriteoff, AdvanceInstallment,
                       Supplier, Customer,
                       Contract, ContractParty, ContractProject, ActionItem,
                       CashPoolConfig, CashPoolTransfer)
from paikuan.models import Payment, PaymentInstallment, ApprovalRecord

# ── AR page keys (must match PAGE_KEYS in paikuan/views.py) ───────────────────
AR_PAGE_KEYS = ['ar_projects', 'ar_records', 'ar_advance', 'ar_analytics', 'ar_cashflow', 'ar_budget']

ADVANCE_DIRECTIONS = ['预收', '预付']

EXAMPLE_ROW_MARKER = '示例-导入前请删除此行'
VALID_CUSTOMER_LEVELS = ['S级', 'A级', 'B级', 'C级', 'D级']
VALID_INVOICE_TYPES = ['专票', '普票', '不开票']


def _norm_header(h):
    """规范化 Excel 表头：去空格、去必填星号、去括号注释，使模板(带*)与
    导出(不带*)的表头可双向匹配。例：'合同名称*'、'开票模式(全额/差额)' →
    '合同名称'、'开票模式'。"""
    s = str(h or '').strip().replace('＊', '').replace('*', '')
    s = s.replace('（', '(').replace('）', ')')
    s = re.sub(r'\([^)]*\)', '', s)   # 去掉括号及其中的注释
    return s.strip()


def _parse_cycle_start_day(v):
    """对账周期起始日：1~28（1=自然月）。返回 (值, 错误文案)。空值按 1。"""
    if v in (None, ''):
        return 1, None
    try:
        d = int(float(v))
    except (TypeError, ValueError):
        return None, f'对账周期起始日"{v}"无效，须为 1~28 的整数（1=自然月）'
    if not (1 <= d <= 28):
        return None, f'对账周期起始日须为 1~28（当前 {d}）。29~31 在短月不存在，请用 28 或改约对账日'
    return d, None


def _match_project_by_short_name(short_name, dept='', allowed_depts=None):
    """Simple lookup used by non-import paths (project search, etc.)."""
    name = (short_name or '').strip()
    if not name:
        return None
    qs = ARProject.objects.filter(short_name=name)
    if dept:
        qs = qs.filter(delivery_dept=dept)
    elif allowed_depts is not None:
        qs = qs.filter(delivery_dept__in=allowed_depts)
    proj = qs.order_by('-id').first()
    if proj:
        return proj
    qs = ARProject.objects.filter(short_name__icontains=name)
    if dept:
        qs = qs.filter(delivery_dept=dept)
    elif allowed_depts is not None:
        qs = qs.filter(delivery_dept__in=allowed_depts)
    return qs.order_by('-id').first()


def _proj_candidate_dict(p):
    """歧义候选项的精简信息，供导入返回让用户挑选/区分。"""
    return {
        'project_no': p.project_no,
        'short_name': p.short_name,
        'customer_name': p.customer_name,
        'delivery_dept': p.delivery_dept,
    }


def _classify_project_for_import(short_name, customer_hint, project_no, allowed_depts, dept_hint=''):
    """只读判定项目归属（不建库、不创建草稿），用于导入前的歧义检查。

    返回 (status, payload)：
      'resolved'  -> payload = ARProject（唯一命中）
      'ambiguous' -> payload = 候选 list[dict]（多命中且无法区分，需人工指定）
      'bad_no'    -> payload = 填写的项目编号（填了但查不到/越权）
      'not_found' -> payload = None（找不到，写入阶段将自动建草稿）

    优先级：项目编号(确定性) > 精确简称(+客户名缩小) > 模糊简称(+客户名缩小)。
    """
    name = (short_name or '').strip()
    pno = (project_no or '').strip()
    dh = (dept_hint or '').strip()

    # ── 0) 项目编号：最高优先级、确定性指定 ────────────────────────────────
    if pno:
        qs = ARProject.objects.filter(project_no=pno)
        if allowed_depts is not None:
            qs = qs.filter(delivery_dept__in=allowed_depts)
        proj = qs.first()
        return ('resolved', proj) if proj else ('bad_no', pno)

    def _narrow(qs):
        # 交付部门是业务主键的一部分（短名+部门），优先用它区分同名跨部门项目
        if dh:
            nd = qs.filter(delivery_dept=dh)
            if nd.exists():
                qs = nd
        if customer_hint:
            n = qs.filter(customer_name__icontains=customer_hint)
            if n.exists():
                return n
        return qs

    # ── 1) 精确简称 ────────────────────────────────────────────────────────
    qs = ARProject.objects.filter(short_name=name)
    if allowed_depts is not None:
        qs = qs.filter(delivery_dept__in=allowed_depts)
    qs = _narrow(qs)
    cnt = qs.count()
    if cnt == 1:
        return 'resolved', qs.first()
    if cnt > 1:
        return 'ambiguous', [_proj_candidate_dict(p) for p in qs.order_by('-id')[:8]]

    # ── 2) 模糊简称 ────────────────────────────────────────────────────────
    qs2 = ARProject.objects.filter(short_name__icontains=name)
    if allowed_depts is not None:
        qs2 = qs2.filter(delivery_dept__in=allowed_depts)
    qs2 = _narrow(qs2)
    cnt2 = qs2.count()
    if cnt2 == 1:
        return 'resolved', qs2.first()
    if cnt2 > 1:
        return 'ambiguous', [_proj_candidate_dict(p) for p in qs2.order_by('-id')[:8]]

    # ── 3) 找不到 → 写入阶段自动建草稿 ─────────────────────────────────────
    return 'not_found', None


def _resolve_project_for_import(short_name, customer_hint, dept, allowed_depts, user,
                                proj_cache, project_no=''):
    """3-stage project resolution for bulk import.

    Returns (proj, match_type, warning_str | None).
    match_type: 'exact' | 'exact_multi' | 'fuzzy' | 'fuzzy_multi' | 'created'

    Uses proj_cache {(short_name, dept) -> result} to avoid creating duplicate
    draft projects when multiple rows share the same project name.
    若提供 project_no（项目编号），优先按编号精确命中，作为确定性指定。
    """
    name = (short_name or '').strip()
    if not name:
        return None, None, None
    cache_key = (name, dept or '')

    # 项目编号确定性命中（不进 proj_cache，按编号逐行解析）
    pno = (project_no or '').strip()
    if pno:
        qs_no = ARProject.objects.filter(project_no=pno)
        if allowed_depts is not None:
            qs_no = qs_no.filter(delivery_dept__in=allowed_depts)
        proj_no = qs_no.first()
        if proj_no:
            return proj_no, 'exact', None

    if cache_key in proj_cache:
        return proj_cache[cache_key]

    # ── Stage 1: 精确匹配 ──────────────────────────────────────────────────
    qs = ARProject.objects.filter(short_name=name)
    if allowed_depts is not None:
        qs = qs.filter(delivery_dept__in=allowed_depts)
    # 交付部门是业务主键的一部分：用它区分同名跨部门项目（在权限范围内进一步缩小）
    if dept:
        qs_dept = qs.filter(delivery_dept=dept)
        if qs_dept.exists():
            qs = qs_dept

    # Customer hint可进一步缩小范围（同名项目不同客户时）
    if customer_hint:
        qs_cust = qs.filter(customer_name__icontains=customer_hint)
        if qs_cust.exists():
            cnt = qs_cust.count()
            proj = qs_cust.order_by('-id').first()
            warn = (f'找到 {cnt} 个同名同客户项目，已取最新的「{proj.customer_name}」，请核查'
                    if cnt > 1 else None)
            result = (proj, 'exact' if cnt == 1 else 'exact_multi', warn)
            proj_cache[cache_key] = result
            return result

    if qs.exists():
        cnt = qs.count()
        proj = qs.order_by('-id').first()
        warn = (f'找到 {cnt} 个同名项目，已取最新的「{proj.customer_name}」，请核查'
                if cnt > 1 else None)
        result = (proj, 'exact' if cnt == 1 else 'exact_multi', warn)
        proj_cache[cache_key] = result
        return result

    # ── Stage 2: 模糊匹配（contains）──────────────────────────────────────
    qs2 = ARProject.objects.filter(short_name__icontains=name)
    if allowed_depts is not None:
        qs2 = qs2.filter(delivery_dept__in=allowed_depts)
    if dept:
        qs2_dept = qs2.filter(delivery_dept=dept)
        if qs2_dept.exists():
            qs2 = qs2_dept
    matches = list(qs2.order_by('-id')[:5])
    if matches:
        proj = matches[0]
        if len(matches) > 1:
            candidates = '、'.join(f'「{m.short_name}」' for m in matches[:3])
            warn = f'"{name}"模糊匹配到多个项目（{candidates}），已取最新的「{proj.short_name}」，建议核查'
            match_type = 'fuzzy_multi'
        else:
            warn = f'"{name}"未精确匹配，模糊命中「{proj.short_name}」，建议核查'
            match_type = 'fuzzy'
        result = (proj, match_type, warn)
        proj_cache[cache_key] = result
        return result

    # ── Stage 3: 自动创建草稿项目 ─────────────────────────────────────────
    proj_dept = dept or (list(allowed_depts)[0] if allowed_depts else '')
    creator_name = user.name if user else '待填'
    proj = ARProject(
        short_name=name,
        customer_name=customer_hint or name,
        delivery_dept=proj_dept,
        is_draft=True,
        sales_contact=creator_name,
        project_manager=creator_name,
        has_contract='无',
    )
    proj.save()
    warn = f'项目台账中未找到「{name}」，已自动创建草稿项目，请到项目台账补充完善'
    result = (proj, 'created', warn)
    proj_cache[cache_key] = result
    return result


def _normalize_date(s):
    """Best-effort normalize a user-typed date string to ISO YYYY-MM-DD.

    Accepts: 2026-01-15, 2026/1/15, 2026.1.15, 2026年1月15日, 26/1/15,
    20260115, datetime/date objects, Excel serial numbers.
    """
    import datetime as _dt
    if s is None:
        return None
    # datetime / date objects (openpyxl can return these)
    if isinstance(s, (_dt.datetime, _dt.date)):
        return s.strftime('%Y-%m-%d')
    # Excel serial date (number of days since 1899-12-30)
    if isinstance(s, (int, float)) and not isinstance(s, bool):
        try:
            base = _dt.date(1899, 12, 30)
            return (base + _dt.timedelta(days=int(s))).strftime('%Y-%m-%d')
        except (OverflowError, ValueError):
            return None
    s = str(s).strip()
    if not s or s.lower() == 'none':
        return None
    # 20260115 → 2026-01-15
    if re.fullmatch(r'\d{8}', s):
        try:
            return f'{s[:4]}-{s[4:6]}-{s[6:]}'
        except ValueError:
            pass
    # Drop chinese unit suffixes / unify separators
    cleaned = (s.replace('/', '-').replace('.', '-')
                .replace('年', '-').replace('月', '-').replace('日', '')
                .replace(' ', '').strip('-'))
    parts = [p for p in cleaned.split('-') if p]
    try:
        if len(parts) >= 3:
            y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
            # 2-digit year → 20xx
            if y < 100:
                y += 2000
            _dt.date(y, m, d)  # validate
            return f'{y:04d}-{m:02d}-{d:02d}'
    except (ValueError, TypeError):
        pass
    return None


def _dec(v, default=Decimal('0')):
    try:
        return Decimal(str(v))
    except (InvalidOperation, TypeError):
        return default


def _int_param(request, key, default):
    """GET 参数安全转 int：非法值回退默认值（避免 ?year=abc 之类把接口打成500）。"""
    try:
        return int(request.GET.get(key, default) or default)
    except (ValueError, TypeError):
        return int(default)


def _parse_body(request):
    try:
        return json.loads(request.body)
    except Exception:
        return {}


def _ar_dept_filter(qs, request, dept_field='delivery_dept', shared_field=None):
    """Filter queryset by the requesting user's allowed departments.
    Optional ?depts=A,B,C further narrows the set (intersected with pk_depts).
    When shared_field is provided and the user has ar_shared_only, restricts to
    shared projects only (shared_field=True)."""
    raw = request.GET.get('depts', '').strip()
    requested = [d for d in raw.split(',') if d.strip()] if raw else []

    if request.pk_role == 'super_admin':
        if requested:
            qs = qs.filter(**{f'{dept_field}__in': requested})
        return qs

    allowed = set(request.pk_depts or [])
    if not allowed:
        return qs.none()
    if requested:
        active = [d for d in requested if d in allowed]
        if active:
            qs = qs.filter(**{f'{dept_field}__in': active})
        else:
            qs = qs.filter(**{f'{dept_field}__in': request.pk_depts})
    else:
        qs = qs.filter(**{f'{dept_field}__in': request.pk_depts})

    if shared_field:
        from paikuan.views import get_request_perms
        perms = get_request_perms(request)
        if perms and perms.get('ar_shared_only'):
            qs = qs.filter(**{shared_field: True})
    return qs


def _page_denied(request, page_key):
    """Return error response if user lacks access to a page; None if allowed."""
    if request.pk_role == 'super_admin':
        return None
    from paikuan.views import get_job_perms
    jt = getattr(request, 'pk_job', '') or ''
    if not jt:
        from paikuan.models import PaikuanUser
        u = PaikuanUser.objects.filter(id=request.pk_uid).only('job_title').first()
        jt = u.job_title if u else ''
    perms = get_job_perms(jt)
    if not perms['pages'].get(page_key, True):
        return err('无权访问此模块', 403)
    return None


def _write_denied(request):
    perms = get_request_perms(request)
    # AR 写入：优先看 ar_can_create（结算会计等聚焦应收的岗位），回退到通用 can_create。
    if perms is not None and not (perms.get('ar_can_create') or perms.get('can_create', False)):
        return err('无写入权限', 403, 403)
    return None


def _action_denied(request, action_key):
    """操作级权限：核销/回款录入等细粒度动作的独立开关。

    actions[key] 显式配置时以其为准（True 放行 / False 拒绝）；
    旧配置无 actions 键时回退通用写权限（向后兼容，不会突然把人锁外面）。
    出纳等岗位可借此单独获得「预付核销」而不必放开整个 can_create。"""
    perms = get_request_perms(request)
    if perms is None:
        return None
    acts = perms.get('actions')
    if isinstance(acts, dict) and action_key in acts:
        if acts.get(action_key):
            return None
        return err('当前职务未开通此操作权限，请联系管理员在「权限配置·操作权限」中开通', 403, 403)
    return _write_denied(request)


def _delete_denied(request):
    perms = get_request_perms(request)
    if perms is not None and not perms.get('can_delete', False):
        return err('无删除权限', 403, 403)
    return None


def _dept_denied(request, dept, msg='无权操作此部门'):
    if request.pk_role == 'super_admin':
        return None
    if dept not in request.pk_depts:
        return err(msg, 403, 403)
    return None


def _sync_project_contracts(proj, contract_ids, request):
    """替换某项目挂靠的合同（项目侧维护多对多）。只挂用户有权部门的合同。"""
    proj.contract_links.all().delete()
    seen = set()
    for cid in (contract_ids or []):
        try:
            cid = int(cid)
        except (ValueError, TypeError):
            continue
        if cid in seen:
            continue
        ct = Contract.objects.filter(pk=cid).first()
        if not ct:
            continue
        if (request.pk_role != 'super_admin' and ct.delivery_dept
                and ct.delivery_dept not in request.pk_depts):
            continue
        seen.add(cid)
        ContractProject.objects.create(contract=ct, project=proj, is_primary=True)


def _project_contracts_list(proj):
    return [
        {'contract_id': cp.contract_id, 'name': cp.contract.name,
         'contract_no': cp.contract.contract_no, 'is_primary': cp.is_primary}
        for cp in proj.contract_links.select_related('contract').all()
    ]


def _object_dept_denied(request, obj, dept_field='delivery_dept'):
    return _dept_denied(request, getattr(obj, dept_field, ''), '无权访问')


def _can_ar_view(request, field_key):
    perms = get_request_perms(request)
    if perms is None:
        return True
    return (perms.get('ar_view') or {}).get(field_key, True)


def _ar_field_denied(request, field_key):
    if not _can_ar_view(request, field_key):
        return err('无权访问此字段', 403, 403)
    return None


_AR_PAYLOAD_DEFS_BY_GROUP = {
    'project': AR_PROJECT_FIELD_DEFS,
    'record': AR_RECORD_FIELD_DEFS,
    'advance': AR_ADVANCE_FIELD_DEFS,
}


def _ar_visible_payload(request, data, group, extra=()):
    perms = get_request_perms(request)
    if perms is None:
        return data
    defs = _AR_PAYLOAD_DEFS_BY_GROUP.get(group, AR_RECORD_FIELD_DEFS)
    ar_view = perms.get('ar_view') or {}
    cols = set(extra)
    for f in defs:
        if ar_view.get(f['key'], True):
            cols.update(f['cols'])
    return {k: v for k, v in data.items() if k in cols}


def _visible_ar_export_cols(request, columns):
    perms = get_request_perms(request)
    if perms is None:
        return columns
    ar_view = perms.get('ar_view') or {}
    return [col for col in columns if col[0] is None or ar_view.get(col[0], True)]


_XL_FORMULA_CHARS = ('=', '+', '-', '@', '\t', '\r')


def _export_response(wb, filename):
    # Excel 公式注入防护（与排款导出口径一致）：全部工作表的文本单元格，
    # 以 =+-@ 等开头的前置单引号转义。集中在导出总出口做，覆盖所有 AR 导出。
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v and v[0] in _XL_FORMULA_CHARS:
                    cell.value = "'" + v
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    safe = quote(filename, safe='')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{safe}"
    resp['Cache-Control'] = 'no-store'
    return resp


def _header_row(ws, headers, color='1565C0'):
    fill = PatternFill('solid', fgColor=color)
    font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')


# ══════════════════════════════════════════════════════════════════════════════
# Projects
# ══════════════════════════════════════════════════════════════════════════════

def _apply_project_list_filters(qs, request):
    """项目台账列表筛选（q/dept/manager/is_shared/is_draft/customer_level/invoice_mode）。
    列表 GET 与批量删除「全选筛选集」共用，保证删除范围与所见列表一致。"""
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(customer_name__icontains=q) |
            Q(short_name__icontains=q) |
            Q(project_no__icontains=q) |
            Q(project_manager__icontains=q) |
            Q(sales_contact__icontains=q)
        )
    dept = request.GET.get('dept', '').strip()
    if dept:
        qs = qs.filter(delivery_dept=dept)
    manager = request.GET.get('manager', '').strip()
    if manager:
        qs = qs.filter(project_manager__icontains=manager)
    is_shared = request.GET.get('is_shared', '').strip()
    if is_shared in ('true', '1'):
        qs = qs.filter(is_shared=True)
    elif is_shared in ('false', '0'):
        qs = qs.filter(is_shared=False)
    is_draft = request.GET.get('is_draft', '').strip()
    if is_draft in ('true', '1'):
        qs = qs.filter(is_draft=True)
    elif is_draft in ('false', '0'):
        qs = qs.filter(is_draft=False)
    level = request.GET.get('customer_level', '').strip()
    if level:
        qs = qs.filter(customer_level=level)
    mode = request.GET.get('invoice_mode', '').strip()
    if mode:
        qs = qs.filter(invoice_mode=mode)
    status = request.GET.get('status', '').strip()
    if status:
        qs = qs.filter(status=status)
    return qs


@csrf_exempt
@pk_required()
def projects(request):
    denied = _page_denied(request, 'ar_projects')
    if denied:
        return denied

    if request.method == 'GET':
        qs = _apply_project_list_filters(
            _ar_dept_filter(ARProject.objects.all(), request, shared_field='is_shared'),
            request)

        page = max(1, int(request.GET.get('page', 1) or 1))
        size = min(100, max(1, int(request.GET.get('size', 50) or 50)))
        total = qs.count()
        items = list(qs.select_related('created_by')[(page - 1) * size: page * size])

        perms = get_request_perms(request)
        rows = [apply_ar_view_mask(p.to_dict(), perms, 'project') for p in items]
        return ok({'items': rows, 'total': total, 'page': page, 'size': size})

    if request.method == 'POST':
        denied = _write_denied(request)
        if denied:
            return denied
        data = _ar_visible_payload(request, _parse_body(request), 'project',
                                   extra=('customer_id', 'contract_ids'))
        dept = data.get('delivery_dept', '')
        if dept not in VALID_DEPARTMENTS:
            return err(f'无效交付部门: {dept}')
        if dept == '集团总部':
            return err('集团总部无项目，不允许新增项目台账')
        denied = _dept_denied(request, dept, '无权操作此部门')
        if denied:
            return denied
        from paikuan.models import PaikuanUser
        user = PaikuanUser.objects.filter(id=request.pk_uid).first()
        csd, csd_err = _parse_cycle_start_day(data.get('cycle_start_day'))
        if csd_err:
            return err(csd_err)
        try:
            p = ARProject(
                customer_name=data.get('customer_name', '').strip(),
                short_name=data.get('short_name', '').strip(),
                delivery_dept=dept,
                sub_dept=data.get('sub_dept', '').strip(),
                business_mode=data.get('business_mode', '').strip(),
                customer_level=data.get('customer_level', '').strip(),
                sales_contact=data.get('sales_contact', '').strip(),
                project_manager=data.get('project_manager', '').strip(),
                has_contract=data.get('has_contract', '无'),
                contract_date=_normalize_date(data.get('contract_date')) or None,
                reconciliation_days=int(data.get('reconciliation_days', 0) or 0),
                invoice_wait_days=int(data.get('invoice_wait_days', 0) or 0),
                post_invoice_days=int(data.get('post_invoice_days', 0) or 0),
                cycle_start_day=csd,
                invoice_mode=data.get('invoice_mode', '全额'),
                invoice_type=data.get('invoice_type', ''),
                tax_rate=_dec(data.get('tax_rate', '0')),
                notes=data.get('notes', '').strip(),
                created_by=user,
            )
            cid = data.get('customer_id')
            if cid:
                p.customer = Customer.objects.filter(pk=cid).first()
            p.save()
            if 'contract_ids' in data:
                _sync_project_contracts(p, data['contract_ids'], request)
        except Exception as e:
            return err(str(e))
        return ok(apply_ar_view_mask(p.to_dict(), get_request_perms(request), 'project'))

    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def project_detail(request, pk):
    denied = _page_denied(request, 'ar_projects')
    if denied:
        return denied
    try:
        proj = ARProject.objects.select_related('created_by').get(pk=pk)
    except ARProject.DoesNotExist:
        return err('项目不存在', 404)
    # dept access
    if request.pk_role != 'super_admin':
        allowed = request.pk_depts
        if proj.delivery_dept not in allowed:
            return err('无权访问', 403)
        perms = get_request_perms(request)
        if perms and perms.get('ar_shared_only') and not proj.is_shared:
            return err('无权访问', 403)

    if request.method == 'GET':
        d = proj.to_dict()
        agg = proj.ar_records.aggregate(
            record_count=Count('id'), total_outstanding=Sum('outstanding_amount'))
        d['record_count'] = agg['record_count'] or 0
        d['total_outstanding'] = str(agg['total_outstanding'] or 0)
        d['contracts'] = _project_contracts_list(proj)
        return ok(apply_ar_view_mask(d, get_request_perms(request), 'project'))

    if request.method == 'PUT':
        denied = _write_denied(request)
        if denied:
            return denied
        data = _ar_visible_payload(request, _parse_body(request), 'project',
                                   extra=('customer_id', 'contract_ids', 'is_draft', 'status'))
        for field in ('customer_name', 'short_name', 'sub_dept',
                      'business_mode', 'customer_level', 'sales_contact', 'project_manager',
                      'has_contract', 'invoice_mode', 'invoice_type', 'notes'):
            if field in data:
                setattr(proj, field, data[field])
        if 'status' in data:
            st = (data['status'] or '').strip()
            if st not in dict(ARProject.STATUS_CHOICES):
                return err(f'无效项目状态「{st}」，可选：运作中/中断/结束')
            proj.status = st
        if 'contract_date' in data:
            proj.contract_date = _normalize_date(data['contract_date']) or None
        for field in ('reconciliation_days', 'invoice_wait_days', 'post_invoice_days'):
            if field in data:
                setattr(proj, field, int(data[field] or 0))
        if 'cycle_start_day' in data:
            csd, csd_err = _parse_cycle_start_day(data['cycle_start_day'])
            if csd_err:
                return err(csd_err)
            proj.cycle_start_day = csd
        if 'tax_rate' in data:
            proj.tax_rate = _dec(data['tax_rate'])
        if 'delivery_dept' in data:
            dept = data['delivery_dept']
            if dept not in VALID_DEPARTMENTS:
                return err(f'无效交付部门: {dept}')
            denied = _dept_denied(request, dept, '无权操作目标部门')
            if denied:
                return denied
            proj.delivery_dept = dept
        if 'is_draft' in data:
            proj.is_draft = bool(data['is_draft'])
        if 'customer_id' in data:
            cid = data.get('customer_id')
            orig_cid = proj.customer_id   # 进入本分支前尚未改动，即数据库原值
            typed = (data.get('customer_name') or '').strip() if 'customer_name' in data else None
            if cid:
                try:
                    new_cid = int(cid)
                    cust = Customer.objects.get(pk=new_cid)
                    if new_cid != orig_cid:
                        # 用户在「关联客户」里改选了不同的客户实体 → 以实体为准
                        proj.customer = cust
                        proj.customer_name = cust.name
                    elif typed and typed != cust.name:
                        # 关联客户未变但修改了客户名称文本 → 以文本为准，
                        # 置空实体由 save() 的 _autolink_customer 按新名称(+部门)重新挂接
                        proj.customer = None
                    else:
                        proj.customer = cust
                        proj.customer_name = cust.name
                except (Customer.DoesNotExist, ValueError, TypeError):
                    pass
            else:
                proj.customer = None
        proj.save()
        if 'contract_ids' in data:
            _sync_project_contracts(proj, data['contract_ids'], request)
        out = proj.to_dict()
        out['contracts'] = _project_contracts_list(proj)
        return ok(apply_ar_view_mask(out, get_request_perms(request), 'project'))

    if request.method == 'DELETE':
        denied = _delete_denied(request)
        if denied:
            return denied
        cust_id = proj.customer_id
        proj.delete()
        purged = _purge_orphan_customers([cust_id] if cust_id else [])
        return ok({'deleted': pk, 'purged_customers': purged})

    return err('Method not allowed', 405)


def _purge_orphan_customers(cust_ids):
    """项目删除后的客户名单同步：客户名单由项目台账派生，项目删尽即随之移除。
    仅清理「名下 0 项目且无合同关联」的客户，避免误删仍被合同引用的客户实体。"""
    ids = [i for i in (cust_ids or []) if i]
    if not ids:
        return 0
    qs = (Customer.objects.filter(pk__in=ids)
          .annotate(_np=Count('projects', distinct=True),
                    _nc=Count('contract_links', distinct=True))
          .filter(_np=0, _nc=0))
    n = qs.count()
    if n:
        qs.delete()
    return n


@csrf_exempt
@pk_required()
def projects_bulk_delete(request):
    """批量删除项目（级联其应收明细/回款）。两种模式：
      - 显式选择：body {ids:[int,...]}
      - 选择全部筛选集：body {all:true} + 查询串里的 q/dept/customer_level/... 与列表同口径
    始终受部门作用域与删除权限约束；单次上限 5000 条，防误删全库。"""
    denied = _page_denied(request, 'ar_projects')
    if denied:
        return denied
    if request.method != 'POST':
        return err('POST only', 405)
    denied = _delete_denied(request)
    if denied:
        return denied

    body = _parse_body(request)
    base = _ar_dept_filter(ARProject.objects.all(), request, shared_field='is_shared')

    if body.get('all'):
        # 与列表 GET 完全相同的筛选口径（走查询串）
        qs = _apply_project_list_filters(base, request)
    else:
        ids = body.get('ids') or []
        if not isinstance(ids, list) or not ids:
            return err('请提供要删除的项目 ids')
        try:
            ids = [int(i) for i in ids]
        except (ValueError, TypeError):
            return err('ids 必须为整数列表')
        qs = base.filter(pk__in=ids)

    # 非超管：再次按可操作部门收紧，杜绝越权删除
    if request.pk_role != 'super_admin':
        qs = qs.filter(delivery_dept__in=(request.pk_depts or []))

    count = qs.count()
    if count == 0:
        return ok({'deleted': 0})
    if count > 5000:
        return err('单次删除上限 5000 条，请先缩小筛选范围')
    del_ids = list(qs.values_list('id', flat=True))
    cust_ids = list(qs.exclude(customer_id__isnull=True)
                    .values_list('customer_id', flat=True).distinct())
    ARProject.objects.filter(pk__in=del_ids).delete()  # 级联删除应收明细/回款
    purged = _purge_orphan_customers(cust_ids)  # 客户名单随项目台账同步收敛
    return ok({'deleted': len(del_ids), 'purged_customers': purged})


@csrf_exempt
@pk_required()
def project_template(request):
    denied = _page_denied(request, 'ar_projects')
    if denied:
        return denied
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '项目信息'
    headers = ['客户名称*', '项目简称*', '交付部门*', '二级部门', '业务模式',
               '客户等级', '销售对接人*', '项目负责人*', '有无合同',
               '签订日期', '合同对账期(天)', '开票等待期(天)', '票后等待期(天)',
               '对账周期起始日(1-28)',
               '开票模式(全额/差额)', '专票/普票/不开票', '税率(如0.06)', '备注']
    _header_row(ws, headers)
    tip_vals = [
        '★必填：客户名称/往来单位全称（即客户公司名；再次导入同名客户+同一事业部时自动更新，不新增；预收录入选此项目时自动带出为往来单位）',
        f'★必填：项目简称，须在事业部内唯一！应收明细、收款预算导入时均以此简称匹配项目，简称为关键桥梁',
        f'★必填：可选值：{"、".join(VALID_DEPARTMENTS)}（每个事业部每天新增客户，每位客户可有多个不同项目）',
        '选填：华南区/华北区等，可空',
        '选填：劳务外包/运输/供应链等，自由填写',
        f'选填：{"/".join(VALID_CUSTOMER_LEVELS)}，默认可空',
        '★必填：销售对接人姓名（与项目负责人不同时自动标记为共享业务）',
        '★必填：项目负责人姓名（应收分析可按项目负责人维度聚合）',
        '"有"或"无"，默认无',
        '选填：格式 2026-01-15 / 2026/1/15 / 2026年1月15日 均可',
        '整数天数，默认0（总账期 = 合同对账期 + 开票等待期 + 票后等待期）',
        '整数天数，默认0（同上，三项之和即为应收日期的延迟天数）',
        '整数天数，默认0（票后等待期：开票后多少天内完成回款；应收日期 = 对账周期结束日 + 总账期天数；修改后自动更新已有明细）',
        '选填：1~28，默认1=自然月（月初到月末）。如月结周期为每月15日到下月14日则填 15；'
        '应收到期 = 运作日期所在账期的结束日 + 总账期天数，修改后已有明细到期日自动重算',
        '"全额"或"差额"（全额：税额自动计算=开票金额/(1+税率)×税率；差额：手填税额）',
        f'{"/".join(VALID_INVOICE_TYPES)}；选"不开票"时税率自动置0',
        '选填：如 0.06 表示6%，范围 0~1；选"不开票"时留空或填0',
        '选填备注',
    ]
    ws.append(tip_vals)
    tip_row = ws.max_row
    tip_fill = PatternFill('solid', fgColor='FFF9E6')
    tip_font = Font(italic=True, color='8B6914', size=9)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(tip_row, c)
        cell.fill = tip_fill
        cell.font = tip_font
        cell.alignment = Alignment(wrap_text=True)
    example = [EXAMPLE_ROW_MARKER, '物流外包A', '劳务事业部', '华南区', '劳务外包',
               'A级', '张三', '李四', '有', '2026-01-01', 30, 0, 60, 1,
               '全额', '专票', 0.06, '示例备注（此行含"示例"标记，导入时自动跳过）']
    ws.append(example)
    col_widths = [28, 18, 16, 14, 14, 12, 16, 16, 10, 18, 20, 20, 20, 20, 18, 18, 14, 24]
    for col, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.row_dimensions[tip_row].height = 60
    return _export_response(wb, '项目信息导入模板.xlsx')


@csrf_exempt
@pk_required()
def project_import(request):
    denied = _page_denied(request, 'ar_projects')
    if denied:
        return denied
    if request.method != 'POST':
        return err('POST only', 405)
    denied = _write_denied(request)
    if denied:
        return denied
    f = request.FILES.get('file')
    if not f:
        return err('请上传文件')
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
    except Exception as e:
        return err(f'无法读取Excel: {e}')

    headers = [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    # 表头规范化：去掉 *、括号注释与空格，使「模板表头(带*)」与「导出表头(不带*)」
    # 双向匹配——否则导出→编辑→再导入会因表头不一致而整表静默跳过。
    col_map = {_norm_header(h): i + 1 for i, h in enumerate(headers)}

    def _cv(row, *names):
        """按规范化表头取值，支持传多个候选别名（任一命中即返回）。"""
        for name in names:
            idx = col_map.get(_norm_header(name))
            if idx is None:
                continue
            v = ws.cell(row, idx).value
            if v is not None and str(v).strip():
                return str(v).strip()
        return ''

    from paikuan.models import PaikuanUser
    user = PaikuanUser.objects.filter(id=request.pk_uid).first()
    errors = []
    plan = []   # 通过校验、待写入的行：{ri, customer_name, dept, short_name_val, field_vals}

    # ══ 阶段一：逐行校验（不写库）。示例/提示/空行静默忽略，问题行收集错误 ══════════
    for ri in range(2, ws.max_row + 1):
        customer_name = _cv(ri, '客户名称', '客户名称*', '合同名称', '合同名称*')
        short_name_val = _cv(ri, '项目简称', '项目简称*')
        dept = _cv(ri, '交付部门', '交付部门*')
        # 提示行/示例行：静默忽略（不计入、不报错）
        if (EXAMPLE_ROW_MARKER in customer_name or customer_name.startswith('★')
                or EXAMPLE_ROW_MARKER in short_name_val or short_name_val.startswith('★')):
            continue
        # 客户名称缺失时回退用项目简称（很多用户只填项目简称）；二者皆空才算空行
        if not customer_name:
            customer_name = short_name_val
        if not customer_name:
            # 整行为空 → 静默忽略；否则明确报错，避免「无声跳过」
            if not any(_cv(ri, h) for h in (
                    '交付部门', '销售对接人', '项目负责人', '业务模式')):
                continue
            errors.append(f'第{ri}行: 缺少「客户名称」或「项目简称」，无法识别项目，请补填后重新导入')
            continue
        if dept not in VALID_DEPARTMENTS:
            errors.append(f'第{ri}行: 交付部门"{dept}"无效，可选值为：{"/".join(VALID_DEPARTMENTS)}')
            continue
        if request.pk_role != 'super_admin':
            allowed = request.pk_depts
            if dept not in allowed:
                errors.append(f'第{ri}行: 无权操作部门"{dept}"，您的授权部门为：{"/".join(allowed)}')
                continue
        # ── field-level validation ────────────────────────────────────────────
        customer_level_val = _cv(ri, '客户等级')
        if customer_level_val and customer_level_val not in VALID_CUSTOMER_LEVELS:
            errors.append(f'第{ri}行: 客户等级"{customer_level_val}"无效，应为：{"/".join(VALID_CUSTOMER_LEVELS)}')
            continue
        invoice_type_val = _cv(ri, '专票/普票/不开票', '专票/普票', '发票类型')
        if invoice_type_val and invoice_type_val not in VALID_INVOICE_TYPES:
            errors.append(f'第{ri}行: 发票类型"{invoice_type_val}"无效，应为：{"/".join(VALID_INVOICE_TYPES)}')
            continue
        invoice_mode_val = _cv(ri, '开票模式(全额/差额)') or '全额'
        if invoice_mode_val not in ('全额', '差额'):
            errors.append(f'第{ri}行: 开票模式"{invoice_mode_val}"无效，应为"全额"或"差额"')
            continue
        tax_raw = _cv(ri, '税率(如0.06)')
        if tax_raw:
            try:
                tax_val = float(tax_raw)
                if not (0 <= tax_val <= 1):
                    errors.append(f'第{ri}行: 税率"{tax_raw}"超出范围，应在0到1之间（如 0.06 表示6%）')
                    continue
            except (ValueError, TypeError):
                errors.append(f'第{ri}行: 税率"{tax_raw}"格式错误，应填数字（如 0.06）')
                continue
        try:
            int_days = lambda col: int(_cv(ri, col) or 0)
        except (ValueError, TypeError):
            errors.append(f'第{ri}行: 账期天数应为整数（合同对账期/开票等待期/票后等待期）')
            continue
        field_vals = dict(
            short_name=short_name_val,
            sub_dept=_cv(ri, '二级部门'),
            business_mode=_cv(ri, '业务模式'),
            customer_level=customer_level_val,
            sales_contact=_cv(ri, '销售对接人', '销售对接人*'),
            project_manager=_cv(ri, '项目负责人', '项目负责人*'),
            has_contract=_cv(ri, '有无合同') or '无',
            contract_date=_normalize_date(
                _cv(ri, '签订日期') or _cv(ri, '签订日期(YYYY-MM-DD)')
            ) or None,
            reconciliation_days=int_days('合同对账期(天)'),
            invoice_wait_days=int_days('开票等待期(天)'),
            post_invoice_days=int_days('票后等待期(天)') or int_days('结算等待期(天)'),
            invoice_mode=invoice_mode_val,
            invoice_type=invoice_type_val,
            notes=_cv(ri, '备注'),
        )
        # 对账周期起始日（_norm_header 去括号注释后列名为「对账周期起始日」）
        csd_raw = _cv(ri, '对账周期起始日(1-28)', '对账周期起始日')
        if csd_raw:
            csd_val, csd_err = _parse_cycle_start_day(csd_raw)
            if csd_err:
                errors.append(f'第{ri}行: {csd_err}')
                continue
            field_vals['cycle_start_day'] = csd_val
        # 仅在明确填写税率时才写入，避免留空时把现有项目的税率清零
        if tax_raw:
            field_vals['tax_rate'] = _dec(tax_raw)
        plan.append({'ri': ri, 'customer_name': customer_name, 'dept': dept,
                     'short_name_val': short_name_val, 'field_vals': field_vals})

    # ══ 有任何问题 → 整表拒绝，不写入（要么修正后重导，要么照提示逐条改）═══════════
    if errors:
        return ok({
            'rejected': True, 'created': 0, 'updated': 0, 'errors': errors,
            'message': (f'导入未执行：发现 {len(errors)} 处问题，已全部列出。'
                        f'请在表格中按提示修正后重新导入（整表全部通过才会写入，不会漏导）。'),
        })

    # ══ 阶段二：全部通过 → 一次性写入（整体事务，任一失败回滚）═══════════════════════
    created = updated = 0
    try:
        with transaction.atomic():
            for p in plan:
                # Upsert: 业务主键「项目简称 + 交付部门」（简称为空才回退合同名）
                if p['short_name_val']:
                    existing = ARProject.objects.filter(
                        short_name=p['short_name_val'], delivery_dept=p['dept']).first()
                else:
                    existing = ARProject.objects.filter(
                        customer_name=p['customer_name'], delivery_dept=p['dept']).first()
                if existing:
                    for k, v in p['field_vals'].items():
                        setattr(existing, k, v)
                    existing.delivery_dept = p['dept']
                    existing.save()  # triggers post_save signal → updates ARRecord due_dates
                    updated += 1
                else:
                    pr = ARProject(customer_name=p['customer_name'], delivery_dept=p['dept'],
                                   created_by=user, **p['field_vals'])
                    pr.save()
                    created += 1
    except Exception as e:
        return ok({
            'rejected': True, 'created': 0, 'updated': 0,
            'errors': [f'写入阶段发生错误并已回滚：{e}。请检查数据后重试。'],
            'message': '导入未执行（写入阶段出错，已整体回滚，不会出现半截数据）。',
        })

    return ok({'created': created, 'updated': updated, 'skipped': 0, 'errors': []})


@csrf_exempt
@pk_required()
def project_export(request):
    denied = _page_denied(request, 'ar_projects')
    if denied:
        return denied
    qs = _ar_dept_filter(ARProject.objects.all(), request, shared_field='is_shared')
    dept = request.GET.get('dept', '').strip()
    if dept:
        qs = qs.filter(delivery_dept=dept)
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(customer_name__icontains=q) | Q(short_name__icontains=q) |
            Q(project_no__icontains=q))

    if qs.count() > 5000:
        return err('导出超过5000行，请缩小筛选范围')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '项目信息'
    columns = _visible_ar_export_cols(request, [
        (None, '项目编号', lambda p: p.project_no),
        ('p_contract_name', '客户名称', lambda p: p.customer_name),
        ('p_short_name', '项目简称', lambda p: p.short_name),
        ('p_delivery_dept', '交付部门', lambda p: p.delivery_dept),
        ('p_sub_dept', '二级部门', lambda p: p.sub_dept),
        ('p_business_mode', '业务模式', lambda p: p.business_mode),
        ('p_customer_level', '客户等级', lambda p: p.customer_level),
        ('p_sales_contact', '销售对接人', lambda p: p.sales_contact),
        ('p_project_manager', '项目负责人', lambda p: p.project_manager),
        ('p_sales_contact', '共享业务', lambda p: '是' if p.is_shared else '否'),
        ('p_has_contract', '有无合同', lambda p: p.has_contract),
        ('p_contract_date', '签订日期', lambda p: str(p.contract_date) if p.contract_date else ''),
        ('p_account_period', '合同对账期(天)', lambda p: p.reconciliation_days),
        ('p_account_period', '开票等待期(天)', lambda p: p.invoice_wait_days),
        ('p_account_period', '票后等待期(天)', lambda p: p.post_invoice_days),
        ('p_account_period', '总账期(天)', lambda p: p.total_days),
        ('p_account_period', '对账周期起始日', lambda p: p.cycle_start_day),
        ('p_invoice_config', '开票模式', lambda p: p.invoice_mode),
        ('p_invoice_config', '专票/普票', lambda p: p.invoice_type),
        ('p_invoice_config', '税率', lambda p: str(p.tax_rate)),
        ('p_notes', '备注', lambda p: p.notes),
    ])
    _header_row(ws, [header for _, header, _ in columns])
    for p in qs:
        ws.append([getter(p) for _, _, getter in columns])
    return _export_response(wb, '项目信息.xlsx')


@csrf_exempt
@pk_required()
def project_stats(request):
    """KPI strip for 项目台账: 项目总数 / S·A 级客户 / 共享业务 / 本月新签环比."""
    denied = _page_denied(request, 'ar_projects')
    if denied:
        return denied
    qs = _ar_dept_filter(ARProject.objects.all(), request, shared_field='is_shared')
    dept = request.GET.get('dept', '').strip()
    if dept:
        qs = qs.filter(delivery_dept=dept)

    total = qs.count()
    shared = qs.filter(is_shared=True).count()
    draft_count = qs.filter(is_draft=True).count()

    # Customer level breakdown
    level_rows = qs.values('customer_level').annotate(c=Count('id'))
    level_map = {(r['customer_level'] or '未分级'): r['c'] for r in level_rows}
    s_count = level_map.get('S级', 0) + level_map.get('S', 0)
    a_count = level_map.get('A级', 0) + level_map.get('A', 0)

    # Month-over-month new signings by contract_date
    today = datetime.date.today()
    this_start = datetime.date(today.year, today.month, 1)
    if today.month == 1:
        last_start = datetime.date(today.year - 1, 12, 1)
    else:
        last_start = datetime.date(today.year, today.month - 1, 1)
    this_count = qs.filter(contract_date__gte=this_start, contract_date__lte=today).count()
    last_end = this_start - datetime.timedelta(days=1)
    last_count = qs.filter(contract_date__gte=last_start, contract_date__lte=last_end).count()
    if last_count:
        mom = round((this_count - last_count) / last_count * 100, 1)
    else:
        mom = None  # no baseline

    return ok({
        'total': total,
        'shared': shared,
        'draft_count': draft_count,
        's_count': s_count,
        'a_count': a_count,
        'level_map': level_map,
        'new_this_month': this_count,
        'new_last_month': last_count,
        'mom_growth': mom,
    })


# ══════════════════════════════════════════════════════════════════════════════
# AR Records
# ══════════════════════════════════════════════════════════════════════════════

@csrf_exempt
@pk_required()
def ar_records_date_bounds(request):
    """Return min/max payment_date across all accessible AR records.

    Used by the frontend to constrain the pay-date range picker so users
    don't see decades of empty calendar space.
    """
    if request.method != 'GET':
        return err('GET only', 405)
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    qs = _ar_dept_filter(ARRecord.objects.all(), request,
                         shared_field='project__is_shared')
    pay_qs = ARPayment.objects.filter(ar_record__in=qs)
    bounds = pay_qs.aggregate(mn=Min('payment_date'), mx=Max('payment_date'))
    mn = bounds['mn']
    mx = bounds['mx']
    # Fallback: derive from operation_year/month when no payments exist
    if mn is None:
        rec_b = qs.aggregate(miny=Min('operation_year'), minm=Min('operation_month'),
                              maxy=Max('operation_year'), maxm=Max('operation_month'))
        if rec_b['miny']:
            mn = datetime.date(rec_b['miny'], rec_b['minm'] or 1, 1)
        if rec_b['maxy']:
            import calendar as _cal
            m = rec_b['maxm'] or 12
            mn = mn or datetime.date(rec_b['miny'] or rec_b['maxy'], 1, 1)
            mx = datetime.date(rec_b['maxy'], m, _cal.monthrange(rec_b['maxy'], m)[1])
    return ok({
        'min': str(mn) if mn else None,
        'max': str(mx) if mx else None,
    })


@csrf_exempt
@pk_required()
def ar_records(request):
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied

    if request.method == 'GET':
        today = datetime.date.today()
        # Listing queryset — select_related for efficient to_dict() rendering
        qs = _ar_dept_filter(ARRecord.objects.select_related('project', 'created_by'), request,
                             shared_field='project__is_shared')
        qs = _apply_record_filters(qs, request)
        qs = _apply_record_state_filters(qs, request, today)
        qs = _apply_conditions(qs, request, today)
        qs = _apply_record_sort(qs, request)

        include_payments = request.GET.get('include_payments', '') in ('1', 'true')
        page = max(1, int(request.GET.get('page', 1) or 1))
        size = min(200, max(1, int(request.GET.get('size', 50) or 50)))

        # Aggregate queryset — plain queryset (no select_related) avoids any
        # accidental extra JOINs and keeps results consistent with the group-summary
        # endpoint which also uses ARRecord.objects.all() as its base.
        qs_agg = _ar_dept_filter(ARRecord.objects.all(), request,
                                 shared_field='project__is_shared')
        qs_agg = _apply_record_filters(qs_agg, request)
        qs_agg = _apply_record_state_filters(qs_agg, request, today)
        qs_agg = _apply_conditions(qs_agg, request, today)

        # 当前筛选全集的金额合计（不止当前页）——支撑"筛选即合计"。
        # 重要：筛选含 payments 关联（回款日期/含未结清）时 qs_agg 带 JOIN，直接
        # aggregate 会把每条记录的金额按其回款笔数重复累加。改用「无 JOIN 基表 +
        # id IN (子查询)」：DB 端去重，避免把全量 id 拉进内存（大数据量下的热点）。
        matched_ids = qs_agg.order_by().values('id')
        base = ARRecord.objects.filter(id__in=matched_ids)
        total = base.count()
        # Only sum outstanding_amount where > 0: matches the table which renders
        # non-positive rows as "—" (settled/overpaid treated identically as 0).
        agg = base.aggregate(
            est=Sum('estimated_amount'),
            inv=Sum('actual_invoice_amount'),
            tax=Sum('tax_amount'),
            out=Sum('outstanding_amount', filter=Q(outstanding_amount__gt=0)),
            adj=Sum('account_diff_adjustment'),
        )
        collected = (ARPayment.objects.filter(ar_record_id__in=matched_ids)
                     .aggregate(s=Sum('amount'))['s'] or 0)

        # ── 时段合计：本月应收/已收 + 本周应收/已收 ─────────────────────────
        # 基准日期 = 所有日期筛选里最晚的一天（都没选则用今天）
        # 注意：today 不加入候选集，只在没有任何筛选时作为兜底；
        # 否则历史月份筛选（如 2024年3月）会被今天覆盖，导致窗口不联动。
        ref_candidates = []
        year_f = request.GET.get('year', '').strip()
        month_f = request.GET.get('month', '').strip()
        pay_end_f = request.GET.get('pay_end', '').strip()
        if year_f:
            try:
                y_i = int(year_f)
                m_i = int(month_f) if month_f else 12
                ref_candidates.append(
                    datetime.date(y_i, m_i, calendar.monthrange(y_i, m_i)[1]))
            except (ValueError, OverflowError):
                pass
        if pay_end_f:
            try:
                ref_candidates.append(datetime.date.fromisoformat(pay_end_f))
            except ValueError:
                pass
        ref_date = max(ref_candidates) if ref_candidates else today

        # 自然月窗口
        mo_start = datetime.date(ref_date.year, ref_date.month, 1)
        mo_end = datetime.date(ref_date.year, ref_date.month,
                               calendar.monthrange(ref_date.year, ref_date.month)[1])
        # 周窗口（周一 ~ 周日）
        wk_start = ref_date - datetime.timedelta(days=ref_date.weekday())
        wk_end = wk_start + datetime.timedelta(days=6)
        # 周标签随基准日联动：基准周==今天所在周时叫"本周"，否则叫"该周"
        today_wk_start = today - datetime.timedelta(days=today.weekday())
        week_label = '本周' if wk_start == today_wk_start else '该周'

        # 应收：按 due_date 落在窗口内的记录预估金额求和（null due_date 自动排除）
        # 当期应收：due_date 落在基准月内
        month_curr_est = (base.filter(due_date__gte=mo_start, due_date__lte=mo_end)
                          .aggregate(s=Sum('estimated_amount'))['s'] or 0)
        # 逾期应收：due_date 早于基准月且仍有未收余额，取 outstanding_amount 之和
        month_overdue_est = (base.filter(due_date__lt=mo_start, outstanding_amount__gt=0)
                             .aggregate(s=Sum('outstanding_amount'))['s'] or 0)
        week_est = (base.filter(due_date__gte=wk_start, due_date__lte=wk_end)
                    .aggregate(s=Sum('estimated_amount'))['s'] or 0)

        # 已收：按 payment_date 落在窗口内，限当前筛选记录集
        # 当期已收：回款记录所属的 AR 明细到期日 >= 基准月首日（当期或未来到期）
        # 逾期已收：回款记录所属的 AR 明细到期日 <  基准月首日（逾期后补收）
        # 两者之和 = 旧的 month_collected（本月全部实收）
        pay_in_set = ARPayment.objects.filter(ar_record_id__in=matched_ids)
        pay_in_month = pay_in_set.filter(payment_date__gte=mo_start, payment_date__lte=mo_end)
        month_curr_collected = (pay_in_month
                                .filter(ar_record__due_date__gte=mo_start)
                                .aggregate(s=Sum('amount'))['s'] or 0)
        month_overdue_collected = (pay_in_month
                                   .filter(ar_record__due_date__lt=mo_start)
                                   .aggregate(s=Sum('amount'))['s'] or 0)
        week_collected = (pay_in_set
                          .filter(payment_date__gte=wk_start, payment_date__lte=wk_end)
                          .aggregate(s=Sum('amount'))['s'] or 0)

        summary = {
            'count': total,
            'estimated': str(agg['est'] or 0),
            'invoiced': str(agg['inv'] or 0),
            'tax': str(agg['tax'] or 0),
            'outstanding': str(agg['out'] or 0),
            'adj': str(agg['adj'] or 0),
            'collected': str(collected),
            'month_curr_est': str(month_curr_est),
            'month_curr_collected': str(month_curr_collected),
            'month_overdue_est': str(month_overdue_est),
            'month_overdue_collected': str(month_overdue_collected),
            'week_est': str(week_est),
            'week_collected': str(week_collected),
            'ref_date': ref_date.isoformat(),
            'ref_month': f'{ref_date.year}年{ref_date.month}月',
            'ref_week': f'{wk_start.month}/{wk_start.day}~{wk_end.month}/{wk_end.day}',
            'ref_week_label': week_label,
        }
        items = list(qs[(page - 1) * size: page * size])

        perms = get_request_perms(request)
        if include_payments:
            # Prefetch payments
            record_ids = [r.id for r in items]
            pay_qs = ARPayment.objects.filter(ar_record_id__in=record_ids).order_by('payment_no')
            pay_map = defaultdict(list)
            for p in pay_qs:
                pay_map[p.ar_record_id].append(p.to_dict())
            rows = []
            for r in items:
                d = r.to_dict(today=today)
                d['payments'] = pay_map.get(r.id, [])
                rows.append(apply_ar_view_mask(d, perms, 'record'))
        else:
            rows = [apply_ar_view_mask(r.to_dict(today=today), perms, 'record') for r in items]

        return ok({'items': rows, 'total': total, 'page': page, 'size': size,
                   'summary': summary})

    if request.method == 'POST':
        denied = _write_denied(request)
        if denied:
            return denied
        data = _ar_visible_payload(
            request, _parse_body(request), 'record',
            extra=('project_id', 'operation_date', 'operation_year', 'operation_month'))
        project_id = data.get('project_id')
        if not project_id:
            return err('缺少 project_id')
        try:
            proj = ARProject.objects.get(pk=int(project_id))
        except ARProject.DoesNotExist:
            return err('项目不存在', 404)
        denied = _dept_denied(request, proj.delivery_dept, '无权操作此部门')
        if denied:
            return denied
        # 运作日期为正源；兼容旧调用方只传年/月（默认当月1日）
        op_raw = _normalize_date(data.get('operation_date'))
        try:
            op_date = datetime.date.fromisoformat(op_raw) if op_raw else None
        except ValueError:
            op_date = None
        if not op_date:
            year = int(data.get('operation_year', 0) or 0)
            month = int(data.get('operation_month', 0) or 0)
            if not (year and 1 <= month <= 12):
                return err('运作日期无效：请提供 operation_date（YYYY-MM-DD）')
            op_date = datetime.date(year, month, 1)
        if not (2000 <= op_date.year <= 2100):
            return err('运作日期超出合理范围（2000–2100）')
        from paikuan.models import PaikuanUser
        user = PaikuanUser.objects.filter(id=request.pk_uid).first()
        try:
            rec = ARRecord(
                project=proj,
                operation_date=op_date,
                estimated_amount=_dec(data.get('estimated_amount', 0)),
                actual_invoice_amount=_dec(data['actual_invoice_amount']) if data.get('actual_invoice_amount') not in (None, '') else None,
                tax_amount=_dec(data['tax_amount']) if data.get('tax_amount') not in (None, '') else None,
                invoice_date=_normalize_date(data.get('invoice_date')) or None,
                target_collection_date=_normalize_date(data.get('target_collection_date')) or None,
                reconciliation_date=_normalize_date(data.get('reconciliation_date')) or None,
                notes=data.get('notes', '').strip(),
                created_by=user,
            )
            # 已填开票金额必须有开票日期：开票状态/到期推算都依赖日期，缺日期会出现
            # 「已开票却无法算账期」的悬空状态
            if rec.actual_invoice_amount is not None and not rec.invoice_date:
                return err('已填实际开票金额，请同时填写开票日期')
            rec.save()
            # 差额调整走明细：创建时带差额 → 生成一条调整明细（原因可一并提交），
            # account_diff_adjustment 由信号派生为明细合计
            init_diff = _dec(data.get('account_diff_adjustment', 0))
            if init_diff:
                ARAdjustment.objects.create(
                    ar_record=rec, amount=init_diff,
                    reason=(data.get('adjustment_reason') or '').strip() or '录入时调整',
                    created_by=user)
                rec.refresh_from_db()
        except Exception as e:
            return err(str(e))
        return ok(apply_ar_view_mask(
            rec.to_dict(today=datetime.date.today(), include_payments=True),
            get_request_perms(request), 'record'))

    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def ar_record_detail(request, pk):
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    try:
        rec = ARRecord.objects.select_related('project', 'created_by').get(pk=pk)
    except ARRecord.DoesNotExist:
        return err('记录不存在', 404)
    if request.pk_role != 'super_admin':
        allowed = request.pk_depts
        if rec.delivery_dept not in allowed:
            return err('无权访问', 403)
        perms = get_request_perms(request)
        if perms and perms.get('ar_shared_only') and not rec.project.is_shared:
            return err('无权访问', 403)

    today = datetime.date.today()

    if request.method == 'GET':
        return ok(apply_ar_view_mask(rec.to_dict(today=today, include_payments=True),
                                     get_request_perms(request), 'record'))

    if request.method == 'PUT':
        denied = _write_denied(request)
        if denied:
            return denied
        data = _ar_visible_payload(request, _parse_body(request), 'record')
        for field in ('estimated_amount',):
            if field in data:
                setattr(rec, field, _dec(data[field]))
        if 'actual_invoice_amount' in data:
            rec.actual_invoice_amount = _dec(data['actual_invoice_amount']) if data['actual_invoice_amount'] not in (None, '') else None
        if 'account_diff_adjustment' in data:
            # 兼容旧调用方按「合计」编辑：与现合计的差值生成一条调整明细，
            # 明细为正源、合计为派生（与运作年月→日期同款的派生列策略）
            new_total = _dec(data['account_diff_adjustment'])
            delta = new_total - (rec.account_diff_adjustment or Decimal('0'))
            if delta:
                from paikuan.models import PaikuanUser as _PU
                ARAdjustment.objects.create(
                    ar_record=rec, amount=delta,
                    reason=(data.get('adjustment_reason') or '').strip() or '人工调整（按合计修改）',
                    created_by=_PU.objects.filter(id=request.pk_uid).first())
                rec.account_diff_adjustment = new_total
        if 'tax_amount' in data:
            rec.tax_amount = _dec(data['tax_amount']) if data['tax_amount'] not in (None, '') else None
        if 'invoice_date' in data:
            rec.invoice_date = _normalize_date(data['invoice_date']) or None
        if 'reconciliation_date' in data:
            rec.reconciliation_date = _normalize_date(data['reconciliation_date']) or None
        if 'target_collection_date' in data:
            rec.target_collection_date = _normalize_date(data['target_collection_date']) or None
        if 'invoice_batch_no' in data:
            rec.invoice_batch_no = (data['invoice_batch_no'] or '').strip()
        if 'notes' in data:
            rec.notes = data['notes'].strip()
        # 触碰了开票字段时校验配套关系（不触碰则不追溯存量数据，避免改备注也被拦）
        if (('actual_invoice_amount' in data or 'invoice_date' in data)
                and rec.actual_invoice_amount is not None and not rec.invoice_date):
            return err('已填实际开票金额，请同时填写开票日期')
        try:
            rec.save()
        except Exception as e:
            return err(str(e))
        return ok(apply_ar_view_mask(
            rec.to_dict(today=today, include_payments=True),
            get_request_perms(request), 'record'))

    if request.method == 'DELETE':
        denied = _delete_denied(request)
        if denied:
            return denied
        rec.delete()
        return ok({'deleted': pk})

    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def ar_record_template(request):
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '应收账款明细'
    headers = ['项目编号', '项目简称*', '交付部门', '客户名称', '运作日期*', '预估上账金额', '实际开票金额',
               '税额(差额模式手填)', '开票日期', '账实差额调整', '差额原因', '目标回款日期', '回款金额', '回款时间', '备注']
    _header_row(ws, headers, color='1B6E35')
    tip_vals = [
        '选填：项目编号（如 YS-20260101-0001）。当同名项目有多个时，填此列可精确指定，避免歧义',
        '★必填：填写项目台账中的"项目简称"（找不到将拒绝导入并提示先建项目；同名多个时需填交付部门/编号区分）',
        '选填：交付部门/事业部。当不同事业部存在同名项目时，填此列即可精确区分（推荐）',
        '选填：客户名称，用于区分同名项目（同部门同名时再用此列）',
        '★必填：应收发生日期，格式 2026-05-01 / 2026/5/1 均可；只填到月（如 2026-05）按当月1日处理。'
        '旧模板的「运作年」「运作月」两列仍兼容',
        '选填：当月预计上账金额（元）；计算公式：未回款 = 上账金额 + 账实差额调整 - 全部回款之和',
        '选填：实际开票金额（元）；全额模式下税额自动计算；对账时以此确认是否已对账',
        '选填：差额模式时手动填写税额（元）；全额模式：税额=开票金额÷(1+税率)×税率，自动算，无需填',
        '选填：格式 2026-01-15 / 2026/1/15 / 2026年1月15日 均可',
        '选填：账实差额调整（元，可正可负）；未回款 = 上账金额 + 此值 - 已回款；导入后生成一条调整明细，后续可在记录上多次追加不同原因的调整',
        '选填：本次差额的原因（如：运费差/客户扣款/补付/四舍五入），留空记为"导入差额调整"',
        '选填：目标回款日期（业务手工设定的回款目标），格式 2026-02-15；与系统按账期推算的「应收日期」并行',
        '选填：本次回款金额（元）；同一明细行可有多次回款，每次导入新回款均追加，不覆盖历史',
        '选填：回款日期，格式同上（月度回款率按回款日期所在月份计算）',
        '选填备注',
    ]
    ws.append(tip_vals)
    tip_row = ws.max_row
    tip_fill = PatternFill('solid', fgColor='E8F5E9')
    tip_font = Font(italic=True, color='1B5E20', size=9)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(tip_row, c)
        cell.fill = tip_fill
        cell.font = tip_font
        cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[tip_row].height = 60
    ws.append(['', EXAMPLE_ROW_MARKER, '运输事业部', '', '2026-01-05', 100000, 100000, '', '2026-01-15', 0,
               '', '2026-02-15', 30000, '2026-01-20', '示例（此行含"示例"标记，导入时自动跳过）'])
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
    return _export_response(wb, '应收账款明细导入模板.xlsx')


@csrf_exempt
@pk_required()
def ar_record_import(request):
    """应收明细批量导入。

    设计原则：
    - 格式错误（年月/金额/日期非法）→ 整表拒绝，列出全部问题
    - 项目找不到 → 自动创建草稿项目（不阻断导入）
    - 模糊匹配 / 多候选 → 取最新项目并在返回摘要里说明，请人工核查
    - 成功后返回三段式摘要：精确匹配 / 模糊匹配 / 自动新建
    """
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    if request.method != 'POST':
        return err('POST only', 405)
    denied = _write_denied(request)
    if denied:
        return denied
    f = request.FILES.get('file')
    if not f:
        return err('请上传文件')
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
    except Exception as e:
        return err(f'无法读取Excel: {e}')

    headers = [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    # 规范化表头（去 *、括号注释），与 project_import 保持一致，使带备注的表头也能匹配
    col_map = {_norm_header(h): i + 1 for i, h in enumerate(headers)}

    _ALIASES = {
        '项目编号':          ['项目编号', '项目编号(可选,精确指定)', '项目编号*'],
        '项目简称*':         ['项目简称*', '项目简称'],
        '交付部门':          ['交付部门', '事业部', '部门'],
        '客户名称':          ['客户名称', '客户', '合同名称'],
        '运作日期*':         ['运作日期*', '运作日期'],
        '运作年*':           ['运作年*', '运作年'],
        '运作月*':           ['运作月*', '运作月'],
        '预估上账金额':      ['预估上账金额', '预估金额', '上账金额'],
        '实际开票金额':      ['实际开票金额', '开票金额'],
        '税额(差额模式手填)':['税额(差额模式手填)', '税额'],
        '开票日期':          ['开票日期', '开票日期(YYYY-MM-DD)', '开票日期*'],
        '账实差额调整':      ['账实差额调整', '账实差额'],
        '差额原因':          ['差额原因', '差额调整原因', '调整原因'],
        '目标回款日期':      ['目标回款日期', '目标回款', '目标回款日'],
        '回款金额':          ['回款金额'],
        '回款时间':          ['回款时间', '回款时间(YYYY-MM-DD)', '回款日期'],
        '备注':              ['备注'],
    }

    def _resolve_idx(name):
        for h in _ALIASES.get(name, [name]):
            if _norm_header(h) in col_map:
                return col_map[_norm_header(h)]
        return None

    def _cv_raw(row, name):
        idx = _resolve_idx(name)
        return ws.cell(row, idx).value if idx else None

    def _cv(row, name):
        v = _cv_raw(row, name)
        return str(v).strip() if v is not None else ''

    def _num_or_err(raw, label, ri):
        if raw is None or str(raw).strip() == '':
            return None, None
        s = str(raw).strip().replace(',', '').replace('，', '')
        try:
            return Decimal(s), None
        except (InvalidOperation, TypeError):
            return None, (f'第{ri}行：「{label}」填的是"{raw}"，不是有效数字。'
                          f'请填纯数字（如 12000，不要带逗号、空格、"元""万"等单位）')

    from paikuan.models import PaikuanUser
    user = PaikuanUser.objects.filter(id=request.pk_uid).first()
    allowed_depts = None if request.pk_role == 'super_admin' else request.pk_depts

    DATA_COLS = ('项目编号', '项目简称*', '交付部门', '客户名称', '运作日期*', '运作年*', '运作月*',
                 '预估上账金额', '实际开票金额', '税额(差额模式手填)', '开票日期', '账实差额调整',
                 '差额原因', '目标回款日期', '回款金额', '回款时间', '备注')

    errors = []      # 格式错误：会导致整表拒绝
    skipped = 0      # 仅保留字段兼容（恒为0）：示例/空行静默忽略，不计入
    plan = []        # 通过格式校验的行，待写入

    # ══ 阶段一：格式校验（不写库，不查项目）══════════════════════════════════
    for ri in range(2, ws.max_row + 1):
        row_vals = {h: _cv(ri, h) for h in DATA_COLS}
        short_name = row_vals['项目简称*']
        has_any = any(v for v in row_vals.values())

        # 示例/提示行、整行空白：静默忽略（不计入、不报错）
        if EXAMPLE_ROW_MARKER in short_name or short_name.startswith('★'):
            continue
        if not has_any:
            continue
        if not short_name:
            errors.append(f'第{ri}行：缺少「项目简称」，请补填（项目简称是关联应收记录与项目的唯一桥梁）')
            continue

        dept_hint = row_vals['交付部门']
        if dept_hint and dept_hint not in VALID_DEPARTMENTS:
            errors.append(f'第{ri}行：交付部门"{dept_hint}"无效，可选值为：{"/".join(VALID_DEPARTMENTS)}（留空则按项目简称匹配）')
            continue

        # 运作日期（正源）：新模板单列日期；旧模板「运作年+运作月」仍兼容（按当月1日）
        op_date = None
        od_raw = _cv_raw(ri, '运作日期*')
        if od_raw not in (None, ''):
            od_norm = _normalize_date(od_raw)
            try:
                op_date = datetime.date.fromisoformat(od_norm) if od_norm else None
            except ValueError:
                op_date = None
            if op_date is None:
                # 只填到月（2026-05 / 2026/5 / 2026年5月）→ 当月1日
                m = re.match(r'^\s*(\d{4})\s*[-/年.]\s*(\d{1,2})\s*月?\s*$', str(od_raw))
                if m and 1 <= int(m.group(2)) <= 12:
                    op_date = datetime.date(int(m.group(1)), int(m.group(2)), 1)
            if op_date is None:
                errors.append(f'第{ri}行：「运作日期」"{od_raw}"格式无效，'
                              f'请用 2026-05-01（或只填到月 2026-05，按当月1日处理）')
                continue
        else:
            y_raw, m_raw = row_vals['运作年*'], row_vals['运作月*']
            if not (y_raw or m_raw):
                errors.append(f'第{ri}行：缺少「运作日期」（如 2026-05-01）。'
                              f'旧模板可继续填「运作年」「运作月」两列，按当月1日处理')
                continue
            try:
                year, month = int(float(y_raw or 0)), int(float(m_raw or 0))
            except (ValueError, TypeError):
                errors.append(f'第{ri}行：运作年/月必须是数字（当前 年="{y_raw}" 月="{m_raw}"）')
                continue
            if not (2000 <= year <= 2100 and 1 <= month <= 12):
                errors.append(f'第{ri}行：运作年/月无效（年="{y_raw}" 月="{m_raw}"）。'
                              f'运作年需为四位年份（如 2026），运作月需为 1-12')
                continue
            op_date = datetime.date(year, month, 1)
        if not (2000 <= op_date.year <= 2100):
            errors.append(f'第{ri}行：运作日期 {op_date} 超出合理范围（2000–2100）')
            continue

        est, e = _num_or_err(row_vals['预估上账金额'], '预估上账金额', ri)
        if e: errors.append(e); continue
        actual, e = _num_or_err(row_vals['实际开票金额'], '实际开票金额', ri)
        if e: errors.append(e); continue
        tax, e = _num_or_err(row_vals['税额(差额模式手填)'], '税额', ri)
        if e: errors.append(e); continue
        diff, e = _num_or_err(row_vals['账实差额调整'], '账实差额调整', ri)
        if e: errors.append(e); continue
        pay_amount, e = _num_or_err(row_vals['回款金额'], '回款金额', ri)
        if e: errors.append(e); continue

        inv_raw = _cv_raw(ri, '开票日期')
        inv_date = _normalize_date(inv_raw)
        if inv_raw not in (None, '') and inv_date is None:
            errors.append(f'第{ri}行：「开票日期」"{inv_raw}"格式无效，请用 2026-01-15 格式')
            continue
        tgt_raw = _cv_raw(ri, '目标回款日期')
        tgt_date = _normalize_date(tgt_raw)
        if tgt_raw not in (None, '') and tgt_date is None:
            errors.append(f'第{ri}行：「目标回款日期」"{tgt_raw}"格式无效，请用 2026-02-15 格式（选填，可留空）')
            continue
        pay_raw = _cv_raw(ri, '回款时间')
        pay_date = _normalize_date(pay_raw)
        if pay_raw not in (None, '') and pay_date is None:
            errors.append(f'第{ri}行：「回款时间」"{pay_raw}"格式无效，请用 2026-01-20 格式')
            continue
        if pay_amount and pay_amount > 0 and not pay_date:
            errors.append(f'第{ri}行：填了「回款金额」却没填「回款时间」，请补填或清空回款金额')
            continue
        if pay_date and not (pay_amount and pay_amount > 0):
            errors.append(f'第{ri}行：填了「回款时间」却没填有效「回款金额」，请补填或清空回款时间')
            continue

        # 超收检查：回款金额 > 预估上账+账实差额 → 未收余额将为负，整表拒绝并给出指引
        if pay_amount and pay_amount > Decimal('0'):
            pay_base = (est or Decimal('0')) + (diff or Decimal('0'))
            if pay_amount > pay_base:
                outstanding_after = pay_base - pay_amount
                errors.append(
                    f'第{ri}行（{short_name}）：回款金额 {pay_amount} 元'
                    f' > 预估上账{(" + 账实差额调整" if diff else "")} {pay_base} 元，'
                    f'未收余额将为 {outstanding_after:.2f} 元（负数），导入被拒绝。\n'
                    f'    ① 核查金额是否有误，修正后重新导入\n'
                    f'    ② 先不填回款列，导入明细后再在「录入回款」里手动补录\n'
                    f'    ③ 如有账实差额，在「账实差额调整」列填 {(pay_amount - pay_base):.2f}'
                    f'，使未收归零'
                )
                continue

        plan.append({
            'ri': ri, 'short_name': short_name,
            'project_no': row_vals['项目编号'],
            'dept_hint': dept_hint,
            'customer_hint': row_vals['客户名称'],
            'op_date': op_date,
            'est': est, 'actual': actual, 'tax': tax, 'inv_date': inv_date,
            'diff': diff, 'diff_reason': (row_vals['差额原因'] or '').strip(),
            'notes': row_vals['备注'], 'tgt_date': tgt_date,
            'pay_amount': pay_amount, 'pay_date': pay_date,
        })

    # ── 任何格式错误 → 整表拒绝 ──────────────────────────────────────────────
    if errors:
        return ok({
            'rejected': True, 'created': 0, 'skipped': skipped, 'errors': errors,
            'message': (f'导入未执行：发现 {len(errors)} 处格式问题，已全部列出。'
                        f'请按提示修正后重新导入（整表全部通过才会写入）。'),
        })

    # ══ 阶段一·补：项目归属检查（只读，不写库）══════════════════════════════
    # 多个同名项目无法区分、项目编号填了却查不到、或项目在台账中不存在 → 整表拒绝
    # 并给出可选项/指导。（已停用「自动建草稿」：找不到的项目要求先在台账建好。）
    ambiguities = []   # [{ri, input, candidates}]
    bad_nos = []       # [{ri, input, project_no}]
    not_founds = []    # [{ri, input}] 台账中不存在的项目
    for p in plan:
        status, payload = _classify_project_for_import(
            p['short_name'], p['customer_hint'], p['project_no'], allowed_depts,
            dept_hint=p['dept_hint'])
        if status == 'ambiguous':
            ambiguities.append({'ri': p['ri'], 'input': p['short_name'], 'candidates': payload})
        elif status == 'bad_no':
            bad_nos.append({'ri': p['ri'], 'input': p['short_name'], 'project_no': payload})
        elif status == 'not_found':
            not_founds.append({'ri': p['ri'], 'input': p['short_name']})

    if ambiguities or bad_nos or not_founds:
        guide = []
        for b in bad_nos:
            guide.append(f'第{b["ri"]}行：填写的「项目编号」{b["project_no"]} 不存在或不在您的权限范围内，'
                         f'请核对编号，或清空编号改用「项目简称」匹配。')
        for nf in not_founds:
            guide.append(f'第{nf["ri"]}行：项目「{nf["input"]}」在项目台账中不存在（或不在您的权限部门内）。'
                         f'请先到「项目台账」新建该项目，再导入其应收明细。')
        for a in ambiguities:
            depts = sorted({c['delivery_dept'] for c in a['candidates'] if c.get('delivery_dept')})
            tip = ('（这几个项目同名但分属不同事业部，填「交付部门」即可区分）'
                   if len(depts) > 1 else '')
            lines = [f'第{a["ri"]}行：「{a["input"]}」匹配到 {len(a["candidates"])} 个项目，无法确定是哪一个。'
                     f'请在导入表填「交付部门」列区分{tip}，或填「项目编号」精确指定；同名同部门时再用「客户名称」区分：']
            for c in a['candidates']:
                lines.append(f'    · 编号 {c["project_no"]}｜简称 {c["short_name"]}'
                             f'｜部门 {c["delivery_dept"]}｜客户 {c["customer_name"] or "—"}')
            guide.append('\n'.join(lines))
        parts = []
        if not_founds:
            parts.append(f'{len(not_founds)} 处项目不存在')
        if ambiguities:
            parts.append(f'{len(ambiguities)} 处项目无法唯一确定')
        if bad_nos:
            parts.append(f'{len(bad_nos)} 处项目编号无效')
        return ok({
            'rejected': True, 'created': 0, 'skipped': skipped,
            'errors': guide, 'ambiguities': ambiguities, 'bad_nos': bad_nos, 'not_founds': not_founds,
            'message': (f'导入未执行：{"、".join(parts)}。'
                        f'请按下方指引先在项目台账补建项目 / 补「项目编号」「客户名称」后重新导入。'),
        })

    # ══ 阶段二：项目匹配 + 写入（事务）══════════════════════════════════════
    # 说明：应收明细导入为【纯增量】——每个计划行写入一条独立 ARRecord，
    # 绝不按客户/合同/简称做合并或去重（同项目同月多条属正常多笔）。
    # proj_cache 防止同一批次重复创建同名草稿
    proj_cache = {}
    # 按 match_type 收集摘要
    match_buckets = {'exact': {}, 'exact_multi': {}, 'fuzzy': {}, 'fuzzy_multi': {}, 'created': {}}

    created = 0
    pay_warnings = []          # 回款超额行（阶段一已拦截，此处仅作双重保险记录）
    reject_errors = []         # 阶段二兜底：项目不存在/越权（停用草稿后整表回滚）
    with transaction.atomic():
        for p in plan:
            proj, match_type, warn = _resolve_project_for_import(
                p['short_name'], p['customer_hint'], p['dept_hint'], allowed_depts, user, proj_cache,
                project_no=p['project_no'])

            # 已停用「自动建草稿」：项目不存在 / 解析为新建 / 越权 → 收集错误，整表回滚
            if (proj is None or match_type == 'created'
                    or (request.pk_role != 'super_admin'
                        and proj.delivery_dept not in request.pk_depts)):
                reject_errors.append(
                    f'第{p["ri"]}行：项目「{p["short_name"]}」在项目台账中不存在或不在您的权限部门内，'
                    f'请先到「项目台账」新建该项目后再导入应收。')
                continue

            # 记录到摘要 bucket
            key = proj.short_name
            bucket = match_buckets.setdefault(match_type, {})
            if key not in bucket:
                bucket[key] = {
                    'short_name': proj.short_name,
                    'customer_name': proj.customer_name,
                    'project_id': proj.id,
                    'is_draft': proj.is_draft,
                    'matched_to': proj.short_name,
                    'input': p['short_name'],
                    'count': 0,
                    'warn': warn,
                }
            bucket[key]['count'] += 1

            rec = ARRecord(project=proj, operation_date=p['op_date'], created_by=user)
            if p['est'] is not None and _can_ar_view(request, 'r_estimated_amount'):
                rec.estimated_amount = p['est']
            if _can_ar_view(request, 'r_actual_invoice_amount'):
                rec.actual_invoice_amount = p['actual']
            if _can_ar_view(request, 'r_tax_amount'):
                rec.tax_amount = p['tax']
            if _can_ar_view(request, 'r_invoice_date'):
                rec.invoice_date = p['inv_date']
            if p['tgt_date'] and _can_ar_view(request, 'r_due_date'):
                rec.target_collection_date = p['tgt_date']
            if _can_ar_view(request, 'r_notes'):
                rec.notes = p['notes']
            rec.save()
            # 差额走调整明细（合计由信号派生）；须在回款写入前生效，
            # 否则回款校验「未收不为负」时差额还没计入
            if p['diff'] and _can_ar_view(request, 'r_account_diff'):
                ARAdjustment.objects.create(
                    ar_record=rec, amount=p['diff'],
                    reason=p['diff_reason'][:200] or '导入差额调整', created_by=user)
            if p['pay_amount'] and p['pay_date']:
                ARPayment.objects.create(ar_record=rec, payment_no=1,
                                         amount=p['pay_amount'], payment_date=p['pay_date'],
                                         notes='导入回款')
            created += 1
        if reject_errors:
            transaction.set_rollback(True)

    if reject_errors:
        return ok({
            'rejected': True, 'created': 0, 'skipped': skipped, 'errors': reject_errors,
            'message': ('导入未执行：存在项目台账中不存在的项目，'
                        '请先在「项目台账」建好项目再重导（系统已停用导入自动建草稿）。'),
        })

    # ── 汇总输出 ──────────────────────────────────────────────────────────────
    def _bucket_list(btype):
        # 行级写入失败会把 count 减回，过滤掉 count<=0 的空壳，避免摘要里出现 0 条项
        return [v for v in match_buckets.get(btype, {}).values() if v['count'] > 0]

    draft_count = len(_bucket_list('created'))
    match_detail = {
        'exact':       _bucket_list('exact'),
        'exact_multi': _bucket_list('exact_multi'),
        'fuzzy':       _bucket_list('fuzzy'),
        'fuzzy_multi': _bucket_list('fuzzy_multi'),
        'created':     _bucket_list('created'),
    }
    warnings = []
    for btype in ('exact_multi', 'fuzzy', 'fuzzy_multi', 'created'):
        for item in match_detail[btype]:
            if item.get('warn'):
                warnings.append(f'【{item["short_name"]}】{item["warn"]}')
    warnings.extend(pay_warnings)   # 回款超额警告优先展示

    tip = '导入后系统已按现规则自动重算未收回金额。'
    if draft_count:
        tip += f' 有 {draft_count} 个草稿项目需到「项目台账」补充完善。'

    return ok({
        'created': created, 'skipped': skipped, 'errors': [],
        'match_detail': match_detail,
        'warnings': warnings,
        'has_draft_projects': draft_count > 0,
        'draft_count': draft_count,
        'tip': tip,
    })


@csrf_exempt
@pk_required()
def ar_record_export(request):
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    today = datetime.date.today()
    # 导出与列表口径一致：复用同一套筛选 + 条件构建器（含日期相对区间/金额运算符）
    qs = _ar_dept_filter(ARRecord.objects.select_related('project'), request,
                         shared_field='project__is_shared')
    qs = _apply_record_filters(qs, request)
    qs = _apply_record_state_filters(qs, request, today)
    qs = _apply_conditions(qs, request, today)
    qs = _apply_record_sort(qs, request)
    if qs.count() > 5000:
        return err('导出超过5000行，请缩小筛选范围')
    qs = qs.prefetch_related('payments')

    def _pay_dates(rec):
        return ' / '.join(str(p.payment_date) for p in rec.payments.all())

    def _pay_amounts(rec):
        return ' / '.join(f'{float(p.amount):.2f}' for p in rec.payments.all())

    def _pay_total(rec):
        s = sum((p.amount for p in rec.payments.all()), Decimal('0'))
        return float(s) if s else 0.0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '应收账款明细'
    columns = _visible_ar_export_cols(request, [
        (None, '项目编号', lambda rec, st: rec.project.project_no),
        ('p_short_name', '项目简称', lambda rec, st: rec.project.short_name),
        ('p_contract_name', '客户名称', lambda rec, st: rec.project.customer_name),
        ('p_delivery_dept', '交付部门', lambda rec, st: rec.delivery_dept),
        ('p_project_manager', '项目负责人', lambda rec, st: rec.project.project_manager),
        ('p_sales_contact', '销售对接人', lambda rec, st: rec.project.sales_contact),
        ('p_account_period', '总账期(天)', lambda rec, st: rec.project.total_days),
        (None, '运作日期', lambda rec, st: str(rec.operation_date) if rec.operation_date else ''),
        ('r_estimated_amount', '预估上账金额', lambda rec, st: float(rec.estimated_amount)),
        ('r_actual_invoice_amount', '实际开票金额',
         lambda rec, st: float(rec.actual_invoice_amount) if rec.actual_invoice_amount is not None else ''),
        ('r_tax_amount', '税额', lambda rec, st: float(rec.tax_amount) if rec.tax_amount is not None else ''),
        ('r_invoice_date', '开票日期', lambda rec, st: str(rec.invoice_date) if rec.invoice_date else ''),
        ('p_invoice_config', '开票模式', lambda rec, st: rec.project.invoice_mode),
        ('r_account_diff', '账实差额调整', lambda rec, st: float(rec.account_diff_adjustment)),
        ('r_account_diff', '差额明细',
         lambda rec, st: '；'.join(f'{a.reason or "未填原因"}:{a.amount}'
                                   for a in rec.adjustments.all())),
        ('r_outstanding', '未回款金额', lambda rec, st: float(rec.outstanding_amount)),
        # 回款明细：同一明细可有多笔回款，日期/金额按 " / " 并列，另给已回款合计
        ('r_payments', '回款日期', lambda rec, st: _pay_dates(rec)),
        ('r_payments', '回款金额', lambda rec, st: _pay_amounts(rec)),
        ('r_payments', '已回款合计', lambda rec, st: _pay_total(rec)),
        ('r_payments', '预收冲抵金额',
         lambda rec, st: float(sum(p.amount for p in rec.payments.all()
                                   if p.source == '预收抵扣')) or ''),
        ('r_payments', '预收冲抵次数',
         lambda rec, st: sum(1 for p in rec.payments.all() if p.source == '预收抵扣') or ''),
        ('r_due_date', '应收日期', lambda rec, st: str(rec.due_date) if rec.due_date else ''),
        ('r_due_date', '目标回款日期',
         lambda rec, st: str(rec.target_collection_date) if rec.target_collection_date else ''),
        ('r_reconciliation', '对账状态', lambda rec, st: rec.reconciliation_status),
        ('r_reconciliation', '对账日期',
         lambda rec, st: rec.reconciliation_date.strftime('%Y-%m-%d') if rec.reconciliation_date else ''),
        ('r_invoice_status', '开票状态', lambda rec, st: rec.invoice_status),
        ('r_due_date', '是否逾期', lambda rec, st: '是' if st['is_overdue'] else ''),
        ('r_due_date', '逾期天数', lambda rec, st: st['overdue_days'] if st['is_overdue'] else ''),
        ('r_notes', '备注', lambda rec, st: rec.notes),
    ])
    _header_row(ws, [header for _, header, _ in columns], color='1B6E35')
    for rec in qs:
        st = rec.status_dict(today)
        ws.append([getter(rec, st) for _, _, getter in columns])
    return _export_response(wb, '应收账款明细.xlsx')


def _apply_record_filters(qs, request):
    """Shared dimension filters for AR records (used by list + kpi)."""
    project_id = request.GET.get('project_id', '').strip()
    if project_id:
        qs = qs.filter(project_id=int(project_id))
    dept = request.GET.get('dept', '').strip()
    if dept:
        qs = qs.filter(delivery_dept=dept)
    year = request.GET.get('year', '').strip()
    if year:
        qs = qs.filter(operation_year=int(year))
    month = request.GET.get('month', '').strip()
    if month:
        qs = qs.filter(operation_month=int(month))
    manager = request.GET.get('manager', '').strip()
    if manager:
        qs = qs.filter(project__project_manager__icontains=manager)
    is_shared = request.GET.get('is_shared', '').strip()
    if is_shared in ('1', 'true'):
        qs = qs.filter(project__is_shared=True)
    elif is_shared in ('0', 'false'):
        qs = qs.filter(project__is_shared=False)
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(project__short_name__icontains=q) |
            Q(project__customer_name__icontains=q) |
            Q(project__project_no__icontains=q) |
            Q(project__project_manager__icontains=q))
    return qs


def _apply_record_state_filters(qs, request, today=None):
    """Status / invoice / reconciliation / due-date filters for AR records.

    Used by the list view, the export view and the group-summary view so the
    same filter semantics apply everywhere. NOT used by the KPI endpoint, which
    computes its own status breakdown from the unfiltered (by-state) set.
    """
    if today is None:
        today = datetime.date.today()
    eomonth_today = datetime.date(today.year, today.month,
                                  calendar.monthrange(today.year, today.month)[1])

    status = request.GET.get('status', '').strip()
    if status == 'overdue':
        qs = qs.filter(outstanding_amount__gt=0, due_date__lt=today)
    elif status == 'current':
        qs = qs.filter(outstanding_amount__gt=0, due_date__gte=today,
                       due_date__lte=eomonth_today)
    elif status == 'not_due':
        qs = qs.filter(outstanding_amount__gt=0, due_date__gt=eomonth_today)
    elif status == 'settled':
        qs = qs.filter(outstanding_amount__lte=0)
    elif status == 'outstanding':
        qs = qs.filter(outstanding_amount__gt=0)

    inv_status = request.GET.get('invoice_status', '').strip()
    # 与 ARRecord.invoice_status 属性保持一致：已结清（outstanding<=0）优先级最高，
    # 因此「未开票」必须排除已结清记录（未开票 = 未开票 且 仍有未收余额）。
    if inv_status == '未开票':
        qs = qs.filter(actual_invoice_amount__isnull=True, outstanding_amount__gt=0)
    elif inv_status == '已结清':
        qs = qs.filter(outstanding_amount__lte=0)
    elif inv_status == '已开票':
        qs = qs.filter(actual_invoice_amount__isnull=False, outstanding_amount__gt=0)

    recon_status = request.GET.get('reconciliation_status', '').strip()
    if recon_status == '已对账':
        qs = qs.filter(Q(reconciliation_date__isnull=False) | Q(actual_invoice_amount__isnull=False))
    elif recon_status == '未对账':
        qs = qs.filter(reconciliation_date__isnull=True, actual_invoice_amount__isnull=True)

    # 责任状态（responsibility phase）— 按责任归属阶段过滤，与明细列的
    # post_invoice_status 责任链一致。逾期/等待中的细分依赖日期运算，这里只过滤到
    # 责任方所属阶段（settled / 对账阶段 / 开票阶段 / 票后回款阶段），可纯 DB 表达。
    responsibility = request.GET.get('responsibility', '').strip()
    if responsibility:
        no_inv = Q(project__invoice_type='不开票')
        if responsibility == 'settled':
            qs = qs.filter(outstanding_amount__lte=0)
        elif responsibility == 'post':
            # 票后/回款阶段（销售对接人责任）
            qs = qs.filter(
                Q(outstanding_amount__gt=0) & (
                    (no_inv & Q(reconciliation_date__isnull=False)) |
                    (~no_inv & Q(invoice_date__isnull=False))
                )
            )
        elif responsibility == 'invoice':
            # 待开票阶段（开票人责任）
            qs = qs.filter(
                Q(outstanding_amount__gt=0) & ~no_inv & Q(invoice_date__isnull=True) &
                (Q(reconciliation_date__isnull=False) | Q(actual_invoice_amount__isnull=False))
            )
        elif responsibility == 'recon':
            # 对账阶段（非销售责任）
            qs = qs.filter(
                Q(outstanding_amount__gt=0) & (
                    (no_inv & Q(reconciliation_date__isnull=True)) |
                    (~no_inv & Q(invoice_date__isnull=True) &
                     Q(reconciliation_date__isnull=True) & Q(actual_invoice_amount__isnull=True))
                )
            )

    # 回款筛选：pay_status='unpaid' 纯无回款；pay_include_unpaid=1 与日期区间做 OR
    # （"3月回款 + 含未结清"→传 pay_start/pay_end + pay_include_unpaid=1）
    pay_status = request.GET.get('pay_status', '').strip()
    pay_include_unpaid = request.GET.get('pay_include_unpaid', '') in ('1', 'true')
    pay_start = request.GET.get('pay_start', '').strip()
    pay_end = request.GET.get('pay_end', '').strip()
    has_date = bool(pay_start or pay_end)

    if has_date:
        date_q = Q()
        if pay_start:
            date_q &= Q(payments__payment_date__gte=pay_start)
        if pay_end:
            date_q &= Q(payments__payment_date__lte=pay_end)
        if pay_include_unpaid:
            qs = qs.filter(date_q | Q(outstanding_amount__gt=0))
        else:
            qs = qs.filter(date_q)
        qs = qs.distinct()
    elif pay_status == 'unpaid':
        qs = qs.filter(payments__isnull=True)
    elif pay_include_unpaid:
        qs = qs.filter(outstanding_amount__gt=0)
    elif pay_status == 'paid':
        qs = qs.filter(payments__isnull=False).distinct()
    return qs


# 应收明细可排序字段白名单：仅暴露已建索引/安全的列，映射到 ORM 字段（值为元组以支持
# 复合排序，如运作年+月）。前端传 sort=key（升序）或 sort=-key（降序）。
_AR_SORT_FIELDS = {
    'operation': ('operation_date',),
    'due_date': ('due_date',),
    'target_date': ('target_collection_date',),
    'outstanding': ('outstanding_amount',),
    'estimated': ('estimated_amount',),
    'invoiced': ('actual_invoice_amount',),
    'invoice_date': ('invoice_date',),
    'reconciliation_date': ('reconciliation_date',),
    'created': ('created_at',),
    'dept': ('delivery_dept',),
    'project_no': ('project__project_no',),
    'short_name': ('project__short_name',),
    'manager': ('project__project_manager',),
}


def _apply_record_sort(qs, request):
    """按白名单字段做服务端排序；空值统一排末尾，并追加 id 作为稳定次序保证分页确定。

    仅作用于列表查询集（不影响合计/聚合）。非法 sort 键安全忽略，回退模型默认排序。
    """
    raw = (request.GET.get('sort') or '').strip()
    if not raw:
        return qs
    desc = raw.startswith('-')
    key = raw[1:] if desc else raw
    fields = _AR_SORT_FIELDS.get(key)
    if not fields:
        return qs
    terms = [F(f).desc(nulls_last=True) if desc else F(f).asc(nulls_last=True) for f in fields]
    terms.append(F('id').desc() if desc else F('id').asc())  # 唯一键兜底，分页稳定
    return qs.order_by(*terms)


# ── 条件构建器（BI 式多条件筛选）─────────────────────────────────────────────
# 维度类条件仍走既有 flat 参数（dept/status/...，语义单一源）；这里只处理两类"新能力"：
#   date：任选日期字段 + 相对区间(本周/本月/上月/下月/本年/去年/自定义) + 含/不含
#   amt ：数值字段 + 运算符(=0/≠0/>0/<0/>/</区间)
# 前端把这两类条件序列化进 conditions(JSON 数组)下发；多条件 AND，可任意叠加。
_COND_DATE_FIELDS = {'due_date', 'payment_date', 'invoice_date', 'reconciliation_date',
                     'operation_date', 'target_collection_date'}
_COND_AMT_FIELDS = {'estimated_amount', 'outstanding_amount', 'tax_amount',
                    'actual_invoice_amount', 'account_diff_adjustment'}


def _relative_date_range(token, today):
    """相对区间 token → (start, end)（含端点）；未知 token 返回 None。服务端按 today 计算。"""
    y, m = today.year, today.month
    if token == 'this_week':
        start = today - datetime.timedelta(days=today.weekday())
        return start, start + datetime.timedelta(days=6)
    if token == 'this_month':
        return datetime.date(y, m, 1), datetime.date(y, m, calendar.monthrange(y, m)[1])
    if token == 'last_month':
        py, pm = (y - 1, 12) if m == 1 else (y, m - 1)
        return datetime.date(py, pm, 1), datetime.date(py, pm, calendar.monthrange(py, pm)[1])
    if token == 'next_month':
        ny, nm = (y + 1, 1) if m == 12 else (y, m + 1)
        return datetime.date(ny, nm, 1), datetime.date(ny, nm, calendar.monthrange(ny, nm)[1])
    if token == 'this_year':
        return datetime.date(y, 1, 1), datetime.date(y, 12, 31)
    if token == 'last_year':
        return datetime.date(y - 1, 1, 1), datetime.date(y - 1, 12, 31)
    if token == 'last_week':
        start = today - datetime.timedelta(days=today.weekday() + 7)
        return start, start + datetime.timedelta(days=6)
    if token == 'this_quarter':
        q = (m - 1) // 3
        qs, qe = q * 3 + 1, q * 3 + 3
        return datetime.date(y, qs, 1), datetime.date(y, qe, calendar.monthrange(y, qe)[1])
    if token == 'last_quarter':
        q = (m - 1) // 3
        if q == 0:
            py, lqs, lqe = y - 1, 10, 12
        else:
            py, lqs, lqe = y, (q - 1) * 3 + 1, (q - 1) * 3 + 3
        return datetime.date(py, lqs, 1), datetime.date(py, lqe, calendar.monthrange(py, lqe)[1])
    return None


def _dec_or_none(v):
    try:
        return Decimal(str(v))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _parse_ym(s):
    """解析 'YYYY-MM' → (year, month)；非法返回 None。"""
    try:
        ys, ms = str(s).split('-')[:2]
        return int(ys), int(ms)
    except (TypeError, ValueError, AttributeError):
        return None


def _ym_range_q(start_ym, end_ym):
    """运作年月区间 → 标量安全 Q（operation_year/month 双列比较）。
    仅给 start 时退化为单月；start/end 任意顺序都按从小到大归一。"""
    s = _parse_ym(start_ym) if start_ym else None
    e = _parse_ym(end_ym) if end_ym else None
    if s and not e:
        e = s
    if e and not s:
        s = e
    if not (s and e):
        return None
    if s > e:
        s, e = e, s
    (y1, m1), (y2, m2) = s, e
    if y1 == y2:
        return Q(operation_year=y1, operation_month__gte=m1, operation_month__lte=m2)
    q = Q(operation_year=y1, operation_month__gte=m1) | Q(operation_year=y2, operation_month__lte=m2)
    if y2 - y1 >= 2:
        q |= Q(operation_year__gt=y1, operation_year__lt=y2)
    return q


def _condition_q(c, today, eomonth_today):
    """把单个条件构建成「标量安全」的 Q（关联字段一律转 id IN 子查询），
    便于在 且/或 任意组合下安全参与布尔运算。非法条目返回 None。"""
    if not isinstance(c, dict):
        return None
    ctype = c.get('t')

    # ── 条件组（括号）：内部按 group.match(且/或) 组合，整体作为一个标量安全 Q ──
    # 支持「(A 且 B) 或 C」这类带括号的复合筛选，可与顶层条件任意嵌套一层。
    if ctype == 'group':
        inner = c.get('conditions')
        if not isinstance(inner, list):
            return None
        inner_any = (c.get('match') or 'all') == 'any'
        combined = None
        for sub in inner:
            q = _condition_q(sub, today, eomonth_today)
            if q is None:
                continue
            combined = q if combined is None else ((combined | q) if inner_any else (combined & q))
        return combined

    # ── 维度类 ─────────────────────────────────────────────────────────────
    if ctype == 'dim':
        field = c.get('field')
        # 运作年月：支持区间(value~end) + 含/不含(exclude)；在空值守卫之前处理
        if field == 'operation_ym':
            q = _ym_range_q(c.get('value'), c.get('end'))
            if q is None:
                return None
            return ~q if c.get('exclude') else q
        v = c.get('value')
        if v in (None, ''):
            return None
        if field == 'dept':
            return Q(delivery_dept=v)
        if field == 'project_id':
            try:
                return Q(project_id=int(v))
            except (TypeError, ValueError):
                return None
        if field == 'manager':
            return Q(project__project_manager__icontains=v)
        if field == 'is_shared':
            return Q(project__is_shared=(str(v) in ('1', 'true', 'True')))
        if field == 'operation_year':
            try:
                return Q(operation_year=int(v))
            except (TypeError, ValueError):
                return None
        if field == 'operation_month':
            try:
                return Q(operation_month=int(v))
            except (TypeError, ValueError):
                return None
        if field == 'q':
            return (Q(project__short_name__icontains=v) | Q(project__customer_name__icontains=v) |
                    Q(project__project_no__icontains=v) | Q(project__project_manager__icontains=v))
        if field == 'status':
            if v == 'overdue':
                return Q(outstanding_amount__gt=0, due_date__lt=today)
            if v == 'current':
                return Q(outstanding_amount__gt=0, due_date__gte=today, due_date__lte=eomonth_today)
            if v == 'not_due':
                return Q(outstanding_amount__gt=0, due_date__gt=eomonth_today)
            if v == 'settled':
                return Q(outstanding_amount__lte=0)
            if v == 'outstanding':
                return Q(outstanding_amount__gt=0)
            return None
        if field == 'reconciliation_status':
            if v == '已对账':
                return Q(reconciliation_date__isnull=False) | Q(actual_invoice_amount__isnull=False)
            if v == '未对账':
                return Q(reconciliation_date__isnull=True, actual_invoice_amount__isnull=True)
            return None
        if field == 'invoice_status':
            if v == '未开票':
                return Q(actual_invoice_amount__isnull=True, outstanding_amount__gt=0)
            if v == '已结清':
                return Q(outstanding_amount__lte=0)
            if v == '已开票':
                return Q(actual_invoice_amount__isnull=False, outstanding_amount__gt=0)
            return None
        if field == 'responsibility':
            no_inv = Q(project__invoice_type='不开票')
            if v == 'settled':
                return Q(outstanding_amount__lte=0)
            if v == 'post':
                return Q(outstanding_amount__gt=0) & (
                    (no_inv & Q(reconciliation_date__isnull=False)) |
                    (~no_inv & Q(invoice_date__isnull=False)))
            if v == 'invoice':
                return (Q(outstanding_amount__gt=0) & ~no_inv & Q(invoice_date__isnull=True) &
                        (Q(reconciliation_date__isnull=False) | Q(actual_invoice_amount__isnull=False)))
            if v == 'recon':
                return Q(outstanding_amount__gt=0) & (
                    (no_inv & Q(reconciliation_date__isnull=True)) |
                    (~no_inv & Q(invoice_date__isnull=True) &
                     Q(reconciliation_date__isnull=True) & Q(actual_invoice_amount__isnull=True)))
            return None
        return None

    # ── 日期类 ─────────────────────────────────────────────────────────────
    if ctype == 'date':
        field = c.get('field')
        if field not in _COND_DATE_FIELDS:
            return None
        rng = c.get('range')
        if rng == 'custom':
            start = _normalize_date(c.get('start')) or None
            end = _normalize_date(c.get('end')) or None
            if not (start or end):
                return None
        else:
            r = _relative_date_range(rng, today)
            if not r:
                return None
            start, end = r
        exclude = bool(c.get('exclude'))
        if field == 'payment_date':
            # 关联回款集 → 转 id IN 子查询，标量安全；不含=无该区间回款
            sub = ARPayment.objects.all()
            if start:
                sub = sub.filter(payment_date__gte=start)
            if end:
                sub = sub.filter(payment_date__lte=end)
            q = Q(id__in=sub.values('ar_record_id'))
        else:
            q = Q()
            if start:
                q &= Q(**{f'{field}__gte': start})
            if end:
                q &= Q(**{f'{field}__lte': end})
        return ~q if exclude else q

    # ── 金额类 ─────────────────────────────────────────────────────────────
    if ctype == 'amt':
        field = c.get('field')
        if field not in _COND_AMT_FIELDS:
            return None
        op = c.get('op')
        if op == 'gt0':
            return Q(**{f'{field}__gt': 0})
        if op == 'lt0':
            return Q(**{f'{field}__lt': 0})
        if op == 'eq0':
            return Q(**{field: 0})
        if op == 'ne0':
            return Q(**{f'{field}__gt': 0}) | Q(**{f'{field}__lt': 0})
        if op in ('gt', 'lt', 'eq'):
            v = _dec_or_none(c.get('value'))
            if v is None:
                return None
            return Q(**{{'gt': f'{field}__gt', 'lt': f'{field}__lt', 'eq': field}[op]: v})
        if op == 'between':
            lo = _dec_or_none(c.get('min'))
            hi = _dec_or_none(c.get('max'))
            q = Q()
            if lo is not None:
                q &= Q(**{f'{field}__gte': lo})
            if hi is not None:
                q &= Q(**{f'{field}__lte': hi})
            return q if q else None
        return None

    return None


def _apply_conditions(qs, request, today=None):
    """统一条件评估：conditions(JSON 数组) 经 match(all=且/any=或) 组合为单个 Q 后应用。

    所有条件均为标量安全 Q（关联字段已转 id IN 子查询），可在且/或下自由组合，
    无需 distinct。非法条目静默跳过，保证健壮与安全。
    """
    raw = (request.GET.get('conditions') or '').strip()
    if not raw:
        return qs
    try:
        conds = json.loads(raw)
        assert isinstance(conds, list)
    except (ValueError, AssertionError):
        return qs
    if today is None:
        today = datetime.date.today()
    eomonth_today = datetime.date(today.year, today.month,
                                  calendar.monthrange(today.year, today.month)[1])
    match_any = (request.GET.get('match') or 'all').strip() == 'any'

    combined = None
    for c in conds:
        q = _condition_q(c, today, eomonth_today)
        if q is None:
            continue
        if combined is None:
            combined = q
        else:
            combined = (combined | q) if match_any else (combined & q)
    if combined is not None:
        qs = qs.filter(combined)
    return qs


@csrf_exempt
@pk_required()
def ar_records_kpi(request):
    """Per-tracking completion KPIs for the current filter (对账/开票/回款)."""
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    today = datetime.date.today()
    qs = _apply_record_filters(
        _ar_dept_filter(ARRecord.objects.all(), request, shared_field='project__is_shared'), request)
    qs = _apply_conditions(qs, request, today)

    total = qs.count()

    # Reconciliation: 已对账 vs 未对账
    recon_done = qs.filter(Q(reconciliation_date__isnull=False) | Q(actual_invoice_amount__isnull=False)).count()
    recon_pending = total - recon_done

    # Invoice: 已开票 vs 未开票（含金额）。已结清记录不计入「待开票」，与
    # invoice_status 属性及列表筛选保持一致（已结清优先级高于未开票）。
    inv_done_qs = qs.filter(actual_invoice_amount__isnull=False)
    inv_done = inv_done_qs.count()
    inv_pending_qs = qs.filter(actual_invoice_amount__isnull=True, outstanding_amount__gt=0)
    inv_pending = inv_pending_qs.count()
    # 未开票预估金额：尚未开票且仍有未收余额记录的预估上账金额合计
    inv_pending_amount = inv_pending_qs.aggregate(s=Sum('estimated_amount'))['s'] or 0
    inv_done_amount = inv_done_qs.aggregate(s=Sum('actual_invoice_amount'))['s'] or 0

    # Collection: 已结清 vs 未收
    settled = qs.filter(outstanding_amount__lte=0).count()
    outstanding_qs = qs.filter(outstanding_amount__gt=0)
    outstanding_count = outstanding_qs.count()
    outstanding_amount = outstanding_qs.aggregate(s=Sum('outstanding_amount'))['s'] or 0
    # 已收总额（当前筛选集所有回款之和）
    collected_amount = qs.aggregate(s=Sum('payments__amount'))['s'] or 0
    estimated_total = qs.aggregate(s=Sum('estimated_amount'))['s'] or 0

    # Overdue (within current filter)
    overdue_qs = qs.filter(outstanding_amount__gt=0, due_date__lt=today)
    overdue_count = overdue_qs.count()
    overdue_amount = overdue_qs.aggregate(s=Sum('outstanding_amount'))['s'] or 0

    def _rate(done, tot):
        return round(done / tot * 100, 1) if tot else 100.0

    return ok({
        'total': total,
        'estimated_total': str(estimated_total),
        'reconciliation': {
            'done': recon_done, 'pending': recon_pending,
            'rate': _rate(recon_done, total),
            'pending_amount': str(qs.filter(reconciliation_date__isnull=True, actual_invoice_amount__isnull=True)
                                  .aggregate(s=Sum('outstanding_amount'))['s'] or 0),
        },
        'invoice': {
            'done': inv_done, 'pending': inv_pending,
            'rate': _rate(inv_done, total),
            'pending_amount': str(inv_pending_amount),
            'done_amount': str(inv_done_amount),
        },
        'collection': {
            'settled': settled,
            'outstanding_count': outstanding_count,
            'outstanding_amount': str(outstanding_amount),
            'collected_amount': str(collected_amount),
            'rate': _rate(settled, total),
        },
        'overdue': {'count': overdue_count, 'amount': str(overdue_amount)},
    })


# ──────────────────────────────────────────────────────────────────────────────
# 回款流水 (payment ledger) — cross-record payments filtered by date range
# ──────────────────────────────────────────────────────────────────────────────

def _payment_ledger_qs(request):
    """Build the filtered ARPayment queryset for the ledger (shared by list+export)."""
    qs = (ARPayment.objects.select_related('ar_record', 'ar_record__project')
          .order_by('-payment_date', '-id'))
    qs = _ar_dept_filter(qs, request, dept_field='ar_record__delivery_dept',
                         shared_field='ar_record__project__is_shared')
    pay_start = request.GET.get('pay_start', '').strip()
    pay_end = request.GET.get('pay_end', '').strip()
    if pay_start:
        qs = qs.filter(payment_date__gte=pay_start)
    if pay_end:
        qs = qs.filter(payment_date__lte=pay_end)
    dept = request.GET.get('dept', '').strip()
    if dept:
        qs = qs.filter(ar_record__delivery_dept=dept)
    project_id = request.GET.get('project_id', '').strip()
    if project_id:
        qs = qs.filter(ar_record__project_id=int(project_id))
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(ar_record__project__short_name__icontains=q) |
            Q(ar_record__project__customer_name__icontains=q) |
            Q(ar_record__project__project_no__icontains=q))
    return qs


def _payment_ledger_row(p):
    rec = p.ar_record
    proj = rec.project
    return {
        'id': p.id,
        'ar_record_id': rec.id,
        'payment_no': p.payment_no,
        'payment_date': str(p.payment_date) if p.payment_date else None,
        'amount': str(p.amount),
        'notes': p.notes,
        'project_no': proj.project_no,
        'short_name': proj.short_name,
        'delivery_dept': rec.delivery_dept,
        'operation_year': rec.operation_year,
        'operation_month': rec.operation_month,
    }


@csrf_exempt
@pk_required()
def ar_payment_ledger(request):
    """跨记录回款流水：按回款日期区间 / 部门 / 项目 / 关键词筛选，并返回合计。"""
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    denied = _ar_field_denied(request, 'r_payments')
    if denied:
        return denied
    if request.method != 'GET':
        return err('Method not allowed', 405)

    qs = _payment_ledger_qs(request)
    total = qs.count()
    total_amount = qs.aggregate(s=Sum('amount'))['s'] or 0
    page = max(1, int(request.GET.get('page', 1) or 1))
    size = min(200, max(1, int(request.GET.get('size', 50) or 50)))
    rows = [_payment_ledger_row(p) for p in qs[(page - 1) * size: page * size]]
    return ok({
        'items': rows, 'total': total, 'page': page, 'size': size,
        'summary': {'count': total, 'total_amount': str(total_amount)},
    })


@csrf_exempt
@pk_required()
def ar_payment_ledger_export(request):
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    denied = _ar_field_denied(request, 'r_payments')
    if denied:
        return denied
    qs = _payment_ledger_qs(request)
    if qs.count() > 5000:
        return err('导出超过5000行，请缩小筛选范围')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '回款流水'
    headers = ['回款日期', '回款金额', '项目编号', '项目简称', '交付部门',
               '运作年', '运作月', '回款序号', '备注']
    _header_row(ws, headers, color='1B6E35')
    for p in qs:
        r = _payment_ledger_row(p)
        ws.append([r['payment_date'], float(p.amount), r['project_no'], r['short_name'],
                   r['delivery_dept'], r['operation_year'], r['operation_month'],
                   r['payment_no'], r['notes']])
    return _export_response(wb, '回款流水.xlsx')


# ──────────────────────────────────────────────────────────────────────────────
# 多维汇总 (group-by pivot) — count + amount sums grouped by a chosen dimension
# ──────────────────────────────────────────────────────────────────────────────

_SUMMARY_GROUP_FIELDS = {
    'dept': ('delivery_dept', '交付部门'),
    'customer_level': ('project__customer_level', '客户等级'),
    'business_mode': ('project__business_mode', '业务模式'),
    'manager': ('project__project_manager', '项目负责人'),
    'month': ('operation_month', '运作年月'),
}

# 客户等级固定展示顺序：S → A → B → C → D，其余/未填写置后
_CUSTOMER_LEVEL_ORDER = {'S级': 0, 'A级': 1, 'B级': 2, 'C级': 3, 'D级': 4}


@csrf_exempt
@pk_required()
def ar_records_group_summary(request):
    """按指定维度对应收记录做计数 + 金额聚合（透视表）。"""
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    if request.method != 'GET':
        return err('Method not allowed', 405)

    today = datetime.date.today()
    qs = _ar_dept_filter(ARRecord.objects.all(), request, shared_field='project__is_shared')
    qs = _apply_record_filters(qs, request)
    qs = _apply_record_state_filters(qs, request, today)
    qs = _apply_conditions(qs, request, today)
    # 若 flat 筛选带 payments JOIN（旧参数），落到无 JOIN 基表避免分组聚合行扇出翻倍；
    # 用 id IN 子查询代替全量拉 id（大数据量友好）。条件构建器条件已是标量安全，无需此步但无害。
    qs = ARRecord.objects.filter(id__in=qs.order_by().values('id'))

    group_by = request.GET.get('group_by', 'dept').strip() or 'dept'

    # NOTE: record-level sums (estimated/invoiced/outstanding/count) must NOT be
    # computed in the same query as Sum('payments__amount') — the payments JOIN
    # causes row fanout that multiplies record-level values. We always compute the
    # 回款 (collected) sum in a SEPARATE aggregate to keep both sides correct.
    def _amounts(sub_qs):
        base = sub_qs.aggregate(
            count=Count('id'),
            estimated=Sum('estimated_amount'),
            invoiced=Sum('actual_invoice_amount'),
            # Only sum positive outstanding — matches table rendering where
            # non-positive rows display "—" (treated as 0).
            outstanding=Sum('outstanding_amount', filter=Q(outstanding_amount__gt=0)),
        )
        collected = sub_qs.aggregate(s=Sum('payments__amount'))['s'] or 0
        return {
            'count': base['count'] or 0,
            'estimated': str(base['estimated'] or 0),
            'invoiced': str(base['invoiced'] or 0),
            'outstanding': str(base['outstanding'] or 0),
            'collected': str(collected),
        }

    rows = []
    if group_by == 'invoice_status':
        buckets = [
            ('未开票', qs.filter(actual_invoice_amount__isnull=True)),
            ('已结清', qs.filter(outstanding_amount__lte=0, actual_invoice_amount__isnull=False)),
            ('已开票', qs.filter(actual_invoice_amount__isnull=False, outstanding_amount__gt=0)),
        ]
        for label, sub in buckets:
            row = {'key': label, 'label': label}
            row.update(_amounts(sub))
            rows.append(row)
    elif group_by == 'month':
        # 运作年月：按 (年, 月) 分组，年月从高到低排序，下钻需带回年+月。
        agg_fields = ('operation_year', 'operation_month')
        base = qs.values(*agg_fields).annotate(
            count=Count('id'),
            estimated=Sum('estimated_amount'),
            invoiced=Sum('actual_invoice_amount'),
            outstanding=Sum('outstanding_amount', filter=Q(outstanding_amount__gt=0)),
        )
        collected_map = {
            (g['operation_year'], g['operation_month']): (g['collected'] or 0)
            for g in qs.values(*agg_fields).annotate(collected=Sum('payments__amount'))
        }
        ym_rows = []
        for g in base:
            y, m = g['operation_year'], g['operation_month']
            ym_rows.append({
                'key': f'{y}-{m}',
                'label': f'{y}年{m}月',
                'year': y, 'month': m,
                'count': g['count'] or 0,
                'estimated': str(g['estimated'] or 0),
                'invoiced': str(g['invoiced'] or 0),
                'outstanding': str(g['outstanding'] or 0),
                'collected': str(collected_map.get((y, m), 0)),
            })
        # 年月从高到低
        ym_rows.sort(key=lambda r: (r['year'], r['month']), reverse=True)
        rows = ym_rows
    elif group_by in _SUMMARY_GROUP_FIELDS:
        field, _ = _SUMMARY_GROUP_FIELDS[group_by]
        # Base record-level sums (no payments join → no fanout).
        # Filter outstanding to > 0 to match the table's "—" treatment of
        # non-positive values and avoid a negative group total.
        base = (qs.values(field).annotate(
            count=Count('id'),
            estimated=Sum('estimated_amount'),
            invoiced=Sum('actual_invoice_amount'),
            outstanding=Sum('outstanding_amount', filter=Q(outstanding_amount__gt=0)),
        ).order_by('-outstanding'))
        # Collected sums computed separately, keyed by the same group field
        collected_map = {
            g[field]: (g['collected'] or 0)
            for g in qs.values(field).annotate(collected=Sum('payments__amount'))
        }
        for g in base:
            raw_key = g[field]
            key = raw_key if raw_key not in (None, '') else '（未填写）'
            rows.append({
                'key': key,
                'label': str(key),
                'count': g['count'] or 0,
                'estimated': str(g['estimated'] or 0),
                'invoiced': str(g['invoiced'] or 0),
                'outstanding': str(g['outstanding'] or 0),
                'collected': str(collected_map.get(raw_key, 0)),
            })
        # 客户等级固定按 S→A→B→C→D 排序（其余置后）
        if group_by == 'customer_level':
            rows.sort(key=lambda r: _CUSTOMER_LEVEL_ORDER.get(r['key'], 99))
    else:
        return err(f'不支持的分组维度: {group_by}')

    return ok({'group_by': group_by, 'rows': rows})


# ══════════════════════════════════════════════════════════════════════════════
# 催款工作台 (collection workbench) — 逾期分桶 + 责任人聚合 + 一键生成催款行动项
# ══════════════════════════════════════════════════════════════════════════════

_DUNNING_BUCKETS = [
    ('d7',   '1-7天',    1,  7),
    ('d30',  '8-30天',   8,  30),
    ('d60',  '31-60天',  31, 60),
    ('d90',  '61-90天',  61, 90),
    ('d90p', '90天以上', 91, None),
]


def _overdue_qs(request, today):
    """当前用户可见范围内、仍有未收余额且已过应收日期的记录。"""
    qs = _ar_dept_filter(ARRecord.objects.select_related('project'), request,
                         shared_field='project__is_shared')
    return qs.filter(outstanding_amount__gt=0, due_date__isnull=False, due_date__lt=today)


def _bucket_cond(today, lo, hi):
    """逾期天数 ∈ [lo, hi] 对应的 due_date 区间条件（hi=None 表示不封顶）。"""
    cond = Q(due_date__lte=today - datetime.timedelta(days=lo))
    if hi:
        cond &= Q(due_date__gte=today - datetime.timedelta(days=hi))
    return cond


def _open_dunning_actions(record_ids=None):
    """record_id → 未关闭催款行动项 id。通过 source_signal.ar_record_id 关联，
    用于工作台标记「已有催款任务」并在批量生成时去重。"""
    qs = (ActionItem.objects.filter(category='collection',
                                    status__in=['open', 'in_progress'])
          .only('id', 'source_signal'))
    out = {}
    for item in qs:
        rid = (item.source_signal or {}).get('ar_record_id')
        if rid is None:
            continue
        if record_ids is not None and rid not in record_ids:
            continue
        out[rid] = item.id
    return out


@csrf_exempt
@pk_required()
def ar_collection_workbench(request):
    """催款工作台：逾期应收按账龄分桶 + 责任人（销售对接人）聚合 + 明细列表。
    明细带「已有催款任务」标记，与决策行动（ActionItem）打通。"""
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    denied = _ar_field_denied(request, 'r_outstanding')
    if denied:
        return denied
    if request.method != 'GET':
        return err('Method not allowed', 405)

    today = datetime.date.today()
    qs = _overdue_qs(request, today)
    dept = request.GET.get('dept', '').strip()
    if dept:
        qs = qs.filter(delivery_dept=dept)
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(project__short_name__icontains=q) |
            Q(project__customer_name__icontains=q) |
            Q(project__project_no__icontains=q) |
            Q(project__sales_contact__icontains=q))

    # 分桶统计（基于 dept/q 筛选后的全量逾期集，不随 bucket/contact 细分变化）
    buckets = []
    for key, label, lo, hi in _DUNNING_BUCKETS:
        agg = qs.filter(_bucket_cond(today, lo, hi)).aggregate(
            count=Count('id'), amount=Sum('outstanding_amount'))
        buckets.append({'key': key, 'label': label,
                        'count': agg['count'] or 0, 'amount': str(agg['amount'] or 0)})

    # 责任人聚合：销售对接人（回款责任人）维度，按未收金额降序
    by_contact = []
    contact_agg = (qs.values('project__sales_contact')
                   .annotate(count=Count('id'), amount=Sum('outstanding_amount'),
                             oldest_due=Min('due_date'))
                   .order_by('-amount'))
    for g in contact_agg:
        by_contact.append({
            'sales_contact': g['project__sales_contact'] or '（未填写）',
            'count': g['count'],
            'amount': str(g['amount'] or 0),
            'max_overdue_days': (today - g['oldest_due']).days if g['oldest_due'] else 0,
        })

    # 明细：可再按 bucket / 责任人 细分
    items_qs = qs
    bucket = request.GET.get('bucket', '').strip()
    if bucket:
        spec = next((b for b in _DUNNING_BUCKETS if b[0] == bucket), None)
        if spec:
            items_qs = items_qs.filter(_bucket_cond(today, spec[2], spec[3]))
    contact = request.GET.get('contact', '').strip()
    if contact:
        if contact == '（未填写）':
            items_qs = items_qs.filter(project__sales_contact='')
        else:
            items_qs = items_qs.filter(project__sales_contact=contact)

    sum_agg = items_qs.aggregate(count=Count('id'), amount=Sum('outstanding_amount'))
    page = max(1, int(request.GET.get('page', 1) or 1))
    size = min(200, max(1, int(request.GET.get('size', 50) or 50)))
    page_qs = (items_qs.annotate(last_pay=Max('payments__payment_date'))
               .order_by('due_date', 'id')[(page - 1) * size: page * size])
    rows = list(page_qs)
    open_actions = _open_dunning_actions({r.id for r in rows})

    items = []
    for rec in rows:
        proj = rec.project
        items.append({
            'id': rec.id,
            'project_id': rec.project_id,
            'project_no': proj.project_no,
            'short_name': proj.short_name,
            'customer_name': proj.customer_name,
            'delivery_dept': rec.delivery_dept,
            'sales_contact': proj.sales_contact,
            'project_manager': proj.project_manager,
            'operation_year': rec.operation_year,
            'operation_month': rec.operation_month,
            'estimated_amount': str(rec.estimated_amount),
            'outstanding_amount': str(rec.outstanding_amount),
            'due_date': str(rec.due_date),
            'overdue_days': (today - rec.due_date).days,
            'invoice_status': rec.invoice_status,
            'last_payment_date': str(rec.last_pay) if rec.last_pay else None,
            'open_action_id': open_actions.get(rec.id),
        })

    return ok({
        'buckets': buckets,
        'by_contact': by_contact,
        'items': items,
        'total': sum_agg['count'] or 0,
        'page': page, 'size': size,
        'summary': {'count': sum_agg['count'] or 0, 'amount': str(sum_agg['amount'] or 0)},
    })


@csrf_exempt
@pk_required()
def ar_collection_dunning(request):
    """批量生成催款行动项：每条逾期应收一条；已有未关闭催款任务的记录自动跳过。
    生成的行动项出现在财务驾驶舱「决策行动」Tab（category=collection）。"""
    if request.method != 'POST':
        return err('Method not allowed', 405)
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    denied = _write_denied(request)
    if denied:
        return denied

    data = _parse_body(request)
    try:
        id_list = [int(i) for i in (data.get('ids') or [])]
    except (ValueError, TypeError):
        return err('ids 参数格式错误')
    if not id_list:
        return err('请先选择要催款的记录')
    if len(id_list) > 200:
        return err('单次最多生成200条催款任务')

    today = datetime.date.today()
    # 走 _overdue_qs 复核：只允许对可见范围内、确实逾期未收的记录生成任务
    recs = list(_overdue_qs(request, today).filter(pk__in=id_list))
    existing = _open_dunning_actions({r.id for r in recs})
    assignee_override = (data.get('assignee') or '').strip()
    follow_days = 7
    try:
        follow_days = max(1, min(90, int(data.get('follow_days', 7))))
    except (ValueError, TypeError):
        pass

    created = []
    skipped = 0
    for rec in recs:
        if rec.id in existing:
            skipped += 1
            continue
        proj = rec.project
        days = (today - rec.due_date).days
        item = ActionItem(
            title=(f'催款：{proj.short_name or proj.customer_name} '
                   f'{rec.operation_year}/{rec.operation_month:02d} '
                   f'逾期{days}天 ¥{rec.outstanding_amount:,.2f}'),
            description=(f'项目 {proj.project_no} · 客户 {proj.customer_name} · '
                         f'应收日期 {rec.due_date} · 未收 {rec.outstanding_amount:,.2f} 元 · '
                         f'销售对接人 {proj.sales_contact or "未填写"}'),
            bu=rec.delivery_dept,
            category='collection',
            priority='high' if days > 30 else 'medium',
            assignee=assignee_override or proj.sales_contact or '',
            due_date=today + datetime.timedelta(days=follow_days),
            source_signal={'ar_record_id': rec.id, 'project_no': proj.project_no,
                           'outstanding': str(rec.outstanding_amount),
                           'overdue_days': days},
            created_by=request.pk_user,
        )
        item.save()
        created.append(item.id)

    return ok({'created': len(created), 'skipped': skipped, 'action_ids': created,
               'missing': len(id_list) - len(recs)})


# ══════════════════════════════════════════════════════════════════════════════
# Payments sub-resource
# ══════════════════════════════════════════════════════════════════════════════

@csrf_exempt
@pk_required()
def ar_payments(request, pk):
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    try:
        rec = ARRecord.objects.select_related('project').get(pk=pk)
    except ARRecord.DoesNotExist:
        return err('记录不存在', 404)
    if request.pk_role != 'super_admin':
        allowed = request.pk_depts
        if rec.delivery_dept not in allowed:
            return err('无权访问', 403)
    denied = _ar_field_denied(request, 'r_payments')
    if denied:
        return denied

    if request.method == 'GET':
        pays = list(rec.payments.order_by('payment_no'))
        return ok([p.to_dict() for p in pays])

    if request.method == 'POST':
        denied = _action_denied(request, 'ar_collect')
        if denied:
            return denied
        data = _parse_body(request)
        amount = _dec(data.get('amount', 0))
        if amount <= 0:
            return err('回款金额必须大于0')
        pay_date = _normalize_date(data.get('payment_date'))
        if not pay_date:
            return err('回款日期无效')
        try:
            with transaction.atomic():
                last = rec.payments.select_for_update().order_by('-payment_no').first()
                next_no = (last.payment_no + 1) if last else 1
                pay = ARPayment.objects.create(
                    ar_record=rec,
                    payment_no=next_no,
                    amount=amount,
                    payment_date=pay_date,
                    notes=data.get('notes', '').strip(),
                )
        except ValidationError as e:
            return err(str(e.message if hasattr(e, 'message') else e), 400)
        return ok(pay.to_dict())

    return err('Method not allowed', 405)


def _adjustments_payload(rec):
    """调整明细变更后的统一回包：明细列表 + 派生合计 + 最新未收（UI 一次刷新）。"""
    rec.refresh_from_db()
    return {
        'items': [a.to_dict() for a in rec.adjustments.all()],
        'total_adjustment': str(rec.account_diff_adjustment or 0),
        'outstanding_amount': str(rec.outstanding_amount or 0),
    }


@csrf_exempt
@pk_required()
def ar_adjustments(request, pk):
    """GET/POST /records/<pk>/adjustments — 差额调整明细（多次、各带原因与金额）。"""
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    try:
        rec = ARRecord.objects.select_related('project').get(pk=pk)
    except ARRecord.DoesNotExist:
        return err('记录不存在', 404)
    if request.pk_role != 'super_admin' and rec.delivery_dept not in request.pk_depts:
        return err('无权访问', 403)
    denied = _ar_field_denied(request, 'r_account_diff')
    if denied:
        return denied

    if request.method == 'GET':
        return ok(_adjustments_payload(rec))

    if request.method == 'POST':
        denied = _write_denied(request)
        if denied:
            return denied
        data = _parse_body(request)
        amount = _dec(data.get('amount', 0))
        if not amount:
            return err('调整金额不能为0（可正可负）')
        reason = (data.get('reason') or '').strip()
        if not reason:
            return err('请填写调整原因（如：运费差、客户扣款、补付、四舍五入）')
        from paikuan.models import PaikuanUser
        try:
            with transaction.atomic():
                ARAdjustment.objects.create(
                    ar_record=rec, amount=amount, reason=reason[:200],
                    adjust_date=_normalize_date(data.get('adjust_date')) or None,
                    created_by=PaikuanUser.objects.filter(id=request.pk_uid).first())
        except ValidationError as e:
            return err(str(e.message if hasattr(e, 'message') else e), 400)
        return ok(_adjustments_payload(rec))

    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def ar_adjustment_detail(request, pk, aid):
    """DELETE /records/<pk>/adjustments/<aid> — 删除一笔调整（合计随之回退）。"""
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    if request.method != 'DELETE':
        return err('Method not allowed', 405)
    denied = _write_denied(request)
    if denied:
        return denied
    try:
        adj = ARAdjustment.objects.select_related('ar_record').get(pk=aid, ar_record_id=pk)
    except ARAdjustment.DoesNotExist:
        return err('调整明细不存在', 404)
    rec = adj.ar_record
    if request.pk_role != 'super_admin' and rec.delivery_dept not in request.pk_depts:
        return err('无权访问', 403)
    denied = _ar_field_denied(request, 'r_account_diff')
    if denied:
        return denied
    try:
        with transaction.atomic():
            adj.delete()
    except ValidationError as e:
        # 删除该调整会使未收为负（累计回款已超过 上账+剩余差额）→ 拒绝
        return err(str(e.message if hasattr(e, 'message') else e), 400)
    return ok(_adjustments_payload(rec))


@csrf_exempt
@pk_required()
def ar_payment_detail(request, pk, ppk):
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    try:
        pay = ARPayment.objects.select_related('ar_record__project').get(
            pk=ppk, ar_record_id=pk)
    except ARPayment.DoesNotExist:
        return err('回款记录不存在', 404)
    if request.pk_role != 'super_admin':
        allowed = request.pk_depts
        if pay.ar_record.delivery_dept not in allowed:
            return err('无权访问', 403)
        perms = get_request_perms(request)
        if perms and perms.get('ar_shared_only') and not pay.ar_record.project.is_shared:
            return err('无权访问', 403)
    denied = _ar_field_denied(request, 'r_payments')
    if denied:
        return denied

    # 「预收抵扣」回款由预收核销自动生成：金额/日期须在核销侧改（PUT 禁止）；
    # DELETE 即「反向核销」——删除对应预收核销，预收余额恢复、回款级联删除。
    if pay.source == '预收抵扣' and request.method == 'PUT':
        return err('该回款由预收核销生成，请在「预收预付」页修改对应核销，或删除后重新核销', 400)
    if pay.source == '预收抵扣' and request.method == 'DELETE':
        denied = _action_denied(request, 'wo_receive')
        if denied:
            return denied
        wo = AdvanceWriteoff.objects.filter(ar_payment_id=pay.pk).first()
        if wo is None:
            # 核销已不存在（历史数据异常）：直接删孤儿回款，恢复应收
            with transaction.atomic():
                pay.delete()
            return ok({'deleted': ppk, 'reversed_writeoff': None})
        wo_id = wo.pk
        try:
            with transaction.atomic():
                wo.delete()   # 信号级联删本回款 + 重算应收未收与预收余额
        except ValidationError as e:
            return err(str(e.message if hasattr(e, 'message') else e), 400)
        return ok({'deleted': ppk, 'reversed_writeoff': wo_id,
                   'message': '已反向核销：预收余额恢复，应收未收回升'})

    if request.method == 'PUT':
        denied = _write_denied(request)
        if denied:
            return denied
        data = _parse_body(request)
        if 'amount' in data:
            amount = _dec(data['amount'])
            if amount <= 0:
                return err('回款金额必须大于0')
            pay.amount = amount
        if 'payment_date' in data:
            pay.payment_date = _normalize_date(data['payment_date']) or pay.payment_date
        if 'notes' in data:
            pay.notes = data['notes'].strip()
        try:
            pay.save()
        except ValidationError as e:
            return err(str(e.message if hasattr(e, 'message') else e), 400)
        return ok(pay.to_dict())

    if request.method == 'DELETE':
        denied = _delete_denied(request)
        if denied:
            return denied
        pay.delete()
        return ok({'deleted': ppk})

    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def ar_record_recompute(request, pk):
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    if request.method != 'POST':
        return err('POST only', 405)
    denied = _write_denied(request)
    if denied:
        return denied
    try:
        rec = ARRecord.objects.select_related('project').get(pk=pk)
    except ARRecord.DoesNotExist:
        return err('记录不存在', 404)
    denied = _object_dept_denied(request, rec)
    if denied:
        return denied
    rec.recompute_derived(save=True)
    rec.save(update_fields=['updated_at'])
    return ok({'id': rec.id, 'outstanding_amount': str(rec.outstanding_amount)})


# ══════════════════════════════════════════════════════════════════════════════
# Data health — detect inconsistent records left by legacy-template imports
# ══════════════════════════════════════════════════════════════════════════════

@csrf_exempt
@pk_required()
def ar_data_health(request):
    """扫描应收明细，找出口径异常的记录（多为旧模板导入产生的废数据）。

    分两类：
    - negative：预估上账 + 账实差额 − 累计回款 < 0（累计回款超过上账口径），
      不能自动修复，需人工核对预估金额/账实差额/回款明细。
    - stale：存储的未收金额与按现规则重算的结果不一致，但重算结果 ≥ 0，
      可一键重算修复。
    """
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    if request.method != 'GET':
        return err('GET only', 405)

    qs = _ar_dept_filter(ARRecord.objects.select_related('project'), request,
                         shared_field='project__is_shared')
    dept = request.GET.get('dept', '').strip()
    if dept:
        qs = qs.filter(delivery_dept=dept)
    qs = qs.annotate(_paid=Sum('payments__amount'))

    LIMIT = 300
    negative, stale = [], []
    neg_count = stale_count = 0
    for rec in qs.iterator():
        base = rec.estimated_amount or Decimal('0')
        adj = rec.account_diff_adjustment or Decimal('0')
        paid = rec._paid or Decimal('0')
        recomputed = base + adj - paid
        stored = rec.outstanding_amount
        if recomputed < Decimal('0'):
            neg_count += 1
            if len(negative) < LIMIT:
                negative.append({
                    'id': rec.id,
                    'project_no': rec.project.project_no,
                    'short_name': rec.project.short_name or rec.project.customer_name,
                    'delivery_dept': rec.delivery_dept,
                    'operation_year': rec.operation_year,
                    'operation_month': rec.operation_month,
                    'estimated_amount': str(base),
                    'account_diff_adjustment': str(adj),
                    'total_paid': str(paid),
                    'stored_outstanding': str(stored) if stored is not None else None,
                    'deficit': str(-recomputed),
                })
        elif stored is None or stored != recomputed:
            stale_count += 1
            if len(stale) < LIMIT:
                stale.append({
                    'id': rec.id,
                    'project_no': rec.project.project_no,
                    'short_name': rec.project.short_name or rec.project.customer_name,
                    'delivery_dept': rec.delivery_dept,
                    'operation_year': rec.operation_year,
                    'operation_month': rec.operation_month,
                    'estimated_amount': str(base),
                    'account_diff_adjustment': str(adj),
                    'total_paid': str(paid),
                    'stored_outstanding': str(stored) if stored is not None else None,
                    'recomputed_outstanding': str(recomputed),
                })
    return ok({
        'negative_count': neg_count, 'stale_count': stale_count,
        'negative': negative, 'stale': stale, 'limit': LIMIT,
        'has_issues': bool(neg_count or stale_count),
    })


@csrf_exempt
@pk_required()
def ar_records_recompute_bulk(request):
    """按现规则批量重算指定记录的未收金额/税额，用于一键修复 stale 记录。
    会跳过重算后未收为负的记录（这类需人工处理）。Body: {ids: [int, ...]}。"""
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    if request.method != 'POST':
        return err('POST only', 405)
    denied = _write_denied(request)
    if denied:
        return denied
    body = _parse_body(request)
    ids = body.get('ids') or []
    if not isinstance(ids, list) or not ids:
        return err('请提供要重算的记录 ids')
    try:
        ids = [int(i) for i in ids][:1000]
    except (ValueError, TypeError):
        return err('ids 必须为整数列表')

    qs = _ar_dept_filter(ARRecord.objects.select_related('project').filter(pk__in=ids),
                         request, shared_field='project__is_shared')
    fixed, failed = 0, []
    for rec in qs:
        if request.pk_role != 'super_admin' and rec.delivery_dept not in (request.pk_depts or []):
            failed.append({'id': rec.id, 'msg': '无权操作此部门'})
            continue
        try:
            rec.recompute_derived(save=True)
            fixed += 1
        except ValidationError as e:
            failed.append({'id': rec.id,
                           'msg': str(e.message if hasattr(e, 'message') else e)})
    return ok({'fixed': fixed, 'failed': failed})


@csrf_exempt
@pk_required()
def ar_records_bulk_delete(request):
    """批量删除应收明细（级联回款）。两种模式：
      - 显式选择：body {ids:[int,...]}
      - 选择全部筛选集：body {all:true} + 查询串里的 conditions/match（与列表同口径）
    始终受部门作用域与删除权限约束；单次上限 5000 条，防误删全库。"""
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    if request.method != 'POST':
        return err('POST only', 405)
    denied = _delete_denied(request)
    if denied:
        return denied

    body = _parse_body(request)
    today = datetime.date.today()
    base = _ar_dept_filter(ARRecord.objects.select_related('project'), request,
                           shared_field='project__is_shared')

    if body.get('all'):
        # 与列表完全相同的筛选口径（conditions/match 走查询串）
        qs = _apply_record_filters(base, request)
        qs = _apply_record_state_filters(qs, request, today)
        qs = _apply_conditions(qs, request, today)
    else:
        ids = body.get('ids') or []
        if not isinstance(ids, list) or not ids:
            return err('请提供要删除的记录 ids')
        try:
            ids = [int(i) for i in ids]
        except (ValueError, TypeError):
            return err('ids 必须为整数列表')
        qs = base.filter(pk__in=ids)

    # 非超管：再次按可操作部门收紧，杜绝越权删除
    if request.pk_role != 'super_admin':
        qs = qs.filter(delivery_dept__in=(request.pk_depts or []))

    count = qs.count()
    if count == 0:
        return ok({'deleted': 0})
    if count > 5000:
        return err('单次删除上限 5000 条，请先缩小筛选范围')
    # 取 id 后用主键集删除：避免子查询/JOIN 在 delete 级联时的边界问题
    del_ids = list(qs.values_list('id', flat=True))
    ARRecord.objects.filter(pk__in=del_ids).delete()  # 级联删除关联回款
    return ok({'deleted': len(del_ids)})


# ══════════════════════════════════════════════════════════════════════════════
# Analytics
# ══════════════════════════════════════════════════════════════════════════════

@csrf_exempt
@pk_required()
def analytics_aging(request):
    denied = _page_denied(request, 'ar_analytics')
    if denied:
        return denied

    today = datetime.date.today()
    eomonth_today = datetime.date(today.year, today.month,
                                  calendar.monthrange(today.year, today.month)[1])

    qs = _ar_dept_filter(
        ARRecord.objects.filter(outstanding_amount__gt=0), request,
        shared_field='project__is_shared')
    dept = request.GET.get('dept', '').strip()
    if dept:
        qs = qs.filter(delivery_dept=dept)

    buckets = [
        ('current', '当期未到期', 0),
        ('1_30', '逾期1-30天', 1),
        ('31_60', '逾期31-60天', 2),
        ('61_90', '逾期61-90天', 3),
        ('90plus', '逾期90天以上', 4),
    ]

    result = []
    # Not overdue
    current_qs = qs.filter(due_date__gte=today)
    current_agg = current_qs.aggregate(count=Count('id'), amount=Sum('outstanding_amount'))
    result.append({'key': 'current', 'label': '当期未到期',
                   'count': current_agg['count'] or 0,
                   'amount': str(current_agg['amount'] or 0)})

    # Overdue buckets
    for key, label, _ in [('1_30', '逾期1-30天', 0), ('31_60', '逾期31-60天', 0),
                           ('61_90', '逾期61-90天', 0), ('90plus', '逾期90天以上', 0)]:
        if key == '1_30':
            bqs = qs.filter(due_date__lt=today, due_date__gte=today - datetime.timedelta(days=30))
        elif key == '31_60':
            bqs = qs.filter(due_date__lt=today - datetime.timedelta(days=30),
                            due_date__gte=today - datetime.timedelta(days=60))
        elif key == '61_90':
            bqs = qs.filter(due_date__lt=today - datetime.timedelta(days=60),
                            due_date__gte=today - datetime.timedelta(days=90))
        else:
            bqs = qs.filter(due_date__lt=today - datetime.timedelta(days=90))
        agg = bqs.aggregate(count=Count('id'), amount=Sum('outstanding_amount'))
        result.append({'key': key, 'label': label,
                       'count': agg['count'] or 0,
                       'amount': str(agg['amount'] or 0)})

    return ok(result)


@csrf_exempt
@pk_required()
def analytics_collection_rate(request):
    denied = _page_denied(request, 'ar_analytics')
    if denied:
        return denied

    year = _int_param(request, 'year', datetime.date.today().year)
    qs = _ar_dept_filter(ARRecord.objects.filter(operation_year=year), request,
                         shared_field='project__is_shared')
    dept = request.GET.get('dept', '').strip()
    if dept:
        qs = qs.filter(delivery_dept=dept)

    # Estimated amounts by operation_month (billed amount entering that month)
    estimated_by_month = {}
    rec_ids = []
    for row in qs.values('id', 'operation_month', 'estimated_amount'):
        m = row['operation_month']
        estimated_by_month[m] = estimated_by_month.get(m, 0) + float(row['estimated_amount'] or 0)
        rec_ids.append(row['id'])

    # Payments by payment_date month (actual cash received in that month)
    payment_by_month = {}
    if rec_ids:
        for pay in (ARPayment.objects
                    .filter(ar_record__in=rec_ids, payment_date__year=year)
                    .values('payment_date', 'amount')):
            m = pay['payment_date'].month
            payment_by_month[m] = payment_by_month.get(m, 0) + float(pay['amount'] or 0)

    # Build monthly stats: AR base = outstanding carried from prior months + new billing
    months = []
    cumulative_outstanding = 0.0
    for m in range(1, 13):
        billed_this_month = estimated_by_month.get(m, 0.0)
        collected_this_month = payment_by_month.get(m, 0.0)
        ar_base = cumulative_outstanding + billed_this_month
        rate = (collected_this_month / ar_base * 100) if ar_base > 0 else 0.0
        cumulative_outstanding = max(0.0, ar_base - collected_this_month)
        months.append({
            'month': m,
            'receivable': round(ar_base, 2),
            'collected': round(collected_this_month, 2),
            'rate': round(min(rate, 999.99), 2),
        })

    return ok({'year': year, 'months': months})


@csrf_exempt
@pk_required()
def analytics_outstanding_top(request):
    denied = _page_denied(request, 'ar_analytics')
    if denied:
        return denied

    n = min(20, int(request.GET.get('n', 10)))
    qs = _ar_dept_filter(
        ARRecord.objects.filter(outstanding_amount__gt=0), request,
        shared_field='project__is_shared')
    dept = request.GET.get('dept', '').strip()
    if dept:
        qs = qs.filter(delivery_dept=dept)

    by_project = (qs.values('project_id', 'project__project_no', 'project__short_name',
                             'project__customer_name', 'delivery_dept')
                  .annotate(total_outstanding=Sum('outstanding_amount'))
                  .order_by('-total_outstanding')[:n])

    result = []
    for r in by_project:
        result.append({
            'project_id': r['project_id'],
            'project_no': r['project__project_no'],
            'short_name': r['project__short_name'] or r['project__customer_name'],
            'delivery_dept': r['delivery_dept'],
            'total_outstanding': str(r['total_outstanding']),
        })
    return ok(result)


@csrf_exempt
@pk_required()
def analytics_status_dist(request):
    denied = _page_denied(request, 'ar_analytics')
    if denied:
        return denied

    today = datetime.date.today()
    eomonth_today = datetime.date(today.year, today.month,
                                  calendar.monthrange(today.year, today.month)[1])
    qs = _ar_dept_filter(ARRecord.objects.all(), request, shared_field='project__is_shared')
    dept = request.GET.get('dept', '').strip()
    if dept:
        qs = qs.filter(delivery_dept=dept)

    total = qs.count()
    settled = qs.filter(outstanding_amount__lte=0).count()
    overdue = qs.filter(outstanding_amount__gt=0, due_date__lt=today).count()
    current = qs.filter(outstanding_amount__gt=0, due_date__gte=today,
                        due_date__lte=eomonth_today).count()
    not_due = qs.filter(outstanding_amount__gt=0, due_date__gt=eomonth_today).count()
    uninvoiced = qs.filter(actual_invoice_amount__isnull=True).count()

    # Amount aggregates
    def _sum(filter_qs):
        return str(filter_qs.aggregate(s=Sum('outstanding_amount'))['s'] or 0)

    return ok({
        'total': total,
        'settled': {'count': settled,
                    'amount': _sum(qs.filter(outstanding_amount__lte=0))},
        'overdue': {'count': overdue,
                    'amount': _sum(qs.filter(outstanding_amount__gt=0, due_date__lt=today))},
        'current': {'count': current,
                    'amount': _sum(qs.filter(outstanding_amount__gt=0, due_date__gte=today,
                                             due_date__lte=eomonth_today))},
        'not_due': {'count': not_due,
                    'amount': _sum(qs.filter(outstanding_amount__gt=0, due_date__gt=eomonth_today))},
        'uninvoiced': {'count': uninvoiced},
    })


@csrf_exempt
@pk_required()
def analytics_by_pm(request):
    denied = _page_denied(request, 'ar_analytics')
    if denied:
        return denied

    year = _int_param(request, 'year', datetime.date.today().year)
    qs = _ar_dept_filter(ARRecord.objects.filter(operation_year=year), request,
                         shared_field='project__is_shared')
    dept = request.GET.get('dept', '').strip()
    if dept:
        qs = qs.filter(delivery_dept=dept)

    # 记录级金额与回款合计必须分两次聚合：payments JOIN 会让多笔回款的记录
    # 在 estimated/outstanding 求和中被乘倍（与 ar_records_group_summary 同坑同解）。
    rows = (qs.values('project__project_manager')
              .annotate(
                  estimated=Sum('estimated_amount'),
                  outstanding=Sum('outstanding_amount'),
                  project_count=Count('project_id', distinct=True),
              )
              .order_by('-outstanding'))
    paid_map = {
        g['project__project_manager']: (g['s'] or 0)
        for g in qs.values('project__project_manager').annotate(s=Sum('payments__amount'))
    }

    result = []
    for r in rows:
        pm = r['project__project_manager'] or '（未填写）'
        estimated = float(r['estimated'] or 0)
        outstanding = float(r['outstanding'] or 0)
        total_paid = float(paid_map.get(r['project__project_manager']) or 0)
        rate = (total_paid / estimated * 100) if estimated > 0 else 0.0
        result.append({
            'pm': pm,
            'project_count': r['project_count'],
            'estimated': round(estimated, 2),
            'outstanding': round(outstanding, 2),
            'collected': round(total_paid, 2),
            'rate': round(min(max(rate, 0), 999.99), 2),
        })
    return ok(result)


_UE_UNLINKED = '未关联客户'


@csrf_exempt
@pk_required()
def analytics_unit_economics(request):
    """客户/项目 单位经济学：客单价 / 客单成本 / 边际贡献。

    口径：收入与成本均取自 caiwu「项目毛利表」(ProjectMargin，金蝶核算维度)，
    边际贡献 = 收入 − 主营成本；客户归属经 ARProject(短名/合同名) 映射。
    入参：year(必填) month(可选,缺省=全年累计) dept(可选) group_by=project|customer
    """
    denied = _page_denied(request, 'ar_projects')
    if denied:
        return denied
    if request.method != 'GET':
        return err('Method not allowed', 405)
    try:
        year = int(request.GET.get('year') or datetime.date.today().year)
    except (TypeError, ValueError):
        return err('年份无效')
    month = None
    month_raw = (request.GET.get('month') or '').strip()
    if month_raw:
        try:
            month = int(month_raw)
            assert 1 <= month <= 12
        except Exception:
            return err('月份无效')
    group_by = (request.GET.get('group_by') or 'project').strip()
    if group_by not in ('project', 'customer'):
        group_by = 'project'
    dept = (request.GET.get('dept') or '').strip()

    from caiwu.models import ProjectMargin
    pm = ProjectMargin.objects.filter(year=year)
    if month:
        pm = pm.filter(month=month)
    if dept:
        pm = pm.filter(business_unit=dept)
    pm = _ar_dept_filter(pm, request, dept_field='business_unit')

    # 聚合到项目（跨月/跨录入求和），business_unit 取贡献最大者
    proj_agg = {}
    for r in pm.values('project_name', 'business_unit').annotate(
            revenue=Sum('revenue'), cost=Sum('cost'),
            sales_exp=Sum('sales_exp'), mgmt_exp=Sum('mgmt_exp')):
        name = (r['project_name'] or '').strip() or '（未命名项目）'
        rev = float(r['revenue'] or 0)
        a = proj_agg.get(name)
        if a is None:
            a = proj_agg[name] = {'project_name': name, 'business_unit': r['business_unit'],
                                  'revenue': 0.0, 'cost': 0.0, 'sales_exp': 0.0, 'mgmt_exp': 0.0,
                                  '_bu_rev': 0.0}
        if rev > a['_bu_rev']:
            a['business_unit'] = r['business_unit']
            a['_bu_rev'] = rev
        a['revenue'] += rev
        a['cost'] += float(r['cost'] or 0)
        a['sales_exp'] += float(r['sales_exp'] or 0)
        a['mgmt_exp'] += float(r['mgmt_exp'] or 0)

    # 客户映射：ProjectMargin.project_name ←→ ARProject.short_name / customer_name
    aqs = ARProject.objects.all()
    if request.pk_role != 'super_admin':
        aqs = aqs.filter(delivery_dept__in=(request.pk_depts or []))
    name_to_cust = {}
    for p in aqs.select_related('customer').only(
            'short_name', 'customer_name', 'customer_level', 'delivery_dept', 'customer'):
        info = {'customer': (p.customer.name if p.customer_id else None),
                'level': p.customer_level or '', 'dept': p.delivery_dept or ''}
        for k in (p.short_name, p.customer_name):
            k = (k or '').strip()
            if k and k not in name_to_cust:
                name_to_cust[k] = info

    projects = []
    for a in proj_agg.values():
        info = name_to_cust.get(a['project_name'])
        cust = (info['customer'] if info else None) or _UE_UNLINKED
        rev, cost = a['revenue'], a['cost']
        margin = rev - cost
        projects.append({
            'key': a['project_name'], 'label': a['project_name'],
            'project_name': a['project_name'], 'business_unit': a['business_unit'],
            'customer': cust, 'customer_level': (info['level'] if info else ''),
            'revenue': round(rev, 2), 'cost': round(cost, 2),
            'sales_exp': round(a['sales_exp'], 2), 'mgmt_exp': round(a['mgmt_exp'], 2),
            'margin': round(margin, 2),
            'margin_rate': round(margin / rev * 100, 2) if rev else None,
        })

    # 客户聚合
    cust_agg = {}
    for p in projects:
        c = p['customer']
        d = cust_agg.get(c)
        if d is None:
            d = cust_agg[c] = {'customer': c, 'level': p['customer_level'],
                               'revenue': 0.0, 'cost': 0.0, 'margin': 0.0, 'project_count': 0}
        d['revenue'] += p['revenue']
        d['cost'] += p['cost']
        d['margin'] += p['margin']
        d['project_count'] += 1
        if not d['level'] and p['customer_level']:
            d['level'] = p['customer_level']

    tot_rev = round(sum(p['revenue'] for p in projects), 2)
    tot_cost = round(sum(p['cost'] for p in projects), 2)
    tot_margin = round(tot_rev - tot_cost, 2)
    real = [d for c, d in cust_agg.items() if c != _UE_UNLINKED]
    cc = len(real)
    linked_rev = sum(d['revenue'] for d in real)
    linked_cost = sum(d['cost'] for d in real)
    linked_margin = sum(d['margin'] for d in real)

    summary = {
        'revenue': tot_rev, 'cost': tot_cost, 'margin': tot_margin,
        'margin_rate': round(tot_margin / tot_rev * 100, 2) if tot_rev else None,
        'project_count': len(projects),
        'customer_count': cc,
        'arpu': round(linked_rev / cc, 2) if cc else None,           # 客单价
        'acpu': round(linked_cost / cc, 2) if cc else None,          # 客单成本
        'margin_per_customer': round(linked_margin / cc, 2) if cc else None,  # 客均边际贡献
        'linked_coverage': round(linked_rev / tot_rev * 100, 1) if tot_rev else None,
    }

    if group_by == 'customer':
        rows = []
        for d in cust_agg.values():
            rev, margin = d['revenue'], d['margin']
            rows.append({
                'key': d['customer'], 'label': d['customer'], 'customer_level': d['level'],
                'revenue': round(rev, 2), 'cost': round(d['cost'], 2), 'margin': round(margin, 2),
                'margin_rate': round(margin / rev * 100, 2) if rev else None,
                'project_count': d['project_count'], 'unlinked': d['customer'] == _UE_UNLINKED,
            })
        rows.sort(key=lambda x: x['margin'], reverse=True)
    else:
        rows = sorted(projects, key=lambda x: x['margin'], reverse=True)

    return ok({'year': year, 'month': month, 'group_by': group_by,
               'rows': rows, 'summary': summary})


# ── 业财损益全景：财务侧毛利 × 应收侧回款，融合在同一项目/客户维度 ──────────────
def _bf_tag(margin_rate, collect_rate, overdue_rate, revenue):
    """给每个项目/客户打一个"业财融合"标签，一眼看穿盈利与回款的错配。
    口径阈值：毛利率<5% 视为薄利；逾期占比>30% 或 回款率<60% 视为回款承压。"""
    thin_margin = margin_rate is not None and margin_rate < 5
    weak_cash = (overdue_rate is not None and overdue_rate > 30) or \
                (collect_rate is not None and collect_rate < 60)
    if revenue is not None and revenue <= 0:
        return ('idle', '无收入')
    if thin_margin and weak_cash:
        return ('critical', '又薄又难收')          # 红牌：低毛利 + 回款差
    if thin_margin:
        return ('low_margin', '赚收入不赚钱')        # 黄牌：规模大但薄利
    if weak_cash:
        return ('cash_risk', '赚利润收不回钱')        # 黄牌：有毛利但回款承压
    return ('healthy', '优质')                       # 绿牌：双优


@csrf_exempt
@pk_required()
def analytics_business_finance(request):
    """业财损益全景：把财务侧「项目毛利」(ProjectMargin) 与应收侧「回款/账期」
    (ARRecord/ARPayment) 嫁接在同一项目/客户维度，一眼识别两类隐性风险——
    「赚收入不赚钱」(薄利) 与「赚利润收不回钱」(回款承压)。

    入参：year(必填) month(可选,缺省=全年累计) dept(可选) group_by=project|customer
    返回 rows：每行含 盈利侧(revenue/cost/margin/margin_rate) +
              应收侧(estimated/invoiced/outstanding/collected/collect_rate/overdue/overdue_rate)
              + 融合标签(tag/tag_label)；外加 summary。
    """
    denied = _page_denied(request, 'ar_analytics')
    if denied:
        return denied
    if request.method != 'GET':
        return err('Method not allowed', 405)
    try:
        year = int(request.GET.get('year') or datetime.date.today().year)
    except (TypeError, ValueError):
        return err('年份无效')
    month = None
    month_raw = (request.GET.get('month') or '').strip()
    if month_raw:
        try:
            month = int(month_raw)
            assert 1 <= month <= 12
        except Exception:
            return err('月份无效')
    group_by = (request.GET.get('group_by') or 'project').strip()
    if group_by not in ('project', 'customer'):
        group_by = 'project'
    dept = (request.GET.get('dept') or '').strip()
    today = datetime.date.today()

    # ── 盈利侧：ProjectMargin 按项目聚合（沿用 unit-economics 口径）──────────────
    from caiwu.models import ProjectMargin
    pm = ProjectMargin.objects.filter(year=year)
    if month:
        pm = pm.filter(month=month)
    if dept:
        pm = pm.filter(business_unit=dept)
    pm = _ar_dept_filter(pm, request, dept_field='business_unit')
    proj_fin = {}
    for r in pm.values('project_name').annotate(revenue=Sum('revenue'), cost=Sum('cost')):
        name = (r['project_name'] or '').strip() or '（未命名项目）'
        a = proj_fin.setdefault(name, {'revenue': 0.0, 'cost': 0.0})
        a['revenue'] += float(r['revenue'] or 0)
        a['cost'] += float(r['cost'] or 0)

    # ── 项目名 → ARProject(身份+客户) 映射，并保留 project_id 反查应收 ────────────
    aqs = ARProject.objects.all()
    if request.pk_role != 'super_admin':
        aqs = aqs.filter(delivery_dept__in=(request.pk_depts or []))
    name_to_proj = {}
    for p in aqs.select_related('customer').only(
            'id', 'short_name', 'customer_name', 'customer_level', 'delivery_dept', 'customer'):
        info = {'pid': p.id, 'customer': (p.customer.name if p.customer_id else None),
                'level': p.customer_level or '', 'dept': p.delivery_dept or ''}
        for k in (p.short_name, p.customer_name):
            k = (k or '').strip()
            if k and k not in name_to_proj:
                name_to_proj[k] = info

    # ── 应收侧：ARRecord 按项目聚合（避免 payments JOIN 扇出，回款单独查）──────────
    ar_qs = _ar_dept_filter(ARRecord.objects.all(), request, shared_field='project__is_shared')
    ar_qs = ar_qs.filter(operation_year=year)
    if month:
        ar_qs = ar_qs.filter(operation_month=month)
    if dept:
        ar_qs = ar_qs.filter(delivery_dept=dept)
    ar_by_pid = {}
    for r in ar_qs.values('project_id').annotate(
            estimated=Sum('estimated_amount'), invoiced=Sum('actual_invoice_amount'),
            outstanding=Sum('outstanding_amount')):
        ar_by_pid[r['project_id']] = {
            'estimated': float(r['estimated'] or 0), 'invoiced': float(r['invoiced'] or 0),
            'outstanding': float(r['outstanding'] or 0), 'collected': 0.0, 'overdue': 0.0}
    # 逾期未收：已过应收日且仍有余额
    for r in ar_qs.filter(due_date__lt=today, outstanding_amount__gt=0).values(
            'project_id').annotate(overdue=Sum('outstanding_amount')):
        if r['project_id'] in ar_by_pid:
            ar_by_pid[r['project_id']]['overdue'] = float(r['overdue'] or 0)
    # 已回款：单独从 ARPayment 聚合，规避反向 JOIN 重复求和
    for r in ARPayment.objects.filter(ar_record__in=ar_qs).values(
            'ar_record__project_id').annotate(collected=Sum('amount')):
        pid = r['ar_record__project_id']
        if pid in ar_by_pid:
            ar_by_pid[pid]['collected'] = float(r['collected'] or 0)

    # ── 融合：以盈利侧项目为主线，并接应收侧 ──────────────────────────────────────
    def _rate(num, den):
        return round(num / den * 100, 1) if den else None

    projects = []
    for name, fin in proj_fin.items():
        info = name_to_proj.get(name)
        ar = ar_by_pid.get(info['pid']) if info else None
        rev, cost = fin['revenue'], fin['cost']
        margin = rev - cost
        mr = _rate(margin, rev)
        est = ar['estimated'] if ar else 0.0
        inv = ar['invoiced'] if ar else 0.0
        out = ar['outstanding'] if ar else 0.0
        col = ar['collected'] if ar else 0.0
        ovd = ar['overdue'] if ar else 0.0
        cr = _rate(col, inv) if inv else None          # 回款率 = 已收/已开票
        ovr = _rate(ovd, out) if out else None          # 逾期占比 = 逾期/未收
        tag, tag_label = _bf_tag(mr, cr, ovr, rev)
        projects.append({
            'key': name, 'label': name, 'project_name': name,
            'business_unit': (info['dept'] if info else ''),
            'customer': (info['customer'] if info else None) or _UE_UNLINKED,
            'customer_level': (info['level'] if info else ''),
            'linked': bool(ar),
            'revenue': round(rev, 2), 'cost': round(cost, 2),
            'margin': round(margin, 2), 'margin_rate': mr,
            'estimated': round(est, 2), 'invoiced': round(inv, 2),
            'outstanding': round(out, 2), 'collected': round(col, 2),
            'collect_rate': cr, 'overdue': round(ovd, 2), 'overdue_rate': ovr,
            'tag': tag, 'tag_label': tag_label,
        })

    # ── 客户聚合 ────────────────────────────────────────────────────────────────
    cust_agg = {}
    for p in projects:
        c = p['customer']
        d = cust_agg.get(c)
        if d is None:
            d = cust_agg[c] = {'customer': c, 'level': p['customer_level'], 'project_count': 0,
                               'revenue': 0.0, 'cost': 0.0, 'margin': 0.0,
                               'invoiced': 0.0, 'outstanding': 0.0, 'collected': 0.0, 'overdue': 0.0}
        for k in ('revenue', 'cost', 'margin', 'invoiced', 'outstanding', 'collected', 'overdue'):
            d[k] += p[k]
        d['project_count'] += 1
        if not d['level'] and p['customer_level']:
            d['level'] = p['customer_level']

    def _pack_customer(d):
        rev, margin = d['revenue'], d['margin']
        mr = _rate(margin, rev)
        cr = _rate(d['collected'], d['invoiced']) if d['invoiced'] else None
        ovr = _rate(d['overdue'], d['outstanding']) if d['outstanding'] else None
        tag, tag_label = _bf_tag(mr, cr, ovr, rev)
        return {
            'key': d['customer'], 'label': d['customer'], 'customer_level': d['level'],
            'project_count': d['project_count'], 'unlinked': d['customer'] == _UE_UNLINKED,
            'revenue': round(rev, 2), 'cost': round(d['cost'], 2),
            'margin': round(margin, 2), 'margin_rate': mr,
            'invoiced': round(d['invoiced'], 2), 'outstanding': round(d['outstanding'], 2),
            'collected': round(d['collected'], 2), 'collect_rate': cr,
            'overdue': round(d['overdue'], 2), 'overdue_rate': ovr,
            'tag': tag, 'tag_label': tag_label,
        }

    # ── 汇总 ────────────────────────────────────────────────────────────────────
    tot_rev = round(sum(p['revenue'] for p in projects), 2)
    tot_cost = round(sum(p['cost'] for p in projects), 2)
    tot_margin = round(tot_rev - tot_cost, 2)
    tot_inv = round(sum(p['invoiced'] for p in projects), 2)
    tot_out = round(sum(p['outstanding'] for p in projects), 2)
    tot_col = round(sum(p['collected'] for p in projects), 2)
    tot_ovd = round(sum(p['overdue'] for p in projects), 2)
    by_tag = {}
    for p in projects:
        t = by_tag.setdefault(p['tag'], {'count': 0, 'revenue': 0.0, 'margin': 0.0, 'overdue': 0.0})
        t['count'] += 1
        t['revenue'] += p['revenue']
        t['margin'] += p['margin']
        t['overdue'] += p['overdue']
    summary = {
        'revenue': tot_rev, 'cost': tot_cost, 'margin': tot_margin,
        'margin_rate': _rate(tot_margin, tot_rev),
        'invoiced': tot_inv, 'outstanding': tot_out, 'collected': tot_col, 'overdue': tot_ovd,
        'collect_rate': _rate(tot_col, tot_inv), 'overdue_rate': _rate(tot_ovd, tot_out),
        'project_count': len(projects),
        'linked_count': sum(1 for p in projects if p['linked']),
        'by_tag': {k: {'count': v['count'], 'revenue': round(v['revenue'], 2),
                       'margin': round(v['margin'], 2), 'overdue': round(v['overdue'], 2)}
                   for k, v in by_tag.items()},
    }

    if group_by == 'customer':
        rows = [_pack_customer(d) for d in cust_agg.values()]
        rows.sort(key=lambda x: x['margin'], reverse=True)
    else:
        rows = sorted(projects, key=lambda x: x['revenue'], reverse=True)

    return ok({'year': year, 'month': month, 'group_by': group_by,
               'rows': rows, 'summary': summary})


@csrf_exempt
@pk_required()
def analytics_project_pnl(request):
    """单项目业财损益卡：一个项目从「合同→开票→回款→毛利」的全链路。
    盈利侧逐月取自 ProjectMargin，应收侧逐月取自 ARRecord/ARPayment。
    入参：name(项目名,必填) year(必填)。返回 project/monthly/payments/totals。
    """
    denied = _page_denied(request, 'ar_analytics')
    if denied:
        return denied
    if request.method != 'GET':
        return err('Method not allowed', 405)
    name = (request.GET.get('name') or '').strip()
    if not name:
        return err('缺少项目名')
    try:
        year = int(request.GET.get('year') or datetime.date.today().year)
    except (TypeError, ValueError):
        return err('年份无效')
    today = datetime.date.today()

    # ── 盈利侧：ProjectMargin 按月 ──────────────────────────────────────────────
    from caiwu.models import ProjectMargin
    pm = _ar_dept_filter(ProjectMargin.objects.filter(year=year, project_name=name),
                         request, dept_field='business_unit')
    fin_by_m = {}
    for r in pm.values('month').annotate(revenue=Sum('revenue'), cost=Sum('cost'),
                                         sales_exp=Sum('sales_exp'), mgmt_exp=Sum('mgmt_exp')):
        fin_by_m[r['month']] = {'revenue': float(r['revenue'] or 0), 'cost': float(r['cost'] or 0),
                                'sales_exp': float(r['sales_exp'] or 0), 'mgmt_exp': float(r['mgmt_exp'] or 0)}

    # ── 项目身份：name → ARProject（短名或合同名匹配）──────────────────────────
    aqs = ARProject.objects.select_related('customer')
    if request.pk_role != 'super_admin':
        aqs = aqs.filter(delivery_dept__in=(request.pk_depts or []))
    projs = list(aqs.filter(Q(short_name=name) | Q(customer_name=name)))
    proj = projs[0] if projs else None
    pids = [p.id for p in projs]

    # ── 应收侧：ARRecord 按月（回款单独从 ARPayment 取，规避扇出）────────────────
    ar_by_m = {}
    arq = ARRecord.objects.none()
    if pids:
        arq = ARRecord.objects.filter(project_id__in=pids, operation_year=year)
        for r in arq.values('operation_month').annotate(
                estimated=Sum('estimated_amount'), invoiced=Sum('actual_invoice_amount'),
                outstanding=Sum('outstanding_amount'), tax=Sum('tax_amount')):
            ar_by_m[r['operation_month']] = {
                'estimated': float(r['estimated'] or 0), 'invoiced': float(r['invoiced'] or 0),
                'outstanding': float(r['outstanding'] or 0), 'tax': float(r['tax'] or 0),
                'collected': 0.0, 'overdue': 0.0}
        for r in arq.filter(due_date__lt=today, outstanding_amount__gt=0).values(
                'operation_month').annotate(overdue=Sum('outstanding_amount')):
            if r['operation_month'] in ar_by_m:
                ar_by_m[r['operation_month']]['overdue'] = float(r['overdue'] or 0)
        for r in ARPayment.objects.filter(ar_record__in=arq).values(
                'ar_record__operation_month').annotate(c=Sum('amount')):
            m = r['ar_record__operation_month']
            if m in ar_by_m:
                ar_by_m[m]['collected'] = float(r['c'] or 0)

    # ── 逐月融合 ────────────────────────────────────────────────────────────────
    def _rate(num, den):
        return round(num / den * 100, 1) if den else None
    monthly = []
    for m in sorted(set(fin_by_m) | set(ar_by_m)):
        f = fin_by_m.get(m, {})
        a = ar_by_m.get(m, {})
        rev, cost = f.get('revenue', 0.0), f.get('cost', 0.0)
        monthly.append({
            'month': m, 'revenue': round(rev, 2), 'cost': round(cost, 2),
            'margin': round(rev - cost, 2), 'margin_rate': _rate(rev - cost, rev),
            'invoiced': round(a.get('invoiced', 0.0), 2), 'outstanding': round(a.get('outstanding', 0.0), 2),
            'collected': round(a.get('collected', 0.0), 2), 'overdue': round(a.get('overdue', 0.0), 2),
        })

    # ── 回款流水（时间线）──────────────────────────────────────────────────────
    payments = []
    if pids:
        for pay in (ARPayment.objects.filter(ar_record__in=arq)
                    .select_related('ar_record').order_by('payment_date', 'id')[:200]):
            payments.append({
                'date': str(pay.payment_date) if pay.payment_date else None,
                'amount': float(pay.amount or 0), 'source': pay.source or '回款',
                'operation_month': pay.ar_record.operation_month})

    # ── 汇总 + 标签 ────────────────────────────────────────────────────────────
    t_rev = round(sum(x['revenue'] for x in monthly), 2)
    t_cost = round(sum(x['cost'] for x in monthly), 2)
    t_margin = round(t_rev - t_cost, 2)
    t_inv = round(sum(x['invoiced'] for x in monthly), 2)
    t_out = round(sum(x['outstanding'] for x in monthly), 2)
    t_col = round(sum(x['collected'] for x in monthly), 2)
    t_ovd = round(sum(x['overdue'] for x in monthly), 2)
    mr = _rate(t_margin, t_rev)
    cr = _rate(t_col, t_inv)
    ovr = _rate(t_ovd, t_out)
    tag, tag_label = _bf_tag(mr, cr, ovr, t_rev)
    totals = {'revenue': t_rev, 'cost': t_cost, 'margin': t_margin, 'margin_rate': mr,
              'invoiced': t_inv, 'outstanding': t_out, 'collected': t_col, 'collect_rate': cr,
              'overdue': t_ovd, 'overdue_rate': ovr, 'tag': tag, 'tag_label': tag_label}

    # ── 付款侧（排款台账，经「项目简称/项目编号」打通）─────────────────────────
    # 排款新增 project_short_name 字段后，项目维度可闭环：回款（流入）对照
    # 排款付款（流出），得到项目净现金。仅匹配非空键，防止空串误匹配全表。
    payment_side = None
    if proj:
        cond = Q()
        if proj.short_name:
            cond |= Q(project_short_name=proj.short_name)
        if proj.project_no:
            cond |= Q(project_no=proj.project_no)
        if cond:
            pq = Payment.objects.filter(cond)
            if request.pk_role != 'super_admin':
                pq = pq.filter(department__in=(request.pk_depts or []))
            agg = pq.annotate(paid_sum=_paid_subq()).aggregate(
                planned=Sum(Coalesce('plan_adjustment', 'total_amount')),
                paid=Sum('paid_sum'), offset=Sum('prepaid_offset_amount'), n=Count('id'))
            planned = float(agg['planned'] or 0)
            paid = float(agg['paid'] or 0)
            offset = float(agg['offset'] or 0)
            payment_side = {
                'count': agg['n'] or 0,
                'planned': round(planned, 2),
                'paid': round(paid, 2),
                'prepaid_offset': round(offset, 2),
                'remaining': round(max(0.0, planned - paid - offset), 2),
                'net_cash': round(t_col - paid, 2),
            }

    project = {
        'name': name,
        'project_no': (proj.project_no if proj else None),
        'short_name': (proj.short_name if proj else name),
        'customer_name': (proj.customer_name if proj else None),
        'customer': (proj.customer.name if (proj and proj.customer_id) else None),
        'customer_level': (proj.customer_level if proj else ''),
        'delivery_dept': (proj.delivery_dept if proj else ''),
        'project_manager': (proj.project_manager if proj else ''),
        'sales_contact': (proj.sales_contact if proj else ''),
        'has_contract': (proj.has_contract if proj else ''),
        'contract_date': (str(proj.contract_date) if (proj and proj.contract_date) else None),
        'total_days': (proj.total_days if proj else None),
        'linked': bool(pids),
    }
    return ok({'year': year, 'project': project, 'monthly': monthly,
               'payments': payments, 'totals': totals, 'payment_side': payment_side})


# ── 现金流多维聚合的维度注册表 ─────────────────────────────────────────────────
# 应收侧键：ARRecord/ARPayment 经 project 关联取值；付款侧键：Payment 自身字段。
# 新增维度（如客户、负责人）只需在两侧模型补字段后在此登记，前端维度切换即生效。
_CASHFLOW_DIMS = {
    'project': {
        'ar_rec_key': 'project__short_name',
        'ar_pay_key': 'ar_record__project__short_name',
        'pk_key': 'payment__project_short_name',
        'label': '项目',
    },
    'secondary_dept': {
        'ar_rec_key': 'project__sub_dept',
        'ar_pay_key': 'ar_record__project__sub_dept',
        'pk_key': 'payment__secondary_dept',
        'label': '二级部门',
    },
}


@csrf_exempt
@pk_required()
def analytics_project_cashflow(request):
    """现金流多维聚合：按维度（项目简称/二级部门）聚合区间内回款（流入）、
    排款付款（流出）、应收敞口（未收）。
    入参：group_by(project|secondary_dept,默认project), dept(可选),
          year(默认今年), date_start/date_end(可选，优先级高于year)
    返回：按净现金降序排列的维度成员列表 + 汇总。"""
    denied = _page_denied(request, 'ar_analytics')
    if denied:
        return denied
    if request.method != 'GET':
        return err('Method not allowed', 405)
    dept = (request.GET.get('dept') or '').strip()
    group_by = (request.GET.get('group_by') or 'project').strip()
    dim = _CASHFLOW_DIMS.get(group_by) or _CASHFLOW_DIMS['project']
    if group_by not in _CASHFLOW_DIMS:
        group_by = 'project'
    try:
        year = int(request.GET.get('year') or datetime.date.today().year)
    except (TypeError, ValueError):
        year = datetime.date.today().year

    # Date range: explicit start/end overrides year-based default
    _ds_raw = _normalize_date(request.GET.get('date_start'))
    _de_raw = _normalize_date(request.GET.get('date_end'))
    try:
        date_start = datetime.date.fromisoformat(_ds_raw) if _ds_raw else datetime.date(year, 1, 1)
        date_end = datetime.date.fromisoformat(_de_raw) if _de_raw else datetime.date(year, 12, 31)
    except (ValueError, AttributeError):
        date_start, date_end = datetime.date(year, 1, 1), datetime.date(year, 12, 31)

    rec_key, arp_key, pk_key = dim['ar_rec_key'], dim['ar_pay_key'], dim['pk_key']

    # ── 应收侧 base queryset ──────────────────────────────────────────────────
    ar_base = _ar_dept_filter(ARRecord.objects.all(), request, shared_field='project__is_shared')
    if dept:
        ar_base = ar_base.filter(delivery_dept=dept)

    # 当前应收敞口（全周期，当前余额；按维度键聚合）
    outstanding_by_key = {}
    for r in (ar_base
              .filter(**{f'{rec_key}__isnull': False})
              .exclude(**{rec_key: ''})
              .values(rec_key, 'delivery_dept')
              .annotate(outstanding=Sum('outstanding_amount'), estimated=Sum('estimated_amount'))):
        key = r[rec_key]
        if key not in outstanding_by_key:
            outstanding_by_key[key] = {'outstanding': 0.0, 'estimated': 0.0, 'dept': r['delivery_dept'] or ''}
        outstanding_by_key[key]['outstanding'] += float(r['outstanding'] or 0)
        outstanding_by_key[key]['estimated'] += float(r['estimated'] or 0)

    # 区间内回款（流入）
    inflow_by_key = {}
    for r in (ARPayment.objects
              .filter(ar_record__in=ar_base,
                      payment_date__gte=date_start,
                      payment_date__lte=date_end)
              .filter(**{f'{arp_key}__isnull': False})
              .exclude(**{arp_key: ''})
              .values(arp_key)
              .annotate(total=Sum('amount'))):
        inflow_by_key[r[arp_key]] = float(r['total'] or 0)

    # 区间内付款（流出）：PaymentInstallment 按维度键聚合
    outflow_by_key = {}
    pi_base = PaymentInstallment.objects.filter(pay_date__gte=date_start, pay_date__lte=date_end)
    if request.pk_role != 'super_admin':
        pi_base = pi_base.filter(payment__department__in=(request.pk_depts or []))
    if dept:
        pi_base = pi_base.filter(payment__department=dept)
    pi_base = pi_base.exclude(**{pk_key: ''})
    for r in pi_base.values(pk_key).annotate(total=Sum('pay_amount')):
        outflow_by_key[r[pk_key]] = float(r['total'] or 0)

    # ── 合并所有维度成员 ──────────────────────────────────────────────────────
    all_names = set(outstanding_by_key) | set(inflow_by_key) | set(outflow_by_key)
    proj_meta = {}
    if group_by == 'project' and all_names:
        for p in ARProject.objects.filter(short_name__in=list(all_names)):
            proj_meta[p.short_name] = {
                'dept': p.delivery_dept or '',
                'project_no': p.project_no or '',
                'customer': p.customer_name or '',
                'manager': p.project_manager or '',
            }

    rows = []
    for name in all_names:
        meta = proj_meta.get(name, {})
        row_dept = meta.get('dept') or outstanding_by_key.get(name, {}).get('dept', '')
        # If dept filter active and this member doesn't match, skip
        if dept and row_dept and row_dept != dept:
            continue
        inflow = inflow_by_key.get(name, 0.0)
        outflow = outflow_by_key.get(name, 0.0)
        out_info = outstanding_by_key.get(name, {})
        rows.append({
            'project': name,            # 维度成员名（project 维=项目简称，secondary_dept 维=二级部门名）
            'dept': row_dept,
            'project_no': meta.get('project_no', ''),
            'customer': meta.get('customer', ''),
            'manager': meta.get('manager', ''),
            'inflow': round(inflow, 2),
            'outflow': round(outflow, 2),
            'net': round(inflow - outflow, 2),
            'outstanding': round(out_info.get('outstanding', 0.0), 2),
            'estimated': round(out_info.get('estimated', 0.0), 2),
        })

    rows.sort(key=lambda x: x['net'], reverse=True)
    total_in = round(sum(r['inflow'] for r in rows), 2)
    total_out = round(sum(r['outflow'] for r in rows), 2)
    return ok({
        'rows': rows,
        'year': year,
        'group_by': group_by,
        'group_label': dim['label'],
        'date_start': str(date_start),
        'date_end': str(date_end),
        'summary': {
            'inflow': total_in,
            'outflow': total_out,
            'net': round(total_in - total_out, 2),
            'outstanding': round(sum(r['outstanding'] for r in rows), 2),
            'count': len(rows),
        },
    })


@csrf_exempt
@pk_required()
def analytics_forecast(request):
    """判未来：①现金流入预测（应收账龄 + 历史回款滞后）②全年落地预测（YTD 年化 vs
    年度目标）③What-if 沙盘基线。入参：year(必填) dept(可选) horizon(默认6,3~12)。"""
    denied = _page_denied(request, 'ar_analytics')
    if denied:
        return denied
    if request.method != 'GET':
        return err('Method not allowed', 405)
    try:
        year = int(request.GET.get('year') or datetime.date.today().year)
    except (TypeError, ValueError):
        return err('年份无效')
    dept = (request.GET.get('dept') or '').strip()
    try:
        horizon = min(max(int(request.GET.get('horizon') or 6), 3), 12)
    except (TypeError, ValueError):
        horizon = 6
    today = datetime.date.today()

    base = _ar_dept_filter(ARRecord.objects.all(), request, shared_field='project__is_shared')
    if dept:
        base = base.filter(delivery_dept=dept)

    # ── 历史回款滞后：实付日 − 应收日 的中位天数（仅用真实回款样本）──────────────
    lag_samples = []
    for p in (ARPayment.objects.filter(source='回款', ar_record__in=base,
                                       ar_record__due_date__isnull=False, payment_date__isnull=False)
              .values('payment_date', 'ar_record__due_date')[:5000]):
        lag_samples.append((p['payment_date'] - p['ar_record__due_date']).days)
    lag_samples.sort()
    avg_lag = lag_samples[len(lag_samples) // 2] if lag_samples else 30
    avg_lag = max(0, min(avg_lag, 120))

    # ── 现金流入预测：未收记录按"应收日+滞后"的预计回款月归集 ────────────────────
    cur = datetime.date(today.year, today.month, 1)
    months, y, m = [], today.year, today.month
    for _ in range(horizon):
        months.append(f'{y:04d}-{m:02d}')
        m += 1
        if m > 12:
            m, y = 1, y + 1
    bucket = {k: 0.0 for k in months}
    opening = beyond = baddebt = 0.0
    for r in base.filter(outstanding_amount__gt=0, due_date__isnull=False).values('due_date', 'outstanding_amount'):
        amt = float(r['outstanding_amount'] or 0)
        opening += amt
        due = r['due_date']
        if due < today and (today - due).days > 90:
            baddebt += amt
        exp = due + datetime.timedelta(days=avg_lag)
        key = f'{exp.year:04d}-{exp.month:02d}'
        if exp < cur:
            bucket[months[0]] += amt        # 预计回款日已过 → 计入首月回收
        elif key in bucket:
            bucket[key] += amt
        else:
            beyond += amt                    # 超出预测窗口（远期）
    cum, cash_rows = 0.0, []
    for k in months:
        cum += bucket[k]
        cash_rows.append({'month': k, 'expected': round(bucket[k], 2), 'cumulative': round(cum, 2)})

    # ── 全年落地预测：YTD 年化 vs 年度目标（财务口径）────────────────────────────
    from caiwu import views as V
    if dept:
        fin_bus = [dept]
    elif request.pk_role == 'super_admin':
        fin_bus = list(DEPARTMENTS)
    else:
        fin_bus = list(request.pk_depts or [])
    actuals = V._collect_actuals(fin_bus, {year})
    mdata = [mm for mm in range(1, 13) if V._period_group(actuals, fin_bus, year, mm)[0] is not None]
    elapsed = len(mdata)
    ytd_rev = V._sum_opt([V._period_group(actuals, fin_bus, year, mm)[0] for mm in mdata]) or 0
    ytd_prof = V._sum_opt([V._period_group(actuals, fin_bus, year, mm)[1] for mm in mdata]) or 0
    ytd_gross = V._sum_opt([V._period_group(actuals, fin_bus, year, mm)[2] for mm in mdata]) or 0
    tgt = V._load_target_index(fin_bus, year)
    ann_t = {k: sum(tgt.get((b, 0), {}).get(k, 0) for b in fin_bus)
             for k in ('target_revenue', 'target_profit', 'target_gross_profit')}

    def _pace(ytd, target):
        proj = (ytd / elapsed * 12) if elapsed else 0
        return {'ytd': round(ytd, 2), 'projected': round(proj, 2), 'target': round(target, 2),
                'rate': round(proj / target * 100, 1) if target else None,
                'gap': round(proj - target, 2)}
    landing = {'elapsed': elapsed,
               'revenue': _pace(ytd_rev, ann_t['target_revenue']),
               'gross': _pace(ytd_gross, ann_t['target_gross_profit']),
               'profit': _pace(ytd_prof, ann_t['target_profit'])}

    proj_rev = (ytd_rev / elapsed * 12) if elapsed else 0
    proj_prof = (ytd_prof / elapsed * 12) if elapsed else 0
    baseline = {
        'proj_revenue': round(proj_rev, 2),
        'gross_margin': round(ytd_gross / ytd_rev * 100, 2) if ytd_rev else None,
        'net_margin': round(ytd_prof / ytd_rev * 100, 2) if ytd_rev else None,
        'outstanding': round(opening, 2), 'avg_lag_days': avg_lag,
        'window_total': round(cum, 2), 'horizon': horizon,
    }
    summary = {
        'proj_profit': round(proj_prof, 2), 'profit_target': round(ann_t['target_profit'], 2),
        'profit_gap': round(proj_prof - ann_t['target_profit'], 2),
        'profit_rate': landing['profit']['rate'],
        'revenue_rate': landing['revenue']['rate'],
        'cash_next3': round(sum(x['expected'] for x in cash_rows[:3]), 2),
        'opening_outstanding': round(opening, 2), 'baddebt_risk': round(baddebt, 2),
        'avg_lag_days': avg_lag,
    }
    return ok({'year': year, 'dept': dept or None, 'horizon': horizon, 'avg_lag_days': avg_lag,
               'cash': {'opening_outstanding': round(opening, 2), 'months': cash_rows,
                        'expected_window_total': round(cum, 2), 'beyond_window': round(beyond, 2),
                        'baddebt_risk': round(baddebt, 2)},
               'landing': landing, 'baseline': baseline, 'summary': summary})


@csrf_exempt
@pk_required()
def analytics_by_dept(request):
    """事业部应收全景：各事业部 上账/开票/已收/未收/逾期/回款率/本月到期目标。

    口径：开放 AR 快照（不按年筛，反映当前应收健康）。本月到期目标 = 当期(本月到期
    且未结清) + 逾期(已过期且未结清) 的上账金额，对应"该月各事业部当期+逾期上账"。
    全集团口径下返回全部可访问事业部（含零值），逐一体现 6 事业部维度。
    """
    denied = _page_denied(request, 'ar_analytics')
    if denied:
        return denied
    today = datetime.date.today()
    month_start = today.replace(day=1)
    month_end = today.replace(day=calendar.monthrange(today.year, today.month)[1])

    base = _ar_dept_filter(ARRecord.objects.all(), request, shared_field='project__is_shared')
    dept = (request.GET.get('dept') or '').strip()
    if dept:
        base = base.filter(delivery_dept=dept)

    # 展示用事业部清单（即使无数据也列出，体现全部事业部维度）
    if dept:
        dept_list = [dept]
    elif request.pk_role == 'super_admin':
        dept_list = list(DEPARTMENTS)
    else:
        dept_list = list(request.pk_depts or [])

    def _bydept(qs, **ann):
        return {r['delivery_dept']: r for r in qs.values('delivery_dept').annotate(**ann)}

    # collected 单独聚合：与记录级 Sum 同窗会因 payments JOIN 行扇出乘倍记录级金额
    overall = _bydept(base, estimated=Sum('estimated_amount'), invoiced=Sum('actual_invoice_amount'),
                      outstanding=Sum('outstanding_amount'),
                      cnt=Count('id', distinct=True))
    collected_by_dept = _bydept(base, collected=Sum('payments__amount'))
    for d, row in overall.items():
        row['collected'] = (collected_by_dept.get(d) or {}).get('collected') or 0
    overdue = _bydept(base.filter(outstanding_amount__gt=0, due_date__lt=today),
                      od_amt=Sum('outstanding_amount'), od_est=Sum('estimated_amount'), od_cnt=Count('id'))
    current = _bydept(base.filter(outstanding_amount__gt=0, due_date__gte=month_start, due_date__lte=month_end),
                      cur_est=Sum('estimated_amount'), cur_cnt=Count('id'))

    rows = []
    for d in dept_list:
        o = overall.get(d, {})
        od = overdue.get(d, {})
        cu = current.get(d, {})
        est = float(o.get('estimated') or 0)
        collected = float(o.get('collected') or 0)
        od_est = float(od.get('od_est') or 0)
        cur_est = float(cu.get('cur_est') or 0)
        rate = round(min(collected / est * 100, 999.99), 1) if est > 0 else None
        rows.append({
            'dept': d,
            'estimated': round(est, 2),
            'invoiced': round(float(o.get('invoiced') or 0), 2),
            'outstanding': round(float(o.get('outstanding') or 0), 2),
            'collected': round(collected, 2),
            'overdue_amount': round(float(od.get('od_amt') or 0), 2),
            'overdue_count': int(od.get('od_cnt') or 0),
            'record_count': int(o.get('cnt') or 0),
            'rate': rate,
            'current_due_est': round(cur_est, 2),
            'overdue_est': round(od_est, 2),
            'month_target': round(cur_est + od_est, 2),   # 本月到期目标 = 当期 + 逾期 上账
        })
    rows.sort(key=lambda r: r['outstanding'], reverse=True)

    totals = {
        'estimated': round(sum(r['estimated'] for r in rows), 2),
        'invoiced': round(sum(r['invoiced'] for r in rows), 2),
        'outstanding': round(sum(r['outstanding'] for r in rows), 2),
        'collected': round(sum(r['collected'] for r in rows), 2),
        'overdue_amount': round(sum(r['overdue_amount'] for r in rows), 2),
        'overdue_count': sum(r['overdue_count'] for r in rows),
        'month_target': round(sum(r['month_target'] for r in rows), 2),
    }
    totals['rate'] = round(min(totals['collected'] / totals['estimated'] * 100, 999.99), 1) if totals['estimated'] > 0 else None
    return ok({'ref_month': f'{today.year}年{today.month}月', 'rows': rows, 'totals': totals})


# ══════════════════════════════════════════════════════════════════════════════
# 预收预付 (Advance receipts / prepayments) — 单表 + direction 判别，挂项目台账
# ══════════════════════════════════════════════════════════════════════════════

def _advance_dept_filter(qs, request):
    """Dept-scope advances. Project-linked rows use project.is_shared for the
    shared-only rule; standalone rows are matched on delivery_dept only."""
    return _ar_dept_filter(qs, request, shared_field='project__is_shared')


def _apply_advance_filters(qs, request):
    """Shared dimension filters for advance list + kpi + summary."""
    direction = request.GET.get('direction', '').strip()
    if direction in ADVANCE_DIRECTIONS:
        qs = qs.filter(direction=direction)
    project_id = request.GET.get('project_id', '').strip()
    if project_id:
        qs = qs.filter(project_id=int(project_id))
    dept = request.GET.get('dept', '').strip()
    if dept:
        qs = qs.filter(delivery_dept=dept)
    year = request.GET.get('year', '').strip()
    if year:
        qs = qs.filter(occur_year=int(year))
    month = request.GET.get('month', '').strip()
    if month:
        qs = qs.filter(occur_month=int(month))
    counterparty = request.GET.get('counterparty', '').strip()
    if counterparty:
        qs = qs.filter(counterparty__icontains=counterparty)
    status = request.GET.get('writeoff_status', '').strip()
    if status == '未核销':
        qs = qs.filter(balance_amount__gt=0, written_off_amount__lte=0)
    elif status == '部分核销':
        qs = qs.filter(balance_amount__gt=0, written_off_amount__gt=0)
    elif status == '已核销':
        qs = qs.filter(balance_amount__lte=0)
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(Q(counterparty__icontains=q) |
                       Q(project__short_name__icontains=q) |
                       Q(project__project_no__icontains=q) |
                       Q(notes__icontains=q))
    return qs


# ══════════════════════════════════════════════════════════════════════════════
# Supplier pool (供应商池)
# ══════════════════════════════════════════════════════════════════════════════

def _supplier_prepaid_agg(s):
    """Return prepaid balance/count for a supplier (one DB query)."""
    qs = AdvanceRecord.objects.filter(
        direction='预付', balance_amount__gt=0, counterparty__iexact=s.name)
    if s.project_id:
        qs = qs.filter(project_id=s.project_id)
    elif s.delivery_dept:
        qs = qs.filter(delivery_dept=s.delivery_dept)
    agg = qs.aggregate(total=Sum('balance_amount'), cnt=Count('id'))
    return str(agg['total'] or Decimal('0')), agg['cnt'] or 0


@csrf_exempt
@pk_required()
def suppliers(request):
    """GET = list, POST = create."""
    denied = _page_denied(request, 'ar_advance')
    if denied:
        return denied

    if request.method == 'GET':
        qs = Supplier.objects.select_related('project', 'created_by')
        qs = _ar_dept_filter(qs, request)
        stype = request.GET.get('type', '').strip()
        if stype in ('private', 'public'):
            qs = qs.filter(supplier_type=stype)
        project_id = request.GET.get('project_id', '').strip()
        if project_id:
            qs = qs.filter(project_id=int(project_id))
        q = request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(contact__icontains=q) | Q(notes__icontains=q))
        items = []
        for s in qs:
            d = s.to_dict()
            d['prepaid_balance'], d['prepaid_count'] = _supplier_prepaid_agg(s)
            items.append(d)
        return ok({'items': items, 'total': len(items)})

    if request.method == 'POST':
        denied = _write_denied(request)
        if denied:
            return denied
        data = _parse_body(request)
        name = (data.get('name') or '').strip()
        if not name:
            return err('供应商名称不能为空')
        stype = data.get('supplier_type', 'public')
        if stype not in ('private', 'public'):
            return err('类型无效')
        project = None
        delivery_dept = (data.get('delivery_dept') or '').strip()
        if stype == 'private':
            pid = data.get('project_id')
            if not pid:
                return err('私有供应商必须关联项目')
            try:
                project = ARProject.objects.get(pk=int(pid))
            except (ARProject.DoesNotExist, ValueError, TypeError):
                return err('项目不存在', 404)
            dep = _dept_denied(request, project.delivery_dept)
            if dep:
                return dep
            delivery_dept = project.delivery_dept
        else:
            if not delivery_dept:
                return err('公共供应商必须指定交付部门')
            if delivery_dept not in VALID_DEPARTMENTS:
                return err('无效部门')
            dep = _dept_denied(request, delivery_dept)
            if dep:
                return dep
        s = Supplier.objects.create(
            name=name, supplier_type=stype, project=project,
            delivery_dept=delivery_dept,
            contact=(data.get('contact') or '').strip(),
            notes=(data.get('notes') or '').strip(),
            created_by_id=request.pk_uid)
        return ok(s.to_dict())

    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def supplier_detail(request, pk):
    denied = _page_denied(request, 'ar_advance')
    if denied:
        return denied
    try:
        s = Supplier.objects.select_related('project', 'created_by').get(pk=pk)
    except Supplier.DoesNotExist:
        return err('供应商不存在', 404)
    if request.pk_role != 'super_admin' and s.delivery_dept not in request.pk_depts:
        return err('无权访问', 403)

    if request.method == 'GET':
        d = s.to_dict()
        d['prepaid_balance'], d['prepaid_count'] = _supplier_prepaid_agg(s)
        adv_qs = AdvanceRecord.objects.filter(
            direction='预付', balance_amount__gt=0, counterparty__iexact=s.name)
        if s.project_id:
            adv_qs = adv_qs.filter(project_id=s.project_id)
        elif s.delivery_dept:
            adv_qs = adv_qs.filter(delivery_dept=s.delivery_dept)
        d['prepaid_advances'] = [a.to_dict() for a in adv_qs.order_by('-occur_date')]
        return ok(d)

    if request.method == 'PUT':
        denied = _write_denied(request)
        if denied:
            return denied
        data = _parse_body(request)
        if 'name' in data:
            v = (data['name'] or '').strip()
            if not v:
                return err('供应商名称不能为空')
            s.name = v
        if 'contact' in data:
            s.contact = (data['contact'] or '').strip()
        if 'notes' in data:
            s.notes = (data['notes'] or '').strip()
        if s.supplier_type == 'public' and 'delivery_dept' in data:
            nd = (data['delivery_dept'] or '').strip()
            if nd and nd in VALID_DEPARTMENTS:
                dep = _dept_denied(request, nd)
                if dep:
                    return dep
                s.delivery_dept = nd
        s.save()
        return ok(s.to_dict())

    if request.method == 'DELETE':
        denied = _delete_denied(request)
        if denied:
            return denied
        s.delete()
        return ok({'deleted': pk})

    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def supplier_search(request):
    """Fuzzy-search suppliers by payee name for PaymentModal matching.
    Returns matched suppliers with prepaid balance and open advances (up to 5 each).
    """
    if request.method != 'GET':
        return err('Method not allowed', 405)
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return ok({'items': []})
    dept = request.GET.get('dept', '').strip()
    qs = Supplier.objects.select_related('project')
    if request.pk_role != 'super_admin':
        allowed = list(request.pk_depts or [])
        if not allowed:
            return ok({'items': []})
        if dept and dept in allowed:
            qs = qs.filter(delivery_dept=dept)
        else:
            qs = qs.filter(delivery_dept__in=allowed)
    elif dept:
        qs = qs.filter(delivery_dept=dept)
    qs = qs.filter(name__icontains=q)[:10]
    items = []
    for s in qs:
        d = s.to_dict()
        adv_qs = AdvanceRecord.objects.filter(
            direction='预付', balance_amount__gt=0, counterparty__iexact=s.name)
        if s.project_id:
            adv_qs = adv_qs.filter(project_id=s.project_id)
        elif s.delivery_dept:
            adv_qs = adv_qs.filter(delivery_dept=s.delivery_dept)
        agg = adv_qs.aggregate(total=Sum('balance_amount'), cnt=Count('id'))
        d['prepaid_balance'] = str(agg['total'] or Decimal('0'))
        d['prepaid_count'] = agg['cnt'] or 0
        d['prepaid_advances'] = [
            {'id': a.id, 'advance_amount': str(a.advance_amount),
             'balance_amount': str(a.balance_amount),
             'occur_date': str(a.occur_date) if a.occur_date else None,
             'notes': a.notes}
            for a in adv_qs.order_by('-occur_date')[:5]
        ]
        items.append(d)
    return ok({'items': items})


@csrf_exempt
@pk_required()
def advances(request):
    denied = _page_denied(request, 'ar_advance')
    if denied:
        return denied

    if request.method == 'GET':
        today = datetime.date.today()
        qs = _advance_dept_filter(
            AdvanceRecord.objects.select_related('project', 'created_by'), request)
        qs = _apply_advance_filters(qs, request)
        page = max(1, int(request.GET.get('page', 1) or 1))
        size = min(200, max(1, int(request.GET.get('size', 50) or 50)))

        qs_agg = _apply_advance_filters(
            _advance_dept_filter(AdvanceRecord.objects.all(), request), request)
        ids = list(qs_agg.order_by().values_list('id', flat=True))
        total = len(ids)
        base = AdvanceRecord.objects.filter(id__in=ids)

        def _dir_summary(direction):
            d_qs = base.filter(direction=direction)
            agg = d_qs.aggregate(
                amt=Sum('advance_amount'),
                wo=Sum('written_off_amount'),
                bal=Sum('balance_amount', filter=Q(balance_amount__gt=0)),
            )
            overdue = (d_qs.filter(balance_amount__gt=0,
                                   expected_writeoff_date__lt=today)
                       .aggregate(s=Sum('balance_amount'))['s'] or 0)
            return {
                'count': d_qs.count(),
                'advance_amount': str(agg['amt'] or 0),
                'written_off': str(agg['wo'] or 0),
                'balance': str(agg['bal'] or 0),
                'overdue_balance': str(overdue),
            }

        summary = {
            'count': total,
            '预收': _dir_summary('预收'),
            '预付': _dir_summary('预付'),
        }

        perms = get_request_perms(request)
        include_wo = request.GET.get('include_writeoffs', '') in ('1', 'true')
        items = list(qs[(page - 1) * size: page * size])
        rows = [apply_ar_view_mask(r.to_dict(today=today, include_writeoffs=include_wo),
                                   perms, 'advance') for r in items]
        return ok({'items': rows, 'total': total, 'page': page, 'size': size,
                   'summary': summary})

    if request.method == 'POST':
        denied = _write_denied(request)
        if denied:
            return denied
        data = _ar_visible_payload(
            request, _parse_body(request), 'advance',
            extra=('project_id', 'direction', 'occur_year', 'occur_month',
                   'occur_date', 'expected_writeoff_date', 'delivery_dept'))
        return _advance_create(request, data)

    return err('Method not allowed', 405)


def _advance_create(request, data):
    direction = (data.get('direction') or '').strip()
    if direction not in ADVANCE_DIRECTIONS:
        return err('方向无效，应为 预收 或 预付')
    year = int(data.get('occur_year', 0) or 0)
    month = int(data.get('occur_month', 0) or 0)
    if not (year and 1 <= month <= 12):
        return err('发生年月无效')

    proj = None
    project_id = data.get('project_id')
    dept = (data.get('delivery_dept') or '').strip()
    if project_id:
        try:
            proj = ARProject.objects.get(pk=int(project_id))
        except ARProject.DoesNotExist:
            return err('项目不存在', 404)
        dept = proj.delivery_dept
    if not dept:
        return err('请选择项目或填写交付部门')
    denied = _dept_denied(request, dept, '无权操作此部门')
    if denied:
        return denied

    from paikuan.models import PaikuanUser
    user = PaikuanUser.objects.filter(id=request.pk_uid).first()
    try:
        rec = AdvanceRecord(
            project=proj,
            delivery_dept=dept,
            direction=direction,
            counterparty=(data.get('counterparty') or '').strip(),
            occur_year=year,
            occur_month=month,
            occur_date=_normalize_date(data.get('occur_date')) or None,
            advance_amount=_dec(data.get('advance_amount', 0)),
            expected_writeoff_date=_normalize_date(data.get('expected_writeoff_date')) or None,
            notes=(data.get('notes') or '').strip(),
            created_by=user,
        )
        rec.save()
        # 总额为派生列（=收付明细之和）：创建时生成首笔收付明细，后续可多次追加
        if rec.advance_amount:
            init_date = rec.occur_date or datetime.date(year, month, 1)
            AdvanceInstallment.objects.create(
                advance_record=rec, install_no=1, amount=rec.advance_amount,
                occur_date=init_date, notes='录入初始金额')
    except Exception as e:
        return err(str(e))
    return ok(apply_ar_view_mask(
        rec.to_dict(today=datetime.date.today(), include_writeoffs=True),
        get_request_perms(request), 'advance'))


@csrf_exempt
@pk_required()
def advance_detail(request, pk):
    denied = _page_denied(request, 'ar_advance')
    if denied:
        return denied
    try:
        rec = AdvanceRecord.objects.select_related('project', 'created_by').get(pk=pk)
    except AdvanceRecord.DoesNotExist:
        return err('记录不存在', 404)
    if request.pk_role != 'super_admin':
        if rec.delivery_dept not in request.pk_depts:
            return err('无权访问', 403)
        perms = get_request_perms(request)
        if perms and perms.get('ar_shared_only') and not (rec.project and rec.project.is_shared):
            return err('无权访问', 403)
    today = datetime.date.today()

    if request.method == 'GET':
        return ok(apply_ar_view_mask(rec.to_dict(today=today, include_writeoffs=True),
                                     get_request_perms(request), 'advance'))

    if request.method == 'PUT':
        denied = _write_denied(request)
        if denied:
            return denied
        data = _ar_visible_payload(request, _parse_body(request), 'advance',
                                   extra=('direction', 'occur_year', 'occur_month',
                                          'occur_date', 'expected_writeoff_date'))
        if 'direction' in data and data['direction'] in ADVANCE_DIRECTIONS \
                and data['direction'] != rec.direction:
            # 方向决定核销语义（预收核销挂应收回款 / 预付核销挂排款）与资金池口径，
            # 已有核销时翻转会让两侧台账互相矛盾
            if rec.writeoffs.exists():
                return err('该记录已有核销，不能修改预收/预付方向；请先删除其全部核销记录')
            rec.direction = data['direction']
        if 'counterparty' in data:
            rec.counterparty = (data['counterparty'] or '').strip()
        if 'occur_year' in data and data['occur_year']:
            rec.occur_year = int(data['occur_year'])
        if 'occur_month' in data and data['occur_month']:
            rec.occur_month = int(data['occur_month'])
        if 'occur_date' in data:
            rec.occur_date = _normalize_date(data['occur_date']) or None
        if 'advance_amount' in data:
            # 兼容旧调用方按「总额」编辑：与现总额的差值生成一笔收付明细
            # （明细为正源、总额为派生，与应收差额调整同款策略）。
            # 减额低于已核销时信号抛 ValidationError → 400（余额不能为负）
            new_total = _dec(data['advance_amount'])
            delta = new_total - (rec.advance_amount or Decimal('0'))
            if delta:
                last_inst = rec.installments.order_by('-install_no').first()
                try:
                    with transaction.atomic():
                        AdvanceInstallment.objects.create(
                            advance_record=rec,
                            install_no=(last_inst.install_no + 1) if last_inst else 1,
                            amount=delta, occur_date=rec.occur_date or datetime.date.today(),
                            notes='人工调整（按总额修改）')
                except ValidationError as e:
                    return err(str(e.message if hasattr(e, 'message') else e), 400)
                rec.advance_amount = new_total
        if 'expected_writeoff_date' in data:
            rec.expected_writeoff_date = _normalize_date(data['expected_writeoff_date']) or None
        if 'notes' in data:
            rec.notes = (data['notes'] or '').strip()
        try:
            rec.save()
        except Exception as e:
            return err(str(e))
        return ok(apply_ar_view_mask(rec.to_dict(today=today, include_writeoffs=True),
                                     get_request_perms(request), 'advance'))

    if request.method == 'DELETE':
        denied = _delete_denied(request)
        if denied:
            return denied
        rec.delete()
        return ok({'deleted': pk})

    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def advances_kpi(request):
    denied = _page_denied(request, 'ar_advance')
    if denied:
        return denied
    today = datetime.date.today()
    qs = _apply_advance_filters(
        _advance_dept_filter(AdvanceRecord.objects.all(), request), request)

    def _block(direction):
        d_qs = qs.filter(direction=direction)
        agg = d_qs.aggregate(amt=Sum('advance_amount'), wo=Sum('written_off_amount'),
                             bal=Sum('balance_amount', filter=Q(balance_amount__gt=0)))
        total_amt = float(agg['amt'] or 0)
        wo = float(agg['wo'] or 0)
        bal = float(agg['bal'] or 0)
        pending = d_qs.filter(balance_amount__gt=0).count()
        overdue_qs = d_qs.filter(balance_amount__gt=0, expected_writeoff_date__lt=today)
        overdue_amt = float(overdue_qs.aggregate(s=Sum('balance_amount'))['s'] or 0)
        return {
            'count': d_qs.count(),
            'advance_amount': total_amt,
            'written_off': wo,
            'balance': bal,
            'writeoff_rate': round(wo / total_amt * 100, 1) if total_amt else 100.0,
            'pending_count': pending,
            'overdue_count': overdue_qs.count(),
            'overdue_balance': overdue_amt,
        }

    return ok({'预收': _block('预收'), '预付': _block('预付')})


@csrf_exempt
@pk_required()
def advances_summary(request):
    """Group-by pivot over advances. group_by ∈ {dept, direction, counterparty, month}."""
    denied = _page_denied(request, 'ar_advance')
    if denied:
        return denied
    qs = _apply_advance_filters(
        _advance_dept_filter(AdvanceRecord.objects.all(), request), request)
    group_by = request.GET.get('group_by', 'dept').strip()
    field_map = {
        'dept': 'delivery_dept',
        'direction': 'direction',
        'counterparty': 'counterparty',
    }
    rows = []
    if group_by == 'month':
        agg = (qs.values('occur_year', 'occur_month')
               .annotate(count=Count('id'), advance_amount=Sum('advance_amount'),
                         written_off=Sum('written_off_amount'),
                         balance=Sum('balance_amount'))
               .order_by('occur_year', 'occur_month'))
        for r in agg:
            rows.append({
                'key': f"{r['occur_year']}-{r['occur_month']:02d}",
                'year': r['occur_year'], 'month': r['occur_month'],
                'count': r['count'],
                'advance_amount': str(r['advance_amount'] or 0),
                'written_off': str(r['written_off'] or 0),
                'balance': str(r['balance'] or 0),
            })
    else:
        field = field_map.get(group_by, 'delivery_dept')
        agg = (qs.values(field)
               .annotate(count=Count('id'), advance_amount=Sum('advance_amount'),
                         written_off=Sum('written_off_amount'),
                         balance=Sum('balance_amount'))
               .order_by('-advance_amount'))
        for r in agg:
            rows.append({
                'key': r[field] or '(未填)',
                'count': r['count'],
                'advance_amount': str(r['advance_amount'] or 0),
                'written_off': str(r['written_off'] or 0),
                'balance': str(r['balance'] or 0),
            })
    return ok({'group_by': group_by, 'rows': rows})


@csrf_exempt
@pk_required()
def advances_available(request):
    """可用预收/预付查询 — 供应收回款 / 排款界面联动弹出。

    返回指定方向(默认预收)下仍有未核销余额(balance_amount > 0)的明细及合计。

    匹配口径（可组合，按 OR 取并集）：
    - project_id：挂在该项目下的预收/预付；
    - customer：未挂项目、且往来单位 = 该客户名（忽略大小写/首尾空格的精确匹配）
      的「散单」预收/预付，用于"未挂项目、只记了客户名"的预收也能在回款界面弹出。
      采用精确匹配以保证与「应收明细→可冲抵预收」双向一致、口径严谨。
    另有 dept / counterparty 作为附加的 AND 过滤（部门精确、往来单位包含）。
    只读，不改动任何账务。
    """
    denied = _page_denied(request, 'ar_advance')
    if denied:
        return denied
    denied = _ar_field_denied(request, 'adv_amount')
    if denied:
        return denied
    if request.method != 'GET':
        return err('Method not allowed', 405)
    direction = (request.GET.get('direction', '预收').strip() or '预收')
    if direction not in ADVANCE_DIRECTIONS:
        return err('方向无效，应为 预收 或 预付')
    today = datetime.date.today()
    qs = _advance_dept_filter(
        AdvanceRecord.objects.select_related('project'), request)
    qs = qs.filter(direction=direction, balance_amount__gt=0)

    # 并集匹配：本项目 ∪ 散单(无项目)且客户名匹配
    match_clauses = []
    project_id = request.GET.get('project_id', '').strip()
    if project_id:
        match_clauses.append(Q(project_id=int(project_id)))
    customer = request.GET.get('customer', '').strip()
    if customer:
        match_clauses.append(Q(project__isnull=True, counterparty__iexact=customer))
    if match_clauses:
        combined = match_clauses[0]
        for c in match_clauses[1:]:
            combined |= c
        qs = qs.filter(combined)

    dept = request.GET.get('dept', '').strip()
    if dept:
        qs = qs.filter(delivery_dept=dept)
    counterparty = request.GET.get('counterparty', '').strip()
    if counterparty:
        qs = qs.filter(counterparty__icontains=counterparty)
    agg = qs.aggregate(total=Sum('balance_amount'), cnt=Count('id'))
    total_balance = (agg['total'] or Decimal('0')).quantize(Decimal('0.01'))
    items = [{
        'id': r.id,
        'project_id': r.project_id,
        'project_no': r.project.project_no if r.project_id else None,
        'short_name': r.project.short_name if r.project_id else None,
        'match_type': 'project' if r.project_id else 'customer',
        'counterparty': r.counterparty,
        'delivery_dept': r.delivery_dept,
        'occur_date': str(r.occur_date) if r.occur_date else None,
        'advance_amount': str(r.advance_amount),
        'balance_amount': str(r.balance_amount),
        'expected_writeoff_date': (str(r.expected_writeoff_date)
                                   if r.expected_writeoff_date else None),
        **r.aging_dict(today),
    } for r in qs.order_by('-balance_amount')[:50]]
    return ok({
        'direction': direction,
        'count': agg['cnt'] or 0,
        'total_balance': str(total_balance),
        'items': items,
    })


@csrf_exempt
@pk_required()
def advance_offsettable_records(request):
    """列出可被预收冲抵的应收明细（未收余额 > 0），供预收核销弹窗选择。

    入参二选一：
    - project_id：该项目下的应收明细（预收挂了项目时）；
    - customer：按项目客户名匹配的应收明细（散单预收，未挂项目时）。
    """
    denied = _page_denied(request, 'ar_advance')
    if denied:
        return denied
    if request.method != 'GET':
        return err('Method not allowed', 405)
    project_id = request.GET.get('project_id', '').strip()
    customer = request.GET.get('customer', '').strip()
    qs = ARRecord.objects.select_related('project').filter(outstanding_amount__gt=0)
    if project_id:
        qs = qs.filter(project_id=int(project_id))
    elif customer:
        # 与 advances_available 对称：散单预收的往来单位 == 项目合同名（即客户名，精确匹配）
        qs = qs.filter(project__customer_name__iexact=customer)
    else:
        return err('缺少 project_id 或 customer')
    if request.pk_role != 'super_admin':
        qs = qs.filter(delivery_dept__in=request.pk_depts)
    qs = qs.order_by('project__short_name', 'operation_year', 'operation_month')
    by_customer = bool(customer) and not project_id
    items = [{
        'id': r.id,
        'project_id': r.project_id,
        'short_name': r.project.short_name,
        'operation_year': r.operation_year,
        'operation_month': r.operation_month,
        'estimated_amount': str(r.estimated_amount),
        'outstanding_amount': str(r.outstanding_amount),
        'label': ((f'{r.project.short_name} · ' if by_customer else '')
                  + f'{r.operation_year}-{r.operation_month:02d} · 未收 '
                  f'{r.outstanding_amount:,.2f}'),
    } for r in qs[:100]]
    return ok({'items': items})


@csrf_exempt
@pk_required()
def advance_template(request):
    denied = _page_denied(request, 'ar_advance')
    if denied:
        return denied
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '预收预付明细'
    headers = ['方向(预收/预付)*', '项目简称', '交付部门', '往来单位*', '发生年*', '发生月*',
               '款项日期', '预收/预付金额', '预计核销日期', '核销金额', '核销日期', '备注']
    _header_row(ws, headers, color='6A1B9A')
    tip_vals = [
        '★必填：预收=客户预付给我们(现金流入)，预付=我们预付供应商(现金流出)',
        '选填：填项目台账中的"项目简称"（精确匹配）；无对应项目可留空，仅填往来单位',
        '选填：未填项目简称时必填本列（交付部门）；填了项目则自动取项目部门',
        '★必填：往来单位（预收填客户、预付填供应商）',
        '★必填：4位整数，如 2026',
        '★必填：1-12 的整数',
        '选填：款项实际收/付日期（驱动现金流），格式 2026-01-15 / 2026年1月15日 均可',
        '选填：本笔预收/预付金额（元）；未核销余额 = 金额 − 累计核销',
        '选填：预计核销日期，超期未核销将预警；格式同上',
        '选填：本次核销(冲减)金额（元）；同一笔可多次核销，多次导入追加不覆盖',
        '选填：核销日期，格式同上',
        '选填备注',
    ]
    ws.append(tip_vals)
    tip_row = ws.max_row
    tip_fill = PatternFill('solid', fgColor='F3E5F5')
    tip_font = Font(italic=True, color='4A148C', size=9)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(tip_row, c)
        cell.fill = tip_fill
        cell.font = tip_font
        cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[tip_row].height = 60
    ws.append(['预收', EXAMPLE_ROW_MARKER, '', '示例客户', 2026, 1, '2026-01-15',
               100000, '2026-06-30', 30000, '2026-03-10',
               '示例（此行含"示例"标记，导入时自动跳过）'])
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    return _export_response(wb, '预收预付导入模板.xlsx')


@csrf_exempt
@pk_required()
def advance_import(request):
    denied = _page_denied(request, 'ar_advance')
    if denied:
        return denied
    if request.method != 'POST':
        return err('POST only', 405)
    denied = _write_denied(request)
    if denied:
        return denied
    f = request.FILES.get('file')
    if not f:
        return err('请上传文件')
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
    except Exception as e:
        return err(f'无法读取Excel: {e}')

    headers = [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(headers)}
    _ALIASES = {
        '方向(预收/预付)*': ['方向(预收/预付)*', '方向(预收/预付)', '方向'],
        '项目简称': ['项目简称', '项目简称*'],
        '交付部门': ['交付部门'],
        '往来单位*': ['往来单位*', '往来单位'],
        '发生年*': ['发生年*', '发生年'],
        '发生月*': ['发生月*', '发生月'],
        '款项日期': ['款项日期', '款项日期(YYYY-MM-DD)'],
        '预收/预付金额': ['预收/预付金额', '金额'],
        '预计核销日期': ['预计核销日期'],
        '核销金额': ['核销金额'],
        '核销日期': ['核销日期'],
        '备注': ['备注'],
    }

    def _resolve_idx(name):
        for h in _ALIASES.get(name, [name]):
            if h in col_map:
                return col_map[h]
        return None

    def _cv_raw(row, name):
        idx = _resolve_idx(name)
        return ws.cell(row, idx).value if idx else None

    def _cv(row, name):
        v = _cv_raw(row, name)
        return str(v).strip() if v is not None else ''

    from paikuan.models import PaikuanUser
    user = PaikuanUser.objects.filter(id=request.pk_uid).first()
    search_depts = None if request.pk_role == 'super_admin' else request.pk_depts
    errors = []
    plan = []   # 通过校验、待写入的行

    # ══ 阶段一：逐行校验（不写库）。示例/提示/空行静默忽略，问题行收集错误 ══════════
    for ri in range(2, ws.max_row + 1):
        direction = _cv(ri, '方向(预收/预付)*')
        short_name = _cv(ri, '项目简称')
        counterparty = _cv(ri, '往来单位*')
        if EXAMPLE_ROW_MARKER in short_name or direction.startswith('★'):
            continue
        if direction not in ADVANCE_DIRECTIONS:
            if not direction and not counterparty and not short_name:
                continue   # 整行空白：静默忽略
            errors.append(f'第{ri}行: 方向"{direction}"无效，应为 预收/预付')
            continue
        proj = None
        dept = _cv(ri, '交付部门')
        if short_name:
            proj = _match_project_by_short_name(short_name, allowed_depts=search_depts)
            if not proj:
                errors.append(f'第{ri}行: 项目简称"{short_name}"未匹配到项目，请核对简称或先在项目台账建项目')
                continue
            dept = proj.delivery_dept
        if not dept:
            errors.append(f'第{ri}行: 未填项目简称时必须填写交付部门')
            continue
        if request.pk_role != 'super_admin' and dept not in request.pk_depts:
            errors.append(f'第{ri}行: 无权操作部门"{dept}"')
            continue
        year = int(_cv(ri, '发生年*') or 0)
        month = int(_cv(ri, '发生月*') or 0)
        if not (year and 1 <= month <= 12):
            errors.append(f'第{ri}行: 发生年月无效（年={_cv(ri, "发生年*") or "空"} 月={_cv(ri, "发生月*") or "空"}），运作月需 1-12')
            continue
        wo_amount = _cv(ri, '核销金额')
        wo_date = _normalize_date(_cv_raw(ri, '核销日期'))
        if wo_amount and not wo_date:
            errors.append(f'第{ri}行: 填了「核销金额」却没填有效「核销日期」，请补填或清空核销金额')
            continue
        plan.append({'ri': ri, 'proj': proj, 'dept': dept, 'direction': direction,
                     'counterparty': counterparty, 'year': year, 'month': month,
                     'occur_date': _normalize_date(_cv_raw(ri, '款项日期')) or None,
                     'advance_amount': _dec(_cv(ri, '预收/预付金额') or 0),
                     'expected_writeoff_date': _normalize_date(_cv_raw(ri, '预计核销日期')) or None,
                     'notes': _cv(ri, '备注'), 'wo_amount': wo_amount, 'wo_date': wo_date})

    if errors:
        return ok({
            'rejected': True, 'created': 0, 'updated': 0, 'errors': errors,
            'message': (f'导入未执行：发现 {len(errors)} 处问题，已全部列出。'
                        f'请在表格中按提示修正后重新导入（整表全部通过才会写入，不会漏导）。'),
        })

    # ══ 阶段二：全部通过 → 一次性写入（整体事务，任一失败回滚）═══════════════════════
    created = 0
    try:
        with transaction.atomic():
            for p in plan:
                rec = AdvanceRecord(
                    project=p['proj'], delivery_dept=p['dept'], direction=p['direction'],
                    counterparty=p['counterparty'], occur_year=p['year'], occur_month=p['month'],
                    occur_date=p['occur_date'], advance_amount=p['advance_amount'],
                    expected_writeoff_date=p['expected_writeoff_date'],
                    notes=p['notes'], created_by=user,
                )
                rec.save()
                # 总额为派生列：导入金额作为首笔收付明细（须先于核销生效）
                if rec.advance_amount:
                    AdvanceInstallment.objects.create(
                        advance_record=rec, install_no=1, amount=rec.advance_amount,
                        occur_date=rec.occur_date or datetime.date(p['year'], p['month'], 1),
                        notes='导入初始金额')
                if p['wo_amount'] and p['wo_date']:
                    amt = _dec(p['wo_amount'])
                    if amt > 0 and not rec.writeoffs.filter(writeoff_date=p['wo_date'], amount=amt).exists():
                        max_no = rec.writeoffs.aggregate(m=Max('writeoff_no')).get('m') or 0
                        AdvanceWriteoff.objects.create(
                            advance_record=rec, writeoff_no=max_no + 1,
                            amount=amt, writeoff_date=p['wo_date'], notes='导入核销')
                created += 1
    except Exception as e:
        return ok({
            'rejected': True, 'created': 0, 'updated': 0,
            'errors': [f'写入阶段发生错误并已回滚：{e}。请检查数据后重试。'],
            'message': '导入未执行（写入阶段出错，已整体回滚，不会出现半截数据）。',
        })

    return ok({'created': created, 'updated': 0, 'skipped': 0, 'errors': []})


@csrf_exempt
@pk_required()
def advance_export(request):
    denied = _page_denied(request, 'ar_advance')
    if denied:
        return denied
    today = datetime.date.today()
    qs = _apply_advance_filters(
        _advance_dept_filter(AdvanceRecord.objects.select_related('project'), request), request)
    if qs.count() > 5000:
        return err('导出超过5000行，请缩小筛选范围')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '预收预付明细'
    columns = _visible_ar_export_cols(request, [
        (None, '方向', lambda rec, st: rec.direction),
        (None, '项目编号', lambda rec, st: rec.project.project_no if rec.project_id else ''),
        ('p_short_name', '项目简称', lambda rec, st: rec.project.short_name if rec.project_id else ''),
        (None, '交付部门', lambda rec, st: rec.delivery_dept),
        ('adv_counterparty', '往来单位', lambda rec, st: rec.counterparty),
        (None, '发生年', lambda rec, st: rec.occur_year),
        (None, '发生月', lambda rec, st: rec.occur_month),
        (None, '款项日期', lambda rec, st: str(rec.occur_date) if rec.occur_date else ''),
        ('adv_amount', '预收/预付金额', lambda rec, st: float(rec.advance_amount)),
        ('adv_amount', '收付明细',
         lambda rec, st: '；'.join(f'{i.occur_date}:{i.amount}'
                                   for i in rec.installments.all())),
        ('adv_writeoff', '已核销金额', lambda rec, st: float(rec.written_off_amount)),
        ('adv_writeoff', '未核销余额', lambda rec, st: float(rec.balance_amount)),
        ('adv_writeoff', '核销状态', lambda rec, st: rec.writeoff_status),
        ('adv_expected_date', '预计核销日期', lambda rec, st: str(rec.expected_writeoff_date) if rec.expected_writeoff_date else ''),
        ('adv_expected_date', '挂账天数', lambda rec, st: st['pending_days']),
        ('adv_expected_date', '是否逾期', lambda rec, st: '是' if st['is_overdue'] else ''),
        ('adv_notes', '备注', lambda rec, st: rec.notes),
    ])
    _header_row(ws, [h for _, h, _ in columns], color='6A1B9A')
    for rec in qs:
        st = rec.aging_dict(today)
        ws.append([getter(rec, st) for _, _, getter in columns])
    return _export_response(wb, '预收预付明细.xlsx')


@csrf_exempt
@pk_required()
def advance_writeoffs(request, pk):
    denied = _page_denied(request, 'ar_advance')
    if denied:
        return denied
    try:
        rec = AdvanceRecord.objects.select_related('project').get(pk=pk)
    except AdvanceRecord.DoesNotExist:
        return err('记录不存在', 404)
    if request.pk_role != 'super_admin' and rec.delivery_dept not in request.pk_depts:
        return err('无权访问', 403)
    denied = _ar_field_denied(request, 'adv_writeoff')
    if denied:
        return denied

    if request.method == 'GET':
        return ok([w.to_dict() for w in
                   rec.writeoffs.select_related('payment').order_by('writeoff_no')])

    if request.method == 'POST':
        denied = _action_denied(request,
                                'wo_receive' if rec.direction == '预收' else 'wo_prepaid')
        if denied:
            return denied
        data = _parse_body(request)
        amount = _dec(data.get('amount', 0))
        if amount <= 0:
            return err('核销金额必须大于0')
        wo_date = _normalize_date(data.get('writeoff_date'))
        if not wo_date:
            return err('核销日期无效')
        ar_rec, offset_err = _resolve_offset_ar_record(
            request, rec, data.get('ar_record_id'), amount)
        if offset_err:
            return offset_err
        # 预付核销可关联排款记录（payment_id 仅限 direction='预付'）
        payment_id = data.get('payment_id')
        payment_obj = None
        if payment_id:
            if rec.direction != '预付':
                return err('仅「预付」核销可关联排款记录')
            try:
                payment_obj = Payment.objects.get(pk=int(payment_id))
            except (Payment.DoesNotExist, ValueError, TypeError):
                return err('排款记录不存在', 404)
            if request.pk_role != 'super_admin' and payment_obj.department not in request.pk_depts:
                return err('无权操作该排款记录', 403)
            # 跨部门核销会让资金池错位：预付流出记在预付方部门，
            # 冲抵却从排款方部门扣——单池余额双向失真（集团合计反而看不出来）
            if payment_obj.department != rec.delivery_dept:
                return err(f'排款所属部门「{payment_obj.department}」与预付所属部门'
                           f'「{rec.delivery_dept}」不一致，不能关联核销')
            # 冲抵上限：累计冲抵 + 已付 不得超过计划金额（否则待付为负、口径失真）
            plan = (payment_obj.plan_adjustment
                    if payment_obj.plan_adjustment is not None else payment_obj.total_amount)
            paid = payment_obj.total_paid
            offset_now = payment_obj.prepaid_offset_amount or Decimal('0')
            room = (plan or Decimal('0')) - paid - offset_now
            if amount > room:
                return err(f'冲抵金额 {amount} 超过该排款剩余待付 {room}'
                           f'（计划 {plan} − 已付 {paid} − 已冲抵 {offset_now}）')
        try:
            with transaction.atomic():
                last = rec.writeoffs.select_for_update().order_by('-writeoff_no').first()
                next_no = (last.writeoff_no + 1) if last else 1
                wo = AdvanceWriteoff.objects.create(
                    advance_record=rec, writeoff_no=next_no, amount=amount,
                    writeoff_date=wo_date, notes=(data.get('notes') or '').strip(),
                    payment=payment_obj)
                if ar_rec is not None:
                    pay = _create_offset_payment(rec, ar_rec, amount, wo_date)
                    AdvanceWriteoff.objects.filter(pk=wo.pk).update(
                        ar_record=ar_rec, ar_payment=pay)
                    wo.ar_record = ar_rec
                    wo.ar_payment = pay
        except ValidationError as e:
            return err(str(e.message if hasattr(e, 'message') else e), 400)
        return ok(wo.to_dict())

    return err('Method not allowed', 405)


def _resolve_offset_ar_record(request, advance, ar_record_id, amount):
    """校验预收核销冲抵的应收明细。返回 (ARRecord|None, err_response|None)。"""
    if not ar_record_id:
        return None, None
    if advance.direction != '预收':
        return None, err('仅「预收」可冲抵应收账款（应收回款）')
    try:
        ar = ARRecord.objects.select_related('project').get(pk=int(ar_record_id))
    except (ARRecord.DoesNotExist, ValueError, TypeError):
        return None, err('所选应收明细不存在', 404)
    if request.pk_role != 'super_admin' and ar.delivery_dept not in request.pk_depts:
        return None, err('无权操作该应收明细所属部门', 403)
    if advance.project_id and ar.project_id != advance.project_id:
        return None, err('所选应收明细与预收所属项目不一致，无法冲抵')
    outstanding = ar.outstanding_amount or Decimal('0')
    if amount > outstanding:
        return None, err(f'冲抵金额 {amount:,.2f} 超过该应收未收余额 {outstanding:,.2f}')
    return ar, None


def _create_offset_payment(advance, ar_record, amount, pay_date):
    """为预收核销生成一笔「预收抵扣」回款，冲减应收 outstanding。"""
    last = ar_record.payments.select_for_update().order_by('-payment_no').first()
    next_no = (last.payment_no + 1) if last else 1
    note = f'预收核销冲抵 · 预收#{advance.id}'
    if advance.counterparty:
        note += f'（{advance.counterparty}）'
    return ARPayment.objects.create(
        ar_record=ar_record, payment_no=next_no, amount=amount,
        payment_date=pay_date, source='预收抵扣', notes=note)


def _adv_installments_payload(rec):
    """收付明细变更后的统一回包：明细 + 派生总额/余额（UI 一次刷新）。"""
    rec.refresh_from_db()
    return {
        'items': [i.to_dict() for i in rec.installments.order_by('install_no')],
        'advance_amount': str(rec.advance_amount or 0),
        'balance_amount': str(rec.balance_amount or 0),
    }


@csrf_exempt
@pk_required()
def advance_installments(request, pk):
    """GET/POST /advances/<pk>/installments — 预收/预付收付明细（多次到账/付出）。"""
    denied = _page_denied(request, 'ar_advance')
    if denied:
        return denied
    try:
        rec = AdvanceRecord.objects.select_related('project').get(pk=pk)
    except AdvanceRecord.DoesNotExist:
        return err('记录不存在', 404)
    if request.pk_role != 'super_admin' and rec.delivery_dept not in request.pk_depts:
        return err('无权访问', 403)
    denied = _ar_field_denied(request, 'adv_amount')
    if denied:
        return denied

    if request.method == 'GET':
        return ok(_adv_installments_payload(rec))

    if request.method == 'POST':
        denied = _action_denied(request, 'adv_installment')
        if denied:
            return denied
        data = _parse_body(request)
        amount = _dec(data.get('amount', 0))
        if not amount:
            return err('收付金额不能为0（可正可负，负数=退回）')
        occur = _normalize_date(data.get('occur_date'))
        if not occur:
            return err('收付日期无效（格式 2026-01-20）')
        try:
            with transaction.atomic():
                last = rec.installments.select_for_update().order_by('-install_no').first()
                AdvanceInstallment.objects.create(
                    advance_record=rec,
                    install_no=(last.install_no + 1) if last else 1,
                    amount=amount, occur_date=occur,
                    notes=(data.get('notes') or '').strip())
        except ValidationError as e:
            return err(str(e.message if hasattr(e, 'message') else e), 400)
        return ok(_adv_installments_payload(rec))

    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def advance_installment_detail(request, pk, iid):
    """DELETE /advances/<pk>/installments/<iid> — 删除一笔收付（总额随之回退；
    删除会使总额低于已核销时拒绝，须先删核销）。"""
    denied = _page_denied(request, 'ar_advance')
    if denied:
        return denied
    if request.method != 'DELETE':
        return err('Method not allowed', 405)
    denied = _action_denied(request, 'adv_installment')
    if denied:
        return denied
    try:
        inst = AdvanceInstallment.objects.select_related('advance_record').get(
            pk=iid, advance_record_id=pk)
    except AdvanceInstallment.DoesNotExist:
        return err('收付明细不存在', 404)
    rec = inst.advance_record
    if request.pk_role != 'super_admin' and rec.delivery_dept not in request.pk_depts:
        return err('无权访问', 403)
    denied = _ar_field_denied(request, 'adv_amount')
    if denied:
        return denied
    try:
        with transaction.atomic():
            inst.delete()
    except ValidationError as e:
        return err(str(e.message if hasattr(e, 'message') else e), 400)
    return ok(_adv_installments_payload(rec))


@csrf_exempt
@pk_required()
def advance_batch_writeoff(request, pk):
    """POST /advances/<pk>/batch-writeoff — 一笔预收按先进先出批量冲抵多条应收。

    解决「一个客户多个项目/多条应收，逐条核销太繁」：
    body: { record_ids: [int,...], amount(选填,默认=min(预收余额, 所选未收合计)),
            writeoff_date(必填) }
    按运作日期先进先出逐条冲抵：每条生成一笔核销 + 一笔「预收抵扣」回款
    （与单笔核销同一底层路径，可在两侧追溯）。
    匹配纪律与单笔核销一致：预收挂项目则只能冲该项目的应收；
    散单预收（未挂项目）则应收的客户名称须与往来单位一致。
    """
    denied = _page_denied(request, 'ar_advance') or _page_denied(request, 'ar_records')
    if denied:
        return denied
    if request.method != 'POST':
        return err('POST only', 405)
    denied = _action_denied(request, 'wo_receive') or _ar_field_denied(request, 'adv_writeoff')
    if denied:
        return denied
    try:
        adv = AdvanceRecord.objects.select_related('project').get(pk=pk)
    except AdvanceRecord.DoesNotExist:
        return err('预收记录不存在', 404)
    if adv.direction != '预收':
        return err('仅「预收」可批量冲抵应收账款')
    if request.pk_role != 'super_admin' and adv.delivery_dept not in request.pk_depts:
        return err('无权访问', 403)

    data = _parse_body(request)
    ids = data.get('record_ids') or []
    if not isinstance(ids, list) or not ids:
        return err('请勾选要冲抵的应收明细（record_ids）')
    try:
        ids = [int(i) for i in ids]
    except (TypeError, ValueError):
        return err('record_ids 必须为整数列表')
    wo_date = _normalize_date(data.get('writeoff_date'))
    if not wo_date:
        return err('核销日期无效（格式 2026-01-20）')

    recs = list(ARRecord.objects.select_related('project')
                .filter(pk__in=ids).order_by('operation_date', 'id'))
    if len(recs) != len(set(ids)):
        return err('部分应收明细不存在')
    for r in recs:
        if request.pk_role != 'super_admin' and r.delivery_dept not in request.pk_depts:
            return err(f'无权操作应收明细 #{r.id} 所属部门', 403)
        if adv.project_id and r.project_id != adv.project_id:
            return err(f'应收「{r.project.short_name} {r.operation_date}」与预收所属项目不一致，无法冲抵')
        if not adv.project_id and adv.counterparty and (
                (r.project.customer_name or '').strip().lower()
                != adv.counterparty.strip().lower()):
            return err(f'应收「{r.project.short_name}」客户为「{r.project.customer_name}」，'
                       f'与预收往来单位「{adv.counterparty}」不一致，无法冲抵')
    open_recs = [r for r in recs if (r.outstanding_amount or Decimal('0')) > 0]
    if not open_recs:
        return err('所选应收明细均已结清，无未收余额可冲抵')

    balance = adv.balance_amount or Decimal('0')
    total_outstanding = sum(r.outstanding_amount for r in open_recs)
    default_amt = min(balance, total_outstanding)
    amount = _dec(data.get('amount', 0)) or default_amt
    if amount <= 0:
        return err('核销金额必须大于0')
    if amount > balance:
        return err(f'核销金额 {amount} 超过预收未核销余额 {balance}')
    if amount > total_outstanding:
        return err(f'核销金额 {amount} 超过所选应收未收合计 {total_outstanding}，'
                   f'请按 {total_outstanding} 以内录入')

    allocations = []
    try:
        with transaction.atomic():
            adv_locked = AdvanceRecord.objects.select_for_update().get(pk=adv.pk)
            remaining = amount
            last_wo = adv_locked.writeoffs.select_for_update().order_by('-writeoff_no').first()
            next_no = (last_wo.writeoff_no + 1) if last_wo else 1
            for r in open_recs:
                if remaining <= 0:
                    break
                alloc = min(remaining, r.outstanding_amount)
                pay = _create_offset_payment(adv_locked, r, alloc, wo_date)
                AdvanceWriteoff.objects.create(
                    advance_record=adv_locked, writeoff_no=next_no, amount=alloc,
                    writeoff_date=wo_date, ar_record=r, ar_payment=pay,
                    notes=f'批量核销（{len(open_recs)}条应收先进先出）')
                next_no += 1
                r.refresh_from_db()
                allocations.append({
                    'record_id': r.id, 'short_name': r.project.short_name,
                    'operation_date': str(r.operation_date) if r.operation_date else None,
                    'allocated': str(alloc),
                    'outstanding_after': str(r.outstanding_amount),
                })
                remaining -= alloc
    except ValidationError as e:
        return err(str(e.message if hasattr(e, 'message') else e), 400)

    adv.refresh_from_db()
    settled = sum(1 for a in allocations if Decimal(a['outstanding_after']) <= 0)
    return ok({
        'advance_id': adv.id, 'amount': str(amount), 'allocations': allocations,
        'advance_balance_after': str(adv.balance_amount),
        'message': (f'预收 {amount} 已按运作日期先进先出冲抵 {len(allocations)} 条应收'
                    f'（{settled} 条就此结清），预收剩余 {adv.balance_amount}'),
    })


@csrf_exempt
@pk_required()
def advance_diff_summary(request):
    """GET /advances/diff-summary — 收付差异：预收 vs 预付经「项目简称」对齐。

    每个挂了项目的预收/预付按项目聚合：预收金额、预付金额、差异（预收−预付）、
    备注（成员备注去重并列），并附两侧逐笔明细（日期/金额/往来单位）供展开。
    入参：dept(可选) q(项目名模糊,可选)。
    """
    denied = _page_denied(request, 'ar_advance')
    if denied:
        return denied
    if request.method != 'GET':
        return err('Method not allowed', 405)
    dept = (request.GET.get('dept') or '').strip()
    q = (request.GET.get('q') or '').strip().lower()

    qs = (_advance_dept_filter(AdvanceRecord.objects.select_related('project'), request)
          .filter(project__isnull=False)
          .prefetch_related('installments'))
    if dept:
        qs = qs.filter(delivery_dept=dept)

    groups = {}
    for a in qs.order_by('occur_date', 'id'):
        name = (a.project.short_name or '').strip()
        if not name:
            continue
        if q and q not in name.lower():
            continue
        g = groups.setdefault(name, {
            'project': name,
            'dept': a.project.delivery_dept or a.delivery_dept or '',
            'in_total': Decimal('0'), 'out_total': Decimal('0'),
            'in_items': [], 'out_items': [], '_notes': [],
        })
        # 明细按「收付明细」逐笔列出（实际发生日期，多次到账/付出各一行）；
        # 无明细的历史直建记录回退用记录级款项日期
        insts = list(a.installments.all())
        balance_left = a.balance_amount or Decimal('0')
        if insts:
            items = [{
                'id': f'{a.id}-{i.id}',
                'occur_date': str(i.occur_date),
                'amount': str(i.amount or 0),
                'counterparty': a.counterparty or '—',
            } for i in sorted(insts, key=lambda x: (x.occur_date, x.install_no))]
        else:
            items = [{
                'id': str(a.id),
                'occur_date': (str(a.occur_date) if a.occur_date
                               else f'{a.occur_year}-{a.occur_month:02d}'),
                'amount': str(a.advance_amount or 0),
                'counterparty': a.counterparty or '—',
            }]
        # 未核销余额标在该记录的最后一笔明细上（余额属于记录而非单笔）
        items[-1]['balance'] = str(balance_left)
        for it in items:
            it.setdefault('balance', '0')
        if a.direction == '预收':
            g['in_total'] += a.advance_amount or Decimal('0')
            g['in_items'].extend(items)
        else:
            g['out_total'] += a.advance_amount or Decimal('0')
            g['out_items'].extend(items)
        note = (a.notes or '').strip()
        if note and note not in g['_notes']:
            g['_notes'].append(note)

    rows = []
    for g in groups.values():
        notes = '；'.join(g.pop('_notes'))[:300]
        rows.append({
            **{k: (str(v) if isinstance(v, Decimal) else v) for k, v in g.items()},
            'diff': str(g['in_total'] - g['out_total']),
            'notes': notes,
        })
    # 差异绝对值大的排前面（最该关注的）
    rows.sort(key=lambda x: abs(Decimal(x['diff'])), reverse=True)

    t_in = sum(Decimal(x['in_total']) for x in rows)
    t_out = sum(Decimal(x['out_total']) for x in rows)
    return ok({
        'rows': rows,
        'summary': {'count': len(rows), 'in_total': str(t_in), 'out_total': str(t_out),
                    'diff': str(t_in - t_out)},
    })


@csrf_exempt
@pk_required()
def advance_offset_workbench(request):
    """GET /advances/offset-workbench — 预收核销工作台（应收账款·预收核销 Tab）。

    按客户聚合：有未核销预收余额的客户 × 其名下仍有未收的应收明细，
    一屏看清「谁的预收还没用、能冲谁的账」，支持勾选批量核销。
    入参：dept(可选) q(客户名模糊,可选)
    """
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    if request.method != 'GET':
        return err('Method not allowed', 405)
    dept = (request.GET.get('dept') or '').strip()
    q = (request.GET.get('q') or '').strip()

    adv_qs = (_advance_dept_filter(AdvanceRecord.objects.select_related('project'), request)
              .filter(direction='预收', balance_amount__gt=0))
    if dept:
        adv_qs = adv_qs.filter(delivery_dept=dept)

    # 客户键：挂项目的取项目客户名，散单取往来单位
    groups = {}
    for a in adv_qs.order_by('occur_date', 'id'):
        cust = ((a.project.customer_name if a.project_id else a.counterparty) or '').strip()
        if not cust:
            continue
        if q and q.lower() not in cust.lower():
            continue
        g = groups.setdefault(cust, {'advances': [], 'records': [], '_proj_ids': set()})
        g['advances'].append({
            'id': a.id, 'counterparty': a.counterparty,
            'project_id': a.project_id,
            'short_name': a.project.short_name if a.project_id else None,
            'occur_date': str(a.occur_date) if a.occur_date else None,
            'balance_amount': str(a.balance_amount),
            'delivery_dept': a.delivery_dept,
        })
        if a.project_id:
            g['_proj_ids'].add(a.project_id)

    if not groups:
        return ok({'groups': [], 'total': 0})

    # 各客户名下仍有未收的应收（部门作用域内）
    rec_qs = (_ar_dept_filter(ARRecord.objects.select_related('project'), request,
                              shared_field='project__is_shared')
              .filter(outstanding_amount__gt=0,
                      project__customer_name__in=list(groups.keys())))
    if dept:
        rec_qs = rec_qs.filter(delivery_dept=dept)
    for r in rec_qs.order_by('operation_date', 'id'):
        cust = (r.project.customer_name or '').strip()
        if cust not in groups:
            continue
        groups[cust]['records'].append({
            'id': r.id, 'project_id': r.project_id,
            'short_name': r.project.short_name,
            'operation_date': str(r.operation_date) if r.operation_date else None,
            'due_date': str(r.due_date) if r.due_date else None,
            'estimated': str(r.estimated_amount or 0),
            'outstanding': str(r.outstanding_amount or 0),
            'delivery_dept': r.delivery_dept,
        })

    rows = []
    for cust, g in sorted(groups.items()):
        bal = sum(Decimal(a['balance_amount']) for a in g['advances'])
        out = sum(Decimal(r['outstanding']) for r in g['records'])
        rows.append({
            'customer': cust,
            'advances': g['advances'],
            'records': g['records'],
            'total_balance': str(bal),
            'total_outstanding': str(out),
            'offsettable': str(min(bal, out)),
        })
    # 有账可冲的排前面
    rows.sort(key=lambda x: Decimal(x['offsettable']), reverse=True)
    return ok({'groups': rows, 'total': len(rows)})


@csrf_exempt
@pk_required()
def advance_writeoff_detail(request, pk, wid):
    denied = _page_denied(request, 'ar_advance')
    if denied:
        return denied
    try:
        wo = AdvanceWriteoff.objects.select_related('advance_record__project').get(
            pk=wid, advance_record_id=pk)
    except AdvanceWriteoff.DoesNotExist:
        return err('核销记录不存在', 404)
    if request.pk_role != 'super_admin' and wo.advance_record.delivery_dept not in request.pk_depts:
        return err('无权访问', 403)
    denied = _ar_field_denied(request, 'adv_writeoff')
    if denied:
        return denied

    _wo_action = 'wo_receive' if wo.advance_record.direction == '预收' else 'wo_prepaid'
    if request.method == 'PUT':
        denied = _action_denied(request, _wo_action)
        if denied:
            return denied
        data = _parse_body(request)
        new_amount = None
        if 'amount' in data:
            new_amount = _dec(data['amount'])
            if new_amount <= 0:
                return err('核销金额必须大于0')
            # 若该核销已生成「预收抵扣」回款，新金额不得超过应收可冲抵额
            # （当前未收余额 + 本笔已冲抵额）。
            if wo.ar_payment_id:
                ar = wo.ar_payment.ar_record
                available = (ar.outstanding_amount or Decimal('0')) + (wo.ar_payment.amount or Decimal('0'))
                if new_amount > available:
                    return err(f'冲抵金额 {new_amount:,.2f} 超过该应收可冲抵额 {available:,.2f}')
            wo.amount = new_amount
        if 'writeoff_date' in data:
            wo.writeoff_date = _normalize_date(data['writeoff_date']) or wo.writeoff_date
        if 'notes' in data:
            wo.notes = (data['notes'] or '').strip()
        try:
            with transaction.atomic():
                if wo.ar_payment_id and (new_amount is not None or 'writeoff_date' in data):
                    pay = wo.ar_payment
                    if new_amount is not None:
                        pay.amount = new_amount
                    pay.payment_date = wo.writeoff_date
                    pay.save()  # 触发应收 outstanding 重算
                wo.save()
        except ValidationError as e:
            return err(str(e.message if hasattr(e, 'message') else e), 400)
        return ok(wo.to_dict())

    if request.method == 'DELETE':
        # 删核销 = 反向核销（撤销自己的操作），按核销操作权限把关，
        # 不要求通用删除权限——出纳能核销就能撤销自己的核销
        denied = _action_denied(request, _wo_action)
        if denied:
            return denied
        # 显式事务：核销删除会级联删「预收抵扣」回款并重算应收/预收，
        # 任一环失败需整体回滚，避免核销没了但回款还在的半套状态。
        try:
            with transaction.atomic():
                wo.delete()
        except ValidationError as e:
            return err(str(e.message if hasattr(e, 'message') else e), 400)
        return ok({'deleted': wid})

    return err('Method not allowed', 405)


# ══════════════════════════════════════════════════════════════════════════════
# Cashflow comparison (AR collected vs AP paid)
# ══════════════════════════════════════════════════════════════════════════════

@csrf_exempt
@pk_required()
def cashflow(request):
    denied = _page_denied(request, 'ar_cashflow')
    if denied:
        return denied

    # Date range — accept day-level start_date/end_date; fall back to year/month
    today = datetime.date.today()
    start_date_raw = (request.GET.get('start_date') or '').strip()
    end_date_raw = (request.GET.get('end_date') or '').strip()
    if start_date_raw and end_date_raw:
        try:
            start_date = datetime.date.fromisoformat(start_date_raw)
            end_date = datetime.date.fromisoformat(end_date_raw)
        except ValueError:
            return err('日期格式错误，应为 YYYY-MM-DD')
        if end_date < start_date:
            return err('结束日期不能早于起始日期')
        start_year, start_month = start_date.year, start_date.month
        end_year, end_month = end_date.year, end_date.month
    else:
        start_year = _int_param(request, 'start_year', today.year)
        start_month = _int_param(request, 'start_month', 1)
        end_year = _int_param(request, 'end_year', today.year)
        end_month = _int_param(request, 'end_month', today.month)
        start_date = datetime.date(start_year, start_month, 1)
        end_date = datetime.date(end_year, end_month,
                                 calendar.monthrange(end_year, end_month)[1])

    # Departments to show
    if request.pk_role == 'super_admin':
        allowed = DEPARTMENTS
    else:
        allowed = request.pk_depts
    depts_param = request.GET.get('depts', '').strip()
    if depts_param:
        requested = [d.strip() for d in depts_param.split(',') if d.strip()]
        depts = [d for d in requested if d in allowed]
    else:
        depts = list(allowed)

    # Build month list
    months = []
    y, m = start_year, start_month
    while (y, m) <= (end_year, end_month):
        months.append((y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1
    month_keys = [f'{y}-{m:02d}' for y, m in months]

    # AR collections by month across all requested depts.
    # 排除「预收抵扣」：其现金已在预收发生时计入流入，重复计会虚增现金。
    ar_coll = (ARPayment.objects
               .filter(payment_date__gte=start_date, payment_date__lte=end_date,
                       ar_record__delivery_dept__in=depts)
               .exclude(source='预收抵扣')
               .annotate(ym=TruncMonth('payment_date'))
               .values('ym', 'ar_record__delivery_dept')
               .annotate(collected=Sum('amount')))
    coll_map = defaultdict(lambda: defaultdict(Decimal))
    for row in ar_coll:
        ym = row['ym'].strftime('%Y-%m')
        dept = row['ar_record__delivery_dept']
        coll_map[dept][ym] += row['collected'] or Decimal('0')

    # AP payments from installments subtable, grouped by month + department
    paid_map = defaultdict(lambda: defaultdict(Decimal))
    inst_qs = (PaymentInstallment.objects
               .filter(pay_date__gte=start_date, pay_date__lte=end_date,
                       payment__department__in=depts)
               .annotate(ym=TruncMonth('pay_date'))
               .values('ym', 'payment__department')
               .annotate(paid=Sum('pay_amount')))
    for row in inst_qs:
        ym = row['ym'].strftime('%Y-%m')
        dept = row['payment__department']
        paid_map[dept][ym] += row['paid'] or Decimal('0')

    # 扣除预付核销冲抵：按最早实付日期（首个 installment 付款日期，否则 planned_date）归月
    from django.db.models import OuterRef, Subquery
    earliest_inst_date = Subquery(
        PaymentInstallment.objects.filter(payment_id=OuterRef('pk'))
            .order_by('pay_date').values('pay_date')[:1]
    )
    po_qs = (Payment.objects
             .filter(department__in=depts, prepaid_offset_amount__gt=0)
             .annotate(first_pay_date=earliest_inst_date)
             .annotate(attr_ym=TruncMonth(Coalesce('first_pay_date', 'planned_date')))
             .filter(attr_ym__gte=start_date, attr_ym__lte=end_date)
             .values('attr_ym', 'department')
             .annotate(offset=Sum('prepaid_offset_amount')))
    for row in po_qs:
        ym = row['attr_ym'].strftime('%Y-%m')
        dept = row['department']
        paid_map[dept][ym] = max(Decimal('0'),
                                 paid_map[dept].get(ym, Decimal('0')) - (row['offset'] or Decimal('0')))

    # 预收(流入) / 预付(流出) by occur_date month — advances move cash on occur_date
    adv_recv_map = defaultdict(lambda: defaultdict(Decimal))
    adv_paid_map = defaultdict(lambda: defaultdict(Decimal))
    adv_qs = (AdvanceRecord.objects
              .filter(occur_date__gte=start_date, occur_date__lte=end_date,
                      delivery_dept__in=depts)
              .annotate(ym=TruncMonth('occur_date'))
              .values('ym', 'delivery_dept', 'direction')
              .annotate(amt=Sum('advance_amount')))
    for row in adv_qs:
        ym = row['ym'].strftime('%Y-%m')
        target = adv_recv_map if row['direction'] == '预收' else adv_paid_map
        target[row['delivery_dept']][ym] += row['amt'] or Decimal('0')

    # Also include budget data for comparison
    budget_coll_map = defaultdict(lambda: defaultdict(Decimal))
    bc_qs = (CollectionBudget.objects
             .filter(expected_date__gte=start_date, expected_date__lte=end_date,
                     delivery_dept__in=depts)
             .annotate(ym=TruncMonth('expected_date'))
             .values('ym', 'delivery_dept')
             .annotate(budget=Sum('amount')))
    for row in bc_qs:
        ym = row['ym'].strftime('%Y-%m')
        budget_coll_map[row['delivery_dept']][ym] += row['budget'] or Decimal('0')

    budget_paid_map = defaultdict(lambda: defaultdict(Decimal))
    bp_qs = (PaymentBudget.objects
             .filter(expected_date__gte=start_date, expected_date__lte=end_date,
                     delivery_dept__in=depts)
             .annotate(ym=TruncMonth('expected_date'))
             .values('ym', 'delivery_dept')
             .annotate(budget=Sum('amount')))
    for row in bp_qs:
        ym = row['ym'].strftime('%Y-%m')
        budget_paid_map[row['delivery_dept']][ym] += row['budget'] or Decimal('0')

    # Build per-dept series + total
    by_dept = []
    total_coll = defaultdict(Decimal)
    total_paid = defaultdict(Decimal)
    total_adv_recv = defaultdict(Decimal)
    total_adv_paid = defaultdict(Decimal)
    total_bcoll = defaultdict(Decimal)
    total_bpaid = defaultdict(Decimal)
    has_alert = False

    for dept in depts:
        series_coll = [float(coll_map[dept].get(ym, 0)) for ym in month_keys]
        series_paid = [float(paid_map[dept].get(ym, 0)) for ym in month_keys]
        series_arecv = [float(adv_recv_map[dept].get(ym, 0)) for ym in month_keys]
        series_apaid = [float(adv_paid_map[dept].get(ym, 0)) for ym in month_keys]
        series_bcoll = [float(budget_coll_map[dept].get(ym, 0)) for ym in month_keys]
        series_bpaid = [float(budget_paid_map[dept].get(ym, 0)) for ym in month_keys]
        # 流入 = 回款 + 预收；流出 = 付款 + 预付
        inflow = [series_coll[i] + series_arecv[i] for i in range(len(month_keys))]
        outflow = [series_paid[i] + series_apaid[i] for i in range(len(month_keys))]
        alert_months = [month_keys[i] for i in range(len(month_keys))
                        if outflow[i] > inflow[i] > 0]
        if alert_months:
            has_alert = True
        by_dept.append({
            'dept': dept,
            'collected': series_coll,
            'paid': series_paid,
            'advance_received': series_arecv,
            'advance_paid': series_apaid,
            'budget_collection': series_bcoll,
            'budget_payment': series_bpaid,
            'alert_months': alert_months,
        })
        for i, ym in enumerate(month_keys):
            total_coll[ym] += coll_map[dept].get(ym, Decimal('0'))
            total_paid[ym] += paid_map[dept].get(ym, Decimal('0'))
            total_adv_recv[ym] += adv_recv_map[dept].get(ym, Decimal('0'))
            total_adv_paid[ym] += adv_paid_map[dept].get(ym, Decimal('0'))
            total_bcoll[ym] += budget_coll_map[dept].get(ym, Decimal('0'))
            total_bpaid[ym] += budget_paid_map[dept].get(ym, Decimal('0'))

    collected_arr = [float(total_coll.get(ym, 0)) for ym in month_keys]
    paid_arr = [float(total_paid.get(ym, 0)) for ym in month_keys]
    adv_recv_arr = [float(total_adv_recv.get(ym, 0)) for ym in month_keys]
    adv_paid_arr = [float(total_adv_paid.get(ym, 0)) for ym in month_keys]
    inflow_arr = [round(collected_arr[i] + adv_recv_arr[i], 2) for i in range(len(month_keys))]
    outflow_arr = [round(paid_arr[i] + adv_paid_arr[i], 2) for i in range(len(month_keys))]
    net_arr = [round(inflow_arr[i] - outflow_arr[i], 2) for i in range(len(month_keys))]

    total_alert = [month_keys[i] for i in range(len(month_keys))
                   if outflow_arr[i] > inflow_arr[i] > 0]
    if total_alert:
        has_alert = True

    cumulative = []
    running = 0.0
    for v in net_arr:
        running += v
        cumulative.append(round(running, 2))

    return ok({
        'months': month_keys,
        'depts': depts,
        'by_dept': by_dept,
        'totals': {
            'collected': collected_arr,
            'paid': paid_arr,
            'advance_received': adv_recv_arr,
            'advance_paid': adv_paid_arr,
            'inflow': inflow_arr,
            'outflow': outflow_arr,
            'net': net_arr,
            'cumulative_net': cumulative,
            'budget_collection': [float(total_bcoll.get(ym, 0)) for ym in month_keys],
            'budget_payment': [float(total_bpaid.get(ym, 0)) for ym in month_keys],
            'alert_months': total_alert,
        },
        'has_alert': has_alert,
        'start_date': str(start_date),
        'end_date': str(end_date),
    })


# ══════════════════════════════════════════════════════════════════════════════
# Budget
# ══════════════════════════════════════════════════════════════════════════════

def _budget_list_create(request, Model, page_key):
    denied = _page_denied(request, page_key)
    if denied:
        return denied

    if request.method == 'GET':
        qs = _ar_dept_filter(Model.objects.select_related('created_by'), request,
                             dept_field='delivery_dept')
        dept = request.GET.get('dept', '').strip()
        if dept:
            qs = qs.filter(delivery_dept=dept)
        _today = datetime.date.today()
        _ds, _de = _parse_budget_date_range(request, _today)
        qs = qs.filter(expected_date__gte=_ds, expected_date__lte=_de)
        page = max(1, int(request.GET.get('page', 1) or 1))
        size = min(200, max(1, int(request.GET.get('size', 50) or 50)))
        total = qs.count()
        items = [obj.to_dict() for obj in qs[(page - 1) * size: page * size]]
        total_amount = str(qs.aggregate(s=Sum('amount'))['s'] or 0)
        return ok({'items': items, 'total': total, 'page': page, 'size': size,
                   'total_amount': total_amount})

    if request.method == 'POST':
        denied = _write_denied(request)
        if denied:
            return denied
        data = _parse_body(request)
        dept = data.get('delivery_dept', '').strip()
        short_name = data.get('short_name', '').strip()
        # Auto-derive dept from matched project when caller omits delivery_dept
        if not dept and short_name:
            search_depts = None if request.pk_role == 'super_admin' else request.pk_depts
            _pre_proj = _match_project_by_short_name(short_name, allowed_depts=search_depts)
            if _pre_proj:
                dept = _pre_proj.delivery_dept
        if dept and dept not in VALID_DEPARTMENTS:
            return err(f'无效交付部门: {dept}')
        denied = _dept_denied(request, dept, '无权操作此部门')
        if denied:
            return denied
        date_str = _normalize_date(data.get('expected_date'))
        if not date_str:
            return err('预计日期无效，请填 2026-06-15 这样的格式（也支持 2026/6/15、2026年6月15日）')
        amount = _dec(data.get('amount', 0))
        if amount <= 0:
            return err('金额必须大于0')
        from paikuan.models import PaikuanUser
        user = PaikuanUser.objects.filter(id=request.pk_uid).first()
        proj = _match_project_by_short_name(short_name, dept)
        auto_project_no = proj.project_no if proj else data.get('project_no', '').strip()
        auto_sub_dept = proj.sub_dept if proj else data.get('sub_dept', '').strip()
        obj = Model.objects.create(
            project_no=auto_project_no,
            short_name=short_name,
            expected_date=date_str,
            sub_dept=auto_sub_dept,
            delivery_dept=dept,
            amount=amount,
            notes=data.get('notes', '').strip(),
            created_by=user,
        )
        return ok(obj.to_dict())

    return err('Method not allowed', 405)


def _budget_detail(request, pk, Model, page_key):
    denied = _page_denied(request, page_key)
    if denied:
        return denied
    try:
        obj = Model.objects.get(pk=pk)
    except Model.DoesNotExist:
        return err('记录不存在', 404)
    denied = _object_dept_denied(request, obj)
    if denied:
        return denied

    if request.method == 'GET':
        return ok(obj.to_dict())

    if request.method == 'PUT':
        denied = _write_denied(request)
        if denied:
            return denied
        data = _parse_body(request)
        for field in ('project_no', 'short_name', 'sub_dept', 'notes'):
            if field in data:
                setattr(obj, field, data[field])
        if 'delivery_dept' in data:
            dept = data['delivery_dept']
            if dept and dept not in VALID_DEPARTMENTS:
                return err(f'无效交付部门: {dept}')
            denied = _dept_denied(request, dept, '无权操作目标部门')
            if denied:
                return denied
            obj.delivery_dept = dept
        if 'expected_date' in data:
            obj.expected_date = _normalize_date(data['expected_date']) or obj.expected_date
        if 'amount' in data:
            amount = _dec(data['amount'])
            if amount <= 0:
                return err('金额必须大于0')
            obj.amount = amount
        obj.save()
        return ok(obj.to_dict())

    if request.method == 'DELETE':
        denied = _delete_denied(request)
        if denied:
            return denied
        obj.delete()
        return ok({'deleted': pk})

    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def budget_collection(request):
    return _budget_list_create(request, CollectionBudget, 'ar_budget')


@csrf_exempt
@pk_required()
def budget_collection_detail(request, pk):
    return _budget_detail(request, pk, CollectionBudget, 'ar_budget')


@csrf_exempt
@pk_required()
def budget_payment(request):
    return _budget_list_create(request, PaymentBudget, 'ar_budget')


@csrf_exempt
@pk_required()
def budget_payment_detail(request, pk):
    return _budget_detail(request, pk, PaymentBudget, 'ar_budget')


def _parse_budget_date_range(request, today):
    """Return (start_date, end_date) from date_start/date_end GET params.
    Defaults to the current month when neither param is supplied."""
    ds = request.GET.get('date_start', '').strip()
    de = request.GET.get('date_end', '').strip()
    try:
        start = datetime.date.fromisoformat(ds) if ds else None
    except ValueError:
        start = None
    try:
        end = datetime.date.fromisoformat(de) if de else None
    except ValueError:
        end = None
    if not start and not end:
        start = datetime.date(today.year, today.month, 1)
        end = datetime.date(today.year, today.month,
                            calendar.monthrange(today.year, today.month)[1])
    elif not start:
        start = datetime.date(end.year, end.month, 1)
    elif not end:
        end = datetime.date(start.year, start.month,
                            calendar.monthrange(start.year, start.month)[1])
    return start, end


def _budget_label(kind):
    return '收款' if kind == 'collection' else '付款'


def _budget_template(request, kind):
    denied = _page_denied(request, 'ar_budget')
    if denied:
        return denied
    lbl = _budget_label(kind)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{lbl}预算'
    headers = ['项目编号(可选)', '项目简称/摘要*', f'预计{lbl}日期*',
               '二级部门', '交付部门', '金额*', '备注']
    color = '1B6E35' if kind == 'collection' else 'B25600'
    _header_row(ws, headers, color=color)

    # 第2行：填写说明（导入时自动跳过——简称列以 ★ 开头）
    tips = [
        '选填：项目编号；留空会按简称自动从台账带入',
        f'★必填：项目台账中的"项目简称"（精确匹配；同名多项目时用"交付部门"列消歧）',
        f'★必填：预计{lbl}日期，格式 2026-06-15 / 2026/6/15 / 2026年6月15日 均可；'
        '若单元格被Excel识别为日期格式也可直接选日期',
        '选填：二级部门；留空或填错会按台账自动更正',
        '选填：交付部门；留空会按台账带入，仅在简称重名时用于区分',
        '★必填：金额（元），须大于 0',
        '选填：备注',
    ]
    ws.append(tips)
    tip_row = ws.max_row
    tip_fill = PatternFill('solid', fgColor='FFF8E1' if kind == 'collection' else 'FFF3E0')
    tip_font = Font(italic=True, color='6D4C00', size=9)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(tip_row, c)
        cell.fill = tip_fill
        cell.font = tip_font
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[tip_row].height = 56

    # 第3行：示例（简称列含"示例"标记，导入时自动跳过）
    ws.append(['', f'{EXAMPLE_ROW_MARKER}示例项目A', '2026-06-15', '华南区',
               '劳务事业部', 100000, '填对项目简称即可，编号/交付部门/二级部门会自动按台账带入或更正'])
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
    return _export_response(wb, f'{lbl}预算导入模板.xlsx')


def _budget_import(request, Model, kind):
    denied = _page_denied(request, 'ar_budget')
    if denied:
        return denied
    if request.method != 'POST':
        return err('POST only', 405)
    denied = _write_denied(request)
    if denied:
        return denied
    f = request.FILES.get('file')
    if not f:
        return err('请上传文件')
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
    except Exception as e:
        return err(f'无法读取Excel: {e}')

    headers = [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(headers)}
    lbl = _budget_label(kind)

    def _cv(row, name):
        idx = col_map.get(name)
        if idx is None:
            return ''
        v = ws.cell(row, idx).value
        return str(v).strip() if v is not None else ''

    # 日期列名：新模板用"预计X日期*"，旧模板用"预计X日期(YYYY-MM-DD)*"，两者都兼容。
    date_col_names = [f'预计{lbl}日期*', f'预计{lbl}日期(YYYY-MM-DD)*',
                      f'预计{lbl}日期']

    def _raw_first(row, names):
        # Return the first matching cell's RAW value (not stringified) so that
        # openpyxl-typed dates reach _normalize_date intact. Excel-recognised date
        # cells come back as datetime objects; str() would yield
        # '2026-06-15 00:00:00', which _normalize_date cannot parse — that was the
        # "预计日期无效" failure on template-filled rows.
        for n in names:
            if n in col_map:
                return ws.cell(row, col_map[n]).value
        return None

    def _cv_first(row, names):
        for n in names:
            if n in col_map:
                v = ws.cell(row, col_map[n]).value
                return str(v).strip() if v is not None else ''
        return ''

    from paikuan.models import PaikuanUser
    user = PaikuanUser.objects.filter(id=request.pk_uid).first()
    search_depts = None if request.pk_role == 'super_admin' else request.pk_depts
    corrected = 0
    errors = []
    warnings = []
    plan = []   # 通过校验、待写入的行

    # ══ 阶段一：逐行校验 + 按台账自动更正（不写库）。示例/空行静默忽略 ══════════════
    for ri in range(2, ws.max_row + 1):
        short_name = _cv(ri, '项目简称/摘要*')
        # 示例/提示行、必填摘要为空：静默忽略（不计入、不报错）
        if not short_name or EXAMPLE_ROW_MARKER in short_name or short_name.startswith('★'):
            continue
        date_str = _normalize_date(_raw_first(ri, date_col_names))
        if not date_str:
            raw_date = _cv_first(ri, date_col_names)
            hint = f'（读到"{raw_date}"）' if raw_date else '（该格为空）'
            errors.append(f'第{ri}行: 预计{lbl}日期无效{hint}，请填 2026-06-15 这样的格式')
            continue
        filled_dept = _cv(ri, '交付部门')

        # 项目简称是权威来源：先只按简称匹配项目（忽略可能填错的部门列）。
        # 简称对应多个部门时，用填写的交付部门消歧。
        name_matches = list(ARProject.objects.filter(short_name=short_name.strip()))
        if search_depts is not None:
            name_matches = [m for m in name_matches if m.delivery_dept in search_depts]
        proj = None
        if len(name_matches) == 1:
            proj = name_matches[0]
        elif len(name_matches) > 1:
            cands = [m for m in name_matches if m.delivery_dept == filled_dept] if filled_dept else []
            if len(cands) == 1:
                proj = cands[0]
            else:
                depts_hint = '、'.join(sorted({m.delivery_dept for m in name_matches}))
                errors.append(f'第{ri}行: 项目简称"{short_name}"对应多个部门（{depts_hint}），'
                              f'请在"交付部门"列填写正确部门以区分')
                continue
        else:
            # 简称无精确匹配，回退到模糊匹配（容许用户填部分简称）
            proj = _match_project_by_short_name(short_name, filled_dept, search_depts)

        if proj:
            # 以台账为准：用项目的编号/部门/二级部门覆盖填写值（自动更正，不算跳过）
            dept = proj.delivery_dept
            project_no = proj.project_no
            sub_dept = proj.sub_dept
            mismatches = []
            if filled_dept and filled_dept != dept:
                mismatches.append(f'交付部门"{filled_dept}"→"{dept}"')
            filled_no = _cv(ri, '项目编号(可选)')
            if filled_no and filled_no != project_no:
                mismatches.append(f'项目编号"{filled_no}"→"{project_no}"')
            filled_sub = _cv(ri, '二级部门')
            if filled_sub and filled_sub != sub_dept:
                mismatches.append(f'二级部门"{filled_sub}"→"{sub_dept}"')
            if mismatches:
                corrected += 1
                warnings.append(f'第{ri}行: 已按项目台账自动更正（{"，".join(mismatches)}）')
        else:
            # 自由文本摘要（非台账项目）：采用填写值
            dept = filled_dept
            project_no = _cv(ri, '项目编号(可选)')
            sub_dept = _cv(ri, '二级部门')

        if dept and dept not in VALID_DEPARTMENTS:
            errors.append(f'第{ri}行: 无效交付部门"{dept}"，可选值为：{"/".join(VALID_DEPARTMENTS)}')
            continue
        if request.pk_role != 'super_admin' and dept not in request.pk_depts:
            errors.append(f'第{ri}行: 无权操作部门"{dept}"')
            continue
        amount = _dec(_cv(ri, '金额*'))
        if amount <= 0:
            errors.append(f'第{ri}行: 金额必须大于0（当前读到"{_cv(ri, "金额*") or "空"}"）')
            continue
        plan.append({'ri': ri, 'project_no': project_no, 'short_name': short_name,
                     'expected_date': date_str, 'sub_dept': sub_dept, 'dept': dept,
                     'amount': amount, 'notes': _cv(ri, '备注')})

    # ══ 有任何问题 → 整表拒绝（自动更正项仍随成功导入一并提示，不阻断）══════════════
    if errors:
        return ok({
            'rejected': True, 'created': 0, 'corrected': 0, 'errors': errors, 'warnings': warnings,
            'message': (f'导入未执行：发现 {len(errors)} 处问题，已全部列出。'
                        f'请在表格中按提示修正后重新导入（整表全部通过才会写入，不会漏导）。'),
        })

    # ══ 阶段二：全部通过 → 一次性写入（整体事务，任一失败回滚）═══════════════════════
    created = 0
    try:
        with transaction.atomic():
            for p in plan:
                Model.objects.create(
                    project_no=p['project_no'], short_name=p['short_name'],
                    expected_date=p['expected_date'], sub_dept=p['sub_dept'],
                    delivery_dept=p['dept'], amount=p['amount'],
                    notes=p['notes'], created_by=user,
                )
                created += 1
    except Exception as e:
        return ok({
            'rejected': True, 'created': 0, 'corrected': 0,
            'errors': [f'写入阶段发生错误并已回滚：{e}。请检查数据后重试。'], 'warnings': warnings,
            'message': '导入未执行（写入阶段出错，已整体回滚，不会出现半截数据）。',
        })

    return ok({'created': created, 'skipped': 0, 'corrected': corrected,
               'errors': [], 'warnings': warnings})


def _budget_export(request, Model, kind):
    denied = _page_denied(request, 'ar_budget')
    if denied:
        return denied
    qs = _ar_dept_filter(Model.objects.all(), request, dept_field='delivery_dept')
    dept = request.GET.get('dept', '').strip()
    if dept:
        qs = qs.filter(delivery_dept=dept)
    _today = datetime.date.today()
    _ds, _de = _parse_budget_date_range(request, _today)
    qs = qs.filter(expected_date__gte=_ds, expected_date__lte=_de)
    if qs.count() > 5000:
        return err('导出超过5000行，请缩小筛选范围')
    lbl = _budget_label(kind)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{lbl}预算'
    headers = ['项目编号', '项目简称/摘要', f'预计{lbl}日期', '二级部门',
               '交付部门', '金额', '备注', '创建人']
    color = '1B6E35' if kind == 'collection' else 'B25600'
    _header_row(ws, headers, color=color)
    for o in qs.select_related('created_by'):
        ws.append([o.project_no, o.short_name, str(o.expected_date), o.sub_dept,
                   o.delivery_dept, float(o.amount), o.notes,
                   o.created_by.name if o.created_by else ''])
    return _export_response(wb, f'{lbl}预算.xlsx')


@csrf_exempt
@pk_required()
def budget_collection_template(request):
    return _budget_template(request, 'collection')


@csrf_exempt
@pk_required()
def budget_collection_import(request):
    return _budget_import(request, CollectionBudget, 'collection')


@csrf_exempt
@pk_required()
def budget_collection_export(request):
    return _budget_export(request, CollectionBudget, 'collection')


@csrf_exempt
@pk_required()
def budget_payment_template(request):
    return _budget_template(request, 'payment')


@csrf_exempt
@pk_required()
def budget_payment_import(request):
    return _budget_import(request, PaymentBudget, 'payment')


@csrf_exempt
@pk_required()
def budget_payment_export(request):
    return _budget_export(request, PaymentBudget, 'payment')


@csrf_exempt
@pk_required()
def budget_summary(request):
    """Monthly budget vs actual summary for the given range and depts."""
    denied = _page_denied(request, 'ar_budget')
    if denied:
        return denied

    today = datetime.date.today()
    start_date, end_date = _parse_budget_date_range(request, today)

    if request.pk_role == 'super_admin':
        depts = list(DEPARTMENTS)
    else:
        depts = list(request.pk_depts)

    # Consume global active-depts scope (injected by axios interceptor as ?depts=A,B,C)
    raw_depts = request.GET.get('depts', '').strip()
    if raw_depts:
        requested = [d for d in raw_depts.split(',') if d.strip()]
        allowed_set = set(depts)
        active = [d for d in requested if d in allowed_set]
        if active:
            depts = active

    # Single-dept page-level filter
    dept_param = request.GET.get('dept', '').strip()
    if dept_param and dept_param in depts:
        depts = [dept_param]

    # Budget totals
    bc = CollectionBudget.objects.filter(
        expected_date__range=(start_date, end_date),
        delivery_dept__in=depts).aggregate(total=Sum('amount'))
    bp = PaymentBudget.objects.filter(
        expected_date__range=(start_date, end_date),
        delivery_dept__in=depts).aggregate(total=Sum('amount'))

    # Actual AR collections (from ARPayment)
    ac = ARPayment.objects.filter(
        payment_date__range=(start_date, end_date),
        ar_record__delivery_dept__in=depts).aggregate(total=Sum('amount'))

    # Actual AP payments (from installments subtable)
    ap_total = (PaymentInstallment.objects
                .filter(pay_date__range=(start_date, end_date),
                        payment__department__in=depts)
                .aggregate(total=Sum('pay_amount'))['total'] or Decimal('0'))

    budget_coll = bc['total'] or Decimal('0')
    budget_paid = bp['total'] or Decimal('0')
    actual_coll = ac['total'] or Decimal('0')
    actual_paid = ap_total

    # Per-dept breakdown for multi-dept comparison chart
    by_dept_result = []
    if len(depts) > 1:
        for d in depts:
            bc_d = CollectionBudget.objects.filter(
                expected_date__range=(start_date, end_date), delivery_dept=d
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
            bp_d = PaymentBudget.objects.filter(
                expected_date__range=(start_date, end_date), delivery_dept=d
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
            ac_d = ARPayment.objects.filter(
                payment_date__range=(start_date, end_date), ar_record__delivery_dept=d
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
            ap_d = (PaymentInstallment.objects
                    .filter(pay_date__range=(start_date, end_date), payment__department=d)
                    .aggregate(total=Sum('pay_amount'))['total'] or Decimal('0'))
            by_dept_result.append({
                'dept': d,
                'budget_collection': float(bc_d),
                'actual_collection': float(ac_d),
                'budget_payment': float(bp_d),
                'actual_payment': float(ap_d),
            })

    return ok({
        'start_date': str(start_date), 'end_date': str(end_date), 'depts': depts,
        'budget_collection': str(budget_coll),
        'budget_payment': str(budget_paid),
        'actual_collection': str(actual_coll),
        'actual_payment': str(actual_paid),
        'collection_achievement_rate': round(float(actual_coll / budget_coll * 100), 2) if budget_coll else None,
        'payment_achievement_rate': round(float(actual_paid / budget_paid * 100), 2) if budget_paid else None,
        'collection_gap': str(budget_coll - actual_coll),
        'payment_gap': str(budget_paid - actual_paid),
        'has_alert': actual_paid > actual_coll,
        'by_dept': by_dept_result,
    })


@csrf_exempt
@pk_required()
def budget_project_compare(request):
    """项目维度预算对照 — 预算（收/付，按项目简称）与实际（应收回款/排款实付，
    按项目简称）同窗对齐，逐项目展示「计划 vs 实际」全貌。

    入参：date_start/date_end（默认本月）、dept（可选）。
    返回 rows：每个涉及项目一行（收款预算/实际收款/达成率/付款预算/实际付款/
    执行率/净现金计划与实际/状态标签）+ summary 汇总。
    """
    denied = _page_denied(request, 'ar_budget')
    if denied:
        return denied
    if request.method != 'GET':
        return err('Method not allowed', 405)

    today = datetime.date.today()
    start_date, end_date = _parse_budget_date_range(request, today)

    if request.pk_role == 'super_admin':
        depts = list(DEPARTMENTS)
    else:
        depts = list(request.pk_depts)
    raw_depts = request.GET.get('depts', '').strip()
    if raw_depts:
        active = [d for d in raw_depts.split(',') if d.strip() and d in set(depts)]
        if active:
            depts = active
    dept_param = request.GET.get('dept', '').strip()
    if dept_param and dept_param in depts:
        depts = [dept_param]

    rows = {}

    def _row(name):
        return rows.setdefault(name, {
            'project': name, 'dept': '', 'customer': '', 'manager': '',
            'budget_in': Decimal('0'), 'actual_in': Decimal('0'),
            'budget_out': Decimal('0'), 'actual_out': Decimal('0'),
        })

    # 收款预算（按简称）
    for g in (CollectionBudget.objects
              .filter(expected_date__range=(start_date, end_date),
                      delivery_dept__in=depts)
              .exclude(short_name='')
              .values('short_name').annotate(s=Sum('amount'))):
        _row(g['short_name'])['budget_in'] += g['s'] or Decimal('0')

    # 付款预算（按简称）
    for g in (PaymentBudget.objects
              .filter(expected_date__range=(start_date, end_date),
                      delivery_dept__in=depts)
              .exclude(short_name='')
              .values('short_name').annotate(s=Sum('amount'))):
        _row(g['short_name'])['budget_out'] += g['s'] or Decimal('0')

    # 实际收款：应收回款经 项目简称（含预收抵扣——预算达成按应收口径）
    for g in (ARPayment.objects
              .filter(payment_date__range=(start_date, end_date),
                      ar_record__delivery_dept__in=depts,
                      ar_record__project__short_name__isnull=False)
              .exclude(ar_record__project__short_name='')
              .values('ar_record__project__short_name').annotate(s=Sum('amount'))):
        _row(g['ar_record__project__short_name'])['actual_in'] += g['s'] or Decimal('0')

    # 实际付款：排款实付分期经 项目简称
    for g in (PaymentInstallment.objects
              .filter(pay_date__range=(start_date, end_date),
                      payment__department__in=depts)
              .exclude(payment__project_short_name='')
              .values('payment__project_short_name').annotate(s=Sum('pay_amount'))):
        _row(g['payment__project_short_name'])['actual_out'] += g['s'] or Decimal('0')

    if rows:
        for p in ARProject.objects.filter(short_name__in=list(rows.keys())):
            r = rows.get(p.short_name)
            if r is not None and not r['dept']:
                r['dept'] = p.delivery_dept or ''
                r['customer'] = p.customer_name or ''
                r['manager'] = p.project_manager or ''

    def _rate(actual, budget):
        return round(float(actual / budget * 100), 1) if budget else None

    out_rows = []
    for r in rows.values():
        bi, ai, bo, ao = r['budget_in'], r['actual_in'], r['budget_out'], r['actual_out']
        tags = []
        if bi > 0 and ai >= bi:
            tags.append('收款达成')
        elif bi > 0 and ai < bi:
            tags.append('收款滞后')
        elif bi == 0 and ai > 0:
            tags.append('计划外收款')
        if bo > 0 and ao > bo:
            tags.append('付款超预算')
        elif bo == 0 and ao > 0:
            tags.append('计划外付款')
        out_rows.append({
            **{k: str(v) if isinstance(v, Decimal) else v for k, v in r.items()},
            'in_rate': _rate(ai, bi),
            'out_rate': _rate(ao, bo),
            'in_gap': str(bi - ai),
            'out_gap': str(bo - ao),
            'budget_net': str(bi - bo),
            'actual_net': str(ai - ao),
            'tags': tags,
        })
    # 体量大的排前面（预算+实际合计）
    out_rows.sort(key=lambda x: (Decimal(x['budget_in']) + Decimal(x['budget_out'])
                                 + Decimal(x['actual_in']) + Decimal(x['actual_out'])),
                  reverse=True)

    t_bi = sum(Decimal(x['budget_in']) for x in out_rows)
    t_ai = sum(Decimal(x['actual_in']) for x in out_rows)
    t_bo = sum(Decimal(x['budget_out']) for x in out_rows)
    t_ao = sum(Decimal(x['actual_out']) for x in out_rows)
    return ok({
        'start_date': str(start_date), 'end_date': str(end_date),
        'rows': out_rows,
        'summary': {
            'count': len(out_rows),
            'budget_in': str(t_bi), 'actual_in': str(t_ai), 'in_rate': _rate(t_ai, t_bi),
            'budget_out': str(t_bo), 'actual_out': str(t_ao), 'out_rate': _rate(t_ao, t_bo),
            'budget_net': str(t_bi - t_bo), 'actual_net': str(t_ai - t_ao),
            'achieved': sum(1 for x in out_rows if '收款达成' in x['tags']),
            'lagging': sum(1 for x in out_rows if '收款滞后' in x['tags']),
            'over_budget': sum(1 for x in out_rows if '付款超预算' in x['tags']),
            'unplanned': sum(1 for x in out_rows
                             if '计划外收款' in x['tags'] or '计划外付款' in x['tags']),
        },
    })


# ──────────────────────────────────────────────────────────────────────────────
# 合并开票批次 — 按 invoice_batch_no 分组汇总 + 批量打标
# ──────────────────────────────────────────────────────────────────────────────

@csrf_exempt
@pk_required()
def ar_invoice_batches(request):
    """GET /records/invoice-batches  —  列出所有已标记开票批次及其汇总。

    每个 invoice_batch_no 非空的批次返回：
      batch_no, count, estimated, invoiced, outstanding, record_ids, dept_names
    支持 dept / year / month 维度筛选，以及 q（简称/合同名/编号模糊匹配）。
    """
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    if request.method != 'GET':
        return err('Method not allowed', 405)

    today = datetime.date.today()
    qs = _ar_dept_filter(ARRecord.objects.all(), request, shared_field='project__is_shared')
    qs = _apply_record_filters(qs, request)
    qs = _apply_conditions(qs, request, today)
    # 只看有批次号的记录
    qs = qs.exclude(invoice_batch_no='')

    # 分组聚合（避免 payments JOIN 行扇出：先算记录级，再单独算已收）
    base = (qs.values('invoice_batch_no').annotate(
        count=Count('id'),
        estimated=Sum('estimated_amount'),
        invoiced=Sum('actual_invoice_amount'),
        outstanding=Sum('outstanding_amount', filter=Q(outstanding_amount__gt=0)),
    ).order_by('invoice_batch_no'))

    collected_map = {
        g['invoice_batch_no']: (g['s'] or 0)
        for g in qs.values('invoice_batch_no').annotate(s=Sum('payments__amount'))
    }

    # 每批次的明细 record_ids、部门与客户列表（小数据量下内存聚合比子查询清晰）
    id_dept_map = defaultdict(lambda: {'ids': [], 'depts': set(), 'customers': set()})
    for r in qs.values('id', 'invoice_batch_no', 'delivery_dept', 'project__customer_name'):
        m = id_dept_map[r['invoice_batch_no']]
        m['ids'].append(r['id'])
        m['depts'].add(r['delivery_dept'])
        if r['project__customer_name']:
            m['customers'].add(r['project__customer_name'])

    rows = []
    for g in base:
        bn = g['invoice_batch_no']
        meta = id_dept_map[bn]
        custs = sorted(meta['customers'])
        rows.append({
            'batch_no': bn,
            'count': g['count'] or 0,
            'estimated': str(g['estimated'] or 0),
            'invoiced': str(g['invoiced'] or 0),
            'outstanding': str(g['outstanding'] or 0),
            'collected': str(collected_map.get(bn, 0)),
            'record_ids': sorted(meta['ids']),
            'dept_names': sorted(meta['depts']),
            'customers': custs[:3] + (['…'] if len(custs) > 3 else []),
        })

    return ok({'batches': rows, 'total': len(rows)})


def _batch_members_qs(request, batch_no):
    """某开票批次在当前用户作用域内的成员记录（含项目信息，按运作日期排序）。"""
    return (_ar_dept_filter(ARRecord.objects.select_related('project'), request,
                            shared_field='project__is_shared')
            .filter(invoice_batch_no=batch_no)
            .order_by('operation_date', 'id'))


def _batch_member_dict(rec):
    billable = (rec.estimated_amount or Decimal('0')) + (rec.account_diff_adjustment or Decimal('0'))
    collected = rec.payments.aggregate(s=Sum('amount'))['s'] or Decimal('0')
    adjustments = [{'reason': a.reason, 'amount': str(a.amount)} for a in rec.adjustments.all()]
    # 税额体现：已开票的取记录税额（全额模式自动算/差额模式手填）；
    # 未开票的按项目税率给预估（全额模式），差额模式无法预估标 None
    rate = rec.project.tax_rate or Decimal('0')
    if rec.actual_invoice_amount is not None:
        tax_show = rec.tax_amount
    elif rec.project.invoice_mode == '全额' and rate:
        tax_show = (billable / (1 + rate) * rate).quantize(Decimal('0.01'))
    else:
        tax_show = None
    return {
        'adjustments': adjustments,
        'tax_rate': str(rate),
        'invoice_mode': rec.project.invoice_mode,
        'tax_amount': str(tax_show) if tax_show is not None else None,
        'id': rec.id,
        'short_name': rec.project.short_name,
        'customer_name': rec.project.customer_name,
        'delivery_dept': rec.delivery_dept,
        'operation_date': str(rec.operation_date) if rec.operation_date else None,
        'estimated': str(rec.estimated_amount or 0),
        'diff': str(rec.account_diff_adjustment or 0),
        'billable': str(billable),
        'invoiced': str(rec.actual_invoice_amount) if rec.actual_invoice_amount is not None else None,
        'invoice_room': str(billable - (rec.actual_invoice_amount or Decimal('0'))),
        'invoice_date': str(rec.invoice_date) if rec.invoice_date else None,
        'collected': str(collected),
        'outstanding': str(rec.outstanding_amount or 0),
    }


@csrf_exempt
@pk_required()
def ar_invoice_batch_detail(request, batch_no):
    """GET /records/invoice-batches/<batch_no> — 批次明细：成员逐条 + 汇总。
    应开口径 = 上账金额 + 账实差额调整（差额已按业务要求落在具体记录上）。"""
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    if request.method != 'GET':
        return err('Method not allowed', 405)
    members = [_batch_member_dict(r) for r in _batch_members_qs(request, batch_no)]
    if not members:
        return err('批次不存在或无权访问', 404)
    # 批次回款事件：成员回款中备注带本批次标记的，按(日期+备注)聚合还原为
    # 「某天的一次批次回款」，支持整次撤销
    member_ids = [m['id'] for m in members]
    tag = f'批次回款[{batch_no}]'
    events = {}
    for p in (ARPayment.objects.filter(ar_record_id__in=member_ids,
                                       notes__startswith=tag)
              .order_by('payment_date', 'id')):
        key = (str(p.payment_date), p.notes)
        ev = events.setdefault(key, {'payment_date': str(p.payment_date),
                                     'notes': p.notes, 'total': Decimal('0'),
                                     'count': 0, 'payment_ids': []})
        ev['total'] += p.amount or Decimal('0')
        ev['count'] += 1
        ev['payment_ids'].append(p.id)
    collections = [{**e, 'total': str(e['total'])} for e in events.values()]
    invoice_events = [e.to_dict() for e in
                      BatchInvoiceEvent.objects.filter(batch_no=batch_no)]
    s = lambda k: sum(Decimal(m[k]) for m in members if m[k] is not None)  # noqa: E731
    return ok({
        'batch_no': batch_no,
        'members': members,
        'collections': collections,
        'invoice_events': invoice_events,
        'summary': {
            'count': len(members),
            'estimated': str(s('estimated')),
            'diff': str(s('diff')),
            'billable': str(s('billable')),
            'invoiced': str(s('invoiced')),
            'tax': str(s('tax_amount')),
            'collected': str(s('collected')),
            'outstanding': str(s('outstanding')),
            'invoice_room': str(s('invoice_room')),
            'pending_invoice': sum(1 for m in members if m['invoiced'] is None),
        },
    })


@csrf_exempt
@pk_required()
def ar_invoice_batch_invoice(request, batch_no):
    """POST /records/invoice-batches/<batch_no>/invoice — 批次开票（金额级分批）。

    body: { invoice_date(必填), amount(必填,价税合计), tax_amount(选填,差额模式手填),
            notes(选填) }
    与批次回款同构：每次开票一个事件，金额按运作日期先进先出分摊到成员的
    「可开余额」（上账+差额−已开）。可多次开票，累计不超过批次可开总额即放行。
    税额：全额模式记录随分摊按项目税率自动重算；差额模式记录按分摊占比
    分配手填税额。每个事件可整次撤销。
    """
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    if request.method != 'POST':
        return err('POST only', 405)
    denied = _write_denied(request)
    if denied:
        return denied
    data = _parse_body(request)
    inv_date = _normalize_date(data.get('invoice_date'))
    if not inv_date:
        return err('开票日期无效（格式 2026-01-15）')
    amount = _dec(data.get('amount') or data.get('invoice_total') or 0)
    if amount <= 0:
        return err('开票金额必须大于0（价税合计）')
    tax_raw = data.get('tax_amount')
    manual_tax = _dec(tax_raw) if tax_raw not in (None, '') else None
    if manual_tax is not None and (manual_tax < 0 or manual_tax >= amount):
        return err('税额无效：须 ≥0 且小于开票金额')

    from paikuan.models import PaikuanUser
    user = PaikuanUser.objects.filter(id=request.pk_uid).first()

    allocations = []
    try:
        with transaction.atomic():
            members = list(_batch_members_qs(request, batch_no).select_for_update())
            if not members:
                return err('批次不存在或无权访问', 404)

            def _room(r):
                billable = ((r.estimated_amount or Decimal('0'))
                            + (r.account_diff_adjustment or Decimal('0')))
                return billable - (r.actual_invoice_amount or Decimal('0'))

            open_recs = [r for r in members if _room(r) > 0]
            total_room = sum(_room(r) for r in open_recs)
            if total_room <= 0:
                return err('该批次可开余额已用尽（累计开票已达 上账+差额 合计）')
            if amount > total_room:
                return err(f'开票金额 {amount} 超过批次可开余额 {total_room}'
                           f'（上账+差额合计 − 累计已开）。'
                           f'如金额确实更大，请先在具体记录上追加「差额调整」')

            # 差额模式记录的手填税额按分摊占比分配；先算分摊再分税
            remaining = amount
            plan = []
            for r in open_recs:
                if remaining <= 0:
                    break
                alloc = min(remaining, _room(r))
                plan.append((r, alloc))
                remaining -= alloc
            diff_alloc_total = sum(a for r, a in plan if r.project.invoice_mode == '差额')
            for r, alloc in plan:
                r.actual_invoice_amount = (r.actual_invoice_amount or Decimal('0')) + alloc
                r.invoice_date = inv_date
                if (manual_tax is not None and r.project.invoice_mode == '差额'
                        and diff_alloc_total > 0):
                    share = (manual_tax * alloc / diff_alloc_total).quantize(Decimal('0.01'))
                    r.tax_amount = (r.tax_amount or Decimal('0')) + share
                r.save()   # 全额模式税额由 save() 按税率自动重算
                allocations.append({'record_id': r.id,
                                    'short_name': r.project.short_name,
                                    'operation_date': str(r.operation_date) if r.operation_date else None,
                                    'amount': str(alloc),
                                    'invoiced_after': str(r.actual_invoice_amount)})
            ev = BatchInvoiceEvent.objects.create(
                batch_no=batch_no, invoice_date=inv_date, amount=amount,
                tax_amount=manual_tax, notes=(data.get('notes') or '').strip()[:200],
                allocations=[{'record_id': a['record_id'], 'amount': a['amount']}
                             for a in allocations],
                created_by=user)
    except ValidationError as e:
        return err(str(e.message if hasattr(e, 'message') else e), 400)

    left = total_room - amount
    return ok({
        'event': ev.to_dict(), 'allocations': allocations,
        'message': (f'批次「{batch_no}」开票 {amount} 已按先进先出分摊到 '
                    f'{len(allocations)} 条记录' +
                    (f'，批次剩余可开 {left}' if left > 0 else '，批次已开满')),
    })


@csrf_exempt
@pk_required()
def ar_invoice_batch_invoice_undo(request, batch_no):
    """POST /records/invoice-batches/<batch_no>/invoice-undo — 整次撤销开票事件。

    body: { event_id }。按事件保存的分摊明细逐条回退各记录的开票金额
    （回退后为0则清空开票金额与日期），全额模式税额自动重算，
    差额模式按事件手填税额比例回退。事务整体执行。
    """
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    if request.method != 'POST':
        return err('POST only', 405)
    denied = _write_denied(request)
    if denied:
        return denied
    data = _parse_body(request)
    try:
        ev = BatchInvoiceEvent.objects.get(pk=int(data.get('event_id') or 0),
                                           batch_no=batch_no)
    except (BatchInvoiceEvent.DoesNotExist, ValueError, TypeError):
        return err('开票事件不存在（可能已撤销）', 404)

    member_ids = set(_batch_members_qs(request, batch_no).values_list('id', flat=True))
    if not member_ids:
        return err('批次不存在或无权访问', 404)
    diff_total = sum(Decimal(a['amount']) for a in ev.allocations or [])
    try:
        with transaction.atomic():
            for a in (ev.allocations or []):
                rid, amt = a['record_id'], Decimal(a['amount'])
                if rid not in member_ids:
                    return err(f'记录 #{rid} 不在当前批次作用域内，无法撤销', 403)
                r = ARRecord.objects.select_related('project').select_for_update().get(pk=rid)
                cur = r.actual_invoice_amount or Decimal('0')
                if cur < amt:
                    return err(f'记录 #{rid} 当前开票额 {cur} 小于本事件分摊 {amt}，'
                               f'（可能已被人工修改）请先核对该记录')
                new_val = cur - amt
                if ev.tax_amount is not None and r.project.invoice_mode == '差额' and diff_total > 0:
                    share = (ev.tax_amount * amt / diff_total).quantize(Decimal('0.01'))
                    r.tax_amount = max(Decimal('0'), (r.tax_amount or Decimal('0')) - share)
                if new_val <= 0:
                    r.actual_invoice_amount = None
                    r.invoice_date = None
                    if r.project.invoice_mode == '全额':
                        r.tax_amount = None
                else:
                    r.actual_invoice_amount = new_val
                r.save()
            ev_id = ev.pk
            ev.delete()
    except ARRecord.DoesNotExist:
        return err('分摊涉及的记录已不存在，无法撤销', 404)
    except ValidationError as e:
        return err(str(e.message if hasattr(e, 'message') else e), 400)
    return ok({'undone': ev_id,
               'message': f'已撤销 {ev.invoice_date} 的开票 {ev.amount}，各记录开票额已回退'})


@csrf_exempt
@pk_required()
def ar_invoice_batch_payment(request, batch_no):
    """POST /records/invoice-batches/<batch_no>/payment — 批次回款自动分摊。

    body: { amount(必填), payment_date(必填), notes(选填) }
    一张发票（批次）到一笔钱：按运作日期先进先出，逐条冲减成员未收余额，
    自动生成各记录的回款行（备注带批次号，可追溯）。支持多次回款，
    每次都从当前仍有未收的成员继续分摊。金额超过批次未收合计 → 拒绝并
    提示超出部分走「预收款」。
    """
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    if request.method != 'POST':
        return err('POST only', 405)
    denied = _action_denied(request, 'ar_collect')
    if denied:
        return denied
    denied = _ar_field_denied(request, 'r_payments')
    if denied:
        return denied
    data = _parse_body(request)
    amount = _dec(data.get('amount', 0))
    if amount <= 0:
        return err('回款金额必须大于0')
    pay_date = _normalize_date(data.get('payment_date'))
    if not pay_date:
        return err('回款日期无效（格式 2026-01-20）')
    user_notes = (data.get('notes') or '').strip()

    with transaction.atomic():
        members = list(_batch_members_qs(request, batch_no).select_for_update())
        if not members:
            return err('批次不存在或无权访问', 404)
        open_recs = [r for r in members if (r.outstanding_amount or Decimal('0')) > 0]
        total_outstanding = sum(r.outstanding_amount for r in open_recs)
        if not open_recs:
            return err('该批次已全部回款结清，无未收余额可分摊')
        if amount > total_outstanding:
            return err(f'回款金额 {amount} 超过批次未收合计 {total_outstanding}。'
                       f'请按 {total_outstanding} 录入本批次回款，'
                       f'超出的 {amount - total_outstanding} 元到「预收预付」录入为该客户预收款')

        remaining = amount
        allocations = []
        note = f'批次回款[{batch_no}]' + (f' {user_notes}' if user_notes else '')
        for r in open_recs:
            if remaining <= 0:
                break
            alloc = min(remaining, r.outstanding_amount)
            last = r.payments.order_by('-payment_no').first()
            ARPayment.objects.create(
                ar_record=r, payment_no=(last.payment_no + 1) if last else 1,
                amount=alloc, payment_date=pay_date, notes=note)
            r.refresh_from_db()
            allocations.append({
                'record_id': r.id, 'short_name': r.project.short_name,
                'operation_date': str(r.operation_date) if r.operation_date else None,
                'allocated': str(alloc), 'outstanding_after': str(r.outstanding_amount),
            })
            remaining -= alloc

    settled = sum(1 for a in allocations if Decimal(a['outstanding_after']) <= 0)
    return ok({
        'batch_no': batch_no, 'amount': str(amount), 'allocations': allocations,
        'message': (f'批次「{batch_no}」回款 {amount} 已按运作日期先进先出分摊到 '
                    f'{len(allocations)} 条记录（{settled} 条就此结清）'),
    })


def _gen_batch_no(qs):
    """自动生成可读开票批次号：{客户简称}-{YYMMDD}-{序号}。

    人脑不该发明编码——号段以所选记录的客户为前缀（同客户取其名，多客户记
    「多户」），日期+当日序号保证唯一，看到号就知道是谁、哪天合并的。"""
    names = set()
    for r in qs.select_related('project')[:200]:
        n = (r.project.customer_name or r.project.short_name or '').strip()
        if n:
            names.add(n)
    base = (list(names)[0][:8] if len(names) == 1 else '多户') or '开票'
    prefix = f'{base}-{datetime.date.today().strftime("%y%m%d")}'
    seq = ARRecord.objects.filter(invoice_batch_no__startswith=prefix)\
        .values('invoice_batch_no').distinct().count() + 1
    return f'{prefix}-{seq:02d}'


@csrf_exempt
@pk_required()
def ar_invoice_batch_payment_undo(request, batch_no):
    """POST /records/invoice-batches/<batch_no>/payment-undo — 整次撤销批次回款。

    body: { payment_ids: [int,...] }（来自批次明细 collections 的一次回款事件）
    校验每笔都属于本批次成员且带本批次回款标记，事务内整体删除，
    各记录未收余额随信号恢复——错一笔即整体拒绝，不留半套状态。
    """
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    if request.method != 'POST':
        return err('POST only', 405)
    denied = _action_denied(request, 'ar_collect')
    if denied:
        return denied
    data = _parse_body(request)
    ids = data.get('payment_ids') or []
    if not isinstance(ids, list) or not ids:
        return err('请提供要撤销的回款 payment_ids')
    try:
        ids = [int(i) for i in ids]
    except (TypeError, ValueError):
        return err('payment_ids 必须为整数列表')

    member_ids = list(_batch_members_qs(request, batch_no).values_list('id', flat=True))
    if not member_ids:
        return err('批次不存在或无权访问', 404)
    tag = f'批次回款[{batch_no}]'
    pays = list(ARPayment.objects.filter(pk__in=ids))
    if len(pays) != len(set(ids)):
        return err('部分回款记录不存在（可能已被撤销）', 404)
    for p in pays:
        if p.ar_record_id not in member_ids:
            return err(f'回款 #{p.id} 不属于批次「{batch_no}」的成员记录，已拒绝')
        if not (p.notes or '').startswith(tag):
            return err(f'回款 #{p.id} 不是本批次的批次回款（{p.notes or "无标记"}），'
                       f'请到该应收记录上单独处理')
        if p.source == '预收抵扣':
            return err(f'回款 #{p.id} 为预收抵扣，请走「撤销核销」')
    total = sum(p.amount or Decimal('0') for p in pays)
    with transaction.atomic():
        ARPayment.objects.filter(pk__in=ids).delete()
    return ok({'undone': len(pays), 'total': str(total),
               'message': f'已撤销该次批次回款：{len(pays)} 笔合计 {total}，各记录未收已恢复'})


@csrf_exempt
@pk_required()
def ar_records_batch_assign(request):
    """POST /records/batch-assign  —  批量设置 invoice_batch_no。

    Body: { "ids": [1,2,3], "invoice_batch_no": "..." } 或 { "ids": [...], "auto": true }
    auto=true 时系统自动生成可读批次号（客户简称-日期-序号），无需人为编码。
    ids 为空 + all=true 时：对当前筛选全集打标（与 bulk-delete 对齐，但限 5000 条）。
    invoice_batch_no 传空字符串且非 auto 则清空批次（取消合并）。
    """
    denied = _page_denied(request, 'ar_records')
    if denied:
        return denied
    denied = _write_denied(request)
    if denied:
        return denied
    if request.method != 'POST':
        return err('POST only', 405)

    data = _parse_body(request)
    batch_no = (data.get('invoice_batch_no') or '').strip()[:50]
    auto = data.get('auto') in (True, 'true', '1')
    all_flag = data.get('all') in (True, 'true', '1')
    ids = data.get('ids')

    if all_flag:
        today = datetime.date.today()
        qs = _ar_dept_filter(ARRecord.objects.all(), request, shared_field='project__is_shared')
        qs = _apply_record_filters(qs, request)
        qs = _apply_record_state_filters(qs, request, today)
        qs = _apply_conditions(qs, request, today)
        if qs.count() > 5000:
            return err('选中记录超过5000条，请缩小筛选范围后再批量操作')
    else:
        if not ids or not isinstance(ids, list):
            return err('请传入 ids 数组或 all=true')
        # 权限：只允许操作自己部门的记录
        qs = ARRecord.objects.filter(pk__in=[int(i) for i in ids])
        if request.pk_role != 'super_admin':
            qs = qs.filter(delivery_dept__in=request.pk_depts)

    if auto and not batch_no:
        batch_no = _gen_batch_no(qs)
    updated = qs.update(invoice_batch_no=batch_no)

    action = f'设置批次号为「{batch_no}」' if batch_no else '清空批次号'
    return ok({'updated': updated, 'invoice_batch_no': batch_no,
               'message': f'{action}，共更新 {updated} 条记录'})


# ──────────────────────────────────────────────────────────────────────────────
# 客户管理 (customers)
# ──────────────────────────────────────────────────────────────────────────────

@csrf_exempt
@pk_required()
def customers(request):
    """GET  /ar/customers   — 分页列表（with_stats=1 附带应收聚合）
    POST /ar/customers   — 新增客户
    """
    # 客户管理隶属「项目台账」模块，与合同接口一致按 ar_projects 页面权限把关。
    # 认证由 @pk_required() 保证；部门隔离由下方 dept 过滤 / 写入由 _write_denied 处理。
    denied = _page_denied(request, 'ar_projects')
    if denied:
        return denied

    if request.method == 'GET':
        from django.db.models.functions import Coalesce
        from django.db.models import Exists, OuterRef
        today = datetime.date.today()
        # 客户按事业部隔离 + 顶部全局范围（?depts）一并生效，与项目/应收口径一致；
        # _ar_dept_filter 对无授权部门的用户返回空集（杜绝「无部门=看全部」漏洞）。
        qs = _ar_dept_filter(Customer.objects.all(), request, dept_field='delivery_dept')
        # 共享业务岗位（ar_shared_only，如销售BP）：仅纳入有共享项目的客户，
        # 且其应收聚合仅统计共享项目，避免经客户视图泄露自营业务财务。
        shared_only = False
        if request.pk_role != 'super_admin':
            _perms = get_request_perms(request)
            shared_only = bool(_perms and _perms.get('ar_shared_only'))
        if shared_only:
            qs = qs.filter(Exists(
                ARProject.objects.filter(customer_id=OuterRef('pk'), is_shared=True)))
        q = request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(contact__icontains=q) | Q(notes__icontains=q))
        level = request.GET.get('level', '').strip()
        if level:
            qs = qs.filter(level=level)
        status_f = request.GET.get('status', '').strip()
        if status_f:
            qs = qs.filter(status=status_f)
        # 事业部筛选：直接按客户自身的交付部门
        dept = request.GET.get('dept', '').strip()
        if dept:
            qs = qs.filter(delivery_dept=dept)

        # 数据库级聚合应收（客户已按事业部隔离 → 其全部应收即本部门，无需逐条过滤部门）
        # shared_only 岗位：所有计数/求和叠加「项目为共享业务」过滤。
        _base = Q(projects__is_shared=True) if shared_only else None
        def _money(extra=None):
            f = _base
            if extra is not None:
                f = (f & extra) if f is not None else extra
            kw = {'filter': f} if f is not None else {}
            return Coalesce(Sum('projects__ar_records__outstanding_amount', **kw),
                            Value(0), output_field=DecimalField())
        _inv_kw = {'filter': _base} if _base is not None else {}
        qs = qs.annotate(
            s_project_count=Count('projects', filter=_base, distinct=True),
            s_invoiced=Coalesce(Sum('projects__ar_records__actual_invoice_amount', **_inv_kw),
                                Value(0), output_field=DecimalField()),
            s_outstanding=_money(),
            s_overdue=_money(Q(projects__ar_records__due_date__lt=today,
                              projects__ar_records__outstanding_amount__gt=0)),
        )

        # 数据库级排序：跨所有分页生效，不再分页时临时算
        SORT_MAP = {
            'outstanding': 's_outstanding', 'overdue': 's_overdue', 'invoiced': 's_invoiced',
            'project_count': 's_project_count', 'name': 'name', 'level': 'level',
            'customer_date': 'customer_date', 'created_at': 'created_at',
        }
        sort = SORT_MAP.get(request.GET.get('sort', 'outstanding'), 's_outstanding')
        direction = '' if request.GET.get('dir', 'desc') == 'asc' else '-'
        qs = qs.order_by(direction + sort, 'name')

        page = max(1, int(request.GET.get('page', 1) or 1))
        size = min(200, max(1, int(request.GET.get('size', 50) or 50)))
        total = qs.count()
        page_customers = list(qs[(page - 1) * size: page * size])
        rows = [{
            'id': c.id, 'name': c.name, 'delivery_dept': c.delivery_dept,
            'level': c.level, 'status': c.status, 'contact': c.contact,
            'customer_date': str(c.customer_date) if c.customer_date else None,
            'notes': c.notes,
            'created_at': c.created_at.isoformat() if c.created_at else None,
            'updated_at': c.updated_at.isoformat() if c.updated_at else None,
            'project_count': c.s_project_count,
            'invoiced': float(c.s_invoiced),
            'outstanding': float(c.s_outstanding),
            'overdue': float(c.s_overdue),
        } for c in page_customers]
        return ok({'items': rows, 'total': total, 'page': page, 'size': size})

    if request.method == 'POST':
        denied = _write_denied(request)
        if denied:
            return denied
        data = _parse_body(request)
        name = (data.get('name') or '').strip()
        if not name:
            return err('客户名称不能为空')
        dept = (data.get('delivery_dept') or '').strip()
        # 客户按事业部隔离 → 新建必须指定事业部，杜绝再产生无部门孤儿客户
        if not dept:
            return err('请选择客户所属事业部（客户按事业部隔离）')
        if dept not in VALID_DEPARTMENTS:
            return err(f'无效交付部门「{dept}」')
        if request.pk_role != 'super_admin' and dept not in request.pk_depts:
            return err(f'无权在事业部「{dept}」下新建客户')
        # 客户按 (名称 + 事业部) 隔离：同名客户在不同部门可各建一个
        if Customer.objects.filter(name=name, delivery_dept=dept).exists():
            return err(f'客户「{name}」在「{dept or "未指定部门"}」已存在')
        st = (data.get('status') or '运作中').strip()
        if st not in dict(Customer.STATUS_CHOICES):
            st = '运作中'
        c = Customer(
            name=name,
            delivery_dept=dept,
            level=(data.get('level') or '').strip(),
            status=st,
            contact=(data.get('contact') or '').strip(),
            customer_date=_normalize_date(data.get('customer_date')) or None,
            notes=(data.get('notes') or '').strip(),
        )
        c.save()
        return ok(c.to_dict())

    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def customers_bulk_tag_level(request):
    """POST /customers/bulk-tag-level  —  批量为客户打/改等级。

    Body: { "ids": [1,2,3], "level": "A级" }
    或   : { "all": true, "level": "A级", "q": "...", "level_filter": "..." }
      all=true 时对当前筛选全集（搜索 q + 当前等级 level_filter）打标，限 5000 条。
    level 传空字符串 = 清空等级（取消分级）。
    """
    denied = _page_denied(request, 'ar_projects') or _write_denied(request)
    if denied:
        return denied
    if request.method != 'POST':
        return err('POST only', 405)
    # 部门隔离由下方 dept 过滤处理。
    # （历史遗留的 'dept_admin'/'user' 并非真实角色，会误伤所有非超管用户）

    data = _parse_body(request)
    level = (data.get('level') or '').strip()
    if level and level not in VALID_CUSTOMER_LEVELS:
        return err(f'无效客户等级「{level}」，可选：{"、".join(VALID_CUSTOMER_LEVELS)}')

    all_flag = data.get('all') in (True, 'true', '1')
    ids = data.get('ids')

    if all_flag:
        # 部门隔离：非超管只能批量操作自己有权部门的客户（无授权部门→空集）
        qs = _ar_dept_filter(Customer.objects.all(), request, dept_field='delivery_dept')
        q = (data.get('q') or '').strip()
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(contact__icontains=q) | Q(notes__icontains=q))
        level_filter = (data.get('level_filter') or '').strip()
        if level_filter:
            qs = qs.filter(level=level_filter)
        dept_filter = (data.get('dept') or '').strip()
        if dept_filter:
            qs = qs.filter(delivery_dept=dept_filter)
        cnt = qs.count()
        if cnt == 0:
            return err('当前筛选没有匹配的客户')
        if cnt > 5000:
            return err('选中客户超过5000个，请缩小筛选范围后再批量操作')
        id_list = list(qs.values_list('id', flat=True))
        updated = qs.update(level=level)
    else:
        if not ids or not isinstance(ids, list):
            return err('请传入 ids 数组或 all=true')
        try:
            id_list = [int(i) for i in ids]
        except (TypeError, ValueError):
            return err('ids 必须为整数数组')
        # 部门隔离：仅保留用户有权部门的客户（无授权部门→空集，杜绝按 id 跨部门越权）
        cqs = _ar_dept_filter(
            Customer.objects.filter(pk__in=id_list), request, dept_field='delivery_dept')
        id_list = list(cqs.values_list('id', flat=True))
        updated = cqs.update(level=level)
    # 以客户为准：客户等级变更后，同步镜像到其名下所有项目
    ARProject.objects.filter(customer_id__in=id_list).update(customer_level=level)

    action = f'设置等级为「{level}」' if level else '清空等级'
    return ok({'updated': updated, 'level': level,
               'message': f'{action}，共更新 {updated} 个客户'})


def _reconcile_customer_levels(customer_ids):
    """以客户为准对齐「客户等级 ↔ 项目等级」。
    每个客户取规范等级 = 客户已有等级；客户为空则取其项目中出现最多的非空等级；
    再把规范等级回写客户、镜像到该客户全部项目。返回（对齐的客户数）。"""
    from collections import Counter
    aligned = 0
    for c in Customer.objects.filter(id__in=list(customer_ids)).iterator():
        proj_levels = [(pl or '').strip() for pl in
                       ARProject.objects.filter(customer_id=c.id).values_list('customer_level', flat=True)]
        canonical = (c.level or '').strip()
        if not canonical:
            nonempty = [pl for pl in proj_levels if pl]
            if nonempty:
                canonical = Counter(nonempty).most_common(1)[0][0]
        if not canonical:
            continue
        touched = False
        if (c.level or '').strip() != canonical:
            c.level = canonical
            c.save(update_fields=['level', 'updated_at'])
            touched = True
        n = (ARProject.objects.filter(customer_id=c.id)
             .exclude(customer_level=canonical).update(customer_level=canonical))
        if touched or n:
            aligned += 1
    return aligned


@csrf_exempt
@pk_required()
def customers_sync_from_projects(request):
    """POST /customers/sync-from-projects — 从项目台账回填客户名单 + 对齐等级。

    历史项目（在「客户正名」之前导入的）可能有客户名称但未挂客户实体。
    本接口：①扫描有客户名称却未挂客户的项目，按唯一客户名 get_or_create 客户并挂接；
    ②以客户为准对齐「客户等级↔项目等级」（同一客户所有项目等级统一）。
    一个客户对应多个项目（不合并金额、不动应收）。幂等：重复点击安全。
    部门权限：仅同步当前用户有权部门的项目。
    """
    denied = _page_denied(request, 'ar_projects') or _write_denied(request)
    if denied:
        return denied
    if request.method != 'POST':
        return err('POST only', 405)
    # 部门隔离由下方 dept 过滤处理。
    # （历史遗留的 'dept_admin'/'user' 并非真实角色，会误伤所有非超管用户）

    qs = _ar_dept_filter(
        ARProject.objects.filter(customer__isnull=True).exclude(customer_name=''), request)
    created = linked = 0
    for p in qs.iterator():
        name = (p.customer_name or '').strip()
        if not name:
            continue
        cust, was_created = Customer.objects.get_or_create(
            name=name, delivery_dept=p.delivery_dept or '',
            defaults={'level': p.customer_level or ''})
        if was_created:
            created += 1
        # 直接 update 挂接，避开 save 的派生重算副作用
        ARProject.objects.filter(pk=p.pk).update(customer=cust)
        linked += 1

    # 对齐等级：本用户有权部门项目所涉及的全部客户
    scoped_proj = _ar_dept_filter(ARProject.objects.filter(customer__isnull=False), request)
    cust_ids = set(scoped_proj.values_list('customer_id', flat=True))
    aligned = _reconcile_customer_levels(cust_ids)

    # 清理孤儿客户：名下 0 项目且无合同关联（项目台账已删但名单残留）。
    # 超管清理全局；事业部用户仅清理本部门的孤儿，避免误删他人数据。
    orphan_qs = Customer.objects.annotate(
        _np=Count('projects', distinct=True),
        _nc=Count('contract_links', distinct=True)).filter(_np=0, _nc=0)
    if request.pk_role != 'super_admin':
        orphan_qs = orphan_qs.filter(delivery_dept__in=(request.pk_depts or []))
    purged = orphan_qs.count()
    orphan_qs.delete()

    parts = []
    if created or linked:
        parts.append(f'新建 {created} 个客户、挂接 {linked} 个项目')
    if aligned:
        parts.append(f'对齐 {aligned} 个客户的等级')
    if purged:
        parts.append(f'清理 {purged} 个名下无项目的孤儿客户')
    msg = ('同步完成：' + '；'.join(parts)) if parts else '客户名单与等级已是最新，无需处理'
    return ok({'created_customers': created, 'linked_projects': linked,
               'aligned_levels': aligned, 'purged_orphans': purged, 'message': msg})


@csrf_exempt
@pk_required()
def customer_detail(request, pk):
    """GET /PUT /DELETE  /ar/customers/<pk>"""
    denied = _page_denied(request, 'ar_projects')
    if denied:
        return denied
    try:
        c = Customer.objects.get(pk=pk)
    except Customer.DoesNotExist:
        return err('客户不存在', 404)
    # 客户按事业部隔离：非超管只能操作自己有权部门的客户（无部门孤儿放行）
    if (request.pk_role != 'super_admin' and c.delivery_dept
            and c.delivery_dept not in request.pk_depts):
        return err('无权访问该客户', 403)

    if request.method == 'GET':
        d = c.to_dict()
        today = datetime.date.today()
        # 该客户的项目 + 每个项目的应收聚合（部门作用域；ar_shared_only 仅见共享业务）
        proj_qs = _ar_dept_filter(
            ARProject.objects.filter(customer_id=pk), request,
            shared_field='is_shared').order_by('short_name')
        proj_ids = [p.id for p in proj_qs]
        rec_qs = _ar_dept_filter(
            ARRecord.objects.filter(project_id__in=proj_ids), request)
        per_proj = {pid: {'invoiced': 0.0, 'outstanding': 0.0, 'overdue': 0.0, 'records': 0}
                    for pid in proj_ids}
        for r in rec_qs.values('project_id').annotate(
                inv=Sum('actual_invoice_amount'), out_amt=Sum('outstanding_amount'),
                cnt=Count('id')):
            pp = per_proj[r['project_id']]
            pp['invoiced'] = float(r['inv'] or 0)
            pp['outstanding'] = float(r['out_amt'] or 0)
            pp['records'] = r['cnt']
        for r in (rec_qs.filter(due_date__lt=today, outstanding_amount__gt=0)
                  .values('project_id').annotate(ov=Sum('outstanding_amount'))):
            per_proj[r['project_id']]['overdue'] = float(r['ov'] or 0)
        projects = []
        for p in proj_qs:
            pp = per_proj.get(p.id, {})
            projects.append({
                'id': p.id, 'project_no': p.project_no,
                'short_name': p.short_name, 'customer_name': p.customer_name,
                'delivery_dept': p.delivery_dept, 'business_mode': p.business_mode,
                'status': p.status,
                'invoiced': pp.get('invoiced', 0.0), 'outstanding': pp.get('outstanding', 0.0),
                'overdue': pp.get('overdue', 0.0), 'records': pp.get('records', 0),
            })
        d['projects'] = projects
        d['stats'] = {
            'project_count': len(projects),
            'invoiced': round(sum(p['invoiced'] for p in projects), 2),
            'outstanding': round(sum(p['outstanding'] for p in projects), 2),
            'overdue': round(sum(p['overdue'] for p in projects), 2),
        }
        d['stats']['collected'] = round(d['stats']['invoiced'] - d['stats']['outstanding'], 2)
        return ok(d)

    if request.method == 'PUT':
        denied = _write_denied(request)
        if denied:
            return denied
        data = _parse_body(request)
        name_changed = False
        if 'name' in data:
            name = (data['name'] or '').strip()
            if not name:
                return err('客户名称不能为空')
            if Customer.objects.filter(name=name, delivery_dept=c.delivery_dept).exclude(pk=pk).exists():
                return err(f'客户名称「{name}」在「{c.delivery_dept or "未指定部门"}」已被占用')
            name_changed = (c.name != name)
            c.name = name
        level_changed = 'level' in data
        for field in ('level', 'contact', 'notes'):
            if field in data:
                setattr(c, field, (data[field] or '').strip())
        if 'status' in data:
            st = (data['status'] or '').strip()
            if st not in dict(Customer.STATUS_CHOICES):
                return err(f'无效客户状态「{st}」，可选：运作中/中断/结束')
            c.status = st
        if 'customer_date' in data:
            c.customer_date = _normalize_date(data['customer_date']) or None
        c.save()
        # 以客户为准：等级变更同步镜像到该客户名下所有项目
        if level_changed:
            ARProject.objects.filter(customer_id=c.id).update(customer_level=c.level)
        # 客户改名 → 同步名下项目的 customer_name 文本字段，保持两边一致
        if name_changed:
            ARProject.objects.filter(customer_id=c.id).update(customer_name=c.name)
        # 一键下发：把客户状态应用到名下所有项目（项目独立状态，仅在显式请求时下发）
        if data.get('push_status'):
            ARProject.objects.filter(customer_id=c.id).update(status=c.status)
        return ok(c.to_dict())

    if request.method == 'DELETE':
        denied = _delete_denied(request)
        if denied:
            return denied
        # 名下还有项目时禁删：否则项目 customer_id 置空、客户名文本悬空，
        # 下次保存项目会按同名重建客户，等级真相源断链
        linked = ARProject.objects.filter(customer_id=c.pk).count()
        if linked:
            return err(f'客户「{c.name}」名下还有 {linked} 个项目，不能删除；'
                       f'请先删除或改挂这些项目（项目台账中修改客户名称即可改挂）')
        c.delete()
        return ok({'deleted': pk})

    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def customers_bulk_delete(request):
    """POST /ar/customers/bulk-delete  body {ids:[int,...]}
    批量删除客户。与单删同口径的保护：名下仍有项目或有合同关联的客户跳过不删，
    并在返回中逐个说明原因；部门作用域与删除权限约束同列表。单次上限 1000。"""
    denied = _page_denied(request, 'ar_projects')
    if denied:
        return denied
    if request.method != 'POST':
        return err('POST only', 405)
    denied = _delete_denied(request)
    if denied:
        return denied

    body = _parse_body(request)
    ids = body.get('ids') or []
    if not isinstance(ids, list) or not ids:
        return err('请提供要删除的客户 ids')
    try:
        ids = [int(i) for i in ids]
    except (ValueError, TypeError):
        return err('ids 必须为整数列表')
    if len(ids) > 1000:
        return err('单次删除上限 1000 个，请分批操作')

    qs = _ar_dept_filter(Customer.objects.filter(pk__in=ids), request,
                         dept_field='delivery_dept')
    qs = qs.annotate(_np=Count('projects', distinct=True),
                     _nc=Count('contract_links', distinct=True))
    deletable, skipped = [], []
    for c in qs:
        if c._np:
            skipped.append(f'「{c.name}」名下还有 {c._np} 个项目，未删除')
        elif c._nc:
            skipped.append(f'「{c.name}」有 {c._nc} 个合同关联，未删除')
        else:
            deletable.append(c.pk)
    if deletable:
        Customer.objects.filter(pk__in=deletable).delete()
    msg = f'已删除 {len(deletable)} 个客户'
    if skipped:
        msg += f'，{len(skipped)} 个被保护跳过'
    return ok({'deleted': len(deletable), 'skipped': len(skipped),
               'skipped_reasons': skipped[:50], 'message': msg})


# ──────────────────────────────────────────────────────────────────────────────
# 草稿项目（待完善）
# ──────────────────────────────────────────────────────────────────────────────

@csrf_exempt
@pk_required()
def project_drafts(request):
    """GET /ar/projects/drafts  —  列出所有 is_draft=True 的项目，供人工完善。"""
    denied = _page_denied(request, 'ar_projects')
    if denied:
        return denied
    if request.method != 'GET':
        return err('Method not allowed', 405)

    qs = _ar_dept_filter(ARProject.objects.filter(is_draft=True), request)
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(Q(short_name__icontains=q) | Q(customer_name__icontains=q))
    dept = request.GET.get('dept', '').strip()
    if dept:
        qs = qs.filter(delivery_dept=dept)

    page = max(1, int(request.GET.get('page', 1) or 1))
    size = min(200, max(1, int(request.GET.get('size', 50) or 50)))
    total = qs.count()
    perms = get_request_perms(request)
    rows = [apply_ar_view_mask(p.to_dict(), perms, 'project')
            for p in qs.order_by('-id')[(page - 1) * size: page * size]]
    return ok({'items': rows, 'total': total, 'page': page, 'size': size})


@csrf_exempt
@pk_required()
def project_drafts_promote(request):
    """POST /ar/projects/drafts/promote  —  把待完善草稿「转正」为正式项目并同步客户。

    草稿是早期应收导入「找不到项目时自动新建」遗留的占位项目（新版已停用自动建草稿）。
    转正：把信息齐全（有交付部门）的草稿 is_draft 置 False，保存即按 (客户名+部门)
    自动挂接/正名客户——既保留应收数据、又把无部门客户收编进正式客户名单。
    交付部门为空的草稿无法自动判定归属，列出供人工到台账补全部门后再转。
    preview=1 只返回数量分布，不改动。
    """
    denied = _page_denied(request, 'ar_projects')
    if denied:
        return denied
    denied = _write_denied(request)
    if denied:
        return denied
    if request.method != 'POST':
        return err('POST only', 405)

    base = _ar_dept_filter(ARProject.objects.filter(is_draft=True), request)
    total = base.count()
    ready = base.exclude(delivery_dept='')          # 有部门 → 可转正
    need_dept = list(base.filter(delivery_dept='')   # 无部门 → 需人工补
                     .values_list('short_name', flat=True)[:50])

    data = _parse_body(request)
    if request.GET.get('preview') or data.get('preview'):
        return ok({'total': total, 'ready': ready.count(), 'need_dept': len(need_dept)})

    promoted = 0
    with transaction.atomic():
        for p in ready:
            p.is_draft = False
            p.save()     # 触发 _autolink_customer：按 (客户名+部门) 挂客户
            promoted += 1

    msg = f'已转正 {promoted} 个草稿为正式项目并同步客户' if promoted else '没有可直接转正的草稿'
    if need_dept:
        msg += f'；另有 {len(need_dept)} 个草稿缺「交付部门」无法自动转正，请到项目台账补部门后再转（或删除）'
    return ok({'promoted': promoted, 'need_dept_count': len(need_dept),
               'need_dept_samples': need_dept[:10], 'message': msg})


# ──────────────────────────────────────────────────────────────────────────────
# 合同管理 (contracts) — 合同实体 + 多对多客户/项目关联
# ──────────────────────────────────────────────────────────────────────────────

def _sync_contract_links(ct, data, request):
    """按请求体同步合同的客户(parties)与项目(projects)关联。
    仅当请求体包含对应键时才替换该侧关联（PATCH 语义，未传则不动）。"""
    # 客户：parties=[{customer_id, role, share}] 或 customer_ids=[id]
    if 'parties' in data or 'customer_ids' in data:
        ct.parties.all().delete()
        parties = data.get('parties')
        if parties is None:
            parties = [{'customer_id': cid, 'role': 'main'} for cid in (data.get('customer_ids') or [])]
        seen = set()
        for p in parties:
            cid = p.get('customer_id')
            if not cid or cid in seen or not Customer.objects.filter(pk=cid).exists():
                continue
            seen.add(cid)
            share = p.get('share')
            ContractParty.objects.create(
                contract=ct, customer_id=cid,
                role=p.get('role') if p.get('role') in ('main', 'sub') else 'main',
                share=_dec(share) if share not in (None, '') else None)
    # 项目：projects=[{project_id, is_primary}] 或 project_ids=[id]
    if 'projects' in data or 'project_ids' in data:
        ct.project_links.all().delete()
        projs = data.get('projects')
        if projs is None:
            projs = [{'project_id': pid, 'is_primary': True} for pid in (data.get('project_ids') or [])]
        seen = set()
        for pr in projs:
            pid = pr.get('project_id')
            if not pid or pid in seen:
                continue
            proj = ARProject.objects.filter(pk=pid).first()
            if not proj:
                continue
            # 只允许关联自己有权限部门的项目
            if request.pk_role != 'super_admin' and proj.delivery_dept not in request.pk_depts:
                continue
            seen.add(pid)
            ContractProject.objects.create(
                contract=ct, project=proj, is_primary=bool(pr.get('is_primary', True)))


@csrf_exempt
@pk_required()
def contracts(request):
    """GET  /ar/contracts  — 合同列表（按部门作用域 + q/dept 筛选）
    POST /ar/contracts  — 新增合同（可同时传 parties / project_ids 建立关联）
    """
    denied = _page_denied(request, 'ar_projects')
    if denied:
        return denied

    if request.method == 'GET':
        qs = _ar_dept_filter(Contract.objects.all(), request, dept_field='delivery_dept')
        q = request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(contract_no__icontains=q))
        dept = request.GET.get('dept', '').strip()
        if dept:
            qs = qs.filter(delivery_dept=dept)
        page = max(1, int(request.GET.get('page', 1) or 1))
        size = min(200, max(1, int(request.GET.get('size', 50) or 50)))
        total = qs.count()
        rows = [c.to_dict() for c in qs[(page - 1) * size: page * size]]
        return ok({'items': rows, 'total': total, 'page': page, 'size': size})

    if request.method == 'POST':
        denied = _write_denied(request)
        if denied:
            return denied
        data = _parse_body(request)
        name = (data.get('name') or '').strip()
        if not name:
            return err('合同名称不能为空')
        dept = (data.get('delivery_dept') or '').strip()
        if dept:
            if dept not in VALID_DEPARTMENTS:
                return err(f'无效交付部门: {dept}')
            denied = _dept_denied(request, dept, '无权操作此部门')
            if denied:
                return denied
        amount = data.get('amount')
        with transaction.atomic():
            ct = Contract(
                name=name,
                contract_no=(data.get('contract_no') or '').strip(),
                delivery_dept=dept,
                sign_date=_normalize_date(data.get('sign_date')) or None,
                amount=_dec(amount) if amount not in (None, '') else None,
                notes=(data.get('notes') or '').strip(),
            )
            ct.save()
            _sync_contract_links(ct, data, request)
        return ok(ct.to_dict(with_links=True))

    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def contract_detail(request, pk):
    """GET /PUT /DELETE  /ar/contracts/<pk>（含 parties/projects 关联维护）"""
    denied = _page_denied(request, 'ar_projects')
    if denied:
        return denied
    try:
        ct = Contract.objects.get(pk=pk)
    except Contract.DoesNotExist:
        return err('合同不存在', 404)
    # 部门访问控制
    if request.pk_role != 'super_admin' and ct.delivery_dept and ct.delivery_dept not in request.pk_depts:
        return err('无权访问', 403)

    if request.method == 'GET':
        return ok(ct.to_dict(with_links=True))

    if request.method == 'PUT':
        denied = _write_denied(request)
        if denied:
            return denied
        data = _parse_body(request)
        if 'name' in data:
            name = (data['name'] or '').strip()
            if not name:
                return err('合同名称不能为空')
            ct.name = name
        if 'contract_no' in data:
            ct.contract_no = (data['contract_no'] or '').strip()
        if 'delivery_dept' in data:
            dept = (data['delivery_dept'] or '').strip()
            if dept and dept not in VALID_DEPARTMENTS:
                return err(f'无效交付部门: {dept}')
            if dept:
                denied = _dept_denied(request, dept, '无权操作目标部门')
                if denied:
                    return denied
            ct.delivery_dept = dept
        if 'sign_date' in data:
            ct.sign_date = _normalize_date(data['sign_date']) or None
        if 'amount' in data:
            amt = data['amount']
            ct.amount = _dec(amt) if amt not in (None, '') else None
        if 'notes' in data:
            ct.notes = (data['notes'] or '').strip()
        with transaction.atomic():
            ct.save()
            _sync_contract_links(ct, data, request)
        return ok(ct.to_dict(with_links=True))

    if request.method == 'DELETE':
        denied = _delete_denied(request)
        if denied:
            return denied
        ct.delete()   # 级联删除 parties / project_links（不删客户与项目本体）
        return ok({'deleted': pk})

    return err('Method not allowed', 405)


# ─────────────────────────────────────────────────────────────────────────────
# P4 决策闭环 — 行动项 (Action Items)
# ─────────────────────────────────────────────────────────────────────────────

@csrf_exempt
@pk_required()
def ar_actions(request):
    """GET: list action items with counts; POST: create one."""
    # dept-scope filter helper（bu='' 为全局事项，对所有人可见；
    # 无授权部门的非超管只看全局事项，杜绝「空部门=看全部」越权）
    def _scope_qs(qs):
        if request.pk_role == 'super_admin':
            return qs
        return qs.filter(Q(bu='') | Q(bu__in=request.pk_depts or []))

    if request.method == 'GET':
        qs = _scope_qs(ActionItem.objects.all())
        status_f = request.GET.get('status')
        bu_f = request.GET.get('bu')
        priority_f = request.GET.get('priority')
        if status_f:
            qs = qs.filter(status=status_f)
        if bu_f:
            qs = qs.filter(bu=bu_f)
        if priority_f:
            qs = qs.filter(priority=priority_f)

        items = [i.to_dict() for i in qs.select_related('created_by', 'resolved_by')[:300]]

        all_qs = _scope_qs(ActionItem.objects.all())
        counts = {s: all_qs.filter(status=s).count()
                  for s in ('open', 'in_progress', 'done', 'dismissed')}
        return ok({'items': items, 'counts': counts})

    if request.method == 'POST':
        denied = _write_denied(request)
        if denied:
            return denied
        data = _parse_body(request)
        title = (data.get('title') or '').strip()
        if not title:
            return err('标题不能为空')
        item = ActionItem(
            title=title,
            description=(data.get('description') or '').strip(),
            bu=(data.get('bu') or '').strip(),
            category=(data.get('category') or '').strip(),
            priority=data.get('priority', 'medium'),
            assignee=(data.get('assignee') or '').strip(),
            source_signal=data.get('source_signal') or {},
        )
        if data.get('due_date'):
            item.due_date = _normalize_date(data['due_date'])
        if hasattr(request, 'pk_user') and request.pk_user:
            item.created_by = request.pk_user
        item.save()
        return ok(item.to_dict())

    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def ar_action_detail(request, pk):
    """GET/PUT/DELETE single action item."""
    try:
        item = ActionItem.objects.select_related('created_by', 'resolved_by').get(pk=pk)
    except ActionItem.DoesNotExist:
        return err('行动项不存在', 404)

    # 与列表口径一致：bu='' 为全局事项；指定 bu 的事项仅本部门（或超管）可见可改
    if (request.pk_role != 'super_admin' and item.bu
            and item.bu not in (request.pk_depts or [])):
        return err('无权访问', 403, 403)

    if request.method == 'GET':
        return ok(item.to_dict())

    if request.method == 'PUT':
        denied = _write_denied(request)
        if denied:
            return denied
        data = _parse_body(request)
        if 'title' in data:
            t = (data['title'] or '').strip()
            if t:
                item.title = t
        if 'description' in data:
            item.description = (data['description'] or '').strip()
        if 'bu' in data:
            item.bu = (data['bu'] or '').strip()
        if 'priority' in data and data['priority'] in ('high', 'medium', 'low'):
            item.priority = data['priority']
        if 'assignee' in data:
            item.assignee = (data['assignee'] or '').strip()
        if 'due_date' in data:
            item.due_date = _normalize_date(data['due_date']) if data['due_date'] else None
        if 'status' in data:
            new_st = data['status']
            if new_st in ('open', 'in_progress', 'done', 'dismissed'):
                old_st = item.status
                item.status = new_st
                if new_st == 'done' and old_st != 'done':
                    item.resolved_at = datetime.datetime.now(datetime.timezone.utc)
                    if hasattr(request, 'pk_user') and request.pk_user:
                        item.resolved_by = request.pk_user
                elif new_st != 'done':
                    item.resolved_at = None
                    item.resolved_by = None
        item.save()
        return ok(item.to_dict())

    if request.method == 'DELETE':
        denied = _delete_denied(request)
        if denied:
            return denied
        item.delete()
        return ok({'deleted': pk})

    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def ar_actions_from_signal(request):
    """POST: auto-create an action item from a today-signal dict.
    Deduplicates: if open/in_progress item exists with same category+bu, returns it.
    """
    if request.method != 'POST':
        return err('Method not allowed', 405)
    denied = _write_denied(request)
    if denied:
        return denied

    data = _parse_body(request)
    signal = data.get('signal') or {}
    category = (signal.get('type') or signal.get('category') or 'general').strip()
    bu = (signal.get('bu') or '').strip()

    # Deduplicate
    existing = ActionItem.objects.filter(
        category=category, bu=bu, status__in=['open', 'in_progress']
    ).first()
    if existing:
        return ok({'item': existing.to_dict(), 'created': False, 'msg': '已有同类待处理行动项'})

    priority_map = {
        'critical': 'high', 'overdue': 'high', 'cash_risk': 'high',
        'low_margin': 'medium', 'forecast': 'medium', 'baddebt': 'high',
    }
    level = signal.get('level') or signal.get('type') or ''
    priority = priority_map.get(level, 'medium')
    title = (signal.get('title') or signal.get('label') or '待处理事项').strip()

    item = ActionItem(
        title=title,
        description=(signal.get('desc') or signal.get('description') or '').strip(),
        bu=bu,
        category=category,
        priority=priority,
        source_signal=signal,
    )
    if hasattr(request, 'pk_user') and request.pk_user:
        item.created_by = request.pk_user
    item.save()
    return ok({'item': item.to_dict(), 'created': True})


# ─────────────────────────────────────────────────────────────────────────────
# P4 目标分解 — 年度目标自上而下分解到事业部 + 月度进度
# ─────────────────────────────────────────────────────────────────────────────

@csrf_exempt
@pk_required()
def analytics_target_decomp(request):
    """年度目标 → 各事业部分解 → 各月目标/实际/完成率。"""
    year = _int_param(request, 'year', datetime.date.today().year)

    from caiwu.models import FinancialTarget, FinancialEntry, ImportBatch, L1Category
    from django.db.models import Sum as _Sum

    def _scope_bu(qs, field='business_unit'):
        if request.pk_role == 'super_admin':
            return qs
        # 无授权部门 → 空集（不再短路跳过过滤，杜绝越权看全部事业部目标）
        return qs.filter(**{field + '__in': request.pk_depts or []})

    # Annual targets (month=0)
    annual_qs = _scope_bu(FinancialTarget.objects.filter(year=year, month=0))
    annual_by_bu = {t.business_unit: t for t in annual_qs}

    # Monthly targets (month 1-12)
    monthly_qs = _scope_bu(FinancialTarget.objects.filter(year=year, month__gt=0))
    monthly_by_bu = {}
    for t in monthly_qs:
        monthly_by_bu.setdefault(t.business_unit, {})[t.month] = t

    # Actuals from published 部门明细表 (FinancialEntry), aggregated by BU+month for 主营业务收入
    rev_l1_ids = list(L1Category.objects.filter(name='主营业务收入').values_list('id', flat=True))
    agg_qs = _scope_bu(FinancialEntry.objects.filter(
        batch__status=ImportBatch.STATUS_PUBLISHED,
        batch__batch_type=ImportBatch.TYPE_DEPT,
        batch__year=year,
        l1_id__in=rev_l1_ids,
    ), field='batch__business_unit').values('batch__business_unit', 'batch__month').annotate(amount=_Sum('amount'))
    actuals = {}
    for r in agg_qs:
        actuals.setdefault(r['batch__business_unit'], {})[r['batch__month']] = float(r['amount'] or 0)

    today = datetime.date.today()
    elapsed = today.month if today.year == year else (12 if today.year > year else 0)

    all_bus = sorted(set(annual_by_bu) | set(monthly_by_bu) | set(actuals))

    rows = []
    for bu in all_bus:
        annual_t = annual_by_bu.get(bu)
        annual_rev = float(annual_t.target_revenue) if annual_t else 0
        annual_profit = float(annual_t.target_profit) if annual_t else 0
        annual_gross = float(annual_t.target_gross_profit) if annual_t else 0

        bu_actuals = actuals.get(bu, {})
        ytd_rev = sum(bu_actuals.get(m, 0) for m in range(1, elapsed + 1))
        projected = round(ytd_rev / elapsed * 12, 2) if elapsed else None

        months = []
        for m in range(1, 13):
            mt = monthly_by_bu.get(bu, {}).get(m)
            actual = bu_actuals.get(m, 0)
            t_rev = float(mt.target_revenue) if mt else None
            achieved = round(actual / t_rev * 100, 1) if t_rev else None
            months.append({
                'month': m,
                'target_revenue': t_rev,
                'actual_revenue': actual,
                'achieved': achieved,
                'is_past': m <= elapsed,
            })

        rows.append({
            'bu': bu,
            'annual_target_revenue': annual_rev,
            'annual_target_profit': annual_profit,
            'annual_target_gross': annual_gross,
            'ytd_actual_revenue': ytd_rev,
            'ytd_achieved': round(ytd_rev / annual_rev * 100, 1) if annual_rev else None,
            'projected': projected,
            'gap': round(annual_rev - projected, 2) if projected is not None and annual_rev else None,
            'months': months,
        })

    # Group-level summary
    total_annual = sum(r['annual_target_revenue'] for r in rows)
    total_ytd = sum(r['ytd_actual_revenue'] for r in rows)
    total_projected = round(total_ytd / elapsed * 12, 2) if elapsed and total_ytd else None

    return ok({
        'year': year,
        'elapsed': elapsed,
        'rows': rows,
        'summary': {
            'total_annual_target': total_annual,
            'total_ytd_actual': total_ytd,
            'total_ytd_achieved': round(total_ytd / total_annual * 100, 1) if total_annual else None,
            'total_projected': total_projected,
            'total_gap': round(total_annual - total_projected, 2) if total_projected is not None else None,
        },
    })


# ══════════════════════════════════════════════════════════════════════════════
# 资金池 (cash pool) — 每事业部一池：期初 + 收支流水推算账面余额，刚性/在途流出做预判
# ══════════════════════════════════════════════════════════════════════════════

def _pool_visible_depts(request):
    """可见池子 = 授权部门 ∩ 全局激活范围（?depts），与其他模块的范围口径一致。"""
    if request.pk_role == 'super_admin':
        allowed = list(DEPARTMENTS)
    else:
        allowed = [d for d in (request.pk_depts or []) if d in VALID_DEPARTMENTS]
    raw = (request.GET.get('depts') or '').strip()
    if raw:
        requested = [d for d in raw.split(',') if d.strip()]
        active = [d for d in allowed if d in requested]
        if active:
            return active
    return allowed


def _pool_actual_flows(dept, start, end):
    """(start, end] 区间的实际现金流（口径与现金流分析一致）。
    返回 (collected, adv_recv, paid, prepaid_offset, adv_paid, t_in, t_out)。"""
    collected = _dec(ARPayment.objects.filter(
        ar_record__delivery_dept=dept,
        payment_date__gt=start, payment_date__lte=end)
        .exclude(source='预收抵扣').aggregate(s=Sum('amount'))['s'])
    adv = (AdvanceRecord.objects.filter(
        delivery_dept=dept, occur_date__gt=start, occur_date__lte=end)
        .values('direction').annotate(s=Sum('advance_amount')))
    adv_map = {r['direction']: _dec(r['s']) for r in adv}
    paid = _dec(PaymentInstallment.objects.filter(
        payment__department=dept,
        pay_date__gt=start, pay_date__lte=end).aggregate(s=Sum('pay_amount'))['s'])
    # 预付核销冲抵：现金已在预付发生时流出，核销时不再是现金事件 → 从实付中扣除
    prepaid_offset = _dec(AdvanceWriteoff.objects.filter(
        payment__department=dept,
        writeoff_date__gt=start, writeoff_date__lte=end).aggregate(s=Sum('amount'))['s'])
    # 调拨：只有已生效（approved）的才是真实现金事件；待审批/已拒绝不动余额
    t_in = _dec(CashPoolTransfer.objects.filter(
        to_dept=dept, status='approved',
        transfer_date__gt=start, transfer_date__lte=end)
        .aggregate(s=Sum('amount'))['s'])
    t_out = _dec(CashPoolTransfer.objects.filter(
        from_dept=dept, status='approved',
        transfer_date__gt=start, transfer_date__lte=end)
        .aggregate(s=Sum('amount'))['s'])
    return (collected, adv_map.get('预收', Decimal('0')), paid, prepaid_offset,
            adv_map.get('预付', Decimal('0')), t_in, t_out)


def _pool_balance(dept, cfg, today):
    """池子当前账面余额 = 期初 + (期初日, 今天] 的净现金流。"""
    c, ar_, p, po, ap, ti, to_ = _pool_actual_flows(dept, cfg.initial_date, today)
    return cfg.initial_amount + c + ar_ - (p - po) - ap + ti - to_


def _pool_metrics(dept, cfg, today):
    """单个池子的全部指标：账面余额、资金预警线、刚性/在途流出、预期流入、余额预测。"""
    start = cfg.initial_date

    c, ar_, p, po, ap, ti, to_ = _pool_actual_flows(dept, start, today)
    balance = (cfg.initial_amount + c + ar_ - (p - po) - ap + ti - to_)

    # ── 刚性流出：付款台账已审批待付（remaining>0），按计划日期分窗。
    #    已用预付核销冲抵的部分不再需要现金，故一并扣除（与余额口径对称）。
    #    已付额必须用关联子查询（_paid_subq）而非 Sum('installments__...')：
    #    JOIN 聚合会让 rem__gt=0 落到 HAVING，其中裸列 plan_adjustment/
    #    total_amount/prepaid_offset_amount 不在 GROUP BY 里——SQLite 容忍，
    #    生产 MySQL 报 1054 Unknown column in 'having clause'。
    #    子查询版无 GROUP BY，过滤走 WHERE，两种库都安全（与付款列表口径同源）──
    pay_qs = (Payment.objects.filter(department=dept)
              .annotate(paid_sum=_paid_subq())
              .annotate(plan=Coalesce('plan_adjustment', 'total_amount'))
              .annotate(rem=F('plan') - F('paid_sum') - F('prepaid_offset_amount'))
              .filter(rem__gt=0))

    def _win(qs):
        return _dec(qs.aggregate(s=Sum('rem'))['s'])
    d30, d60, d90 = (today + datetime.timedelta(days=n) for n in (30, 60, 90))
    committed = {
        'overdue': _win(pay_qs.filter(planned_date__lte=today)),
        'd30': _win(pay_qs.filter(planned_date__gt=today, planned_date__lte=d30)),
        'd60': _win(pay_qs.filter(planned_date__gt=d30, planned_date__lte=d60)),
        'd90': _win(pay_qs.filter(planned_date__gt=d60, planned_date__lte=d90)),
    }
    committed['total'] = sum(committed.values())

    # ── 在途流出：审批记录（未排款）。已批待排=大概率近期变刚性；审批中=不确定。
    #    另含待审批的调拨出款申请——批准即出账，不计会高估可用余额 ─────────────
    pipe = (ApprovalRecord.objects.filter(department=dept, archived=False)
            .values('status').annotate(s=Sum('amount')))
    pipe_map = {r['status']: _dec(r['s']) for r in pipe}
    pipeline = {'approved': pipe_map.get('approved', Decimal('0')),
                'pending': pipe_map.get('pending', Decimal('0')),
                'transfer_out_pending': _dec(CashPoolTransfer.objects.filter(
                    from_dept=dept, status='pending').aggregate(s=Sum('amount'))['s'])}

    # ── 预期流入：未结清应收按到期日分窗；已逾期的单列（不计入预判，视作上行空间）──
    ar_qs = ARRecord.objects.filter(delivery_dept=dept, outstanding_amount__gt=0)

    def _ein(lo, hi):
        return _dec(ar_qs.filter(due_date__gt=lo, due_date__lte=hi)
                    .aggregate(s=Sum('outstanding_amount'))['s'])
    expected_in = {
        'd30': _ein(today, d30),
        'd60': _ein(d30, d60),
        'd90': _ein(d60, d90),
        'overdue_outstanding': _dec(ar_qs.filter(due_date__lte=today)
                                    .aggregate(s=Sum('outstanding_amount'))['s']),
    }

    # ── 余额预测：现余额 + 预期流入 − 刚性流出（逾期刚性视作立即要付）────────────
    level30 = balance + expected_in['d30'] - committed['overdue'] - committed['d30']
    level60 = level30 + expected_in['d60'] - committed['d60']
    level90 = level60 + expected_in['d90'] - committed['d90']
    # 审慎口径：已批待排的在途支出与待批调拨出款，按30天内全部付出计算
    d30_pess = level30 - pipeline['approved'] - pipeline['transfer_out_pending']
    projection = {'d30': str(level30), 'd60': str(level60), 'd90': str(level90),
                  'd30_with_pipeline': str(d30_pess)}

    # ── 资金预警线：超管手设固定额度优先；未设则按未来 warning_days 天
    #    刚性流出（含逾期未付）动态推算 ────────────────────────────────────────
    if cfg.warning_amount is not None:
        warn_amount = cfg.warning_amount
        warn_mode = 'fixed'
    else:
        wd = today + datetime.timedelta(days=cfg.warning_days or 30)
        warn_amount = committed['overdue'] + _win(
            pay_qs.filter(planned_date__gt=today, planned_date__lte=wd))
        warn_mode = 'dynamic'
    if balance < warn_amount:
        status = 'danger'
    elif level60 < 0 or d30_pess < 0:
        status = 'warn'
    else:
        status = 'ok'

    # ── 健康指标：近90天口径 ─────────────────────────────────────────────────
    t90 = today - datetime.timedelta(days=90)
    s90 = max(start, t90)
    c9, ar9, p9, po9, ap9, ti9, to9 = _pool_actual_flows(dept, s90, today)
    span = max(1, (today - s90).days)
    in90 = c9 + ar9
    out90 = (p9 - po9) + ap9
    runway = float(balance) / (float(out90) / span) if out90 > 0 and balance > 0 else None
    health = {
        'runway_days': round(runway) if runway is not None else None,
        'self_rate': round(float(in90) / float(out90) * 100, 1) if out90 > 0 else None,
        'net90': str(in90 - out90),
    }

    return {
        'dept': dept,
        'configured': True,
        'config': cfg.to_dict(),
        'balance': str(balance),
        'parts': {
            'initial': str(cfg.initial_amount),
            'collected': str(c), 'advance_received': str(ar_),
            'paid': str(p - po), 'advance_paid': str(ap),
            'transfer_in': str(ti), 'transfer_out': str(to_),
        },
        'warning': {'amount': str(warn_amount), 'status': status, 'mode': warn_mode},
        'committed': {k: str(v) for k, v in committed.items()},
        'pipeline': {k: str(v) for k, v in pipeline.items()},
        'expected_in': {k: str(v) for k, v in expected_in.items()},
        'projection': projection,
        'health': health,
    }


@csrf_exempt
@pk_required()
def cash_pool(request):
    """资金池总览：可见事业部各一池（账面余额/资金预警线/刚性+在途流出/预期流入/余额预测）。"""
    denied = _page_denied(request, 'ar_cashflow')
    if denied:
        return denied
    if request.method != 'GET':
        return err('Method not allowed', 405)

    today = datetime.date.today()
    depts = _pool_visible_depts(request)
    try:
        cfgs = {c.delivery_dept: c for c in CashPoolConfig.objects.filter(delivery_dept__in=depts)}
    except Exception as exc:
        import traceback
        logger.error('cash_pool: config query failed: %s\n%s', exc, traceback.format_exc())
        return err(f'资金池配置读取失败：{exc}。'
                   f'若提示字段/列不存在，说明数据库迁移未应用到当前服务所连的库——'
                   f'请在服务器运行环境（与启动服务相同的环境变量，含 MYSQL_ADDRESS）'
                   f'执行 python manage.py migrate 后重启服务', 500)

    pools = []
    for dept in depts:
        cfg = cfgs.get(dept)
        if cfg is None:
            pools.append({'dept': dept, 'configured': False})
            continue
        try:
            pools.append(_pool_metrics(dept, cfg, today))
        except Exception as exc:
            import traceback
            logger.error('cash_pool: _pool_metrics failed dept=%s: %s\n%s',
                         dept, exc, traceback.format_exc())
            pools.append({'dept': dept, 'configured': True, 'error': str(exc)})

    configured = [p for p in pools if p.get('configured') and not p.get('error')]
    group = None
    if configured:
        def _sumf(getter):
            return str(sum(Decimal(getter(p)) for p in configured))
        group = {
            'balance': _sumf(lambda p: p['balance']),
            'warning_amount': _sumf(lambda p: p['warning']['amount']),
            'committed_total': _sumf(lambda p: p['committed']['total']),
            'pipeline_approved': _sumf(lambda p: p['pipeline']['approved']),
            'pipeline_pending': _sumf(lambda p: p['pipeline']['pending']),
            'projection_d30': _sumf(lambda p: p['projection']['d30']),
            'projection_d60': _sumf(lambda p: p['projection']['d60']),
            'projection_d90': _sumf(lambda p: p['projection']['d90']),
            'danger_count': sum(1 for p in configured if p['warning']['status'] == 'danger'),
            'warn_count': sum(1 for p in configured if p['warning']['status'] == 'warn'),
        }
    return ok({'pools': pools, 'group': group, 'today': str(today)})


@csrf_exempt
@pk_required(roles=['super_admin'])
def cash_pool_config(request):
    """池配置（仅超管）：GET 列表 / POST 按事业部 upsert 期初基准。"""
    if request.method == 'GET':
        return ok({'items': [c.to_dict() for c in CashPoolConfig.objects.all()],
                   'departments': DEPARTMENTS})
    if request.method == 'POST':
        data = _parse_body(request)
        dept = (data.get('delivery_dept') or '').strip()
        if dept not in VALID_DEPARTMENTS:
            return err('无效的事业部')
        initial_date = _normalize_date(data.get('initial_date'))
        if not initial_date:
            return err('期初基准日必填（YYYY-MM-DD）')
        try:
            init_d = datetime.date.fromisoformat(str(initial_date)[:10])
        except (ValueError, TypeError):
            return err('期初基准日格式错误')
        if init_d > datetime.date.today():
            return err('期初基准日不能晚于今天（期初是已发生的账面事实）')
        try:
            initial_amount = Decimal(str(data.get('initial_amount', 0) or 0))
        except (InvalidOperation, ValueError):
            return err('期初金额格式错误')
        try:
            warning_days = max(7, min(120, int(data.get('warning_days', 30) or 30)))
        except (ValueError, TypeError):
            warning_days = 30
        warning_amount = None
        raw_warn = data.get('warning_amount')
        if raw_warn not in (None, ''):
            try:
                warning_amount = Decimal(str(raw_warn))
            except (InvalidOperation, ValueError):
                return err('资金预警线金额格式错误')
            if warning_amount < 0:
                return err('资金预警线不能为负数')
        cfg, _ = CashPoolConfig.objects.update_or_create(
            delivery_dept=dept,
            defaults={'initial_date': initial_date, 'initial_amount': initial_amount,
                      'warning_days': warning_days, 'warning_amount': warning_amount,
                      'notes': (data.get('notes') or '').strip(),
                      'updated_by': request.pk_user})
        return ok(cfg.to_dict())
    return err('Method not allowed', 405)


def _validate_transfer_payload(data):
    """调拨公共校验。返回 (from_dept, to_dept, amount, tr_date, err_response)。"""
    f, t = (data.get('from_dept') or '').strip(), (data.get('to_dept') or '').strip()
    if f not in VALID_DEPARTMENTS or t not in VALID_DEPARTMENTS:
        return None, None, None, None, err('无效的事业部')
    if f == t:
        return None, None, None, None, err('调出与调入不能是同一个池子')
    try:
        amount = Decimal(str(data.get('amount') or 0))
    except (InvalidOperation, ValueError):
        return None, None, None, None, err('金额格式错误')
    if amount <= 0:
        return None, None, None, None, err('调拨金额必须大于0')
    tr_date = _normalize_date(data.get('transfer_date')) or str(datetime.date.today())
    try:
        tr_date_d = datetime.date.fromisoformat(str(tr_date)[:10])
    except (ValueError, TypeError):
        return None, None, None, None, err('调拨日期格式错误')
    if tr_date_d > datetime.date.today():
        return None, None, None, None, err('调拨日期不能晚于今天（调拨以实际发生日入账）')
    return f, t, amount, tr_date_d, None


def _transfer_guards(f, t, tr_date_d, lock=False):
    """期初配置与日期守恒校验。lock=True 时锁定两侧配置行（须在事务内），
    串行化并发的余额检查，防止两笔同时通过校验后合计透支。
    返回 (cfg_f, cfg_t, err_response)。"""
    qs = CashPoolConfig.objects.select_for_update() if lock else CashPoolConfig.objects
    cfgs = {c.delivery_dept: c for c in qs.filter(delivery_dept__in=[f, t])}
    cfg_f, cfg_t = cfgs.get(f), cfgs.get(t)
    # 两池都须已配置期初，且调拨日期不得早于任一侧期初基准日——
    # 否则一侧计入一侧忽略，集团合计≠各池之和（台账不守恒）
    if not cfg_f or not cfg_t:
        return None, None, err('调出/调入池需先在「池配置」中设置期初基准')
    if tr_date_d < cfg_f.initial_date or tr_date_d < cfg_t.initial_date:
        return None, None, err('调拨日期不能早于两侧池子的期初基准日（会造成台账不守恒）')
    return cfg_f, cfg_t, None


def _transfer_overdraft_err(f, cfg_f, amount):
    """不允许透支调拨：调出额不得超过调出池当前账面余额。"""
    balance = _pool_balance(f, cfg_f, datetime.date.today())
    if amount > balance:
        return err(f'调出池「{f}」当前可用余额 {balance:,.2f} 元，'
                   f'不足以调出 {amount:,.2f} 元（不允许透支调拨）')
    return None


@csrf_exempt
@pk_required()
def cash_pool_transfers(request):
    """池间资金调拨。
    GET：列表（非超管只看与自己部门相关的）。
    POST：超管直接调拨即时生效；事业部用户提交「调拨申请」（status=pending），
          须经调出方事业部（或超管）审批后生效，且发起人须与调出/调入一侧同部门。"""
    if request.method == 'GET':
        denied = _page_denied(request, 'ar_cashflow')
        if denied:
            return denied
        qs = CashPoolTransfer.objects.all()
        if request.pk_role != 'super_admin':
            mine = request.pk_depts or []
            qs = qs.filter(Q(from_dept__in=mine) | Q(to_dept__in=mine))
        try:
            return ok({'items': [t.to_dict() for t in qs[:200]]})
        except Exception as exc:
            import traceback
            logger.error('cash_pool_transfers: list failed: %s\n%s', exc, traceback.format_exc())
            return err(f'调拨台账读取失败：{exc}。'
                       f'若提示字段/列不存在，请在服务器运行环境执行 '
                       f'python manage.py migrate 后重启服务', 500)
    if request.method == 'POST':
        data = _parse_body(request)
        f, t, amount, tr_date_d, bad = _validate_transfer_payload(data)
        if bad:
            return bad
        common = {'expected_return_date': _normalize_date(data.get('expected_return_date')),
                  'notes': (data.get('notes') or '').strip(), 'created_by': request.pk_user}

        if request.pk_role == 'super_admin':
            # 集团统筹：直接生效。锁配置行串行化余额检查，防并发合计透支
            with transaction.atomic():
                cfg_f, cfg_t, bad = _transfer_guards(f, t, tr_date_d, lock=True)
                if bad:
                    return bad
                bad = _transfer_overdraft_err(f, cfg_f, amount)
                if bad:
                    return bad
                tr = CashPoolTransfer.objects.create(
                    from_dept=f, to_dept=t, amount=amount, transfer_date=tr_date_d,
                    status='approved', reviewed_by=request.pk_user,
                    reviewed_at=timezone.now(), **common)
            return ok(tr.to_dict())

        # 事业部用户：提交调拨申请，待调出方审批
        denied = _write_denied(request)
        if denied:
            return denied
        mine = request.pk_depts or []
        if f not in mine and t not in mine:
            return err('只能发起与本部门相关的调拨申请（调出或调入一侧须为本部门）', 403, 403)
        cfg_f, cfg_t, bad = _transfer_guards(f, t, tr_date_d)
        if bad:
            return bad
        # 申请阶段不冻结资金、不动余额；透支在审批生效时校验
        tr = CashPoolTransfer.objects.create(
            from_dept=f, to_dept=t, amount=amount, transfer_date=tr_date_d,
            status='pending', **common)
        return ok(tr.to_dict())
    return err('Method not allowed', 405)


@csrf_exempt
@pk_required()
def cash_pool_transfer_review(request, pk):
    """调拨申请审批：批准/拒绝。仅超管或调出方事业部（有写入权限）可审；
    不能审批自己发起的申请。批准时以审批日为实际生效日并校验余额充足。"""
    if request.method != 'POST':
        return err('Method not allowed', 405)
    try:
        tr = CashPoolTransfer.objects.get(pk=pk)
    except CashPoolTransfer.DoesNotExist:
        return err('调拨申请不存在', 404)
    if tr.status != 'pending':
        return err('该申请已处理（只有待审批的申请可以审批）')

    if request.pk_role != 'super_admin':
        if tr.from_dept not in (request.pk_depts or []):
            return err('只有调出方事业部（或超管）可以审批此申请', 403, 403)
        denied = _write_denied(request)
        if denied:
            return denied
        if tr.created_by_id and tr.created_by_id == request.pk_user.id:
            return err('不能审批自己发起的调拨申请', 403, 403)

    data = _parse_body(request)
    action = (data.get('action') or '').strip()
    review_notes = (data.get('review_notes') or '').strip()
    if action == 'reject':
        tr.status = 'rejected'
        tr.reviewed_by = request.pk_user
        tr.reviewed_at = timezone.now()
        tr.review_notes = review_notes
        tr.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'review_notes'])
        return ok(tr.to_dict())
    if action != 'approve':
        return err("action 须为 'approve' 或 'reject'")

    today = datetime.date.today()
    with transaction.atomic():
        cfg_f, cfg_t, bad = _transfer_guards(tr.from_dept, tr.to_dept, today, lock=True)
        if bad:
            return bad
        bad = _transfer_overdraft_err(tr.from_dept, cfg_f, tr.amount)
        if bad:
            return bad
        # 生效日以审批日为准：现金事件记在它实际发生的那天
        tr.status = 'approved'
        tr.transfer_date = today
        tr.reviewed_by = request.pk_user
        tr.reviewed_at = timezone.now()
        tr.review_notes = review_notes
        tr.save(update_fields=['status', 'transfer_date', 'reviewed_by',
                               'reviewed_at', 'review_notes'])
    return ok(tr.to_dict())


@csrf_exempt
@pk_required()
def cash_pool_transfer_detail(request, pk):
    """删除调拨：待审批的申请发起人可撤回；已生效/已拒绝的仅超管可删（余额回退）。"""
    if request.method != 'DELETE':
        return err('Method not allowed', 405)
    try:
        tr = CashPoolTransfer.objects.get(pk=pk)
    except CashPoolTransfer.DoesNotExist:
        return err('调拨记录不存在', 404)
    if request.pk_role != 'super_admin':
        if not (tr.status == 'pending' and tr.created_by_id
                and tr.created_by_id == request.pk_user.id):
            return err('只能撤回自己发起且尚未审批的调拨申请', 403, 403)
    tr.delete()
    return ok({'deleted': pk})
