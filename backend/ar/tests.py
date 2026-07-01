import io
import json
from datetime import date, timedelta
from decimal import Decimal

import openpyxl
from django.test import Client, TestCase

from ar.models import (ARPayment, ARProject, ARRecord, ARAdjustment, AdvanceInstallment, CollectionBudget, PaymentBudget,
                       AdvanceRecord, AdvanceWriteoff, Customer,
                       Contract, ContractParty, ContractProject, ActionItem)
from paikuan.models import JobPermission, PaikuanUser
from paikuan.views import DEPARTMENTS, default_job_config, make_token, _invalidate_perm_cache


class ARPermissionRegressionTests(TestCase):
    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.dept = '劳务事业部'
        self.other_dept = '运输事业部'

    def tearDown(self):
        _invalidate_perm_cache()

    def make_user(self, phone, job_title, departments=None, role='operator'):
        user = PaikuanUser(
            phone=phone,
            name=f'{job_title} user',
            role=role,
            job_title=job_title,
            departments=departments or [self.dept],
            is_active=True,
            is_approved=True,
        )
        user.set_password('Test123456')
        user.save()
        return user

    def auth(self, user):
        return {'HTTP_AUTHORIZATION': f'Bearer {make_token(user)}'}

    def json_post(self, url, payload, user):
        return self.client.post(
            url, data=json.dumps(payload), content_type='application/json',
            **self.auth(user))

    def json_put(self, url, payload, user):
        return self.client.put(
            url, data=json.dumps(payload), content_type='application/json',
            **self.auth(user))

    def create_project(self, dept=None, short_name='Project A', delivery_dept=None):
        return ARProject.objects.create(
            customer_name='Contract A',
            short_name=short_name,
            delivery_dept=delivery_dept or dept or self.dept,
            sub_dept='Sub A',
            business_mode='Mode A',
            customer_level='A级',
            sales_contact='Sales A',
            project_manager='PM A',
            has_contract='有',
            contract_date=date(2026, 1, 1),
            reconciliation_days=10,
            invoice_wait_days=5,
            post_invoice_days=15,
            invoice_mode='全额',
            invoice_type='专票',
            tax_rate=Decimal('0.0600'),
        )

    def create_record(self, project=None):
        return ARRecord.objects.create(
            project=project or self.create_project(),
            operation_year=2026,
            operation_month=5,
            estimated_amount=Decimal('1000.00'),
            actual_invoice_amount=Decimal('1060.00'),
            invoice_date=date(2026, 5, 31),
        )

    def project_payload(self):
        return {
            'customer_name': 'Contract B',
            'short_name': 'Project B',
            'delivery_dept': self.dept,
            'sub_dept': 'Sub B',
            'business_mode': 'Mode B',
            'customer_level': 'A级',
            'sales_contact': 'Sales B',
            'project_manager': 'PM B',
            'has_contract': '有',
            'contract_date': '2026-02-01',
            'reconciliation_days': 10,
            'invoice_wait_days': 5,
            'post_invoice_days': 15,
            'invoice_mode': '全额',
            'invoice_type': '专票',
            'tax_rate': '0.06',
        }

    def headers_from_xlsx(self, response):
        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
        return [cell.value for cell in wb.active[1]]

    def test_cashflow_excludes_soft_deleted_payments(self):
        """现金流分析（驾驶舱同源接口）：已软删除付款的实付分期不得计入现金流出。"""
        from paikuan.models import Payment, PaymentInstallment
        from django.utils import timezone
        admin = self.make_user('13900000399', 'finance_director', role='super_admin')
        p = Payment.objects.create(
            department=self.dept, project_desc='现金流测试', payee='供应商X',
            total_amount=Decimal('800'), planned_date=date(2026, 6, 1))
        PaymentInstallment.objects.create(payment=p, seq=1, pay_date=date(2026, 6, 10),
                                          pay_amount=Decimal('800'))
        params = {'start_year': 2026, 'start_month': 6, 'end_year': 2026, 'end_month': 6}
        d = self.client.get('/api/pk/ar/cashflow', params, **self.auth(admin)).json()['data']
        self.assertEqual(d['totals']['paid'], [800.0])       # 在册付款计入流出
        # 软删除该付款 → 流出归零（回收站中的付款不构成现金流出）
        p.deleted_at = timezone.now()
        p.save(update_fields=['deleted_at'])
        d2 = self.client.get('/api/pk/ar/cashflow', params, **self.auth(admin)).json()['data']
        self.assertEqual(d2['totals']['paid'], [0.0])

    def test_actual_receivable_and_invoice_mismatch(self):
        proj = self.create_project()
        # 实际应收 = 预估 + 账实差额；已开票且 ≠ 实际应收 → invoice_mismatch=True
        r = ARRecord.objects.create(
            project=proj, operation_year=2026, operation_month=5,
            estimated_amount=Decimal('1000.00'),
            account_diff_adjustment=Decimal('60.00'),
            actual_invoice_amount=Decimal('1000.00'),   # 开票1000 ≠ 实际应收1060
            invoice_date=date(2026, 5, 31))
        d = r.to_dict()
        self.assertEqual(d['actual_receivable'], '1060.00')
        self.assertTrue(d['invoice_mismatch'])
        # 开票金额等于实际应收 → 不提醒
        r.actual_invoice_amount = Decimal('1060.00')
        r.save()
        self.assertFalse(r.to_dict()['invoice_mismatch'])
        # 未开票 → 即使预估≠账面也不提醒；实际应收照常计算
        r2 = ARRecord.objects.create(
            project=proj, operation_year=2026, operation_month=6,
            estimated_amount=Decimal('500.00'), account_diff_adjustment=Decimal('0.00'))
        self.assertIsNone(r2.actual_invoice_amount)
        self.assertFalse(r2.to_dict()['invoice_mismatch'])
        self.assertEqual(r2.to_dict()['actual_receivable'], '500.00')

    def test_project_post_invoice_days_update_persists(self):
        """票后等待期(post_invoice_days)编辑后应真正落库——回归 settlement_wait_days
        旧列名残留导致 _ar_visible_payload 把该字段从 PUT 载荷里剥掉的问题。"""
        admin = self.make_user('13900000311', 'finance_director', role='super_admin')
        resp = self.json_post('/api/pk/ar/projects', self.project_payload(), admin)
        self.assertEqual(resp.status_code, 200, resp.content)
        pid = resp.json()['data']['id']
        put = self.json_put(f'/api/pk/ar/projects/{pid}', {'post_invoice_days': 30}, admin)
        self.assertEqual(put.status_code, 200, put.content)
        self.assertEqual(put.json()['data']['post_invoice_days'], 30)
        # 重新拉取确认落库（而非仅响应回显）
        got = self.client.get(f'/api/pk/ar/projects/{pid}', **self.auth(admin))
        self.assertEqual(got.json()['data']['post_invoice_days'], 30)

    def test_projects_list_exposes_customer_name_for_advance_autofill(self):
        """预收新增的关联项目下拉用 /ar/projects，选中后以 customer_name 自动带出
        往来单位。该字段须随列表返回且非空，否则前端 pickProject 拿不到值。
        「合同名正名为客户」上线后：新建项目按合同名自动建/挂客户（_autolink_customer），
        列表返回的 customer_id/customer_name 应非空且客户名=合同名。"""
        cfg = default_job_config('cashier')
        cfg['pages']['ar_projects'] = True
        JobPermission.objects.create(job_title='cashier', config=cfg)
        _invalidate_perm_cache('cashier')
        user = self.make_user('13910000099', 'cashier')
        self.create_project(short_name='福佑物流')
        resp = self.client.get('/api/pk/ar/projects', {'q': '福佑物流'}, **self.auth(user))
        self.assertEqual(resp.status_code, 200, resp.content)
        item = resp.json()['data']['items'][0]
        self.assertEqual(item['customer_name'], 'Contract A')
        # 合同名自动正名为客户：customer 非空且名称=合同名
        self.assertIsNotNone(item['customer_id'])
        self.assertEqual(item['customer_name'], 'Contract A')

    def test_cashier_page_access_cannot_write_ar_records(self):
        cfg = default_job_config('cashier')
        cfg['pages']['ar_projects'] = True
        JobPermission.objects.create(job_title='cashier', config=cfg)
        _invalidate_perm_cache('cashier')
        user = self.make_user('13910000001', 'cashier')
        project = self.create_project()
        record = self.create_record(project)

        checks = [
            self.json_post('/api/pk/ar/projects', self.project_payload(), user),
            self.json_post('/api/pk/ar/records', {
                'project_id': project.id,
                'operation_year': 2026,
                'operation_month': 6,
                'estimated_amount': '100.00',
            }, user),
            self.json_post('/api/pk/ar/budget/collection', {
                'short_name': 'Budget A',
                'expected_date': '2026-06-10',
                'delivery_dept': self.dept,
                'amount': '100.00',
            }, user),
        ]

        self.assertTrue(all(resp.status_code == 403 for resp in checks))

        # 回款录入由操作权限 ar_collect 单独管控（出纳基线默认开通）：
        # 无 can_create 也可录回款——这是出纳的本职动作，不再绑死通用写权限
        resp = self.json_post(f'/api/pk/ar/records/{record.id}/payments', {
            'amount': '100.00',
            'payment_date': '2026-06-11',
        }, user)
        self.assertEqual(resp.status_code, 200, resp.content)

    def test_budget_detail_is_department_scoped_and_delete_requires_permission(self):
        user = self.make_user('13910000002', 'finance_bp')
        own_budget = CollectionBudget.objects.create(
            short_name='Own Budget',
            expected_date=date(2026, 6, 1),
            delivery_dept=self.dept,
            amount=Decimal('100.00'),
        )
        other_budget = CollectionBudget.objects.create(
            short_name='Other Budget',
            expected_date=date(2026, 6, 1),
            delivery_dept=self.other_dept,
            amount=Decimal('200.00'),
        )

        self.assertEqual(
            self.client.get(f'/api/pk/ar/budget/collection/{other_budget.id}', **self.auth(user)).status_code,
            403,
        )
        self.assertEqual(
            self.json_put(f'/api/pk/ar/budget/collection/{other_budget.id}', {'amount': '300.00'}, user).status_code,
            403,
        )
        self.assertEqual(
            self.client.delete(f'/api/pk/ar/budget/collection/{other_budget.id}', **self.auth(user)).status_code,
            403,
        )

        ok_resp = self.json_put(
            f'/api/pk/ar/budget/collection/{own_budget.id}',
            {'amount': '150.00'},
            user,
        )
        self.assertEqual(ok_resp.status_code, 200)
        self.assertEqual(
            self.client.delete(f'/api/pk/ar/budget/collection/{own_budget.id}', **self.auth(user)).status_code,
            403,
        )

    def test_ar_exports_respect_hidden_field_permissions(self):
        cfg = default_job_config('cashier')
        cfg['pages']['ar_projects'] = True
        cfg['ar_view']['p_customer_name'] = False
        JobPermission.objects.create(job_title='cashier', config=cfg)
        _invalidate_perm_cache('cashier')
        user = self.make_user('13910000003', 'cashier')
        project = self.create_project()
        self.create_record(project)

        project_resp = self.client.get('/api/pk/ar/projects/export', **self.auth(user))
        record_resp = self.client.get('/api/pk/ar/records/export', **self.auth(user))

        self.assertEqual(project_resp.status_code, 200)
        self.assertNotIn('合同名称', self.headers_from_xlsx(project_resp))
        self.assertEqual(record_resp.status_code, 200)
        record_headers = self.headers_from_xlsx(record_resp)
        self.assertNotIn('税额', record_headers)
        self.assertNotIn('账实差额调整', record_headers)

    def test_hidden_payments_field_blocks_payment_subresource(self):
        cfg = default_job_config('finance_bp')
        cfg['ar_view']['r_payments'] = False
        JobPermission.objects.create(job_title='finance_bp', config=cfg)
        _invalidate_perm_cache('finance_bp')
        user = self.make_user('13910000004', 'finance_bp')
        record = self.create_record()

        get_resp = self.client.get(f'/api/pk/ar/records/{record.id}/payments', **self.auth(user))
        post_resp = self.json_post(f'/api/pk/ar/records/{record.id}/payments', {
            'amount': '100.00',
            'payment_date': '2026-06-11',
        }, user)

        self.assertEqual(get_resp.status_code, 403)
        self.assertEqual(post_resp.status_code, 403)

    def test_payment_signal_recomputes_tax_and_outstanding_amount(self):
        record = self.create_record()

        self.assertEqual(record.tax_amount, Decimal('60.00'))
        # outstanding is based on estimated_amount (1000), not invoice amount
        self.assertEqual(record.outstanding_amount, Decimal('1000.00'))

        ARPayment.objects.create(
            ar_record=record,
            payment_no=1,
            amount=Decimal('160.00'),
            payment_date=date(2026, 6, 10),
        )

        record.refresh_from_db()
        self.assertEqual(record.tax_amount, Decimal('60.00'))
        self.assertEqual(record.outstanding_amount, Decimal('840.00'))

    def test_data_health_flags_stale_and_negative_records(self):
        admin = self.make_user('13900000099', 'finance_director', role='super_admin')
        project = self.create_project()

        # healthy record — should not be flagged
        healthy = self.create_record(project)
        ARPayment.objects.create(ar_record=healthy, payment_no=1,
                                 amount=Decimal('300.00'), payment_date=date(2026, 6, 1))

        # stale record — tamper stored outstanding directly (bypassing recompute)
        stale = ARRecord.objects.create(project=project, operation_year=2026,
                                        operation_month=6, estimated_amount=Decimal('770000.00'))
        ARPayment.objects.create(ar_record=stale, payment_no=1,
                                 amount=Decimal('100000.00'), payment_date=date(2026, 6, 2))
        ARRecord.objects.filter(pk=stale.pk).update(outstanding_amount=Decimal('999999.00'))

        # negative record — payments exceed estimated (raw insert to bypass validation)
        neg = ARRecord.objects.create(project=project, operation_year=2026,
                                      operation_month=7, estimated_amount=Decimal('270000.00'))
        ARPayment.objects.bulk_create([
            ARPayment(ar_record=neg, payment_no=1, amount=Decimal('100000.00'),
                      payment_date=date(2026, 6, 3)),
            ARPayment(ar_record=neg, payment_no=2, amount=Decimal('200000.00'),
                      payment_date=date(2026, 6, 4)),
        ])

        resp = self.client.get('/api/pk/ar/records/health', **self.auth(admin))
        self.assertEqual(resp.status_code, 200)
        data = resp.json()['data']
        self.assertTrue(data['has_issues'])
        self.assertEqual(data['stale_count'], 1)
        self.assertEqual(data['negative_count'], 1)
        self.assertEqual(data['stale'][0]['id'], stale.id)
        self.assertEqual(data['negative'][0]['id'], neg.id)
        self.assertEqual(data['negative'][0]['deficit'], '30000.00')

        # bulk recompute fixes the stale one and reports the negative one as failed
        resp = self.json_post('/api/pk/ar/records/recompute',
                              {'ids': [stale.id, neg.id]}, admin)
        self.assertEqual(resp.status_code, 200)
        result = resp.json()['data']
        self.assertEqual(result['fixed'], 1)
        self.assertEqual(len(result['failed']), 1)
        self.assertEqual(result['failed'][0]['id'], neg.id)

        stale.refresh_from_db()
        self.assertEqual(stale.outstanding_amount, Decimal('670000.00'))

    def test_project_update_syncs_budget_dept_fields(self):
        project = self.create_project()
        cb = CollectionBudget.objects.create(
            project_no=project.project_no, short_name=project.short_name,
            expected_date=date(2026, 6, 1), delivery_dept=project.delivery_dept,
            sub_dept=project.sub_dept or '', amount=Decimal('100000'))
        pb = PaymentBudget.objects.create(
            project_no=project.project_no, short_name=project.short_name,
            expected_date=date(2026, 6, 1), delivery_dept=project.delivery_dept,
            sub_dept=project.sub_dept or '', amount=Decimal('50000'))

        # Change delivery_dept and sub_dept on the project — signal should propagate
        project.delivery_dept = '运输事业部'
        project.sub_dept = '新二级部门'
        project.save()

        cb.refresh_from_db(); pb.refresh_from_db()
        self.assertEqual(cb.delivery_dept, '运输事业部')
        self.assertEqual(cb.sub_dept, '新二级部门')
        self.assertEqual(pb.delivery_dept, '运输事业部')
        self.assertEqual(pb.sub_dept, '新二级部门')

    def test_budget_import_corrects_wrong_fields_from_project(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        admin = self.make_user('13900000088', 'finance_director', role='super_admin')
        project = self.create_project()  # short_name 'Project A', dept 劳务事业部, sub 'Sub A'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['项目编号(可选)', '项目简称/摘要*', '预计收款日期(YYYY-MM-DD)*',
                   '二级部门', '交付部门', '金额*', '备注'])
        # correct short_name, but WRONG project_no / sub_dept / delivery_dept
        ws.append(['WRONG-001', 'Project A', '2026-06-15',
                   '错误二级部门', '运输事业部', 100000, '测试'])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        upload = SimpleUploadedFile(
            'b.xlsx', buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        resp = self.client.post('/api/pk/ar/budget/collection/import',
                                data={'file': upload}, **self.auth(admin))
        self.assertEqual(resp.status_code, 200)
        data = resp.json()['data']
        self.assertEqual(data['created'], 1)
        self.assertEqual(data['corrected'], 1)

        cb = CollectionBudget.objects.get(short_name='Project A')
        # All three wrong fields overridden with the project's authoritative values
        self.assertEqual(cb.delivery_dept, project.delivery_dept)
        self.assertEqual(cb.sub_dept, project.sub_dept)
        self.assertEqual(cb.project_no, project.project_no)

    def test_budget_import_accepts_excel_date_cell(self):
        """回归：Excel 把日期列识别为日期类型时（openpyxl 返回 datetime），
        导入不应再报"预计日期无效"。这是"按模板填都报错"的根因。"""
        import datetime as dt
        from django.core.files.uploadedfile import SimpleUploadedFile
        admin = self.make_user('13900000201', 'finance_director', role='super_admin')
        self.create_project(short_name='日期格式项目', delivery_dept='劳务事业部')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['项目编号(可选)', '项目简称/摘要*', '预计收款日期*',
                   '二级部门', '交付部门', '金额*', '备注'])
        # 关键：日期作为真正的 datetime 单元格写入（模拟 Excel 自动日期识别）
        ws.append(['', '日期格式项目', dt.datetime(2026, 6, 15), '', '劳务事业部', 50000, ''])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        resp = self.client.post('/api/pk/ar/budget/collection/import',
                                {'file': SimpleUploadedFile('b.xlsx', buf.read())},
                                **self.auth(admin))
        self.assertEqual(resp.status_code, 200)
        d = resp.json()['data']
        self.assertEqual(d['created'], 1, d.get('errors'))
        self.assertEqual(d['errors'], [])
        self.assertEqual(str(CollectionBudget.objects.get().expected_date), '2026-06-15')

    def test_budget_import_accepts_multiple_date_formats(self):
        """回归：模板支持的多种日期格式（斜杠/中文/紧凑/2位年）导入均应通过。"""
        from django.core.files.uploadedfile import SimpleUploadedFile
        admin = self.make_user('13900000203', 'finance_director', role='super_admin')
        self.create_project(short_name='多格式项目', delivery_dept='劳务事业部')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['项目编号(可选)', '项目简称/摘要*', '预计收款日期*',
                   '二级部门', '交付部门', '金额*', '备注'])
        for d in ['2026/6/15', '2026年6月15日', '20260615', '26-6-15']:
            ws.append(['', '多格式项目', d, '', '劳务事业部', 1000, ''])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        resp = self.client.post('/api/pk/ar/budget/collection/import',
                                {'file': SimpleUploadedFile('b.xlsx', buf.read())},
                                **self.auth(admin))
        self.assertEqual(resp.status_code, 200)
        d = resp.json()['data']
        self.assertEqual(d['created'], 4, d.get('errors'))
        self.assertEqual(d['errors'], [])
        dates = sorted(str(b.expected_date) for b in CollectionBudget.objects.all())
        self.assertEqual(dates, ['2026-06-15'] * 4)

    def test_ar_record_import_rejects_whole_file_on_error(self):
        """应收明细导入：任一行【格式错误】→ 整表拒绝、零写入、列出全部问题；
        全部合法 → 整表写入。（找不到的项目不算错误，写入阶段自动建草稿）"""
        from django.core.files.uploadedfile import SimpleUploadedFile
        admin = self.make_user('13900000301', 'finance_director', role='super_admin')
        self.create_project(short_name='应收导入项目', delivery_dept='劳务事业部')
        HEAD = ['项目简称*', '运作年*', '运作月*', '预估上账金额', '实际开票金额',
                '税额(差额模式手填)', '开票日期', '账实差额调整', '回款金额', '回款时间', '备注']

        # ① 含格式错误：第2行金额非法 → 整表拒绝（第3行项目不存在不算错误）
        wb = openpyxl.Workbook(); ws = wb.active; ws.append(HEAD)
        ws.append(['应收导入项目', 2026, 1, 'abc', '', '', '', '', '', '', ''])   # 金额非法
        ws.append(['不存在的项目', 2026, 1, 1000, '', '', '', '', '', '', ''])    # 项目缺失(自动建草稿)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        resp = self.client.post('/api/pk/ar/records/import',
                                {'file': SimpleUploadedFile('r.xlsx', buf.read())}, **self.auth(admin))
        self.assertEqual(resp.status_code, 200)
        d = resp.json()['data']
        self.assertTrue(d['rejected'])
        self.assertEqual(d['created'], 0)
        self.assertEqual(len(d['errors']), 1, d['errors'])   # 仅金额格式 1 处
        self.assertEqual(ARRecord.objects.count(), 0)   # 零写入

        # ② 全部合法 → 写入
        wb2 = openpyxl.Workbook(); ws2 = wb2.active; ws2.append(HEAD)
        ws2.append(['应收导入项目', 2026, 1, 100000, 100000, '', '2026-01-15', 0, 30000, '2026-01-20', ''])
        ws2.append(['应收导入项目', 2026, 2, 80000, '', '', '', '', '', '', '同月可多条'])
        buf2 = io.BytesIO(); wb2.save(buf2); buf2.seek(0)
        resp2 = self.client.post('/api/pk/ar/records/import',
                                 {'file': SimpleUploadedFile('r2.xlsx', buf2.read())}, **self.auth(admin))
        d2 = resp2.json()['data']
        self.assertFalse(d2.get('rejected'))
        self.assertEqual(d2['created'], 2, d2.get('errors'))
        self.assertEqual(ARRecord.objects.count(), 2)

    def test_budget_import_skips_template_tip_and_example_rows(self):
        """回归：模板自带的"说明行"(★开头)与"示例行"(含示例标记)应被跳过，
        不会被当作真实数据导入或产生日期错误。"""
        from ar.views import _budget_template
        from django.test import RequestFactory
        from django.core.files.uploadedfile import SimpleUploadedFile
        admin = self.make_user('13900000202', 'finance_director', role='super_admin')
        # 直接下载模板，原样回传导入——应零创建、零错误
        rf = RequestFactory()
        req = rf.get('/api/pk/ar/budget/collection/template')
        req.pk_uid = admin.id; req.pk_role = 'super_admin'; req.pk_depts = []
        tmpl_resp = _budget_template(req, 'collection')
        buf = io.BytesIO(tmpl_resp.content); buf.seek(0)
        resp = self.client.post('/api/pk/ar/budget/collection/import',
                                {'file': SimpleUploadedFile('t.xlsx', buf.read())},
                                **self.auth(admin))
        self.assertEqual(resp.status_code, 200)
        d = resp.json()['data']
        self.assertEqual(d['created'], 0)
        self.assertEqual(d['errors'], [])

    def test_records_filter_by_payment_date_and_unpaid(self):
        admin = self.make_user('13900000077', 'finance_director', role='super_admin')
        project = self.create_project()

        paid = self.create_record(project)
        ARPayment.objects.create(ar_record=paid, payment_no=1,
                                 amount=Decimal('200.00'), payment_date=date(2026, 6, 10))
        # second record with no payments at all
        unpaid = ARRecord.objects.create(project=project, operation_year=2026,
                                         operation_month=7, estimated_amount=Decimal('500.00'))

        def ids(params):
            resp = self.client.get('/api/pk/ar/records', params, **self.auth(admin))
            self.assertEqual(resp.status_code, 200)
            return {r['id'] for r in resp.json()['data']['items']}

        # pay date range covering the payment → only the paid record
        in_range = ids({'pay_start': '2026-06-01', 'pay_end': '2026-06-30'})
        self.assertIn(paid.id, in_range)
        self.assertNotIn(unpaid.id, in_range)

        # pay date range outside the payment → neither record
        out_range = ids({'pay_start': '2026-08-01', 'pay_end': '2026-08-31'})
        self.assertNotIn(paid.id, out_range)

        # unpaid filter → only the record with no payments
        unpaid_only = ids({'pay_status': 'unpaid'})
        self.assertIn(unpaid.id, unpaid_only)
        self.assertNotIn(paid.id, unpaid_only)

        # date range + pay_include_unpaid=1 → OR: June-paid OR outstanding_amount>0
        combined = ids({'pay_start': '2026-06-01', 'pay_end': '2026-06-30', 'pay_include_unpaid': '1'})
        self.assertIn(paid.id, combined)    # has June payment
        self.assertIn(unpaid.id, combined)  # outstanding > 0

    def test_records_server_side_sort(self):
        """服务端排序：白名单字段升/降序、空值排末尾、非法 sort 安全回退。"""
        admin = self.make_user('13900000079', 'finance_director', role='super_admin')
        project = self.create_project()
        # actual_invoice_amount 可空且非派生，用于验证升降序 + 空值排末尾
        a = ARRecord.objects.create(project=project, operation_year=2026, operation_month=5,
                                    estimated_amount=Decimal('300.00'), actual_invoice_amount=Decimal('30.00'))
        b = ARRecord.objects.create(project=project, operation_year=2026, operation_month=5,
                                    estimated_amount=Decimal('100.00'), actual_invoice_amount=Decimal('10.00'))
        c = ARRecord.objects.create(project=project, operation_year=2026, operation_month=5,
                                    estimated_amount=Decimal('200.00'))  # actual_invoice 为空

        def order(params):
            resp = self.client.get('/api/pk/ar/records', params, **self.auth(admin))
            self.assertEqual(resp.status_code, 200, resp.content)
            return [r['id'] for r in resp.json()['data']['items']]

        self.assertEqual(order({'sort': 'estimated'}), [b.id, c.id, a.id])      # 升序
        self.assertEqual(order({'sort': '-estimated'}), [a.id, c.id, b.id])     # 降序
        # 空值排末尾：invoiced 升序时非空在前(b=10,a=30)、空值 c 殿后
        asc_inv = order({'sort': 'invoiced'})
        self.assertEqual(asc_inv, [b.id, a.id, c.id])
        self.assertEqual(asc_inv[-1], c.id)
        # 非法 sort 键安全忽略（仍返回三条，不报错）
        self.assertEqual(len(order({'sort': 'drop_table'})), 3)

    def test_bulk_delete_by_ids_and_by_filter(self):
        """批量删除：显式 ids + 选择全部筛选集(all+conditions)，级联回款、受权限约束。"""
        import json as _json
        admin = self.make_user('13900000083', 'finance_director', role='super_admin')
        project = self.create_project()
        r1 = ARRecord.objects.create(project=project, operation_year=2026, operation_month=5,
                                     estimated_amount=Decimal('100.00'))
        ARPayment.objects.create(ar_record=r1, payment_no=1, amount=Decimal('10.00'),
                                 payment_date=date(2026, 5, 9))
        r2 = ARRecord.objects.create(project=project, operation_year=2026, operation_month=6,
                                     estimated_amount=Decimal('200.00'))
        r3 = ARRecord.objects.create(project=project, operation_year=2026, operation_month=7,
                                     estimated_amount=Decimal('300.00'))

        # 显式 ids 删除 r1（级联其回款）
        resp = self.client.post('/api/pk/ar/records/bulk-delete',
                                data=_json.dumps({'ids': [r1.id]}),
                                content_type='application/json', **self.auth(admin))
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data']['deleted'], 1)
        self.assertFalse(ARRecord.objects.filter(pk=r1.id).exists())
        self.assertEqual(ARPayment.objects.filter(ar_record_id=r1.id).count(), 0)

        # all + 条件：删除 预估>150 的全部（r2/r3），r 不在条件外
        conds = _json.dumps([{'t': 'amt', 'field': 'estimated_amount', 'op': 'gt', 'value': 150}])
        resp = self.client.post(f'/api/pk/ar/records/bulk-delete?conditions={conds}',
                                data=_json.dumps({'all': True}),
                                content_type='application/json', **self.auth(admin))
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data']['deleted'], 2)
        self.assertEqual(ARRecord.objects.filter(pk__in=[r2.id, r3.id]).count(), 0)

    def test_project_bulk_delete_by_ids_and_filter(self):
        """项目批量删除：显式 ids + 选择全部筛选集(all+dept)，级联应收明细、受权限约束。"""
        import json as _json
        admin = self.make_user('13900000084', 'finance_director', role='super_admin')
        p1 = self.create_project(short_name='P1')
        p2 = self.create_project(short_name='P2')
        p3 = self.create_project(short_name='P3', delivery_dept=self.other_dept)
        ARRecord.objects.create(project=p1, operation_year=2026, operation_month=5,
                                estimated_amount=Decimal('100.00'))

        # 显式 ids 删除 p1（级联其应收明细）
        resp = self.client.post('/api/pk/ar/projects/bulk-delete',
                                data=_json.dumps({'ids': [p1.id]}),
                                content_type='application/json', **self.auth(admin))
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data']['deleted'], 1)
        self.assertFalse(ARProject.objects.filter(pk=p1.id).exists())
        self.assertEqual(ARRecord.objects.filter(project_id=p1.id).count(), 0)

        # all + dept 筛选：仅删本部门(p2)，p3 在其他部门不受影响
        resp = self.client.post(f'/api/pk/ar/projects/bulk-delete?dept={self.dept}',
                                data=_json.dumps({'all': True}),
                                content_type='application/json', **self.auth(admin))
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data']['deleted'], 1)
        self.assertFalse(ARProject.objects.filter(pk=p2.id).exists())
        self.assertTrue(ARProject.objects.filter(pk=p3.id).exists())

    def test_project_bulk_delete_requires_permission(self):
        """无删除权限不得批量删除项目。"""
        import json as _json
        cfg = default_job_config('cashier')
        cfg['pages']['ar_projects'] = True
        cfg['can_delete'] = False
        JobPermission.objects.create(job_title='cashier', config=cfg)
        _invalidate_perm_cache('cashier')
        user = self.make_user('13910000097', 'cashier')
        p = self.create_project()
        resp = self.client.post('/api/pk/ar/projects/bulk-delete',
                                data=_json.dumps({'ids': [p.id]}),
                                content_type='application/json', **self.auth(user))
        self.assertEqual(resp.status_code, 403, resp.content)
        self.assertTrue(ARProject.objects.filter(pk=p.id).exists())

    def test_bulk_delete_allowed_with_can_delete(self):
        """非超管但 can_delete=true 应被允许批量删除（验证授权链路正确）。"""
        import json as _json
        cfg = default_job_config('cashier')
        cfg['pages']['ar_records'] = True
        cfg['can_delete'] = True
        JobPermission.objects.create(job_title='cashier', config=cfg)
        _invalidate_perm_cache('cashier')
        user = self.make_user('13910000098', 'cashier')
        rec = self.create_record(self.create_project())
        resp = self.client.post('/api/pk/ar/records/bulk-delete',
                                data=_json.dumps({'ids': [rec.id]}),
                                content_type='application/json', **self.auth(user))
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data']['deleted'], 1)
        self.assertFalse(ARRecord.objects.filter(pk=rec.id).exists())

    def test_bulk_delete_requires_permission(self):
        """无删除权限不得批量删除。"""
        import json as _json
        cfg = default_job_config('cashier')
        cfg['pages']['ar_records'] = True
        cfg['can_delete'] = False
        JobPermission.objects.create(job_title='cashier', config=cfg)
        _invalidate_perm_cache('cashier')
        user = self.make_user('13910000099', 'cashier')
        rec = self.create_record(self.create_project())
        resp = self.client.post('/api/pk/ar/records/bulk-delete',
                                data=_json.dumps({'ids': [rec.id]}),
                                content_type='application/json', **self.auth(user))
        self.assertEqual(resp.status_code, 403, resp.content)
        self.assertTrue(ARRecord.objects.filter(pk=rec.id).exists())

    def test_conditions_match_any_or_logic(self):
        """统一条件 + match=all(且)/any(或)：维度/金额条件组合。"""
        import json as _json
        admin = self.make_user('13900000082', 'finance_director', role='super_admin')
        project = self.create_project()
        # a: 未开票(无开票额、有未收余额)，预估 100
        a = ARRecord.objects.create(project=project, operation_year=2026, operation_month=5,
                                    estimated_amount=Decimal('100.00'))
        # b: 已开票，预估 200
        b = ARRecord.objects.create(project=project, operation_year=2026, operation_month=6,
                                    estimated_amount=Decimal('200.00'),
                                    actual_invoice_amount=Decimal('200.00'))

        def ids(conditions, match=None):
            p = {'conditions': _json.dumps(conditions)}
            if match:
                p['match'] = match
            resp = self.client.get('/api/pk/ar/records', p, **self.auth(admin))
            self.assertEqual(resp.status_code, 200, resp.content)
            return {r['id'] for r in resp.json()['data']['items']}

        conds = [{'t': 'dim', 'field': 'invoice_status', 'value': '未开票'},
                 {'t': 'amt', 'field': 'estimated_amount', 'op': 'gt', 'value': 150}]
        # 且：未开票 且 预估>150 → 空（a未开票但=100，b预估200但已开票）
        self.assertEqual(ids(conds, 'all'), set())
        # 或：未开票 或 预估>150 → a(未开票) + b(预估200) = 两条
        self.assertEqual(ids(conds, 'any'), {a.id, b.id})

    def test_dim_condition_multi_value_union(self):
        """列头枚举多选：dim 条件 value 为数组时各桶取并集（OR）。"""
        import json as _json
        admin = self.make_user('13900000089', 'finance_director', role='super_admin')
        project = self.create_project()
        today = date.today()
        # a: 逾期（有未收 + 应收到期已过）
        a = ARRecord.objects.create(project=project, operation_year=2026, operation_month=5,
                                    estimated_amount=Decimal('100.00'),
                                    due_date=today - timedelta(days=10))
        # b: 已结清（全额回款 → outstanding<=0）
        b = ARRecord.objects.create(project=project, operation_year=2026, operation_month=6,
                                    estimated_amount=Decimal('200.00'),
                                    actual_invoice_amount=Decimal('200.00'))
        ARPayment.objects.create(ar_record=b, payment_no=1, amount=Decimal('200.00'),
                                 payment_date=today)
        # c: 未到期（有未收 + 到期日在远期）
        c = ARRecord.objects.create(project=project, operation_year=2026, operation_month=7,
                                    estimated_amount=Decimal('300.00'),
                                    due_date=today + timedelta(days=120))

        def ids(value):
            conds = [{'t': 'dim', 'field': 'status', 'value': value}]
            resp = self.client.get('/api/pk/ar/records', {'conditions': _json.dumps(conds)},
                                   **self.auth(admin))
            self.assertEqual(resp.status_code, 200, resp.content)
            return {r['id'] for r in resp.json()['data']['items']}

        # 单值仍生效（向后兼容）
        self.assertEqual(ids('overdue'), {a.id})
        self.assertEqual(ids('settled'), {b.id})
        # 多选（数组）= 并集
        self.assertEqual(ids(['overdue', 'settled']), {a.id, b.id})
        self.assertEqual(ids(['overdue', 'not_due']), {a.id, c.id})

    def test_conditions_builder_date_and_amount(self):
        """条件构建器：金额运算符(≠0) + 回款日期相对区间(含/不含本月)。"""
        import json as _json
        admin = self.make_user('13900000081', 'finance_director', role='super_admin')
        project = self.create_project()
        today = date.today()
        this_m = date(today.year, today.month, 15)
        last_m = (date(today.year - 1, 12, 15) if today.month == 1
                  else date(today.year, today.month - 1, 15))

        # r1: 有未收余额(outstanding>0)，本月回款
        r1 = ARRecord.objects.create(project=project, operation_year=2026, operation_month=5,
                                     estimated_amount=Decimal('500.00'))
        ARPayment.objects.create(ar_record=r1, payment_no=1, amount=Decimal('100.00'), payment_date=this_m)
        # r2: 上月回款，无未收（全额回款）
        r2 = ARRecord.objects.create(project=project, operation_year=2026, operation_month=6,
                                     estimated_amount=Decimal('200.00'))
        ARPayment.objects.create(ar_record=r2, payment_no=1, amount=Decimal('200.00'), payment_date=last_m)

        def ids(conditions):
            resp = self.client.get('/api/pk/ar/records',
                                   {'conditions': _json.dumps(conditions)}, **self.auth(admin))
            self.assertEqual(resp.status_code, 200, resp.content)
            return {r['id'] for r in resp.json()['data']['items']}

        # 未收金额≠0 → 仅 r1
        ne0 = ids([{'t': 'amt', 'field': 'outstanding_amount', 'op': 'ne0'}])
        self.assertEqual(ne0, {r1.id})
        # 回款日期含本月 → 仅 r1
        in_m = ids([{'t': 'date', 'field': 'payment_date', 'range': 'this_month'}])
        self.assertEqual(in_m, {r1.id})
        # 回款日期不含本月 → 排除有本月回款的 r1 → 仅 r2
        not_m = ids([{'t': 'date', 'field': 'payment_date', 'range': 'this_month', 'exclude': True}])
        self.assertEqual(not_m, {r2.id})
        # 组合：回款不含本月 + 未收≠0 → 空（r2 无未收，r1 被日期排除）
        combo = ids([{'t': 'date', 'field': 'payment_date', 'range': 'this_month', 'exclude': True},
                     {'t': 'amt', 'field': 'outstanding_amount', 'op': 'ne0'}])
        self.assertEqual(combo, set())
        # 非法 conditions 安全忽略（返回全部 2 条）
        self.assertEqual(len(ids('not-a-list' and [{'t': 'bogus'}])), 2)

    def test_invoice_status_unbilled_filter_excludes_settled(self):
        # 开票跟踪「未开票」筛选不得包含已结清（已全额回款）的记录，与
        # ARRecord.invoice_status 属性一致（已结清优先级高于未开票）。
        admin = self.make_user('13900000078', 'finance_director', role='super_admin')
        project = self.create_project()

        # 未开票且仍有未收余额 → 应出现在「未开票」结果中
        unbilled = ARRecord.objects.create(project=project, operation_year=2026,
                                           operation_month=7, estimated_amount=Decimal('500.00'))

        # 未开票但已全额回款 → invoice_status 计为「已结清」，不应出现在「未开票」
        settled = ARRecord.objects.create(project=project, operation_year=2026,
                                          operation_month=8, estimated_amount=Decimal('300.00'))
        ARPayment.objects.create(ar_record=settled, payment_no=1,
                                 amount=Decimal('300.00'), payment_date=date(2026, 8, 15))
        settled.refresh_from_db()
        self.assertLessEqual(settled.outstanding_amount, Decimal('0'))
        self.assertEqual(settled.invoice_status, '已结清')

        resp = self.client.get('/api/pk/ar/records', {'invoice_status': '未开票'}, **self.auth(admin))
        self.assertEqual(resp.status_code, 200, resp.content)
        result_ids = {r['id'] for r in resp.json()['data']['items']}
        self.assertIn(unbilled.id, result_ids)
        self.assertNotIn(settled.id, result_ids)

        # KPI「待开票」笔数同样应排除已结清记录
        kpi = self.client.get('/api/pk/ar/records/kpi', **self.auth(admin))
        self.assertEqual(kpi.status_code, 200, kpi.content)
        inv = kpi.json()['data']['invoice']
        self.assertEqual(inv['pending'], 1)
        # 完成度 = (total−待开票)/total = (2−1)/2 = 50%，已结清算「已完成」，
        # 不再出现「完成度低却待开票为0」的矛盾
        self.assertEqual(inv['rate'], 50.0)

    def test_focus_param_matches_kpi_pending(self):
        # 金蝶式聚焦查询：focus=reconciliation/invoice/collection 只返回本环节待办子集，
        # 且条数与 KPI 进度条的 pending 分子完全一致（列表 = 进度待办，不打架）。
        admin = self.make_user('13900000079', 'finance_director', role='super_admin')
        project = self.create_project()

        # A: 未对账 + 未开票 + 有未收 → 三个 focus 都应命中
        a = ARRecord.objects.create(project=project, operation_year=2026,
                                    operation_month=7, estimated_amount=Decimal('500.00'))
        # B: 已开票（有发票额）但仍有未收 → recon 已对账口径、invoice 已开票，collection 仍待收
        b = ARRecord.objects.create(project=project, operation_year=2026,
                                    operation_month=8, estimated_amount=Decimal('400.00'),
                                    actual_invoice_amount=Decimal('400.00'),
                                    reconciliation_date=date(2026, 8, 1))
        # C: 已全额回款 → 已结清，任何 focus 都不应命中
        c = ARRecord.objects.create(project=project, operation_year=2026,
                                    operation_month=9, estimated_amount=Decimal('300.00'))
        ARPayment.objects.create(ar_record=c, payment_no=1,
                                 amount=Decimal('300.00'), payment_date=date(2026, 9, 10))
        c.refresh_from_db()
        self.assertLessEqual(c.outstanding_amount, Decimal('0'))

        def focus_ids(f):
            resp = self.client.get('/api/pk/ar/records', {'focus': f}, **self.auth(admin))
            self.assertEqual(resp.status_code, 200, resp.content)
            return {r['id'] for r in resp.json()['data']['items']}, resp.json()['data']['total']

        recon_ids, recon_total = focus_ids('reconciliation')
        inv_ids, inv_total = focus_ids('invoice')
        coll_ids, coll_total = focus_ids('collection')

        # 对账待办：仅 A（未对账且未开票且有未收）
        self.assertEqual(recon_ids, {a.id})
        # 开票待办：仅 A（未开票且有未收）；B 已开票、C 已结清均排除
        self.assertEqual(inv_ids, {a.id})
        # 回款待办：A、B（仍有未收）；C 已结清排除
        self.assertEqual(coll_ids, {a.id, b.id})

        # 与 KPI pending 分子对齐
        kpi = self.client.get('/api/pk/ar/records/kpi', **self.auth(admin)).json()['data']
        self.assertEqual(recon_total, kpi['reconciliation']['pending'])
        self.assertEqual(inv_total, kpi['invoice']['pending'])
        self.assertEqual(coll_total, kpi['collection']['outstanding_count'])

    def test_kpi_respects_column_filters(self):
        # 列头筛选应同步进 KPI（进度条/待办与列表同口径）
        admin = self.make_user('13900000079', 'finance_director', role='super_admin')
        project = self.create_project()
        a = ARRecord.objects.create(project=project, operation_year=2026,
                                    operation_month=7, estimated_amount=Decimal('500.00'))
        ARRecord.objects.create(project=project, operation_year=2026,
                                operation_month=7, estimated_amount=Decimal('9000.00'))
        # 不带筛选：两条都未开票 → 待开票 2
        k0 = self.client.get('/api/pk/ar/records/kpi', **self.auth(admin)).json()['data']
        self.assertEqual(k0['invoice']['pending'], 2)
        # 列头筛选 估算金额 ≤ 1000：只剩 a 一条 → 待开票 1、total 1
        filters = json.dumps({'estimated_amount': {'op': 'lte', 'value': 1000}})
        k1 = self.client.get('/api/pk/ar/records/kpi', {'filters': filters},
                             **self.auth(admin)).json()['data']
        self.assertEqual(k1['total'], 1)
        self.assertEqual(k1['invoice']['pending'], 1)

    def test_column_filter_extended_text_operators(self):
        """金蝶式文本运算符：不等于/不包含/开头/结尾/为空/不为空/选值(in)。"""
        admin = self.make_user('13900000201', 'finance_director', role='super_admin')
        p_alpha = self.create_project(short_name='Alpha物流')
        p_beta = self.create_project(short_name='Beta仓储')
        p_gamma = self.create_project(short_name='Gamma物流')
        for p in (p_alpha, p_beta, p_gamma):
            ARRecord.objects.create(project=p, operation_year=2026,
                                    operation_month=7, estimated_amount=Decimal('100.00'))

        def names(filters):
            resp = self.client.get('/api/pk/ar/records', {'filters': json.dumps(filters)},
                                   **self.auth(admin))
            self.assertEqual(resp.status_code, 200)
            return sorted(r['short_name'] for r in resp.json()['data']['items'])

        # 开头是 "Alpha"
        self.assertEqual(names({'short_name': {'op': 'startswith', 'value': 'Alpha'}}), ['Alpha物流'])
        # 结尾是 "物流"
        self.assertEqual(names({'short_name': {'op': 'endswith', 'value': '物流'}}), ['Alpha物流', 'Gamma物流'])
        # 不包含 "物流"
        self.assertEqual(names({'short_name': {'op': 'not_contains', 'value': '物流'}}), ['Beta仓储'])
        # 不等于 "Beta仓储"
        self.assertEqual(names({'short_name': {'op': 'ne', 'value': 'Beta仓储'}}), ['Alpha物流', 'Gamma物流'])
        # 选值（Excel 自动筛选）：勾选 Alpha物流 + Beta仓储
        self.assertEqual(names({'short_name': {'op': 'in', 'value': ['Alpha物流', 'Beta仓储']}}),
                         ['Alpha物流', 'Beta仓储'])
        # 选值排除：not_in 排除 Alpha物流
        self.assertEqual(names({'short_name': {'op': 'not_in', 'value': ['Alpha物流']}}),
                         ['Beta仓储', 'Gamma物流'])

    def test_column_filter_empty_not_empty(self):
        """为空/不为空（一元运算符，不读 value）。"""
        admin = self.make_user('13900000202', 'finance_director', role='super_admin')
        p = self.create_project()
        with_notes = ARRecord.objects.create(project=p, operation_year=2026,
                                             operation_month=7, estimated_amount=Decimal('100.00'),
                                             notes='催过一次')
        ARRecord.objects.create(project=p, operation_year=2026, operation_month=7,
                                estimated_amount=Decimal('100.00'), notes='')

        def ids(filters):
            resp = self.client.get('/api/pk/ar/records', {'filters': json.dumps(filters)},
                                   **self.auth(admin))
            return {r['id'] for r in resp.json()['data']['items']}
        # 备注不为空 → 只剩有备注那条
        self.assertEqual(ids({'notes': {'op': 'not_empty'}}), {with_notes.id})
        # 备注为空 → 不含有备注那条
        self.assertNotIn(with_notes.id, ids({'notes': {'op': 'empty'}}))

    def test_distinct_values_endpoint(self):
        """列头「选值」清单：返回该列在可见数据里出现过的去重值。"""
        admin = self.make_user('13900000203', 'finance_director', role='super_admin')
        for nm in ('物流甲', '物流甲', '物流乙'):
            p = self.create_project(short_name=nm)
            ARRecord.objects.create(project=p, operation_year=2026,
                                    operation_month=7, estimated_amount=Decimal('100.00'))
        resp = self.client.get('/api/pk/ar/records/distinct-values',
                               {'field': 'short_name'}, **self.auth(admin))
        self.assertEqual(resp.status_code, 200)
        data = resp.json()['data']
        self.assertEqual(data['values'], ['物流乙', '物流甲'])   # 去重 + 排序
        self.assertFalse(data['truncated'])
        # 非白名单/计算列拒绝取值
        bad = self.client.get('/api/pk/ar/records/distinct-values',
                              {'field': 'outstanding_amount'}, **self.auth(admin))
        self.assertEqual(bad.status_code, 400)

    def test_filter_schemes_private_vs_public_visibility(self):
        """筛选方案：私有仅本人可见，公共团队共享；跨用户可见性 + 删除权限。"""
        a = self.make_user('13900000211', 'finance_director', role='super_admin')
        b = self.make_user('13900000212', 'finance_director', role='super_admin')
        conds = [{'t': 'dim', 'field': 'status', 'value': 'overdue'}]

        # A 建一个私有 + 一个公共
        r1 = self.json_post('/api/pk/ar/filter-schemes',
                            {'name': '我的逾期', 'scope': 'private', 'conditions': conds, 'match': 'all'}, a)
        self.assertEqual(r1.status_code, 200)
        r2 = self.json_post('/api/pk/ar/filter-schemes',
                            {'name': '团队逾期', 'scope': 'public', 'conditions': conds, 'match': 'all'}, a)
        self.assertEqual(r2.status_code, 200)
        pub_id = r2.json()['data']['id']
        priv_id = r1.json()['data']['id']

        # A 看到两个
        la = self.client.get('/api/pk/ar/filter-schemes', **self.auth(a)).json()['data']['items']
        self.assertEqual({s['name'] for s in la}, {'我的逾期', '团队逾期'})
        # B 只看到公共那个
        lb = self.client.get('/api/pk/ar/filter-schemes', **self.auth(b)).json()['data']['items']
        self.assertEqual([s['name'] for s in lb], ['团队逾期'])
        # 公共方案条件快照原样带回
        self.assertEqual(lb[0]['conditions'], conds)

        # B 无法访问/删除 A 的私有方案（不可见 → 403）
        d = self.client.delete(f'/api/pk/ar/filter-schemes/{priv_id}', **self.auth(b))
        self.assertEqual(d.status_code, 403)
        # 公共方案：A 可重命名（PUT）
        up = self.json_put(f'/api/pk/ar/filter-schemes/{pub_id}', {'name': '团队-逾期清单'}, a)
        self.assertEqual(up.status_code, 200)
        self.assertEqual(up.json()['data']['name'], '团队-逾期清单')

    def test_filter_scheme_default_follows_account(self):
        """默认方案存服务端、跟随账号：set-default 后列表标 is_default；方案删除自动清默认。"""
        a = self.make_user('13900000214', 'finance_director', role='super_admin')
        conds = [{'t': 'dim', 'field': 'status', 'value': 'overdue'}]
        sid = self.json_post('/api/pk/ar/filter-schemes',
                             {'name': '默认逾期', 'scope': 'private', 'conditions': conds}, a
                             ).json()['data']['id']
        # 设为默认
        r = self.json_post('/api/pk/ar/filter-schemes/set-default',
                           {'module': 'ar_records', 'scheme_id': sid}, a)
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.json()['data']['default_id'], sid)
        # 列表带回 default_id + 该方案 is_default=True（跨设备/会话只认服务端）
        lst = self.client.get('/api/pk/ar/filter-schemes', **self.auth(a)).json()['data']
        self.assertEqual(lst['default_id'], sid)
        self.assertTrue(next(s for s in lst['items'] if s['id'] == sid)['is_default'])
        # 取消默认
        self.json_post('/api/pk/ar/filter-schemes/set-default',
                       {'module': 'ar_records', 'scheme_id': None}, a)
        lst2 = self.client.get('/api/pk/ar/filter-schemes', **self.auth(a)).json()['data']
        self.assertIsNone(lst2['default_id'])
        # 重新设默认后删除方案 → 默认随级联清理
        self.json_post('/api/pk/ar/filter-schemes/set-default',
                       {'module': 'ar_records', 'scheme_id': sid}, a)
        self.client.delete(f'/api/pk/ar/filter-schemes/{sid}', **self.auth(a))
        lst3 = self.client.get('/api/pk/ar/filter-schemes', **self.auth(a)).json()['data']
        self.assertIsNone(lst3['default_id'])

    def test_filter_scheme_default_requires_visibility(self):
        """不能把别人的私有方案设为自己的默认。"""
        a = self.make_user('13900000215', 'finance_director', role='super_admin')
        b = self.make_user('13900000216', 'finance_director', role='super_admin')
        sid = self.json_post('/api/pk/ar/filter-schemes',
                             {'name': 'A私有', 'scope': 'private', 'conditions': []}, a
                             ).json()['data']['id']
        r = self.json_post('/api/pk/ar/filter-schemes/set-default',
                           {'module': 'ar_records', 'scheme_id': sid}, b)
        self.assertEqual(r.status_code, 403)

    def test_filter_scheme_same_name_overwrites(self):
        """同名同范围方案覆盖更新，不堆重复。"""
        a = self.make_user('13900000213', 'finance_director', role='super_admin')
        c1 = [{'t': 'amt', 'field': 'outstanding_amount', 'op': 'gt0'}]
        c2 = [{'t': 'dim', 'field': 'status', 'value': 'overdue'}]
        self.json_post('/api/pk/ar/filter-schemes',
                       {'name': '常用', 'scope': 'private', 'conditions': c1}, a)
        self.json_post('/api/pk/ar/filter-schemes',
                       {'name': '常用', 'scope': 'private', 'conditions': c2}, a)
        items = self.client.get('/api/pk/ar/filter-schemes', **self.auth(a)).json()['data']['items']
        same = [s for s in items if s['name'] == '常用']
        self.assertEqual(len(same), 1)            # 覆盖，不重复
        self.assertEqual(same[0]['conditions'], c2)   # 取最新快照

    def test_summary_not_inflated_by_multiple_payments(self):
        admin = self.make_user('13900000055', 'finance_director', role='super_admin')
        project = self.create_project()
        rec = ARRecord.objects.create(project=project, operation_year=2026,
                                      operation_month=8, estimated_amount=Decimal('1000.00'))
        # two payments on the same record — must not multiply estimated/count
        ARPayment.objects.create(ar_record=rec, payment_no=1,
                                 amount=Decimal('300.00'), payment_date=date(2026, 9, 1))
        ARPayment.objects.create(ar_record=rec, payment_no=2,
                                 amount=Decimal('200.00'), payment_date=date(2026, 9, 5))

        # filter by pay date range that matches both payments (JOIN fanout risk)
        resp = self.client.get('/api/pk/ar/records',
                               {'pay_start': '2026-09-01', 'pay_end': '2026-09-30'},
                               **self.auth(admin))
        self.assertEqual(resp.status_code, 200)
        s = resp.json()['data']['summary']
        self.assertEqual(s['count'], 1)
        self.assertEqual(Decimal(s['estimated']), Decimal('1000.00'))   # not 2000
        self.assertEqual(Decimal(s['collected']), Decimal('500.00'))    # 300 + 200

    def test_summary_period_fields_month_and_week(self):
        import calendar as cal
        admin = self.make_user('13900000044', 'finance_director', role='super_admin')
        project = self.create_project()

        # Record whose due_date is the last day of 2026-06
        last_june = date(2026, 6, 30)
        rec_june = ARRecord.objects.create(
            project=project, operation_year=2026, operation_month=6,
            estimated_amount=Decimal('1000.00'))
        ARRecord.objects.filter(pk=rec_june.pk).update(due_date=last_june)

        # Record whose due_date is in a different month (July)
        rec_july = ARRecord.objects.create(
            project=project, operation_year=2026, operation_month=7,
            estimated_amount=Decimal('2000.00'))
        ARRecord.objects.filter(pk=rec_july.pk).update(due_date=date(2026, 7, 15))

        # Payment on 2026-07-01 → inside the week (6/26~7/2, 周五~周四口径) but NOT in June month
        ARPayment.objects.create(ar_record=rec_june, payment_no=1,
                                 amount=Decimal('400.00'), payment_date=date(2026, 7, 1))
        # Payment on 2026-06-15 → inside June month but NOT in the week
        ARPayment.objects.create(ar_record=rec_june, payment_no=2,
                                 amount=Decimal('250.00'), payment_date=date(2026, 6, 15))

        # Request with year=2026 month=6 → ref_date=2026-06-30 (month end)
        resp = self.client.get('/api/pk/ar/records', {'year': 2026, 'month': 6},
                               **self.auth(admin))
        self.assertEqual(resp.status_code, 200)
        s = resp.json()['data']['summary']

        # 当期应收（期初存量）：due_date=2026-06-30 落在6月，rec_june 预估1000
        # 期前（payment_date < 2026-06-01）无回款 → 期初应收 = 1000（毛预估）
        self.assertEqual(Decimal(s['month_curr_est']), Decimal('1000.00'))
        # 当期调整：rec_june 无账实差额 → 0
        self.assertEqual(Decimal(s['month_curr_adjust']), Decimal('0'))
        self.assertEqual(s['ref_month'], '2026年6月')
        # 当期已收：payment_date 在6月(06-15) 且 ar_record.due_date>=mo_start(06-30>=06-01) → 250
        self.assertEqual(Decimal(s['month_curr_collected']), Decimal('250.00'))
        # 逾期应收：rec_june due_date=06-30 不早于mo_start(06-01) → 0（无逾期记录）
        self.assertEqual(Decimal(s['month_overdue_est']), Decimal('0'))
        # 逾期已收：payment_date 在6月(06-15) 且 ar_record.due_date<mo_start(06-30<06-01? No) → 0
        self.assertEqual(Decimal(s['month_overdue_collected']), Decimal('0'))

        # 周窗口（不跨月口径）：2026-06-30（周二）所在自然周 Fri=6/26～Thu=7/2
        # 月份裁剪后：wk=[6/26, 6/30]（末周 5 天，不延伸到 7 月）
        # rec_june due_date=06-30 落在该周; 期前(payment_date < 6/26) = 250(06-15)
        # week_est = 1000 − 250 = 750
        self.assertEqual(Decimal(s['week_est']), Decimal('750.00'))
        self.assertEqual(Decimal(s['week_adjust']), Decimal('0'))
        # week_collected: 周窗口 [6/26,6/30]；07-01 不在此区间，06-15 也不在 → 0
        self.assertEqual(Decimal(s['week_collected']), Decimal('0'))
        self.assertEqual(s['ref_week'], '6/26~6/30')          # 月份裁剪：不延伸到 7 月
        # 上周环比窗口：上个自然周 Fri=6/19～Thu=6/25（全在 6 月，无裁剪）
        self.assertEqual(s['prev_ref_week'], '6/19~6/25')
        self.assertEqual(Decimal(s['prev_week_est']), Decimal('0'))
        self.assertEqual(Decimal(s['prev_week_adjust']), Decimal('0'))
        self.assertEqual(Decimal(s['prev_week_collected']), Decimal('0'))

        # ── 关键回归：历史月份筛选（早于今天）时 ref_date 不应被 today 覆盖 ──
        # ref_date=2026-03-31（周二），自然周 Fri=3/27～Thu=4/2；月份裁剪到 [3/27,3/31]
        resp2 = self.client.get('/api/pk/ar/records', {'pay_end': '2026-03-31'},
                                **self.auth(admin))
        self.assertEqual(resp2.status_code, 200)
        s2 = resp2.json()['data']['summary']
        self.assertEqual(s2['ref_date'], '2026-03-31')
        self.assertEqual(s2['ref_month'], '2026年3月')
        self.assertEqual(s2['ref_week'], '3/27~3/31')          # 月份裁剪：不延伸到 4 月
        self.assertEqual(s2['prev_ref_week'], '3/20~3/26')     # 上周环比（全在 3 月）
        # 基准周非今天所在周 → 标签为"该周"
        self.assertEqual(s2['ref_week_label'], '该周')

        # 无任何日期筛选 → ref_date=today，周标签为"本周"
        resp3 = self.client.get('/api/pk/ar/records', **self.auth(admin))
        s3 = resp3.json()['data']['summary']
        self.assertEqual(s3['ref_date'], date.today().isoformat())
        self.assertEqual(s3['ref_week_label'], '本周')

    def test_summary_period_net_est_and_adjust(self):
        """时段合计：应收=期初存量（毛预估−期前已收−期前已调），调整=本期调，收=本期收。
        恒等式：期初应收 + 本期调整 − 本期已收 = 期末未收(outstanding)。"""
        from ar.models import ARAdjustment
        admin = self.make_user('13900000077', 'finance_director', role='super_admin')
        project = self.create_project()

        # 当月到期记录：预估1000，两次回款共300（均在6月），账实差额+50（6月）
        rec = ARRecord.objects.create(
            project=project, operation_year=2026, operation_month=6,
            estimated_amount=Decimal('1000.00'))
        ARRecord.objects.filter(pk=rec.pk).update(due_date=date(2026, 6, 30))
        ARPayment.objects.create(ar_record=rec, payment_no=1,
                                 amount=Decimal('200.00'), payment_date=date(2026, 6, 10))
        ARPayment.objects.create(ar_record=rec, payment_no=2,
                                 amount=Decimal('100.00'), payment_date=date(2026, 6, 20))
        ARAdjustment.objects.create(ar_record=rec, amount=Decimal('50.00'),
                                    reason='运费差', adjust_date=date(2026, 6, 12))
        rec.refresh_from_db()

        resp = self.client.get('/api/pk/ar/records', {'year': 2026, 'month': 6},
                               **self.auth(admin))
        s = resp.json()['data']['summary']
        # 应收(期初存量) = 毛预估1000 − 期前回款0 − 期前调整0 = 1000
        # （两笔回款 payment_date 均在6月内，不早于 mo_start=06-01，不计入"期前"）
        self.assertEqual(Decimal(s['month_curr_est']), Decimal('1000.00'))
        # 调整 = 本期调整合计 +50（adjust_date=06-12 在6月）
        self.assertEqual(Decimal(s['month_curr_adjust']), Decimal('50.00'))
        # 恒等式：期初应收 + 本期调整 − 本期已收 = 期末未收(outstanding)
        # 1000 + 50 − 300 = 750
        month_curr_collected = Decimal(s['month_curr_collected'])
        self.assertEqual(month_curr_collected, Decimal('300.00'))
        self.assertEqual(
            Decimal(s['month_curr_est']) + Decimal(s['month_curr_adjust']) - month_curr_collected,
            rec.outstanding_amount)
        self.assertEqual(rec.outstanding_amount, Decimal('750.00'))

    def test_summary_adjust_buckets_by_adjust_date(self):
        """差额调整按其 adjust_date 归入月/周，与记录到期日无关。"""
        from ar.models import ARAdjustment
        admin = self.make_user('13900000088', 'finance_director', role='super_admin')
        project = self.create_project()
        # 记录 6 月到期；两笔调整：一笔 6 月（当月），一笔 5 月（不应入当月）
        rec = ARRecord.objects.create(
            project=project, operation_year=2026, operation_month=6,
            estimated_amount=Decimal('1000.00'))
        ARRecord.objects.filter(pk=rec.pk).update(due_date=date(2026, 6, 30))
        ARAdjustment.objects.create(ar_record=rec, amount=Decimal('30.00'),
                                    reason='6月调整', adjust_date=date(2026, 6, 5))
        ARAdjustment.objects.create(ar_record=rec, amount=Decimal('70.00'),
                                    reason='5月调整', adjust_date=date(2026, 5, 20))

        resp = self.client.get('/api/pk/ar/records', {'year': 2026, 'month': 6},
                               **self.auth(admin))
        s = resp.json()['data']['summary']
        # 仅 6/5 那笔(30) 落在基准月，5/20 那笔(70)不计——尽管两笔都挂在同一条 6 月到期记录上
        self.assertEqual(Decimal(s['month_curr_adjust']), Decimal('30.00'))

    def test_records_search_matches_project_manager(self):
        admin = self.make_user('13900000066', 'finance_director', role='super_admin')
        project = self.create_project()  # project_manager 'PM A'
        rec = self.create_record(project)

        resp = self.client.get('/api/pk/ar/records', {'q': 'PM A'}, **self.auth(admin))
        self.assertEqual(resp.status_code, 200)
        self.assertIn(rec.id, {r['id'] for r in resp.json()['data']['items']})


class AdvanceModuleTests(TestCase):
    """预收预付：CRUD、核销重算、导入、现金流打通、KPI、权限遮蔽。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.dept = '劳务事业部'

    def tearDown(self):
        _invalidate_perm_cache()

    def make_user(self, phone, job_title, departments=None, role='operator'):
        u = PaikuanUser(phone=phone, name=f'{job_title} u', role=role, job_title=job_title,
                        departments=departments or [self.dept], is_active=True, is_approved=True)
        u.set_password('Test123456'); u.save()
        return u

    def auth(self, user):
        return {'HTTP_AUTHORIZATION': f'Bearer {make_token(user)}'}

    def create_project(self, short_name='预收项目A'):
        return ARProject.objects.create(
            customer_name='合同A', short_name=short_name, delivery_dept=self.dept,
            sales_contact='Sales A', project_manager='PM A', has_contract='有',
            contract_date=date(2026, 1, 1))

    def post(self, url, payload, user):
        return self.client.post(url, data=json.dumps(payload),
                                content_type='application/json', **self.auth(user))

    # ── 创建（挂项目 / 独立往来单位）────────────────────────────────────────────
    def test_create_with_and_without_project(self):
        admin = self.make_user('13911100001', 'finance_director', role='super_admin')
        proj = self.create_project()
        # 挂项目：交付部门自动取项目部门
        r1 = self.post('/api/pk/ar/advances', {
            'direction': '预收', 'project_id': proj.id, 'counterparty': '客户甲',
            'occur_year': 2026, 'occur_month': 3, 'occur_date': '2026-03-10',
            'advance_amount': 100000}, admin)
        self.assertEqual(r1.status_code, 200, r1.content)
        self.assertEqual(r1.json()['data']['delivery_dept'], self.dept)
        self.assertEqual(r1.json()['data']['balance_amount'], '100000.00')
        # 独立（预付供应商，无项目，手填部门）
        r2 = self.post('/api/pk/ar/advances', {
            'direction': '预付', 'delivery_dept': self.dept, 'counterparty': '供应商乙',
            'occur_year': 2026, 'occur_month': 3, 'occur_date': '2026-03-12',
            'advance_amount': 50000}, admin)
        self.assertEqual(r2.status_code, 200, r2.content)
        self.assertIsNone(r2.json()['data']['project_id'])
        # 方向校验
        bad = self.post('/api/pk/ar/advances', {
            'direction': 'X', 'delivery_dept': self.dept, 'occur_year': 2026,
            'occur_month': 3, 'advance_amount': 1}, admin)
        self.assertEqual(bad.status_code, 400)

    # ── 核销重算 + 余额非负 ────────────────────────────────────────────────────
    def test_writeoff_recompute_and_non_negative(self):
        admin = self.make_user('13911100002', 'finance_director', role='super_admin')
        rec = AdvanceRecord.objects.create(
            direction='预收', delivery_dept=self.dept, counterparty='客户甲',
            occur_year=2026, occur_month=3, occur_date=date(2026, 3, 10),
            advance_amount=Decimal('100000'))
        # 核销 30000 → 余额 70000，部分核销
        w = self.post(f'/api/pk/ar/advances/{rec.id}/writeoffs',
                      {'amount': 30000, 'writeoff_date': '2026-04-01'}, admin)
        self.assertEqual(w.status_code, 200, w.content)
        rec.refresh_from_db()
        self.assertEqual(rec.balance_amount, Decimal('70000.00'))
        self.assertEqual(rec.written_off_amount, Decimal('30000.00'))
        self.assertEqual(rec.writeoff_status, '部分核销')
        # 超额核销 → 拒绝（余额不能为负）
        over = self.post(f'/api/pk/ar/advances/{rec.id}/writeoffs',
                         {'amount': 999999, 'writeoff_date': '2026-04-02'}, admin)
        self.assertEqual(over.status_code, 400, over.content)
        rec.refresh_from_db()
        self.assertEqual(rec.balance_amount, Decimal('70000.00'))  # unchanged
        # 删除核销 → 余额恢复
        wid = w.json()['data']['id']
        d = self.client.delete(f'/api/pk/ar/advances/{rec.id}/writeoffs/{wid}', **self.auth(admin))
        self.assertEqual(d.status_code, 200)
        rec.refresh_from_db()
        self.assertEqual(rec.balance_amount, Decimal('100000.00'))
        self.assertEqual(rec.writeoff_status, '未核销')

    # ── 导入（含核销）+ 模板可解析 ─────────────────────────────────────────────
    def test_import_creates_advances_and_writeoff(self):
        admin = self.make_user('13911100003', 'finance_director', role='super_admin')
        self.create_project(short_name='预收项目A')
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(['方向(预收/预付)*', '项目简称', '交付部门', '往来单位*', '发生年*', '发生月*',
                   '款项日期', '预收/预付金额', '预计核销日期', '核销金额', '核销日期', '备注'])
        ws.append(['预收', '预收项目A', '', '客户甲', 2026, 3, '2026-03-10',
                   100000, '2026-06-30', 40000, '2026-04-01', ''])
        ws.append(['预付', '', self.dept, '供应商乙', 2026, 3, '2026-03-12', 50000, '', '', '', ''])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        buf.name = 'adv.xlsx'
        resp = self.client.post('/api/pk/ar/advances/import', {'file': buf}, **self.auth(admin))
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data']['created'], 2)
        recv = AdvanceRecord.objects.get(direction='预收', counterparty='客户甲')
        self.assertEqual(recv.balance_amount, Decimal('60000.00'))  # 100000 - 40000
        self.assertEqual(recv.delivery_dept, self.dept)

    # ── 收付差异 · 时间维度（月/周）分解 ────────────────────────────────────────
    def test_diff_timeline_month_and_week(self):
        admin = self.make_user('13911100009', 'finance_director', role='super_admin')
        proj = self.create_project(short_name='时间线项目')
        # 1月预收10万、2月预付4万、3月预收3万 → 累计差异应为 10/6/9（万）
        AdvanceRecord.objects.create(direction='预收', project=proj, delivery_dept=self.dept,
                                     counterparty='客户甲', occur_year=2026, occur_month=1,
                                     occur_date=date(2026, 1, 10), advance_amount=Decimal('100000'))
        AdvanceRecord.objects.create(direction='预付', project=proj, delivery_dept=self.dept,
                                     counterparty='供应商乙', occur_year=2026, occur_month=2,
                                     occur_date=date(2026, 2, 15), advance_amount=Decimal('40000'))
        AdvanceRecord.objects.create(direction='预收', project=proj, delivery_dept=self.dept,
                                     counterparty='客户丙', occur_year=2026, occur_month=3,
                                     occur_date=date(2026, 3, 5), advance_amount=Decimal('30000'))
        # 缺日期的记录 → 归入 null 桶
        AdvanceRecord.objects.create(direction='预收', project=proj, delivery_dept=self.dept,
                                     counterparty='客户丁', occur_year=2026, occur_month=4,
                                     occur_date=None, advance_amount=Decimal('5000'))

        r = self.client.get('/api/pk/ar/advances/diff-timeline',
                            {'grain': 'month'}, **self.auth(admin))
        self.assertEqual(r.status_code, 200, r.content)
        d = r.json()['data']
        self.assertEqual(d['grain'], 'month')
        # 连续补桶：1/2/3 月三期（含中间无收付的月也应在序列中）
        periods = {p['period']: p for p in d['periods']}
        self.assertIn('2026-01', periods)
        self.assertIn('2026-02', periods)
        self.assertIn('2026-03', periods)
        # 本期差异
        self.assertEqual(periods['2026-01']['diff'], '100000.00')
        self.assertEqual(periods['2026-02']['diff'], '-40000.00')
        self.assertEqual(periods['2026-03']['diff'], '30000.00')
        # 累计差异滚动求和：10 / 6 / 9 万
        self.assertEqual(periods['2026-01']['cum_diff'], '100000.00')
        self.assertEqual(periods['2026-02']['cum_diff'], '60000.00')
        self.assertEqual(periods['2026-03']['cum_diff'], '90000.00')
        # null 桶（缺日期）单列，不混入时间序
        self.assertIsNotNone(d['null_period'])
        self.assertEqual(d['null_period']['in_total'], '5000.00')
        # 汇总：总预收 13.5 万、总预付 4 万
        self.assertEqual(d['summary']['in_total'], '135000.00')
        self.assertEqual(d['summary']['out_total'], '40000.00')

        # 周粒度：仍能返回且 grain 标记正确
        rw = self.client.get('/api/pk/ar/advances/diff-timeline',
                            {'grain': 'week'}, **self.auth(admin))
        self.assertEqual(rw.status_code, 200, rw.content)
        self.assertEqual(rw.json()['data']['grain'], 'week')

        # 日期窗口过滤：仅 2 月 → 只剩预付一期
        rf = self.client.get('/api/pk/ar/advances/diff-timeline',
                            {'grain': 'month', 'start': '2026-02-01', 'end': '2026-02-28'},
                            **self.auth(admin))
        df = rf.json()['data']
        self.assertEqual(df['summary']['out_total'], '40000.00')
        self.assertEqual(df['summary']['in_total'], '0')

        # 导出 Excel：返回 xlsx 附件
        re = self.client.get('/api/pk/ar/advances/diff-timeline/export',
                            {'grain': 'month'}, **self.auth(admin))
        self.assertEqual(re.status_code, 200, re.content)
        self.assertIn('spreadsheet', re['Content-Type'])

    def test_diff_timeline_period_projects_detail(self):
        """按月/按周视图依赖：每期 projects 拆分须带项目级当期差异 + 逐笔明细，
        供前端选定某期后渲染与「按项目」一致的可展开表格。"""
        admin = self.make_user('13911100019', 'finance_director', role='super_admin')
        proj = self.create_project(short_name='透视项目')
        # 1月预收10万、2月预付4万、3月预收3万
        AdvanceRecord.objects.create(direction='预收', project=proj, delivery_dept=self.dept,
                                     counterparty='客户甲', occur_year=2026, occur_month=1,
                                     occur_date=date(2026, 1, 10), advance_amount=Decimal('100000'))
        AdvanceRecord.objects.create(direction='预付', project=proj, delivery_dept=self.dept,
                                     counterparty='供应商乙', occur_year=2026, occur_month=2,
                                     occur_date=date(2026, 2, 15), advance_amount=Decimal('40000'))
        AdvanceRecord.objects.create(direction='预收', project=proj, delivery_dept=self.dept,
                                     counterparty='客户丙', occur_year=2026, occur_month=3,
                                     occur_date=date(2026, 3, 5), advance_amount=Decimal('30000'))

        r = self.client.get('/api/pk/ar/advances/diff-timeline',
                            {'grain': 'month'}, **self.auth(admin))
        self.assertEqual(r.status_code, 200, r.content)
        d = r.json()['data']
        periods = {p['period']: p for p in d['periods']}
        # 1月：项目级当期差异 = +10 万，预收明细带日期/金额/往来单位
        jan = periods['2026-01']['projects']
        self.assertEqual(len(jan), 1)
        self.assertEqual(jan[0]['project'], '透视项目')
        self.assertEqual(jan[0]['diff'], '100000.00')
        self.assertEqual(len(jan[0]['in_items']), 1)
        self.assertEqual(jan[0]['in_items'][0]['amount'], '100000.00')
        self.assertEqual(jan[0]['in_items'][0]['counterparty'], '客户甲')
        self.assertEqual(jan[0]['out_items'], [])
        # 2月：项目级当期差异 = −4 万（预付侧）
        feb = periods['2026-02']['projects']
        self.assertEqual(feb[0]['diff'], '-40000.00')
        self.assertEqual(len(feb[0]['out_items']), 1)
        self.assertEqual(feb[0]['out_items'][0]['amount'], '40000.00')

        # 周粒度：仍按项目拆分并带当期差异
        rw = self.client.get('/api/pk/ar/advances/diff-timeline',
                            {'grain': 'week'}, **self.auth(admin))
        self.assertEqual(rw.status_code, 200, rw.content)
        dw = rw.json()['data']
        self.assertEqual(dw['grain'], 'week')
        self.assertTrue(any(p['projects'] for p in dw['periods']))

    # ── 现金流打通：净额含预收(流入)与预付(流出) ───────────────────────────────
    def test_cashflow_includes_advances(self):
        admin = self.make_user('13911100004', 'finance_director', role='super_admin')
        AdvanceRecord.objects.create(direction='预收', delivery_dept=self.dept,
                                     counterparty='客户甲', occur_year=2026, occur_month=3,
                                     occur_date=date(2026, 3, 10), advance_amount=Decimal('100000'))
        AdvanceRecord.objects.create(direction='预付', delivery_dept=self.dept,
                                     counterparty='供应商乙', occur_year=2026, occur_month=3,
                                     occur_date=date(2026, 3, 12), advance_amount=Decimal('30000'))
        resp = self.client.get('/api/pk/ar/cashflow',
                               {'start_date': '2026-03-01', 'end_date': '2026-03-31',
                                'depts': self.dept}, **self.auth(admin))
        self.assertEqual(resp.status_code, 200, resp.content)
        t = resp.json()['data']['totals']
        self.assertEqual(t['advance_received'][0], 100000.0)
        self.assertEqual(t['advance_paid'][0], 30000.0)
        self.assertEqual(t['inflow'][0], 100000.0)
        self.assertEqual(t['outflow'][0], 30000.0)
        self.assertEqual(t['net'][0], 70000.0)

    # ── KPI 分预收/预付 ────────────────────────────────────────────────────────
    def test_kpi_blocks(self):
        admin = self.make_user('13911100005', 'finance_director', role='super_admin')
        rec = AdvanceRecord.objects.create(direction='预收', delivery_dept=self.dept,
                                           counterparty='客户甲', occur_year=2026, occur_month=3,
                                           occur_date=date(2026, 3, 10), advance_amount=Decimal('100000'))
        AdvanceWriteoff.objects.create(advance_record=rec, writeoff_no=1,
                                       amount=Decimal('25000'), writeoff_date=date(2026, 4, 1))
        rec.recompute_derived()
        resp = self.client.get('/api/pk/ar/advances/kpi', **self.auth(admin))
        self.assertEqual(resp.status_code, 200, resp.content)
        b = resp.json()['data']['预收']
        self.assertEqual(b['advance_amount'], 100000.0)
        self.assertEqual(b['written_off'], 25000.0)
        self.assertEqual(b['balance'], 75000.0)
        self.assertEqual(b['writeoff_rate'], 25.0)

    # ── 字段级权限遮蔽：隐藏金额 ───────────────────────────────────────────────
    def test_field_permission_masks_amount(self):
        # 自定义职务：隐藏 adv_amount
        cfg = default_job_config('operator')
        cfg['ar_view']['adv_amount'] = False
        JobPermission.objects.update_or_create(job_title='operator', defaults={'config': cfg})
        _invalidate_perm_cache()
        user = self.make_user('13911100006', 'operator', role='operator')
        AdvanceRecord.objects.create(direction='预收', delivery_dept=self.dept,
                                     counterparty='客户甲', occur_year=2026, occur_month=3,
                                     occur_date=date(2026, 3, 10), advance_amount=Decimal('100000'))
        resp = self.client.get('/api/pk/ar/advances', **self.auth(user))
        self.assertEqual(resp.status_code, 200, resp.content)
        row = resp.json()['data']['items'][0]
        self.assertIsNone(row['advance_amount'])      # masked
        self.assertEqual(row['counterparty'], '客户甲')  # still visible

    # ── 可用预收/预付联动查询（供回款/排款页弹出）─────────────────────────────
    def test_available_lookup_by_project(self):
        admin = self.make_user('13911100007', 'finance_director', role='super_admin')
        proj = self.create_project()
        # 一笔部分核销的预收（余额 70000）+ 一笔已核销的预收（余额 0，应排除）
        a1 = AdvanceRecord.objects.create(
            direction='预收', project=proj, delivery_dept=self.dept, counterparty='客户甲',
            occur_year=2026, occur_month=3, occur_date=date(2026, 3, 10),
            advance_amount=Decimal('100000'))
        AdvanceWriteoff.objects.create(advance_record=a1, writeoff_no=1,
                                       amount=Decimal('30000'), writeoff_date=date(2026, 4, 1))
        a1.recompute_derived()
        a2 = AdvanceRecord.objects.create(
            direction='预收', project=proj, delivery_dept=self.dept, counterparty='客户甲',
            occur_year=2026, occur_month=2, occur_date=date(2026, 2, 1),
            advance_amount=Decimal('20000'))
        AdvanceWriteoff.objects.create(advance_record=a2, writeoff_no=1,
                                       amount=Decimal('20000'), writeoff_date=date(2026, 3, 1))
        a2.recompute_derived()
        # 一笔预付（方向不同，预收查询应排除）
        AdvanceRecord.objects.create(
            direction='预付', project=proj, delivery_dept=self.dept, counterparty='供应商乙',
            occur_year=2026, occur_month=3, occur_date=date(2026, 3, 12),
            advance_amount=Decimal('50000'))
        resp = self.client.get('/api/pk/ar/advances/available',
                               {'project_id': proj.id, 'direction': '预收'}, **self.auth(admin))
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['count'], 1)                 # 仅 a1 仍有余额
        self.assertEqual(d['total_balance'], '70000.00')
        self.assertEqual(d['items'][0]['id'], a1.id)
        # 预付方向
        resp_p = self.client.get('/api/pk/ar/advances/available',
                                 {'project_id': proj.id, 'direction': '预付'}, **self.auth(admin))
        self.assertEqual(resp_p.json()['data']['total_balance'], '50000.00')

    # ── 预收核销自动转回款（冲减应收）+ 现金流不重复计 ─────────────────────────
    def _ar_record(self, proj, est, year=2026, month=2):
        return ARRecord.objects.create(
            project=proj, operation_year=year, operation_month=month,
            estimated_amount=Decimal(str(est)))

    def test_writeoff_offsets_ar_record_and_reverses_on_delete(self):
        admin = self.make_user('13911100010', 'finance_director', role='super_admin')
        proj = self.create_project()
        ar = self._ar_record(proj, 100000)
        adv = AdvanceRecord.objects.create(
            direction='预收', project=proj, delivery_dept=self.dept, counterparty='客户甲',
            occur_year=2026, occur_month=3, occur_date=date(2026, 3, 10),
            advance_amount=Decimal('100000'))
        # 核销 40000 并冲抵该应收明细
        w = self.post(f'/api/pk/ar/advances/{adv.id}/writeoffs',
                      {'amount': 40000, 'writeoff_date': '2026-03-20',
                       'ar_record_id': ar.id}, admin)
        self.assertEqual(w.status_code, 200, w.content)
        wj = w.json()['data']
        self.assertEqual(wj['ar_record_id'], ar.id)
        self.assertIsNotNone(wj['ar_payment_id'])
        adv.refresh_from_db(); ar.refresh_from_db()
        self.assertEqual(adv.balance_amount, Decimal('60000.00'))   # 预收余额减少
        self.assertEqual(ar.outstanding_amount, Decimal('60000.00'))  # 应收 outstanding 减少
        pay = ARPayment.objects.get(pk=wj['ar_payment_id'])
        self.assertEqual(pay.source, '预收抵扣')
        self.assertEqual(pay.amount, Decimal('40000.00'))
        # 现金流：预收 100000 计入流入；预收抵扣 40000 不重复计现金（collected 排除）
        resp = self.client.get('/api/pk/ar/cashflow',
                               {'start_date': '2026-03-01', 'end_date': '2026-03-31',
                                'depts': self.dept}, **self.auth(admin))
        t = resp.json()['data']['totals']
        self.assertEqual(t['advance_received'][0], 100000.0)
        self.assertEqual(t['collected'][0], 0.0)       # 预收抵扣不计入现金回款
        self.assertEqual(t['inflow'][0], 100000.0)     # 不重复计
        # 删除核销 → 应收 outstanding 与预收余额均恢复，预收抵扣回款被删除
        d = self.client.delete(
            f'/api/pk/ar/advances/{adv.id}/writeoffs/{wj["id"]}', **self.auth(admin))
        self.assertEqual(d.status_code, 200, d.content)
        adv.refresh_from_db(); ar.refresh_from_db()
        self.assertEqual(adv.balance_amount, Decimal('100000.00'))
        self.assertEqual(ar.outstanding_amount, Decimal('100000.00'))
        self.assertFalse(ARPayment.objects.filter(pk=wj['ar_payment_id']).exists())

    def test_offset_amount_cannot_exceed_outstanding(self):
        admin = self.make_user('13911100011', 'finance_director', role='super_admin')
        proj = self.create_project()
        ar = self._ar_record(proj, 30000)
        adv = AdvanceRecord.objects.create(
            direction='预收', project=proj, delivery_dept=self.dept, counterparty='客户甲',
            occur_year=2026, occur_month=3, occur_date=date(2026, 3, 10),
            advance_amount=Decimal('100000'))
        w = self.post(f'/api/pk/ar/advances/{adv.id}/writeoffs',
                      {'amount': 40000, 'writeoff_date': '2026-03-20',
                       'ar_record_id': ar.id}, admin)
        self.assertEqual(w.status_code, 400, w.content)
        ar.refresh_from_db(); adv.refresh_from_db()
        self.assertEqual(ar.outstanding_amount, Decimal('30000.00'))  # 未变
        self.assertEqual(adv.balance_amount, Decimal('100000.00'))    # 未生成核销

    def test_offset_only_for_receivable_direction(self):
        admin = self.make_user('13911100012', 'finance_director', role='super_admin')
        proj = self.create_project()
        ar = self._ar_record(proj, 100000)
        adv = AdvanceRecord.objects.create(
            direction='预付', project=proj, delivery_dept=self.dept, counterparty='供应商乙',
            occur_year=2026, occur_month=3, occur_date=date(2026, 3, 10),
            advance_amount=Decimal('100000'))
        w = self.post(f'/api/pk/ar/advances/{adv.id}/writeoffs',
                      {'amount': 10000, 'writeoff_date': '2026-03-20',
                       'ar_record_id': ar.id}, admin)
        self.assertEqual(w.status_code, 400, w.content)

    def test_offset_payment_delete_reverses_writeoff(self):
        # 行为升级：删除「预收抵扣」回款 = 反向核销（删对应核销，预收余额恢复）；
        # PUT 编辑仍然禁止（金额/日期须在核销侧改）。
        admin = self.make_user('13911100013', 'finance_director', role='super_admin')
        proj = self.create_project()
        ar = self._ar_record(proj, 100000)
        adv = AdvanceRecord.objects.create(
            direction='预收', project=proj, delivery_dept=self.dept, counterparty='客户甲',
            occur_year=2026, occur_month=3, occur_date=date(2026, 3, 10),
            advance_amount=Decimal('100000'))
        w = self.post(f'/api/pk/ar/advances/{adv.id}/writeoffs',
                      {'amount': 40000, 'writeoff_date': '2026-03-20',
                       'ar_record_id': ar.id}, admin)
        pid = w.json()['data']['ar_payment_id']
        # PUT 仍禁止
        pu = self.client.put(f'/api/pk/ar/records/{ar.id}/payments/{pid}',
                             data=json.dumps({'amount': '1'}),
                             content_type='application/json', **self.auth(admin))
        self.assertEqual(pu.status_code, 400, pu.content)
        # DELETE = 反向核销
        d = self.client.delete(f'/api/pk/ar/records/{ar.id}/payments/{pid}', **self.auth(admin))
        self.assertEqual(d.status_code, 200, d.content)
        adv.refresh_from_db()
        self.assertEqual(adv.balance_amount, Decimal('100000'))
        self.assertEqual(adv.writeoffs.count(), 0)

    # ── 可用预收：本项目 ∪ 散单客户匹配 ────────────────────────────────────────
    def test_available_union_project_and_customer(self):
        admin = self.make_user('13911100014', 'finance_director', role='super_admin')
        proj = ARProject.objects.create(
            customer_name='合同P', short_name='项目P', delivery_dept=self.dept,
            sales_contact='S', project_manager='M')
        other = ARProject.objects.create(
            customer_name='合同Q', short_name='项目Q', delivery_dept=self.dept,
            sales_contact='S', project_manager='M')
        # A1：挂本项目
        a1 = AdvanceRecord.objects.create(
            direction='预收', project=proj, delivery_dept=self.dept, counterparty='ACME物流',
            occur_year=2026, occur_month=3, occur_date=date(2026, 3, 1),
            advance_amount=Decimal('50000'))
        # A2：散单（无项目），往来单位 == 客户名（精确匹配）
        a2 = AdvanceRecord.objects.create(
            direction='预收', delivery_dept=self.dept, counterparty='ACME物流',
            occur_year=2026, occur_month=3, occur_date=date(2026, 3, 2),
            advance_amount=Decimal('30000'))
        # A2b：散单但往来单位仅部分包含客户名 → 精确匹配下排除
        AdvanceRecord.objects.create(
            direction='预收', delivery_dept=self.dept, counterparty='ACME物流有限公司',
            occur_year=2026, occur_month=3, occur_date=date(2026, 3, 5),
            advance_amount=Decimal('11111'))
        # A3：散单但别的客户 → 排除
        AdvanceRecord.objects.create(
            direction='预收', delivery_dept=self.dept, counterparty='别的客户',
            occur_year=2026, occur_month=3, occur_date=date(2026, 3, 3),
            advance_amount=Decimal('99999'))
        # A4：挂另一项目（同客户名）→ 排除（不能冲抵本项目应收）
        AdvanceRecord.objects.create(
            direction='预收', project=other, delivery_dept=self.dept, counterparty='ACME物流',
            occur_year=2026, occur_month=3, occur_date=date(2026, 3, 4),
            advance_amount=Decimal('77777'))
        resp = self.client.get('/api/pk/ar/advances/available',
                               {'project_id': proj.id, 'customer': 'ACME物流',
                                'direction': '预收'}, **self.auth(admin))
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['count'], 2)
        self.assertEqual(d['total_balance'], '80000.00')
        ids = {it['id'] for it in d['items']}
        self.assertEqual(ids, {a1.id, a2.id})
        mt = {it['id']: it['match_type'] for it in d['items']}
        self.assertEqual(mt[a1.id], 'project')
        self.assertEqual(mt[a2.id], 'customer')

    def test_offset_with_standalone_advance(self):
        admin = self.make_user('13911100015', 'finance_director', role='super_admin')
        proj = ARProject.objects.create(
            customer_name='合同R', short_name='项目R', delivery_dept=self.dept,
            sales_contact='S', project_manager='M')
        ar = self._ar_record(proj, 100000)
        adv = AdvanceRecord.objects.create(   # 散单预收，无项目
            direction='预收', delivery_dept=self.dept, counterparty='ACME',
            occur_year=2026, occur_month=3, occur_date=date(2026, 3, 1),
            advance_amount=Decimal('60000'))
        w = self.post(f'/api/pk/ar/advances/{adv.id}/writeoffs',
                      {'amount': 60000, 'writeoff_date': '2026-03-20',
                       'ar_record_id': ar.id}, admin)
        self.assertEqual(w.status_code, 200, w.content)
        adv.refresh_from_db(); ar.refresh_from_db()
        self.assertEqual(adv.balance_amount, Decimal('0.00'))
        self.assertEqual(ar.outstanding_amount, Decimal('40000.00'))
        pay = ARPayment.objects.get(pk=w.json()['data']['ar_payment_id'])
        self.assertEqual(pay.source, '预收抵扣')

    def test_offsettable_records_by_customer(self):
        admin = self.make_user('13911100016', 'finance_director', role='super_admin')
        proj = ARProject.objects.create(
            customer_name='ACME物流', short_name='项目S', delivery_dept=self.dept,
            sales_contact='S', project_manager='M')
        other = ARProject.objects.create(
            customer_name='别的客户', short_name='项目T', delivery_dept=self.dept,
            sales_contact='S', project_manager='M')
        r1 = self._ar_record(proj, 50000)
        self._ar_record(other, 70000)  # 别的客户 → 不应出现
        resp = self.client.get('/api/pk/ar/advances/offsettable',
                               {'customer': 'ACME物流'}, **self.auth(admin))
        self.assertEqual(resp.status_code, 200, resp.content)
        items = resp.json()['data']['items']
        self.assertEqual([it['id'] for it in items], [r1.id])

    def test_available_lookup_respects_field_permission(self):
        cfg = default_job_config('operator')
        cfg['ar_view']['adv_amount'] = False
        JobPermission.objects.update_or_create(job_title='operator', defaults={'config': cfg})
        _invalidate_perm_cache()
        user = self.make_user('13911100008', 'operator', role='operator')
        proj = self.create_project()
        AdvanceRecord.objects.create(
            direction='预收', project=proj, delivery_dept=self.dept, counterparty='客户甲',
            occur_year=2026, occur_month=3, occur_date=date(2026, 3, 10),
            advance_amount=Decimal('100000'))
        resp = self.client.get('/api/pk/ar/advances/available',
                               {'project_id': proj.id}, **self.auth(user))
        self.assertEqual(resp.status_code, 403)         # 金额字段被遮蔽 → 拒绝


class ProjectImportRoundtripTests(TestCase):
    """项目台账导入：导出表头(不带*)应能再次导入；缺合同名时回退项目简称。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.dept = '供应链事业部'
        u = PaikuanUser(phone='13922200001', name='Admin', role='super_admin',
                        job_title='finance_director', departments=[self.dept],
                        is_active=True, is_approved=True)
        u.set_password('Test123456'); u.save()
        self.token = make_token(u)

    def tearDown(self):
        _invalidate_perm_cache()

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _upload(self, ws_rows):
        wb = openpyxl.Workbook(); ws = wb.active
        for r in ws_rows:
            ws.append(r)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0); buf.name = 'proj.xlsx'
        return self.client.post('/api/pk/ar/projects/import', {'file': buf}, **self.auth())

    def test_export_style_headers_without_asterisk_import(self):
        # 导出文件表头：无 *、开票模式/专票普票为短名 —— 之前会整表静默跳过
        headers = ['项目编号', '客户名称', '项目简称', '交付部门', '二级部门',
                   '业务模式', '客户等级', '销售对接人', '项目负责人', '共享业务', '有无合同',
                   '签订日期', '合同对账期(天)', '开票等待期(天)', '票后等待期(天)', '总账期(天)',
                   '开票模式', '专票/普票', '税率', '备注']
        row = ['GYL-X', '南京福佑物流合同', '南京福佑', self.dept, '',
               '整车', 'A级', '张三', '李四', '是', '有',
               '2026-01-01', 30, 0, 60, 90, '全额', '专票', '0.06', '']
        resp = self._upload([headers, row])
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['created'], 1, d)
        self.assertEqual(d['skipped'], 0, d)
        p = ARProject.objects.get(short_name='南京福佑')
        self.assertEqual(p.customer_name, '南京福佑物流合同')
        self.assertEqual(p.invoice_type, '专票')

    def test_missing_customer_name_falls_back_to_short_name(self):
        headers = ['合同名称*', '项目简称*', '交付部门*', '销售对接人*', '项目负责人*']
        row = ['', '南京福佑', self.dept, '张三', '李四']
        resp = self._upload([headers, row])
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['created'], 1, d)
        p = ARProject.objects.get(short_name='南京福佑')
        self.assertEqual(p.customer_name, '南京福佑')   # 回退用项目简称

    def test_same_customer_name_different_short_names_not_merged(self):
        """1合同多项目：同一合同名下不同项目简称必须各自建档，不能被合并去重
        （回归：去重主键须为项目简称+部门，而非合同名+部门）。"""
        headers = ['合同名称*', '项目简称*', '交付部门*', '销售对接人*', '项目负责人*']
        rows = [
            ['南京福佑在线电子商务有限公司', '南京福佑', self.dept, '张三', '李四'],
            ['南京福佑在线电子商务有限公司', '福佑顺心', self.dept, '张三', '李四'],
        ]
        resp = self._upload([headers] + rows)
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['created'], 2, d)      # 两个项目都应创建
        self.assertEqual(d['updated'], 0, d)
        names = set(ARProject.objects.filter(
            customer_name='南京福佑在线电子商务有限公司').values_list('short_name', flat=True))
        self.assertEqual(names, {'南京福佑', '福佑顺心'})

    def test_reimport_same_short_name_updates_not_duplicates(self):
        """同项目简称+部门重复导入应更新而非重复建档。"""
        headers = ['合同名称*', '项目简称*', '交付部门*', '销售对接人*', '项目负责人*', '业务模式']
        self._upload([headers, ['合同A', '项目X', self.dept, '张三', '李四', '整车']])
        resp = self._upload([headers, ['合同A改名', '项目X', self.dept, '王五', '赵六', '零担']])
        d = resp.json()['data']
        self.assertEqual(d['created'], 0, d)
        self.assertEqual(d['updated'], 1, d)
        self.assertEqual(ARProject.objects.filter(short_name='项目X', delivery_dept=self.dept).count(), 1)

    def test_row_with_data_but_no_name_reports_error(self):
        headers = ['合同名称*', '项目简称*', '交付部门*', '销售对接人*', '项目负责人*']
        row = ['', '', self.dept, '张三', '李四']   # 有数据但无名称
        resp = self._upload([headers, row])
        d = resp.json()['data']
        self.assertEqual(d['created'], 0, d)
        self.assertEqual(len(d['errors']), 1, d)
        self.assertIn('客户名称', d['errors'][0])


class InvoiceBatchTests(TestCase):
    """合并开票批次号：批量设置、按批次查询汇总、单条编辑。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.dept = '供应链事业部'
        u = PaikuanUser(phone='13933300001', name='Admin', role='super_admin',
                        job_title='finance_director', departments=[self.dept],
                        is_active=True, is_approved=True)
        u.set_password('Test123456'); u.save()
        self.token = make_token(u)
        self.proj = ARProject.objects.create(
            customer_name='合并开票测试合同', short_name='批次测试项目',
            delivery_dept=self.dept, sales_contact='S', project_manager='M')

    def tearDown(self):
        _invalidate_perm_cache()

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _rec(self, month, estimated=100000):
        r = ARRecord(project=self.proj, operation_year=2026, operation_month=month,
                     estimated_amount=Decimal(str(estimated)))
        r.save()
        return r

    def _put(self, url, payload):
        return self.client.put(url, data=json.dumps(payload),
                               content_type='application/json', **self.auth())

    def _post(self, url, payload):
        return self.client.post(url, data=json.dumps(payload),
                                content_type='application/json', **self.auth())

    def test_batch_assign_by_ids(self):
        r1 = self._rec(1); r2 = self._rec(2)
        resp = self._post('/api/pk/ar/records/batch-assign',
                          {'ids': [r1.id, r2.id], 'invoice_batch_no': 'PF-2026-001'})
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['updated'], 2)
        r1.refresh_from_db(); r2.refresh_from_db()
        self.assertEqual(r1.invoice_batch_no, 'PF-2026-001')
        self.assertEqual(r2.invoice_batch_no, 'PF-2026-001')

    def test_batch_assign_clear(self):
        r1 = self._rec(3)
        r1.invoice_batch_no = 'PF-2026-001'; r1.save()
        resp = self._post('/api/pk/ar/records/batch-assign',
                          {'ids': [r1.id], 'invoice_batch_no': ''})
        self.assertEqual(resp.status_code, 200)
        r1.refresh_from_db()
        self.assertEqual(r1.invoice_batch_no, '')

    def test_invoice_batches_endpoint(self):
        r1 = self._rec(4); r2 = self._rec(5)
        r1.invoice_batch_no = 'PF-2026-002'; r1.save()
        r2.invoice_batch_no = 'PF-2026-002'; r2.save()
        resp = self.client.get('/api/pk/ar/records/invoice-batches', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        batches = resp.json()['data']['batches']
        self.assertEqual(len(batches), 1)
        b = batches[0]
        self.assertEqual(b['batch_no'], 'PF-2026-002')
        self.assertEqual(b['count'], 2)
        self.assertAlmostEqual(float(b['estimated']), 200000.0, places=2)

    def test_put_record_sets_batch_no(self):
        r1 = self._rec(6)
        resp = self._put(f'/api/pk/ar/records/{r1.id}',
                         {'invoice_batch_no': 'PF-2026-003'})
        self.assertEqual(resp.status_code, 200, resp.content)
        r1.refresh_from_db()
        self.assertEqual(r1.invoice_batch_no, 'PF-2026-003')

class ContractAndImportAmbiguityTests(TestCase):
    """合同实体（1合同多客户多项目 / 1项目多合同）+ 应收导入歧义处理。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.dept = '供应链事业部'
        u = PaikuanUser(phone='13944400001', name='Admin', role='super_admin',
                        job_title='finance_director', departments=[self.dept],
                        is_active=True, is_approved=True)
        u.set_password('Test123456'); u.save()
        self.token = make_token(u)

    def tearDown(self):
        _invalidate_perm_cache()

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _proj(self, short_name, customer_name='合同X'):
        return ARProject.objects.create(
            customer_name=customer_name, short_name=short_name,
            delivery_dept=self.dept, sales_contact='S', project_manager='M')

    def _upload_records(self, head, rows):
        from django.core.files.uploadedfile import SimpleUploadedFile
        wb = openpyxl.Workbook(); ws = wb.active; ws.append(head)
        for r in rows:
            ws.append(r)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return self.client.post('/api/pk/ar/records/import',
                                {'file': SimpleUploadedFile('r.xlsx', buf.read())}, **self.auth())

    # ── 合同多对多 ────────────────────────────────────────────────────────────
    def test_contract_many_customers_and_projects(self):
        c1 = Customer.objects.create(name='客户甲')
        c2 = Customer.objects.create(name='客户乙')
        p1 = self._proj('项目A'); p2 = self._proj('项目B')
        ct = Contract.objects.create(name='联合服务合同', delivery_dept=self.dept)
        ContractParty.objects.create(contract=ct, customer=c1, role='main')
        ContractParty.objects.create(contract=ct, customer=c2, role='sub')
        ContractProject.objects.create(contract=ct, project=p1)
        ContractProject.objects.create(contract=ct, project=p2)
        self.assertEqual(ct.customers.count(), 2)
        self.assertEqual(ct.projects.count(), 2)
        # 1项目多合同
        ct2 = Contract.objects.create(name='补充协议', delivery_dept=self.dept)
        ContractProject.objects.create(contract=ct2, project=p1)
        self.assertEqual(p1.contracts.count(), 2)

    # ── 导入歧义：多个同名项目且未消歧 → 整表拒绝 + 候选项 ─────────────────────
    def test_import_ambiguous_same_name_rejects_with_candidates(self):
        self._proj('同名项目', customer_name='合同甲')
        self._proj('同名项目', customer_name='合同乙')
        HEAD = ['项目编号', '项目简称*', '客户名称', '运作年*', '运作月*', '预估上账金额',
                '实际开票金额', '税额(差额模式手填)', '开票日期', '账实差额调整',
                '回款金额', '回款时间', '备注']
        d = self._upload_records(HEAD, [
            ['', '同名项目', '', 2026, 1, 100000, '', '', '', '', '', '', ''],
        ]).json()['data']
        self.assertTrue(d['rejected'])
        self.assertEqual(d['created'], 0)
        self.assertEqual(len(d['ambiguities']), 1)
        self.assertEqual(len(d['ambiguities'][0]['candidates']), 2)
        self.assertEqual(ARRecord.objects.count(), 0)

    # ── 项目编号精确指定 → 消除歧义、正常写入 ─────────────────────────────────
    def test_import_project_no_disambiguates(self):
        p1 = self._proj('同名项目', customer_name='合同甲')
        self._proj('同名项目', customer_name='合同乙')
        HEAD = ['项目编号', '项目简称*', '客户名称', '运作年*', '运作月*', '预估上账金额',
                '实际开票金额', '税额(差额模式手填)', '开票日期', '账实差额调整',
                '回款金额', '回款时间', '备注']
        d = self._upload_records(HEAD, [
            [p1.project_no, '同名项目', '', 2026, 1, 100000, '', '', '', '', '', '', ''],
        ]).json()['data']
        self.assertFalse(d.get('rejected'), d)
        self.assertEqual(d['created'], 1)
        self.assertEqual(ARRecord.objects.filter(project=p1).count(), 1)

    # ── 客户名称消歧 → 正常写入 ──────────────────────────────────────────────
    def test_import_customer_hint_disambiguates(self):
        self._proj('同名项目', customer_name='合同甲')
        p2 = self._proj('同名项目', customer_name='合同乙')
        HEAD = ['项目编号', '项目简称*', '客户名称', '运作年*', '运作月*', '预估上账金额',
                '实际开票金额', '税额(差额模式手填)', '开票日期', '账实差额调整',
                '回款金额', '回款时间', '备注']
        d = self._upload_records(HEAD, [
            ['', '同名项目', '合同乙', 2026, 1, 100000, '', '', '', '', '', '', ''],
        ]).json()['data']
        self.assertFalse(d.get('rejected'), d)
        self.assertEqual(d['created'], 1)
        self.assertEqual(ARRecord.objects.filter(project=p2).count(), 1)

    # ── 项目编号填了但不存在 → 拒绝 + 指导 ───────────────────────────────────
    def test_import_bad_project_no_rejects(self):
        self._proj('唯一项目')
        HEAD = ['项目编号', '项目简称*', '客户名称', '运作年*', '运作月*', '预估上账金额',
                '实际开票金额', '税额(差额模式手填)', '开票日期', '账实差额调整',
                '回款金额', '回款时间', '备注']
        d = self._upload_records(HEAD, [
            ['YS-99999999-9999', '唯一项目', '', 2026, 1, 100000, '', '', '', '', '', '', ''],
        ]).json()['data']
        self.assertTrue(d['rejected'])
        self.assertEqual(len(d['bad_nos']), 1)
        self.assertEqual(ARRecord.objects.count(), 0)

    # ── 纯增量：同项目同月多行不合并 ─────────────────────────────────────────
    def test_import_is_additive_no_merge(self):
        p = self._proj('增量项目')
        HEAD = ['项目编号', '项目简称*', '客户名称', '运作年*', '运作月*', '预估上账金额',
                '实际开票金额', '税额(差额模式手填)', '开票日期', '账实差额调整',
                '回款金额', '回款时间', '备注']
        d = self._upload_records(HEAD, [
            ['', '增量项目', '', 2026, 1, 100000, '', '', '', '', '', '', '第一笔'],
            ['', '增量项目', '', 2026, 1, 80000, '', '', '', '', '', '', '同月第二笔'],
        ]).json()['data']
        self.assertFalse(d.get('rejected'), d)
        self.assertEqual(d['created'], 2)
        self.assertEqual(ARRecord.objects.filter(project=p, operation_month=1).count(), 2)

    # ── 合同 CRUD API（含 parties / projects 关联同步）─────────────────────────
    def _cpost(self, url, payload):
        return self.client.post(url, data=json.dumps(payload),
                                content_type='application/json', **self.auth())

    def _cput(self, url, payload):
        return self.client.put(url, data=json.dumps(payload),
                               content_type='application/json', **self.auth())

    def test_contract_api_create_with_links(self):
        c1 = Customer.objects.create(name='API客户甲')
        c2 = Customer.objects.create(name='API客户乙')
        p1 = self._proj('API项目A'); p2 = self._proj('API项目B')
        resp = self._cpost('/api/pk/ar/contracts', {
            'name': 'API合同', 'delivery_dept': self.dept, 'amount': '500000',
            'parties': [{'customer_id': c1.id, 'role': 'main', 'share': '70'},
                        {'customer_id': c2.id, 'role': 'sub', 'share': '30'}],
            'projects': [{'project_id': p1.id, 'is_primary': True},
                         {'project_id': p2.id, 'is_primary': False}],
        })
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(len(d['parties']), 2)
        self.assertEqual(len(d['projects']), 2)
        ct = Contract.objects.get(pk=d['id'])
        self.assertEqual(ct.customers.count(), 2)
        self.assertEqual(ct.projects.count(), 2)

    def test_contract_api_update_replaces_parties(self):
        c1 = Customer.objects.create(name='U客户甲')
        c2 = Customer.objects.create(name='U客户乙')
        ct = Contract.objects.create(name='U合同', delivery_dept=self.dept)
        ContractParty.objects.create(contract=ct, customer=c1, role='main')
        resp = self._cput(f'/api/pk/ar/contracts/{ct.id}', {
            'parties': [{'customer_id': c2.id, 'role': 'main'}],
        })
        self.assertEqual(resp.status_code, 200, resp.content)
        ct.refresh_from_db()
        self.assertEqual(list(ct.customers.values_list('name', flat=True)), ['U客户乙'])

    def test_contract_api_partial_update_keeps_links(self):
        c1 = Customer.objects.create(name='K客户')
        ct = Contract.objects.create(name='K合同', delivery_dept=self.dept)
        ContractParty.objects.create(contract=ct, customer=c1, role='main')
        # 不传 parties/projects → 关联不应被清空
        resp = self._cput(f'/api/pk/ar/contracts/{ct.id}', {'notes': '仅改备注'})
        self.assertEqual(resp.status_code, 200, resp.content)
        ct.refresh_from_db()
        self.assertEqual(ct.customers.count(), 1)
        self.assertEqual(ct.notes, '仅改备注')

    def test_contract_api_list_and_delete(self):
        ct = Contract.objects.create(name='L合同', delivery_dept=self.dept)
        Customer.objects.create(name='L客户')
        resp = self.client.get('/api/pk/ar/contracts', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertTrue(any(x['name'] == 'L合同' for x in resp.json()['data']['items']))
        # delete
        resp2 = self.client.delete(f'/api/pk/ar/contracts/{ct.id}', **self.auth())
        self.assertEqual(resp2.status_code, 200, resp2.content)
        self.assertFalse(Contract.objects.filter(pk=ct.id).exists())

    def test_contract_no_auto_generated(self):
        ct = Contract.objects.create(name='自动编号合同', delivery_dept='运输事业部')
        self.assertTrue(ct.contract_no.startswith('P-YS-'), ct.contract_no)
        ct2 = Contract.objects.create(name='自动编号合同2', delivery_dept='运输事业部')
        # 同部门同日序号自增
        self.assertNotEqual(ct.contract_no, ct2.contract_no)
        # 显式指定则不覆盖
        ct3 = Contract.objects.create(name='手填编号', delivery_dept='运输事业部', contract_no='MY-001')
        self.assertEqual(ct3.contract_no, 'MY-001')

    def test_contract_no_auto_without_dept(self):
        ct = Contract.objects.create(name='无部门合同')
        self.assertTrue(ct.contract_no.startswith('P-'), ct.contract_no)
        self.assertNotIn('--', ct.contract_no)  # 无部门段不应出现空段

    def test_project_put_attaches_contracts_and_customer(self):
        # 客户按事业部隔离：显式挂的客户需与项目同部门
        cust = Customer.objects.create(name='挂靠客户', delivery_dept=self.dept)
        ct = Contract.objects.create(name='挂靠合同', delivery_dept=self.dept)
        p = self._proj('挂靠项目')
        resp = self._cput(f'/api/pk/ar/projects/{p.id}', {
            'customer_id': cust.id, 'contract_ids': [ct.id],
        })
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['customer_id'], cust.id)
        self.assertEqual(len(d['contracts']), 1)
        self.assertEqual(d['contracts'][0]['contract_id'], ct.id)
        # GET 详情也带 contracts
        g = self.client.get(f'/api/pk/ar/projects/{p.id}', **self.auth()).json()['data']
        self.assertEqual(len(g['contracts']), 1)

    def test_project_put_clearing_contracts(self):
        ct = Contract.objects.create(name='待清合同', delivery_dept=self.dept)
        p = self._proj('待清项目')
        ContractProject.objects.create(contract=ct, project=p)
        resp = self._cput(f'/api/pk/ar/projects/{p.id}', {'contract_ids': []})
        self.assertEqual(resp.status_code, 200, resp.content)
        p.refresh_from_db()
        self.assertEqual(p.contracts.count(), 0)

    # ── 导出带回款日期/金额 + 合并年月筛选 ────────────────────────────────────
    def test_export_includes_payment_date_amount(self):
        from openpyxl import load_workbook
        p = self._proj('导出回款项目')
        r = ARRecord.objects.create(project=p, operation_year=2026, operation_month=3,
                                    estimated_amount=Decimal('100000'))
        ARPayment.objects.create(ar_record=r, payment_no=1, amount=Decimal('30000'),
                                 payment_date=date(2026, 3, 20))
        ARPayment.objects.create(ar_record=r, payment_no=2, amount=Decimal('20000'),
                                 payment_date=date(2026, 4, 10))
        resp = self.client.get('/api/pk/ar/records/export', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        self.assertIn('回款日期', headers)
        self.assertIn('回款金额', headers)
        self.assertIn('已回款合计', headers)
        # 找到数据行，校验两笔回款并列 + 合计
        di, dl = headers.index('回款日期'), headers.index('回款金额')
        ti = headers.index('已回款合计')
        row = [c.value for c in ws[2]]
        self.assertIn('2026-03-20', str(row[di]))
        self.assertIn('2026-04-10', str(row[di]))
        self.assertIn('30000', str(row[dl]))
        self.assertEqual(float(row[ti]), 50000.0)

    def test_combined_year_month_condition(self):
        p = self._proj('年月筛选项目')
        ARRecord.objects.create(project=p, operation_year=2026, operation_month=1,
                                estimated_amount=Decimal('1000'))
        ARRecord.objects.create(project=p, operation_year=2026, operation_month=2,
                                estimated_amount=Decimal('2000'))
        conds = json.dumps([{'t': 'dim', 'field': 'operation_ym', 'value': '2026-01'}])
        resp = self.client.get('/api/pk/ar/records', {'conditions': conds, 'match': 'all'},
                               **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['total'], 1)
        self.assertEqual(d['items'][0]['operation_month'], 1)
        # summary 合计也只统计该月
        self.assertEqual(d['summary']['count'], 1)
        self.assertAlmostEqual(float(d['summary']['estimated']), 1000.0, places=2)

    def test_operation_ym_range_and_exclude(self):
        """运作年月支持区间(start~end)与含/不含反选。"""
        p = self._proj('年月区间项目')
        for ym in [(2025, 12), (2026, 1), (2026, 3), (2026, 6)]:
            ARRecord.objects.create(project=p, operation_year=ym[0], operation_month=ym[1],
                                    estimated_amount=Decimal('100'))

        def months(cond):
            conds = json.dumps([cond])
            resp = self.client.get('/api/pk/ar/records', {'conditions': conds}, **self.auth())
            self.assertEqual(resp.status_code, 200, resp.content)
            return sorted((r['operation_year'], r['operation_month'])
                          for r in resp.json()['data']['items'])

        # 区间 2026-01 ~ 2026-03 → 含 1月、3月
        self.assertEqual(months({'t': 'dim', 'field': 'operation_ym',
                                  'value': '2026-01', 'end': '2026-03'}),
                         [(2026, 1), (2026, 3)])
        # 跨年区间 2025-12 ~ 2026-01
        self.assertEqual(months({'t': 'dim', 'field': 'operation_ym',
                                  'value': '2025-12', 'end': '2026-01'}),
                         [(2025, 12), (2026, 1)])
        # 不含 2026-01 ~ 2026-03 → 排除区间内，留 2025-12 与 2026-06
        self.assertEqual(months({'t': 'dim', 'field': 'operation_ym',
                                  'value': '2026-01', 'end': '2026-03', 'exclude': True}),
                         [(2025, 12), (2026, 6)])
        # 仅 start（无 end）退化为单月
        self.assertEqual(months({'t': 'dim', 'field': 'operation_ym', 'value': '2026-06'}),
                         [(2026, 6)])

    def test_condition_group_parentheses(self):
        """条件组（括号）：(A 且 B) 或 C 复合筛选。"""
        p = self._proj('括号筛选项目')
        # a: 2026-01 预估 100；b: 2026-01 预估 500；c: 2026-09 预估 500
        a = ARRecord.objects.create(project=p, operation_year=2026, operation_month=1,
                                    estimated_amount=Decimal('100'))
        b = ARRecord.objects.create(project=p, operation_year=2026, operation_month=1,
                                    estimated_amount=Decimal('500'))
        c = ARRecord.objects.create(project=p, operation_year=2026, operation_month=9,
                                    estimated_amount=Decimal('500'))

        # (运作=2026-01 且 预估>200) 或 (运作=2026-09)
        conds = json.dumps([
            {'t': 'group', 'match': 'all', 'conditions': [
                {'t': 'dim', 'field': 'operation_ym', 'value': '2026-01'},
                {'t': 'amt', 'field': 'estimated_amount', 'op': 'gt', 'value': 200},
            ]},
            {'t': 'dim', 'field': 'operation_ym', 'value': '2026-09'},
        ])
        resp = self.client.get('/api/pk/ar/records', {'conditions': conds, 'match': 'any'},
                               **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        got = {r['id'] for r in resp.json()['data']['items']}
        self.assertEqual(got, {b.id, c.id})   # a 被组内 预估>200 排除
        self.assertNotIn(a.id, got)

    def test_same_month_same_project_multiple_rows_kept_distinct(self):
        """同项目同年同月多条不同金额：各自独立入库，绝不合并、不取第一条覆盖。"""
        p = self._proj('同月多笔项目')
        HEAD = ['项目编号', '项目简称*', '客户名称', '运作年*', '运作月*', '预估上账金额',
                '实际开票金额', '税额(差额模式手填)', '开票日期', '账实差额调整',
                '回款金额', '回款时间', '备注']
        d = self._upload_records(HEAD, [
            ['', '同月多笔项目', '', 2026, 4, 59545.44, '', '', '', '', '', '', '第一笔'],
            ['', '同月多笔项目', '', 2026, 4, 251264.44, '', '', '', '', '', '', '第二笔'],
            ['', '同月多笔项目', '', 2026, 4, 107377.24, '', '', '', '', '', '', '第三笔'],
        ]).json()['data']
        self.assertFalse(d.get('rejected'), d)
        self.assertEqual(d['created'], 3, d)
        recs = ARRecord.objects.filter(project=p, operation_year=2026, operation_month=4)
        self.assertEqual(recs.count(), 3)
        amounts = sorted(float(r.estimated_amount) for r in recs)
        self.assertEqual(amounts, [59545.44, 107377.24, 251264.44])   # 三个金额都在，未被覆盖
        total = sum(float(r.estimated_amount) for r in recs)
        self.assertAlmostEqual(total, 418187.12, places=2)


class CollectionWorkbenchTests(TestCase):
    """催款工作台：分桶/部门隔离/生成催款任务（去重 + 权限）。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.admin = PaikuanUser(phone='13800000001', name='超管', role='super_admin',
                                 job_title='', departments=[], is_active=True, is_approved=True)
        self.admin.set_password('Test123456'); self.admin.save()
        self.acct = PaikuanUser(phone='13800000002', name='结算会计', role='operator',
                                job_title='settlement_accountant',
                                departments=['运输事业部'], is_active=True, is_approved=True)
        self.acct.set_password('Test123456'); self.acct.save()
        self.viewer = PaikuanUser(phone='13800000003', name='出纳', role='operator',
                                  job_title='cashier', departments=['运输事业部'],
                                  is_active=True, is_approved=True)
        self.viewer.set_password('Test123456'); self.viewer.save()

        def proj(dept, name):
            return ARProject.objects.create(
                customer_name=f'{name}客户', short_name=name, delivery_dept=dept,
                sales_contact='销售甲', project_manager='经理乙')
        p_ys = proj('运输事业部', '运输项目')
        p_lw = proj('劳务事业部', '劳务项目')
        # 逾期 10 天（d30 桶）与逾期 100 天（d90p 桶）各一条 + 他部门一条
        from datetime import timedelta
        today = date.today()
        self.r1 = ARRecord.objects.create(project=p_ys, operation_year=2026, operation_month=1,
                                          estimated_amount=Decimal('1000'),
                                          due_date=today - timedelta(days=10))
        self.r2 = ARRecord.objects.create(project=p_ys, operation_year=2026, operation_month=2,
                                          estimated_amount=Decimal('2000'),
                                          due_date=today - timedelta(days=100))
        self.r3 = ARRecord.objects.create(project=p_lw, operation_year=2026, operation_month=1,
                                          estimated_amount=Decimal('3000'),
                                          due_date=today - timedelta(days=50))

    def tearDown(self):
        _invalidate_perm_cache()

    def auth(self, user):
        return {'HTTP_AUTHORIZATION': f'Bearer {make_token(user)}'}

    def test_buckets_and_total(self):
        res = self.client.get('/api/pk/ar/records/collection', **self.auth(self.admin))
        self.assertEqual(res.status_code, 200)
        d = res.json()['data']
        buckets = {b['key']: b['count'] for b in d['buckets']}
        self.assertEqual(buckets['d30'], 1)
        self.assertEqual(buckets['d60'], 1)
        self.assertEqual(buckets['d90p'], 1)
        self.assertEqual(d['total'], 3)

    def test_bucket_filter_narrows_items(self):
        res = self.client.get('/api/pk/ar/records/collection?bucket=d90p', **self.auth(self.admin))
        d = res.json()['data']
        self.assertEqual(d['total'], 1)
        self.assertEqual(d['items'][0]['id'], self.r2.id)
        # 分桶统计仍是全量口径
        self.assertEqual(sum(b['count'] for b in d['buckets']), 3)

    def test_dept_isolation(self):
        res = self.client.get('/api/pk/ar/records/collection', **self.auth(self.acct))
        d = res.json()['data']
        self.assertEqual(d['total'], 2)
        self.assertEqual({i['delivery_dept'] for i in d['items']}, {'运输事业部'})

    def test_dunning_create_dedupe_and_scope(self):
        # 结算会计（ar_can_create）可生成；跨部门 id 被静默过滤
        res = self.client.post(
            '/api/pk/ar/records/collection/dunning',
            data=json.dumps({'ids': [self.r1.id, self.r3.id]}),
            content_type='application/json', **self.auth(self.acct))
        self.assertEqual(res.status_code, 200)
        d = res.json()['data']
        self.assertEqual(d['created'], 1)        # 只有本部门的 r1
        self.assertEqual(d['missing'], 1)        # r3 不在可见范围
        item = ActionItem.objects.get(pk=d['action_ids'][0])
        self.assertEqual(item.category, 'collection')
        self.assertEqual(item.bu, '运输事业部')
        self.assertEqual(item.assignee, '销售甲')
        self.assertEqual(item.source_signal.get('ar_record_id'), self.r1.id)
        # 重复生成被跳过
        res2 = self.client.post(
            '/api/pk/ar/records/collection/dunning',
            data=json.dumps({'ids': [self.r1.id]}),
            content_type='application/json', **self.auth(self.acct))
        self.assertEqual(res2.json()['data']['skipped'], 1)
        # 工作台明细带已建标记
        res3 = self.client.get('/api/pk/ar/records/collection', **self.auth(self.acct))
        marked = {i['id']: i['open_action_id'] for i in res3.json()['data']['items']}
        self.assertEqual(marked[self.r1.id], item.id)

    def test_dunning_requires_write_permission(self):
        res = self.client.post(
            '/api/pk/ar/records/collection/dunning',
            data=json.dumps({'ids': [self.r1.id]}),
            content_type='application/json', **self.auth(self.viewer))
        self.assertEqual(res.status_code, 403)


class AuditLogTests(TestCase):
    """审计中间件 + 查询接口：写操作留痕、敏感字段脱敏、仅超管可查。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.admin = PaikuanUser(phone='13800000011', name='超管', role='super_admin',
                                 job_title='', departments=[], is_active=True, is_approved=True)
        self.admin.set_password('Test123456'); self.admin.save()
        self.op = PaikuanUser(phone='13800000012', name='普通用户', role='operator',
                              job_title='settlement_accountant', departments=['运输事业部'],
                              is_active=True, is_approved=True)
        self.op.set_password('Test123456'); self.op.save()

    def tearDown(self):
        _invalidate_perm_cache()

    def auth(self, user):
        return {'HTTP_AUTHORIZATION': f'Bearer {make_token(user)}'}

    def test_write_operations_are_logged_with_user(self):
        from paikuan.models import AuditLog
        self.client.post('/api/pk/ar/projects', data=json.dumps({
            'customer_name': '客户A', 'short_name': '项目A', 'delivery_dept': '运输事业部',
            'sales_contact': '甲', 'project_manager': '乙'}),
            content_type='application/json', **self.auth(self.op))
        log = AuditLog.objects.order_by('-id').first()
        self.assertIsNotNone(log)
        self.assertEqual(log.method, 'POST')
        self.assertEqual(log.path, '/api/pk/ar/projects')
        self.assertEqual(log.module, 'ar')
        self.assertEqual(log.user_name, '普通用户')
        self.assertEqual(log.payload.get('short_name'), '项目A')

    def test_login_password_masked(self):
        from paikuan.models import AuditLog
        self.client.post('/api/pk/login', data=json.dumps(
            {'phone': '13800000011', 'password': 'Test123456'}),
            content_type='application/json')
        log = AuditLog.objects.filter(path='/api/pk/login').order_by('-id').first()
        self.assertIsNotNone(log)
        self.assertEqual(log.payload.get('password'), '***')

    def test_get_requests_not_logged(self):
        from paikuan.models import AuditLog
        before = AuditLog.objects.count()
        self.client.get('/api/pk/ar/projects', **self.auth(self.op))
        self.assertEqual(AuditLog.objects.count(), before)

    def test_audit_api_super_admin_only(self):
        res = self.client.get('/api/pk/audit-logs', **self.auth(self.op))
        self.assertEqual(res.status_code, 403)
        res2 = self.client.get('/api/pk/audit-logs', **self.auth(self.admin))
        self.assertEqual(res2.status_code, 200)

    def test_login_attributed_by_phone(self):
        from paikuan.models import AuditLog
        # 登录成功与失败都应归因到对应账号（异常登录可审计）
        self.client.post('/api/pk/login', data=json.dumps(
            {'phone': '13800000012', 'password': 'Test123456'}),
            content_type='application/json')
        log = AuditLog.objects.filter(path='/api/pk/login').order_by('-id').first()
        self.assertEqual(log.user_name, '普通用户')
        self.assertEqual(log.status_code, 200)
        self.client.post('/api/pk/login', data=json.dumps(
            {'phone': '13800000012', 'password': 'wrong'}),
            content_type='application/json')
        log2 = AuditLog.objects.filter(path='/api/pk/login').order_by('-id').first()
        self.assertEqual(log2.user_name, '普通用户')
        self.assertGreaterEqual(log2.status_code, 400)

    def test_jwt_fallback_attribution_without_view(self):
        from paikuan.models import AuditLog
        # 404（无视图执行 → request.pk_user 不存在）也应通过 JWT 兜底归因
        self.client.post('/api/pk/ar/nonexist', data='{}',
                         content_type='application/json', **self.auth(self.op))
        log = AuditLog.objects.order_by('-id').first()
        self.assertEqual(log.status_code, 404)
        self.assertEqual(log.user_name, '普通用户')


class CashPoolTests(TestCase):
    """资金池：余额口径 / 刚性+在途分窗 / 预测 / 配置与调拨权限。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.today = date.today()
        from datetime import timedelta
        self.td = timedelta
        self.admin = PaikuanUser(phone='13800000021', name='超管', role='super_admin',
                                 job_title='', departments=[], is_active=True, is_approved=True)
        self.admin.set_password('Test123456'); self.admin.save()
        self.cashier = PaikuanUser(phone='13800000022', name='出纳', role='operator',
                                   job_title='cashier', departments=['运输事业部'],
                                   is_active=True, is_approved=True)
        self.cashier.set_password('Test123456'); self.cashier.save()

        proj = ARProject.objects.create(
            customer_name='池客户', short_name='池项目', delivery_dept='运输事业部',
            sales_contact='甲', project_manager='乙')
        # 记录A：1000，逾期20天；回款200 + 预收抵扣100 → 未收700（逾期在外余额）
        self.rec_a = ARRecord.objects.create(project=proj, operation_year=2026, operation_month=1,
                                             estimated_amount=Decimal('1000'),
                                             due_date=self.today - self.td(days=20))
        ARPayment.objects.create(ar_record=self.rec_a, payment_no=1, amount=Decimal('200'),
                                 payment_date=self.today - self.td(days=10), source='回款')
        ARPayment.objects.create(ar_record=self.rec_a, payment_no=2, amount=Decimal('100'),
                                 payment_date=self.today - self.td(days=9), source='预收抵扣')
        # 记录B：300，15天后到期 → 30天窗预期流入
        ARRecord.objects.create(project=proj, operation_year=2026, operation_month=2,
                                estimated_amount=Decimal('300'),
                                due_date=self.today + self.td(days=15))
        # 预收50 / 预付30
        AdvanceRecord.objects.create(delivery_dept='运输事业部', direction='预收',
                                     occur_year=2026, occur_month=5,
                                     occur_date=self.today - self.td(days=5),
                                     advance_amount=Decimal('50'))
        AdvanceRecord.objects.create(delivery_dept='运输事业部', direction='预付',
                                     occur_year=2026, occur_month=5,
                                     occur_date=self.today - self.td(days=3),
                                     advance_amount=Decimal('30'))
        # 排款500，已付80（实际流出），余420计划10天后付（30天窗刚性流出）
        from paikuan.models import Payment as PkPayment, PaymentInstallment, ApprovalRecord
        pay = PkPayment.objects.create(department='运输事业部', project_desc='池支出',
                                       payee='供应商', total_amount=Decimal('500'),
                                       planned_date=self.today + self.td(days=10))
        PaymentInstallment.objects.create(payment=pay, seq=1,
                                          pay_date=self.today - self.td(days=2),
                                          pay_amount=Decimal('80'))
        # 管道：已批待排700 + 审批中600
        ApprovalRecord.objects.create(applicant='甲', department='运输事业部',
                                      approval_number='A' * 21, summary='已批待排',
                                      amount=Decimal('700'), payee='供应商', status='approved')
        ApprovalRecord.objects.create(applicant='甲', department='运输事业部',
                                      approval_number='B' * 21, summary='审批中',
                                      amount=Decimal('600'), payee='供应商', status='pending')

    def tearDown(self):
        _invalidate_perm_cache()

    def auth(self, user):
        return {'HTTP_AUTHORIZATION': f'Bearer {make_token(user)}'}

    def _config(self, user=None, dept='运输事业部', amount='10000'):  # noqa: D401
        return self.client.post(
            '/api/pk/ar/pool/config',
            data=json.dumps({'delivery_dept': dept,
                             'initial_date': str(self.today - self.td(days=100)),
                             'initial_amount': amount}),
            content_type='application/json', **self.auth(user or self.admin))

    def test_config_super_admin_only(self):
        self.assertEqual(self._config(user=self.cashier).status_code, 403)
        res = self._config()
        self.assertEqual(res.status_code, 200)
        # upsert 覆盖
        res2 = self._config(amount='20000')
        self.assertEqual(res2.json()['data']['initial_amount'], '20000')
        from ar.models import CashPoolConfig
        self.assertEqual(CashPoolConfig.objects.count(), 1)

    def test_pool_balance_and_projection(self):
        self._config()
        self._config(dept='劳务事业部', amount='1000')   # 调拨两侧池子都须已配置
        # 调拨：劳务→运输 10，运输→劳务 5
        for f, t, amt in (('劳务事业部', '运输事业部', '10'), ('运输事业部', '劳务事业部', '5')):
            res = self.client.post('/api/pk/ar/pool/transfers',
                                   data=json.dumps({'from_dept': f, 'to_dept': t, 'amount': amt,
                                                    'transfer_date': str(self.today - self.td(days=1))}),
                                   content_type='application/json', **self.auth(self.admin))
            self.assertEqual(res.status_code, 200)

        res = self.client.get('/api/pk/ar/pool', **self.auth(self.admin))
        self.assertEqual(res.status_code, 200)
        pool = next(p for p in res.json()['data']['pools'] if p['dept'] == '运输事业部')
        # 10000期初 +200回款(预收抵扣100不算现金) +50预收 −80实付 −30预付 +10调入 −5调出
        self.assertEqual(Decimal(pool['balance']), Decimal('10145'))
        self.assertEqual(Decimal(pool['committed']['d30']), Decimal('420'))   # 500−80
        self.assertEqual(Decimal(pool['pipeline']['approved']), Decimal('700'))
        self.assertEqual(Decimal(pool['pipeline']['pending']), Decimal('600'))
        self.assertEqual(Decimal(pool['expected_in']['d30']), Decimal('300'))
        self.assertEqual(Decimal(pool['expected_in']['overdue_outstanding']), Decimal('700'))
        # 预判30天 = 10145 + 300 − 420；悲观口径再减管道700
        self.assertEqual(Decimal(pool['projection']['d30']), Decimal('10025'))
        self.assertEqual(Decimal(pool['projection']['d30_with_pipeline']), Decimal('9325'))
        self.assertEqual(pool['warning']['status'], 'ok')

    def test_dept_scoping_and_transfer_permission(self):
        self._config()
        res = self.client.get('/api/pk/ar/pool', **self.auth(self.cashier))
        self.assertEqual(res.status_code, 200)
        depts = [p['dept'] for p in res.json()['data']['pools']]
        self.assertEqual(depts, ['运输事业部'])
        # 非超管不可调拨
        res2 = self.client.post('/api/pk/ar/pool/transfers',
                                data=json.dumps({'from_dept': '运输事业部', 'to_dept': '劳务事业部',
                                                 'amount': '1'}),
                                content_type='application/json', **self.auth(self.cashier))
        self.assertEqual(res2.status_code, 403)

    def test_unconfigured_pool_flagged(self):
        res = self.client.get('/api/pk/ar/pool', **self.auth(self.admin))
        pool = next(p for p in res.json()['data']['pools'] if p['dept'] == '运输事业部')
        self.assertFalse(pool['configured'])


class AuditHardeningTests(TestCase):
    """体检修复回归：行动项隔离 / 分析扇出 / 导出闸口 / 登录限速 / 改密踢token / 催款闭环。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.today = date.today()
        from datetime import timedelta
        self.td = timedelta
        self.admin = PaikuanUser(phone='13800000031', name='超管', role='super_admin',
                                 job_title='', departments=[], is_active=True, is_approved=True)
        self.admin.set_password('Test1234x'); self.admin.save()
        self.acct = PaikuanUser(phone='13800000032', name='运输会计', role='operator',
                                job_title='settlement_accountant', departments=['运输事业部'],
                                is_active=True, is_approved=True)
        self.acct.set_password('Test1234x'); self.acct.save()
        self.cashier = PaikuanUser(phone='13800000033', name='出纳', role='operator',
                                   job_title='cashier', departments=['运输事业部'],
                                   is_active=True, is_approved=True)
        self.cashier.set_password('Test1234x'); self.cashier.save()

    def tearDown(self):
        _invalidate_perm_cache()

    def auth(self, user):
        return {'HTTP_AUTHORIZATION': f'Bearer {make_token(user)}'}

    # ── 1. 行动项详情隔离 ────────────────────────────────────────────────────
    def test_action_detail_dept_isolated_and_write_gated(self):
        other = ActionItem.objects.create(title='他部门事项', bu='劳务事业部')
        mine = ActionItem.objects.create(title='本部门事项', bu='运输事业部')
        glob = ActionItem.objects.create(title='全局事项', bu='')
        # 跨部门：读/改/删全部 403
        self.assertEqual(self.client.get(f'/api/pk/ar/actions/{other.id}',
                                         **self.auth(self.acct)).status_code, 403)
        self.assertEqual(self.client.put(f'/api/pk/ar/actions/{other.id}',
                                         data=json.dumps({'status': 'done'}),
                                         content_type='application/json',
                                         **self.auth(self.acct)).status_code, 403)
        self.assertEqual(self.client.delete(f'/api/pk/ar/actions/{other.id}',
                                            **self.auth(self.acct)).status_code, 403)
        # 本部门/全局：可读；有 ar_can_create 的结算会计可改
        self.assertEqual(self.client.get(f'/api/pk/ar/actions/{mine.id}',
                                         **self.auth(self.acct)).status_code, 200)
        self.assertEqual(self.client.get(f'/api/pk/ar/actions/{glob.id}',
                                         **self.auth(self.acct)).status_code, 200)
        self.assertEqual(self.client.put(f'/api/pk/ar/actions/{mine.id}',
                                         data=json.dumps({'status': 'in_progress'}),
                                         content_type='application/json',
                                         **self.auth(self.acct)).status_code, 200)
        # 只读岗位（出纳 can_create=False 且无 ar_can_create）改不动
        self.assertEqual(self.client.put(f'/api/pk/ar/actions/{mine.id}',
                                         data=json.dumps({'status': 'done'}),
                                         content_type='application/json',
                                         **self.auth(self.cashier)).status_code, 403)
        # 无删除权限删不动
        self.assertEqual(self.client.delete(f'/api/pk/ar/actions/{mine.id}',
                                            **self.auth(self.acct)).status_code, 403)

    # ── 2. 分析聚合扇出 ──────────────────────────────────────────────────────
    def test_analytics_no_payment_join_fanout(self):
        proj = ARProject.objects.create(customer_name='扇出客户', short_name='扇出项目',
                                        delivery_dept='运输事业部',
                                        sales_contact='甲', project_manager='扇出负责人')
        rec = ARRecord.objects.create(project=proj, operation_year=self.today.year,
                                      operation_month=1, estimated_amount=Decimal('1000'))
        # 同一条记录两笔回款：扇出 BUG 会把 estimated 算成 2000
        ARPayment.objects.create(ar_record=rec, payment_no=1, amount=Decimal('100'),
                                 payment_date=self.today)
        ARPayment.objects.create(ar_record=rec, payment_no=2, amount=Decimal('200'),
                                 payment_date=self.today)
        res = self.client.get(f'/api/pk/ar/analytics/by-pm?year={self.today.year}',
                              **self.auth(self.admin))
        row = next(r for r in res.json()['data'] if r['pm'] == '扇出负责人')
        self.assertEqual(row['estimated'], 1000.0)
        self.assertEqual(row['collected'], 300.0)

        res2 = self.client.get('/api/pk/ar/analytics/by-dept', **self.auth(self.admin))
        d = next(r for r in res2.json()['data']['rows'] if r['dept'] == '运输事业部')
        self.assertEqual(d['estimated'], 1000.0)
        self.assertEqual(d['collected'], 300.0)

    # ── 3. approval_export 三连：页面闸口 / 行数上限存在性由代码保证，测闸口 ──
    def test_approval_export_requires_page_permission(self):
        JobPermission.objects.create(job_title='cashier',
                                     config={'pages': {'approval_records': False}})
        _invalidate_perm_cache()
        res = self.client.get('/api/pk/approvals/export', **self.auth(self.cashier))
        self.assertEqual(res.status_code, 403)
        res2 = self.client.get('/api/pk/approvals/export', **self.auth(self.admin))
        self.assertEqual(res2.status_code, 200)

    # ── 4. 登录限速 ─────────────────────────────────────────────────────────
    def test_login_lockout_after_failures(self):
        for _ in range(5):
            r = self.client.post('/api/pk/login', data=json.dumps(
                {'phone': '13800000032', 'password': 'wrong'}),
                content_type='application/json')
            self.assertEqual(r.status_code, 401)
        # 第6次即使密码正确也被锁
        r6 = self.client.post('/api/pk/login', data=json.dumps(
            {'phone': '13800000032', 'password': 'Test1234x'}),
            content_type='application/json')
        self.assertEqual(r6.status_code, 429)
        # 其他账号不受影响
        r7 = self.client.post('/api/pk/login', data=json.dumps(
            {'phone': '13800000033', 'password': 'Test1234x'}),
            content_type='application/json')
        self.assertEqual(r7.status_code, 200)

    # ── 5. 改密码踢旧 token ──────────────────────────────────────────────────
    def test_password_change_invalidates_old_token(self):
        from django.utils import timezone as _tz
        old_headers = self.auth(self.acct)
        # 管理员重置密码会写 pwd_changed_at
        res = self.client.put(f'/api/pk/users/{self.acct.id}',
                              data=json.dumps({'password': 'Newpass123'}),
                              content_type='application/json', **self.auth(self.admin))
        self.assertEqual(res.status_code, 200)
        self.acct.refresh_from_db()
        self.assertIsNotNone(self.acct.pwd_changed_at)
        # 把改密时间推后2秒，规避同秒签发的边界，断言旧 token 失效逻辑生效
        self.acct.pwd_changed_at = _tz.now() + self.td(seconds=2)
        self.acct.save(update_fields=['pwd_changed_at'])
        r = self.client.get('/api/pk/me', **old_headers)
        self.assertEqual(r.status_code, 401)
        # 弱密码被拒
        res2 = self.client.put(f'/api/pk/users/{self.acct.id}',
                               data=json.dumps({'password': 'short1'}),
                               content_type='application/json', **self.auth(self.admin))
        self.assertEqual(res2.status_code, 400)

    # ── 6. 应收结清 → 催款任务自动完成；记录删除 → 自动忽略 ────────────────────
    def test_dunning_auto_close_on_settle_and_delete(self):
        proj = ARProject.objects.create(customer_name='闭环客户', short_name='闭环项目',
                                        delivery_dept='运输事业部',
                                        sales_contact='甲', project_manager='乙')
        rec = ARRecord.objects.create(project=proj, operation_year=2026, operation_month=1,
                                      estimated_amount=Decimal('100'),
                                      due_date=self.today - self.td(days=10))
        act = ActionItem.objects.create(title='催款A', category='collection', status='open',
                                        bu='运输事业部', source_signal={'ar_record_id': rec.id})
        ARPayment.objects.create(ar_record=rec, payment_no=1, amount=Decimal('100'),
                                 payment_date=self.today)   # 结清
        act.refresh_from_db()
        self.assertEqual(act.status, 'done')
        self.assertIn('已结清', act.description)

        rec2 = ARRecord.objects.create(project=proj, operation_year=2026, operation_month=2,
                                       estimated_amount=Decimal('50'),
                                       due_date=self.today - self.td(days=5))
        act2 = ActionItem.objects.create(title='催款B', category='collection', status='open',
                                         bu='运输事业部', source_signal={'ar_record_id': rec2.id})
        rec2.delete()
        act2.refresh_from_db()
        self.assertEqual(act2.status, 'dismissed')

    # ── 7. 级联删除恢复预收余额 ──────────────────────────────────────────────
    def test_record_cascade_restores_advance_balance(self):
        proj = ARProject.objects.create(customer_name='级联客户', short_name='级联项目',
                                        delivery_dept='运输事业部',
                                        sales_contact='甲', project_manager='乙')
        rec = ARRecord.objects.create(project=proj, operation_year=2026, operation_month=3,
                                      estimated_amount=Decimal('200'))
        adv = AdvanceRecord.objects.create(delivery_dept='运输事业部', direction='预收',
                                           occur_year=2026, occur_month=1,
                                           occur_date=self.today - self.td(days=30),
                                           advance_amount=Decimal('150'))
        pay = ARPayment.objects.create(ar_record=rec, payment_no=1, amount=Decimal('80'),
                                       payment_date=self.today, source='预收抵扣')
        AdvanceWriteoff.objects.create(advance_record=adv, writeoff_no=1,
                                       amount=Decimal('80'), writeoff_date=self.today,
                                       ar_record=rec, ar_payment=pay)
        adv.refresh_from_db()
        self.assertEqual(adv.balance_amount, Decimal('70.00'))
        # 删除应收记录 → 级联删回款 → 反向删核销 → 预收余额恢复
        rec.delete()
        adv.refresh_from_db()
        self.assertEqual(AdvanceWriteoff.objects.filter(advance_record=adv).count(), 0)
        self.assertEqual(adv.balance_amount, Decimal('150.00'))

    # ── 8. 资金池：?depts 范围 + 刚性待付扣预付冲抵 ─────────────────────────
    def test_pool_depts_scope_and_committed_offset(self):
        from ar.models import CashPoolConfig
        from paikuan.models import Payment as PkPayment
        for d in ('运输事业部', '劳务事业部'):
            CashPoolConfig.objects.create(delivery_dept=d,
                                          initial_date=self.today - self.td(days=30),
                                          initial_amount=Decimal('1000'))
        res = self.client.get('/api/pk/ar/pool?depts=运输事业部', **self.auth(self.admin))
        depts = [p['dept'] for p in res.json()['data']['pools']]
        self.assertEqual(depts, ['运输事业部'])
        # 排款 500，已付 0，预付冲抵 120 → 刚性待付 380
        PkPayment.objects.create(department='运输事业部', project_desc='x', payee='y',
                                 total_amount=Decimal('500'),
                                 planned_date=self.today + self.td(days=10),
                                 prepaid_offset_amount=Decimal('120'))
        res2 = self.client.get('/api/pk/ar/pool?depts=运输事业部', **self.auth(self.admin))
        pool = res2.json()['data']['pools'][0]
        self.assertEqual(Decimal(pool['committed']['d30']), Decimal('380'))


class CashPoolTransferGuardTests(TestCase):
    """池间调拨防呆：禁透支 / 禁早于期初 / 两侧须已配置。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.today = date.today()
        from datetime import timedelta
        self.td = timedelta
        self.admin = PaikuanUser(phone='13800000041', name='超管', role='super_admin',
                                 job_title='', departments=[], is_active=True, is_approved=True)
        self.admin.set_password('Test1234x'); self.admin.save()
        from ar.models import CashPoolConfig
        CashPoolConfig.objects.create(delivery_dept='运输事业部',
                                      initial_date=self.today - self.td(days=30),
                                      initial_amount=Decimal('100000'))
        CashPoolConfig.objects.create(delivery_dept='劳务事业部',
                                      initial_date=self.today - self.td(days=20),
                                      initial_amount=Decimal('50000'))

    def tearDown(self):
        _invalidate_perm_cache()

    def _post(self, body):
        return self.client.post('/api/pk/ar/pool/transfers', data=json.dumps(body),
                                content_type='application/json',
                                HTTP_AUTHORIZATION=f'Bearer {make_token(self.admin)}')

    def test_overdraft_rejected(self):
        # 池里 10 万，要调出 15 万 → 拒绝并报当前可用余额
        res = self._post({'from_dept': '运输事业部', 'to_dept': '劳务事业部',
                          'amount': '150000', 'transfer_date': str(self.today)})
        self.assertEqual(res.status_code, 400)
        self.assertIn('不足以调出', res.json()['error'])
        # 10 万以内放行
        res2 = self._post({'from_dept': '运输事业部', 'to_dept': '劳务事业部',
                           'amount': '60000', 'transfer_date': str(self.today)})
        self.assertEqual(res2.status_code, 200)
        # 再调 5 万：剩 4 万，不够 → 拒
        res3 = self._post({'from_dept': '运输事业部', 'to_dept': '劳务事业部',
                           'amount': '50000', 'transfer_date': str(self.today)})
        self.assertEqual(res3.status_code, 400)

    def test_date_before_initial_rejected(self):
        res = self._post({'from_dept': '运输事业部', 'to_dept': '劳务事业部',
                          'amount': '1000',
                          'transfer_date': str(self.today - self.td(days=25))})  # 晚于运输期初但早于劳务期初
        self.assertEqual(res.status_code, 400)
        self.assertIn('期初基准日', res.json()['error'])

    def test_unconfigured_dept_rejected(self):
        res = self._post({'from_dept': '运输事业部', 'to_dept': '阔展事业部',
                          'amount': '1000', 'transfer_date': str(self.today)})
        self.assertEqual(res.status_code, 400)
        self.assertIn('池配置', res.json()['error'])

    def test_future_transfer_date_rejected(self):
        res = self._post({'from_dept': '运输事业部', 'to_dept': '劳务事业部',
                          'amount': '1000',
                          'transfer_date': str(self.today + self.td(days=3))})
        self.assertEqual(res.status_code, 400)
        self.assertIn('不能晚于今天', res.json()['error'])


class CashPoolTransferWorkflowTests(TestCase):
    """调拨申请两阶段审批：发起 / 调出方审批 / 拒绝 / 防自批 / 撤回 / 生效校验。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.today = date.today()
        from datetime import timedelta
        self.td = timedelta
        self.admin = PaikuanUser(phone='13800000051', name='超管', role='super_admin',
                                 job_title='', departments=[], is_active=True, is_approved=True)
        self.admin.set_password('Test1234x'); self.admin.save()
        # 劳务侧申请人（gm_assistant 默认 can_create=True）
        self.laowu = PaikuanUser(phone='13800000052', name='劳务申请人', role='operator',
                                 job_title='gm_assistant', departments=['劳务事业部'],
                                 is_active=True, is_approved=True)
        self.laowu.set_password('Test1234x'); self.laowu.save()
        # 运输侧审批人
        self.yunshu = PaikuanUser(phone='13800000053', name='运输审批人', role='operator',
                                  job_title='gm_assistant', departments=['运输事业部'],
                                  is_active=True, is_approved=True)
        self.yunshu.set_password('Test1234x'); self.yunshu.save()
        # 运输侧只读岗（cashier can_create=False）
        self.cashier = PaikuanUser(phone='13800000054', name='运输出纳', role='operator',
                                   job_title='cashier', departments=['运输事业部'],
                                   is_active=True, is_approved=True)
        self.cashier.set_password('Test1234x'); self.cashier.save()
        from ar.models import CashPoolConfig
        CashPoolConfig.objects.create(delivery_dept='运输事业部',
                                      initial_date=self.today - self.td(days=30),
                                      initial_amount=Decimal('100000'))
        CashPoolConfig.objects.create(delivery_dept='劳务事业部',
                                      initial_date=self.today - self.td(days=30),
                                      initial_amount=Decimal('50000'))

    def tearDown(self):
        _invalidate_perm_cache()

    def auth(self, user):
        return {'HTTP_AUTHORIZATION': f'Bearer {make_token(user)}'}

    def _request_transfer(self, user, amount='20000', f='运输事业部', t='劳务事业部'):
        return self.client.post('/api/pk/ar/pool/transfers',
                                data=json.dumps({'from_dept': f, 'to_dept': t,
                                                 'amount': amount,
                                                 'transfer_date': str(self.today)}),
                                content_type='application/json', **self.auth(user))

    def _review(self, user, tid, action, notes=''):
        return self.client.post(f'/api/pk/ar/pool/transfers/{tid}/review',
                                data=json.dumps({'action': action, 'review_notes': notes}),
                                content_type='application/json', **self.auth(user))

    def _balance(self, dept):
        res = self.client.get('/api/pk/ar/pool', **self.auth(self.admin))
        pool = next(p for p in res.json()['data']['pools'] if p['dept'] == dept)
        return Decimal(pool['balance']), pool

    def test_dept_request_pending_no_balance_impact(self):
        # 劳务向运输申请调拨2万 → 待审批，不动两侧余额；计入运输的待批调拨出款
        res = self._request_transfer(self.laowu)
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.json()['data']['status'], 'pending')
        bal_y, pool_y = self._balance('运输事业部')
        self.assertEqual(bal_y, Decimal('100000'))
        self.assertEqual(Decimal(pool_y['pipeline']['transfer_out_pending']),
                         Decimal('20000'))
        bal_l, _ = self._balance('劳务事业部')
        self.assertEqual(bal_l, Decimal('50000'))

    def test_from_dept_approve_moves_balance(self):
        tid = self._request_transfer(self.laowu).json()['data']['id']
        res = self._review(self.yunshu, tid, 'approve', '同意拆借')
        self.assertEqual(res.status_code, 200)
        d = res.json()['data']
        self.assertEqual(d['status'], 'approved')
        self.assertEqual(d['transfer_date'], str(self.today))  # 生效日=审批日
        self.assertEqual(d['reviewed_by_name'], '运输审批人')
        self.assertEqual(self._balance('运输事业部')[0], Decimal('80000'))
        self.assertEqual(self._balance('劳务事业部')[0], Decimal('70000'))

    def test_from_dept_reject_keeps_balance(self):
        tid = self._request_transfer(self.laowu).json()['data']['id']
        res = self._review(self.yunshu, tid, 'reject', '本月资金紧张，暂不同意')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.json()['data']['status'], 'rejected')
        self.assertEqual(self._balance('运输事业部')[0], Decimal('100000'))
        self.assertEqual(self._balance('劳务事业部')[0], Decimal('50000'))
        # 已处理的申请不能再审
        self.assertEqual(self._review(self.yunshu, tid, 'approve').status_code, 400)

    def test_review_permission_isolation(self):
        tid = self._request_transfer(self.laowu).json()['data']['id']
        # 调入方（劳务）不能审批；只读岗（运输出纳）无写入权限不能审批
        self.assertEqual(self._review(self.laowu, tid, 'approve').status_code, 403)
        self.assertEqual(self._review(self.cashier, tid, 'approve').status_code, 403)
        # 超管可以审
        self.assertEqual(self._review(self.admin, tid, 'approve').status_code, 200)

    def test_self_review_forbidden(self):
        # 运输审批人自己发起的调出申请，不能自批；超管可代批
        tid = self._request_transfer(self.yunshu).json()['data']['id']
        self.assertEqual(self._review(self.yunshu, tid, 'approve').status_code, 403)
        self.assertEqual(self._review(self.admin, tid, 'approve').status_code, 200)

    def test_request_must_relate_to_own_dept(self):
        # 劳务用户不能发起与劳务无关的调拨（运输→阔展）
        res = self._request_transfer(self.laowu, f='运输事业部', t='阔展事业部')
        self.assertEqual(res.status_code, 403)
        # 只读岗位（无写入权限）不能发起申请
        res2 = self._request_transfer(self.cashier)
        self.assertEqual(res2.status_code, 403)

    def test_approve_overdraft_rejected(self):
        # 申请15万 > 运输余额10万：申请可以提，批准时拦截
        tid = self._request_transfer(self.laowu, amount='150000').json()['data']['id']
        res = self._review(self.yunshu, tid, 'approve')
        self.assertEqual(res.status_code, 400)
        self.assertIn('不足以调出', res.json()['error'])
        from ar.models import CashPoolTransfer
        self.assertEqual(CashPoolTransfer.objects.get(pk=tid).status, 'pending')

    def test_requester_cancel_pending_only(self):
        tid = self._request_transfer(self.laowu).json()['data']['id']
        # 他人不能撤回
        self.assertEqual(self.client.delete(f'/api/pk/ar/pool/transfers/{tid}',
                                            **self.auth(self.yunshu)).status_code, 403)
        # 发起人可撤回
        self.assertEqual(self.client.delete(f'/api/pk/ar/pool/transfers/{tid}',
                                            **self.auth(self.laowu)).status_code, 200)
        # 已生效的只有超管能删
        tid2 = self._request_transfer(self.laowu).json()['data']['id']
        self._review(self.yunshu, tid2, 'approve')
        self.assertEqual(self.client.delete(f'/api/pk/ar/pool/transfers/{tid2}',
                                            **self.auth(self.laowu)).status_code, 403)
        self.assertEqual(self.client.delete(f'/api/pk/ar/pool/transfers/{tid2}',
                                            **self.auth(self.admin)).status_code, 200)

    def test_fixed_warning_amount_mode(self):
        # 手设资金预警线优先于按天数推算；余额低于预警线 → danger
        res = self.client.post('/api/pk/ar/pool/config',
                               data=json.dumps({'delivery_dept': '运输事业部',
                                                'initial_date': str(self.today - self.td(days=30)),
                                                'initial_amount': '100000',
                                                'warning_amount': '200000'}),
                               content_type='application/json', **self.auth(self.admin))
        self.assertEqual(res.status_code, 200)
        _, pool = self._balance('运输事业部')
        self.assertEqual(pool['warning']['mode'], 'fixed')
        self.assertEqual(Decimal(pool['warning']['amount']), Decimal('200000'))
        self.assertEqual(pool['warning']['status'], 'danger')

    def test_config_future_initial_date_rejected(self):
        res = self.client.post('/api/pk/ar/pool/config',
                               data=json.dumps({'delivery_dept': '运输事业部',
                                                'initial_date': str(self.today + self.td(days=1)),
                                                'initial_amount': '1'}),
                               content_type='application/json', **self.auth(self.admin))
        self.assertEqual(res.status_code, 400)
        self.assertIn('不能晚于今天', res.json()['error'])


class ChainIntegrityTests(TestCase):
    """全链路体检回归：客户删除守卫 / 换挂等级方向 / 预收预付方向锁 /
    开票金额配日期 / 导出公式转义。"""

    def setUp(self):
        _invalidate_perm_cache()
        self.client = Client()
        self.today = date.today()
        from datetime import timedelta
        self.td = timedelta
        self.admin = PaikuanUser(phone='13800000061', name='超管', role='super_admin',
                                 job_title='', departments=[], is_active=True, is_approved=True)
        self.admin.set_password('Test1234x'); self.admin.save()

    def tearDown(self):
        _invalidate_perm_cache()

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {make_token(self.admin)}'}

    # ── 客户名下有项目时禁删 ────────────────────────────────────────────────
    def test_customer_delete_blocked_when_projects_linked(self):
        proj = ARProject.objects.create(
            customer_name='守卫客户', short_name='守卫项目', delivery_dept='运输事业部',
            sales_contact='甲', project_manager='乙')
        proj.refresh_from_db()
        cid = proj.customer_id
        self.assertIsNotNone(cid)
        res = self.client.delete(f'/api/pk/ar/customers/{cid}', **self.auth())
        self.assertEqual(res.status_code, 400)
        self.assertIn('名下还有', res.json()['error'])
        # 删掉项目后客户可删
        proj.delete()
        res2 = self.client.delete(f'/api/pk/ar/customers/{cid}', **self.auth())
        self.assertEqual(res2.status_code, 200)

    # ── 换挂客户：等级以目标客户为准，不被旧等级冲掉 ────────────────────────
    def test_relink_mirrors_target_customer_level(self):
        from ar.models import Customer
        a = ARProject.objects.create(customer_name='客户A', short_name='项目A',
                                     delivery_dept='运输事业部', customer_level='C级',
                                     sales_contact='甲', project_manager='乙')
        b1 = ARProject.objects.create(customer_name='客户B', short_name='项目B1',
                                      delivery_dept='运输事业部', customer_level='S级',
                                      sales_contact='甲', project_manager='乙')
        cust_b = Customer.objects.get(name='客户B', delivery_dept='运输事业部')
        self.assertEqual(cust_b.level, 'S级')
        # 项目A 改名挂到 客户B：等级应镜像 B 的 S级，而不是把 C级 推给 B
        a.customer_name = '客户B'
        a.save()
        a.refresh_from_db(); cust_b.refresh_from_db(); b1.refresh_from_db()
        self.assertEqual(a.customer_id, cust_b.id)
        self.assertEqual(a.customer_level, 'S级')
        self.assertEqual(cust_b.level, 'S级')
        self.assertEqual(b1.customer_level, 'S级')

    # ── 预收/预付：已有核销不可翻方向 ───────────────────────────────────────
    def test_advance_direction_locked_after_writeoff(self):
        adv = AdvanceRecord.objects.create(delivery_dept='运输事业部', direction='预付',
                                           occur_year=2026, occur_month=1,
                                           occur_date=self.today - self.td(days=10),
                                           advance_amount=Decimal('100'))
        AdvanceWriteoff.objects.create(advance_record=adv, writeoff_no=1,
                                       amount=Decimal('40'), writeoff_date=self.today)
        res = self.client.put(f'/api/pk/ar/advances/{adv.id}',
                              data=json.dumps({'direction': '预收'}),
                              content_type='application/json', **self.auth())
        self.assertEqual(res.status_code, 400)
        self.assertIn('不能修改预收/预付方向', res.json()['error'])
        # 金额改到低于已核销额也被拦（余额不能为负）
        res2 = self.client.put(f'/api/pk/ar/advances/{adv.id}',
                               data=json.dumps({'advance_amount': '30'}),
                               content_type='application/json', **self.auth())
        self.assertEqual(res2.status_code, 400)

    # ── 开票金额必须配开票日期（创建与编辑两个入口）─────────────────────────
    def test_invoice_amount_requires_invoice_date(self):
        proj = ARProject.objects.create(customer_name='开票客户', short_name='开票项目',
                                        delivery_dept='运输事业部',
                                        sales_contact='甲', project_manager='乙')
        res = self.client.post('/api/pk/ar/records',
                               data=json.dumps({'project_id': proj.id,
                                                'operation_year': 2026, 'operation_month': 3,
                                                'estimated_amount': '100',
                                                'actual_invoice_amount': '100'}),
                               content_type='application/json', **self.auth())
        self.assertEqual(res.status_code, 400)
        self.assertIn('开票日期', res.json()['error'])
        res2 = self.client.post('/api/pk/ar/records',
                                data=json.dumps({'project_id': proj.id,
                                                 'operation_year': 2026, 'operation_month': 3,
                                                 'estimated_amount': '100',
                                                 'actual_invoice_amount': '100',
                                                 'invoice_date': str(self.today)}),
                                content_type='application/json', **self.auth())
        self.assertEqual(res2.status_code, 200)
        rid = res2.json()['data']['id']
        # 清空日期但留着金额 → 拦截；只改备注不追溯 → 放行
        res3 = self.client.put(f'/api/pk/ar/records/{rid}',
                               data=json.dumps({'invoice_date': ''}),
                               content_type='application/json', **self.auth())
        self.assertEqual(res3.status_code, 400)
        res4 = self.client.put(f'/api/pk/ar/records/{rid}',
                               data=json.dumps({'notes': '只改备注'}),
                               content_type='application/json', **self.auth())
        self.assertEqual(res4.status_code, 200)

    # ── 导出公式注入转义（集中出口覆盖所有 AR 导出）─────────────────────────
    def test_export_escapes_formula_cells(self):
        import openpyxl as _xl
        ARProject.objects.create(customer_name='=HYPERLINK("http://evil")',
                                 short_name='=1+1', delivery_dept='运输事业部',
                                 sales_contact='甲', project_manager='乙')
        res = self.client.get('/api/pk/ar/projects/export', **self.auth())
        self.assertEqual(res.status_code, 200)
        wb = _xl.load_workbook(io.BytesIO(res.content))
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        self.assertFalse(cell.value.startswith('='),
                                         f'未转义公式单元格: {cell.value!r}')


class OperationDateTests(TestCase):
    """运作年/月 → 运作日期（operation_date）全链路：创建、派生、导入、导出。"""

    def setUp(self):
        self.client = Client()
        self.dept = '运输事业部'
        admin = PaikuanUser(phone='13900000700', name='OpDateAdmin', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.token = make_token(admin)
        self.proj = ARProject.objects.create(
            customer_name='客户丙', short_name='日期化项目', delivery_dept=self.dept,
            sales_contact='S', project_manager='M', project_no='OPD-0001')

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _post(self, payload):
        return self.client.post('/api/pk/ar/records', data=json.dumps(payload),
                                content_type='application/json', **self.auth())

    # ── API 创建：传 operation_date，年/月自动派生 ────────────────────────
    def test_create_with_operation_date(self):
        resp = self._post({'project_id': self.proj.id, 'operation_date': '2026-05-18',
                           'estimated_amount': '1000'})
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['operation_date'], '2026-05-18')
        self.assertEqual(d['operation_year'], 2026)
        self.assertEqual(d['operation_month'], 5)

    # ── API 创建：旧调用方只传年/月 → 默认当月1日 ─────────────────────────
    def test_create_with_legacy_year_month_defaults_to_first_day(self):
        resp = self._post({'project_id': self.proj.id, 'operation_year': 2026,
                           'operation_month': 5, 'estimated_amount': '1000'})
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data']['operation_date'], '2026-05-01')

    # ── 模型层：仅给年/月直接建模（≈迁移回填语义）→ 派生当月1日 ──────────
    def test_model_save_derives_date_from_year_month(self):
        rec = ARRecord.objects.create(project=self.proj, operation_year=2025,
                                      operation_month=11, estimated_amount=Decimal('1'))
        self.assertEqual(str(rec.operation_date), '2025-11-01')

    # ── 导入：新模板「运作日期」列；只填到月按当月1日；坏日期整表拒绝 ─────
    def test_import_with_operation_date_column(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['项目简称*', '运作日期*', '预估上账金额'])
        ws.append(['日期化项目', '2026-05-18', 800])
        ws.append(['日期化项目', '2026-06', 900])     # 只填到月 → 06-01
        buf = io.BytesIO(); wb.save(buf); buf.seek(0); buf.name = 'r.xlsx'
        resp = self.client.post('/api/pk/ar/records/import', {'file': buf}, **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data'].get('created'), 2, resp.json())
        dates = sorted(str(r.operation_date) for r in ARRecord.objects.all())
        self.assertEqual(dates, ['2026-05-18', '2026-06-01'])

    def test_import_bad_date_rejected_with_guidance(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['项目简称*', '运作日期*', '预估上账金额'])
        ws.append(['日期化项目', '不是日期', 800])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0); buf.name = 'r.xlsx'
        resp = self.client.post('/api/pk/ar/records/import', {'file': buf}, **self.auth())
        self.assertEqual(resp.status_code, 200)
        d = resp.json()['data']
        self.assertTrue(d.get('rejected'), d)
        self.assertTrue(any('运作日期' in e for e in d['errors']), d['errors'])

    # ── 导入：旧模板「运作年/运作月」两列仍兼容 ──────────────────────────
    def test_import_legacy_year_month_columns_still_work(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['项目简称*', '运作年*', '运作月*', '预估上账金额'])
        ws.append(['日期化项目', 2026, 7, 700])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0); buf.name = 'r.xlsx'
        resp = self.client.post('/api/pk/ar/records/import', {'file': buf}, **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data'].get('created'), 1, resp.json())
        self.assertEqual(str(ARRecord.objects.first().operation_date), '2026-07-01')

    # ── 导出：包含「运作日期」单列，不再有年/月两列 ──────────────────────
    def test_export_has_operation_date_column(self):
        ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 18),
                                estimated_amount=Decimal('1'))
        resp = self.client.get('/api/pk/ar/records/export', **self.auth())
        self.assertEqual(resp.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        headers = [c.value for c in wb.active[1]]
        self.assertIn('运作日期', headers)
        self.assertNotIn('运作年', headers)
        self.assertNotIn('运作月', headers)


class CustomerProjectSyncTests(TestCase):
    """项目台账删除 → 客户名单同步收敛（孤儿客户随之清理）。"""

    def setUp(self):
        self.client = Client()
        self.dept = '运输事业部'
        admin = PaikuanUser(phone='13900000800', name='CustSyncAdmin', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.token = make_token(admin)

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _mk_proj(self, short_name, cust, no):
        return ARProject.objects.create(
            customer_name=cust.name, short_name=short_name, delivery_dept=self.dept,
            sales_contact='S', project_manager='M', project_no=no, customer=cust)

    # 单删：客户唯一项目删除后，客户随之从名单移除
    def test_delete_last_project_purges_customer(self):
        cust = Customer.objects.create(name='孤儿客户A', delivery_dept=self.dept)
        proj = self._mk_proj('同步项目A', cust, 'SYNC-0001')
        resp = self.client.delete(f'/api/pk/ar/projects/{proj.id}', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data'].get('purged_customers'), 1)
        self.assertFalse(Customer.objects.filter(pk=cust.pk).exists())

    # 单删：客户名下还有其他项目时保留
    def test_delete_one_of_many_projects_keeps_customer(self):
        cust = Customer.objects.create(name='多项目客户B', delivery_dept=self.dept)
        p1 = self._mk_proj('同步项目B1', cust, 'SYNC-0002')
        self._mk_proj('同步项目B2', cust, 'SYNC-0003')
        resp = self.client.delete(f'/api/pk/ar/projects/{p1.id}', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data'].get('purged_customers'), 0)
        self.assertTrue(Customer.objects.filter(pk=cust.pk).exists())

    # 批删：删尽的客户清理，未删尽的保留
    def test_bulk_delete_purges_only_fully_orphaned(self):
        c1 = Customer.objects.create(name='批删客户C1', delivery_dept=self.dept)
        c2 = Customer.objects.create(name='批删客户C2', delivery_dept=self.dept)
        p1 = self._mk_proj('批删项目C1', c1, 'SYNC-0004')
        p2 = self._mk_proj('批删项目C2a', c2, 'SYNC-0005')
        self._mk_proj('批删项目C2b', c2, 'SYNC-0006')   # c2 还剩一个项目
        resp = self.client.post('/api/pk/ar/projects/bulk-delete',
                                data=json.dumps({'ids': [p1.id, p2.id]}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['deleted'], 2)
        self.assertEqual(d.get('purged_customers'), 1)
        self.assertFalse(Customer.objects.filter(pk=c1.pk).exists())
        self.assertTrue(Customer.objects.filter(pk=c2.pk).exists())

    # 有合同关联的客户即使项目删尽也保留（避免断掉合同台账）
    def test_customer_with_contract_link_survives(self):
        cust = Customer.objects.create(name='有合同客户D', delivery_dept=self.dept)
        contract = Contract.objects.create(name='合同D', delivery_dept=self.dept)
        ContractParty.objects.create(contract=contract, customer=cust)
        proj = self._mk_proj('同步项目D', cust, 'SYNC-0007')
        resp = self.client.delete(f'/api/pk/ar/projects/{proj.id}', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data'].get('purged_customers'), 0)
        self.assertTrue(Customer.objects.filter(pk=cust.pk).exists())

    # 同步接口：历史遗留的孤儿客户一键清理
    def test_sync_endpoint_purges_existing_orphans(self):
        Customer.objects.create(name='历史孤儿E', delivery_dept=self.dept)
        cust_live = Customer.objects.create(name='在用客户F', delivery_dept=self.dept)
        self._mk_proj('在用项目F', cust_live, 'SYNC-0008')
        resp = self.client.post('/api/pk/ar/customers/sync-from-projects', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data'].get('purged_orphans'), 1)
        self.assertFalse(Customer.objects.filter(name='历史孤儿E').exists())
        self.assertTrue(Customer.objects.filter(pk=cust_live.pk).exists())

    # 批量删除客户：孤儿删除；有项目/合同关联的保护跳过并说明原因
    def test_customers_bulk_delete_with_protection(self):
        orphan = Customer.objects.create(name='可删客户G', delivery_dept=self.dept)
        busy = Customer.objects.create(name='有项目客户H', delivery_dept=self.dept)
        self._mk_proj('在用项目H', busy, 'SYNC-0009')
        contracted = Customer.objects.create(name='有合同客户I', delivery_dept=self.dept)
        contract = Contract.objects.create(name='合同I', delivery_dept=self.dept)
        ContractParty.objects.create(contract=contract, customer=contracted)
        resp = self.client.post('/api/pk/ar/customers/bulk-delete',
                                data=json.dumps({'ids': [orphan.id, busy.id, contracted.id]}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['deleted'], 1)
        self.assertEqual(d['skipped'], 2)
        self.assertTrue(any('项目' in r for r in d['skipped_reasons']), d)
        self.assertTrue(any('合同' in r for r in d['skipped_reasons']), d)
        self.assertFalse(Customer.objects.filter(pk=orphan.pk).exists())
        self.assertTrue(Customer.objects.filter(pk=busy.pk).exists())
        self.assertTrue(Customer.objects.filter(pk=contracted.pk).exists())

    def test_customers_bulk_delete_requires_ids(self):
        resp = self.client.post('/api/pk/ar/customers/bulk-delete',
                                data=json.dumps({}), content_type='application/json',
                                **self.auth())
        self.assertEqual(resp.status_code, 400)


class TargetCollectionDateTests(TestCase):
    """目标回款日期（target_collection_date，选填手工字段）全链路。"""

    def setUp(self):
        self.client = Client()
        self.dept = '运输事业部'
        admin = PaikuanUser(phone='13900000900', name='TgtDateAdmin', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.token = make_token(admin)
        self.proj = ARProject.objects.create(
            customer_name='客户T', short_name='目标日期项目', delivery_dept=self.dept,
            sales_contact='S', project_manager='M', project_no='TGT-0001')

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def test_create_and_update_target_date(self):
        resp = self.client.post('/api/pk/ar/records', data=json.dumps({
            'project_id': self.proj.id, 'operation_date': '2026-05-01',
            'estimated_amount': '1000', 'target_collection_date': '2026-06-15'}),
            content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['target_collection_date'], '2026-06-15')
        rid = d['id']
        # 更新与清空
        resp = self.client.put(f'/api/pk/ar/records/{rid}', data=json.dumps({
            'target_collection_date': ''}), content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertIsNone(resp.json()['data']['target_collection_date'])

    def test_import_with_target_date_column(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['项目简称*', '运作日期*', '预估上账金额', '目标回款日期'])
        ws.append(['目标日期项目', '2026-05-01', 800, '2026-06-20'])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0); buf.name = 'r.xlsx'
        resp = self.client.post('/api/pk/ar/records/import', {'file': buf}, **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data'].get('created'), 1, resp.json())
        rec = ARRecord.objects.get(project=self.proj)
        self.assertEqual(str(rec.target_collection_date), '2026-06-20')

    def test_import_bad_target_date_rejected(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['项目简称*', '运作日期*', '预估上账金额', '目标回款日期'])
        ws.append(['目标日期项目', '2026-05-01', 800, '不是日期'])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0); buf.name = 'r.xlsx'
        resp = self.client.post('/api/pk/ar/records/import', {'file': buf}, **self.auth())
        self.assertEqual(resp.status_code, 200)
        d = resp.json()['data']
        self.assertTrue(d.get('rejected'), d)
        self.assertTrue(any('目标回款日期' in e for e in d['errors']), d['errors'])

    def test_export_has_target_date_column(self):
        ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 1),
                                estimated_amount=Decimal('100'),
                                target_collection_date=date(2026, 6, 15))
        resp = self.client.get('/api/pk/ar/records/export', **self.auth())
        self.assertEqual(resp.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        headers = [c.value for c in wb.active[1]]
        self.assertIn('目标回款日期', headers)
        idx = headers.index('目标回款日期')
        self.assertEqual(wb.active.cell(2, idx + 1).value, '2026-06-15')


class InvoiceBatchWorkbenchTests(TestCase):
    """合并开票批次工作台：批次明细 / 批次开票落账 / 批次回款先进先出分摊。"""

    def setUp(self):
        self.client = Client()
        self.dept = '运输事业部'
        admin = PaikuanUser(phone='13900001000', name='BatchAdmin', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.token = make_token(admin)
        self.proj = ARProject.objects.create(
            customer_name='批次客户', short_name='批次项目', delivery_dept=self.dept,
            sales_contact='S', project_manager='M', project_no='BAT-0001')
        # 3条记录合并为一个批次：5月/6月/7月，6月那条带差额调整 +500
        self.r1 = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 1),
                                          estimated_amount=Decimal('1000'),
                                          invoice_batch_no='PF-001')
        self.r2 = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 6, 1),
                                          estimated_amount=Decimal('2000'),
                                          account_diff_adjustment=Decimal('500'),
                                          invoice_batch_no='PF-001')
        self.r3 = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 7, 1),
                                          estimated_amount=Decimal('3000'),
                                          invoice_batch_no='PF-001')

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def test_batch_detail(self):
        resp = self.client.get('/api/pk/ar/records/invoice-batches/PF-001', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(d['summary']['count'], 3)
        self.assertEqual(Decimal(d['summary']['billable']), Decimal('6500'))  # 6000 + 500差额
        self.assertEqual(d['summary']['pending_invoice'], 3)
        # 成员按运作日期排序
        self.assertEqual([m['operation_date'] for m in d['members']],
                         ['2026-05-01', '2026-06-01', '2026-07-01'])

    def test_batch_invoice_partial_amount_allowed_over_room_rejected(self):
        # 金额级分批：6000 ≤ 可开 6500 → 放行；再开 600 超剩余 500 → 拒绝
        resp = self.client.post('/api/pk/ar/records/invoice-batches/PF-001/invoice',
                                data=json.dumps({'invoice_date': '2026-07-10',
                                                 'amount': '6000'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        resp = self.client.post('/api/pk/ar/records/invoice-batches/PF-001/invoice',
                                data=json.dumps({'invoice_date': '2026-07-20',
                                                 'amount': '600'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 400)
        self.assertIn('可开余额', resp.json()['error'])

    def test_batch_invoice_fifo_tops_up_partially_invoiced(self):
        # r1 已人工开票 800（可开余 200）：开满全批剩余 5700 → r1 补到 1000，r2/r3 开满
        self.r1.actual_invoice_amount = Decimal('800')
        self.r1.invoice_date = date(2026, 6, 1)
        self.r1.save()
        resp = self.client.post('/api/pk/ar/records/invoice-batches/PF-001/invoice',
                                data=json.dumps({'invoice_date': '2026-07-10',
                                                 'amount': '5700'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.r1.refresh_from_db(); self.r2.refresh_from_db(); self.r3.refresh_from_db()
        self.assertEqual(self.r1.actual_invoice_amount, Decimal('1000'))
        self.assertEqual(self.r2.actual_invoice_amount, Decimal('2500'))     # 2000+500差额
        self.assertEqual(self.r3.actual_invoice_amount, Decimal('3000'))
        self.assertEqual(str(self.r3.invoice_date), '2026-07-10')

    def test_batch_payment_fifo_allocation(self):
        # 回款 3200：先结清 r1(1000)，再结清 r2(2500口径中的2200... 注意未收=上账+差额)
        # r1 未收1000，r2 未收2500，r3 未收3000；3200 → r1全收1000 + r2收2200（剩300）
        resp = self.client.post('/api/pk/ar/records/invoice-batches/PF-001/payment',
                                data=json.dumps({'amount': '3200', 'payment_date': '2026-07-20',
                                                 'notes': '建行到账'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        allocs = resp.json()['data']['allocations']
        self.assertEqual(len(allocs), 2)
        self.assertEqual(Decimal(allocs[0]['allocated']), Decimal('1000'))
        self.assertEqual(Decimal(allocs[0]['outstanding_after']), Decimal('0'))
        self.assertEqual(Decimal(allocs[1]['allocated']), Decimal('2200'))
        self.assertEqual(Decimal(allocs[1]['outstanding_after']), Decimal('300'))
        # 回款行带批次标记，可追溯
        self.assertTrue(self.r1.payments.filter(notes__contains='PF-001').exists())
        # 第二次回款 300+3000=3300 全部结清
        resp = self.client.post('/api/pk/ar/records/invoice-batches/PF-001/payment',
                                data=json.dumps({'amount': '3300', 'payment_date': '2026-08-05'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.r2.refresh_from_db(); self.r3.refresh_from_db()
        self.assertEqual(self.r2.outstanding_amount, Decimal('0'))
        self.assertEqual(self.r3.outstanding_amount, Decimal('0'))

    def test_batch_payment_over_outstanding_rejected(self):
        resp = self.client.post('/api/pk/ar/records/invoice-batches/PF-001/payment',
                                data=json.dumps({'amount': '99999', 'payment_date': '2026-07-20'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 400)
        self.assertIn('预收', resp.json()['error'])

    def test_batch_not_found(self):
        resp = self.client.get('/api/pk/ar/records/invoice-batches/不存在', **self.auth())
        self.assertEqual(resp.status_code, 404)


class AdjustmentAndCycleTests(TestCase):
    """差额调整多笔化（原因+金额，合计派生）+ 对账周期起始日（到期日锚点）。"""

    def setUp(self):
        self.client = Client()
        self.dept = '运输事业部'
        admin = PaikuanUser(phone='13900001100', name='AdjAdmin', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.token = make_token(admin)
        self.proj = ARProject.objects.create(
            customer_name='调整客户', short_name='调整项目', delivery_dept=self.dept,
            sales_contact='S', project_manager='M', project_no='ADJ-0001',
            reconciliation_days=30)

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    # ── 差额调整：多笔追加/删除，合计与未收派生 ─────────────────────────────
    def test_multiple_adjustments_derive_total(self):
        rec = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 1),
                                      estimated_amount=Decimal('1000'))
        r1 = self.client.post(f'/api/pk/ar/records/{rec.id}/adjustments',
                              data=json.dumps({'amount': '200', 'reason': '运费差'}),
                              content_type='application/json', **self.auth())
        self.assertEqual(r1.status_code, 200, r1.content)
        r2 = self.client.post(f'/api/pk/ar/records/{rec.id}/adjustments',
                              data=json.dumps({'amount': '-50', 'reason': '客户扣款'}),
                              content_type='application/json', **self.auth())
        self.assertEqual(r2.status_code, 200, r2.content)
        d = r2.json()['data']
        self.assertEqual(Decimal(d['total_adjustment']), Decimal('150'))
        self.assertEqual(Decimal(d['outstanding_amount']), Decimal('1150'))
        self.assertEqual(len(d['items']), 2)
        # 删除一笔 → 合计回退
        aid = d['items'][0]['id']
        r3 = self.client.delete(f'/api/pk/ar/records/{rec.id}/adjustments/{aid}', **self.auth())
        self.assertEqual(r3.status_code, 200, r3.content)
        self.assertEqual(Decimal(r3.json()['data']['total_adjustment']), Decimal('-50'))
        rec.refresh_from_db()
        self.assertEqual(rec.account_diff_adjustment, Decimal('-50'))
        self.assertEqual(rec.outstanding_amount, Decimal('950'))

    def test_adjustment_requires_reason_and_nonzero(self):
        rec = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 1),
                                      estimated_amount=Decimal('1000'))
        r = self.client.post(f'/api/pk/ar/records/{rec.id}/adjustments',
                             data=json.dumps({'amount': '0', 'reason': 'x'}),
                             content_type='application/json', **self.auth())
        self.assertEqual(r.status_code, 400)
        r = self.client.post(f'/api/pk/ar/records/{rec.id}/adjustments',
                             data=json.dumps({'amount': '10', 'reason': ''}),
                             content_type='application/json', **self.auth())
        self.assertEqual(r.status_code, 400)

    def test_adjustment_delete_blocked_when_outstanding_would_go_negative(self):
        rec = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 1),
                                      estimated_amount=Decimal('1000'))
        adj = ARAdjustment.objects.create(ar_record=rec, amount=Decimal('500'), reason='补付')
        ARPayment.objects.create(ar_record=rec, payment_no=1, amount=Decimal('1300'),
                                 payment_date=date(2026, 6, 1))
        # 删除 +500 调整 → 未收 = 1000 - 1300 = -300 → 拒绝
        r = self.client.delete(f'/api/pk/ar/records/{rec.id}/adjustments/{adj.id}', **self.auth())
        self.assertEqual(r.status_code, 400)
        self.assertTrue(ARAdjustment.objects.filter(pk=adj.pk).exists())

    def test_legacy_put_total_creates_delta_adjustment(self):
        rec = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 1),
                                      estimated_amount=Decimal('1000'))
        ARAdjustment.objects.create(ar_record=rec, amount=Decimal('100'), reason='初始')
        # 旧调用方按合计改为 250 → 自动生成 +150 的差值明细
        r = self.client.put(f'/api/pk/ar/records/{rec.id}',
                            data=json.dumps({'account_diff_adjustment': '250'}),
                            content_type='application/json', **self.auth())
        self.assertEqual(r.status_code, 200, r.content)
        rec.refresh_from_db()
        self.assertEqual(rec.account_diff_adjustment, Decimal('250'))
        self.assertEqual(rec.adjustments.count(), 2)

    def test_import_diff_with_reason_creates_adjustment(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['项目简称*', '运作日期*', '预估上账金额', '账实差额调整', '差额原因'])
        ws.append(['调整项目', '2026-05-01', 800, -30, '四舍五入'])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0); buf.name = 'r.xlsx'
        resp = self.client.post('/api/pk/ar/records/import', {'file': buf}, **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['data'].get('created'), 1, resp.json())
        rec = ARRecord.objects.get(project=self.proj)
        self.assertEqual(rec.account_diff_adjustment, Decimal('-30'))
        adj = rec.adjustments.get()
        self.assertEqual(adj.reason, '四舍五入')

    def test_export_has_adjustment_detail_column(self):
        rec = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 1),
                                      estimated_amount=Decimal('1000'))
        ARAdjustment.objects.create(ar_record=rec, amount=Decimal('80'), reason='运费差')
        resp = self.client.get('/api/pk/ar/records/export', **self.auth())
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        headers = [c.value for c in wb.active[1]]
        self.assertIn('差额明细', headers)
        idx = headers.index('差额明细')
        self.assertIn('运费差', str(wb.active.cell(2, idx + 1).value))

    # ── 对账周期起始日 → 到期日锚点 ─────────────────────────────────────────
    def test_cycle_start_day_due_date(self):
        # 周期15日~下月14日：运作 5/20 落在 5/15~6/14，账期结束 6/14，+30天账期
        self.proj.cycle_start_day = 15
        self.proj.total_days = 30
        self.proj.save()
        rec = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 20),
                                      estimated_amount=Decimal('100'))
        self.assertEqual(str(rec.due_date), '2026-07-14')   # 6/14 + 30
        # 运作 5/10 落在上一周期 4/15~5/14，账期结束 5/14
        rec2 = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 10),
                                       estimated_amount=Decimal('100'))
        self.assertEqual(str(rec2.due_date), '2026-06-13')  # 5/14 + 30

    def test_cycle_default_is_natural_month(self):
        self.proj.total_days = 30
        self.proj.save()
        rec = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 20),
                                      estimated_amount=Decimal('100'))
        self.assertEqual(str(rec.due_date), '2026-06-30')   # 5/31 + 30

    def test_cycle_change_recomputes_existing_due_dates(self):
        self.proj.total_days = 30
        self.proj.save()
        rec = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 20),
                                      estimated_amount=Decimal('100'))
        self.assertEqual(str(rec.due_date), '2026-06-30')
        # 项目改为月中结 → 信号自动重算存量到期日
        resp = self.client.put(f'/api/pk/ar/projects/{self.proj.id}',
                               data=json.dumps({'cycle_start_day': 15}),
                               content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        rec.refresh_from_db()
        self.assertEqual(str(rec.due_date), '2026-07-14')

    def test_cycle_start_day_validation(self):
        resp = self.client.put(f'/api/pk/ar/projects/{self.proj.id}',
                               data=json.dumps({'cycle_start_day': 31}),
                               content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 400)
        self.assertIn('1~28', resp.json()['error'])

    def test_project_import_with_cycle_column(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['客户名称*', '项目简称*', '交付部门*', '销售对接人*', '项目负责人*',
                   '对账周期起始日(1-28)'])
        ws.append(['周期客户', '周期项目', self.dept, 'A', 'B', 15])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0); buf.name = 'p.xlsx'
        resp = self.client.post('/api/pk/ar/projects/import', {'file': buf}, **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        p = ARProject.objects.get(short_name='周期项目')
        self.assertEqual(p.cycle_start_day, 15)


class OffsetChainTests(TestCase):
    """预收/预付核销口径前移：批量核销、核销工作台、预付冲抵上限、简称查询。"""

    def setUp(self):
        self.client = Client()
        self.dept = '运输事业部'
        admin = PaikuanUser(phone='13900001200', name='OffsetAdmin', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.token = make_token(admin)
        self.proj = ARProject.objects.create(
            customer_name='核销客户', short_name='核销项目', delivery_dept=self.dept,
            sales_contact='S', project_manager='M', project_no='OFF-0001')
        self.r1 = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 4, 1),
                                          estimated_amount=Decimal('1000'))
        self.r2 = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 1),
                                          estimated_amount=Decimal('2000'))
        self.adv = AdvanceRecord.objects.create(
            direction='预收', counterparty='核销客户', project=self.proj,
            delivery_dept=self.dept, occur_date=date(2026, 3, 1),
            occur_year=2026, occur_month=3,
            advance_amount=Decimal('2500'))

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def test_batch_writeoff_fifo(self):
        resp = self.client.post(f'/api/pk/ar/advances/{self.adv.id}/batch-writeoff',
                                data=json.dumps({'record_ids': [self.r1.id, self.r2.id],
                                                 'amount': '2200', 'writeoff_date': '2026-06-01'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        # 先进先出：r1(4月,1000)先结清，r2 分到 1200
        self.assertEqual(Decimal(d['allocations'][0]['allocated']), Decimal('1000'))
        self.assertEqual(Decimal(d['allocations'][1]['allocated']), Decimal('1200'))
        self.assertEqual(Decimal(d['advance_balance_after']), Decimal('300'))
        self.r1.refresh_from_db(); self.r2.refresh_from_db()
        self.assertEqual(self.r1.outstanding_amount, Decimal('0'))
        self.assertEqual(self.r2.outstanding_amount, Decimal('800'))
        # 每条都有「预收抵扣」回款 + 核销行可追溯
        self.assertEqual(self.r1.payments.filter(source='预收抵扣').count(), 1)
        self.assertEqual(self.adv.writeoffs.count(), 2)

    def test_batch_writeoff_default_amount(self):
        # 不传金额 → 默认 min(余额2500, 未收合计3000) = 2500
        resp = self.client.post(f'/api/pk/ar/advances/{self.adv.id}/batch-writeoff',
                                data=json.dumps({'record_ids': [self.r1.id, self.r2.id],
                                                 'writeoff_date': '2026-06-01'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.adv.refresh_from_db()
        self.assertEqual(self.adv.balance_amount, Decimal('0'))

    def test_batch_writeoff_over_balance_rejected(self):
        resp = self.client.post(f'/api/pk/ar/advances/{self.adv.id}/batch-writeoff',
                                data=json.dumps({'record_ids': [self.r1.id, self.r2.id],
                                                 'amount': '9999', 'writeoff_date': '2026-06-01'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 400)

    def test_batch_writeoff_cross_project_rejected(self):
        other = ARProject.objects.create(
            customer_name='别家客户', short_name='别家项目', delivery_dept=self.dept,
            sales_contact='S', project_manager='M', project_no='OFF-0002')
        r3 = ARRecord.objects.create(project=other, operation_date=date(2026, 5, 1),
                                     estimated_amount=Decimal('500'))
        resp = self.client.post(f'/api/pk/ar/advances/{self.adv.id}/batch-writeoff',
                                data=json.dumps({'record_ids': [r3.id],
                                                 'writeoff_date': '2026-06-01'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 400)
        self.assertIn('不一致', resp.json()['error'])

    def test_offset_workbench_groups_by_customer(self):
        resp = self.client.get('/api/pk/ar/advances/offset-workbench', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        groups = resp.json()['data']['groups']
        self.assertEqual(len(groups), 1)
        g = groups[0]
        self.assertEqual(g['customer'], '核销客户')
        self.assertEqual(len(g['advances']), 1)
        self.assertEqual(len(g['records']), 2)
        self.assertEqual(Decimal(g['offsettable']), Decimal('2500'))

    def test_prepaid_balance_by_short_name(self):
        AdvanceRecord.objects.create(
            direction='预付', counterparty='某供应商', project=self.proj,
            delivery_dept=self.dept, occur_date=date(2026, 3, 1),
            occur_year=2026, occur_month=3,
            advance_amount=Decimal('800'))
        resp = self.client.get('/api/pk/payments/prepaid-balance',
                               {'short_name': '核销项目'}, **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertTrue(d['matched'])
        self.assertEqual(Decimal(d['total_balance']), Decimal('800'))

    def test_prepaid_writeoff_over_payment_remaining_rejected(self):
        from paikuan.models import Payment
        prepaid = AdvanceRecord.objects.create(
            direction='预付', counterparty='某供应商', project=self.proj,
            delivery_dept=self.dept, occur_date=date(2026, 3, 1),
            occur_year=2026, occur_month=3,
            advance_amount=Decimal('5000'))
        pay = Payment.objects.create(
            department=self.dept, project_desc='设备款', payee='某供应商',
            total_amount=Decimal('1000'), planned_date=date(2026, 7, 1),
            project_short_name='核销项目')
        # 冲抵 1200 > 待付 1000 → 拒绝
        resp = self.client.post(f'/api/pk/ar/advances/{prepaid.id}/writeoffs',
                                data=json.dumps({'amount': '1200', 'writeoff_date': '2026-07-01',
                                                 'payment_id': pay.id}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 400)
        self.assertIn('剩余待付', resp.json()['error'])
        # 冲抵 600 通过，prepaid_offset_amount 信号更新
        resp = self.client.post(f'/api/pk/ar/advances/{prepaid.id}/writeoffs',
                                data=json.dumps({'amount': '600', 'writeoff_date': '2026-07-01',
                                                 'payment_id': pay.id}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        pay.refresh_from_db()
        self.assertEqual(pay.prepaid_offset_amount, Decimal('600'))
        # 第二次核销 500 > 剩余 400 → 拒绝；400 通过（多次核销）
        resp = self.client.post(f'/api/pk/ar/advances/{prepaid.id}/writeoffs',
                                data=json.dumps({'amount': '500', 'writeoff_date': '2026-07-02',
                                                 'payment_id': pay.id}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 400)
        resp = self.client.post(f'/api/pk/ar/advances/{prepaid.id}/writeoffs',
                                data=json.dumps({'amount': '400', 'writeoff_date': '2026-07-02',
                                                 'payment_id': pay.id}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        pay.refresh_from_db()
        self.assertEqual(pay.prepaid_offset_amount, Decimal('1000'))

    def test_ar_export_has_offset_columns(self):
        self.client.post(f'/api/pk/ar/advances/{self.adv.id}/batch-writeoff',
                         data=json.dumps({'record_ids': [self.r1.id],
                                          'writeoff_date': '2026-06-01'}),
                         content_type='application/json', **self.auth())
        resp = self.client.get('/api/pk/ar/records/export', **self.auth())
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        headers = [c.value for c in wb.active[1]]
        self.assertIn('预收冲抵金额', headers)
        self.assertIn('预收冲抵次数', headers)


class ReverseOffsetAndInstallmentTests(TestCase):
    """反向核销（两侧）+ 预付散单匹配 + 预收预付多次收付明细。"""

    def setUp(self):
        self.client = Client()
        self.dept = '运输事业部'
        admin = PaikuanUser(phone='13900001300', name='RevAdmin', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.token = make_token(admin)
        self.proj = ARProject.objects.create(
            customer_name='反核客户', short_name='反核项目', delivery_dept=self.dept,
            sales_contact='S', project_manager='M', project_no='REV-0001')

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _adv(self, direction, amount, **kw):
        return AdvanceRecord.objects.create(
            direction=direction, delivery_dept=self.dept,
            occur_year=2026, occur_month=3, occur_date=date(2026, 3, 1),
            advance_amount=Decimal(amount), **kw)

    # ── 反向核销：应收侧（删除预收抵扣回款 = 撤销核销）────────────────────────
    def test_delete_offset_payment_reverses_writeoff(self):
        adv = self._adv('预收', '1000', project=self.proj, counterparty='反核客户')
        rec = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 1),
                                      estimated_amount=Decimal('800'))
        resp = self.client.post(f'/api/pk/ar/advances/{adv.id}/batch-writeoff',
                                data=json.dumps({'record_ids': [rec.id],
                                                 'writeoff_date': '2026-06-01'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        rec.refresh_from_db(); adv.refresh_from_db()
        self.assertEqual(rec.outstanding_amount, Decimal('0'))
        self.assertEqual(adv.balance_amount, Decimal('200'))
        pay = rec.payments.get(source='预收抵扣')
        # 在应收侧删除该抵扣回款 → 反向核销
        resp = self.client.delete(f'/api/pk/ar/records/{rec.id}/payments/{pay.id}', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertIsNotNone(resp.json()['data'].get('reversed_writeoff'))
        rec.refresh_from_db(); adv.refresh_from_db()
        self.assertEqual(rec.outstanding_amount, Decimal('800'))   # 未收回升
        self.assertEqual(adv.balance_amount, Decimal('1000'))      # 预收余额恢复
        self.assertEqual(adv.writeoffs.count(), 0)

    # ── 反向核销：付款侧（删核销 → 冲抵额回退）────────────────────────────────
    def test_payment_offsets_list_and_reverse(self):
        from paikuan.models import Payment
        prepaid = self._adv('预付', '900', project=self.proj, counterparty='某供应商')
        pay = Payment.objects.create(
            department=self.dept, project_desc='设备款', payee='某供应商',
            total_amount=Decimal('1000'), planned_date=date(2026, 7, 1),
            project_short_name='反核项目')
        resp = self.client.post(f'/api/pk/ar/advances/{prepaid.id}/writeoffs',
                                data=json.dumps({'amount': '600', 'writeoff_date': '2026-07-01',
                                                 'payment_id': pay.id}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        # 列表接口
        resp = self.client.get(f'/api/pk/payments/{pay.id}/offsets', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(len(d['items']), 1)
        self.assertEqual(Decimal(d['total_offset']), Decimal('600'))
        wid = d['items'][0]['id']
        # 反向核销
        resp = self.client.delete(f'/api/pk/ar/advances/{prepaid.id}/writeoffs/{wid}', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        pay.refresh_from_db(); prepaid.refresh_from_db()
        self.assertEqual(pay.prepaid_offset_amount, Decimal('0'))
        self.assertEqual(prepaid.balance_amount, Decimal('900'))

    # ── 预付散单（未挂项目）按收款方匹配 ─────────────────────────────────────
    def test_prepaid_balance_matches_loose_by_payee(self):
        self._adv('预付', '500', counterparty='散单供应商')   # 不挂项目
        resp = self.client.get('/api/pk/payments/prepaid-balance',
                               {'payee': '散单供应商'}, **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertTrue(d['matched'])
        self.assertEqual(Decimal(d['total_balance']), Decimal('500'))

    # ── 多次收付明细：追加/删除，总额与余额派生 ──────────────────────────────
    def test_advance_multiple_installments(self):
        # API 创建 → 自动生成首笔收付明细
        resp = self.client.post('/api/pk/ar/advances', data=json.dumps({
            'direction': '预收', 'counterparty': '分期客户', 'delivery_dept': self.dept,
            'occur_year': 2026, 'occur_month': 4, 'occur_date': '2026-04-01',
            'advance_amount': '1000'}), content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        aid = resp.json()['data']['id']
        # 第二笔到账 +500
        resp = self.client.post(f'/api/pk/ar/advances/{aid}/installments',
                                data=json.dumps({'amount': '500', 'occur_date': '2026-05-01',
                                                 'notes': '第二笔预收'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(len(d['items']), 2)
        self.assertEqual(Decimal(d['advance_amount']), Decimal('1500'))
        self.assertEqual(Decimal(d['balance_amount']), Decimal('1500'))
        # 删除第二笔 → 总额回退
        iid = d['items'][1]['id']
        resp = self.client.delete(f'/api/pk/ar/advances/{aid}/installments/{iid}', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(Decimal(resp.json()['data']['advance_amount']), Decimal('1000'))

    def test_installment_delete_blocked_below_writeoffs(self):
        adv = self._adv('预收', '0', counterparty='分期客户2')
        i1 = AdvanceInstallment.objects.create(advance_record=adv, install_no=1,
                                               amount=Decimal('600'), occur_date=date(2026, 3, 1))
        AdvanceInstallment.objects.create(advance_record=adv, install_no=2,
                                          amount=Decimal('400'), occur_date=date(2026, 4, 1))
        AdvanceWriteoff.objects.create(advance_record=adv, writeoff_no=1,
                                       amount=Decimal('700'), writeoff_date=date(2026, 5, 1))
        # 删 600 的首笔 → 总额 400 < 已核销 700 → 拒绝
        resp = self.client.delete(f'/api/pk/ar/advances/{adv.id}/installments/{i1.id}', **self.auth())
        self.assertEqual(resp.status_code, 400)
        self.assertTrue(AdvanceInstallment.objects.filter(pk=i1.pk).exists())

    def test_legacy_put_amount_creates_delta_installment(self):
        adv = self._adv('预付', '0', counterparty='调额供应商')
        AdvanceInstallment.objects.create(advance_record=adv, install_no=1,
                                          amount=Decimal('300'), occur_date=date(2026, 3, 1))
        resp = self.client.put(f'/api/pk/ar/advances/{adv.id}',
                               data=json.dumps({'advance_amount': '800'}),
                               content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        adv.refresh_from_db()
        self.assertEqual(adv.advance_amount, Decimal('800'))
        self.assertEqual(adv.installments.count(), 2)

    def test_advance_import_creates_initial_installment(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['方向*', '往来单位*', '交付部门', '发生年*', '发生月*', '金额*'])
        ws.append(['预收', '导入分期客户', self.dept, 2026, 5, 1200])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0); buf.name = 'a.xlsx'
        resp = self.client.post('/api/pk/ar/advances/import', {'file': buf}, **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        if resp.json()['data'].get('created'):
            adv = AdvanceRecord.objects.get(counterparty='导入分期客户')
            self.assertEqual(adv.installments.count(), 1)
            self.assertEqual(adv.installments.first().amount, Decimal('1200'))


class BudgetProjectCompareTests(TestCase):
    """预算项目对照：收/付预算与实际收付经项目简称同窗对齐。"""

    def setUp(self):
        self.client = Client()
        self.dept = '运输事业部'
        admin = PaikuanUser(phone='13900001400', name='BgtAdmin', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.token = make_token(admin)
        self.proj = ARProject.objects.create(
            customer_name='预算客户', short_name='预算项目', delivery_dept=self.dept,
            sales_contact='S', project_manager='M', project_no='BGT-0001')

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def test_compare_aligns_budget_and_actuals(self):
        from paikuan.models import Payment, PaymentInstallment
        # 预算：收 10000 / 付 4000（6月）
        CollectionBudget.objects.create(short_name='预算项目', project_no='BGT-0001',
                                        delivery_dept=self.dept,
                                        expected_date=date(2026, 6, 15),
                                        amount=Decimal('10000'))
        PaymentBudget.objects.create(short_name='预算项目', project_no='BGT-0001',
                                     delivery_dept=self.dept,
                                     expected_date=date(2026, 6, 20),
                                     amount=Decimal('4000'))
        # 实际：回款 6000；排款实付 5000（超付款预算）
        rec = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 1),
                                      estimated_amount=Decimal('20000'))
        ARPayment.objects.create(ar_record=rec, payment_no=1, amount=Decimal('6000'),
                                 payment_date=date(2026, 6, 10))
        pay = Payment.objects.create(department=self.dept, project_desc='成本',
                                     payee='供应商', total_amount=Decimal('8000'),
                                     planned_date=date(2026, 6, 1),
                                     project_short_name='预算项目')
        PaymentInstallment.objects.create(payment=pay, seq=1, pay_date=date(2026, 6, 12),
                                          pay_amount=Decimal('5000'))
        resp = self.client.get('/api/pk/ar/budget/project-compare',
                               {'date_start': '2026-06-01', 'date_end': '2026-06-30'},
                               **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(len(d['rows']), 1)
        r = d['rows'][0]
        self.assertEqual(r['project'], '预算项目')
        self.assertEqual(Decimal(r['budget_in']), Decimal('10000'))
        self.assertEqual(Decimal(r['actual_in']), Decimal('6000'))
        self.assertEqual(r['in_rate'], 60.0)
        self.assertEqual(Decimal(r['budget_out']), Decimal('4000'))
        self.assertEqual(Decimal(r['actual_out']), Decimal('5000'))
        self.assertEqual(r['out_rate'], 125.0)
        self.assertIn('收款滞后', r['tags'])
        self.assertIn('付款超预算', r['tags'])
        self.assertEqual(Decimal(r['actual_net']), Decimal('1000'))
        s = d['summary']
        self.assertEqual(s['lagging'], 1)
        self.assertEqual(s['over_budget'], 1)

    def test_compare_unplanned_actuals(self):
        # 无预算但有实际回款 → 计划外收款
        rec = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 6, 1),
                                      estimated_amount=Decimal('5000'))
        ARPayment.objects.create(ar_record=rec, payment_no=1, amount=Decimal('2000'),
                                 payment_date=date(2026, 6, 5))
        resp = self.client.get('/api/pk/ar/budget/project-compare',
                               {'date_start': '2026-06-01', 'date_end': '2026-06-30'},
                               **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(len(d['rows']), 1)
        self.assertIn('计划外收款', d['rows'][0]['tags'])
        self.assertEqual(d['summary']['unplanned'], 1)

    def test_compare_window_excludes_outside(self):
        CollectionBudget.objects.create(short_name='预算项目', delivery_dept=self.dept,
                                        expected_date=date(2026, 7, 1),
                                        amount=Decimal('9999'))   # 窗外
        resp = self.client.get('/api/pk/ar/budget/project-compare',
                               {'date_start': '2026-06-01', 'date_end': '2026-06-30'},
                               **self.auth())
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.json()['data']['rows']), 0)


class AdvanceDiffSummaryTests(TestCase):
    """收付差异：预收 vs 预付按项目简称对齐 + 两侧逐笔明细。"""

    def setUp(self):
        self.client = Client()
        self.dept = '运输事业部'
        admin = PaikuanUser(phone='13900001500', name='DiffAdmin', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.token = make_token(admin)
        self.proj = ARProject.objects.create(
            customer_name='差异客户', short_name='差异项目', delivery_dept=self.dept,
            sales_contact='S', project_manager='M', project_no='DIF-0001')

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _adv(self, direction, amount, day, cp, notes=''):
        return AdvanceRecord.objects.create(
            direction=direction, project=self.proj, delivery_dept=self.dept,
            counterparty=cp, occur_year=2026, occur_month=day.month,
            occur_date=day, advance_amount=Decimal(amount), notes=notes)

    def test_diff_groups_and_details(self):
        self._adv('预收', '10000', date(2026, 3, 1), '客户A', notes='首期预收')
        self._adv('预收', '5000', date(2026, 4, 1), '客户A')
        self._adv('预付', '6000', date(2026, 3, 15), '供应商B', notes='设备预付')
        resp = self.client.get('/api/pk/ar/advances/diff-summary', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = resp.json()['data']
        self.assertEqual(len(d['rows']), 1)
        r = d['rows'][0]
        self.assertEqual(r['project'], '差异项目')
        self.assertEqual(Decimal(r['in_total']), Decimal('15000'))
        self.assertEqual(Decimal(r['out_total']), Decimal('6000'))
        self.assertEqual(Decimal(r['diff']), Decimal('9000'))
        self.assertEqual(len(r['in_items']), 2)
        self.assertEqual(len(r['out_items']), 1)
        self.assertEqual(r['in_items'][0]['counterparty'], '客户A')
        self.assertEqual(r['out_items'][0]['counterparty'], '供应商B')
        self.assertIn('首期预收', r['notes'])
        self.assertIn('设备预付', r['notes'])
        s = d['summary']
        self.assertEqual(Decimal(s['diff']), Decimal('9000'))

    def test_diff_excludes_loose_advances(self):
        # 散单（未挂项目）不参与差异对照
        AdvanceRecord.objects.create(
            direction='预收', delivery_dept=self.dept, counterparty='散单客户',
            occur_year=2026, occur_month=3, occur_date=date(2026, 3, 1),
            advance_amount=Decimal('999'))
        resp = self.client.get('/api/pk/ar/advances/diff-summary', **self.auth())
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.json()['data']['rows']), 0)

    def test_diff_q_filter(self):
        self._adv('预收', '100', date(2026, 3, 1), '客户A')
        resp = self.client.get('/api/pk/ar/advances/diff-summary', {'q': '不存在'}, **self.auth())
        self.assertEqual(len(resp.json()['data']['rows']), 0)
        resp = self.client.get('/api/pk/ar/advances/diff-summary', {'q': '差异'}, **self.auth())
        self.assertEqual(len(resp.json()['data']['rows']), 1)

    def test_diff_details_use_installment_dates(self):
        # 多次到账：明细按收付明细的实际发生日期逐笔列出（非首笔/签约日）
        adv = self._adv('预收', '0', date(2026, 3, 1), '客户A')
        AdvanceInstallment.objects.create(advance_record=adv, install_no=1,
                                          amount=Decimal('3000'), occur_date=date(2026, 3, 5))
        AdvanceInstallment.objects.create(advance_record=adv, install_no=2,
                                          amount=Decimal('2000'), occur_date=date(2026, 5, 18))
        resp = self.client.get('/api/pk/ar/advances/diff-summary', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        r = resp.json()['data']['rows'][0]
        self.assertEqual(Decimal(r['in_total']), Decimal('5000'))
        self.assertEqual([i['occur_date'] for i in r['in_items']],
                         ['2026-03-05', '2026-05-18'])
        # 未核销余额标在该记录最后一笔明细上
        self.assertEqual(Decimal(r['in_items'][-1]['balance']), Decimal('5000'))
        self.assertEqual(Decimal(r['in_items'][0]['balance']), Decimal('0'))


class ActionPermissionTests(TestCase):
    """操作级权限：出纳无 can_create 也能单独获得预付核销等动作。"""

    def setUp(self):
        self.client = Client()
        self.dept = '运输事业部'
        self.cashier = PaikuanUser(phone='13900001600', name='出纳小王', role='user',
                                   job_title='cashier', departments=[self.dept],
                                   is_active=True, is_approved=True)
        self.cashier.set_password('Test123456')
        self.cashier.save()
        self.token = make_token(self.cashier)
        self.proj = ARProject.objects.create(
            customer_name='权限客户', short_name='权限项目', delivery_dept=self.dept,
            sales_contact='S', project_manager='M', project_no='PRM-0001')
        _invalidate_perm_cache()

    def tearDown(self):
        _invalidate_perm_cache()

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _prepaid(self, amount='1000'):
        return AdvanceRecord.objects.create(
            direction='预付', project=self.proj, delivery_dept=self.dept,
            counterparty='供应商X', occur_year=2026, occur_month=3,
            occur_date=date(2026, 3, 1), advance_amount=Decimal(amount))

    # 出纳默认：无 can_create，但 actions.wo_prepaid=True → 预付核销放行
    def test_cashier_can_writeoff_prepaid_without_can_create(self):
        from paikuan.views import get_job_perms
        perms = get_job_perms('cashier')
        self.assertFalse(perms['can_create'])
        self.assertTrue(perms['actions']['wo_prepaid'])
        prepaid = self._prepaid()
        resp = self.client.post(f'/api/pk/ar/advances/{prepaid.id}/writeoffs',
                                data=json.dumps({'amount': '300', 'writeoff_date': '2026-04-01'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        prepaid.refresh_from_db()
        self.assertEqual(prepaid.balance_amount, Decimal('700'))
        # 撤销自己的核销也放行（同动作键，不要求 can_delete）
        wid = resp.json()['data']['id']
        resp = self.client.delete(f'/api/pk/ar/advances/{prepaid.id}/writeoffs/{wid}', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)

    # 出纳默认 wo_receive=False → 预收核销被拒，提示去权限配置开通
    def test_cashier_receive_writeoff_denied_by_default(self):
        adv = AdvanceRecord.objects.create(
            direction='预收', project=self.proj, delivery_dept=self.dept,
            counterparty='权限客户', occur_year=2026, occur_month=3,
            occur_date=date(2026, 3, 1), advance_amount=Decimal('1000'))
        rec = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 1),
                                      estimated_amount=Decimal('500'))
        resp = self.client.post(f'/api/pk/ar/advances/{adv.id}/batch-writeoff',
                                data=json.dumps({'record_ids': [rec.id],
                                                 'writeoff_date': '2026-06-01'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 403)
        self.assertIn('操作权限', resp.json()['error'])

    # 超管在权限配置开通 wo_receive 后 → 放行（独立配置生效）
    def test_action_grant_via_permission_config(self):
        JobPermission.objects.update_or_create(
            job_title='cashier',
            defaults={'config': {'actions': {'wo_receive': True}}})
        _invalidate_perm_cache('cashier')
        adv = AdvanceRecord.objects.create(
            direction='预收', project=self.proj, delivery_dept=self.dept,
            counterparty='权限客户', occur_year=2026, occur_month=3,
            occur_date=date(2026, 3, 1), advance_amount=Decimal('1000'))
        rec = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 1),
                                      estimated_amount=Decimal('500'))
        resp = self.client.post(f'/api/pk/ar/advances/{adv.id}/batch-writeoff',
                                data=json.dumps({'record_ids': [rec.id],
                                                 'writeoff_date': '2026-06-01'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)

    # 出纳默认 ar_collect=True → 录回款放行（无 can_create 依然可录）
    def test_cashier_can_record_collection(self):
        rec = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 1),
                                      estimated_amount=Decimal('800'))
        resp = self.client.post(f'/api/pk/ar/records/{rec.id}/payments',
                                data=json.dumps({'amount': '200', 'payment_date': '2026-06-01'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)

    # 显式 False 覆盖：即使有 can_create 的职务也可被禁用某动作
    def test_explicit_false_overrides_can_create(self):
        JobPermission.objects.update_or_create(
            job_title='cashier',
            defaults={'config': {'can_create': True, 'actions': {'wo_prepaid': False}}})
        _invalidate_perm_cache('cashier')
        prepaid = self._prepaid()
        resp = self.client.post(f'/api/pk/ar/advances/{prepaid.id}/writeoffs',
                                data=json.dumps({'amount': '100', 'writeoff_date': '2026-04-01'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 403)


class AutoBatchNoTests(TestCase):
    """批次号自动生成：客户简称-日期-序号，免人为编码。"""

    def setUp(self):
        self.client = Client()
        self.dept = '运输事业部'
        admin = PaikuanUser(phone='13900001700', name='AutoBatchAdmin', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.token = make_token(admin)
        self.proj = ARProject.objects.create(
            customer_name='华为物流', short_name='华为项目', delivery_dept=self.dept,
            sales_contact='S', project_manager='M', project_no='AB-0001')

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _rec(self, month):
        return ARRecord.objects.create(project=self.proj, operation_date=date(2026, month, 1),
                                       estimated_amount=Decimal('1000'))

    def test_auto_generates_readable_batch_no(self):
        r1, r2 = self._rec(4), self._rec(5)
        resp = self.client.post('/api/pk/ar/records/batch-assign',
                                data=json.dumps({'ids': [r1.id, r2.id], 'auto': True}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        bn = resp.json()['data']['invoice_batch_no']
        today = date.today().strftime('%y%m%d')
        self.assertEqual(bn, f'华为物流-{today}-01')
        r1.refresh_from_db()
        self.assertEqual(r1.invoice_batch_no, bn)
        # 同日再合并一批 → 序号递增
        r3 = self._rec(6)
        resp = self.client.post('/api/pk/ar/records/batch-assign',
                                data=json.dumps({'ids': [r3.id], 'auto': True}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.json()['data']['invoice_batch_no'], f'华为物流-{today}-02')

    def test_auto_multi_customer_prefix(self):
        other = ARProject.objects.create(
            customer_name='别家客户', short_name='别家项目', delivery_dept=self.dept,
            sales_contact='S', project_manager='M', project_no='AB-0002')
        r1 = self._rec(4)
        r2 = ARRecord.objects.create(project=other, operation_date=date(2026, 4, 1),
                                     estimated_amount=Decimal('500'))
        resp = self.client.post('/api/pk/ar/records/batch-assign',
                                data=json.dumps({'ids': [r1.id, r2.id], 'auto': True}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertTrue(resp.json()['data']['invoice_batch_no'].startswith('多户-'))

    def test_manual_and_clear_still_work(self):
        r1 = self._rec(4)
        resp = self.client.post('/api/pk/ar/records/batch-assign',
                                data=json.dumps({'ids': [r1.id], 'invoice_batch_no': '自定义-01'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.json()['data']['invoice_batch_no'], '自定义-01')
        resp = self.client.post('/api/pk/ar/records/batch-assign',
                                data=json.dumps({'ids': [r1.id], 'invoice_batch_no': ''}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200)
        r1.refresh_from_db()
        self.assertEqual(r1.invoice_batch_no, '')

    def test_batch_list_includes_customers(self):
        r1 = self._rec(4)
        self.client.post('/api/pk/ar/records/batch-assign',
                         data=json.dumps({'ids': [r1.id], 'auto': True}),
                         content_type='application/json', **self.auth())
        resp = self.client.get('/api/pk/ar/records/invoice-batches', **self.auth())
        self.assertEqual(resp.status_code, 200)
        b = resp.json()['data']['batches'][0]
        self.assertIn('华为物流', b['customers'])


class PartialBatchInvoiceTests(TestCase):
    """金额级分批开票：多次开票事件先进先出分摊 + 税额 + 整次撤销。"""

    def setUp(self):
        self.client = Client()
        self.dept = '运输事业部'
        admin = PaikuanUser(phone='13900001900', name='PInvAdmin', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.token = make_token(admin)
        self.proj = ARProject.objects.create(
            customer_name='分批客户', short_name='分批项目', delivery_dept=self.dept,
            sales_contact='S', project_manager='M', project_no='PIV-0001',
            invoice_mode='全额', tax_rate=Decimal('0.06'))
        self.r1 = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 4, 1),
                                          estimated_amount=Decimal('1060'),
                                          invoice_batch_no='PB-01')
        self.r2 = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 1),
                                          estimated_amount=Decimal('2120'),
                                          invoice_batch_no='PB-01')

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _inv(self, amount, day, **extra):
        return self.client.post('/api/pk/ar/records/invoice-batches/PB-01/invoice',
                                data=json.dumps({'invoice_date': day, 'amount': amount, **extra}),
                                content_type='application/json', **self.auth())

    def test_multiple_invoice_events_fifo(self):
        # 第一次开 1500：r1 开满 1060，r2 分到 440
        resp = self._inv('1500', '2026-06-01')
        self.assertEqual(resp.status_code, 200, resp.content)
        allocs = resp.json()['data']['allocations']
        self.assertEqual(Decimal(allocs[0]['amount']), Decimal('1060'))
        self.assertEqual(Decimal(allocs[1]['amount']), Decimal('440'))
        self.r1.refresh_from_db(); self.r2.refresh_from_db()
        self.assertEqual(self.r1.actual_invoice_amount, Decimal('1060'))
        self.assertEqual(self.r1.tax_amount, Decimal('60.00'))   # 全额自动算
        self.assertEqual(self.r2.actual_invoice_amount, Decimal('440'))
        # 第二次开 1680：r2 开满（440+1680=2120）
        resp = self._inv('1680', '2026-07-01')
        self.assertEqual(resp.status_code, 200, resp.content)
        self.r2.refresh_from_db()
        self.assertEqual(self.r2.actual_invoice_amount, Decimal('2120'))
        self.assertIn('已开满', resp.json()['data']['message'])
        # 事件留痕
        d = self.client.get('/api/pk/ar/records/invoice-batches/PB-01', **self.auth()).json()['data']
        self.assertEqual(len(d['invoice_events']), 2)
        self.assertEqual(Decimal(d['summary']['invoice_room']), Decimal('0'))

    def test_over_room_rejected(self):
        resp = self._inv('5000', '2026-06-01')
        self.assertEqual(resp.status_code, 400)
        self.assertIn('可开余额', resp.json()['error'])

    def test_undo_invoice_event(self):
        ev = self._inv('1500', '2026-06-01').json()['data']['event']
        resp = self.client.post('/api/pk/ar/records/invoice-batches/PB-01/invoice-undo',
                                data=json.dumps({'event_id': ev['id']}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.r1.refresh_from_db(); self.r2.refresh_from_db()
        self.assertIsNone(self.r1.actual_invoice_amount)
        self.assertIsNone(self.r1.invoice_date)
        self.assertIsNone(self.r1.tax_amount)
        self.assertIsNone(self.r2.actual_invoice_amount)

    def test_manual_tax_distributed_to_diff_mode(self):
        # 差额模式项目：手填税额按分摊占比分配
        dproj = ARProject.objects.create(
            customer_name='差额客户', short_name='差额项目', delivery_dept=self.dept,
            sales_contact='S', project_manager='M', project_no='PIV-0002',
            invoice_mode='差额')
        r3 = ARRecord.objects.create(project=dproj, operation_date=date(2026, 3, 1),
                                     estimated_amount=Decimal('1000'),
                                     invoice_batch_no='PB-02')
        resp = self.client.post('/api/pk/ar/records/invoice-batches/PB-02/invoice',
                                data=json.dumps({'invoice_date': '2026-06-01',
                                                 'amount': '600', 'tax_amount': '36'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        r3.refresh_from_db()
        self.assertEqual(r3.actual_invoice_amount, Decimal('600'))
        self.assertEqual(r3.tax_amount, Decimal('36.00'))
        # 可继续开剩余 400
        resp = self.client.post('/api/pk/ar/records/invoice-batches/PB-02/invoice',
                                data=json.dumps({'invoice_date': '2026-07-01', 'amount': '400'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        r3.refresh_from_db()
        self.assertEqual(r3.actual_invoice_amount, Decimal('1000'))


class BatchPaymentUndoTests(TestCase):
    """批次管理：批次回款事件还原 + 整次撤销。"""

    def setUp(self):
        self.client = Client()
        self.dept = '运输事业部'
        admin = PaikuanUser(phone='13900001800', name='UndoAdmin', role='super_admin',
                            job_title='finance_director', departments=[self.dept],
                            is_active=True, is_approved=True)
        admin.set_password('Test123456')
        admin.save()
        self.token = make_token(admin)
        self.proj = ARProject.objects.create(
            customer_name='撤销客户', short_name='撤销项目', delivery_dept=self.dept,
            sales_contact='S', project_manager='M', project_no='UND-0001')
        self.r1 = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 4, 1),
                                          estimated_amount=Decimal('1000'),
                                          invoice_batch_no='B-01')
        self.r2 = ARRecord.objects.create(project=self.proj, operation_date=date(2026, 5, 1),
                                          estimated_amount=Decimal('2000'),
                                          invoice_batch_no='B-01')

    def auth(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def test_collections_event_and_undo(self):
        # 批次回款 1500 → 事件出现在批次明细，整次撤销后未收恢复
        resp = self.client.post('/api/pk/ar/records/invoice-batches/B-01/payment',
                                data=json.dumps({'amount': '1500', 'payment_date': '2026-06-01'}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d = self.client.get('/api/pk/ar/records/invoice-batches/B-01', **self.auth()).json()['data']
        self.assertEqual(len(d['collections']), 1)
        ev = d['collections'][0]
        self.assertEqual(Decimal(ev['total']), Decimal('1500'))
        self.assertEqual(ev['count'], 2)   # r1 结清1000 + r2 分到500
        resp = self.client.post('/api/pk/ar/records/invoice-batches/B-01/payment-undo',
                                data=json.dumps({'payment_ids': ev['payment_ids']}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        self.r1.refresh_from_db(); self.r2.refresh_from_db()
        self.assertEqual(self.r1.outstanding_amount, Decimal('1000'))
        self.assertEqual(self.r2.outstanding_amount, Decimal('2000'))

    def test_undo_rejects_non_batch_payment(self):
        # 单笔手工回款不能经批次撤销入口删
        pay = ARPayment.objects.create(ar_record=self.r1, payment_no=1,
                                       amount=Decimal('100'), payment_date=date(2026, 6, 1))
        resp = self.client.post('/api/pk/ar/records/invoice-batches/B-01/payment-undo',
                                data=json.dumps({'payment_ids': [pay.id]}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 400)
        self.assertTrue(ARPayment.objects.filter(pk=pay.pk).exists())

    def test_same_day_same_note_payments_stay_separate(self):
        # 同日、相同（空）备注的两次批次回款，应还原为两个独立事件，
        # 撤销其一只回退该次（唯一事件标记防止合并）
        for amt in ('300', '500'):
            r = self.client.post('/api/pk/ar/records/invoice-batches/B-01/payment',
                                 data=json.dumps({'amount': amt, 'payment_date': '2026-06-01'}),
                                 content_type='application/json', **self.auth())
            self.assertEqual(r.status_code, 200, r.content)
        d = self.client.get('/api/pk/ar/records/invoice-batches/B-01', **self.auth()).json()['data']
        self.assertEqual(len(d['collections']), 2)   # 两次未被合并
        totals = sorted(Decimal(ev['total']) for ev in d['collections'])
        self.assertEqual(totals, [Decimal('300'), Decimal('500')])
        ev300 = next(ev for ev in d['collections'] if Decimal(ev['total']) == Decimal('300'))
        resp = self.client.post('/api/pk/ar/records/invoice-batches/B-01/payment-undo',
                                data=json.dumps({'payment_ids': ev300['payment_ids']}),
                                content_type='application/json', **self.auth())
        self.assertEqual(resp.status_code, 200, resp.content)
        d2 = self.client.get('/api/pk/ar/records/invoice-batches/B-01', **self.auth()).json()['data']
        self.assertEqual(len(d2['collections']), 1)
        self.assertEqual(Decimal(d2['collections'][0]['total']), Decimal('500'))


