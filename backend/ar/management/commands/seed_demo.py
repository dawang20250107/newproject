"""一键重置 / 载入演示数据（6 个事业部的真实应收 + 财务报表）。

用法：
    python manage.py seed_demo                # 清空并重新载入应收 + 财务
    python manage.py seed_demo --ar-only      # 仅载入应收明细
    python manage.py seed_demo --fin-only     # 仅载入财务报表
    python manage.py seed_demo --no-clear     # 追加（不先清空）

数据源：backend/seed_data/ar_records.xlsx、financials_2026.xlsx
  · 应收：逐行建项目/应收记录/回款（支持「a / b」多笔回款单元格），并重算未收/税额。
  · 财务：按事业部 × 月 建已发布的部门明细批次 + 一级科目金额；毛利/净利由口径公式
          自动推导（与报表自带计算行一致）。
"""
import datetime
from decimal import Decimal, InvalidOperation

from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone

FIN_YEAR = 2026
MONTH_COLS = {1: 3, 2: 4, 3: 5, 4: 6, 5: 7}  # 财务报表 1月..5月 列序
BUS = ['集团总部', '劳务事业部', '运输事业部', '自营事业部',
       '阔展事业部', '多式联运事业部', '供应链事业部']


def _dec(v):
    if v is None or v == '':
        return None
    s = str(v).replace(',', '').replace('，', '').replace('¥', '').strip()
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _date(v):
    if v in (None, ''):
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = str(v).strip().split(' ')[0].replace('/', '-')
    try:
        return datetime.date.fromisoformat(s)
    except Exception:
        return None


class Command(BaseCommand):
    help = '清空并载入 6 个事业部的演示数据（应收明细 + 财务报表）'

    def add_arguments(self, parser):
        parser.add_argument('--ar-only', action='store_true', help='仅载入应收明细')
        parser.add_argument('--fin-only', action='store_true', help='仅载入财务报表')
        parser.add_argument('--no-clear', action='store_true', help='不先清空（追加）')

    def handle(self, *args, **opts):
        try:
            import openpyxl  # noqa
        except ImportError:
            self.stderr.write('需要 openpyxl：pip install openpyxl')
            return
        data_dir = settings.BASE_DIR / 'seed_data'
        do_ar = not opts['fin_only']
        do_fin = not opts['ar_only']
        if do_ar:
            self._seed_ar(data_dir / 'ar_records.xlsx', clear=not opts['no_clear'])
            self._seed_customers(clear=not opts['no_clear'])
        if do_fin:
            self._seed_fin(data_dir / 'financials_2026.xlsx', clear=not opts['no_clear'])
            self._seed_targets(clear=not opts['no_clear'])
        # 项目毛利依赖应收(取项目名/开票额)+财务(取BU毛利率)，两侧都灌完再点亮业财融合
        if do_ar and do_fin:
            self._seed_project_margin(clear=not opts['no_clear'])
        self.stdout.write(self.style.SUCCESS('演示数据载入完成。'))

    # ── 应收明细 ────────────────────────────────────────────────────────────────
    def _seed_ar(self, path, clear):
        import openpyxl
        from ar.models import (ARProject, ARRecord, ARPayment, ARAdjustment,
                               AdvanceRecord, AdvanceWriteoff)
        if not path.exists():
            self.stderr.write(f'缺少数据文件：{path}')
            return
        if clear:
            with transaction.atomic():
                AdvanceWriteoff.objects.all().delete()
                AdvanceRecord.objects.all().delete()
                ARPayment.objects.all().delete()
                ARAdjustment.objects.all().delete()
                ARRecord.objects.all().delete()
                ARProject.objects.all().delete()
            self.stdout.write('已清空应收数据表。')

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb['应收账款明细'] if '应收账款明细' in wb.sheetnames else wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        header = [str(c).strip() if c is not None else '' for c in rows[0]]
        idx = {h: i for i, h in enumerate(header)}

        def cell(r, name):
            i = idx.get(name)
            return r[i] if (i is not None and i < len(r)) else None

        cache, cp, cr, cpay, skip = {}, 0, 0, 0, 0
        for r in rows[1:]:
            pno = (str(cell(r, '项目编号')).strip() if cell(r, '项目编号') else '')
            short = (str(cell(r, '项目简称')).strip() if cell(r, '项目简称') else '')
            dept = (str(cell(r, '交付部门')).strip() if cell(r, '交付部门') else '')
            if not (pno or short) or not dept:
                continue
            contract = (str(cell(r, '合同名称')).strip() if cell(r, '合同名称') else '') or short or pno
            mgr = (str(cell(r, '项目负责人')).strip() if cell(r, '项目负责人') else '—') or '—'
            sales = (str(cell(r, '销售对接人')).strip() if cell(r, '销售对接人') else '') or mgr
            try:
                oy, om = int(float(cell(r, '运作年'))), int(float(cell(r, '运作月')))
            except Exception:
                continue
            if not (1 <= om <= 12):
                continue
            try:
                total_days = int(float(cell(r, '总账期(天)') or 0))
            except Exception:
                total_days = 0
            inv_mode = (str(cell(r, '开票模式')).strip() if cell(r, '开票模式') else '') or '全额'

            pkey = pno or f'{short}|{dept}'
            proj = cache.get(pkey)
            if proj is None:
                proj = (ARProject.objects.filter(project_no=pno).first() if pno else None)
                if proj is None:
                    proj = ARProject.objects.create(
                        project_no=(pno or f'AUTO-{len(cache) + 1:04d}'),
                        customer_name=contract, short_name=short or contract,
                        delivery_dept=dept, sales_contact=sales, project_manager=mgr,
                        reconciliation_days=total_days, invoice_wait_days=0,
                        post_invoice_days=0, invoice_mode=inv_mode)
                    cp += 1
                cache[pkey] = proj

            est = _dec(cell(r, '预估上账金额')) or Decimal('0')
            inv_raw = cell(r, '实际开票金额')
            inv_amt = _dec(inv_raw) if inv_raw not in (None, '') else None
            tax_raw = cell(r, '税额')
            tax_amt = (_dec(tax_raw) or Decimal('0')) if tax_raw not in (None, '') else Decimal('0')
            adj = _dec(cell(r, '账实差额调整')) or Decimal('0')
            pay_dates = [s.strip() for s in str(cell(r, '回款日期') or '').split('/') if s.strip()]
            pay_amts = [s.strip() for s in str(cell(r, '回款金额') or '').split('/') if s.strip()]
            due = _date(cell(r, '应收日期'))
            pays = []
            for i, a in enumerate(pay_amts):
                amt = _dec(a)
                if not amt or amt <= 0:
                    continue
                d = _date(pay_dates[i]) if i < len(pay_dates) else due
                pays.append((d or due or datetime.date(oy, om, 28), amt))
            try:
                with transaction.atomic():
                    rec = ARRecord.objects.create(
                        project=proj, operation_year=oy, operation_month=om,
                        estimated_amount=est, actual_invoice_amount=inv_amt, tax_amount=tax_amt,
                        invoice_date=_date(cell(r, '开票日期')), account_diff_adjustment=adj,
                        reconciliation_date=_date(cell(r, '对账日期')), due_date=due,
                        notes=(str(cell(r, '备注')).strip() if cell(r, '备注') else ''),
                        delivery_dept=dept)
                    # 差额走明细链路（带 adjust_date），与导入口径一致，时段合计方能归集
                    if adj:
                        ARAdjustment.objects.create(
                            ar_record=rec, amount=adj, reason='导入差额调整',
                            adjust_date=rec.operation_date or due)
                    for n, (pd, pa) in enumerate(pays, start=1):
                        ARPayment.objects.create(ar_record=rec, payment_no=n, amount=pa,
                                                 payment_date=pd, source='回款')
                        cpay += 1
                    rec.recompute_derived(save=True)
                    cr += 1
            except Exception:
                skip += 1
        self.stdout.write(self.style.SUCCESS(
            f'应收：项目 {cp} · 记录 {cr} · 回款 {cpay} · 跳过 {skip}'))

    # ── 客户正名（合同名称 → 客户）：点亮客商画像 ────────────────────────────────
    def _seed_customers(self, clear):
        """把"合同名称"字段(实为客户公司全称)正名为客户实体并回填 ARProject.customer。
        每个不同合同名 = 一个客户；客户等级取项目上的 customer_level。源数据无独立客户
        列，公司名本就装在合同名字段里，故此为"正名"而非"造数据"。"""
        from ar.models import ARProject, Customer
        if clear:
            ARProject.objects.exclude(customer__isnull=True).update(customer=None)
            Customer.objects.all().delete()
        cache, made, linked = {}, 0, 0
        for p in ARProject.objects.all().only('id', 'customer_name', 'customer_level'):
            name = (p.customer_name or '').strip()
            if not name:
                continue
            # 客户按 (名称 + 交付部门) 隔离
            key = (name, p.delivery_dept or '')
            cust = cache.get(key)
            if cust is None:
                cust, created = Customer.objects.get_or_create(
                    name=name, delivery_dept=p.delivery_dept or '',
                    defaults={'level': p.customer_level or ''})
                cache[key] = cust
                made += int(created)
            # 用 update 绕开 ARProject.save 的派生逻辑，仅落 customer 外键
            ARProject.objects.filter(pk=p.pk).update(customer=cust)
            linked += 1
        self.stdout.write(self.style.SUCCESS(f'客户正名：客户 {made} 个 · 回填项目 {linked} 个'))

    # ── 财务报表 ────────────────────────────────────────────────────────────────
    def _seed_fin(self, path, clear):
        import openpyxl
        from caiwu.models import ImportBatch, FinancialEntry, L1Category
        if not path.exists():
            self.stderr.write(f'缺少数据文件：{path}')
            return
        l1_by_name = {c.name: c for c in L1Category.objects.all()}
        raw_names = set(c.name for c in L1Category.objects.filter(is_calculated=False))
        if clear:
            ImportBatch.objects.filter(year=FIN_YEAR, batch_type=ImportBatch.TYPE_DEPT).delete()
            self.stdout.write(f'已清空 {FIN_YEAR} 年部门明细批次。')

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        nb = ne = 0
        for bu in BUS:
            if bu not in wb.sheetnames:
                continue
            rows = list(wb[bu].iter_rows(values_only=True))
            per_month = {m: {} for m in MONTH_COLS}
            for r in rows[3:]:
                name = r[0]
                if name in (None, '') or name not in raw_names:
                    continue
                l1 = l1_by_name[name]
                for m, ci in MONTH_COLS.items():
                    amt = _dec(r[ci]) if ci < len(r) else None
                    if amt is not None and amt != 0:
                        per_month[m][l1.id] = per_month[m].get(l1.id, Decimal('0')) + amt
            with transaction.atomic():
                for m in MONTH_COLS:
                    if not per_month[m]:
                        continue
                    batch = ImportBatch.objects.create(
                        business_unit=bu, year=FIN_YEAR, month=m,
                        batch_type=ImportBatch.TYPE_DEPT, status=ImportBatch.STATUS_PUBLISHED,
                        uploaded_by=None, published_at=timezone.now(),
                        row_count=len(per_month[m]), file_name='seed_demo.xlsx')
                    nb += 1
                    for l1_id, amt in per_month[m].items():
                        FinancialEntry.objects.create(batch=batch, l1_id=l1_id, l2=None, l3=None, amount=amt)
                        ne += 1
        wb.close()
        self.stdout.write(self.style.SUCCESS(f'财务：批次 {nb} · 科目分录 {ne}'))

    # ── 经营目标（按已发布实际的运行率推导，使达成率口径合理）──────────────────────
    def _seed_targets(self, clear):
        """收入目标取 1-5 月实际的年化（≈与时间进度对齐）；毛利/净利目标取理想正向
        (收入×6% / 收入×3%)，从而真实呈现"收入达标、盈利远未达标"的经营故事。"""
        from caiwu.models import FinancialTarget
        from caiwu import views as V
        if clear:
            FinancialTarget.objects.filter(year=FIN_YEAR).delete()
        actuals = V._collect_actuals(BUS, {FIN_YEAR})
        elapsed = len(MONTH_COLS)  # 已发布月数
        made = 0
        for bu in BUS:
            ytd_rev = sum((V._rp(actuals, bu, FIN_YEAR, m)[0] or 0) for m in MONTH_COLS)
            if ytd_rev <= 0:
                continue
            ann_rev = round(ytd_rev / elapsed * 12)
            ann_gross = round(ann_rev * 0.06)
            ann_profit = round(ann_rev * 0.03)
            FinancialTarget.objects.update_or_create(
                business_unit=bu, year=FIN_YEAR, month=FinancialTarget.MONTH_ANNUAL,
                defaults=dict(target_revenue=ann_rev, target_gross_profit=ann_gross,
                              target_profit=ann_profit))
            made += 1
            for m in MONTH_COLS:
                FinancialTarget.objects.update_or_create(
                    business_unit=bu, year=FIN_YEAR, month=m,
                    defaults=dict(target_revenue=round(ann_rev / 12),
                                  target_gross_profit=round(ann_gross / 12),
                                  target_profit=round(ann_profit / 12)))
        self.stdout.write(self.style.SUCCESS(f'目标：{made} 个事业部（年度 + 1-5 月）'))

    # ── 项目毛利（业财融合原料）：应收侧项目×月开票额作收入，BU 当月毛利率反推成本 ──
    def _seed_project_margin(self, clear):
        """点亮「业财损益全景」：财务侧原本只有 BU 级数据，项目级毛利表(ProjectMargin)
        来自金蝶按项目核算、需单独上传。演示库里据应收项目逐月开票额(无则预估额)作收入，
        按所属事业部当月真实毛利率(来自已发布财务报表)反推成本，项目名取应收项目简称，
        使项目级业财融合在演示数据上即可端到端点亮。仅演示数据、可逆。"""
        from ar.models import ARProject, ARRecord
        from caiwu.models import ProjectMargin
        from caiwu import views as V
        if clear:
            ProjectMargin.objects.filter(year=FIN_YEAR).delete()
        actuals = V._collect_actuals(BUS, {FIN_YEAR})

        def cost_ratio(bu, m):
            rev, _prof, gross = V._rp(actuals, bu, FIN_YEAR, m)
            if rev and gross is not None:
                return max(0.0, 1 - gross / rev)   # 成本占比 = 1 − 毛利率
            return 0.94                              # 缺财务数据时按 6% 毛利兜底

        # 应收侧：每个项目逐月的收入近似（开票额优先，缺失用预估额）
        proj_name = {p.id: (p.short_name or p.customer_name or p.project_no)
                     for p in ARProject.objects.only('id', 'short_name', 'customer_name', 'project_no')}
        agg = {}  # (project_id, dept, month) -> revenue
        for r in (ARRecord.objects.filter(operation_year=FIN_YEAR)
                  .values('project_id', 'delivery_dept', 'operation_month',
                          'actual_invoice_amount', 'estimated_amount')):
            rev = r['actual_invoice_amount'] or r['estimated_amount'] or 0
            if not rev or rev <= 0:
                continue
            key = (r['project_id'], r['delivery_dept'], r['operation_month'])
            agg[key] = agg.get(key, Decimal('0')) + Decimal(str(rev))

        rows, made = [], 0
        for (pid, dept, m), rev in agg.items():
            name = (proj_name.get(pid) or '').strip()
            if not name or not (1 <= (m or 0) <= 12):
                continue
            cost = (rev * Decimal(str(cost_ratio(dept, m)))).quantize(Decimal('0.01'))
            rows.append(ProjectMargin(
                business_unit=dept, year=FIN_YEAR, month=m, project_name=name,
                revenue=rev.quantize(Decimal('0.01')), cost=cost,
                sales_exp=Decimal('0'), mgmt_exp=Decimal('0')))
            made += 1
        with transaction.atomic():
            ProjectMargin.objects.bulk_create(rows, batch_size=500)
        self.stdout.write(self.style.SUCCESS(f'项目毛利：{made} 条（{FIN_YEAR} 年项目×月）'))
